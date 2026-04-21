"""
New Sanla Page
──────────────
Tracks new "sanla" (pawn) transactions — Jewelry Empeno (NEW) and Storage Empeno (NEW).
Pulls data from daily_reports_brand_a.
Columns: Branch | Jewelry Lotes | Jewelry Capital | Storage Lotes | Storage Capital
"""

from PyQt5.QtWidgets import (
    QWidget, QLabel, QVBoxLayout, QHBoxLayout, QComboBox, QTableWidget,
    QTableWidgetItem, QDateEdit, QMessageBox, QHeaderView, QSizePolicy,
    QPushButton, QScrollArea, QFrame, QFileDialog
)
from PyQt5.QtCore import Qt, QDate, QTimer
from PyQt5.QtGui import QFont, QColor, QBrush, QPainter
from PyQt5.QtPrintSupport import QPrinter, QPrintDialog
from db_connect_pooled import db_manager
from db_worker import run_query_async
from date_range_widget import DateRangeWidget


# Column definitions  ──────────────────────────────────────────────────────
#   (header, db_cols_to_sum, is_lotes)
COLUMNS = [
    ("JEWELRY EMPENO\nLotes",   ["empeno_jew_new_lotes"],  True),
    ("JEWELRY EMPENO\nCapital", ["empeno_jew_new"],         False),
    ("STORAGE EMPENO\nLotes",   ["empeno_sto_new_lotes"],   True),
    ("STORAGE EMPENO\nCapital", ["empeno_sto_new"],          False),
]

HEADERS = ["Branch"] + [c[0] for c in COLUMNS]

# Header colours
COL_COLORS = {
    0: QColor("#495057"),  # Branch
    1: QColor("#dc3545"),  # Jewelry Lotes
    2: QColor("#dc3545"),  # Jewelry Capital
    3: QColor("#e67e22"),  # Storage Lotes
    4: QColor("#e67e22"),  # Storage Capital
}


class ColoredHeaderView(QHeaderView):
    def __init__(self, orientation, parent=None):
        super().__init__(orientation, parent)
        self.colors = {}
        self.setFont(QFont("", 9, QFont.Bold))

    def paintSection(self, painter, rect, logicalIndex):
        painter.save()
        bg = self.colors.get(logicalIndex, QColor("#495057"))
        painter.fillRect(rect, bg)
        painter.setPen(QColor("#333"))
        painter.drawRect(rect.adjusted(0, 0, -1, -1))
        painter.setPen(QColor("white"))
        painter.setFont(QFont("", 9, QFont.Bold))
        text = self.model().headerData(logicalIndex, self.orientation(), Qt.DisplayRole)
        painter.drawText(rect, Qt.AlignCenter | Qt.TextWordWrap, str(text) if text else "")
        painter.restore()


class NewSanlaPage(QWidget):
    """New Sanla — new pawn (Jewelry + Storage) per branch."""

    def __init__(self):
        super().__init__()
        self._is_loading = False
        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(8)

        self._build_controls(layout)
        self._build_table(layout)
        self._build_buttons(layout)
        self.load_corporations()

    # ── Controls ──────────────────────────────────────────────────────────
    def _build_controls(self, parent):
        frame = QFrame()
        frame.setStyleSheet(
            "QFrame{background:#f8f9fa;border:1px solid #dee2e6;"
            "border-radius:5px;padding:10px;}"
        )
        row = QHBoxLayout(frame)
        row.setSpacing(15)

        row.addWidget(self._bold_label("Group:"))
        self.group_selector = self._combo(220)
        self.group_selector.currentTextChanged.connect(self.populate_table)
        row.addWidget(self.group_selector)

        row.addSpacing(20)
        row.addWidget(self._bold_label("Date:"))
        self.date_range_widget = DateRangeWidget()
        self.date_range_widget.dateRangeChanged.connect(self.populate_table)
        self.date_selector = self.date_range_widget  # backward-compat
        row.addWidget(self.date_range_widget)

        row.addSpacing(20)
        row.addWidget(self._bold_label("Status:"))
        self.reg_filter_selector = self._combo(150)
        self.reg_filter_selector.addItem("Registered Only", "registered")
        self.reg_filter_selector.addItem("Not Registered", "not_registered")
        self.reg_filter_selector.addItem("All Branches", "all")
        self.reg_filter_selector.currentIndexChanged.connect(self.populate_table)
        row.addWidget(self.reg_filter_selector)

        # Keep os_filter_selector for backward compat but hidden
        self.os_filter_selector = self._combo(180)
        self.os_filter_selector.addItem("All (by Corporation)", None)
        self.os_filter_selector.setVisible(False)
        # Keep corp_selector for backward compat but hidden
        self.corp_selector = self._combo(220)
        self.corp_selector.setVisible(False)

        row.addStretch()
        parent.addWidget(frame, 0)

    # ── Table ─────────────────────────────────────────────────────────────
    def _build_table(self, parent):
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        self.table = QTableWidget()
        self.table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.table.setColumnCount(len(HEADERS))
        self.table.setHorizontalHeaderLabels(HEADERS)
        self.table.setMinimumHeight(300)
        self.table.verticalHeader().setDefaultSectionSize(35)
        self.table.setAlternatingRowColors(True)

        self.table.setColumnWidth(0, 160)
        for i in range(1, len(HEADERS)):
            self.table.setColumnWidth(i, 130)

        header = self.table.horizontalHeader()
        header.setStretchLastSection(True)
        for i in range(len(HEADERS)):
            header.setSectionResizeMode(i, QHeaderView.Stretch)

        self._apply_header_colors()
        self._style_table()

        scroll.setWidget(self.table)
        parent.addWidget(scroll, 1)

    def _apply_header_colors(self):
        ch = ColoredHeaderView(Qt.Horizontal, self.table)
        ch.colors = dict(COL_COLORS)
        ch.setDefaultAlignment(Qt.AlignCenter)
        self.table.setHorizontalHeader(ch)

    def _style_table(self):
        self.table.setStyleSheet("""
            QTableWidget {
                gridline-color:#d0d0d0; border:1px solid #c0c0c0;
                background:white; alternate-background-color:#f8f9fa;
                font-size:12px; selection-background-color:#e3f2fd;
            }
            QTableWidget::item { border:1px solid #e0e0e0; padding:6px; }
            QTableWidget::item:selected { background:#e3f2fd; color:black; }
            QHeaderView::section {
                padding:8px 4px; font-weight:bold; font-size:10px;
                border:1px solid #666; color:white;
            }
        """)

    # ── Buttons ───────────────────────────────────────────────────────────
    def _build_buttons(self, parent):
        frame = QFrame()
        frame.setFixedHeight(70)
        lay = QHBoxLayout(frame)

        self.export_btn = self._action_btn("📊 Export to Excel", "#217346", "#1a5c38", "#155724")
        self.export_btn.clicked.connect(self.export_to_excel)
        self.print_btn = self._action_btn("🖨️ Print Report", "#6f42c1", "#5a2d91", "#4c1f75")
        self.print_btn.clicked.connect(self.print_table)

        lay.addStretch()
        lay.addWidget(self.export_btn)
        lay.addSpacing(15)
        lay.addWidget(self.print_btn)
        lay.addStretch()
        parent.addWidget(frame, 0)

    # ── Data loading ──────────────────────────────────────────────────────
    def load_corporations(self):
        self.group_selector.blockSignals(True)
        self.group_selector.clear()
        try:
            rows = db_manager.execute_query(
                "SELECT DISTINCT os_name FROM branches "
                "WHERE os_name IS NOT NULL AND os_name != '' ORDER BY os_name"
            )
            if rows:
                for r in rows:
                    name = r['os_name'] if isinstance(r, dict) else r[0]
                    self.group_selector.addItem(name)
        except Exception as e:
            print(f"Error loading groups: {e}")
        finally:
            self.group_selector.blockSignals(False)

    def _load_os_options(self):
        self.os_filter_selector.blockSignals(True)
        prev = self.os_filter_selector.currentData()
        self.os_filter_selector.clear()
        self.os_filter_selector.addItem("All (by Corporation)", None)
        try:
            rows = db_manager.execute_query(
                "SELECT DISTINCT os_name FROM branches "
                "WHERE os_name IS NOT NULL AND os_name != '' ORDER BY os_name"
            )
            if rows:
                for r in rows:
                    name = r['os_name'] if isinstance(r, dict) else r[0]
                    self.os_filter_selector.addItem(name, name)
            if prev:
                idx = self.os_filter_selector.findData(prev)
                if idx >= 0:
                    self.os_filter_selector.setCurrentIndex(idx)
        except Exception as e:
            print(f"Error loading OS options: {e}")
        finally:
            self.os_filter_selector.blockSignals(False)

    # ── Populate table ────────────────────────────────────────────────────
    def populate_table(self):
        self._is_loading = True
        self.table.setRowCount(0)

        group = self.group_selector.currentText().strip() if hasattr(self, 'group_selector') else ""
        date_start, date_end = self.date_range_widget.get_date_range()
        is_range = self.date_range_widget.is_range_mode()
        reg_value = self.reg_filter_selector.currentData()

        if not group:
            self._is_loading = False
            return

        needed = set()
        for _, db_cols, _ in COLUMNS:
            needed.update(db_cols)

        select_parts = ["b.name AS branch"]
        for col in sorted(needed):
            if is_range:
                select_parts.append(f"SUM(COALESCE(dr.`{col}`, 0)) AS `{col}`")
            else:
                select_parts.append(f"COALESCE(dr.`{col}`, 0) AS `{col}`")

        if is_range:
            where_parts = ["dr.date >= %s", "dr.date <= %s"]
            params = [date_start, date_end]
        else:
            where_parts = ["dr.date = %s"]
            params = [date_start]

        if group:
            where_parts.append("b.os_name = %s")
            params.append(group)
        if reg_value == "registered":
            where_parts.append("b.is_registered = 1")
        elif reg_value == "not_registered":
            where_parts.append("(b.is_registered = 0 OR b.is_registered IS NULL)")

        group_by = " GROUP BY b.name" if is_range else ""
        sql = (
            f"SELECT {', '.join(select_parts)} "
            f"FROM branches b "
            f"LEFT JOIN daily_reports_brand_a dr ON b.name COLLATE utf8mb4_general_ci = dr.branch COLLATE utf8mb4_general_ci "
            f"WHERE {' AND '.join(where_parts)}"
            f"{group_by} "
            f"ORDER BY b.name"
        )

        run_query_async(
            parent=self,
            query=sql,
            params=tuple(params),
            on_result=self._on_populate_result,
            on_error=self._on_populate_error,
            loading_message="\u23f3  Loading sanla data\u2026",
        )

    def _on_populate_error(self, err):
        QMessageBox.critical(self, "Query Error", err)
        self._is_loading = False

    def _on_populate_result(self, rows):
        if not rows:
            self._is_loading = False
            return

        totals = [0.0] * len(HEADERS)

        for row_data in rows:
            r = self.table.rowCount()
            self.table.insertRow(r)

            branch_item = QTableWidgetItem(row_data['branch'])
            branch_item.setFlags(branch_item.flags() & ~Qt.ItemIsEditable)
            branch_item.setFont(QFont("", 10, QFont.Bold))
            self.table.setItem(r, 0, branch_item)

            for ci, (_, db_cols, is_lotes) in enumerate(COLUMNS, start=1):
                val = sum(float(row_data.get(c, 0) or 0) for c in db_cols)
                if is_lotes:
                    val = int(val)
                    txt = str(val)
                else:
                    txt = f"{val:,.2f}"
                item = QTableWidgetItem(txt)
                item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                self.table.setItem(r, ci, item)
                totals[ci] += val

        # Totals row
        r = self.table.rowCount()
        self.table.insertRow(r)
        total_item = QTableWidgetItem("TOTAL")
        total_item.setFont(QFont("", 10, QFont.Bold))
        total_item.setData(Qt.BackgroundRole, QBrush(QColor("#343a40")))
        total_item.setData(Qt.ForegroundRole, QBrush(QColor("#343a40")))
        total_item.setFlags(total_item.flags() & ~Qt.ItemIsEditable)
        self.table.setItem(r, 0, total_item)

        for ci, (_, _, is_lotes) in enumerate(COLUMNS, start=1):
            v = totals[ci]
            txt = str(int(v)) if is_lotes else f"{v:,.2f}"
            item = QTableWidgetItem(txt)
            item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            item.setFont(QFont("", 10, QFont.Bold))
            item.setData(Qt.BackgroundRole, QBrush(QColor("#343a40")))
            item.setData(Qt.ForegroundRole, QBrush(QColor("#343a40")))
            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
            self.table.setItem(r, ci, item)

        self._is_loading = False

    # ── Export ────────────────────────────────────────────────────────────
    def export_to_excel(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            QMessageBox.warning(self, "Missing Package", "Install openpyxl:\npip install openpyxl")
            return

        if self.table.rowCount() == 0:
            QMessageBox.warning(self, "No Data", "Please load data first.")
            return

        date_start, date_end = self.date_range_widget.get_date_range()
        date_label = date_start if date_start == date_end else f"{date_start}_to_{date_end}"
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Excel",
            f"New_Sanla_{date_label}.xlsx",
            "Excel Files (*.xlsx)"
        )
        if not path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "New Sanla"
        hfont = Font(bold=True, color="FFFFFF", size=10)
        hfill = PatternFill("solid", fgColor="495057")
        thin = Side(style='thin')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for c in range(self.table.columnCount()):
            text = self.table.horizontalHeaderItem(c).text().replace("\n", " ")
            cell = ws.cell(row=1, column=c + 1, value=text)
            cell.font = hfont
            cell.fill = hfill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

        for r in range(self.table.rowCount()):
            for c in range(self.table.columnCount()):
                item = self.table.item(r, c)
                val = item.text() if item else ""
                try:
                    num = float(val.replace(",", ""))
                    cell = ws.cell(row=r + 2, column=c + 1, value=num)
                    cell.number_format = '#,##0.00' if '.' in val else '#,##0'
                except ValueError:
                    cell = ws.cell(row=r + 2, column=c + 1, value=val)
                cell.border = border

        for col_cells in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col_cells), default=8)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 25)

        wb.save(path)
        QMessageBox.information(self, "Export Successful", f"Saved to:\n{path}")

    # ── Print ─────────────────────────────────────────────────────────────
    def print_table(self):
        if self.table.rowCount() == 0:
            QMessageBox.warning(self, "No Data", "Please load data first.")
            return
        printer = QPrinter(QPrinter.HighResolution)
        printer.setOrientation(QPrinter.Landscape)
        dlg = QPrintDialog(printer, self)
        if dlg.exec_() != QPrintDialog.Accepted:
            return

        painter = QPainter(printer)
        try:
            page = printer.pageRect()
            cols = self.table.columnCount()
            rows = self.table.rowCount()
            col_w = page.width() / max(cols, 1)
            row_h = 400
            y = 0
            painter.setFont(QFont("Arial", 8, QFont.Bold))
            for c in range(cols):
                text = self.table.horizontalHeaderItem(c).text().replace("\n", " ")
                painter.drawText(int(c * col_w), y, int(col_w), row_h,
                                 Qt.AlignCenter, text)
            y += row_h
            painter.setFont(QFont("Arial", 7))
            for r in range(rows):
                if y + row_h > page.height():
                    printer.newPage()
                    y = 0
                for c in range(cols):
                    item = self.table.item(r, c)
                    painter.drawText(int(c * col_w), y, int(col_w), row_h,
                                     Qt.AlignCenter, item.text() if item else "")
                y += row_h
        finally:
            painter.end()

    # ── Helpers ───────────────────────────────────────────────────────────
    @staticmethod
    def _bold_label(text):
        lbl = QLabel(text)
        lbl.setFont(QFont("Arial", 11, QFont.Bold))
        return lbl

    @staticmethod
    def _combo(min_w=160):
        cb = QComboBox()
        cb.setMinimumWidth(min_w)
        cb.setMinimumHeight(38)
        cb.setStyleSheet(
            "QComboBox{padding:8px;border:2px solid #dee2e6;border-radius:6px;"
            "background:white;font-size:13px;}"
            "QComboBox:focus{border-color:#007bff;}"
            "QComboBox::drop-down{width:28px;}"
        )
        return cb

    @staticmethod
    def _action_btn(text, c1, c2, c3):
        b = QPushButton(text)
        b.setStyleSheet(f"""
            QPushButton{{padding:12px 24px;border:none;border-radius:8px;
                background:{c1};color:white;font-weight:bold;font-size:12px;
                min-width:130px;min-height:40px;}}
            QPushButton:hover{{background:{c2};}}
            QPushButton:pressed{{background:{c3};}}
        """)
        return b
