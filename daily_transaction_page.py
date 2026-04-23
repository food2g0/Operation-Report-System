from PyQt5.QtWidgets import (
    QWidget, QLabel, QVBoxLayout, QHBoxLayout, QComboBox, QTableWidget,
    QTableWidgetItem, QDateEdit, QMessageBox, QHeaderView, QSizePolicy,
    QPushButton, QScrollArea, QFrame, QFileDialog, QApplication
)
from PyQt5.QtCore import Qt, QDate, QTimer
from PyQt5.QtGui import QFont, QColor, QBrush, QPainter
from PyQt5.QtPrintSupport import QPrinter, QPrintDialog
from db_connect_pooled import db_manager
from db_worker import run_query_async
from date_range_widget import DateRangeWidget
import datetime
import json
import os


_db_columns_cache = {}  
_db_cache_time = 0
_DB_CACHE_TTL = 300  


def _get_table_columns(table_name: str) -> set:

    global _db_columns_cache, _db_cache_time
    import time
    
    current_time = time.time()
    

    if table_name in _db_columns_cache and (current_time - _db_cache_time) < _DB_CACHE_TTL:
        return _db_columns_cache[table_name]
    
    try:

        result = db_manager.execute_query(
            f"SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = %s AND TABLE_SCHEMA = DATABASE()",
            (table_name,)
        )
        if result and isinstance(result, (list, tuple)):
            columns = set()
            for row in result:
        
                if isinstance(row, dict):
                    col_name = row.get('COLUMN_NAME') or row.get('column_name') or ''
                elif isinstance(row, (list, tuple)) and len(row) > 0:
                    col_name = row[0]
                else:
                    col_name = ''
                if col_name:
                    columns.add(col_name)
            if columns:
                _db_columns_cache[table_name] = columns
                _db_cache_time = current_time
                print(f"Loaded {len(columns)} columns from {table_name}")
                return columns
    except Exception as e:
        print(f"Error getting columns for {table_name}: {e}")
    

    if table_name in _db_columns_cache:
        return _db_columns_cache[table_name]
    
    return set()


def _filter_column_groups(groups: list, table_name: str = "daily_reports_brand_a") -> list:

    available_columns = _get_table_columns(table_name)
    
    if not available_columns:
  
        return groups
    
    filtered_groups = []
    
    for group_name, subs in groups:
        filtered_subs = []
        for sub_header, db_cols, is_lotes in subs:
   
            existing_cols = [col for col in db_cols if col in available_columns]
            if existing_cols:
                filtered_subs.append((sub_header, existing_cols, is_lotes))
        
   
        if filtered_subs:
            filtered_groups.append((group_name, filtered_subs))
    
    return filtered_groups


def refresh_schema_cache():

    global _db_columns_cache, _db_cache_time
    _db_columns_cache = {}
    _db_cache_time = 0


def _load_dynamic_fields_for_report(report_type: str) -> list:

    dynamic_groups = []
    
    try:

        result = db_manager.execute_query(
            "SELECT config_value FROM field_config WHERE config_key = 'field_definitions' ORDER BY id DESC LIMIT 1"
        )
        if result and result[0].get('config_value'):
            config = json.loads(result[0]['config_value'])
        else:
          
            import os
            config_path = os.path.join(os.path.dirname(__file__), "field_config.json")
            if os.path.exists(config_path):
                with open(config_path, 'r', encoding='utf-8') as f:
                    config = json.load(f)
            else:
                return []
        
        for section in ['debit', 'credit']:
            fields = config.get('Brand A', {}).get(section, [])
            for field in fields:
                if len(field) >= 4 and isinstance(field[3], dict):
                    reports = field[3].get('reports', [])
                    if report_type in reports:
                        label = field[0]
                        db_col = field[2]
                    
                        group = (
                            label.upper(),
                            [
                                ("LOTES", [f"{db_col}_lotes"], True),
                                ("CAPITAL", [db_col], False),
                            ]
                        )
             
                        if group not in dynamic_groups:
                            dynamic_groups.append(group)
    except Exception as e:
        print(f"Error loading dynamic fields for {report_type}: {e}")
    
    return dynamic_groups

COLUMN_GROUPS = [
    ("JEWELRY", [
        ("Lotes", ["empeno_jew_new_lotes", "empeno_jew_renew_lotes"], True),
        ("Capital", ["empeno_jew_new", "empeno_jew_renew"], False),
    ]),
    ("STORAGE", [
        ("Lotes", ["empeno_sto_new_lotes", "fund_empeno_sto_renew_lotes"], True),
        ("Capital", ["empeno_sto_new", "fund_empeno_sto_renew"], False),
    ]),
    ("MOTOR/CAR", [
        ("Lotes", ["empeno_motor_car_lotes"], True),
        ("Capital", ["empeno_motor_car"], False),
    ]),
    ("MC", [
        ("Lotes", ["mc_out_lotes"], True),
        ("Capital", ["mc_out"], False),
    ]),
    ("SILVER", [
        ("Lotes", ["empeno_silver_lotes"], True),
        ("Capital", ["empeno_silver"], False),
    ]),
    ("PALAWAN", [
        ("Lotes", ["palawan_send_out_lotes", "palawan_sc_lotes",
                    "palawan_pay_out_lotes", "palawan_pay_out_incentives_lotes"], True),
        ("Capital", ["palawan_send_out", "palawan_sc",
                      "palawan_pay_out", "palawan_pay_out_incentives"], False),
    ]),
    ("INSURANCE", [
        ("20s", ["insurance_20"], False),
        ("30s", ["insurance_philam_30"], False),
        ("60s", ["insurance_philam_60"], False),
        ("90s", ["insurance_philam_90"], False),
    ]),
    ("O.S.F", [
        ("Lotes", ["osf_storage_lotes", "osf_silver_lotes", "osf_motor_lotes"], True),
        ("Capital", ["osf_storage", "osf_silver", "osf_motor"], False),
    ]),
    ("RESCATE JEW.", [
        ("Lotes", ["rescate_jewelry_lotes"], True),
        ("Capital", ["rescate_jewelry"], False),
    ]),
    ("RESCATE STO.", [
        ("Lotes", ["cr_storage_lotes", "rescate_silver_lotes",
                    "res_storage_lotes", "res_motor_lotes"], True),
        ("Capital", ["cr_storage", "rescate_silver", "res_storage", "res_motor"], False),
    ]),
    ("GCASH IN", [
        ("Lotes", ["gcash_in_lotes"], True),
        ("Capital", ["gcash_in"], False),
    ]),
    ("GCASH OUT", [
        ("Lotes", ["gcash_out_lotes"], True),
        ("Capital", ["gcash_out"], False),
    ]),
    ("MONEYGRAM", [
        ("Lotes", ["moneygram_lotes"], True),
        ("Capital", ["moneygram"], False),
    ]),
    ("TRANSFAST", [
        ("Lotes", ["transfast_lotes"], True),
        ("Capital", ["transfast"], False),
    ]),
    ("RIA", [
        ("Lotes", ["ria_out_lotes"], True),
        ("Capital", ["ria_out"], False),
    ]),
    ("I2I REM. IN", [
        ("Lotes", ["i2i_remittance_in_lotes"], True),
        ("Capital", ["i2i_remittance_in"], False),
    ]),
    ("I2I BILLS", [
        ("Lotes", ["i2i_bills_payment_lotes"], True),
        ("Capital", ["i2i_bills_payment"], False),
    ]),
    ("I2I INSTAPAY", [
        ("Lotes", ["i2i_instapay_lotes"], True),
        ("Capital", ["i2i_instapay"], False),
    ]),
    ("SENDAH LOAD", [
        ("Lotes", ["sendah_load_sc_lotes"], True),
        ("Capital", ["sendah_load_sc"], False),
    ]),
    ("SENDAH BILLS", [
        ("Lotes", ["sendah_bills_sc_lotes"], True),
        ("Capital", ["sendah_bills_sc"], False),
    ]),
    ("PAYMAYA", [
        ("Lotes", ["paymaya_in_lotes"], True),
        ("Capital", ["paymaya_in"], False),
    ]),
    ("SMART $ IN", [
        ("Lotes", ["smart_money_sc_lotes"], True),
        ("Capital", ["smart_money_sc"], False),
    ]),
    ("SMART $ OUT", [
        ("Lotes", ["smart_money_po_lotes"], True),
        ("Capital", ["smart_money_po"], False),
    ]),
    ("GCASH PADALA", [
        ("Lotes", ["gcash_padala_sendah_lotes"], True),
        ("Capital", ["gcash_padala_sendah"], False),
    ]),
    ("PAL PAY IN", [
        ("Lotes", ["palawan_pay_cash_in_sc_lotes"], True),
        ("Capital", ["palawan_pay_cash_in_sc"], False),
    ]),
    ("PAL PAY OUT", [
        ("Lotes", ["palawan_pay_cash_out_lotes"], True),
        ("Capital", ["palawan_pay_cash_out"], False),
    ]),
    ("REMITLY", [
        ("Lotes", ["remitly_lotes"], True),
        ("Capital", ["remitly"], False),
    ]),
]


OTHER_SERVICES_COLUMN_GROUPS = [
    ("PAL SEND OUT",  [("Lotes", ["palawan_send_out_lotes"], True),
                       ("Amount", ["palawan_send_out"], False)]),
    ("PAL SC",        [("Lotes", ["palawan_sc_lotes"], True),
                       ("Amount", ["palawan_sc"], False)]),
    ("PAL PAY OUT",   [("Lotes", ["palawan_pay_out_lotes"], True),
                       ("Amount", ["palawan_pay_out"], False)]),
    ("PAL INC.",      [("Lotes", ["palawan_pay_out_incentives_lotes"], True),
                       ("Amount", ["palawan_pay_out_incentives"], False)]),
    ("PAL PAY IN",    [("Lotes", ["palawan_pay_cash_in_sc_lotes"], True),
                       ("Amount", ["palawan_pay_cash_in_sc"], False)]),
    ("PAL PAY BILLS", [("Amount", ["palawan_pay_bills_sc"], False)]),
    ("PAL LOAD",      [("Amount", ["palawan_load_sc"], False)]),
    ("PAL PAY OUT",   [("Lotes", ["palawan_pay_cash_out_lotes"], True),
                       ("Amount", ["palawan_pay_cash_out"], False)]),
    ("SUKI CARD",     [("Amount", ["palawan_suki_card"], False)]),
    ("PPAY OUT SC",   [("Amount", ["palawan_pay_cash_out_sc"], False)]),
    ("SENDAH LOAD",   [("Lotes", ["sendah_load_sc_lotes"], True),
                       ("Amount", ["sendah_load_sc"], False)]),
    ("SENDAH BILLS",  [("Lotes", ["sendah_bills_sc_lotes"], True),
                       ("Amount", ["sendah_bills_sc"], False)]),
    ("SMART $ IN",    [("Lotes", ["smart_money_sc_lotes"], True),
                       ("Amount", ["smart_money_sc"], False)]),
    ("SMART $ OUT",   [("Lotes", ["smart_money_po_lotes"], True),
                       ("Amount", ["smart_money_po"], False)]),
    ("GCASH IN",      [("Lotes", ["gcash_in_lotes"], True),
                       ("Amount", ["gcash_in"], False)]),
    ("GCASH OUT",     [("Lotes", ["gcash_out_lotes"], True),
                       ("Amount", ["gcash_out"], False)]),
    ("GCASH PADALA",  [("Lotes", ["gcash_padala_sendah_lotes"], True),
                       ("Amount", ["gcash_padala_sendah"], False)]),
    ("ABRA IN",       [("Amount", ["abra_so_sc"], False)]),
    ("ABRA OUT",      [("Amount", ["abra_po"], False)]),
    ("REMITLY",       [("Lotes", ["remitly_lotes"], True),
                       ("Amount", ["remitly"], False)]),
    ("PAYMAYA IN",    [("Lotes", ["paymaya_in_lotes"], True),
                       ("Amount", ["paymaya_in"], False)]),
    ("PAYMAYA OUT",   [("Amount", ["paymaya_out"], False)]),
    ("RIA IN",        [("Lotes", ["ria_in_sc_lotes"], True),
                       ("Amount", ["ria_in_sc"], False)]),
    ("RIA OUT",       [("Amount", ["ria_out"], False)]),
    ("BDO SC",        [("Amount", ["bdo_sc"], False)]),
    ("BDO OUT",       [("Amount", ["bdo_po"], False)]),
    ("TRANSFAST",     [("Lotes", ["transfast_lotes"], True),
                       ("Amount", ["transfast"], False)]),
    ("AYANAH SC",     [("Amount", ["ayanah_sc"], False)]),
    ("AYANAH OUT",    [("Amount", ["ayanah_out"], False)]),
    ("MONEYGRAM",     [("Lotes", ["moneygram_lotes"], True),
                       ("Amount", ["moneygram"], False)]),
    ("I2I REM IN",    [("Lotes", ["i2i_remittance_in_lotes"], True),
                       ("Amount", ["i2i_remittance_in"], False)]),
    ("I2I BILLS",     [("Lotes", ["i2i_bills_payment_lotes"], True),
                       ("Amount", ["i2i_bills_payment"], False)]),
    ("I2I BANK TRF",  [("Amount", ["i2i_bank_transfer"], False)]),
    ("I2I PESONET",   [("Amount", ["i2i_pesonet"], False)]),
    ("I2I INSTAPAY",  [("Lotes", ["i2i_instapay_lotes"], True),
                       ("Amount", ["i2i_instapay"], False)]),
    ("FIXCO OUT",     [("Amount", ["fixco"], False)]),
    ("I2I REM OUT",   [("Amount", ["i2i_remittance_out"], False)]),
    ("ECBILLS",       [("Amount", ["ecbills"], False)]),
    ("ECLOAD",        [("Amount", ["ecload"], False)]),
    ("ECPINS",        [("Amount", ["ecpins"], False)]),
    ("ECCASH",        [("Amount", ["eccash"], False)]),
    ("ECTICKETS",     [("Amount", ["ectickets"], False)]),
    ("ECCATALOGUE",   [("Amount", ["eccatalogue"], False)]),
    ("ECLINK",        [("Amount", ["eclink"], False)]),
    ("ECGIFT",        [("Amount", ["ecgift"], False)]),
    ("ECCASH-OUT",    [("Amount", ["eccash_out"], False)]),
]


PL_COLUMN_GROUPS = [
    ("INTEREST",          [("Amount", ["interest"], False)]),
    ("PENALTY",           [("Amount", ["penalty"], False)]),
    ("STAMP",             [("Amount", ["stamp"], False)]),
    ("RESCUARDO",         [("Amount", ["rescuardo_affidavit"], False)]),
    ("AI JEWELRY",        [("Amount", ["jew_ai"], False)]),
    ("S.C. JEWELRY",      [("Amount", ["service_charge"], False)]),
    ("SUBASTA INT.",      [("Amount", ["habol_renew_tubos",
                                       "habol_rt_interest_stamp"], False)]),
    ("AI STORAGE",        [("Amount", ["storage_ai"], False)]),
    ("OSF STORAGE",       [("Amount", ["osf_storage"], False)]),
    ("PENALTY STO.",      [("Amount", ["cr_storage_int_penalty"], False)]),
    ("AI SILVER",         [("Amount", ["silver_ai"], False)]),
    ("OSF SILVER",        [("Amount", ["osf_silver"], False)]),
    ("PENALTY SILVER",    [("Amount", ["res_storage_int_penalty"], False)]),
    ("AI MOTOR",          [("Amount", ["motor_ai"], False)]),
    ("OSF MOTOR",         [("Amount", ["osf_motor"], False)]),
    ("PENALTY MOTOR",     [("Amount", ["penalty_motor"], False)]),
    ("MISC. FEE",         [("Amount", ["miscellaneous_fee"], False)]),
    ("DISC SUKI",         [("Amount", ["palawan_suki_discounts"], False)]),
    ("REBATES SUKI",      [("Amount", ["palawan_suki_rebates"], False)]),
    ("REBATES STO.",      [("Amount", ["storage_rebates"], False)]),
    ("REBATES SILVER",    [("Amount", ["silver_rebates"], False)]),
    ("SUKI CARD",         [("Amount", ["palawan_suki_card"], False)]),
    ("PC TRANSPO",        [("Amount", ["pc_transpo"], False)]),
    ("SALARY",            [("Amount", ["pc_salary"], False)]),
    ("INC. MOTOR",        [("Amount", ["pc_inc_motor"], False)]),
    ("INC. EMP",          [("Amount", ["pc_inc_emp"], False)]),
    ("INC. SUKI",         [("Amount", ["pc_inc_suki_card"], False)]),
    ("INC. INS.",         [("Amount", ["pc_inc_insurance"], False)]),
    ("INC. MC",           [("Amount", ["pc_inc_mc"], False)]),
    ("PC SUPPLIES",       [("Amount", ["pc_supplies_xerox_maintenance"], False)]),
    ("ELECTRIC",          [("Amount", ["pc_electric"], False)]),
    ("WATER",             [("Amount", ["pc_water"], False)]),
    ("INTERNET",          [("Amount", ["pc_internet"], False)]),
    ("RENT",              [("Amount", ["pc_rental"], False)]),
    ("BUS. PERMIT",       [("Amount", ["pc_permits_bir_payments"], False)]),
    ("PC LBC",            [("Amount", ["pc_lbc_jrs_jnt"], False)]),
]


_BASE_VIEW_MAP = {
    "Daily Transaction": COLUMN_GROUPS,
    "Other Services":      OTHER_SERVICES_COLUMN_GROUPS,
    "P&L":                 PL_COLUMN_GROUPS,
}

_REPORT_TYPE_MAP = {
    "Daily Transaction": "daily_transaction",
    "Other Services": "other_services",
    "P&L": "pnl",
}

_VIEW_TABLE_MAP = {
    "Daily Transaction": "daily_reports_brand_a",
    "Other Services": "daily_reports_brand_a",
    "P&L": "daily_reports_brand_a",
}

def get_view_columns(view_name: str) -> list:
    """Get column groups for a view, filtered by actual database schema."""
    base_groups = list(_BASE_VIEW_MAP.get(view_name, COLUMN_GROUPS))

    report_type = _REPORT_TYPE_MAP.get(view_name)
    if report_type:
        dynamic_groups = _load_dynamic_fields_for_report(report_type)
        if dynamic_groups:
            base_groups.extend(dynamic_groups)
    
    table_name = _VIEW_TABLE_MAP.get(view_name, "daily_reports_brand_a")
    filtered_groups = _filter_column_groups(base_groups, table_name)
    
    return filtered_groups
VIEW_MAP = _BASE_VIEW_MAP

def _build_columns(groups=None):

    if groups is None:
        groups = COLUMN_GROUPS
    headers = [""]  
    col_meta = []
    for group_name, subs in groups:
        for sub_header, db_cols, is_lotes in subs:
            headers.append(sub_header.upper())  
            col_meta.append({
                "group": group_name,
                "sub": sub_header,
                "db_cols": db_cols,
                "is_lotes": is_lotes,
            })
    return headers, col_meta

HEADERS, COL_META = _build_columns()

GROUP_COLORS = {

    "JEWELRY":      QColor("#dc3545"),
    "STORAGE":      QColor("#e67e22"),
    "MOTOR/CAR":    QColor("#8e44ad"),
    "MC":           QColor("#2980b9"),
    "SILVER":       QColor("#7f8c8d"),
    "PALAWAN":      QColor("#28a745"),
    "INSURANCE":    QColor("#17a2b8"),
    "O.S.F":        QColor("#6f42c1"),
    "RESCATE JEW.": QColor("#c0392b"),
    "RESCATE STO.": QColor("#d35400"),
    "GCASH IN":     QColor("#16a085"),
    "GCASH OUT":    QColor("#1abc9c"),
    "MONEYGRAM":    QColor("#2c3e50"),
    "TRANSFAST":    QColor("#34495e"),
    "RIA":          QColor("#e74c3c"),
    "I2I REM. IN":  QColor("#3498db"),
    "I2I BILLS":    QColor("#2471a3"),
    "I2I INSTAPAY": QColor("#1a5276"),
    "SENDAH LOAD":  QColor("#f39c12"),
    "SENDAH BILLS": QColor("#d4ac0d"),
    "PAYMAYA":      QColor("#27ae60"),
    "SMART $ IN":   QColor("#0e6655"),
    "SMART $ OUT":  QColor("#117864"),
    "GCASH PADALA": QColor("#148f77"),
    "PAL PAY IN":   QColor("#1e8449"),
    "PAL PAY OUT":  QColor("#196f3d"),
    "REMITLY":      QColor("#7d3c98"),

    "PAL SEND OUT": QColor("#1e8449"),
    "PAL SC":       QColor("#196f3d"),
    "PAL INC.":     QColor("#145a32"),
    "PAL PAY BILLS":QColor("#0b5345"),
    "PAL LOAD":     QColor("#0a3d2e"),
    "SUKI CARD":    QColor("#1a6b55"),
    "PPAY OUT SC":  QColor("#117a65"),
    "SENDAH LOAD":  QColor("#f39c12"),
    "SENDAH BILLS": QColor("#d4ac0d"),
    "SMART $ IN":   QColor("#0e6655"),
    "SMART $ OUT":  QColor("#117864"),
    "GCASH IN":     QColor("#16a085"),
    "GCASH OUT":    QColor("#1abc9c"),
    "GCASH PADALA": QColor("#148f77"),
    "ABRA IN":      QColor("#2e86c1"),
    "ABRA OUT":     QColor("#1a5276"),
    "REMITLY":      QColor("#7d3c98"),
    "PAYMAYA IN":   QColor("#27ae60"),
    "PAYMAYA OUT":  QColor("#1e8449"),
    "RIA IN":       QColor("#e74c3c"),
    "RIA OUT":      QColor("#c0392b"),
    "BDO SC":       QColor("#2471a3"),
    "BDO OUT":      QColor("#1a5276"),
    "TRANSFAST":    QColor("#34495e"),
    "AYANAH SC":    QColor("#d4ac0d"),
    "AYANAH OUT":   QColor("#b7950b"),
    "MONEYGRAM":    QColor("#2c3e50"),
    "I2I REM IN":   QColor("#3498db"),
    "I2I BILLS":    QColor("#2471a3"),
    "I2I BANK TRF": QColor("#21618c"),
    "I2I PESONET":  QColor("#1b4f72"),
    "I2I INSTAPAY": QColor("#1a5276"),
    "FIXCO OUT":    QColor("#6c3483"),
    "I2I REM OUT":  QColor("#4a235a"),

    "INTEREST":     QColor("#c0392b"),
    "PENALTY":      QColor("#922b21"),
    "STAMP":        QColor("#7b241c"),
    "RESCUARDO":    QColor("#641e16"),
    "AI JEWELRY":   QColor("#1a5276"),
    "S.C. JEWELRY": QColor("#154360"),
    "SUBASTA INT.": QColor("#0d3349"),
    "AI STORAGE":   QColor("#1e8449"),
    "OSF STORAGE":  QColor("#196f3d"),
    "PENALTY STO.": QColor("#145a32"),
    "AI SILVER":    QColor("#7f8c8d"),
    "OSF SILVER":   QColor("#717d7e"),
    "PENALTY SILVER":QColor("#616a6b"),
    "AI MOTOR":     QColor("#8e44ad"),
    "OSF MOTOR":    QColor("#76448a"),
    "PENALTY MOTOR":QColor("#6c3483"),
    "MISC. FEE":    QColor("#2e4057"),
    "DISC SUKI":    QColor("#117864"),
    "REBATES SUKI": QColor("#0e6655"),
    "REBATES STO.": QColor("#0b5345"),
    "REBATES SILVER":QColor("#0a3d2e"),
    "SUKI CARD":    QColor("#1a6b55"),
    "PC TRANSPO":   QColor("#5b2c6f"),
    "SALARY":       QColor("#512e5f"),
    "INC. MOTOR":   QColor("#472e5f"),
    "INC. EMP":     QColor("#3d2b5e"),
    "INC. SUKI":    QColor("#342555"),
    "INC. INS.":    QColor("#2b1f4e"),
    "INC. MC":      QColor("#1f1547"),
    "PC SUPPLIES":  QColor("#34495e"),
    "ELECTRIC":     QColor("#2c3e50"),
    "WATER":        QColor("#1a252f"),
    "INTERNET":     QColor("#212f3c"),
    "RENT":         QColor("#17202a"),
    "BUS. PERMIT":  QColor("#1c2833"),
    "PC LBC":       QColor("#212f3d"),
}


class ColoredHeaderView(QHeaderView):


    def __init__(self, orientation, parent=None):
        super().__init__(orientation, parent)
        self.colors = {}
        self.setFont(QFont("", 8, QFont.Bold))

    def paintSection(self, painter, rect, logicalIndex):
        painter.save()
        bg = self.colors.get(logicalIndex, QColor("#495057"))
        painter.fillRect(rect, bg)
        pen = painter.pen()
        pen.setColor(QColor("#333"))
        painter.setPen(pen)
        painter.drawRect(rect.adjusted(0, 0, -1, -1))
        painter.setPen(QColor("white"))
        painter.setFont(QFont("", 8, QFont.Bold))
        text = self.model().headerData(logicalIndex, self.orientation(), Qt.DisplayRole)
        painter.drawText(rect, Qt.AlignCenter | Qt.TextWordWrap, str(text) if text else "")
        painter.restore()


class MergedGroupHeaderView(QHeaderView):
    def __init__(self, orientation, parent=None, groups=None):
        super().__init__(orientation, parent)
        self.colors = {}
        self.groups = groups or []
        self.setFont(QFont("", 9, QFont.Bold))

        self._merge_info = {}
        self._col_to_group = {}  
        if groups:
            col = 1  
            for group_name, subs in groups:
                span_width = len(subs)
                self._merge_info[col] = (group_name, span_width)
                for i in range(span_width):
                    self._col_to_group[col + i] = col 
                col += span_width

    def paintSection(self, painter, rect, logicalIndex):
        painter.save()
        bg = self.colors.get(logicalIndex, QColor("#495057"))

        if logicalIndex in self._col_to_group:
            start_col = self._col_to_group[logicalIndex]
            group_name, span_width = self._merge_info.get(start_col, ("", 1))
            
            if logicalIndex == start_col:
                merged_rect = rect
                for i in range(1, span_width):
                    next_col = start_col + i
                    section_size = self.sectionSize(next_col)
                    merged_rect = merged_rect.adjusted(0, 0, section_size, 0)
                
                painter.fillRect(merged_rect, bg)
                pen = painter.pen()
                pen.setColor(QColor("#333"))
                painter.setPen(pen)
                painter.drawRect(merged_rect.adjusted(0, 0, -1, -1))
                painter.setPen(QColor("white"))
                painter.setFont(QFont("", 9, QFont.Bold))
                painter.drawText(merged_rect, Qt.AlignCenter, group_name)

        else:
     
            painter.fillRect(rect, bg)
            pen = painter.pen()
            pen.setColor(QColor("#333"))
            painter.setPen(pen)
            painter.drawRect(rect.adjusted(0, 0, -1, -1))
            painter.setPen(QColor("white"))
            painter.setFont(QFont("", 9, QFont.Bold))
            text = self.model().headerData(logicalIndex, self.orientation(), Qt.DisplayRole)
            painter.drawText(rect, Qt.AlignCenter, str(text) if text else "")
        
        painter.restore()

class DailyTransactionPage(QWidget):

    def __init__(self):
        super().__init__()
        self._is_loading = False

        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(8)

        self._build_controls(layout)
        self._build_report_header(layout)
        self._build_table(layout)
        self._build_buttons(layout)

        self.load_corporations()

    def _build_report_header(self, parent):

        self.report_header_frame = QFrame()
        self.report_header_frame.setStyleSheet("""
            QFrame { background: white; border: 1px solid #dee2e6;
                     border-radius: 5px; padding: 8px 15px; }
        """)
        header_layout = QVBoxLayout(self.report_header_frame)
        header_layout.setContentsMargins(15, 10, 15, 10)
        header_layout.setSpacing(2)

        self.report_title_label = QLabel("Daily Transaction")
        self.report_title_label.setFont(QFont("Arial", 14, QFont.Bold))
        self.report_title_label.setStyleSheet("color: #333; border: none;")
        header_layout.addWidget(self.report_title_label)

        self.report_date_label = QLabel("Date: ")
        self.report_date_label.setFont(QFont("Arial", 11, QFont.Bold))
        self.report_date_label.setStyleSheet("color: #555; border: none;")
        header_layout.addWidget(self.report_date_label)

        self.report_os_label = QLabel("Operation Supervisor: ")
        self.report_os_label.setFont(QFont("Arial", 11, QFont.Bold))
        self.report_os_label.setStyleSheet("color: #555; border: none;")
        header_layout.addWidget(self.report_os_label)

        parent.addWidget(self.report_header_frame, 0)

    def _update_report_header(self):
        """Update report header labels with current date and OS selection"""
        date_start, date_end = self.date_range_widget.get_date_range()
        date_display = date_start if date_start == date_end else f"{date_start} to {date_end}"
        self.report_date_label.setText(f"Date: {date_display}")

        group_name = self.group_selector.currentText() if hasattr(self, 'group_selector') else ""
        os_name = self.os_filter_selector.currentText() if not group_name else group_name
        self.report_os_label.setText(f"Group: {os_name or 'All'}")

    def _build_controls(self, parent):
        frame = QFrame()
        frame.setStyleSheet("""
            QFrame { background:#f8f9fa; border:1px solid #dee2e6;
                     border-radius:5px; padding:10px; }
        """)
        row = QHBoxLayout(frame)
        row.setSpacing(15)

   
        self.corp_label = self._bold_label("Corporation:")
        self.corp_selector = self._combo(220)
        self.corp_selector.currentTextChanged.connect(self.populate_table)
        self.corp_label.setVisible(False)
        self.corp_selector.setVisible(False)
        row.addWidget(self.corp_label)
        row.addWidget(self.corp_selector)


        row.addWidget(self._bold_label("Group:"))
        self.group_selector = self._combo(220)
        self.group_selector.currentTextChanged.connect(self.populate_table)
        row.addWidget(self.group_selector)


        row.addSpacing(20)
        self.date_range_widget = DateRangeWidget()
        self.date_range_widget.dateRangeChanged.connect(self.populate_table)
        row.addWidget(self.date_range_widget)

        self.date_selector = self.date_range_widget


        row.addSpacing(20)
        row.addWidget(self._bold_label("Status:"))
        self.reg_filter_selector = self._combo(150)
        self.reg_filter_selector.addItem("Registered Only", "registered")
        self.reg_filter_selector.addItem("Not Registered", "not_registered")
        self.reg_filter_selector.addItem("All Branches", "all")
        self.reg_filter_selector.currentIndexChanged.connect(self.populate_table)
        row.addWidget(self.reg_filter_selector)

        self.os_filter_selector = self._combo(180)
        self.os_filter_selector.addItem("All (by Corporation)", None)
        self.os_filter_selector.setVisible(False)

        row.addSpacing(20)
        row.addWidget(self._bold_label("View:"))
        self.view_selector = self._combo(190)
        for view_name in VIEW_MAP:
            self.view_selector.addItem(view_name, view_name)
        self.view_selector.currentIndexChanged.connect(self._on_view_changed)
        row.addWidget(self.view_selector)

        row.addStretch()
        parent.addWidget(frame, 0)

    def _build_table(self, parent):
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        scroll.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        self.table = QTableWidget()
        self.table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.table.setColumnCount(len(HEADERS))
        self.table.setMinimumHeight(300)
        self.table.verticalHeader().setDefaultSectionSize(32)
        self.table.setAlternatingRowColors(False) 

        initial_groups = get_view_columns("Daily Transaction")
        initial_headers, _ = _build_columns(initial_groups)
        self.table.setColumnCount(len(initial_headers))
        self.table.setColumnWidth(0, 130)  # Branch
        for i in range(1, len(initial_headers)):
            self.table.setColumnWidth(i, 120)
        header = self.table.horizontalHeader()
        header.setStretchLastSection(False)
        for i in range(len(initial_headers)):
            header.setSectionResizeMode(i, QHeaderView.Interactive)
        header.setMinimumSectionSize(55)

        self._apply_header_colors(initial_groups) 
        self._style_table()

        scroll.setWidget(self.table)
        parent.addWidget(scroll, 1)

    def _apply_header_colors(self, groups):

        headers, _ = _build_columns(groups)
        group_headers = ["Branch"]
        col = 1
        for group_name, subs in groups:
            span_width = len(subs)
         
            group_headers.append(group_name)
            for _ in range(span_width - 1):
                group_headers.append("")  
        
        ch = MergedGroupHeaderView(Qt.Horizontal, self.table, groups)
        ch.colors[0] = QColor("#495057")
        col = 1
        for group_name, subs in groups:
            base = GROUP_COLORS.get(group_name, QColor("#495057"))
            for _ in subs:
                ch.colors[col] = base
                col += 1
        ch.setDefaultAlignment(Qt.AlignCenter)
        ch.setMinimumSectionSize(55)
        ch.setDefaultSectionSize(120)
        ch.setStretchLastSection(False)
        self.table.setHorizontalHeader(ch)
        self.table.setHorizontalHeaderLabels(group_headers)
        ch.setVisible(True)
        ch.setFixedHeight(32)
        for i in range(len(group_headers)):
            ch.setSectionResizeMode(i, QHeaderView.Interactive)
        
        self._current_groups = groups
        self._sub_headers = headers  
        self._group_header_row_exists = False

    def _create_group_header_row(self, groups):

        if self.table.rowCount() == 0:
            return
        
        self.table.clearSpans()
        
        self.table.setRowHeight(0, 28)
        
        headers, _ = _build_columns(groups)
        

        branch_item = QTableWidgetItem("")
        branch_item.setTextAlignment(Qt.AlignCenter)
        branch_item.setBackground(QColor("#495057"))
        branch_item.setForeground(QColor("white"))
        font = branch_item.font()
        font.setBold(True)
        font.setPointSize(9)
        branch_item.setFont(font)
        branch_item.setFlags(branch_item.flags() & ~Qt.ItemIsEditable)
        self.table.setItem(0, 0, branch_item)
        
        col = 1
        for group_name, subs in groups:
            group_color = GROUP_COLORS.get(group_name, QColor("#495057"))
            
            for sub_header, _, _ in subs:
                cell_item = QTableWidgetItem(sub_header.upper())
                cell_item.setTextAlignment(Qt.AlignCenter)
                cell_item.setBackground(group_color)
                cell_item.setForeground(QColor("white"))
                font = cell_item.font()
                font.setBold(True)
                font.setPointSize(9)
                cell_item.setFont(font)
                cell_item.setFlags(cell_item.flags() & ~Qt.ItemIsEditable)
                self.table.setItem(0, col, cell_item)
                col += 1
        
        self._group_header_row_exists = True

    def _style_table(self):

        self.table.setStyleSheet("""
            QTableWidget {
                gridline-color:#d0d0d0; border:1px solid #c0c0c0;
                font-size:11px; selection-background-color:#e3f2fd;
                background-color: white;
            }
            QTableWidget::item:selected { background-color:#e3f2fd; color:black; }
            QHeaderView::section {
                padding:6px 2px; font-weight:bold; font-size:9px;
                border:1px solid #666; color:white;
            }
        """)

    def _build_buttons(self, parent):
        frame = QFrame()
        frame.setFixedHeight(70)
        lay = QHBoxLayout(frame)

        def _btn(text, color1, color2, color3):
            b = QPushButton(text)
            b.setStyleSheet(f"""
                QPushButton{{padding:12px 24px;border:none;border-radius:8px;
                    background:{color1};color:white;font-weight:bold;font-size:12px;
                    min-width:130px;min-height:40px;}}
                QPushButton:hover{{background:{color2};}}
                QPushButton:pressed{{background:{color3};}}
            """)
            return b

        self.export_btn = _btn("Export to Excel", "#217346", "#1a5c38", "#155724")
        self.export_btn.clicked.connect(self.export_to_excel)
        self.print_btn = _btn("Print Report", "#6f42c1", "#5a2d91", "#4c1f75")
        self.print_btn.clicked.connect(self.print_table)

        lay.addStretch()
        lay.addWidget(self.export_btn)
        lay.addSpacing(15)
        lay.addWidget(self.print_btn)
        lay.addStretch()
        parent.addWidget(frame, 0)

    def _on_view_changed(self):
        view_name = self.view_selector.currentData() or "Daily Transaction"
        groups = get_view_columns(view_name)
        headers, _ = _build_columns(groups)
        self.table.setRowCount(0)
        self.table.setColumnCount(len(headers))
        self.table.setColumnWidth(0, 130)

        col_width = 200 if view_name == "P&L" else 120
        for i in range(1, len(headers)):
            self.table.setColumnWidth(i, col_width)
        self._apply_header_colors(groups) 
        hdr = self.table.horizontalHeader()
        for i in range(len(headers)):
            hdr.setSectionResizeMode(i, QHeaderView.Interactive)
        self.populate_table()

    def _active_groups(self):
    
        view_name = self.view_selector.currentData() if hasattr(self, 'view_selector') else None
        return get_view_columns(view_name) if view_name else COLUMN_GROUPS

    def load_corporations(self):
        self.corp_selector.blockSignals(True)
        self.corp_selector.clear()
        self.corp_selector.addItem("")
        try:
            rows = db_manager.execute_query(
                "SELECT DISTINCT corporation FROM daily_reports_brand_a ORDER BY corporation"
            )
            for r in rows:
                self.corp_selector.addItem(r['corporation'])
        except Exception as e:
            print(f"Error loading corporations: {e}")
        finally:
            self.corp_selector.blockSignals(False)
        self._load_groups()
        self._load_os_options()

    def _load_groups(self):

        self.group_selector.blockSignals(True)
        self.group_selector.clear()
        try:
            rows = db_manager.execute_query(
                "SELECT DISTINCT os_name FROM branches "
                "WHERE os_name IS NOT NULL AND os_name != '' ORDER BY os_name"
            )
            if rows:
                for r in rows:
                    name = r['os_name'] if isinstance(r, dict) else r[0]
                    self.group_selector.addItem(name)
        except Exception as e:
            print(f"Error loading groups: {e}")
        finally:
            self.group_selector.blockSignals(False)

    def _load_os_options(self):
        self.os_filter_selector.blockSignals(True)
        prev = self.os_filter_selector.currentData()
        self.os_filter_selector.clear()
        self.os_filter_selector.addItem("All (by Corporation)", None)
        try:
            rows = db_manager.execute_query(
                "SELECT DISTINCT os_name FROM branches "
                "WHERE os_name IS NOT NULL AND os_name != '' ORDER BY os_name"
            )
            if rows:
                for r in rows:
                    name = r['os_name'] if isinstance(r, dict) else r[0]
                    self.os_filter_selector.addItem(name, name)
            if prev:
                idx = self.os_filter_selector.findData(prev)
                if idx >= 0:
                    self.os_filter_selector.setCurrentIndex(idx)
        except Exception as e:
            print(f"Error loading OS options: {e}")
        finally:
            self.os_filter_selector.blockSignals(False)

    def populate_table(self):
        self._is_loading = True
        self.table.setRowCount(0)

        corp = self.corp_selector.currentText().strip()
        group = self.group_selector.currentText().strip() if hasattr(self, 'group_selector') else ""
        date_start, date_end = self.date_range_widget.get_date_range()
        is_range = self.date_range_widget.is_range_mode()
        os_filter = self.os_filter_selector.currentData()
        reg_value = self.reg_filter_selector.currentData()


        self._update_report_header()


        if not group and not corp and not os_filter:
            self._is_loading = False
            return

        active_groups = self._active_groups()
        headers, col_meta = _build_columns(active_groups)

        view_name = self.view_selector.currentData() if hasattr(self, 'view_selector') else None
        if self.table.columnCount() != len(headers):
            self.table.setColumnCount(len(headers))
            self.table.setColumnWidth(0, 130)
  
            col_width = 200 if view_name == "P&L" else 120
            for i in range(1, len(headers)):
                self.table.setColumnWidth(i, col_width)
            self._apply_header_colors(active_groups)  

        self.table.insertRow(0)
        self._create_group_header_row(active_groups)

   
        needed_cols = set()
        for _, subs in active_groups:
            for _, db_cols, _ in subs:
                needed_cols.update(db_cols)

        available_columns = _get_table_columns("daily_reports_brand_a")
        if available_columns:
            original_count = len(needed_cols)
            needed_cols = needed_cols & available_columns
            if len(needed_cols) < original_count:
                print(f"Filtered out {original_count - len(needed_cols)} non-existent columns")
        
        if not needed_cols:
            print("Warning: No valid columns to query after filtering")
            self._is_loading = False
            return

        select_parts = ["b.name AS branch"]
        for col in sorted(needed_cols):
            if is_range:
                select_parts.append(f"SUM(COALESCE(dr.`{col}`, 0)) AS `{col}`")
            else:
                select_parts.append(f"COALESCE(dr.`{col}`, 0) AS `{col}`")

        # Date condition goes into the JOIN ON clause so branches without entries still appear
        if is_range:
            join_date_condition = "AND dr.date >= %s AND dr.date <= %s"
            params = [date_start, date_end]
        else:
            join_date_condition = "AND dr.date = %s"
            params = [date_start]

        where_parts = []
        if corp:
            where_parts.append("(b.corporation_id = (SELECT id FROM corporations WHERE name = %s) OR b.sub_corporation_id = (SELECT id FROM corporations WHERE name = %s))")
            params.extend([corp, corp])
        if group:
            where_parts.append("b.os_name = %s")
            params.append(group)
        if os_filter:
            where_parts.append("b.os_name = %s")
            params.append(os_filter)

        if reg_value == "registered":
            where_parts.append("b.is_registered = 1")
        elif reg_value == "not_registered":
            where_parts.append("(b.is_registered = 0 OR b.is_registered IS NULL)")

        where_clause = f"WHERE {' AND '.join(where_parts)}" if where_parts else ""
        group_by = " GROUP BY b.name" if is_range else ""
        sql = (
            f"SELECT {', '.join(select_parts)} "
            f"FROM branches b "
            f"LEFT JOIN daily_reports_brand_a dr ON b.name COLLATE utf8mb4_general_ci = dr.branch COLLATE utf8mb4_general_ci "
            f"{join_date_condition} "
            f"{where_clause} "
            f"{group_by} "
            f"ORDER BY b.name"
        )

        self._pending_populate = {
            'headers': headers,
            'col_meta': col_meta,
            'active_groups': active_groups,
        }

        run_query_async(
            parent=self,
            query=sql,
            params=tuple(params),
            on_result=self._on_populate_result,
            on_error=self._on_populate_error,
            loading_message="⏳  Loading transactions…",
        )

    def _on_populate_error(self, err):
        QMessageBox.critical(self, "Query Error", err)
        self._is_loading = False

    def _on_populate_result(self, rows):
        ctx = self._pending_populate
        headers = ctx['headers']
        col_meta = ctx['col_meta']

        if not rows:
            self._is_loading = False
            return

        col_count = len(headers)
        totals = [0.0] * col_count

        for row_data in rows:
            r = self.table.rowCount()
            self.table.insertRow(r)

            branch_item = QTableWidgetItem(row_data['branch'])
            branch_item.setFlags(branch_item.flags() & ~Qt.ItemIsEditable)
            branch_item.setFont(QFont("", 10, QFont.Bold))
            self.table.setItem(r, 0, branch_item)

            for ci, meta in enumerate(col_meta, start=1):
                val = sum(float(row_data.get(c, 0) or 0) for c in meta['db_cols'])
                if meta['is_lotes']:
                    val = int(val)
                    txt = str(val)
                else:
                    txt = f"{val:,.2f}"
                item = QTableWidgetItem(txt)
                item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                self.table.setItem(r, ci, item)
                totals[ci] += val

        r = self.table.rowCount()
        self.table.insertRow(r)
        total_label = QTableWidgetItem("TOTAL")
        total_label.setFont(QFont("", 10, QFont.Bold))
        total_label.setData(Qt.BackgroundRole, QBrush(QColor("#e2e6eb")))
        total_label.setData(Qt.ForegroundRole, QBrush(QColor("#050505")))
        total_label.setFlags(total_label.flags() & ~Qt.ItemIsEditable)
        self.table.setItem(r, 0, total_label)

        for ci, meta in enumerate(col_meta, start=1):
            v = totals[ci]
            if meta['is_lotes']:
                txt = str(int(v))
            else:
                txt = f"{v:,.2f}"
            item = QTableWidgetItem(txt)
            item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            item.setFont(QFont("", 10, QFont.Bold))
            item.setData(Qt.BackgroundRole, QBrush(QColor("#e2e6eb")))
            item.setData(Qt.ForegroundRole, QBrush(QColor("#050505")))
            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
            self.table.setItem(r, ci, item)

        self._is_loading = False

    def export_to_excel(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            QMessageBox.warning(self, "Missing Package", "Install openpyxl:\npip install openpyxl")
            return

        if self.table.rowCount() <= 1:
            QMessageBox.warning(self, "No Data", "Please load data first.")
            return

        date_start, date_end = self.date_range_widget.get_date_range()
        date_label_str = date_start if date_start == date_end else f"{date_start}_to_{date_end}"
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Excel", f"Daily_Transaction_{date_label_str}.xlsx",
            "Excel Files (*.xlsx)"
        )
        if not path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Daily Transaction"

        header_font = Font(bold=True, color="FFFFFF", size=9)
        header_fill = PatternFill("solid", fgColor="495057")
        thin = Side(style='thin')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        title_font = Font(bold=True, size=14)
        info_font = Font(bold=True, size=11)

        ws.cell(row=1, column=1, value="Daily Transaction").font = title_font
        
        date_start, date_end = self.date_range_widget.get_date_range()
        date_display = date_start if date_start == date_end else f"{date_start} to {date_end}"
        ws.cell(row=2, column=1, value=f"Date: {date_display}").font = info_font
        
        os_name = self.os_filter_selector.currentText() or "All"
        ws.cell(row=3, column=1, value=f"OS: {os_name}").font = info_font

        group_row = 5  
        sub_row = 6    
        data_start = 7   

        active_groups = self._active_groups()
        col = 1

        cell = ws.cell(row=group_row, column=col, value="Branch")
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(start_row=group_row, start_column=col, end_row=sub_row, end_column=col)
        col += 1
        
        for group_name, subs in active_groups:
            span_width = len(subs)
            group_color = GROUP_COLORS.get(group_name, QColor("#495057"))

            hex_color = f"{group_color.red():02X}{group_color.green():02X}{group_color.blue():02X}"
            group_fill = PatternFill("solid", fgColor=hex_color)

            cell = ws.cell(row=group_row, column=col, value=group_name)
            cell.font = header_font
            cell.fill = group_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            

            if span_width > 1:
                ws.merge_cells(start_row=group_row, start_column=col, end_row=group_row, end_column=col + span_width - 1)
            

            for sub_idx, (sub_header, _, _) in enumerate(subs):
                sub_cell = ws.cell(row=sub_row, column=col + sub_idx, value=sub_header.upper())
                sub_cell.font = header_font
                sub_cell.fill = group_fill
                sub_cell.border = border
                sub_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            col += span_width


        for r in range(1, self.table.rowCount()):
            excel_row = data_start + (r - 1)
            for c in range(self.table.columnCount()):
                item = self.table.item(r, c)
                val = item.text() if item else ""
    
                try:
                    num = float(val.replace(",", ""))
                    cell = ws.cell(row=excel_row, column=c + 1, value=num)
                    cell.number_format = '#,##0.00' if '.' in val else '#,##0'
                except ValueError:
                    cell = ws.cell(row=excel_row, column=c + 1, value=val)
                cell.border = border


        for col_cells in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col_cells), default=8)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 18)

        wb.save(path)
        QMessageBox.information(self, "Export Successful", f"Saved to:\n{path}")


    def print_table(self):
        if self.table.rowCount() <= 1:  
            QMessageBox.warning(self, "No Data", "Please load data first.")
            return
        printer = QPrinter(QPrinter.HighResolution)
        printer.setOrientation(QPrinter.Landscape)
        dlg = QPrintDialog(printer, self)
        if dlg.exec_() != QPrintDialog.Accepted:
            return

        painter = QPainter(printer)
        try:
            page = printer.pageRect()
            cols = self.table.columnCount()
            rows = self.table.rowCount()
            col_w = page.width() / max(cols, 1)
            row_h = 300
            y = 0
            

            painter.setFont(QFont("Arial", 6, QFont.Bold))
            for c in range(cols):
                item = self.table.item(0, c)
                text = item.text() if item else ""
                painter.drawText(int(c * col_w), y, int(col_w), row_h,
                                 Qt.AlignCenter | Qt.TextWordWrap, text)
            y += row_h
            

            for c in range(cols):
                text = self.table.horizontalHeaderItem(c).text().replace("\n", " ") if self.table.horizontalHeaderItem(c) else ""
                painter.drawText(int(c * col_w), y, int(col_w), row_h,
                                 Qt.AlignCenter | Qt.TextWordWrap, text)
            y += row_h
            
            painter.setFont(QFont("Arial", 5))
            for r in range(1, rows):  # Start from row 1
                if y + row_h > page.height():
                    printer.newPage()
                    y = 0
                for c in range(cols):
                    item = self.table.item(r, c)
                    painter.drawText(int(c * col_w), y, int(col_w), row_h,
                                     Qt.AlignCenter, item.text() if item else "")
                y += row_h
        finally:
            painter.end()


    @staticmethod
    def _bold_label(text):
        lbl = QLabel(text)
        lbl.setFont(QFont("Arial", 11, QFont.Bold))
        return lbl

    @staticmethod
    def _combo(min_w=160):
        cb = QComboBox()
        cb.setMinimumWidth(min_w)
        cb.setMinimumHeight(38)
        cb.setStyleSheet(
            "QComboBox{padding:8px;border:2px solid #dee2e6;border-radius:6px;"
            "background:white;font-size:13px;}"
            "QComboBox:focus{border-color:#007bff;}"
            "QComboBox::drop-down{width:28px;}"
        )
        return cb
