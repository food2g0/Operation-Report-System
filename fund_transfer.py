from PyQt5.QtWidgets import (
    QWidget, QLabel, QVBoxLayout, QComboBox, QTableWidget, QTableWidgetItem,
    QDateEdit, QMessageBox, QHeaderView, QSizePolicy, QPushButton, QHBoxLayout,
    QAbstractItemView, QFileDialog, QTabWidget, QLineEdit, QGroupBox, QFrame,
)
from PyQt5.QtPrintSupport import QPrinter, QPrintDialog
from PyQt5.QtGui import QFont, QColor, QBrush, QTextDocument, QDoubleValidator
from PyQt5.QtCore import Qt, QDate
from db_connect_pooled import db_manager
from db_worker import run_func_async
from date_range_widget import DateRangeWidget
import datetime


class FundTransferPage(QWidget):
    def __init__(self, account_type=2):
        super().__init__()
        self.account_type = account_type
        self.setWindowTitle("Fund Transfer")
        self.resize(1100, 650)

        self.layout = QVBoxLayout()
        self.layout.setContentsMargins(5, 5, 5, 5)
        self.layout.setSpacing(4)
        self.setLayout(self.layout)

        self.report_type_selector = QComboBox()
        self.report_type_selector.addItem("Corporation", "corporation")
        self.report_type_selector.addItem("Group", "os")
        self.report_type_selector.currentIndexChanged.connect(self.on_report_type_changed)

        self.corp_label = QLabel("Corporation:")
        self.corp_selector = QComboBox()
        self.corp_selector.currentTextChanged.connect(self.populate_table)

        self.os_label = QLabel("Group:")
        self.os_selector = QComboBox()
        self.os_selector.currentTextChanged.connect(self.populate_table)
        self.os_label.setVisible(False)
        self.os_selector.setVisible(False)

        self.daily_table = "daily_reports_brand_a"

        self.date_range_widget = DateRangeWidget()
        self.date_range_widget.dateRangeChanged.connect(self.populate_table)
        self.date_selector = self.date_range_widget  # backward-compat

        self.reg_filter_label = QLabel("Status:")
        self.reg_filter_selector = QComboBox()
        self.reg_filter_selector.addItem("Registered Only", "registered")
        self.reg_filter_selector.addItem("Not Registered", "not_registered")
        self.reg_filter_selector.addItem("All Branches", "all")
        self.reg_filter_selector.currentIndexChanged.connect(self.populate_table)

        filter_row = QHBoxLayout()
        filter_row.setSpacing(8)
        self.report_type_label = QLabel("Report Type:")
        filter_row.addWidget(self.report_type_label)
        filter_row.addWidget(self.report_type_selector)
        filter_row.addWidget(self.corp_label)
        filter_row.addWidget(self.corp_selector, 1)
        filter_row.addWidget(self.os_label)
        filter_row.addWidget(self.os_selector, 1)
        filter_row.addWidget(self.date_range_widget)
        filter_row.addWidget(self.reg_filter_label)
        filter_row.addWidget(self.reg_filter_selector)
        self.layout.addLayout(filter_row)

        self.report_type_selector.setCurrentIndex(1) 
        self.report_type_label.setVisible(False)
        self.report_type_selector.setVisible(False)
        self.corp_label.setVisible(False)
        self.corp_selector.setVisible(False)
        self.os_label.setVisible(True)
        self.os_selector.setVisible(True)

        self.tab_widget = QTabWidget()
        self.layout.addWidget(self.tab_widget)

        fund_tab = QWidget()
        fund_tab_layout = QVBoxLayout()
        fund_tab.setLayout(fund_tab_layout)

        # Table
        self.table = QTableWidget()
        self.table.setColumnCount(12)
        self.table.setHorizontalHeaderLabels([
            "AREA", "#", "CORPORATION", "LOB", "GLOBAL", "SUNDAY",
            "Branch Name", "Invty", "CASH FLOAT", "CASH COUNT", "BR to HO", "BR to BR"
        ])
        self.table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        # Set edit triggers - only allow editing for specific columns
        self.table.setEditTriggers(QAbstractItemView.DoubleClicked | QAbstractItemView.SelectedClicked)

        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(6, QHeaderView.Stretch)  # Branch Name stretches

        # Style the table
        self.table.setStyleSheet("""
            QTableWidget {
                gridline-color: #000;
                font-size: 11px;
            }
            QHeaderView::section {
                background-color: #f0f0f0;
                font-weight: bold;
                border: 1px solid black;
                padding: 4px;
            }
        """)

        self.table.itemChanged.connect(self.on_item_changed)

        fund_tab_layout.addWidget(self.table)

        self._extra_space_row = -1

        button_layout = QHBoxLayout()
        self.export_btn = QPushButton("Export to Excel")
        self.export_btn.clicked.connect(self.export_to_excel)
        self.export_btn.setStyleSheet("""
            QPushButton {
                background-color: #217346;
                color: white;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #1a5c38;
            }
        """)

        self.print_btn = QPushButton("Print")
        self.print_btn.clicked.connect(self.print_table)

        button_layout.addWidget(self.export_btn)
        button_layout.addWidget(self.print_btn)

        fund_tab_layout.addLayout(button_layout)

        signature_layout = QHBoxLayout()
        signature_layout.addWidget(QLabel("Prepared by: ________________________"))
        signature_layout.addSpacing(40)
        signature_layout.addWidget(QLabel("Approved by: ________________________"))
        signature_layout.addStretch()
        fund_tab_layout.addLayout(signature_layout)

        self.tab_widget.addTab(fund_tab, "Fund Transfer")

        extra_tab = QWidget()
        extra_tab_layout = QVBoxLayout()
        extra_tab.setLayout(extra_tab_layout)

        input_group = QGroupBox("Post Extra Space Entry")
        input_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                font-size: 13px;
                border: 1px solid #ccc;
                border-radius: 6px;
                margin-top: 10px;
                padding-top: 18px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 12px;
                padding: 0 6px;
            }
        """)
        input_form = QHBoxLayout()
        input_group.setLayout(input_form)

        # Date picker
        input_form.addWidget(QLabel("Date:"))
        self.es_date_picker = QDateEdit()
        self.es_date_picker.setCalendarPopup(True)
        self.es_date_picker.setDate(QDate.currentDate())
        self.es_date_picker.setDisplayFormat("yyyy-MM-dd")
        self.es_date_picker.setFixedWidth(160)
        self.es_date_picker.setStyleSheet("""
            QDateEdit {
                padding: 6px 10px;
                border: 1px solid #aaa;
                border-radius: 4px;
                font-size: 12px;
            }
        """)
        input_form.addWidget(self.es_date_picker)

        input_form.addSpacing(16)

        # Amount input
        input_form.addWidget(QLabel("Amount:"))
        self.es_amount_input = QLineEdit()
        self.es_amount_input.setPlaceholderText("Enter amount…")
        self.es_amount_input.setFixedWidth(180)
        self.es_amount_input.setValidator(QDoubleValidator(0, 999999999.99, 2))
        self.es_amount_input.setStyleSheet("""
            QLineEdit {
                padding: 6px 10px;
                border: 1px solid #aaa;
                border-radius: 4px;
                font-size: 12px;
            }
        """)
        input_form.addWidget(self.es_amount_input)

        input_form.addSpacing(16)

        # Post button
        self.es_post_btn = QPushButton("Post")
        self.es_post_btn.setFixedWidth(110)
        self.es_post_btn.setCursor(Qt.PointingHandCursor)
        self.es_post_btn.clicked.connect(self._post_extra_space)
        self.es_post_btn.setStyleSheet("""
            QPushButton {
                background-color: #4472C4;
                color: white;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
                font-size: 12px;
            }
            QPushButton:hover { background-color: #3461a8; }
        """)
        input_form.addWidget(self.es_post_btn)
        input_form.addStretch()

        extra_tab_layout.addWidget(input_group)

        # ── History table ─────────────────────────────────────────────────
        history_label = QLabel("Extra Space History")
        history_label.setStyleSheet("font-weight: bold; font-size: 13px; margin-top: 8px;")
        extra_tab_layout.addWidget(history_label)

        self.es_history_table = QTableWidget()
        self.es_history_table.setColumnCount(3)
        self.es_history_table.setHorizontalHeaderLabels([
            "Date", "Amount", "Action"
        ])
        self.es_history_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.es_history_table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.es_history_table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.es_history_table.setAlternatingRowColors(True)
        self.es_history_table.setStyleSheet("""
            QTableWidget {
                gridline-color: #ddd;
                font-size: 11px;
            }
            QTableWidget::item:alternate {
                background-color: #f7f9fc;
            }
            QHeaderView::section {
                background-color: #4472C4;
                color: white;
                font-weight: bold;
                border: 1px solid #3461a8;
                padding: 6px;
            }
        """)
        extra_tab_layout.addWidget(self.es_history_table)

        self.tab_widget.addTab(extra_tab, "💰 Extra Space")

        # Load history when switching to Extra Space tab
        self.tab_widget.currentChanged.connect(self._on_tab_changed)

        self.load_corporations()
        self.load_os_list()
        self._ensure_extra_space_table()

    def _on_tab_changed(self, index):
        """Refresh extra-space history when switching to the Extra Space tab"""
        if index == 1:  # Extra Space tab
            self._load_extra_space_history()

    def on_report_type_changed(self):
        """Toggle between Corporation and OS mode"""
        report_type = self.report_type_selector.currentData()
        if report_type == "corporation":
            self.corp_label.setVisible(True)
            self.corp_selector.setVisible(True)
            self.os_label.setVisible(False)
            self.os_selector.setVisible(False)
        else:
            self.corp_label.setVisible(False)
            self.corp_selector.setVisible(False)
            self.os_label.setVisible(True)
            self.os_selector.setVisible(True)
        self.populate_table()

    def load_corporations(self):
        """Load unique corporations from the daily_reports_brand_a table"""
        self.corp_selector.clear()
        try:
            query = "SELECT DISTINCT corporation FROM daily_reports_brand_a ORDER BY corporation"
            corporations = db_manager.execute_query(query)

            if not corporations:
                print("No corporations found in database")
                return

            self.corp_selector.blockSignals(True)
            for corp in corporations:
                self.corp_selector.addItem(corp['corporation'])
            self.corp_selector.blockSignals(False)

        except Exception as e:
            self.corp_selector.blockSignals(False)
            print(f"Error loading corporations: {e}")
            QMessageBox.critical(self, "Database Error", f"Error connecting to database: {str(e)}")

    def load_os_list(self):

        self.os_selector.clear()
        try:
            query = """
                SELECT DISTINCT os_name FROM branches 
                WHERE os_name IS NOT NULL AND os_name != '' 
                ORDER BY os_name
            """
            os_list = db_manager.execute_query(query)

            if not os_list:
                print("No OS found in database")
                return

            self.os_selector.blockSignals(True)
            for os_row in os_list:
                os_name = os_row['os_name'] if isinstance(os_row, dict) else os_row[0]
                self.os_selector.addItem(os_name)
            self.os_selector.blockSignals(False)

        except Exception as e:
            self.os_selector.blockSignals(False)
            print(f"Error loading OS list: {e}")

    def _ensure_extra_space_table(self):

        try:
            db_manager.execute_query("""
                CREATE TABLE IF NOT EXISTS extra_space_fund_transfer (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    report_type VARCHAR(50) DEFAULT NULL,
                    filter_value VARCHAR(255) DEFAULT NULL,
                    report_date DATE NOT NULL,
                    amount DECIMAL(15,2) DEFAULT 0.00,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    UNIQUE KEY uq_extra_space_date (report_date)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
            """)
        except Exception as e:
            print(f"Extra space table create (may already exist): {e}")


        try:
            db_manager.execute_query(
                "ALTER TABLE extra_space_fund_transfer "
                "MODIFY COLUMN report_type VARCHAR(50) DEFAULT NULL, "
                "MODIFY COLUMN filter_value VARCHAR(255) DEFAULT NULL"
            )
        except Exception:
            pass
        try:
            existing = db_manager.execute_query(
                "SELECT INDEX_NAME FROM INFORMATION_SCHEMA.STATISTICS "
                "WHERE TABLE_SCHEMA = DATABASE() "
                "AND TABLE_NAME = 'extra_space_fund_transfer' "
                "AND INDEX_NAME IN ('uq_extra_space', 'uq_extra_space_date')"
            )
            existing_names = {r['INDEX_NAME'] for r in existing} if existing else set()

            if 'uq_extra_space' in existing_names:
                db_manager.execute_query(
                    "ALTER TABLE extra_space_fund_transfer DROP INDEX uq_extra_space"
                )

            if 'uq_extra_space_date' not in existing_names:
                db_manager.execute_query(
                    "ALTER TABLE extra_space_fund_transfer ADD UNIQUE KEY uq_extra_space_date (report_date)"
                )
        except Exception:
            pass

    def _load_extra_space_from_db(self):
        """Load extra space amount by date only.
        Single date → single entry; range mode → SUM of entries in range.
        """
        try:
            date_start, date_end = self.date_range_widget.get_date_range()
            is_range = self.date_range_widget.is_range_mode()

            if is_range and date_start != date_end:
                result = db_manager.execute_query(
                    "SELECT COALESCE(SUM(amount), 0) AS total_amount "
                    "FROM extra_space_fund_transfer WHERE report_date >= %s AND report_date <= %s",
                    (date_start, date_end)
                )
                if result and len(result) > 0:
                    return float(result[0].get('total_amount', 0) or 0)
            else:
                result = db_manager.execute_query(
                    "SELECT amount FROM extra_space_fund_transfer WHERE report_date = %s",
                    (date_start,)
                )
                if result and len(result) > 0:
                    return float(result[0].get('amount', 0) or 0)
        except Exception as e:
            print(f"Error loading extra space: {e}")
        return 0.0

    def _save_extra_space_to_db(self, amount):
        """Save extra space amount by date only"""
        try:
            date_start, _ = self.date_range_widget.get_date_range()
            db_manager.execute_query(
                """INSERT INTO extra_space_fund_transfer (report_date, amount)
                   VALUES (%s, %s)
                   ON DUPLICATE KEY UPDATE amount = VALUES(amount)""",
                (date_start, amount)
            )
        except Exception as e:
            print(f"Error saving extra space: {e}")

    def _post_extra_space(self):
        """Post an extra-space entry (date-based only)"""
        amount_text = self.es_amount_input.text().strip()
        if not amount_text:
            QMessageBox.warning(self, "Missing Amount",
                                "Please enter an amount.")
            return

        try:
            amount = float(amount_text)
        except ValueError:
            QMessageBox.warning(self, "Invalid Amount",
                                "Please enter a valid number.")
            return

        selected_date = self.es_date_picker.date().toString("yyyy-MM-dd")

        try:
            db_manager.execute_query(
                """INSERT INTO extra_space_fund_transfer (report_date, amount)
                   VALUES (%s, %s)
                   ON DUPLICATE KEY UPDATE amount = VALUES(amount)""",
                (selected_date, amount)
            )
            QMessageBox.information(self, "Posted",
                                    f"Extra Space entry posted successfully.\n"
                                    f"Date: {selected_date}\nAmount: {amount:,.2f}")
            self.es_amount_input.clear()
            self._load_extra_space_history()
            # Refresh fund transfer table
            self.populate_table()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to post entry: {str(e)}")

    def _delete_extra_space_entry(self, entry_id):
        """Delete an extra-space entry by id"""
        reply = QMessageBox.question(
            self, "Confirm Delete",
            "Are you sure you want to delete this extra space entry?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return
        try:
            db_manager.execute_query(
                "DELETE FROM extra_space_fund_transfer WHERE id = %s", (entry_id,)
            )
            self._load_extra_space_history()
            self.populate_table()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to delete entry: {str(e)}")

    def _load_extra_space_history(self):

        self.es_history_table.setRowCount(0)

        try:
            rows = db_manager.execute_query(
                """SELECT id, report_date, amount
                   FROM extra_space_fund_transfer
                   ORDER BY report_date DESC"""
            )

            if not rows:
                return

            for r in rows:
                row_idx = self.es_history_table.rowCount()
                self.es_history_table.insertRow(row_idx)

                date_str = str(r.get('report_date', ''))
                self.es_history_table.setItem(row_idx, 0, QTableWidgetItem(date_str))

                amt = float(r.get('amount', 0) or 0)
                amt_item = QTableWidgetItem(f"{amt:,.2f}")
                amt_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.es_history_table.setItem(row_idx, 1, amt_item)

                # Delete button
                del_btn = QPushButton("🗑 Delete")
                del_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #e74c3c;
                        color: white;
                        padding: 4px 10px;
                        border-radius: 3px;
                        font-size: 11px;
                    }
                    QPushButton:hover { background-color: #c0392b; }
                """)
                entry_id = r.get('id')
                del_btn.clicked.connect(lambda checked, eid=entry_id: self._delete_extra_space_entry(eid))
                self.es_history_table.setCellWidget(row_idx, 2, del_btn)

        except Exception as e:
            print(f"Error loading extra space history: {e}")

    def on_item_changed(self, item):
        row = item.row()
        col = item.column()

        first_item = self.table.item(row, 0)
        row_type = first_item.data(Qt.UserRole) if first_item else None

        if row_type in ['header', 'total', 'grand_total']:
            return

        if row_type == 'extra_space' and col == 9:
            try:
                val = float(item.text().replace(',', '')) if item.text() else 0.0
                self._save_extra_space_to_db(val)
            except ValueError:
                QMessageBox.warning(self, "Invalid Input", "Please enter a valid number.")
                item.setText("0.00")
            return

        if col not in [7, 10, 11]:
            return

        try:
            float(item.text())
        except ValueError:
            QMessageBox.warning(self, "Invalid Input", "Please enter a valid number.")
            item.setText("0.00")

    def populate_table(self):

        date_start, date_end = self.date_range_widget.get_date_range()
        is_range = self.date_range_widget.is_range_mode()
        report_type = self.report_type_selector.currentData()
        reg_filter = self.reg_filter_selector.currentData()

        # Build registration clause
        reg_clause = ""
        if reg_filter == "registered":
            reg_clause = "AND b.is_registered = 1"
        elif reg_filter == "not_registered":
            reg_clause = "AND (b.is_registered = 0 OR b.is_registered IS NULL)"


        if report_type == "corporation":
            filter_value = self.corp_selector.currentText()
            if not filter_value:
                return
            if is_range:
                query = f"""
                    SELECT b.name as branch,
                           COALESCE(b.area, 'UNASSIGNED') as area,
                           COALESCE(MAX(c.name), '') as corporation_name,
                           COALESCE(b.line_of_business, '') as line_of_business,
                           COALESCE(b.global_tag, '') as global_tag,
                           COALESCE(b.sunday, '') as sunday,
                           SUM(COALESCE(dr.cash_count, 0)) as cash_count,
                           SUM(COALESCE(cf.cash_float, 0)) as cash_float
                    FROM branches b
                    LEFT JOIN corporations c ON c.id = COALESCE(b.sub_corporation_id, b.corporation_id)
                    LEFT JOIN {self.daily_table} dr ON b.name COLLATE utf8mb4_general_ci = dr.branch COLLATE utf8mb4_general_ci
                        AND dr.corporation = %s
                        AND dr.date >= %s AND dr.date <= %s
                    LEFT JOIN cash_float_tbl cf ON b.name COLLATE utf8mb4_general_ci = cf.branch COLLATE utf8mb4_general_ci
                        AND dr.corporation COLLATE utf8mb4_general_ci = cf.corporation COLLATE utf8mb4_general_ci AND dr.date = cf.date
                    WHERE (b.corporation_id = (SELECT id FROM corporations WHERE name = %s)
                           OR b.sub_corporation_id = (SELECT id FROM corporations WHERE name = %s))
                    {reg_clause}
                    GROUP BY b.name, b.area, b.line_of_business, b.global_tag, b.sunday
                    ORDER BY COALESCE(b.area, 'ZZZZZ'), b.name
                """
                params = (filter_value, date_start, date_end, filter_value, filter_value)
            else:
                query = f"""
                    SELECT b.name as branch,
                           COALESCE(b.area, 'UNASSIGNED') as area,
                           COALESCE(MAX(c.name), '') as corporation_name,
                           COALESCE(b.line_of_business, '') as line_of_business,
                           COALESCE(b.global_tag, '') as global_tag,
                           COALESCE(b.sunday, '') as sunday,
                           COALESCE(SUM(dr.cash_count), 0) as cash_count,
                           COALESCE(SUM(cf.cash_float), 0) as cash_float
                    FROM branches b
                    LEFT JOIN corporations c ON c.id = COALESCE(b.sub_corporation_id, b.corporation_id)
                    LEFT JOIN {self.daily_table} dr ON b.name COLLATE utf8mb4_general_ci = dr.branch COLLATE utf8mb4_general_ci
                        AND dr.corporation = %s
                        AND dr.date = %s
                    LEFT JOIN cash_float_tbl cf ON b.name COLLATE utf8mb4_general_ci = cf.branch COLLATE utf8mb4_general_ci
                        AND dr.corporation COLLATE utf8mb4_general_ci = cf.corporation COLLATE utf8mb4_general_ci AND dr.date = cf.date
                    WHERE (b.corporation_id = (SELECT id FROM corporations WHERE name = %s)
                           OR b.sub_corporation_id = (SELECT id FROM corporations WHERE name = %s))
                    {reg_clause}
                    GROUP BY b.name, b.area, b.line_of_business, b.global_tag, b.sunday
                    ORDER BY COALESCE(b.area, 'ZZZZZ'), b.name
                """
                params = (filter_value, date_start, filter_value, filter_value)
        else:  # OS mode
            filter_value = self.os_selector.currentText()
            if not filter_value:
                return
            if is_range:
                query = f"""
                    SELECT b.name as branch,
                           COALESCE(b.area, 'UNASSIGNED') as area,
                           COALESCE(MAX(c.name), '') as corporation_name,
                           COALESCE(b.line_of_business, '') as line_of_business,
                           COALESCE(b.global_tag, '') as global_tag,
                           COALESCE(b.sunday, '') as sunday,
                           SUM(COALESCE(dr.cash_count, 0)) as cash_count,
                           SUM(COALESCE(cf.cash_float, 0)) as cash_float
                    FROM branches b
                    LEFT JOIN corporations c ON c.id = COALESCE(b.sub_corporation_id, b.corporation_id)
                    LEFT JOIN {self.daily_table} dr ON b.name COLLATE utf8mb4_general_ci = dr.branch COLLATE utf8mb4_general_ci
                        AND dr.date >= %s AND dr.date <= %s
                    LEFT JOIN cash_float_tbl cf ON b.name COLLATE utf8mb4_general_ci = cf.branch COLLATE utf8mb4_general_ci
                        AND dr.corporation COLLATE utf8mb4_general_ci = cf.corporation COLLATE utf8mb4_general_ci AND dr.date = cf.date
                    WHERE b.os_name = %s
                      {reg_clause}
                    GROUP BY b.name, b.area, b.line_of_business, b.global_tag, b.sunday
                    ORDER BY COALESCE(b.area, 'ZZZZZ'), b.name
                """
                params = (date_start, date_end, filter_value)
            else:
                query = f"""
                    SELECT b.name as branch,
                           COALESCE(b.area, 'UNASSIGNED') as area,
                           COALESCE(MAX(c.name), '') as corporation_name,
                           COALESCE(b.line_of_business, '') as line_of_business,
                           COALESCE(b.global_tag, '') as global_tag,
                           COALESCE(b.sunday, '') as sunday,
                           COALESCE(SUM(dr.cash_count), 0) as cash_count,
                           COALESCE(SUM(cf.cash_float), 0) as cash_float
                    FROM branches b
                    LEFT JOIN corporations c ON c.id = COALESCE(b.sub_corporation_id, b.corporation_id)
                    LEFT JOIN {self.daily_table} dr ON b.name COLLATE utf8mb4_general_ci = dr.branch COLLATE utf8mb4_general_ci
                        AND dr.date = %s
                    LEFT JOIN cash_float_tbl cf ON b.name COLLATE utf8mb4_general_ci = cf.branch COLLATE utf8mb4_general_ci
                        AND dr.corporation COLLATE utf8mb4_general_ci = cf.corporation COLLATE utf8mb4_general_ci AND dr.date = cf.date
                    WHERE b.os_name = %s
                      {reg_clause}
                    GROUP BY b.name, b.area, b.line_of_business, b.global_tag, b.sunday
                    ORDER BY COALESCE(b.area, 'ZZZZZ'), b.name
                """
                params = (date_start, filter_value)

        self._pending_is_range_mode = is_range and date_start != date_end
        _ds, _de, _ir = date_start, date_end, is_range

        def _fetch():
            rows = db_manager.execute_query(query, params) or []
            # Load extra-space in the same background thread
            es_val = 0.0
            try:
                if _ir and _ds != _de:
                    es = db_manager.execute_query(
                        "SELECT COALESCE(SUM(amount), 0) AS total_amount "
                        "FROM extra_space_fund_transfer WHERE report_date >= %s AND report_date <= %s",
                        (_ds, _de))
                    if es:
                        es_val = float(es[0].get('total_amount', 0) or 0)
                else:
                    es = db_manager.execute_query(
                        "SELECT amount FROM extra_space_fund_transfer WHERE report_date = %s",
                        (_ds,))
                    if es:
                        es_val = float(es[0].get('amount', 0) or 0)
            except Exception:
                pass
            return {'results': rows, 'extra_space': es_val}

        run_func_async(
            parent=self,
            func=_fetch,
            on_result=self._on_populate_result,
            on_error=self._on_populate_error,
            loading_message="\u23f3  Loading fund transfer data\u2026",
        )

    def _on_populate_error(self, err):
        print(f"Error loading fund transfer data: {err}")

    def _on_populate_result(self, data):
        results = data['results']
        extra_space_val = data['extra_space']
        is_range_mode = self._pending_is_range_mode

        try:
            # Temporarily disconnect signals
            self.table.itemChanged.disconnect()
        except TypeError:
            pass
        self.table.setRowCount(0)

        # Group data by area
        from collections import OrderedDict
        area_groups = OrderedDict()
        for row_data in results:
            area = row_data['area'] or 'UNASSIGNED'
            if area not in area_groups:
                area_groups[area] = []
            area_groups[area].append(row_data)

        row_idx = 0
        grand_total_cash_count = 0.0
        continuous_branch_num = 1  # Continuous numbering across all areas

        for area_name, branches in area_groups.items():
            # Add area header row (centered across all columns)
            self.table.insertRow(row_idx)
            # Put header text in center column for visual centering
            for col in range(12):
                if col == 6:  
                    header_item = QTableWidgetItem(f"{area_name} AREA")
                    header_font = QFont()
                    header_font.setBold(True)
                    header_font.setItalic(True)
                    header_item.setFont(header_font)
                else:
                    header_item = QTableWidgetItem("")
                header_item.setData(Qt.UserRole, 'header')
                header_item.setTextAlignment(Qt.AlignCenter)
                header_item.setFlags(header_item.flags() & ~Qt.ItemIsEditable)
                self.table.setItem(row_idx, col, header_item)
            row_idx += 1

            area_total_cash_count = 0.0

            for row_data in branches:
                branch_name = row_data['branch']
                area = row_data['area'] or ''
                corporation_name = row_data.get('corporation_name', '') or ''
                lob = row_data['line_of_business'] or ''
                global_tag = row_data['global_tag'] or ''
                sunday = row_data['sunday'] or ''
                if sunday == 'NO':
                    sunday = 'NO SUNDAY'
                cash_count = float(row_data['cash_count'] or 0)
                cash_float = float(row_data.get('cash_float', 0) or 0)

                self.table.insertRow(row_idx)

                # Col 0: Area
                area_item = QTableWidgetItem(area)
                area_item.setTextAlignment(Qt.AlignCenter)
                area_item.setFlags(area_item.flags() & ~Qt.ItemIsEditable)
                self.table.setItem(row_idx, 0, area_item)

                # Col 1: # (continuous row number)
                num_item = QTableWidgetItem(str(continuous_branch_num))
                num_item.setTextAlignment(Qt.AlignCenter)
                num_item.setFlags(num_item.flags() & ~Qt.ItemIsEditable)
                self.table.setItem(row_idx, 1, num_item)

                # Col 2: Corporation
                corp_item = QTableWidgetItem(corporation_name)
                corp_item.setTextAlignment(Qt.AlignCenter)
                corp_item.setFlags(corp_item.flags() & ~Qt.ItemIsEditable)
                self.table.setItem(row_idx, 2, corp_item)

                # Col 3: Line of Business
                lob_item = QTableWidgetItem(lob)
                lob_item.setTextAlignment(Qt.AlignCenter)
                lob_item.setFlags(lob_item.flags() & ~Qt.ItemIsEditable)
                self.table.setItem(row_idx, 3, lob_item)

                # Col 4: Global
                global_item = QTableWidgetItem(global_tag)
                global_item.setTextAlignment(Qt.AlignCenter)
                global_item.setFlags(global_item.flags() & ~Qt.ItemIsEditable)
                self.table.setItem(row_idx, 4, global_item)

                # Col 5: Sunday
                sunday_item = QTableWidgetItem(sunday)
                sunday_item.setTextAlignment(Qt.AlignCenter)
                sunday_item.setFlags(sunday_item.flags() & ~Qt.ItemIsEditable)
                self.table.setItem(row_idx, 5, sunday_item)

                # Col 6: Branch Name
                branch_item = QTableWidgetItem(branch_name)
                branch_item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                branch_item.setFlags(branch_item.flags() & ~Qt.ItemIsEditable)
                self.table.setItem(row_idx, 6, branch_item)

                # Col 7: Invty (editable)
                invty_item = QTableWidgetItem("")
                invty_item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(row_idx, 7, invty_item)

                # Col 8: Cash Float (read-only) - show "-" if zero
                if cash_float == 0:
                    cash_float_item = QTableWidgetItem("-")
                else:
                    cash_float_item = QTableWidgetItem(f"{cash_float:,.2f}")
                cash_float_item.setTextAlignment(Qt.AlignCenter)
                cash_float_item.setFlags(cash_float_item.flags() & ~Qt.ItemIsEditable)
                self.table.setItem(row_idx, 8, cash_float_item)

                # Col 9: Cash Count (read-only)
                cash_item = QTableWidgetItem(f"{cash_count:,.2f}")
                cash_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                cash_item.setFlags(cash_item.flags() & ~Qt.ItemIsEditable)
                self.table.setItem(row_idx, 9, cash_item)

                # Col 10: BR to HO (editable)
                br_ho_item = QTableWidgetItem("")
                br_ho_item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(row_idx, 10, br_ho_item)

                # Col 11: BR to BR (editable)
                br_br_item = QTableWidgetItem("")
                br_br_item.setTextAlignment(Qt.AlignCenter)
                self.table.setItem(row_idx, 11, br_br_item)

                area_total_cash_count += cash_count
                continuous_branch_num += 1
                row_idx += 1

            # Add area total row
            self.table.insertRow(row_idx)
            total_font = QFont()
            total_font.setBold(True)

            for col in range(12):
                if col == 6:
                    total_item = QTableWidgetItem(f"TOTAL {area_name}")
                    total_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                elif col == 9:
                    total_item = QTableWidgetItem(f"{area_total_cash_count:,.2f}")
                    total_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                else:
                    total_item = QTableWidgetItem("")
                    total_item.setTextAlignment(Qt.AlignCenter)
                total_item.setFont(total_font)
                total_item.setFlags(total_item.flags() & ~Qt.ItemIsEditable)
                total_item.setData(Qt.UserRole, 'total')
                total_item.setBackground(QBrush(QColor("#E9ECEF")))
                self.table.setItem(row_idx, col, total_item)
            
            grand_total_cash_count += area_total_cash_count
            row_idx += 1

        # ── EXTRA SPACE row (NOT included in grand total) ───────────
        self.table.insertRow(row_idx)
        self._extra_space_row = row_idx
        es_font = QFont()
        es_font.setBold(True)

        for col in range(12):
            if col == 6:
                label_text = "EXTRA SPACE"
                if is_range_mode:
                    label_text = "EXTRA SPACE (Total)"
                es_item = QTableWidgetItem(label_text)
                es_item.setTextAlignment(Qt.AlignCenter)
                es_item.setFont(es_font)
                es_item.setFlags(es_item.flags() & ~Qt.ItemIsEditable)
            elif col == 9:
                if extra_space_val > 0:
                    es_item = QTableWidgetItem(f"{extra_space_val:,.2f}")
                else:
                    es_item = QTableWidgetItem("")
                es_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                es_item.setFont(es_font)
               
                if is_range_mode:
                    es_item.setFlags(es_item.flags() & ~Qt.ItemIsEditable)
            else:
                es_item = QTableWidgetItem("")
                es_item.setTextAlignment(Qt.AlignCenter)
                es_item.setFlags(es_item.flags() & ~Qt.ItemIsEditable)
            es_item.setData(Qt.UserRole, 'extra_space')
            es_item.setBackground(QBrush(QColor("#FFF8E1")))
            self.table.setItem(row_idx, col, es_item)
        row_idx += 1

    
        self.table.insertRow(row_idx)
        for col in range(12):
            spacer_item = QTableWidgetItem("")
            spacer_item.setFlags(spacer_item.flags() & ~Qt.ItemIsEditable)
            spacer_item.setData(Qt.UserRole, 'spacer')
            self.table.setItem(row_idx, col, spacer_item)
        row_idx += 1


        self.table.insertRow(row_idx)
        gt_font = QFont()
        gt_font.setBold(True)
        gt_font.setPointSize(gt_font.pointSize() + 1)

        for col in range(12):
            if col == 6:
                gt_item = QTableWidgetItem("GRAND TOTAL")
                gt_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            elif col == 9:
                gt_item = QTableWidgetItem(f"{grand_total_cash_count:,.2f}")
                gt_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            else:
                gt_item = QTableWidgetItem("")
                gt_item.setTextAlignment(Qt.AlignCenter)
            gt_item.setFont(gt_font)
            gt_item.setFlags(gt_item.flags() & ~Qt.ItemIsEditable)
            gt_item.setData(Qt.UserRole, 'grand_total')
            gt_item.setBackground(QBrush(QColor("#4472C4")))
            gt_item.setForeground(QBrush(QColor("#FFFFFF")))
            self.table.setItem(row_idx, col, gt_item)
        row_idx += 1

        # Reconnect signal
        self.table.itemChanged.connect(self.on_item_changed)

    def export_to_excel(self):
        """Export table data to Excel with professional formatting matching the screenshot"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from datetime import datetime
        except ImportError:
            QMessageBox.critical(
                self,
                "Missing Dependency",
                "The openpyxl package is required to export to Excel.\nInstall with: pip install openpyxl"
            )
            return
        
        report_type = self.report_type_selector.currentData()
        if report_type == "corporation":
            filter_name = self.corp_selector.currentText()
        else:
            filter_name = self.os_selector.currentText()
        
        date_start, date_end = self.date_range_widget.get_date_range()
        date = date_start if date_start == date_end else f"{date_start}_to_{date_end}"
        
        rows = self.table.rowCount()
        cols = self.table.columnCount()
        
        if rows == 0:
            QMessageBox.warning(self, "No Data", "No data to export.")
            return
        
        # File dialog for save location
        default_filename = f"Fund_Transfer_Report_{filter_name}_{date}.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Excel File",
            default_filename,
            "Excel Files (*.xlsx);;All Files (*)"
        )
        
        if not file_path:
            return  # User cancelled
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Fund Transfer"
            
            # Define styles
            title_font = Font(bold=True, size=16)
            subtitle_font = Font(bold=True, size=12)
            date_font = Font(size=11)
            header_font = Font(bold=True, size=11, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            area_header_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
            area_header_font = Font(bold=True, size=11)
            total_fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
            total_font = Font(bold=True)
            extra_space_fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
            extra_space_font = Font(bold=True)
            grand_total_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            grand_total_font = Font(bold=True, size=12, color="FFFFFF")
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # ------ Row 1: Title ------
            ws.merge_cells('A1:L1')
            ws['A1'] = "FUND TRANSFER"
            ws['A1'].font = title_font
            ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
            ws.row_dimensions[1].height = 20
            
            # ------ Row 2: Group/Filter Name ------
            ws.merge_cells('A2:L2')
            ws['A2'] = f"GROUP {filter_name.upper()}"
            ws['A2'].font = subtitle_font
            ws['A2'].alignment = Alignment(horizontal='left', vertical='center')
            
            # ------ Row 3: Date ------
            ws.merge_cells('A3:L3')
            try:
                date_obj = datetime.strptime(str(date_start), '%Y-%m-%d')
                formatted_date = date_obj.strftime('%A, %B %d, %Y')
            except:
                formatted_date = str(date_start)
            ws['A3'] = formatted_date
            ws['A3'].font = date_font
            ws['A3'].alignment = Alignment(horizontal='left', vertical='center')
            
            # ------ Column Headers (Row 5) ------
            header_row = 5
            
            # Header values for all columns
            header_values = {
                'A': 'AREA',
                'B': '#',
                'C': 'CORPORATION',
                'D': 'LOB',
                'E': 'GLOBAL',
                'F': 'SUNDAY',
                'G': 'Branch Name',
                'H': 'Invty',
                'I': 'CASH FLOAT',
                'J': 'CASH COUNT',
                'K': 'BR to HO',
                'L': 'BR to BR',
            }
            
            col_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']
            for col_letter in col_letters:
                cell = ws[f'{col_letter}{header_row}']
                cell.value = header_values.get(col_letter, '')
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
            
            # Add sub-header for FUND TRANSFER columns (row 6)
            sub_header_row = 6
            for col in col_letters:
                cell = ws[f'{col}{sub_header_row}']
                if col == 'K':
                    cell.value = 'BR to HO'
                elif col == 'L':
                    cell.value = 'BR to BR'
                else:
                    cell.value = ''
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            
            # Data rows
            excel_row = sub_header_row + 1
            for row in range(rows):
                first_item = self.table.item(row, 0)
                row_type = first_item.data(Qt.UserRole) if first_item else None
                
                # Handle area header rows - merge cells and set value only in first cell
                if row_type == 'header':
                    ws.merge_cells(f'A{excel_row}:L{excel_row}')
                    first_item_cell = self.table.item(row, 0)
                    header_text = first_item_cell.text() if first_item_cell else ''
                    
                    # Only set value in A (first cell of merged range)
                    cell_a = ws[f'A{excel_row}']
                    cell_a.value = header_text
                    cell_a.fill = area_header_fill
                    cell_a.font = area_header_font
                    cell_a.alignment = Alignment(horizontal='center', vertical='center')
                    cell_a.border = thin_border
                    
                    # Apply styling to remaining merged cells
                    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
                        cell = ws[f'{col_letter}{excel_row}']
                        cell.fill = area_header_fill
                        cell.font = area_header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = thin_border
                else:
                    # Regular data rows
                    for col in range(cols):
                        cell = ws.cell(row=excel_row, column=col+1)
                        item = self.table.item(row, col)
                        
                        if item:
                            text = item.text()
                            
                            # Handle numeric columns
                            if col in [9, 10, 11]:  # Cash Count, BR to HO, BR to BR
                                try:
                                    clean_text = text.replace(',', '').replace('-', '')
                                    if clean_text:
                                        cell.value = float(clean_text)
                                        cell.number_format = '#,##0.00'
                                    else:
                                        cell.value = ''
                                except ValueError:
                                    cell.value = text
                            else:
                                cell.value = text
                            
                            # Apply formatting based on row type
                            if row_type == 'total':  # Area total
                                cell.fill = total_fill
                                cell.font = total_font
                                if col == 9:  # Cash Count column
                                    cell.alignment = Alignment(horizontal='right')
                                else:
                                    cell.alignment = Alignment(horizontal='center')
                            elif row_type == 'extra_space':
                                cell.fill = extra_space_fill
                                cell.font = extra_space_font
                                if col == 9:
                                    cell.alignment = Alignment(horizontal='right')
                                else:
                                    cell.alignment = Alignment(horizontal='center')
                            elif row_type == 'grand_total':
                                cell.fill = grand_total_fill
                                cell.font = grand_total_font
                                if col == 9:
                                    cell.alignment = Alignment(horizontal='right')
                                else:
                                    cell.alignment = Alignment(horizontal='center')
                            elif row_type == 'spacer':
                                cell.alignment = Alignment(horizontal='center')
                            else:
                                if col == 9 or col == 10 or col == 11:
                                    cell.alignment = Alignment(horizontal='right')
                                elif col == 6:  # Branch name
                                    cell.alignment = Alignment(horizontal='left')
                                else:
                                    cell.alignment = Alignment(horizontal='center')
                            
                            cell.border = thin_border
                        else:
                            cell.border = thin_border
                
                excel_row += 1
            
            # Set column widths
            ws.column_dimensions['A'].width = 18  # AREA
            ws.column_dimensions['B'].width = 5   # #
            ws.column_dimensions['C'].width = 18  # CORPORATION
            ws.column_dimensions['D'].width = 12  # LOB
            ws.column_dimensions['E'].width = 12  # GLOBAL
            ws.column_dimensions['F'].width = 12  # SUNDAY
            ws.column_dimensions['G'].width = 20  # Branch Name
            ws.column_dimensions['H'].width = 12  # Invty
            ws.column_dimensions['I'].width = 15  # Cash Float
            ws.column_dimensions['J'].width = 15  # Cash Count
            ws.column_dimensions['K'].width = 12  # BR to HO
            ws.column_dimensions['L'].width = 12  # BR to BR
            
            wb.save(file_path)
            QMessageBox.information(self, "Export Successful", f"Report exported to:\n{file_path}")
            
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Error exporting to Excel: {str(e)}")

    def print_table(self):
        """Print the table"""
        try:
            if self.table.rowCount() == 0:
                QMessageBox.warning(self, "No Data", "No data to print.")
                return

            report_type = self.report_type_selector.currentData()
            if report_type == "corporation":
                filter_name = self.corp_selector.currentText()
                filter_label = "Corporation"
            else:
                filter_name = self.os_selector.currentText()
                filter_label = "OS"

            date_start, date_end = self.date_range_widget.get_date_range()
            date_display = date_start if date_start == date_end else f"{date_start} to {date_end}"

            doc = QTextDocument()
            html = "<h2>Fund Transfer Report</h2>"
            html += f"<p><b>{filter_label}:</b> {filter_name}<br>"
            html += f"<b>Date:</b> {date_display}</p>"
            html += "<table border='1' cellspacing='0' cellpadding='4' style='border-collapse: collapse; font-size: 10px;'>"

            # Headers
            html += "<tr style='background-color: #4472C4; color: white;'>"
            for col in range(self.table.columnCount()):
                header_item = self.table.horizontalHeaderItem(col)
                header_text = header_item.text() if header_item else ""
                html += f"<th style='padding: 6px; text-align: center;'>{header_text}</th>"
            html += "</tr>"

            # Data rows
            for row in range(self.table.rowCount()):
                first_item = self.table.item(row, 0)
                row_type = first_item.data(Qt.UserRole) if first_item else None
                
                if row_type == 'header':
                    html += "<tr style='background-color: #d4edda;'>"
                elif row_type == 'total':
                    html += "<tr style='background-color: #E9ECEF; font-weight: bold;'>"
                elif row_type == 'extra_space':
                    html += "<tr style='background-color: #FFF8E1; font-weight: bold;'>"
                elif row_type == 'grand_total':
                    html += "<tr style='background-color: #4472C4; color: white; font-weight: bold; font-size: 12px;'>"
                elif row_type == 'spacer':
                    html += "<tr>"
                else:
                    html += "<tr>"
                
                for col in range(self.table.columnCount()):
                    item = self.table.item(row, col)
                    cell_text = item.text() if item else ""
                    
                    if row_type == 'header':
                        html += f"<td style='padding: 6px; text-align: center; font-weight: bold; font-style: italic;'>{cell_text}</td>"
                    elif row_type == 'total':
                        html += f"<td style='padding: 6px; text-align: center; font-weight: bold;'>{cell_text}</td>"
                    elif row_type == 'extra_space':
                        html += f"<td style='padding: 6px; text-align: center; font-weight: bold;'>{cell_text}</td>"
                    elif row_type == 'grand_total':
                        html += f"<td style='padding: 8px; text-align: center; font-weight: bold; color: white;'>{cell_text}</td>"
                    else:
                        html += f"<td style='padding: 6px; text-align: center;'>{cell_text}</td>"
                html += "</tr>"

            html += "</table>"

            # Add signature lines to HTML
            html += "<br><br>"
            html += "<p>Prepared by: ________________________</p>"
            html += "<p>Approved by: ________________________</p>"

            doc.setHtml(html)

            printer = QPrinter()
            dialog = QPrintDialog(printer, self)
            if dialog.exec_() == QPrintDialog.Accepted:
                doc.print_(printer)

        except Exception as e:
            QMessageBox.critical(self, "Print Error", f"Error printing: {str(e)}")
