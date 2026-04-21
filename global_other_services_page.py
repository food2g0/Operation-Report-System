from PyQt5.QtWidgets import (
    QWidget, QLabel, QVBoxLayout, QHBoxLayout, QComboBox, QTableWidget,
    QTableWidgetItem, QDateEdit, QMessageBox, QHeaderView, QSizePolicy,
    QPushButton, QScrollArea, QFrame, QFileDialog, QApplication
)
from PyQt5.QtCore import Qt, QDate
from PyQt5.QtGui import QFont, QColor, QBrush, QPainter, QTextDocument
from PyQt5.QtPrintSupport import QPrinter, QPrintDialog
from db_connect_pooled import db_manager
from db_worker import run_query_async
from date_range_widget import DateRangeWidget

COLUMN_GROUPS = [
    ("GCASH OUT",         [("Lotes", ["gcash_out_lotes"], True),
                           ("Capital", ["gcash_out"], False)]),
    ("MONEYGRAM",         [("Lotes", ["moneygram_lotes"], True),
                           ("Capital", ["moneygram"], False)]),
    ("TRANSFAST",         [("Lotes", ["transfast_lotes"], True),
                           ("Capital", ["transfast"], False)]),
    ("RIA",               [("Lotes", ["ria_lotes"], True),
                           ("Capital", ["ria"], False)]),
    ("SMART MONEY OUT",   [("Lotes", ["smart_money_out_lotes"], True),
                           ("Capital", ["smart_money_out"], False)]),
    ("GCASH PADALA",      [("Lotes", ["gcash_padala_lotes"], True),
                           ("Capital", ["gcash_padala"], False)]),
    ("ABRA OUT",          [("Lotes", ["abra_out_lotes"], True),
                           ("Capital", ["abra_out"], False)]),
    ("REMITLY",           [("Lotes", ["remitly_lotes"], True),
                           ("Capital", ["remitly"], False)]),
    ("PAL PAY CASH OUT",  [("Lotes", ["pal_pay_cash_out_lotes"], True),
                           ("Capital", ["pal_pay_cash_out"], False)]),
    ("MC OUT",            [("Lotes", ["mc_out_lotes"], True),
                           ("Capital", ["mc_out"], False)]),
]

GROUP_COLORS = {
    "GCASH OUT":        QColor("#1abc9c"),
    "MONEYGRAM":        QColor("#2c3e50"),
    "TRANSFAST":        QColor("#34495e"),
    "RIA":              QColor("#e74c3c"),
    "SMART MONEY OUT":  QColor("#117864"),
    "GCASH PADALA":     QColor("#148f77"),
    "ABRA OUT":         QColor("#1a5276"),
    "REMITLY":          QColor("#7d3c98"),
    "PAL PAY CASH OUT": QColor("#196f3d"),
    "MC OUT":           QColor("#2980b9"),
}

TABLE_NAME = "global_other_services_tbl"


def _build_columns(groups=None):
    """Return (headers, col_meta) flat list from group definitions."""
    if groups is None:
        groups = COLUMN_GROUPS
    headers = [""]  # Branch
    col_meta = []
    for group_name, subs in groups:
        for sub_header, db_cols, is_lotes in subs:
            headers.append(sub_header.upper())
            col_meta.append({
                "group": group_name,
                "sub": sub_header,
                "db_cols": db_cols,
                "is_lotes": is_lotes,
            })
    return headers, col_meta


# ── Custom headers (reuse pattern from daily_transaction_page) ────────────

class _ColoredHeader(QHeaderView):
    def __init__(self, orientation, parent=None):
        super().__init__(orientation, parent)
        self.colors = {}
        self.setFont(QFont("", 8, QFont.Bold))

    def paintSection(self, painter, rect, logicalIndex):
        painter.save()
        bg = self.colors.get(logicalIndex, QColor("#495057"))
        painter.fillRect(rect, bg)
        pen = painter.pen(); pen.setColor(QColor("#333")); painter.setPen(pen)
        painter.drawRect(rect.adjusted(0, 0, -1, -1))
        painter.setPen(QColor("white"))
        painter.setFont(QFont("", 8, QFont.Bold))
        text = self.model().headerData(logicalIndex, self.orientation(), Qt.DisplayRole)
        painter.drawText(rect, Qt.AlignCenter | Qt.TextWordWrap, str(text) if text else "")
        painter.restore()


class _MergedGroupHeader(QHeaderView):
    def __init__(self, orientation, parent=None, groups=None):
        super().__init__(orientation, parent)
        self.colors = {}
        self._merge_info = {}
        self._col_to_group = {}
        if groups:
            col = 1
            for group_name, subs in groups:
                span = len(subs)
                self._merge_info[col] = (group_name, span)
                for i in range(span):
                    self._col_to_group[col + i] = col
                col += span

    def paintSection(self, painter, rect, logicalIndex):
        painter.save()
        bg = self.colors.get(logicalIndex, QColor("#495057"))
        if logicalIndex in self._col_to_group:
            start = self._col_to_group[logicalIndex]
            name, span = self._merge_info.get(start, ("", 1))
            if logicalIndex == start:
                merged = rect
                for i in range(1, span):
                    merged = merged.adjusted(0, 0, self.sectionSize(start + i), 0)
                painter.fillRect(merged, bg)
                pen = painter.pen(); pen.setColor(QColor("#333")); painter.setPen(pen)
                painter.drawRect(merged.adjusted(0, 0, -1, -1))
                painter.setPen(QColor("white"))
                painter.setFont(QFont("", 9, QFont.Bold))
                painter.drawText(merged, Qt.AlignCenter, name)
        else:
            painter.fillRect(rect, bg)
            pen = painter.pen(); pen.setColor(QColor("#333")); painter.setPen(pen)
            painter.drawRect(rect.adjusted(0, 0, -1, -1))
            painter.setPen(QColor("white"))
            painter.setFont(QFont("", 9, QFont.Bold))
            text = self.model().headerData(logicalIndex, self.orientation(), Qt.DisplayRole)
            painter.drawText(rect, Qt.AlignCenter, str(text) if text else "")
        painter.restore()


# ══════════════════════════════════════════════════════════════════════════
class GlobalOtherServicesPage(QWidget):
    """Global Other Services — Brand A, global-tagged branches only."""

    def __init__(self):
        super().__init__()
        self._is_loading = False

        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(8)

        self._ensure_table()
        self._build_controls(layout)
        self._build_report_header(layout)
        self._build_table(layout)
        self._build_buttons(layout)
        self.load_corporations()

    # ── Ensure DB table ───────────────────────────────────────────────────
    def _ensure_table(self):
        try:
            db_manager.execute_query(f"""
                CREATE TABLE IF NOT EXISTS {TABLE_NAME} (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    branch VARCHAR(255) NOT NULL,
                    corporation VARCHAR(255) NOT NULL,
                    date DATE NOT NULL,
                    gcash_out_lotes INT DEFAULT 0,
                    gcash_out DECIMAL(15,2) DEFAULT 0.00,
                    moneygram_lotes INT DEFAULT 0,
                    moneygram DECIMAL(15,2) DEFAULT 0.00,
                    transfast_lotes INT DEFAULT 0,
                    transfast DECIMAL(15,2) DEFAULT 0.00,
                    ria_lotes INT DEFAULT 0,
                    ria DECIMAL(15,2) DEFAULT 0.00,
                    smart_money_out_lotes INT DEFAULT 0,
                    smart_money_out DECIMAL(15,2) DEFAULT 0.00,
                    gcash_padala_lotes INT DEFAULT 0,
                    gcash_padala DECIMAL(15,2) DEFAULT 0.00,
                    abra_out_lotes INT DEFAULT 0,
                    abra_out DECIMAL(15,2) DEFAULT 0.00,
                    remitly_lotes INT DEFAULT 0,
                    remitly DECIMAL(15,2) DEFAULT 0.00,
                    pal_pay_cash_out_lotes INT DEFAULT 0,
                    pal_pay_cash_out DECIMAL(15,2) DEFAULT 0.00,
                    mc_out_lotes INT DEFAULT 0,
                    mc_out DECIMAL(15,2) DEFAULT 0.00,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    UNIQUE KEY uq_branch_date (branch, date)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
            """)
        except Exception as e:
            print(f"Global Other Services table setup: {e}")

    # ── Report header ─────────────────────────────────────────────────────
    def _build_report_header(self, parent):
        frame = QFrame()
        frame.setStyleSheet("""
            QFrame { background: white; border: 1px solid #dee2e6;
                     border-radius: 5px; padding: 8px 15px; }
        """)
        h = QVBoxLayout(frame)
        h.setContentsMargins(15, 10, 15, 10)
        h.setSpacing(2)

        self.report_title_label = QLabel("Global Other Services")
        self.report_title_label.setFont(QFont("Arial", 14, QFont.Bold))
        self.report_title_label.setStyleSheet("color: #333; border: none;")
        h.addWidget(self.report_title_label)

        self.report_corp_label = QLabel("Corporation: ")
        self.report_corp_label.setFont(QFont("Arial", 11, QFont.Bold))
        self.report_corp_label.setStyleSheet("color: #555; border: none;")
        h.addWidget(self.report_corp_label)

        self.report_date_label = QLabel("Date: ")
        self.report_date_label.setFont(QFont("Arial", 11, QFont.Bold))
        self.report_date_label.setStyleSheet("color: #555; border: none;")
        h.addWidget(self.report_date_label)

        self.report_os_label = QLabel("Operation Supervisor: ")
        self.report_os_label.setFont(QFont("Arial", 11, QFont.Bold))
        self.report_os_label.setStyleSheet("color: #555; border: none;")
        h.addWidget(self.report_os_label)

        parent.addWidget(frame, 0)

    def _update_report_header(self):
        group_name = self.group_selector.currentText() if hasattr(self, 'group_selector') else ""
        self.report_corp_label.setText(f"Group: {group_name}")
        ds, de = self.date_range_widget.get_date_range()
        date_str = ds if ds == de else f"{ds} to {de}"
        self.report_date_label.setText(f"Date: {date_str}")
        self.report_os_label.setText(f"Group: {group_name or 'All'}")

    # ── Controls ──────────────────────────────────────────────────────────
    def _build_controls(self, parent):
        frame = QFrame()
        frame.setStyleSheet("""
            QFrame { background:#f8f9fa; border:1px solid #dee2e6;
                     border-radius:5px; padding:10px; }
        """)
        row = QHBoxLayout(frame)
        row.setSpacing(15)

        row.addWidget(self._bold("Group:"))
        self.group_selector = self._combo(220)
        self.group_selector.currentTextChanged.connect(self.populate_table)
        row.addWidget(self.group_selector)

        row.addSpacing(20)
        self.date_range_widget = DateRangeWidget()
        self.date_range_widget.dateRangeChanged.connect(self.populate_table)
        row.addWidget(self.date_range_widget)

        row.addSpacing(20)
        row.addWidget(self._bold("Status:"))
        self.reg_filter_selector = self._combo(150)
        self.reg_filter_selector.addItem("Registered Only", "registered")
        self.reg_filter_selector.addItem("Not Registered", "not_registered")
        self.reg_filter_selector.addItem("All Branches", "all")
        self.reg_filter_selector.currentIndexChanged.connect(self.populate_table)
        row.addWidget(self.reg_filter_selector)

        # Keep corp_selector and os_filter_selector for backward compat but hidden
        self.corp_selector = self._combo(220)
        self.corp_selector.setVisible(False)
        self.os_filter_selector = self._combo(180)
        self.os_filter_selector.addItem("All (by Corporation)", None)
        self.os_filter_selector.setVisible(False)

        row.addStretch()
        parent.addWidget(frame, 0)

    # ── Table ─────────────────────────────────────────────────────────────
    def _build_table(self, parent):
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        scroll.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        self.table = QTableWidget()
        self.table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.table.setMinimumHeight(300)
        self.table.verticalHeader().setDefaultSectionSize(32)
        self.table.setAlternatingRowColors(False)

        headers, _ = _build_columns()
        self.table.setColumnCount(len(headers))
        self.table.setColumnWidth(0, 160)
        for i in range(1, len(headers)):
            self.table.setColumnWidth(i, 120)
        hdr = self.table.horizontalHeader()
        hdr.setStretchLastSection(False)
        for i in range(len(headers)):
            hdr.setSectionResizeMode(i, QHeaderView.Interactive)
        hdr.setMinimumSectionSize(55)

        self._apply_header_colors(COLUMN_GROUPS)
        self._style_table()

        scroll.setWidget(self.table)
        parent.addWidget(scroll, 1)

    def _apply_header_colors(self, groups):
        headers, _ = _build_columns(groups)
        group_headers = ["Branch"]
        for gn, subs in groups:
            group_headers.append(gn)
            for _ in range(len(subs) - 1):
                group_headers.append("")

        ch = _MergedGroupHeader(Qt.Horizontal, self.table, groups)
        ch.colors[0] = QColor("#495057")
        col = 1
        for gn, subs in groups:
            base = GROUP_COLORS.get(gn, QColor("#495057"))
            for _ in subs:
                ch.colors[col] = base
                col += 1
        ch.setDefaultAlignment(Qt.AlignCenter)
        ch.setMinimumSectionSize(55)
        ch.setDefaultSectionSize(120)
        ch.setStretchLastSection(False)
        self.table.setHorizontalHeader(ch)
        self.table.setHorizontalHeaderLabels(group_headers)
        ch.setVisible(True)
        ch.setFixedHeight(32)
        for i in range(len(group_headers)):
            ch.setSectionResizeMode(i, QHeaderView.Interactive)

        self._current_groups = groups
        self._sub_headers = headers

    def _create_sub_header_row(self, groups):
        if self.table.rowCount() == 0:
            return
        self.table.clearSpans()
        self.table.setRowHeight(0, 28)
        headers, _ = _build_columns(groups)

        item = QTableWidgetItem("")
        item.setTextAlignment(Qt.AlignCenter)
        item.setBackground(QColor("#495057"))
        item.setForeground(QColor("white"))
        f = item.font(); f.setBold(True); f.setPointSize(9); item.setFont(f)
        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
        self.table.setItem(0, 0, item)

        col = 1
        for gn, subs in groups:
            gc = GROUP_COLORS.get(gn, QColor("#495057"))
            for sub_header, _, _ in subs:
                ci = QTableWidgetItem(sub_header.upper())
                ci.setTextAlignment(Qt.AlignCenter)
                ci.setBackground(gc)
                ci.setForeground(QColor("white"))
                f2 = ci.font(); f2.setBold(True); f2.setPointSize(9); ci.setFont(f2)
                ci.setFlags(ci.flags() & ~Qt.ItemIsEditable)
                self.table.setItem(0, col, ci)
                col += 1

    def _style_table(self):
        self.table.setStyleSheet("""
            QTableWidget {
                gridline-color:#d0d0d0; border:1px solid #c0c0c0;
                font-size:11px; selection-background-color:#e3f2fd;
                background-color: white;
            }
            QTableWidget::item:selected { background-color:#e3f2fd; color:black; }
            QHeaderView::section {
                padding:6px 2px; font-weight:bold; font-size:9px;
                border:1px solid #666; color:white;
            }
        """)

    # ── Buttons ───────────────────────────────────────────────────────────
    def _build_buttons(self, parent):
        frame = QFrame()
        frame.setFixedHeight(70)
        lay = QHBoxLayout(frame)

        def _btn(text, c1, c2, c3):
            b = QPushButton(text)
            b.setStyleSheet(f"""
                QPushButton{{padding:12px 24px;border:none;border-radius:8px;
                    background:{c1};color:white;font-weight:bold;font-size:12px;
                    min-width:130px;min-height:40px;}}
                QPushButton:hover{{background:{c2};}}
                QPushButton:pressed{{background:{c3};}}
            """)
            return b

        self.export_btn = _btn("📊 Export to Excel", "#217346", "#1a5c38", "#155724")
        self.export_btn.clicked.connect(self.export_to_excel)
        self.print_btn = _btn("🖨️ Print Report", "#6f42c1", "#5a2d91", "#4c1f75")
        self.print_btn.clicked.connect(self.print_table)

        lay.addStretch()
        lay.addWidget(self.export_btn)
        lay.addSpacing(15)
        lay.addWidget(self.print_btn)
        lay.addStretch()
        parent.addWidget(frame, 0)

    # ── Helpers ───────────────────────────────────────────────────────────
    def _bold(self, text):
        lbl = QLabel(text)
        lbl.setFont(QFont("", 11, QFont.Bold))
        return lbl

    def _combo(self, width):
        c = QComboBox()
        c.setFixedWidth(width)
        return c

    # ── Data loading ──────────────────────────────────────────────────────
    def load_corporations(self):
        self.group_selector.blockSignals(True)
        self.group_selector.clear()
        try:
            rows = db_manager.execute_query(
                "SELECT DISTINCT os_name FROM branches "
                "WHERE os_name IS NOT NULL AND os_name != '' ORDER BY os_name"
            )
            if rows:
                for r in rows:
                    name = r['os_name'] if isinstance(r, dict) else r[0]
                    self.group_selector.addItem(name)
        except Exception as e:
            print(f"Error loading groups: {e}")
        finally:
            self.group_selector.blockSignals(False)

    def _load_os_options(self):
        self.os_filter_selector.blockSignals(True)
        prev = self.os_filter_selector.currentData()
        self.os_filter_selector.clear()
        self.os_filter_selector.addItem("All (by Corporation)", None)
        try:
            rows = db_manager.execute_query(
                "SELECT DISTINCT os_name FROM branches "
                "WHERE os_name IS NOT NULL AND os_name != '' ORDER BY os_name"
            )
            if rows:
                for r in rows:
                    name = r['os_name'] if isinstance(r, dict) else r[0]
                    self.os_filter_selector.addItem(name, name)
            if prev:
                idx = self.os_filter_selector.findData(prev)
                if idx >= 0:
                    self.os_filter_selector.setCurrentIndex(idx)
        except Exception as e:
            print(f"Error loading OS options: {e}")
        finally:
            self.os_filter_selector.blockSignals(False)

    # ── Populate ──────────────────────────────────────────────────────────
    def populate_table(self):
        self._is_loading = True
        self.table.setRowCount(0)

        group = self.group_selector.currentText().strip() if hasattr(self, 'group_selector') else ""
        date_start, date_end = self.date_range_widget.get_date_range()
        is_range = self.date_range_widget.is_range_mode()
        reg_value = self.reg_filter_selector.currentData() if hasattr(self, 'reg_filter_selector') else "all"

        self._update_report_header()

        if not group:
            self._is_loading = False
            return

        groups = COLUMN_GROUPS
        headers, col_meta = _build_columns(groups)

        # Insert sub-header row (row 0)
        self.table.insertRow(0)
        self._create_sub_header_row(groups)

        # Gather needed DB columns
        needed = set()
        for _, subs in groups:
            for _, db_cols, _ in subs:
                needed.update(db_cols)

        # Build SELECT
        parts = ["b.name AS branch"]
        for c in sorted(needed):
            if is_range:
                parts.append(f"SUM(COALESCE(dr.`{c}`, 0)) AS `{c}`")
            else:
                parts.append(f"COALESCE(dr.`{c}`, 0) AS `{c}`")

        # Date condition in JOIN ON so branches without entries still appear
        if is_range:
            join_date_cond = "AND dr.date >= %s AND dr.date <= %s"
            params = [date_start, date_end]
        else:
            join_date_cond = "AND dr.date = %s"
            params = [date_start]

        where = []

        if group:
            where.append("b.os_name = %s")
            params.append(group)

        # Registration filter
        if reg_value == "registered":
            where.append("b.is_registered = 1")
        elif reg_value == "not_registered":
            where.append("(b.is_registered = 0 OR b.is_registered IS NULL)")

        # Only global-tagged branches
        where.append("b.global_tag = 'GLOBAL'")

        where_clause = f"WHERE {' AND '.join(where)}" if where else ""
        group_by = " GROUP BY b.name" if is_range else ""
        sql = (
            f"SELECT {', '.join(parts)} "
            f"FROM branches b "
            f"LEFT JOIN {TABLE_NAME} dr ON b.name COLLATE utf8mb4_general_ci = dr.branch COLLATE utf8mb4_general_ci "
            f"{join_date_cond} "
            f"{where_clause} "
            f"{group_by} "
            f"ORDER BY b.name"
        )

        # Store context for async callback
        self._pending_populate = {
            'headers': headers,
            'col_meta': col_meta,
            'groups': groups,
        }

        run_query_async(
            parent=self,
            query=sql,
            params=tuple(params),
            on_result=self._on_populate_result,
            on_error=self._on_populate_error,
            loading_message="\u23f3  Loading services…",
        )

    def _on_populate_error(self, err):
        QMessageBox.critical(self, "Query Error", err)
        self._is_loading = False

    def _on_populate_result(self, rows):
        ctx = self._pending_populate
        headers = ctx['headers']
        col_meta = ctx['col_meta']

        if not rows:
            self._is_loading = False
            return

        col_count = len(headers)
        totals = [0.0] * col_count

        for row_data in rows:
            r = self.table.rowCount()
            self.table.insertRow(r)

            branch_item = QTableWidgetItem(row_data['branch'])
            branch_item.setFlags(branch_item.flags() & ~Qt.ItemIsEditable)
            branch_item.setFont(QFont("", 10, QFont.Bold))
            self.table.setItem(r, 0, branch_item)

            for ci, meta in enumerate(col_meta, start=1):
                val = sum(float(row_data.get(c, 0) or 0) for c in meta['db_cols'])
                if meta['is_lotes']:
                    val = int(val)
                    txt = str(val)
                else:
                    txt = f"{val:,.2f}"
                item = QTableWidgetItem(txt)
                item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                self.table.setItem(r, ci, item)
                totals[ci] += val

        # Total row
        tr = self.table.rowCount()
        self.table.insertRow(tr)
        self.table.setRowHeight(tr, 36)
        total_item = QTableWidgetItem("TOTAL")
        total_item.setFont(QFont("", 10, QFont.Bold))
        total_item.setBackground(QColor("#343a40"))
        total_item.setForeground(QColor("white"))
        total_item.setTextAlignment(Qt.AlignCenter)
        total_item.setFlags(total_item.flags() & ~Qt.ItemIsEditable)
        self.table.setItem(tr, 0, total_item)

        for ci, meta in enumerate(col_meta, start=1):
            val = totals[ci]
            if meta['is_lotes']:
                txt = str(int(val))
            else:
                txt = f"{val:,.2f}"
            ti = QTableWidgetItem(txt)
            ti.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            ti.setFont(QFont("", 10, QFont.Bold))
            ti.setBackground(QColor("#343a40"))
            ti.setForeground(QColor("white"))
            ti.setFlags(ti.flags() & ~Qt.ItemIsEditable)
            self.table.setItem(tr, ci, ti)

        self._is_loading = False

    # ── Export ────────────────────────────────────────────────────────────
    def export_to_excel(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font as XlFont, Alignment, PatternFill, Border, Side
        except ImportError:
            QMessageBox.critical(self, "Missing Dependency",
                                 "openpyxl is required.\nInstall: pip install openpyxl")
            return

        rows_count = self.table.rowCount()
        if rows_count == 0:
            QMessageBox.warning(self, "No Data", "Nothing to export.")
            return

        corp = self.corp_selector.currentText() or "All"
        ds, de = self.date_range_widget.get_date_range()
        date_tag = ds if ds == de else f"{ds}_to_{de}"
        date_display = ds if ds == de else f"{ds} to {de}"
        default_name = f"Global_Other_Services_{corp}_{date_tag}.xlsx"

        path, _ = QFileDialog.getSaveFileName(self, "Save Excel", default_name,
                                              "Excel Files (*.xlsx);;All Files (*)")
        if not path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Global Other Services"

            title_font = XlFont(bold=True, size=16)
            hdr_font = XlFont(bold=True, size=11, color="FFFFFF")
            hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))

            ws.merge_cells('A1:K1')
            ws['A1'] = "Global Other Services Report"
            ws['A1'].font = title_font
            ws['A1'].alignment = Alignment(horizontal='center')
            ws['A3'] = "Corporation:"
            ws['B3'] = corp
            ws['A4'] = "Date:"
            ws['B4'] = date_display
            ws['A3'].font = XlFont(bold=True)
            ws['A4'].font = XlFont(bold=True)

            groups = COLUMN_GROUPS
            headers, col_meta = _build_columns(groups)

            # Group header row (row 6)
            gr = 6
            col_idx = 2  # B column (A = Branch)
            ws.cell(row=gr, column=1, value="Branch").font = hdr_font
            ws.cell(row=gr, column=1).fill = hdr_fill
            ws.cell(row=gr, column=1).border = border
            for gn, subs in groups:
                start_col = col_idx
                end_col = col_idx + len(subs) - 1
                if start_col != end_col:
                    ws.merge_cells(start_row=gr, start_column=start_col,
                                   end_row=gr, end_column=end_col)
                cell = ws.cell(row=gr, column=start_col, value=gn)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
                for c in range(start_col, end_col + 1):
                    ws.cell(row=gr, column=c).border = border
                    ws.cell(row=gr, column=c).fill = hdr_fill
                col_idx = end_col + 1

            # Sub-header row (row 7)
            sr = 7
            ws.cell(row=sr, column=1, value="").font = hdr_font
            ws.cell(row=sr, column=1).fill = hdr_fill
            ws.cell(row=sr, column=1).border = border
            for ci, h in enumerate(headers[1:], start=2):
                cell = ws.cell(row=sr, column=ci, value=h)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = border

            # Data rows (skip table row 0 = sub-header)
            for tr in range(1, rows_count):
                er = sr + tr
                for ci in range(len(headers)):
                    item = self.table.item(tr, ci)
                    cell = ws.cell(row=er, column=ci + 1)
                    if item:
                        txt = item.text()
                        if ci >= 1:
                            try:
                                cell.value = float(txt.replace(',', ''))
                                cell.number_format = '#,##0.00' if '.' in txt else '#,##0'
                            except ValueError:
                                cell.value = txt
                        else:
                            cell.value = txt
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')

            for i in range(1, len(headers) + 1):
                ws.column_dimensions[chr(64 + min(i, 26))].width = 15 if i > 1 else 22

            wb.save(path)
            QMessageBox.information(self, "Exported", f"Report saved to:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", str(e))

    # ── Print ─────────────────────────────────────────────────────────────
    def print_table(self):
        rows_count = self.table.rowCount()
        if rows_count == 0:
            QMessageBox.warning(self, "No Data", "Nothing to print.")
            return

        corp = self.corp_selector.currentText() or "All"
        ds, de = self.date_range_widget.get_date_range()
        date_display = ds if ds == de else f"{ds} to {de}"

        groups = COLUMN_GROUPS
        headers, _ = _build_columns(groups)

        doc = QTextDocument()
        html = "<h2>Global Other Services Report</h2>"
        html += f"<p><b>Corporation:</b> {corp}<br><b>Date:</b> {date_display}</p>"
        html += "<table border='1' cellspacing='0' cellpadding='4' style='border-collapse:collapse;font-size:9px;'>"

        # Group header
        html += "<tr style='background:#4472C4;color:white;font-weight:bold;'>"
        html += "<th>Branch</th>"
        for gn, subs in groups:
            html += f"<th colspan='{len(subs)}'>{gn}</th>"
        html += "</tr>"

        # Sub header (row 0)
        html += "<tr style='background:#5a6d7e;color:white;font-weight:bold;'>"
        html += "<th></th>"
        for h in headers[1:]:
            html += f"<th>{h}</th>"
        html += "</tr>"

        # Data
        for r in range(1, rows_count):
            is_total = False
            item0 = self.table.item(r, 0)
            if item0 and item0.text() == "TOTAL":
                is_total = True
                html += "<tr style='background:#343a40;color:white;font-weight:bold;'>"
            else:
                html += "<tr>"
            for c in range(len(headers)):
                item = self.table.item(r, c)
                txt = item.text() if item else ""
                html += f"<td style='padding:4px;text-align:{'center' if c == 0 else 'right'};'>{txt}</td>"
            html += "</tr>"

        html += "</table>"
        doc.setHtml(html)

        printer = QPrinter()
        dlg = QPrintDialog(printer, self)
        if dlg.exec_() == QPrintDialog.Accepted:
            doc.print_(printer)
