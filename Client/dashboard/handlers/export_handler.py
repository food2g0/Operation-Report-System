"""Export handler - manages report export to Excel."""

import logging
import os
import datetime

logger = logging.getLogger(__name__)


class ExportHandler:
    """Handles report export to Excel files."""

    def __init__(self, branch, corporation, line_of_business=''):
        """Initialize export handler.

        Args:
            branch: Branch name
            corporation: Corporation name
            line_of_business: Branch LOB tag; VAT column added when 'GROUP 1'
        """
        self.branch = branch
        self.corporation = corporation
        self.line_of_business = line_of_business
        self._has_vat = str(line_of_business or '').strip().lower().replace(' ', '') == 'group1'

    def validate_file_path(self, file_path):
        """Validate file path for security (prevent directory traversal).

        Args:
            file_path: Path to validate

        Returns:
            tuple: (is_valid, error_message)
        """
        file_path = os.path.abspath(file_path)
        home_dir = os.path.expanduser("~")
        documents_dir = os.path.join(home_dir, "Documents")

        if not (file_path.startswith(documents_dir) or file_path.startswith(home_dir)):
            logger.error(f"Invalid file path: {file_path}")
            return False, "File must be saved in Documents or home directory"

        return True, ""

    def build_workbook(self, wb, data_dict, selected_date, user_email, brands_to_export):
        """Build Excel workbook with report data.

        Args:
            wb: Workbook instance
            data_dict: Dict with brand data
            selected_date: Report date (YYYY-MM-DD)
            user_email: User email
            brands_to_export: List of (brand_name, brand_data) tuples

        Returns:
            None (modifies wb in place)
        """
        try:
            from openpyxl.styles import Font, Alignment, PatternFill, Border
        except ImportError:
            raise ImportError("openpyxl package is required")

        title_font = Font(bold=True, size=16, color="000000")
        header_font = Font(bold=True, size=11, color="000000")
        header_fill = PatternFill(fill_type=None)
        summary_fill = PatternFill(fill_type=None)
        total_fill = PatternFill(fill_type=None)
        border = Border()

        first_sheet = True
        for brand_name, data in brands_to_export:
            if first_sheet:
                ws = wb.active
                ws.title = f"{brand_name} Report"
                first_sheet = False
            else:
                ws = wb.create_sheet(title=f"{brand_name} Report")

            try:
                ws.sheet_view.showGridLines = False
            except Exception:
                pass

            self._write_header(ws, brand_name, selected_date, user_email, title_font)
            current_row = self._write_summary(ws, data, summary_fill, border, header_font)
            current_row = self._write_debit_section(
                ws, data, current_row, header_font, header_fill, border, total_fill
            )
            current_row = self._write_credit_section(
                ws, data, current_row, header_font, header_fill, border, total_fill
            )

            ws.column_dimensions['A'].width = 35
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 5

    def _write_header(self, ws, brand_name, selected_date, user_email, title_font):
        """Write report header section.

        Args:
            ws: Worksheet
            brand_name: Brand name
            selected_date: Report date
            user_email: User email
            title_font: Title font style
        """
        from openpyxl.styles import Font, Alignment

        ws.merge_cells('A1:D1')
        ws['A1'] = f"Daily Cash Report - {brand_name}"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')

        ws['A3'] = "Date:"
        ws['B3'] = selected_date
        ws['A4'] = "Branch:"
        ws['B4'] = self.branch
        ws['A5'] = "Corporation:"
        ws['B5'] = self.corporation
        ws['A6'] = "User:"
        ws['B6'] = user_email
        ws['A7'] = "Generated:"
        ws['B7'] = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        for row in range(3, 8):
            ws.cell(row=row, column=1).font = Font(bold=True)

    def _write_summary(self, ws, data, summary_fill, border, header_font):
        """Write summary section.

        Args:
            ws: Worksheet
            data: Brand data
            summary_fill: Fill style
            border: Border style
            header_font: Header font

        Returns:
            int: Current row after summary
        """
        from openpyxl.styles import Font, Alignment

        current_row = 9

        ws.merge_cells(f'A{current_row}:D{current_row}')
        summary_header = ws.cell(row=current_row, column=1)
        summary_header.value = "Summary"
        summary_header.font = Font(bold=True, size=14, color="000000")
        summary_header.fill = summary_fill
        summary_header.alignment = Alignment(horizontal='center')
        current_row += 1

        summary_items = [
            ("Beginning Balance", data["beginning_balance"]),
            ("Ending Balance", data["ending_balance"]),
            ("Cash Count", data["cash_count"]),
            ("Variance", data["variance"]),
        ]

        for label, value in summary_items:
            label_cell = ws.cell(row=current_row, column=1)
            label_cell.value = label
            label_cell.font = Font(bold=True, color="000000")
            label_cell.border = border

            value_cell = ws.cell(row=current_row, column=2)
            value_cell.value = value
            value_cell.number_format = '#,##0.00'
            value_cell.alignment = Alignment(horizontal='right')
            value_cell.border = border

            if label == "Variance":
                status_cell = ws.cell(row=current_row, column=3)
                if value > 0:
                    status_cell.value = "(Over)"
                elif value < 0:
                    status_cell.value = "(Short)"
                else:
                    status_cell.value = "(Balanced)"
                status_cell.font = Font(bold=True, color="000000")
                status_cell.border = border

            current_row += 1

        return current_row + 1

    def _write_debit_section(self, ws, data, current_row, header_font, header_fill, border, total_fill):
        """Write debit (cash receipt) section.

        Args:
            ws: Worksheet
            data: Brand data
            current_row: Starting row
            header_font: Header font
            header_fill: Header fill
            border: Border style
            total_fill: Total fill style

        Returns:
            int: Current row after section
        """
        from openpyxl.styles import Font, Alignment

        if not data["debit"]:
            return current_row

        for col, hdr in enumerate(["Field", "Amount", "Lotes", ""], 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = hdr
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        current_row += 1

        debit_total = 0.0
        for label, amount, lotes in data["debit"]:
            debit_total += amount
            ws.cell(row=current_row, column=1).value = label
            ws.cell(row=current_row, column=1).border = border

            amount_cell = ws.cell(row=current_row, column=2)
            amount_cell.value = amount
            amount_cell.number_format = '#,##0.00'
            amount_cell.alignment = Alignment(horizontal='right')
            amount_cell.border = border

            lotes_cell = ws.cell(row=current_row, column=3)
            lotes_cell.value = lotes if lotes else "-"
            lotes_cell.alignment = Alignment(horizontal='center')
            lotes_cell.border = border

            current_row += 1

        total_label = ws.cell(row=current_row, column=1)
        total_label.value = "Total Cash Receipt"
        total_label.font = Font(bold=True, color="000000")
        total_label.fill = total_fill
        total_label.border = border

        total_amount = ws.cell(row=current_row, column=2)
        total_amount.value = debit_total
        total_amount.number_format = '#,##0.00'
        total_amount.font = Font(bold=True)
        total_amount.fill = total_fill
        total_amount.alignment = Alignment(horizontal='right')
        total_amount.border = border

        ws.cell(row=current_row, column=3).fill = total_fill
        ws.cell(row=current_row, column=3).border = border

        return current_row + 2

    def _write_credit_section(self, ws, data, current_row, header_font, header_fill, border, total_fill):
        """Write credit (cash out) section.

        Args:
            ws: Worksheet
            data: Brand data
            current_row: Starting row
            header_font: Header font
            header_fill: Header fill
            border: Border style
            total_fill: Total fill style

        Returns:
            int: Current row after section
        """
        from openpyxl.styles import Font, Alignment

        if not data["credit"]:
            return current_row

        for col, hdr in enumerate(["Field", "Amount", "Lotes", ""], 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = hdr
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        current_row += 1

        credit_total = 0.0
        for label, amount, lotes in data["credit"]:
            credit_total += amount
            ws.cell(row=current_row, column=1).value = label
            ws.cell(row=current_row, column=1).border = border

            amount_cell = ws.cell(row=current_row, column=2)
            amount_cell.value = amount
            amount_cell.number_format = '#,##0.00'
            amount_cell.alignment = Alignment(horizontal='right')
            amount_cell.border = border

            lotes_cell = ws.cell(row=current_row, column=3)
            lotes_cell.value = lotes if lotes else "-"
            lotes_cell.alignment = Alignment(horizontal='center')
            lotes_cell.border = border

            current_row += 1

        total_label = ws.cell(row=current_row, column=1)
        total_label.value = "Total Cash Out"
        total_label.font = Font(bold=True, color="000000")
        total_label.fill = total_fill
        total_label.border = border

        total_amount = ws.cell(row=current_row, column=2)
        total_amount.value = credit_total
        total_amount.number_format = '#,##0.00'
        total_amount.font = Font(bold=True)
        total_amount.fill = total_fill
        total_amount.alignment = Alignment(horizontal='right')
        total_amount.border = border

        ws.cell(row=current_row, column=3).fill = total_fill
        ws.cell(row=current_row, column=3).border = border

        return current_row + 2

    # ── Palawan Details sheet ─────────────────────────────────────────────────

    def write_palawan_sheet(self, wb, db_manager, selected_date):
        """Add a 'Palawan Details' sheet — SEND-OUT (yellow) | PAY-OUT (green) side-by-side."""
        import json as _json

        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            import openpyxl.utils as _xlutil
        except ImportError:
            return

        result = db_manager.execute_query(
            "SELECT sendout_detailed_principal, payout_detailed_principal "
            "FROM payable_tbl_brand_a "
            "WHERE date=%s AND branch=%s AND corporation=%s LIMIT 1",
            (selected_date, self.branch, self.corporation),
        )

        if not result:
            return

        row = result[0]

        def _parse(col):
            raw = row.get(col)
            if not raw:
                return []
            try:
                return _json.loads(raw)
            except Exception:
                return []

        sendout_txns = _parse("sendout_detailed_principal")
        payout_txns  = _parse("payout_detailed_principal")

        if not sendout_txns and not payout_txns:
            return

        ws = wb.create_sheet(title="Palawan Details")

        # ── Styles ───────────────────────────────────────────────────────────
        thin = Side(style='thin')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        NUM_FMT = '#,##0.00'

        # Section header fills
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        green_fill  = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        # Column header fill (dark blue-grey)
        col_hdr_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        col_hdr_font = Font(bold=True, color="FFFFFF", size=10)
        col_hdr_aln  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # Formula description row fill (light blue-grey)
        formula_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        formula_font = Font(bold=True, size=9, color="1F3864")
        formula_aln  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # Totals row
        tot_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        tot_font = Font(bold=True, size=10)
        tot_aln_r = Alignment(horizontal="right", vertical="center")
        tot_aln_l = Alignment(horizontal="left",  vertical="center")
        # Data rows
        data_aln_l = Alignment(horizontal="left",   vertical="center", wrap_text=True)
        data_aln_r = Alignment(horizontal="right",  vertical="center")
        data_aln_c = Alignment(horizontal="center", vertical="center")

        # ── Column layout ─────────────────────────────────────────────────────
        # Without VAT: SO cols 1-18, gap 19-20, PO cols 21-37
        # With VAT:    SO cols 1-19, gap 20-21, PO cols 22-38
        has_vat = self._has_vat
        SO_START = 1
        SO_COLS  = 19 if has_vat else 18   # SO section width
        PO_COLS  = 18 if has_vat else 17   # PO section width
        PO_START = SO_START + SO_COLS + 2  # 2-col gap

        # ── Row 1: Section headers ─────────────────────────────────────────────
        ws.merge_cells(start_row=1, start_column=SO_START,
                       end_row=1, end_column=SO_START + SO_COLS - 1)
        c = ws.cell(row=1, column=SO_START)
        c.value = "REMITTANCE (SEND-OUT)"
        c.fill  = yellow_fill
        c.font  = Font(bold=True, size=13, color="000000")
        c.alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells(start_row=1, start_column=PO_START,
                       end_row=1, end_column=PO_START + PO_COLS - 1)
        c = ws.cell(row=1, column=PO_START)
        c.value = "PAY-OUT"
        c.fill  = green_fill
        c.font  = Font(bold=True, size=13, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[1].height = 22

        def _formula_cell(r, c_idx, text="", merged_end=None):
            if merged_end:
                ws.merge_cells(start_row=r, start_column=c_idx,
                               end_row=r, end_column=merged_end)
            cell = ws.cell(row=r, column=c_idx)
            cell.value     = text
            cell.fill      = formula_fill
            cell.font      = formula_font
            cell.alignment = formula_aln
            cell.border    = border

        # ── Row 2: Formula description row ────────────────────────────────────
        for ci in range(SO_START, SO_START + SO_COLS):
            _formula_cell(2, ci)
        for ci in range(PO_START, PO_START + PO_COLS):
            _formula_cell(2, ci)

        # SO formula labels
        _formula_cell(2, SO_START + 3, "COMPUTATION", merged_end=SO_START + 5)
        _formula_cell(2, SO_START + 6, "COM. + SC")
        _formula_cell(2, SO_START + 7, "PRINCIPAL +\nTOTAL SC")  # cash on hand
        _formula_cell(2, SO_START + 8, "COM. × 39%")             # income
        if has_vat:
            _formula_cell(2, SO_START + 9,  "INCOME / 1.12\n× 12%")  # VAT
            _formula_cell(2, SO_START + 10, "COH - INCOME")           # A/P
        else:
            _formula_cell(2, SO_START + 9, "COH - INCOME")            # A/P

        # PO formula labels (no Cash on Hand)
        _formula_cell(2, PO_START + 3, "COMPUTATION", merged_end=PO_START + 5)
        _formula_cell(2, PO_START + 6, "COM. + SC")
        _formula_cell(2, PO_START + 7, "COM. × 43%")              # income
        if has_vat:
            _formula_cell(2, PO_START + 8, "INCOME / 1.12\n× 12%")   # VAT
            _formula_cell(2, PO_START + 9, "PRIN. + INCOME")          # A/R
        else:
            _formula_cell(2, PO_START + 8, "PRIN. + INCOME")          # A/R

        ws.row_dimensions[2].height = 28

        # ── Row 3: Column headers ─────────────────────────────────────────────
        _text_hdrs = [
            "RELATIONSHIP\n(RECEIVER /\nSENDER)",
            "SOURCE OF\nFUNDS",
            "PURPOSE\nOF TRANS.",
            "KYC DOCS.\nPHOTO ID",
            "POSITION IN\nTHE BUS. CO.",
            "BUSINESS\nCOMPANY NAME",
            "EVALUATION\nREMARKS",
            "TOTAL\nTRANSACTION",
        ]
        SO_HDRS = [
            "TRANSACTION\nCODE", "SENDER", "RECEIVER",
            "PRINCIPAL", "COMMISSION", "SC",
            "TOTAL SC", "CASH ON\nHAND", "INCOME",
        ] + (["VAT\n(12%)"] if has_vat else []) + ["A/P\nPALAWAN"] + _text_hdrs

        PO_HDRS = [
            "TRANSACTION\nCODE", "SENDER", "RECEIVER",
            "PRINCIPAL", "COMMISSION", "SC",
            "TOTAL SC", "INCOME",
        ] + (["VAT\n(12%)"] if has_vat else []) + ["A/R\nPALAWAN"] + _text_hdrs

        for i, hdr in enumerate(SO_HDRS):
            c = ws.cell(row=3, column=SO_START + i)
            c.value, c.fill, c.font, c.alignment, c.border = (
                hdr, col_hdr_fill, col_hdr_font, col_hdr_aln, border)
        for i, hdr in enumerate(PO_HDRS):
            c = ws.cell(row=3, column=PO_START + i)
            c.value, c.fill, c.font, c.alignment, c.border = (
                hdr, col_hdr_fill, col_hdr_font, col_hdr_aln, border)

        ws.row_dimensions[3].height = 40

        # ── Data writing helpers ──────────────────────────────────────────────
        def _dc(r, c_idx, value, is_num=False):
            cell = ws.cell(row=r, column=c_idx)
            cell.value  = value
            cell.border = border
            if is_num:
                cell.number_format = NUM_FMT
                cell.alignment = data_aln_r
            else:
                cell.alignment = data_aln_l

        def _write_so_row(r, txn):
            commission   = float(txn.get("commission", 0))
            income_39    = round(commission * 0.39, 2)
            vat_so       = float(txn.get("vat", income_39 / 1.12 * 0.12))
            principal    = float(txn.get("principal", 0))
            total_sc     = float(txn.get("total_sc", 0))
            cash_on_hand = principal + total_sc
            ap_palawan   = cash_on_hand - income_39

            _dc(r, SO_START + 0, txn.get("code", ""))
            _dc(r, SO_START + 1, txn.get("sender", ""))
            _dc(r, SO_START + 2, txn.get("receiver", ""))
            _dc(r, SO_START + 3, principal,    True)
            _dc(r, SO_START + 4, commission,   True)
            _dc(r, SO_START + 5, float(txn.get("sc", 0)), True)
            _dc(r, SO_START + 6, total_sc,     True)
            _dc(r, SO_START + 7, cash_on_hand, True)
            _dc(r, SO_START + 8, income_39,    True)
            o = 1 if has_vat else 0          # column offset after optional VAT
            if has_vat:
                _dc(r, SO_START + 9, vat_so, True)
            _dc(r, SO_START + 9  + o, ap_palawan,                   True)
            _dc(r, SO_START + 10 + o, txn.get("relationship", ""))
            _dc(r, SO_START + 11 + o, txn.get("source_funds", ""))
            _dc(r, SO_START + 12 + o, txn.get("purpose", ""))
            _dc(r, SO_START + 13 + o, txn.get("kyc_docs", ""))
            _dc(r, SO_START + 14 + o, txn.get("position", ""))
            _dc(r, SO_START + 15 + o, txn.get("business_name", ""))
            _dc(r, SO_START + 16 + o, txn.get("evaluation", ""))
            _dc(r, SO_START + 17 + o, cash_on_hand, True)           # TOTAL = COH

        def _write_po_row(r, txn):
            commission = float(txn.get("commission", 0))
            income_43  = round(commission * 0.43, 2)
            vat_po     = float(txn.get("vat", income_43 / 1.12 * 0.12))
            principal  = float(txn.get("principal", 0))
            total_sc   = float(txn.get("total_sc", 0))
            ar_palawan = principal + income_43
            total_txn  = principal + total_sc

            _dc(r, PO_START + 0, txn.get("code", ""))
            _dc(r, PO_START + 1, txn.get("sender", ""))
            _dc(r, PO_START + 2, txn.get("receiver", ""))
            _dc(r, PO_START + 3, principal,  True)
            _dc(r, PO_START + 4, commission, True)
            _dc(r, PO_START + 5, float(txn.get("sc", 0)), True)
            _dc(r, PO_START + 6, total_sc,   True)
            _dc(r, PO_START + 7, income_43,  True)
            o = 1 if has_vat else 0
            if has_vat:
                _dc(r, PO_START + 8, vat_po, True)
            _dc(r, PO_START + 8  + o, ar_palawan,                   True)
            _dc(r, PO_START + 9  + o, txn.get("relationship", ""))
            _dc(r, PO_START + 10 + o, txn.get("source_funds", ""))
            _dc(r, PO_START + 11 + o, txn.get("purpose", ""))
            _dc(r, PO_START + 12 + o, txn.get("kyc_docs", ""))
            _dc(r, PO_START + 13 + o, txn.get("position", ""))
            _dc(r, PO_START + 14 + o, txn.get("business_name", ""))
            _dc(r, PO_START + 15 + o, txn.get("evaluation", ""))
            _dc(r, PO_START + 16 + o, total_txn, True)

        # ── Write data rows ───────────────────────────────────────────────────
        max_rows = max(len(sendout_txns), len(payout_txns))
        data_start = 4
        for i in range(max_rows):
            r = data_start + i
            ws.row_dimensions[r].height = 18
            if i < len(sendout_txns):
                _write_so_row(r, sendout_txns[i])
            if i < len(payout_txns):
                _write_po_row(r, payout_txns[i])

        # ── Totals row ────────────────────────────────────────────────────────
        tot_row = data_start + max_rows + 1
        ws.row_dimensions[tot_row].height = 20

        def _tot_cell(r, c_idx, value=None, label=None):
            cell = ws.cell(row=r, column=c_idx)
            cell.fill   = tot_fill
            cell.font   = tot_font
            cell.border = border
            if value is not None:
                cell.value         = value
                cell.number_format = NUM_FMT
                cell.alignment     = tot_aln_r
            else:
                cell.value     = label or ""
                cell.alignment = tot_aln_l

        for ci in range(SO_START, SO_START + SO_COLS):
            _tot_cell(tot_row, ci)
        for ci in range(PO_START, PO_START + PO_COLS):
            _tot_cell(tot_row, ci)

        o = 1 if has_vat else 0

        if sendout_txns:
            _tot_cell(tot_row, SO_START, label="TOTAL")
            comm_so = sum(float(t.get("commission", 0)) for t in sendout_txns)
            inc_39  = round(comm_so * 0.39, 2)
            prin_so = sum(float(t.get("principal", 0)) for t in sendout_txns)
            sc_so   = sum(float(t.get("sc", 0))        for t in sendout_txns)
            tsc_so  = sum(float(t.get("total_sc", 0))  for t in sendout_txns)
            coh_so  = prin_so + tsc_so
            ap_so   = coh_so - inc_39
            _tot_cell(tot_row, SO_START + 3, prin_so)
            _tot_cell(tot_row, SO_START + 4, comm_so)
            _tot_cell(tot_row, SO_START + 5, sc_so)
            _tot_cell(tot_row, SO_START + 6, tsc_so)
            _tot_cell(tot_row, SO_START + 7, coh_so)
            _tot_cell(tot_row, SO_START + 8, inc_39)
            if has_vat:
                vat_so_tot = sum(float(t.get("vat", float(t.get("income", 0)) / 1.12 * 0.12))
                                 for t in sendout_txns)
                _tot_cell(tot_row, SO_START + 9, vat_so_tot)
            _tot_cell(tot_row, SO_START + 9  + o, ap_so)
            _tot_cell(tot_row, SO_START + 17 + o, coh_so)   # TOTAL TRANSACTION

        if payout_txns:
            _tot_cell(tot_row, PO_START, label="TOTAL")
            comm_po = sum(float(t.get("commission", 0)) for t in payout_txns)
            inc_43  = round(comm_po * 0.43, 2)
            prin_po = sum(float(t.get("principal", 0)) for t in payout_txns)
            sc_po   = sum(float(t.get("sc", 0))        for t in payout_txns)
            tsc_po  = sum(float(t.get("total_sc", 0))  for t in payout_txns)
            ar_po   = prin_po + inc_43
            if has_vat:
                vat_po_tot = sum(float(t.get("vat", float(t.get("income", 0)) / 1.12 * 0.12))
                                 for t in payout_txns)
                _tot_cell(tot_row, PO_START + 8, vat_po_tot)
            _tot_cell(tot_row, PO_START + 3,      prin_po)
            _tot_cell(tot_row, PO_START + 4,      comm_po)
            _tot_cell(tot_row, PO_START + 5,      sc_po)
            _tot_cell(tot_row, PO_START + 6,      tsc_po)
            _tot_cell(tot_row, PO_START + 7,      inc_43)
            _tot_cell(tot_row, PO_START + 8  + o, ar_po)
            _tot_cell(tot_row, PO_START + 16 + o, prin_po + tsc_po)

        # ── Column widths ─────────────────────────────────────────────────────
        if has_vat:
            SO_W = [18, 15, 15, 12, 12, 8, 12, 14, 12, 10, 12, 18, 15, 15, 14, 18, 20, 16, 14]
            PO_W = [18, 15, 15, 12, 12, 8, 12, 12, 10, 12, 18, 15, 15, 14, 18, 20, 16, 14]
        else:
            SO_W = [18, 15, 15, 12, 12, 8, 12, 14, 12, 12, 18, 15, 15, 14, 18, 20, 16, 14]
            PO_W = [18, 15, 15, 12, 12, 8, 12, 12, 12, 18, 15, 15, 14, 18, 20, 16, 14]
        for i, w in enumerate(SO_W):
            ws.column_dimensions[_xlutil.get_column_letter(SO_START + i)].width = w
        # gap cols (2 narrow cols after SO section)
        ws.column_dimensions[_xlutil.get_column_letter(SO_START + SO_COLS)].width = 2
        ws.column_dimensions[_xlutil.get_column_letter(SO_START + SO_COLS + 1)].width = 2
        for i, w in enumerate(PO_W):
            ws.column_dimensions[_xlutil.get_column_letter(PO_START + i)].width = w

        # Freeze pane at row 4 (headers stay visible)
        ws.freeze_panes = "A4"

        logger.info(
            "Palawan Details sheet: %d sendout, %d payout",
            len(sendout_txns), len(payout_txns),
        )
