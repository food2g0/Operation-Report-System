
from PyQt5.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QPushButton,
    QFrame, QTableWidget, QTableWidgetItem, QHeaderView, QGroupBox,
    QAbstractItemView, QDialogButtonBox, QFormLayout, QScrollArea, QWidget,
    QComboBox
)
from PyQt5.QtGui import QDoubleValidator
from PyQt5.QtCore import Qt, QEvent
from PyQt5.QtWidgets import QApplication


class PalawanTransactionDetailDialog(QDialog):


    def __init__(self, field_label="Transaction Details", transactions=None, parent=None, code_name_lookup=None, transaction_type=None, has_payroll=False, line_of_business=''):

        super().__init__(parent)
        self.setWindowTitle(f"Palawan Transactions – {field_label}")
        self.setMinimumSize(800, 480)
        self.setModal(True)
        self._transactions = list(transactions) if transactions else []
        self._field_label = field_label
        # transaction_type: 'so' (send-out, 49%), 'po' (pay-out, 43%), 'int' (international, default 43%)
        self._transaction_type = transaction_type or 'po'
        self._income_rate = 0.39 if transaction_type == 'so' else 0.43
        self._has_payroll = has_payroll
        # VAT (12%) applies only to Line of Business Group 1 branches
        # Normalise: strip spaces so 'Group 1', 'group 1', 'Group1', 'group1' all match
        self._has_vat = str(line_of_business or '').strip().lower().replace(' ', '') == 'group1'

        # Size to 90% of available screen so it fits on any monitor
        from PyQt5.QtWidgets import QApplication
        screen = QApplication.primaryScreen().availableGeometry()
        self.resize(
            min(1200, int(screen.width() * 0.92)),
            min(860, int(screen.height() * 0.92)),
        )

        self._zoom_factor = 1.0
        self._orig_widget_styles = None  # captured lazily on first zoom

        # Install self as app-level event filter (same pattern as core.py dashboard)
        QApplication.instance().installEventFilter(self)
        self.finished.connect(lambda: QApplication.instance().removeEventFilter(self))

        # Base font — ensures all labels/inputs are readable on small monitors
        self.setStyleSheet("""
            QLabel       { font-size: 13px; font-weight: 600; color: #1E293B; }
            QLineEdit    { font-size: 13px; }
            QComboBox    { font-size: 13px; }
            QPushButton  { font-size: 13px; }
        """)

        root = QVBoxLayout(self)
        root.setSpacing(8)
        root.setContentsMargins(16, 12, 16, 12)

        # ── Top bar: title + zoom controls ──────────────────────────────────
        top_bar = QHBoxLayout()
        title = QLabel(f"Palawan Transactions – {field_label}")
        title.setStyleSheet("font-size: 15px; font-weight: 800; color: #1E293B;")
        top_bar.addWidget(title, 1)

        top_bar.addStretch()
        zoom_out_btn = QPushButton("−")
        zoom_out_btn.setFixedSize(26, 26)
        zoom_out_btn.setToolTip("Zoom out")
        zoom_out_btn.setStyleSheet(
            "QPushButton { font-size: 16px; font-weight: 700; border: 1px solid #CBD5E1;"
            " border-radius: 4px; background: #F1F5F9; color: #334155; }"
            "QPushButton:hover { background: #E2E8F0; }"
        )
        zoom_out_btn.clicked.connect(self._zoom_out)

        self._zoom_pct_label = QLabel("100%")
        self._zoom_pct_label.setFixedWidth(46)
        self._zoom_pct_label.setAlignment(Qt.AlignCenter)
        self._zoom_pct_label.setStyleSheet(
            "font-size: 12px; font-weight: 700; color: #64748B; background: transparent;"
        )

        zoom_in_btn = QPushButton("+")
        zoom_in_btn.setFixedSize(26, 26)
        zoom_in_btn.setToolTip("Zoom in")
        zoom_in_btn.setStyleSheet(
            "QPushButton { font-size: 16px; font-weight: 700; border: 1px solid #CBD5E1;"
            " border-radius: 4px; background: #F1F5F9; color: #334155; }"
            "QPushButton:hover { background: #E2E8F0; }"
        )
        zoom_in_btn.clicked.connect(self._zoom_in)

        top_bar.addWidget(zoom_out_btn)
        top_bar.addWidget(self._zoom_pct_label)
        top_bar.addWidget(zoom_in_btn)
        root.addLayout(top_bar)

        # Single scroll area wrapping ALL content except the Save/Cancel buttons.
        # This guarantees the buttons are always visible at the bottom.
        form_scroll = QScrollArea()
        form_scroll.setWidgetResizable(True)
        form_scroll.setStyleSheet("QScrollArea { border: none; background-color: white; }")

        form_container = QWidget()
        form_layout = QVBoxLayout(form_container)
        form_layout.setSpacing(10)
        form_layout.setContentsMargins(0, 0, 6, 0)

   
        basic_group = QGroupBox("Transaction Information")
        basic_group.setStyleSheet("""
            QGroupBox {
                border: 1px solid #E2E8F0;
                border-radius: 6px;
                margin-top: 12px;
                padding: 14px;
                font-weight: 600;
                font-size: 13px;
            }
            QGroupBox::title {
                padding: 0px 8px;
                font-weight: 700;
                font-size: 13px;
            }
        """)
        basic_form = QFormLayout()
        basic_form.setSpacing(10)
        basic_form.setContentsMargins(10, 10, 10, 10)
        basic_form.setLabelAlignment(Qt.AlignRight | Qt.AlignVCenter)

        # Transaction Code
        self.code_edit = QLineEdit()
        self.code_edit.setPlaceholderText("e.g., TXN001")
        self.code_edit.setMinimumHeight(36)
        basic_form.addRow("Transaction Code:", self.code_edit)

        # Receiver
        self.receiver_edit = QLineEdit()
        self.receiver_edit.setPlaceholderText("Receiver name")
        self.receiver_edit.setMinimumHeight(36)
        basic_form.addRow("Receiver:", self.receiver_edit)

        # Sender
        self.sender_edit = QLineEdit()
        self.sender_edit.setPlaceholderText("Sender name")
        self.sender_edit.setMinimumHeight(36)
        basic_form.addRow("Sender:", self.sender_edit)

        basic_group.setLayout(basic_form)
        form_layout.addWidget(basic_group)

        # ========== SECTION 2: AMOUNTS ==========
        amounts_group = QGroupBox("Transaction Amounts")
        amounts_group.setStyleSheet("""
            QGroupBox {
                border: 1px solid #E2E8F0;
                border-radius: 6px;
                margin-top: 12px;
                padding: 14px;
                font-weight: 600;
                font-size: 13px;
            }
            QGroupBox::title {
                padding: 0px 8px;
                font-weight: 700;
                font-size: 13px;
            }
        """)
        amounts_form = QFormLayout()
        amounts_form.setSpacing(10)
        amounts_form.setContentsMargins(10, 10, 10, 10)

        # Three columns for amounts
        amounts_h = QHBoxLayout()
        amounts_h.setSpacing(16)
        amounts_h.setContentsMargins(0, 0, 0, 0)

        # Principal column
        principal_v = QVBoxLayout()
        principal_v.setSpacing(6)
        principal_lbl = QLabel("Principal")
        principal_lbl.setStyleSheet("font-weight: 700; font-size: 13px; color: #1E293B;")
        self.principal_edit = QLineEdit()
        self.principal_edit.setValidator(QDoubleValidator(0.0, 1e12, 2))
        self.principal_edit.setPlaceholderText("0.00")
        self.principal_edit.setMinimumHeight(36)
        self.principal_edit.textChanged.connect(self._calculate_derived)
        principal_v.addWidget(principal_lbl)
        principal_v.addWidget(self.principal_edit)
        amounts_h.addLayout(principal_v, 1)

        # Commission column
        commission_v = QVBoxLayout()
        commission_v.setSpacing(6)
        commission_lbl = QLabel("Commission")
        commission_lbl.setStyleSheet("font-weight: 700; font-size: 13px; color: #1E293B;")
        self.commission_edit = QLineEdit()
        self.commission_edit.setValidator(QDoubleValidator(0.0, 1e12, 2))
        self.commission_edit.setPlaceholderText("0.00")
        self.commission_edit.setMinimumHeight(36)
        self.commission_edit.textChanged.connect(self._calculate_derived)
        commission_v.addWidget(commission_lbl)
        commission_v.addWidget(self.commission_edit)
        amounts_h.addLayout(commission_v, 1)

        # SC column
        sc_v = QVBoxLayout()
        sc_v.setSpacing(6)
        sc_lbl = QLabel("SC")
        sc_lbl.setStyleSheet("font-weight: 700; font-size: 13px; color: #1E293B;")
        self.sc_edit = QLineEdit()
        self.sc_edit.setValidator(QDoubleValidator(0.0, 1e12, 2))
        self.sc_edit.setPlaceholderText("0.00")
        self.sc_edit.setMinimumHeight(36)
        self.sc_edit.textChanged.connect(self._calculate_derived)
        sc_v.addWidget(sc_lbl)
        sc_v.addWidget(self.sc_edit)
        amounts_h.addLayout(sc_v, 1)

        amounts_form.addRow(amounts_h)
        amounts_group.setLayout(amounts_form)
        form_layout.addWidget(amounts_group)

        # ========== SECTION 3: CALCULATED FIELDS ==========
        calc_group = QGroupBox("Calculated Values")
        calc_group.setStyleSheet("""
            QGroupBox {
                border: 1px solid #E2E8F0;
                border-radius: 6px;
                margin-top: 12px;
                padding: 14px;
                font-weight: 600;
                font-size: 13px;
                background-color: #F8FAFC;
            }
            QGroupBox::title {
                padding: 0px 8px;
                font-weight: 700;
                font-size: 13px;
            }
        """)
        calc_form = QFormLayout()
        calc_form.setSpacing(10)
        calc_form.setContentsMargins(10, 10, 10, 10)

        # Three columns for calculated fields
        calc_h = QHBoxLayout()
        calc_h.setSpacing(16)
        calc_h.setContentsMargins(0, 0, 0, 0)

        # Total SC column
        total_sc_v = QVBoxLayout()
        total_sc_v.setSpacing(6)
        total_sc_lbl = QLabel("Total SC (Commission + SC)")
        total_sc_lbl.setStyleSheet("font-weight: 700; font-size: 13px; color: #0EA5E9;")
        self.total_sc_lbl = QLineEdit()
        self.total_sc_lbl.setReadOnly(True)
        self.total_sc_lbl.setText("0.00")
        self.total_sc_lbl.setStyleSheet("background-color: white; color: #0EA5E9; font-size: 13px; font-weight: 700; border: 1px solid #CBD5E1;")
        self.total_sc_lbl.setMinimumHeight(36)
        total_sc_v.addWidget(total_sc_lbl)
        total_sc_v.addWidget(self.total_sc_lbl)
        calc_h.addLayout(total_sc_v, 1)

        # Cash on Hand column (for Pay-Out only)
        cash_on_hand_v = QVBoxLayout()
        cash_on_hand_v.setSpacing(6)
        cash_on_hand_lbl = QLabel("Cash on Hand (Principal + Total SC)")
        cash_on_hand_lbl.setStyleSheet("font-weight: 700; font-size: 13px; color: #F59E0B;")
        self.cash_on_hand_lbl = QLineEdit()
        self.cash_on_hand_lbl.setReadOnly(True)
        self.cash_on_hand_lbl.setText("0.00")
        self.cash_on_hand_lbl.setStyleSheet("background-color: white; color: #F59E0B; font-size: 13px; font-weight: 700; border: 1px solid #CBD5E1;")
        self.cash_on_hand_lbl.setMinimumHeight(36)
        cash_on_hand_v.addWidget(cash_on_hand_lbl)
        cash_on_hand_v.addWidget(self.cash_on_hand_lbl)
        self.cash_on_hand_label = cash_on_hand_lbl  # Store label for show/hide
        self.cash_on_hand_widget_layout = cash_on_hand_v  # Store layout for show/hide
        calc_h.addLayout(cash_on_hand_v, 1)

        # Income column (dynamic percentage based on transaction type)
        income_v = QVBoxLayout()
        income_v.setSpacing(6)
        # Label will be updated in __init__ based on transaction_type
        self.income_label_text = QLabel("Income (49% of Commission)")
        self.income_label_text.setStyleSheet("font-weight: 700; font-size: 13px; color: #10B981;")
        self.income_lbl = QLineEdit()
        self.income_lbl.setReadOnly(True)
        self.income_lbl.setText("0.00")
        self.income_lbl.setStyleSheet("background-color: white; color: #10B981; font-size: 13px; font-weight: 700; border: 1px solid #CBD5E1;")
        self.income_lbl.setMinimumHeight(36)
        income_v.addWidget(self.income_label_text)
        income_v.addWidget(self.income_lbl)
        calc_h.addLayout(income_v, 1)

        # VAT (12%) column — only for Line of Business Group 1 branches
        vat_v = QVBoxLayout()
        vat_v.setSpacing(6)
        self._vat_label = QLabel("VAT (12%)")
        self._vat_label.setStyleSheet("font-weight: 700; font-size: 13px; color: #D97706;")
        self._vat_lbl = QLineEdit()
        self._vat_lbl.setReadOnly(True)
        self._vat_lbl.setText("0.00")
        self._vat_lbl.setStyleSheet(
            "background-color: white; color: #D97706; font-size: 13px;"
            "font-weight: 700; border: 1px solid #CBD5E1;"
        )
        self._vat_lbl.setMinimumHeight(36)
        vat_v.addWidget(self._vat_label)
        vat_v.addWidget(self._vat_lbl)
        calc_h.addLayout(vat_v, 1)

        # A/R Palawan or A/P Palawan column (depends on transaction type)
        ar_ap_v = QVBoxLayout()
        ar_ap_v.setSpacing(6)
        self.ar_ap_label_text = QLabel("A/R PALAWAN (Principal + Income)")
        self.ar_ap_label_text.setStyleSheet("font-weight: 700; font-size: 13px; color: #8B5CF6;")
        self.ar_palawan_lbl = QLineEdit()
        self.ar_palawan_lbl.setReadOnly(True)
        self.ar_palawan_lbl.setText("0.00")
        self.ar_palawan_lbl.setStyleSheet("background-color: white; color: #8B5CF6; font-size: 13px; font-weight: 700; border: 1px solid #CBD5E1;")
        self.ar_palawan_lbl.setMinimumHeight(36)
        ar_ap_v.addWidget(self.ar_ap_label_text)
        ar_ap_v.addWidget(self.ar_palawan_lbl)
        calc_h.addLayout(ar_ap_v, 1)

        calc_form.addRow(calc_h)
        calc_group.setLayout(calc_form)
        form_layout.addWidget(calc_group)

        # Update labels and visibility based on transaction type
        if self._transaction_type == 'so':
            # Send-Out: Cash on Hand shown, 39% income, A/P Palawan = Cash on Hand - Income
            self.income_label_text.setText("Income (39% of Commission)")
            self.ar_ap_label_text.setText("A/P PALAWAN (Cash on Hand - Income)")
            self.cash_on_hand_label.setVisible(True)
            self.cash_on_hand_lbl.setVisible(True)
        elif self._transaction_type == 'po':
            # Pay-Out: no Cash on Hand, 43% income, A/R Palawan = Principal + Income
            self.income_label_text.setText("Income (43% of Commission)")
            self.ar_ap_label_text.setText("A/R PALAWAN (Principal + Income)")
            self.cash_on_hand_label.setVisible(False)
            self.cash_on_hand_lbl.setVisible(False)
        else:
            # International: no Cash on Hand, 43% income, A/R Palawan = Principal + Income
            self.income_label_text.setText("Income (43% of Commission)")
            self.ar_ap_label_text.setText("A/R PALAWAN (Principal + Income)")
            self.cash_on_hand_label.setVisible(False)
            self.cash_on_hand_lbl.setVisible(False)

        # VAT field: only visible for Line of Business Group 1
        self._vat_label.setVisible(self._has_vat)
        self._vat_lbl.setVisible(self._has_vat)

        # ========== SECTION 4: DOCUMENTATION & DETAILS ==========
        doc_group = QGroupBox("Documentation & Transaction Details")
        doc_group.setStyleSheet("""
            QGroupBox {
                border: 1px solid #E2E8F0;
                border-radius: 6px;
                margin-top: 12px;
                padding: 14px;
                font-weight: 600;
                font-size: 13px;
            }
            QGroupBox::title {
                padding: 0px 8px;
                font-weight: 700;
                font-size: 13px;
            }
        """)
        doc_form = QFormLayout()
        doc_form.setSpacing(10)
        doc_form.setContentsMargins(10, 10, 10, 10)
        doc_form.setLabelAlignment(Qt.AlignRight | Qt.AlignVCenter)

        # KYC Docs
        self.kyc_docs_edit = QLineEdit()
        self.kyc_docs_edit.setPlaceholderText("e.g., Valid ID, Passport")
        self.kyc_docs_edit.setMinimumHeight(36)
        doc_form.addRow("KYC Docs/Photo ID:", self.kyc_docs_edit)

        # Business Name
        self.business_name_edit = QLineEdit()
        self.business_name_edit.setPlaceholderText("Business/Company name")
        self.business_name_edit.setMinimumHeight(36)
        doc_form.addRow("Business/Company Name:", self.business_name_edit)

        # Position in Company
        self.position_edit = QLineEdit()
        self.position_edit.setPlaceholderText("e.g., Owner, Manager, Employee")
        self.position_edit.setMinimumHeight(36)
        doc_form.addRow("Position in Company:", self.position_edit)

        # Relationship
        self.relationship_edit = QLineEdit()
        self.relationship_edit.setPlaceholderText("e.g., Friend, Family, Client")
        self.relationship_edit.setMinimumHeight(36)
        doc_form.addRow("Relationship to Sender/Receiver:", self.relationship_edit)

        # Source of Funds
        self.source_funds_edit = QLineEdit()
        self.source_funds_edit.setPlaceholderText("e.g., Salary, Business Income")
        self.source_funds_edit.setMinimumHeight(36)
        doc_form.addRow("Source of Funds:", self.source_funds_edit)

        # Purpose
        self.purpose_edit = QLineEdit()
        self.purpose_edit.setPlaceholderText("e.g., Payment for services, Gift")
        self.purpose_edit.setMinimumHeight(36)
        doc_form.addRow("Purpose of Transaction:", self.purpose_edit)

        # Evaluation
        self.evaluation_edit = QComboBox()
        self.evaluation_edit.addItems(["LOW RISK", "MEDIUM RISK", "HIGH RISK"])
        self.evaluation_edit.setMinimumHeight(36)
        doc_form.addRow("Evaluation:", self.evaluation_edit)

        doc_group.setLayout(doc_form)
        form_layout.addWidget(doc_group)

        # ── Action buttons row (inside scroll) ────────────────────────────────
        action_row = QHBoxLayout()
        action_row.setSpacing(10)

        add_btn = QPushButton("+ Add Transaction")
        add_btn.setStyleSheet("""
            QPushButton {
                background: #0EA5E9; color: white; border: none;
                border-radius: 6px; padding: 8px 20px;
                font-weight: 700; font-size: 13px;
                min-height: 34px;
            }
            QPushButton:hover { background: #0284C7; }
        """)
        add_btn.clicked.connect(self._add_transaction)
        action_row.addWidget(add_btn)

        if self._has_payroll:
            import_btn = QPushButton("↑ Import from Excel")
            import_btn.setStyleSheet("""
                QPushButton {
                    background: #16A34A; color: white; border: none;
                    border-radius: 6px; padding: 8px 20px;
                    font-weight: 700; font-size: 13px;
                    min-height: 34px;
                }
                QPushButton:hover { background: #15803D; }
            """)
            import_btn.setToolTip(
                "Import transactions from an Excel file.\n"
                "Required columns: Code, Principal, Commission, SC\n"
                "Optional: Receiver, Sender, Position, Relationship,\n"
                "          Source of Funds, Purpose, KYC Docs,\n"
                "          Business Name, Evaluation"
            )
            import_btn.clicked.connect(self._import_from_excel)
            action_row.addWidget(import_btn)

        action_row.addStretch()
        form_layout.addLayout(action_row)

        # ── Transactions table (inside scroll) ───────────────────────────────
        # 13 columns: col 8 = Cash on Hand (so only), col 10 = VAT (Group 1 only)
        _ar_ap_col_lbl = "A/P Palawan" if self._transaction_type == 'so' else "A/R Palawan"
        self.table = QTableWidget(0, 13)
        self.table.setHorizontalHeaderLabels([
            "Code", "Receiver", "Sender", "Position", "Principal", "Commission", "SC",
            "Total SC", "Cash on Hand", "Income", "VAT (12%)", _ar_ap_col_lbl, ""
        ])
        if self._transaction_type != 'so':
            self.table.setColumnHidden(8, True)
        if not self._has_vat:
            self.table.setColumnHidden(10, True)
        hh = self.table.horizontalHeader()
        hh.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        for i in range(1, 13):
            hh.setSectionResizeMode(i, QHeaderView.Stretch)
        hh.setSectionResizeMode(12, QHeaderView.ResizeToContents)
        self.table.setSelectionMode(QAbstractItemView.NoSelection)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self.table.setMinimumHeight(260)
        self.table.setStyleSheet("""
            QTableWidget { border: 1px solid #E2E8F0; border-radius: 6px; }
            QHeaderView::section { background: #F1F5F9; font-weight: 700;
                                   font-size: 11px; padding: 6px; border: none; }
            QTableWidget::item { padding: 4px 8px; }
        """)
        form_layout.addWidget(self.table, stretch=3)

        # ── Totals frame (inside scroll) ─────────────────────────────────────
        totals_frame = QFrame()
        totals_frame.setStyleSheet(
            "background:#F8FAFC; border:1px solid #E2E8F0;"
            "border-radius:8px; padding:8px;"
        )
        totals_layout = QHBoxLayout(totals_frame)
        totals_layout.setContentsMargins(12, 6, 12, 6)
        totals_layout.setSpacing(20)

        _ar_ap_lbl = "A/P Total" if self._transaction_type == 'so' else "A/R Total"
        _totals_fields = [
            ("Principal Total",    "_principal_total"),
            ("Commission Total",   "_commission_total"),
            ("SC Total",           "_sc_total"),
            ("Total SC",           "_total_sc_total"),
        ]
        if self._transaction_type == 'so':
            _totals_fields.append(("Cash on Hand Total", "_cash_on_hand_total"))
        _totals_fields.append(("Income Total", "_income_total"))
        if self._has_vat:
            _totals_fields.append(("VAT (12%)", "_vat_total"))
        _totals_fields.append((_ar_ap_lbl, "_ar_total"))
        for label_text, attr_name in _totals_fields:
            box = QVBoxLayout()
            lbl_title = QLabel(label_text)
            lbl_title.setStyleSheet("font-size: 12px; font-weight: 700; color: #64748B;")
            lbl_value = QLabel("0.00")
            lbl_value.setStyleSheet("font-size: 13px; font-weight: 800; color: #1E293B;")
            box.addWidget(lbl_title)
            box.addWidget(lbl_value)
            totals_layout.addLayout(box)
            setattr(self, attr_name, lbl_value)

        totals_layout.addStretch()
        form_layout.addWidget(totals_frame)
        form_layout.addStretch()

        # Attach scroll container
        form_scroll.setWidget(form_container)
        root.addWidget(form_scroll, stretch=1)

        # ── Dialog buttons — OUTSIDE scroll so always visible ─────────────────
        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        self._save_btn = btns.button(QDialogButtonBox.Ok)
        self._save_btn.setText("Save")
        self._save_btn.setStyleSheet(
            "background:#16A34A;color:white;border:none;border-radius:5px;"
            "padding:6px 18px;font-weight:700;min-height:36px;"
        )
        btns.button(QDialogButtonBox.Cancel).setStyleSheet(
            "background:#64748B;color:white;border:none;border-radius:5px;"
            "padding:6px 18px;font-weight:700;min-height:36px;"
        )
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        root.addWidget(btns)

        # Load existing transactions if provided
        if self._transactions:
            self._refresh_table()
        else:
            self._update_save_btn()

    def _update_save_btn(self):
        self._save_btn.setEnabled(True)
        self._save_btn.setStyleSheet(
            "background:#16A34A;color:white;border:none;border-radius:5px;"
            "padding:6px 18px;font-weight:700;min-height:36px;"
        )

    def accept(self):
        if not self._transactions:
            from PyQt5.QtWidgets import QMessageBox
            QMessageBox.warning(
                self, "No Transactions",
                "Please add at least one transaction before saving."
            )
        super().accept()

    def _calculate_derived(self):
        """Calculate derived fields based on transaction type."""
        try:
            commission = float(self.commission_edit.text() or 0)
            sc = float(self.sc_edit.text() or 0)
            principal = float(self.principal_edit.text() or 0)
        except ValueError:
            commission = sc = principal = 0.0

        total_sc = commission + sc
        income = commission * self._income_rate
        vat = income / 1.12 * 0.12

        self.total_sc_lbl.setText(f"{total_sc:.2f}")
        self.income_lbl.setText(f"{income:.2f}")
        if self._has_vat:
            self._vat_lbl.setText(f"{vat:.2f}")

        if self._transaction_type == 'so':
            # Send-Out: Cash on Hand = Principal + Total SC; A/R = Cash on Hand - Income
            cash_on_hand = principal + total_sc
            self.cash_on_hand_lbl.setText(f"{cash_on_hand:.2f}")
            self.ar_palawan_lbl.setText(f"{cash_on_hand - income:.2f}")
        elif self._transaction_type == 'po':
            # Pay-Out: no Cash on Hand; A/R = Principal + Income
            self.ar_palawan_lbl.setText(f"{principal + income:.2f}")
        else:
            # International: A/R = Principal + Income
            self.ar_palawan_lbl.setText(f"{principal + income:.2f}")

    def _add_transaction(self):
        """Add the current form data as a transaction."""
        try:
            principal = float(self.principal_edit.text() or 0)
            commission = float(self.commission_edit.text() or 0)
            sc = float(self.sc_edit.text() or 0)
        except ValueError:
            principal = commission = sc = 0.0

        _income = commission * self._income_rate
        txn = {
            "code": self.code_edit.text().strip(),
            "receiver": self.receiver_edit.text().strip(),
            "sender": self.sender_edit.text().strip(),
            "principal": principal,
            "commission": commission,
            "sc": sc,
            "total_sc": commission + sc,
            "income": _income,
            "vat": _income / 1.12 * 0.12,
            # so: Cash on Hand - Income  |  po/int: Principal + Income
            "ar_palawan": (
                principal + commission + sc - _income
                if self._transaction_type == 'so' else
                principal + _income
            ),
            "kyc_docs": self.kyc_docs_edit.text().strip(),
            "business_name": self.business_name_edit.text().strip(),
            "position": self.position_edit.text().strip(),
            "relationship": self.relationship_edit.text().strip(),
            "source_funds": self.source_funds_edit.text().strip(),
            "purpose": self.purpose_edit.text().strip(),
            "evaluation": self.evaluation_edit.currentText(),
        }

        if txn["code"]:  # Code is required
            # Prevent duplicate transaction codes
            existing_codes = {t.get("code", "") for t in self._transactions}
            if txn["code"] in existing_codes:
                from PyQt5.QtWidgets import QMessageBox
                QMessageBox.warning(self, "Duplicate Code",
                    f"Transaction code '{txn['code']}' already exists in this list.\n"
                    "Please use a different code or delete the existing entry first.")
                return
            self._transactions.append(txn)
            self._refresh_table()
            self._clear_form()

    def _clear_form(self):
        """Clear the form fields."""
        self.code_edit.clear()
        self.receiver_edit.clear()
        self.sender_edit.clear()
        self.principal_edit.clear()
        self.commission_edit.clear()
        self.sc_edit.clear()
        self.kyc_docs_edit.clear()
        self.business_name_edit.clear()
        self.position_edit.clear()
        self.relationship_edit.clear()
        self.source_funds_edit.clear()
        self.purpose_edit.clear()
        self.evaluation_edit.setCurrentIndex(0)
        self._calculate_derived()

    # ── Column keyword mapping for Excel import ───────────────────────────────
    _IMPORT_KEYWORDS = {
        "type":         ["type", "transaction type", "txn type", "trans type",
                         "trans. type", "category", "kind"],
        "code":         ["code", "txn", "transaction code", "control", "reference",
                         "ref no", "control no", "trans. code", "trans code"],
        "receiver":     ["receiver", "beneficiary", "recipient", "payee"],
        "sender":       ["sender", "remitter", "originator", "remittance by"],
        "principal":    ["principal", "amount", "principal amount", "remittance amount",
                         "capital"],
        "commission":   ["commission", "comm", "comm.", "fee", "service fee"],
        "sc":           ["sc", "service charge", "s.c", "s.c.", "svc charge"],
        "kyc_docs":     ["kyc", "photo id", "kyc docs", "identification", "id presented",
                         "valid id"],
        "business_name":["business", "company", "business name", "company name",
                         "business/company"],
        "position":     ["position", "position in company", "job", "occupation",
                         "position in the"],
        "relationship": ["relationship", "relation", "relationship to",
                         "relationship to sender", "relationship to receiver"],
        "source_funds": ["source", "source of funds", "fund source", "funds",
                         "source of fund"],
        "purpose":      ["purpose", "reason", "purpose of transaction",
                         "nature of transaction", "purpose of trans"],
        "evaluation":   ["evaluation", "risk", "risk level", "remarks", "risk evaluation"],
        "vat":          ["vat", "vat amount", "vat (12%)", "vat(12%)", "value added tax",
                         "tax", "vat 12%"],
    }

    # Keywords that identify each transaction type value inside the "type" column
    _TYPE_KEYWORDS = {
        "so":  ["send", "sendout", "send out", "send-out", "so", "remittance"],
        "po":  ["pay", "payout", "pay out", "pay-out", "po"],
        "int": ["international", "intl", "int"],
    }

    @staticmethod
    def _match_column(header: str) -> "str | None":
        """Return the field name for an Excel column header, or None if unrecognised."""
        h = header.lower().strip()
        if not h:
            return None
        for field, keywords in PalawanTransactionDetailDialog._IMPORT_KEYWORDS.items():
            for kw in keywords:
                # Exact match
                if h == kw:
                    return field
                # Header starts with keyword  e.g. "position in company" → "position"
                if h.startswith(kw):
                    return field
                # Keyword starts with header  e.g. "relation" → "relationship",
                #                                  "commiss" → "commission"
                if kw.startswith(h):
                    return field
        return None

    def _classify_type(self, raw_value: str) -> "str | None":
        """Return 'so', 'po', 'int', or None for an unrecognised type cell value."""
        v = raw_value.lower().strip()
        for txn_type, keywords in self._TYPE_KEYWORDS.items():
            for kw in keywords:
                if kw in v or v in kw:
                    return txn_type
        return None

    def _import_from_excel(self):
        """Import transactions from an Excel file.

        Handles three formats:
        - Single-type file: all rows are one transaction type
        - Side-by-side combined: separate column groups per type in the same sheet
          e.g. "RELEASED (PAYOUT)" columns on the left,
               "REMITTANCE (SENDOUT)" columns on the right
        - Type-column combined: a dedicated column identifies each row's type
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            from PyQt5.QtWidgets import QMessageBox
            QMessageBox.critical(
                self, "Missing Dependency",
                "The openpyxl package is required to import Excel files.\n"
                "Install it with:  pip install openpyxl"
            )
            return

        from PyQt5.QtWidgets import QFileDialog, QMessageBox

        file_path, _ = QFileDialog.getOpenFileName(
            self, "Select Excel File", "",
            "Excel Files (*.xlsx *.xls *.xlsm);;All Files (*)"
        )
        if not file_path:
            return

        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
        except Exception as exc:
            QMessageBox.critical(self, "File Error", f"Could not open the file:\n{exc}")
            return

        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not rows:
            QMessageBox.warning(self, "Empty File", "The selected file has no data.")
            return

        # ── Step 1: detect side-by-side section headers ───────────────────────
        # Scan the first 5 rows for cells that identify each section
        # e.g. "RELEASED (PAYOUT)"  → po
        #      "REMITTANCE (SENDOUT)" → so
        _SECTION_MARKERS = {
            'so':  ('sendout', 'send-out', 'send out', 'remittance'),
            'po':  ('payout',  'pay-out',  'pay out',  'released'),
            'int': ('international', 'intl'),
        }
        section_col_starts = {}   # type_key → column index where that section title starts
        for row in rows[:5]:
            for col_idx, cell in enumerate(row):
                if cell is None:
                    continue
                v = str(cell).lower()
                for t_key, markers in _SECTION_MARKERS.items():
                    if t_key not in section_col_starts and any(m in v for m in markers):
                        section_col_starts[t_key] = col_idx

        # ── Step 2: determine column slice for the target type ────────────────
        target    = self._transaction_type   # 'so', 'po', or 'int'
        col_start = 0       # inclusive
        col_end   = None    # exclusive; None = to end of row

        if len(section_col_starts) >= 2:
            # Side-by-side layout detected — restrict to matching section columns
            sorted_secs = sorted(section_col_starts.items(), key=lambda x: x[1])
            found_target = False
            for i, (sec_type, sec_col) in enumerate(sorted_secs):
                if sec_type == target:
                    col_start    = sec_col
                    col_end      = sorted_secs[i + 1][1] if i + 1 < len(sorted_secs) else None
                    found_target = True
                    break

            if not found_target:
                _lbl = {"so": "Send-Out", "po": "Pay-Out", "int": "International"}
                found_names = [_lbl.get(k, k) for k in section_col_starts]
                QMessageBox.warning(
                    self, "Section Not Found",
                    f"This file contains sections for: {', '.join(found_names)}\n\n"
                    f"No '{_lbl.get(target, target)}' section was detected.\n"
                    "Please open the correct dialog before importing."
                )
                return

        def _slice(row):
            s = row[col_start:col_end] if col_end is not None else row[col_start:]
            return s

        # ── Step 3: find header row inside the column slice ───────────────────
        best_row_idx = 0
        best_score   = -1
        for idx, raw_row in enumerate(rows):
            sliced = _slice(raw_row)
            score  = sum(
                1 for cell in sliced
                if cell and self._match_column(str(cell)) is not None
            )
            if score > best_score:
                best_score   = score
                best_row_idx = idx

        if best_score == 0:
            QMessageBox.warning(
                self, "No Recognised Columns",
                "Could not find recognised column headers in the file.\n\n"
                "Expected columns include:\n"
                "Code, Principal, Commission, SC, Receiver, Sender,\n"
                "KYC Docs, Business Name, Position, Relationship,\n"
                "Source of Funds, Purpose, Evaluation"
            )
            return

        # Build col_map relative to the slice (0-based within the slice)
        header_sliced = _slice(rows[best_row_idx])
        col_map = {}
        for col_idx, cell in enumerate(header_sliced):
            if cell is None:
                continue
            field = self._match_column(str(cell))
            if field and field not in col_map:
                col_map[field] = col_idx

        # ── Step 4: check required columns ───────────────────────────────────
        missing = [f for f in ("code", "principal", "commission", "sc") if f not in col_map]
        if missing:
            QMessageBox.warning(
                self, "Missing Required Columns",
                "The following required columns were not found:\n"
                + ", ".join(missing).upper()
                + "\n\nPlease make sure your Excel file has these headers."
            )
            return

        has_type_col = "type" in col_map

        # ── Step 5: parse data rows ───────────────────────────────────────────
        existing_codes = {t.get("code", "") for t in self._transactions}
        imported      = 0
        skipped_dup   = []
        skipped_empty = 0
        skipped_type  = 0

        def _str(sliced_row, field, default=""):
            idx = col_map.get(field)
            if idx is None or idx >= len(sliced_row):
                return default
            val = sliced_row[idx]
            return str(val).strip() if val is not None else default

        def _num(sliced_row, field):
            idx = col_map.get(field)
            if idx is None or idx >= len(sliced_row):
                return 0.0
            val = sliced_row[idx]
            if val is None:
                return 0.0
            try:
                return float(str(val).replace(",", "").strip() or 0)
            except (ValueError, TypeError):
                return 0.0

        for raw_row in rows[best_row_idx + 1:]:
            sliced = _slice(raw_row)

            # Skip completely empty rows in this section
            if all(cell is None or str(cell).strip() == "" for cell in sliced):
                skipped_empty += 1
                continue

            # If there is a type column, filter by it (type-column combined files)
            if has_type_col:
                raw_type = _str(sliced, "type")
                if raw_type and self._classify_type(raw_type) != target:
                    skipped_type += 1
                    continue

            code       = _str(sliced, "code")
            principal  = _num(sliced, "principal")
            commission = _num(sliced, "commission")
            sc         = _num(sliced, "sc")

            # Skip rows that are effectively empty data
            if not code and principal == 0 and commission == 0:
                skipped_empty += 1
                continue

            # Skip duplicate codes
            if code and code in existing_codes:
                skipped_dup.append(code)
                continue

            # Calculate derived fields
            total_sc = commission + sc
            income   = commission * self._income_rate
            # Use VAT from Excel if the column is present and non-zero,
            # otherwise calculate it (income / 1.12 × 12%)
            if "vat" in col_map:
                vat = _num(sliced, "vat") or (income / 1.12 * 0.12)
            else:
                vat = income / 1.12 * 0.12
            ar_palawan = (
                (principal + total_sc) - income   # so: cash_on_hand - income
                if target == 'so' else
                principal + income                 # po / int: principal + income
            )

            eval_raw = _str(sliced, "evaluation", "LOW RISK").upper()
            if "HIGH" in eval_raw:
                evaluation = "HIGH RISK"
            elif "MEDIUM" in eval_raw or "MED" in eval_raw:
                evaluation = "MEDIUM RISK"
            else:
                evaluation = "LOW RISK"

            txn = {
                "code":          code,
                "receiver":      _str(sliced, "receiver"),
                "sender":        _str(sliced, "sender"),
                "principal":     principal,
                "commission":    commission,
                "sc":            sc,
                "total_sc":      total_sc,
                "income":        income,
                "vat":           vat,
                "ar_palawan":    ar_palawan,
                "kyc_docs":      _str(sliced, "kyc_docs"),
                "business_name": _str(sliced, "business_name"),
                "position":      _str(sliced, "position"),
                "relationship":  _str(sliced, "relationship"),
                "source_funds":  _str(sliced, "source_funds"),
                "purpose":       _str(sliced, "purpose"),
                "evaluation":    evaluation,
            }
            self._transactions.append(txn)
            if code:
                existing_codes.add(code)
            imported += 1

        # ── Step 6: refresh UI and show summary ───────────────────────────────
        if imported:
            self._refresh_table()

        _lbl = {"so": "Send-Out", "po": "Pay-Out", "int": "International"}
        current_label = _lbl.get(target, target)
        side_by_side  = len(section_col_starts) >= 2

        parts = [f"{imported} {current_label} transaction(s) imported successfully."]
        if side_by_side:
            parts[0] += f"  (read from '{current_label}' section of combined file)"
        if has_type_col and skipped_type:
            parts.append(f"{skipped_type} row(s) skipped — not {current_label} type.")
        if skipped_dup:
            parts.append(
                f"{len(skipped_dup)} skipped (duplicate codes): "
                + ", ".join(skipped_dup[:5])
                + (" …" if len(skipped_dup) > 5 else "")
            )
        if skipped_empty:
            parts.append(f"{skipped_empty} empty row(s) skipped.")

        if imported == 0 and not skipped_dup:
            QMessageBox.warning(
                self, "No Data Imported",
                f"No valid {current_label} rows were found in the file.\n\n"
                + ("Check that the section columns contain data rows."
                   if side_by_side else
                   "Please check that the file has data rows below the header.")
            )
        else:
            QMessageBox.information(self, "Import Complete", "\n".join(parts))

    def _refresh_table(self):

        self.table.setRowCount(0)

        p_total = c_total = s_total = ts_total = i_total = vat_total = ar_total = coh_total = 0.0

        for txn in self._transactions:
            row = self.table.rowCount()
            self.table.insertRow(row)

            _p    = float(txn.get("principal", 0))
            _ts   = float(txn.get("total_sc", 0))
            _inc  = float(txn.get("income", 0))
            _vat  = float(txn.get("vat", _inc / 1.12 * 0.12))

            self.table.setItem(row, 0, QTableWidgetItem(txn.get("code", "")))
            self.table.setItem(row, 1, QTableWidgetItem(txn.get("receiver", "")))
            self.table.setItem(row, 2, QTableWidgetItem(txn.get("sender", "")))
            self.table.setItem(row, 3, QTableWidgetItem(txn.get("position", "")))
            self.table.setItem(row, 4, QTableWidgetItem(f"{_p:.2f}"))
            self.table.setItem(row, 5, QTableWidgetItem(f"{txn.get('commission', 0):.2f}"))
            self.table.setItem(row, 6, QTableWidgetItem(f"{txn.get('sc', 0):.2f}"))
            self.table.setItem(row, 7, QTableWidgetItem(f"{_ts:.2f}"))
            # col 8: Cash on Hand (sendout only)
            self.table.setItem(row, 8, QTableWidgetItem(f"{_p + _ts:.2f}" if self._transaction_type == 'so' else ""))
            self.table.setItem(row, 9, QTableWidgetItem(f"{_inc:.2f}"))
            # col 10: VAT (Group 1 only — column hidden for other branches)
            self.table.setItem(row, 10, QTableWidgetItem(f"{_vat:.2f}"))
            self.table.setItem(row, 11, QTableWidgetItem(f"{txn.get('ar_palawan', 0):.2f}"))

            # Delete button at col 12
            rem_btn = QPushButton("✕")
            rem_btn.setFixedWidth(28)
            rem_btn.setStyleSheet(
                "QPushButton { color: #EF4444; font-weight: 900; border: none; font-size: 13px; }"
                "QPushButton:hover { color: #DC2626; }"
            )
            rem_btn.clicked.connect(lambda checked=False, r=row: self._remove_transaction(r))
            self.table.setCellWidget(row, 12, rem_btn)

            # Accumulate totals
            p_total   += txn.get("principal", 0)
            c_total   += txn.get("commission", 0)
            s_total   += txn.get("sc", 0)
            ts_total  += txn.get("total_sc", 0)
            i_total   += _inc
            vat_total += _vat
            ar_total  += txn.get("ar_palawan", 0)
            coh_total += _p + _ts

        # Update total labels
        self._principal_total.setText(f"{p_total:.2f}")
        self._commission_total.setText(f"{c_total:.2f}")
        self._sc_total.setText(f"{s_total:.2f}")
        self._total_sc_total.setText(f"{ts_total:.2f}")
        if self._transaction_type == 'so':
            self._cash_on_hand_total.setText(f"{coh_total:.2f}")
        self._income_total.setText(f"{i_total:.2f}")
        if self._has_vat:
            self._vat_total.setText(f"{vat_total:.2f}")
        self._ar_total.setText(f"{ar_total:.2f}")

        self._update_save_btn()

    def _remove_transaction(self, row):
        """Remove a transaction by row index."""
        if 0 <= row < len(self._transactions):
            self._transactions.pop(row)
            self._refresh_table()

    def get_transactions_data(self) -> list:
        """Get all transactions with complete details."""
        return self._transactions

    def get_total_principal(self) -> float:
        """Get total principal amount from all transactions."""
        return sum(t.get("principal", 0) for t in self._transactions)

    def get_total_commission(self) -> float:
        """Get total commission amount from all transactions."""
        return sum(t.get("commission", 0) for t in self._transactions)

    def get_total_sc(self) -> float:
        """Get total SC amount from all transactions."""
        return sum(t.get("sc", 0) for t in self._transactions)

    def get_total_amount(self) -> float:
        """Get the primary total (for compatibility)."""
        return self.get_total_principal()

    # ── Zoom ─────────────────────────────────────────────────────────────────

    def eventFilter(self, obj, event):
        if event.type() == QEvent.Wheel:
            if event.modifiers() & Qt.ControlModifier:
                delta = event.angleDelta().y()
                if delta > 0:
                    self._zoom_in()
                else:
                    self._zoom_out()
                return True
        return super().eventFilter(obj, event)

    def wheelEvent(self, event):
        if event.modifiers() & Qt.ControlModifier:
            delta = event.angleDelta().y()
            if delta > 0:
                self._zoom_in()
            else:
                self._zoom_out()
            event.accept()
        else:
            super().wheelEvent(event)

    def _zoom_in(self):
        if self._zoom_factor < 2.0:
            self._zoom_factor = round(self._zoom_factor + 0.1, 1)
            self._apply_zoom()

    def _zoom_out(self):
        if self._zoom_factor > 0.6:
            self._zoom_factor = round(self._zoom_factor - 0.1, 1)
            self._apply_zoom()

    def _apply_zoom(self):
        import re
        from PyQt5.QtWidgets import QLineEdit as _QLE, QComboBox as _QCB, QPushButton as _QPB

        # Capture every widget's original stylesheet once (before any zoom modifies them)
        if self._orig_widget_styles is None:
            self._orig_widget_styles = {}
            for w in self.findChildren(QLabel) + self.findChildren(QGroupBox) + \
                     self.findChildren(_QLE) + self.findChildren(_QCB) + \
                     self.findChildren(_QPB):
                ss = w.styleSheet()
                if ss and 'font-size' in ss:
                    self._orig_widget_styles[id(w)] = (w, ss)

        # Recalculate dialog-level base sizes
        base = 13
        new_base = max(9, round(base * self._zoom_factor))
        self.setStyleSheet(f"""
            QLabel       {{ font-size: {new_base}px; font-weight: 600; color: #1E293B; }}
            QLineEdit    {{ font-size: {new_base}px; }}
            QComboBox    {{ font-size: {new_base}px; }}
            QPushButton  {{ font-size: {new_base}px; }}
        """)

        # Scale every individually-styled widget proportionally from its original
        def _scale(m):
            orig_px = int(m.group(1))
            return f'font-size: {max(9, round(orig_px * self._zoom_factor))}px'

        for widget, orig_ss in self._orig_widget_styles.values():
            widget.setStyleSheet(re.sub(r'font-size:\s*(\d+)px', _scale, orig_ss))

        self._zoom_pct_label.setText(f"{int(self._zoom_factor * 100)}%")
