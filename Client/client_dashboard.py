import datetime
import json
import logging
import os
import sys
import time
import re

logger = logging.getLogger(__name__)

from PyQt5.QtWidgets import (
    QWidget, QLabel, QLineEdit, QVBoxLayout, QHBoxLayout,
    QGroupBox, QPushButton, QSizePolicy, QDateEdit,
    QMessageBox, QScrollArea, QFrame, QGridLayout, QTabWidget,
    QComboBox, QApplication, QDialog, QTableWidget, QTableWidgetItem,
    QHeaderView, QDialogButtonBox, QAbstractItemView, QFileDialog,
    QProgressBar, QGraphicsOpacityEffect
)
from PyQt5.QtGui import QDoubleValidator, QFont, QFontDatabase, QTextDocument, QMovie
from PyQt5.QtCore import Qt, QDate, pyqtSignal, QTimer, QPropertyAnimation
from PyQt5.QtPrintSupport import QPrinter, QPrintDialog

from Client.cash_flow_tab import CashFlowTab
from Client.palawan_details_tab import PalawanDetailsTab
from security import SessionManager
from Client.ui_components import LoadingOverlay, NoWheelDateEdit
from Client.ui_styles import (
    _SLATE_50, _SLATE_100, _SLATE_200, _SLATE_300, _SLATE_400,
    _SLATE_500, _SLATE_600, _SLATE_700, _SLATE_800, _SLATE_900,
    _INDIGO_50, _INDIGO_100, _INDIGO_400, _INDIGO_500, _INDIGO_600,
    _INDIGO_700, _EMERALD_50, _EMERALD_400, _EMERALD_500, _EMERALD_600,
    _AMBER_400, _AMBER_500, _RED_400, _RED_500, _WHITE,
    _BG_APP, _BG_CARD, _BG_INPUT, _BG_RDONLY, _BG_HEADER, _BORDER,
    _TEXT_PRI, _TEXT_SEC, _TEXT_MUTED, _PRIMARY, _PRIMARY_DK,
    _PRIMARY_PR, _SUCCESS, _SUCCESS_DK, _build_global_qss
)
from Client.ui_scaling import _sz


try:
    from offline_manager import offline_manager
    OFFLINE_SUPPORT = True
except ImportError:
    OFFLINE_SUPPORT = False
    offline_manager = None

try:
    from auto_updater import check_for_updates, check_update_success
    from version import __version__
    AUTO_UPDATE_ENABLED = True
except ImportError:
    AUTO_UPDATE_ENABLED = False
    __version__ = "1.0.0"
    check_update_success = None


def _s(px: int) -> int:
    return _sz(px)


def _hline():
    f = QFrame()
    f.setFrameShape(QFrame.HLine)
    f.setFrameShadow(QFrame.Plain)
    f.setStyleSheet(f"color: {_BORDER}; max-height: 1px; background: {_BORDER};")
    return f


def _vline(dark=False):
    f = QFrame()
    f.setFrameShape(QFrame.VLine)
    f.setFrameShadow(QFrame.Plain)
    col = "#1E293B" if dark else _BORDER
    f.setStyleSheet(f"color: {col}; max-width: 1px; background: {col};")
    return f


def _micro_label(text, color=_TEXT_MUTED):
    lbl = QLabel(text.upper())
    lbl.setStyleSheet(
        f"font-size: 11px; font-weight: 700; color: {color}; "
        f"letter-spacing: 1.1px; background: transparent;"
    )
    return lbl


from connection_watcher import ConnectionBanner


class FundTransferHODialog(QDialog):

    BANK_ACCOUNTS = [
        {"id": 1, "bank_name": "CIB-BDO", "account_name": "Global Reliance", "account_number": "0077-9002-3923"},
        {"id": 2, "bank_name": "CIB-BPI", "account_name": "Kristal Clear Diamond and Gold Pawnshop", "account_number": "0091-0692-29"},
        {"id": 3, "bank_name": "CIB-BDO", "account_name": "Kristal Clear", "account_number": "0077-9001-8784"},
        {"id": 4, "bank_name": "CIB-Union Bank", "account_name": "Golbal Reliance Mgmt and Holdings Corp", "account_number": "0015-6000-5790"},
        {"id": 5, "bank_name": "CIB-BDO", "account_name": "Europacific Management & Holdings Corp", "account_number": "0038-1801-5838"},
        {"id": 6, "bank_name": "CIB-BPI", "account_name": "Europacific Management & Holdings Corp", "account_number": "3541-0035-67"},
        {"id": 7, "bank_name": "CIB-UB", "account_name": "Europacific Management & Holdings Corp", "account_number": "0021-7001-7921"},
        {"id": 8, "bank_name": "CIB-UB", "account_name": "BPI BILLS  PAYMENT SAN RAMON", "account_number": ""},
        {"id": 9, "bank_name": "CIB-UB", "account_name": "BPI  BILLS PAYMENT SILVERSTAR", "account_number": ""},
        {"id": 10, "bank_name": "CIB-UB", "account_name": "BPI  BILLS PAYMENT ALLEXITE", "account_number": ""},
        {"id": 11, "bank_name": "CIB-UB", "account_name": "BPI BILLS PAYMENT MEGAWORLD", "account_number": ""},
        {"id": 12, "bank_name": "CIB-UB", "account_name": "BPI BILLS PAYMENT HOMENEEDS", "account_number": ""},
        {"id": 13, "bank_name": "CIB-UB", "account_name": "BPI  BILLS PAYMENT KRISTAL CLEAR", "account_number": ""},
        {"id": 14, "bank_name": "CIB-UB", "account_name": "BPI BILLS PAYMENT SAFELOCK", "account_number": ""},
    ]

    def __init__(self, field_label="Fund Transfer to HEAD OFFICE", parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"Detail – {field_label}")
        self.setMinimumSize(620, 460)
        self.setModal(True)
        self._rows_data = [] 

        root = QVBoxLayout(self)
        root.setSpacing(12)
        root.setContentsMargins(16, 16, 16, 16)

        title = QLabel(f"Breakdown for  {field_label}")
        title.setStyleSheet("font-size: 15px; font-weight: 800; color: #1E293B;")
        root.addWidget(title)

        self.table = QTableWidget(0, 3)
        self.table.setHorizontalHeaderLabels(["Bank Account", "Amount", ""])
        hh = self.table.horizontalHeader()
        hh.setSectionResizeMode(0, QHeaderView.Stretch)
        hh.setSectionResizeMode(1, QHeaderView.Stretch)
        hh.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.table.setSelectionMode(QAbstractItemView.NoSelection)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self.table.setStyleSheet("""
            QTableWidget { border: 1px solid #E2E8F0; border-radius: 6px; }
            QHeaderView::section { background: #F1F5F9; font-weight: 700;
                                   font-size: 11px; padding: 6px; border: none; }
            QTableWidget::item { padding: 4px 8px; }
        """)
        root.addWidget(self.table)

        add_btn = QPushButton("+ Add Transfer")
        add_btn.setStyleSheet("""
            QPushButton { background: #8B5CF6; color: white; border: none;
                          border-radius: 6px; padding: 7px 18px;
                          font-weight: 700; font-size: 12px; }
            QPushButton:hover { background: #7C3AED; }
        """)
        add_btn.clicked.connect(self._add_row)
        root.addWidget(add_btn, alignment=Qt.AlignLeft)

        totals_frame = QFrame()
        totals_frame.setStyleSheet(
            "background:#F8FAFC; border:1px solid #E2E8F0;"
            "border-radius:8px; padding:6px;"
        )
        totals_layout = QHBoxLayout(totals_frame)
        totals_layout.setContentsMargins(12, 8, 12, 8)
        totals_layout.setSpacing(40)

        ta_box = QVBoxLayout()
        ta_title = QLabel("Total Amount")
        ta_title.setStyleSheet("font-size: 11px; font-weight: 700; color: #64748B;")
        self._total_amount_lbl = QLabel("0.00")
        self._total_amount_lbl.setStyleSheet("font-size: 18px; font-weight: 800; color: #8B5CF6;")
        ta_sub = QLabel("(carried to field)")
        ta_sub.setStyleSheet("font-size: 12px; color: #94A3B8;")
        ta_box.addWidget(ta_title)
        ta_box.addWidget(self._total_amount_lbl)
        ta_box.addWidget(ta_sub)

        count_box = QVBoxLayout()
        count_title = QLabel("Transactions")
        count_title.setStyleSheet("font-size: 11px; font-weight: 700; color: #64748B;")
        self._count_lbl = QLabel("0")
        self._count_lbl.setStyleSheet("font-size: 18px; font-weight: 800; color: #2563EB;")
        count_sub = QLabel("(number of transfers)")
        count_sub.setStyleSheet("font-size: 12px; color: #94A3B8;")
        count_box.addWidget(count_title)
        count_box.addWidget(self._count_lbl)
        count_box.addWidget(count_sub)

        totals_layout.addLayout(ta_box)
        totals_layout.addLayout(count_box)
        totals_layout.addStretch()
        root.addWidget(totals_frame)

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.button(QDialogButtonBox.Ok).setText("Post")
        btns.button(QDialogButtonBox.Ok).setStyleSheet(
            "background:#16A34A;color:white;border:none;border-radius:5px;"
            "padding:6px 18px;font-weight:700;"
        )
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        root.addWidget(btns)

        self._add_row()

    def _add_row(self):
        row = self.table.rowCount()
        self.table.insertRow(row)

        bank_combo = QComboBox()
        bank_combo.setMinimumWidth(200)
        for bank in self.BANK_ACCOUNTS:
            display = f"{bank['bank_name']} - {bank['account_name']}"
            bank_combo.addItem(display, bank['id'])
        bank_combo.setStyleSheet("padding: 4px 6px; font-size: 12px;")
        self.table.setCellWidget(row, 0, bank_combo)

        amt_edit = QLineEdit()
        amt_edit.setPlaceholderText("0.00")
        amt_edit.setValidator(QDoubleValidator(0.0, 1e12, 2))
        amt_edit.setStyleSheet(
            "padding: 4px 8px; font-size: 13px; font-weight: 600;"
        )
        self.table.setCellWidget(row, 1, amt_edit)

        rem_btn = QPushButton("✕")
        rem_btn.setFixedWidth(28)
        rem_btn.setStyleSheet(
            "QPushButton { color: #EF4444; font-weight: 900; border: none; font-size: 13px; }"
            "QPushButton:hover { color: #DC2626; }"
        )
        rem_btn.clicked.connect(lambda _, b=rem_btn: self._remove_row_by_widget(b))
        self.table.setCellWidget(row, 2, rem_btn)

        self._rows_data.append((bank_combo, amt_edit))
        amt_edit.textChanged.connect(self._recalc)
        self.table.resizeRowsToContents()

    def _remove_row_by_widget(self, btn):
        for r in range(self.table.rowCount()):
            if self.table.cellWidget(r, 2) is btn:
                combo = self.table.cellWidget(r, 0)
                edit = self.table.cellWidget(r, 1)
                self._rows_data = [
                    (c, e) for c, e in self._rows_data
                    if c is not combo
                ]
                self.table.removeRow(r)
                self._recalc()
                return

    def _recalc(self, *_):
        total_amt = 0.0
        for bank_combo, amt_edit in self._rows_data:
            try:
                amt = float(amt_edit.text().strip().replace(',', '') or 0)
            except ValueError:
                amt = 0.0
            total_amt += amt
        self._total_amount_lbl.setText(f"{total_amt:,.2f}")
        self._count_lbl.setText(str(len(self._rows_data)))

    def get_total_amount(self) -> float:
        try:
            return float(self._total_amount_lbl.text().replace(',', ''))
        except ValueError:
            return 0.0

    def get_row_count(self) -> int:
        return len(self._rows_data)

    def get_breakdown_data(self) -> list:

        data = []
        for bank_combo, amt_edit in self._rows_data:
            try:
                amt = float(amt_edit.text().strip().replace(',', '') or 0)
            except ValueError:
                amt = 0.0
            bank_display = bank_combo.currentText()
            bank_id = bank_combo.currentData()
            data.append([bank_display, bank_id, amt])
        return data

class MotorCarDetailDialog(QDialog):

    PERCENTAGES = ["10.0%", "20.0%"]

    def __init__(self, field_label, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"Detail – {field_label}")
        self.setMinimumSize(580, 460)
        self.setModal(True)
        self._rows_data = []

        root = QVBoxLayout(self)
        root.setSpacing(12)
        root.setContentsMargins(16, 16, 16, 16)

        title = QLabel(f"Breakdown for  {field_label}")
        title.setStyleSheet("font-size: 15px; font-weight: 800; color: #1E293B;")
        root.addWidget(title)

        self.table = QTableWidget(0, 4)
        self.table.setHorizontalHeaderLabels(["Percentage", "Amount", "Computed", ""])
        hh = self.table.horizontalHeader()
        hh.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        hh.setSectionResizeMode(1, QHeaderView.Stretch)
        hh.setSectionResizeMode(2, QHeaderView.Stretch)
        hh.setSectionResizeMode(3, QHeaderView.ResizeToContents)
        self.table.setSelectionMode(QAbstractItemView.NoSelection)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self.table.setStyleSheet("""
            QTableWidget { border: 1px solid #E2E8F0; border-radius: 6px; }
            QHeaderView::section { background: #F1F5F9; font-weight: 700;
                                   font-size: 11px; padding: 6px; border: none; }
            QTableWidget::item { padding: 4px 8px; }
        """)
        root.addWidget(self.table)

        add_btn = QPushButton("+ Add Item")
        add_btn.setStyleSheet("""
            QPushButton { background: #3B82F6; color: white; border: none;
                          border-radius: 6px; padding: 7px 18px;
                          font-weight: 700; font-size: 12px; }
            QPushButton:hover { background: #2563EB; }
        """)
        add_btn.clicked.connect(self._add_row)
        root.addWidget(add_btn, alignment=Qt.AlignLeft)

        totals_frame = QFrame()
        totals_frame.setStyleSheet(
            "background:#F8FAFC; border:1px solid #E2E8F0;"
            "border-radius:8px; padding:6px;"
        )
        totals_layout = QHBoxLayout(totals_frame)
        totals_layout.setContentsMargins(12, 8, 12, 8)
        totals_layout.setSpacing(40)

        ta_box = QVBoxLayout()
        ta_title = QLabel("Total Amount")
        ta_title.setStyleSheet("font-size: 11px; font-weight: 700; color: #64748B;")
        self._total_amount_lbl = QLabel("0.00")
        self._total_amount_lbl.setStyleSheet("font-size: 18px; font-weight: 800; color: #0F766E;")
        ta_sub = QLabel("(pasted to field)")
        ta_sub.setStyleSheet("font-size: 12px; color: #94A3B8;")
        ta_box.addWidget(ta_title)
        ta_box.addWidget(self._total_amount_lbl)
        ta_box.addWidget(ta_sub)

        ct_box = QVBoxLayout()
        ct_title = QLabel("Computed Total")
        ct_title.setStyleSheet("font-size: 11px; font-weight: 700; color: #64748B;")
        self._computed_total_lbl = QLabel("0.00")
        self._computed_total_lbl.setStyleSheet("font-size: 18px; font-weight: 800; color: #2563EB;")
        ct_sub = QLabel("(added to Motor A.I)")
        ct_sub.setStyleSheet("font-size: 12px; color: #94A3B8;")
        ct_box.addWidget(ct_title)
        ct_box.addWidget(self._computed_total_lbl)
        ct_box.addWidget(ct_sub)

        totals_layout.addLayout(ta_box)
        totals_layout.addLayout(ct_box)
        totals_layout.addStretch()
        root.addWidget(totals_frame)

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.button(QDialogButtonBox.Ok).setText("Post")
        btns.button(QDialogButtonBox.Ok).setStyleSheet(
            "background:#16A34A;color:white;border:none;border-radius:5px;"
            "padding:6px 18px;font-weight:700;"
        )
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        root.addWidget(btns)

        self._add_row()

    def _add_row(self):
        row = self.table.rowCount()
        self.table.insertRow(row)

        pct_combo = QComboBox()
        pct_combo.addItems(self.PERCENTAGES)
        pct_combo.setStyleSheet("padding: 4px 6px; font-size: 12px;")
        self.table.setCellWidget(row, 0, pct_combo)

        amt_edit = QLineEdit()
        amt_edit.setPlaceholderText("0.00")
        amt_edit.setValidator(QDoubleValidator(0.0, 1e12, 2))
        amt_edit.setStyleSheet(
            "border: 1px solid #CBD5E1; border-radius: 5px;"
            "padding: 5px 8px; font-size: 12px;"
        )
        self.table.setCellWidget(row, 1, amt_edit)

        computed_lbl = QLabel("0.00")
        computed_lbl.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        computed_lbl.setStyleSheet("font-weight: 700; color: #2563EB; padding: 0 8px;")
        self.table.setCellWidget(row, 2, computed_lbl)

        rem_btn = QPushButton("\u2715")
        rem_btn.setFixedWidth(28)
        rem_btn.setStyleSheet(
            "QPushButton{background:#FEE2E2;color:#DC2626;border:none;"
            "border-radius:4px;font-weight:700;font-size:11px;}"
            "QPushButton:hover{background:#FECACA;}"
        )
        rem_btn.clicked.connect(lambda _, b=rem_btn: self._remove_row_by_widget(b))
        self.table.setCellWidget(row, 3, rem_btn)

        self._rows_data.append((pct_combo, amt_edit, computed_lbl))
        pct_combo.currentIndexChanged.connect(self._recalc)
        amt_edit.textChanged.connect(self._recalc)
        self.table.resizeRowsToContents()

    def _remove_row_by_widget(self, btn):
        for r in range(self.table.rowCount()):
            if self.table.cellWidget(r, 3) is btn:
                self.table.removeRow(r)
                self._rows_data = []
                for rr in range(self.table.rowCount()):
                    pc = self.table.cellWidget(rr, 0)
                    ae = self.table.cellWidget(rr, 1)
                    cl = self.table.cellWidget(rr, 2)
                    if pc and ae and cl:
                        self._rows_data.append((pc, ae, cl))
                self._recalc()
                return

    def _recalc(self, *_):
        total_amt = 0.0
        total_cmp = 0.0
        for pct_combo, amt_edit, computed_lbl in self._rows_data:
            try:
                pct = float(pct_combo.currentText().replace('%', '').strip()) / 100.0
                amt = float(amt_edit.text().strip().replace(',', '') or 0)
                cmp = amt * pct
            except ValueError:
                amt = cmp = 0.0
            computed_lbl.setText(f"{cmp:,.2f}")
            total_amt += amt
            total_cmp += cmp
        self._total_amount_lbl.setText(f"{total_amt:,.2f}")
        self._computed_total_lbl.setText(f"{total_cmp:,.2f}")

    def get_total_amount(self) -> float:
        try:
            return float(self._total_amount_lbl.text().replace(',', ''))
        except ValueError:
            return 0.0

    def get_computed_total(self) -> float:
        try:
            return float(self._computed_total_lbl.text().replace(',', ''))
        except ValueError:
            return 0.0

    def get_row_count(self) -> int:
        return len(self._rows_data)

    def get_breakdown_data(self) -> list:
        """Return list of [percentage_str, amount] for each row."""
        data = []
        for pct_combo, amt_edit, _ in self._rows_data:
            try:
                pct_str = pct_combo.currentText()
                amt = float(amt_edit.text().strip().replace(',', '') or 0)
            except ValueError:
                amt = 0.0
            data.append([pct_str, amt])
        return data

class EmpenaDetailDialog(QDialog):

    PERCENTAGES = ["2.5%", "3.0%", "4.0%", "5.0%", "20.0%"]

    def __init__(self, field_label, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"Detail – {field_label}")
        self.setMinimumSize(580, 460)
        self.setModal(True)
        self._rows_data = [] 

        root = QVBoxLayout(self)
        root.setSpacing(12)
        root.setContentsMargins(16, 16, 16, 16)

 
        title = QLabel(f"Breakdown for  {field_label}")
        title.setStyleSheet("font-size: 15px; font-weight: 800; color: #1E293B;")
        root.addWidget(title)

   
        self.table = QTableWidget(0, 4)
        self.table.setHorizontalHeaderLabels(["Percentage", "Amount", "Computed", ""])
        hh = self.table.horizontalHeader()
        hh.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        hh.setSectionResizeMode(1, QHeaderView.Stretch)
        hh.setSectionResizeMode(2, QHeaderView.Stretch)
        hh.setSectionResizeMode(3, QHeaderView.ResizeToContents)
        self.table.setSelectionMode(QAbstractItemView.NoSelection)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self.table.setStyleSheet("""
            QTableWidget { border: 1px solid #E2E8F0; border-radius: 6px; }
            QHeaderView::section { background: #F1F5F9; font-weight: 700;
                                   font-size: 11px; padding: 6px; border: none; }
            QTableWidget::item { padding: 4px 8px; }
        """)
        root.addWidget(self.table)


        add_btn = QPushButton("+ Add Item")
        add_btn.setStyleSheet("""
            QPushButton { background: #3B82F6; color: white; border: none;
                          border-radius: 6px; padding: 7px 18px;
                          font-weight: 700; font-size: 12px; }
            QPushButton:hover { background: #2563EB; }
        """)
        add_btn.clicked.connect(self._add_row)
        root.addWidget(add_btn, alignment=Qt.AlignLeft)

     
        totals_frame = QFrame()
        totals_frame.setStyleSheet(
            "background:#F8FAFC; border:1px solid #E2E8F0;"
            "border-radius:8px; padding:6px;"
        )
        totals_layout = QHBoxLayout(totals_frame)
        totals_layout.setContentsMargins(12, 8, 12, 8)
        totals_layout.setSpacing(40)

       
        ta_box = QVBoxLayout()
        ta_title = QLabel("Total Amount")
        ta_title.setStyleSheet("font-size: 11px; font-weight: 700; color: #64748B;")
        self._total_amount_lbl = QLabel("0.00")
        self._total_amount_lbl.setStyleSheet(
            "font-size: 18px; font-weight: 800; color: #0F766E;"
        )
        ta_sub = QLabel("(pasted to field)")
        ta_sub.setStyleSheet("font-size: 12px; color: #94A3B8;")
        ta_box.addWidget(ta_title)
        ta_box.addWidget(self._total_amount_lbl)
        ta_box.addWidget(ta_sub)


        ct_box = QVBoxLayout()
        ct_title = QLabel("Computed Total")
        ct_title.setStyleSheet("font-size: 11px; font-weight: 700; color: #64748B;")
        self._computed_total_lbl = QLabel("0.00")
        self._computed_total_lbl.setStyleSheet(
            "font-size: 18px; font-weight: 800; color: #2563EB;"
        )
        ct_sub = QLabel("(added to Jew. A.I)")
        ct_sub.setStyleSheet("font-size: 12px; color: #94A3B8;")
        ct_box.addWidget(ct_title)
        ct_box.addWidget(self._computed_total_lbl)
        ct_box.addWidget(ct_sub)

        totals_layout.addLayout(ta_box)
        totals_layout.addLayout(ct_box)
        totals_layout.addStretch()
        root.addWidget(totals_frame)

   
        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.button(QDialogButtonBox.Ok).setText("Post")
        btns.button(QDialogButtonBox.Ok).setStyleSheet(
            "background:#16A34A;color:white;border:none;border-radius:5px;"
            "padding:6px 18px;font-weight:700;"
        )
        btns.accepted.connect(self.accept)
        btns.rejected.connect(self.reject)
        root.addWidget(btns)

        self._add_row()  

   
    def _add_row(self):
        row = self.table.rowCount()
        self.table.insertRow(row)

        pct_combo = QComboBox()
        pct_combo.addItems(self.PERCENTAGES)
        pct_combo.setStyleSheet("padding: 4px 6px; font-size: 12px;")
        self.table.setCellWidget(row, 0, pct_combo)

        amt_edit = QLineEdit()
        amt_edit.setPlaceholderText("0.00")
        amt_edit.setValidator(QDoubleValidator(0.0, 1e12, 2))
        amt_edit.setStyleSheet(
            "border: 1px solid #CBD5E1; border-radius: 5px;"
            "padding: 5px 8px; font-size: 12px;"
        )
        self.table.setCellWidget(row, 1, amt_edit)

        computed_lbl = QLabel("0.00")
        computed_lbl.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        computed_lbl.setStyleSheet("font-weight: 700; color: #2563EB; padding: 0 8px;")
        self.table.setCellWidget(row, 2, computed_lbl)

        rem_btn = QPushButton("✕")
        rem_btn.setFixedWidth(28)
        rem_btn.setStyleSheet(
            "QPushButton{background:#FEE2E2;color:#DC2626;border:none;"
            "border-radius:4px;font-weight:700;font-size:11px;}"
            "QPushButton:hover{background:#FECACA;}"
        )
        rem_btn.clicked.connect(lambda _, b=rem_btn: self._remove_row_by_widget(b))
        self.table.setCellWidget(row, 3, rem_btn)

        self._rows_data.append((pct_combo, amt_edit, computed_lbl))
        pct_combo.currentIndexChanged.connect(self._recalc)
        amt_edit.textChanged.connect(self._recalc)
        self.table.resizeRowsToContents()

    def _remove_row_by_widget(self, btn):
        for r in range(self.table.rowCount()):
            if self.table.cellWidget(r, 3) is btn:
                self.table.removeRow(r)
                self._rows_data = []
                for rr in range(self.table.rowCount()):
                    pc = self.table.cellWidget(rr, 0)
                    ae = self.table.cellWidget(rr, 1)
                    cl = self.table.cellWidget(rr, 2)
                    if pc and ae and cl:
                        self._rows_data.append((pc, ae, cl))
                self._recalc()
                return

    def _recalc(self, *_):
        total_amt = 0.0
        total_cmp = 0.0
        for pct_combo, amt_edit, computed_lbl in self._rows_data:
            try:
                pct = float(pct_combo.currentText().replace('%', '').strip()) / 100.0
                amt = float(amt_edit.text().strip().replace(',', '') or 0)
                cmp = amt * pct
            except ValueError:
                amt = cmp = 0.0
            computed_lbl.setText(f"{cmp:,.2f}")
            total_amt += amt
            total_cmp += cmp
        self._total_amount_lbl.setText(f"{total_amt:,.2f}")
        self._computed_total_lbl.setText(f"{total_cmp:,.2f}")


    def get_total_amount(self) -> float:
        """Sum of raw amounts – pasted to the Empeno JEW field."""
        try:
            return float(self._total_amount_lbl.text().replace(',', ''))
        except ValueError:
            return 0.0

    def get_computed_total(self) -> float:
        """Sum of (amount × %) – contributed to Jew. A.I."""
        try:
            return float(self._computed_total_lbl.text().replace(',', ''))
        except ValueError:
            return 0.0

    def get_row_count(self) -> int:
        """Number of item rows – used as the lotes (quantity) value."""
        return len(self._rows_data)



class ClientDashboard(QWidget):
    logout_requested = pyqtSignal()
    brand_changed    = pyqtSignal(str)

    def __init__(self, username, branch, corporation, db_manager, offline_mode=False):
        super().__init__()
        self.user_email  = username
        self.corporation = corporation
        self.branch      = branch
        self.db_manager  = db_manager
        self.offline_mode = offline_mode
        self._update_checker_threads = []
        
        # Zoom functionality
        self.zoom_level = 100
        self.setFocusPolicy(Qt.StrongFocus)

        self.session = SessionManager(inactivity_timeout=1800)
        self._session_timer = QTimer(self)
        self._session_timer.timeout.connect(self._check_session_timeout)
        self._session_timer.start(60_000)

        # Network connectivity state — check every 5s via QThreadPool (non-blocking)
        # Defer first check by 3 seconds to not block app startup
        self._is_connected = True
        self._conn_timer = QTimer(self)
        self._conn_timer.timeout.connect(self._check_internet)
        QTimer.singleShot(3_000, lambda: self._conn_timer.start(5_000))

        self.beginning_balance_auto_filled_a = False
        self.beginning_balance_auto_filled_b = False
        self.previous_day_balance_a = None
        self.previous_day_date_a    = None
        self.previous_day_balance_b = None
        self.previous_day_date_b    = None

        self.setWindowTitle("Daily Cash Report")
        self.setMinimumSize(960, 600)
        self.setStyleSheet(_build_global_qss())

        lay = QVBoxLayout(self)
        lay.setSpacing(_sz(6))
        lay.setContentsMargins(_sz(12), _sz(10), _sz(12), _sz(8))

        # ── Connectivity banner (hidden until connection drops) ───────────
        self._conn_banner = ConnectionBanner()
        self._conn_banner.retry_btn.clicked.connect(self.on_date_changed)
        lay.addWidget(self._conn_banner)

        lay.addWidget(self._build_header(username, branch, corporation))
        lay.addWidget(self._build_toolbar())
        lay.addWidget(self._build_tabs(), stretch=1)

        lay.addWidget(self._build_summary_strip())
        lay.addWidget(self._build_footer())
        
        # Load global_tag for this branch
        self.global_tag = ''
        try:
            _gt = self.db_manager.execute_query(
                "SELECT COALESCE(global_tag, '') AS global_tag FROM branches WHERE name = %s LIMIT 1",
                (self.branch,)
            )
            if _gt:
                self.global_tag = str(_gt[0].get('global_tag') or '').strip()
        except Exception:
            pass

        self._connect_shared_fields()
        self._connect_palawan_adjustments_to_brand_b()
        self._connect_brand_a_auto_calculations()
        self._setup_empeno_jew_buttons()
        self._setup_empeno_motor_button()
        self._setup_ft_ho_button()
        self._setup_pc_salary_button()
        
        # Capture base fonts for zoom
        self._capture_base_fonts()

        # Install event filter on application for zoom
        QApplication.instance().installEventFilter(self)

        # Defer the initial date-load DB queries until after the window is painted
        QTimer.singleShot(0, self.on_date_changed)

    def _setup_pc_salary_button(self):
   
        try:
            from Client.salary_detail_dialog import SalaryDetailDialog
            for tab, label in [
                (self.cash_flow_tab_a, "PC-Salary"),
            ]:
                field = tab.credit_inputs.get(label)
                if field is None:
                    continue
                field.setReadOnly(True)
                field.setStyleSheet(
                    field.styleSheet() + "background-color: #EFF6FF; color: #1D4ED8;"
                )
                field.setToolTip("Click + to enter salary breakdown")
                field.setPlaceholderText("Click + to enter salary breakdown")

                container = field.parentWidget()
                layout = container.layout() if container else None
                if layout is None:
                    continue

                if not hasattr(self, '_salary_dialogs'):
                    self._salary_dialogs = {}
                key = f"{tab}_{label}"
                def _open_dialog():
                    if key not in self._salary_dialogs:
                        self._salary_dialogs[key] = SalaryDetailDialog(parent=self)
                    dlg = self._salary_dialogs[key]
                    if dlg.exec_() == QDialog.Accepted:
                        val = f"{dlg.get_total_salary():.2f}"
                        field.blockSignals(True)
                        field.setText(val)
                        field.blockSignals(False)
                       
                        self._pc_salary_breakdown = dlg.get_salary_breakdown()

                plus_btn = QPushButton("+")
                plus_btn.setFixedSize(28, 28)
                plus_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #2563EB; color: white;
                        border: none; border-radius: 5px;
                        font-size: 14px; font-weight: 900;
                    }
                    QPushButton:hover { background-color: #1D4ED8; }
                """)
                plus_btn.setToolTip("Open salary breakdown dialog")
                plus_btn.clicked.connect(_open_dialog)
                layout.insertWidget(0, plus_btn)
        except Exception as e:
            logger.error("_setup_pc_salary_button error: %s", e)

        from PyQt5.QtCore import QTimer as _QT
        _QT.singleShot(300, self._load_draft)
        
       
        self.loading_overlay = LoadingOverlay(self)
        
        self.showMaximized()

        if AUTO_UPDATE_ENABLED and check_update_success:
            check_update_success(parent=self)
        
        
        if not self.offline_mode and OFFLINE_SUPPORT and offline_manager:
            _QT.singleShot(1000, self._check_pending_entries_sync)

    def _check_pending_entries_sync(self):
        
        try:
            if not OFFLINE_SUPPORT or not offline_manager:
                return
            
            pending = offline_manager.get_pending_entries(self.user_email)
            if not pending:
                return
            
            count = len(pending)
            reply = QMessageBox.question(
                self,
                "Pending Entries Found",
                f"📤 You have {count} pending entries saved while offline.\n\n"
                f"Would you like to sync them now?\n\n"
                f"• Click 'Yes' to post all pending entries\n"
                f"• Click 'No' to sync later",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.Yes
            )
            
            if reply == QMessageBox.Yes:
                self._sync_pending_entries(pending)
                
        except Exception as e:
            logger.error("Error checking pending entries: %s", e)

    def _sync_pending_entries(self, pending_entries):
      
        try:
            self.loading_overlay.set_status("Syncing entries...", "Please wait")
            self.loading_overlay.show()
            self.loading_overlay.raise_()
            QApplication.processEvents()
            
            success_count = 0
            error_count = 0
            errors = []
            
            for i, entry in enumerate(pending_entries, 1):
                try:
                    self.loading_overlay.set_status(
                        f"Syncing entry {i}/{len(pending_entries)}...",
                        f"Date: {entry.get('entry_data', {}).get('date', 'Unknown')}"
                    )
                    QApplication.processEvents()
                    
                    success = self._post_pending_entry(entry)
                    
                    if success:
                        offline_manager.mark_entry_synced(entry['id'])
                        success_count += 1
                    else:
                        error_count += 1
                        errors.append(entry['id'][:15])
                        
                except Exception as e:
                    error_count += 1
                    errors.append(f"{entry['id'][:15]}: {str(e)[:30]}")
                    offline_manager.mark_entry_failed(entry['id'], str(e))
            
            self.loading_overlay.hide()
            
            if error_count == 0:
                self._msg_success(
                    f"Sync Complete!\n\n"
                    f"Successfully posted {success_count} entries."
                )
            else:
                self._msg(
                    "Sync Partially Complete",
                    f"Posted: {success_count}\n"
                    f"Failed: {error_count}\n\n"
                    f"Failed entries will remain in queue for retry.",
                    QMessageBox.Warning
                )
                
            # Refresh the view
            self.on_date_changed()
            
        except Exception as e:
            self.loading_overlay.hide()
            self._msg("Sync Error", f"Failed to sync: {e}", QMessageBox.Critical)

    def _post_pending_entry(self, pending_entry):

        try:
            entry_data = pending_entry.get('entry_data', {})
            selected_date = entry_data.get('date')
            username = entry_data.get('username')
            branch = entry_data.get('branch')
            corporation = entry_data.get('corporation')
            palawan_data = entry_data.get('palawan_data', {})
            brand_data = entry_data.get('brand_data', {})
            
            results = []
            
            for brand_full, data in brand_data.items():
                table_name = data.get('table_name')
                
            
                check_query = f"""
                    SELECT 1 FROM {table_name} 
                    WHERE date = %s AND username = %s 
                    LIMIT 1
                """
                existing = self.db_manager.execute_query(
                    check_query, [selected_date, username]
                )
                if existing:
                 
                    continue
                
                all_vals = data.get('all_values', {})
                filtered = self._filter_vals_for_table(all_vals, table_name)
                cols = [
                    'date', 'username', 'branch', 'corporation',
                    'beginning_balance', 'debit_total', 'credit_total',
                    'ending_balance', 'cash_count', 'cash_result', 'variance_status'
                ] + list(filtered.keys())
                
                vals = [
                    selected_date, username, branch, corporation,
                    data.get('beginning_balance', 0),
                    data.get('beginning_balance', 0) + data.get('debit_total', 0),
                    data.get('credit_total', 0),
                    data.get('ending_balance', 0),
                    data.get('cash_count', 0),
                    data.get('cash_result', 0),
                    data.get('variance_status', 'balanced')
                ] + list(filtered.values())
                
                ph = ', '.join(['%s'] * len(cols))
                query = f"INSERT INTO {table_name} ({', '.join(cols)}) VALUES ({ph})"
                upd = ', '.join(f"{c}=VALUES({c})" for c in cols
                               if c not in ('date', 'username'))
                if upd:
                    query += " ON DUPLICATE KEY UPDATE " + upd
                
                result = self.db_manager.execute_query(query, vals)
                if result is not None:
                    results.append(brand_full)
                    self._save_palawan_to_payable(selected_date, brand_full, palawan_data)
            
            return len(results) > 0
            
        except Exception as e:
            logger.error("Error posting pending entry: %s", e)
            return False


    def _build_header(self, username, branch, corporation):
        bar = QFrame()
        bar.setFixedHeight(_sz(54))
        bar.setStyleSheet(f"""
            QFrame {{
                background-color: {_SLATE_900};
                border-radius: {_sz(8)}px;
            }}
            QLabel {{ background: transparent; border: none; }}
        """)

        h = QHBoxLayout(bar)
        h.setContentsMargins(_sz(18), 0, _sz(14), 0)
        h.setSpacing(_sz(12))


        app_lbl = QLabel("Daily Cash Report")
        app_lbl.setStyleSheet(
            f"font-size: {_sz(15)}px; font-weight: 800; color: {_WHITE}; "
            f"letter-spacing: 0.3px; background: transparent;"
        )
        h.addWidget(app_lbl)


        dot = QLabel("·")
        dot.setStyleSheet(f"font-size: {_sz(18)}px; color: {_SLATE_500}; background: transparent;")
        h.addWidget(dot)


        info_lbl = QLabel(f"{username}  ·  {branch}  ·  {corporation}")
        info_lbl.setStyleSheet(
            f"font-size: {_sz(13)}px; font-weight: 500; color: {_SLATE_300}; background: transparent;"
        )
        h.addWidget(info_lbl)
        

        if self.offline_mode:
            offline_badge = QLabel("OFFLINE")
            offline_badge.setStyleSheet("""
                QLabel {
                    background-color: #F59E0B;
                    color: #1E293B;
                    font-size: 12px;
                    font-weight: bold;
                    padding: 5px 12px;
                    border-radius: 12px;
                }
            """)
            offline_badge.setToolTip("Entries will be saved locally for later posting")
            h.addWidget(offline_badge, alignment=Qt.AlignVCenter)
            

            if OFFLINE_SUPPORT and offline_manager:
                pending_count = offline_manager.get_pending_count(username)
                if pending_count > 0:
                    pending_badge = QLabel(f"{pending_count} pending")
                    pending_badge.setStyleSheet("""
                        QLabel {
                            background-color: #3B82F6;
                            color: white;
                            font-size: 12px;
                            font-weight: bold;
                            padding: 5px 10px;
                            border-radius: 12px;
                        }
                    """)
                    pending_badge.setToolTip(f"{pending_count} entries waiting to sync")
                    h.addWidget(pending_badge, alignment=Qt.AlignVCenter)
        
        h.addStretch()

        ver_lbl = QLabel(f"v{__version__}")
        ver_lbl.setStyleSheet(f"""
            QLabel {{
                color: {_TEXT_MUTED};
                font-size: {_sz(11)}px;
                font-weight: 600;
                padding: 0 6px;
                background: transparent;
            }}
        """)
        ver_lbl.setToolTip(f"Operation Report System v{__version__}")
        h.addWidget(ver_lbl)

        if AUTO_UPDATE_ENABLED:
            upd_btn = QPushButton("🔄 Update")
            upd_btn.setFixedSize(_sz(90), _sz(30))
            upd_btn.setStyleSheet(f"""
                QPushButton {{
                    background: rgba(99,102,241,0.15);
                    color: {_INDIGO_400};
                    border: 1px solid rgba(99,102,241,0.3);
                    border-radius: {_sz(5)}px;
                    font-size: {_sz(12)}px; font-weight: 700;
                }}
                QPushButton:hover {{ background: rgba(99,102,241,0.25); }}
            """)
            upd_btn.clicked.connect(self.check_for_updates)
            upd_btn.setToolTip("Check for updates")
            h.addWidget(upd_btn, alignment=Qt.AlignVCenter)
            h.addSpacing(_sz(8))

        logout_btn = QPushButton("Sign out")
        logout_btn.setFixedSize(_sz(88), _sz(30))
        logout_btn.setStyleSheet(f"""
            QPushButton {{
                background: transparent;
                color: {_SLATE_300};
                border: 1px solid {_SLATE_600};
                border-radius: {_sz(5)}px;
                font-size: {_sz(12)}px; font-weight: 600;
            }}
            QPushButton:hover {{
                background: rgba(248,113,113,0.1);
                color: {_RED_400};
                border-color: rgba(248,113,113,0.3);
            }}
        """)
        logout_btn.clicked.connect(self.handle_logout)
        h.addWidget(logout_btn, alignment=Qt.AlignVCenter)

        return bar

    def _build_toolbar(self):
        bar = QFrame()
        bar.setObjectName("toolbar")
        bar.setStyleSheet(f"""
            QFrame#toolbar {{
                background-color: {_BG_CARD};
                border: 1px solid {_BORDER};
                border-radius: {_sz(10)}px;
            }}
        """)

        outer = QHBoxLayout(bar)
        outer.setContentsMargins(_sz(20), _sz(14), _sz(20), _sz(14))
        outer.setSpacing(0)

        date_card = QFrame()
        date_card.setObjectName("dateCard")
        date_card.setFixedWidth(_sz(210))
        date_card.setStyleSheet(f"""
            QFrame#dateCard {{
                background-color: {_SLATE_50};
                border: 1px solid {_BORDER};
                border-radius: {_sz(8)}px;
                padding: 0px;
            }}
        """)
        date_card_layout = QVBoxLayout(date_card)
        date_card_layout.setContentsMargins(_sz(14), _sz(12), _sz(14), _sz(12))
        date_card_layout.setSpacing(_sz(8))

        date_title = QLabel("Report Date")
        date_title.setStyleSheet(
            f"font-size: {_sz(8)}px; font-weight: 600; color: {_TEXT_SEC}; background: transparent;"
            f" text-transform: uppercase; letter-spacing: 1px;"
        )
        date_card_layout.addWidget(date_title)

        # Use NoWheelDateEdit to prevent accidental scroll changes
        self.date_picker = NoWheelDateEdit()
        self.date_picker.setCalendarPopup(True)
        self.date_picker.setDisplayFormat("dd MMM yyyy")
        self.date_picker.setDate(QDate.currentDate())
        self.date_picker.setFixedHeight(_sz(34))
        self.date_picker.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.date_picker.setMinimumWidth(_sz(140))
        self.date_picker.dateChanged.connect(self.on_date_changed)
        _cal = os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            'assets', 'calendar.png'
        ).replace('\\', '/')
        self.date_picker.setStyleSheet(f"""
            QDateEdit {{
                background-color: {_WHITE};
                border: 1px solid {_BORDER};
                border-radius: 6px;
                padding: 4px 10px;
                font-size: 13px;
                font-weight: 600;
                color: {_TEXT_PRI};
            }}
            QDateEdit::drop-down {{
                subcontrol-origin: border;
                subcontrol-position: center right;
                width: 28px;
                border-left: 1px solid {_BORDER};
                background-color: #f0f2f5;
                border-top-right-radius: 6px;
                border-bottom-right-radius: 6px;
            }}
            QDateEdit::drop-down:hover {{
                background-color: #dde1e7;
            }}
            QDateEdit::down-arrow {{
                image: url({_cal});
                width: 14px;
                height: 14px;
            }}
            QCalendarWidget {{
                min-width: 340px;
                min-height: 280px;
                background: white;
                border: 1px solid {_BORDER};
                border-radius: 6px;
            }}
            QCalendarWidget QWidget#qt_calendar_navigationbar {{
                background-color: #343a40;
                min-height: 42px;
                padding: 4px 6px;
                border-radius: 4px 4px 0 0;
            }}
            QCalendarWidget QToolButton {{
                color: #ecf0f1;
                font-size: 14px;
                font-weight: bold;
                background-color: transparent;
                padding: 6px 10px;
                border-radius: 4px;
                margin: 2px;
            }}
            QCalendarWidget QToolButton:hover {{
                background-color: #007bff;
                color: white;
            }}
            QCalendarWidget QToolButton:pressed {{
                background-color: #0056b3;
                color: white;
            }}
            QCalendarWidget QSpinBox {{
                color: #2c3e50;
                background-color: #ecf0f1;
                font-size: 13px;
                font-weight: bold;
                border: 1px solid #bdc3c7;
                border-radius: 4px;
                padding: 4px 8px;
                selection-background-color: #007bff;
                selection-color: white;
            }}
            QCalendarWidget QAbstractItemView {{
                background: white;
                selection-background-color: #007bff;
                selection-color: white;
                font-size: 12px;
                alternate-background-color: #f8f9fa;
            }}
            QCalendarWidget QAbstractItemView::item {{
                padding: 6px;
                border-radius: 4px;
            }}
            QCalendarWidget QAbstractItemView::item:alternate {{
                background-color: #f8f9fa;
            }}
            QCalendarWidget QAbstractItemView::item:selected {{
                background-color: #007bff;
                color: white;
                font-weight: bold;
            }}
        """)

        # Add the date picker directly; auto-loads entry when date changes
        self.date_picker.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        date_card_layout.addWidget(self.date_picker)

        outer.addWidget(date_card)
        outer.addSpacing(_sz(20))

  
        divider = QFrame()
        divider.setFrameShape(QFrame.VLine)
        divider.setStyleSheet(f"color: {_BORDER}; background: {_BORDER}; max-width: 1px;")
        divider.setFixedHeight(_sz(80))
        outer.addWidget(divider)
        outer.addSpacing(_sz(20))

       
        self.bb_container_a = QWidget()
        self.bb_container_a.setStyleSheet("background: transparent;")
        bb_layout_a = QHBoxLayout(self.bb_container_a)
        bb_layout_a.setContentsMargins(0, 0, 0, 0)
        bb_layout_a.setSpacing(0)
        bb_layout_a.addLayout(self._build_balance_column("A"))
        outer.addWidget(self.bb_container_a, stretch=1)

        outer.addSpacing(_sz(20))

       
        self.bb_divider_b = QFrame()
        self.bb_divider_b.setFrameShape(QFrame.VLine)
        self.bb_divider_b.setStyleSheet(f"color: {_BORDER}; background: {_BORDER}; max-width: 1px;")
        self.bb_divider_b.setFixedHeight(_sz(80))
        outer.addWidget(self.bb_divider_b)
        outer.addSpacing(_sz(20))

    
        self.bb_container_b = QWidget()
        self.bb_container_b.setStyleSheet("background: transparent;")
        bb_layout_b = QHBoxLayout(self.bb_container_b)
        bb_layout_b.setContentsMargins(0, 0, 0, 0)
        bb_layout_b.setSpacing(0)
        bb_layout_b.addLayout(self._build_balance_column("B"))
        outer.addWidget(self.bb_container_b, stretch=1)

  
        self.bb_container_b.hide()
        self.bb_divider_b.hide()

 
        self.beginning_balance_input = self.beginning_balance_input_a
        self.balance_status_label    = self.balance_status_label_a
        self.auto_fill_button        = self.auto_fill_button_a

        return bar

    def _build_balance_column(self, brand: str) -> QVBoxLayout:

        is_a  = brand == "A"
        label = "Brand A" if is_a else "Brand B"

        col = QVBoxLayout()
        col.setSpacing(3)

        # Title row with refresh button on the right
        title_row = QHBoxLayout()
        title_row.setContentsMargins(0, 0, 0, 0)
        title_row.setSpacing(8)

        title = QLabel(f"{label}  —  Opening Balance")
        title.setStyleSheet(
            f"font-size: 10px; font-weight: 600; color: {_TEXT_SEC}; background: transparent;"
            f" text-transform: uppercase; letter-spacing: 1px;"
        )
        title_row.addWidget(title)
        title_row.addStretch()

        # Add refresh button on top right (only for Brand A)
        if brand == "A" and not hasattr(self, 'refresh_button'):
            self.refresh_button = QPushButton("↻")
            self.refresh_button.setFixedSize(_sz(32), _sz(28))
            self.refresh_button.setStyleSheet(f"""
                QPushButton {{
                    background-color: #6366F1;
                    color: {_WHITE};
                    font-size: {_sz(14)}px;
                    font-weight: 700;
                    border: none;
                    border-radius: {_sz(6)}px;
                }}
                QPushButton:hover {{ background-color: #4F46E5; }}
                QPushButton:pressed {{ background-color: #4338CA; }}
            """)
            self.refresh_button.clicked.connect(self.on_date_changed)
            self.refresh_button.setToolTip("Reload the report for the current date.")
            title_row.addWidget(self.refresh_button)

        col.addLayout(title_row)

        bb_input = QLineEdit()
        bb_input.setValidator(QDoubleValidator(0.0, 1e12, 2))
        bb_input.setPlaceholderText("0.00")
        bb_input.setReadOnly(True)
        bb_input.setFixedHeight(_sz(38))
        bb_input.setStyleSheet(f"""
            QLineEdit {{
                background-color: {_SLATE_50};
                border: 1.5px solid {_BORDER};
                border-radius: {_sz(8)}px;
                padding: {_sz(6)}px {_sz(14)}px;
                font-size: {_sz(15)}px;
                font-weight: 700;
                color: {_TEXT_PRI};
            }}
        """)
        bb_input.textChanged.connect(self.recalculate_all)
        col.addWidget(bb_input)

    
        action_row = QHBoxLayout()
        action_row.setSpacing(_sz(8))
        action_row.setContentsMargins(0, 2, 0, 0)

        load_btn = QPushButton("↓ Load Previous Day")
        load_btn.setFixedHeight(_sz(26))
        load_btn.setCursor(Qt.PointingHandCursor)
        load_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: transparent;
                color: {_PRIMARY};
                border: 1px solid {_BORDER};
                border-radius: {_sz(5)}px;
                font-size: {_sz(11)}px;
                font-weight: 600;
                padding: {_sz(3)}px {_sz(10)}px;
                min-width: 0;
            }}
            QPushButton:hover {{
                background-color: {_INDIGO_50};
                border-color: {_PRIMARY};
            }}
            QPushButton:pressed {{
                background-color: {_INDIGO_100};
            }}
            QPushButton:disabled {{
                color: {_TEXT_MUTED};
                border-color: {_SLATE_200};
                background-color: transparent;
            }}
        """)
        load_btn.clicked.connect(lambda checked, b=brand: self.auto_fill_beginning_balance(b))
        action_row.addWidget(load_btn)

        status_lbl = QLabel("")
        status_lbl.setStyleSheet(
            f"font-size: 10px; color: {_TEXT_MUTED}; background: transparent;"
        )
        status_lbl.setWordWrap(True)
        action_row.addWidget(status_lbl, stretch=1)

        col.addLayout(action_row)

        # Store references
        if is_a:
            self.beginning_balance_input_a = bb_input
            self.auto_fill_button_a        = load_btn
            self.balance_status_label_a    = status_lbl
        else:
            self.beginning_balance_input_b = bb_input
            self.auto_fill_button_b        = load_btn
            self.balance_status_label_b    = status_lbl

        return col

    def _build_tabs(self):
        tw = QTabWidget()
        tw.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        self.cash_flow_tab_a = CashFlowTab(self, "Brand A")
        self.cash_flow_tab_b = CashFlowTab(self, "Brand B")
        self.palawan_tab     = PalawanDetailsTab(self)

        self.cash_flow_tab = self.cash_flow_tab_a

        tw.addTab(self.cash_flow_tab_a, "Brand A")
        tw.addTab(self.cash_flow_tab_b, "Brand B")
        tw.addTab(self.palawan_tab,     "Palawan Details B")


        self.tab_widget = tw
        tw.currentChanged.connect(self._on_tab_changed)

        return tw

    def _on_tab_changed(self, index):

        if not hasattr(self, 'summary_container_a') or not hasattr(self, 'summary_container_b'):
            return

       
        if hasattr(self, 'bb_container_a') and hasattr(self, 'bb_container_b'):
            if index == 0:   # Brand A
                self.bb_container_a.show()
                self.bb_container_b.hide()
                self.bb_divider_b.hide()
            elif index == 1: # Brand B
                self.bb_container_a.hide()
                self.bb_container_b.show()
                self.bb_divider_b.show()
            else:            # Palawan Details B — show both
                self.bb_container_a.show()
                self.bb_container_b.show()
                self.bb_divider_b.show()

        if index == 0:   # Brand A
            self.summary_container_a.show()
            self.summary_container_b.hide()
        elif index == 1: # Brand B
            self.summary_container_a.hide()
            self.summary_container_b.show()
        else:            # Palawan Details B — hide all summaries
            self.summary_container_a.hide()
            self.summary_container_b.hide()

    def _build_summary_strip(self):
        strip = QFrame()
        strip.setFixedHeight(_sz(110))
        strip.setStyleSheet(f"""
            QFrame {{
                background-color: {_SLATE_900};
                border-radius: {_sz(8)}px;
            }}
        """)

        h = QHBoxLayout(strip)
        h.setContentsMargins(_sz(20), _sz(10), _sz(20), _sz(10))
        h.setSpacing(0)

        col_a, self.ending_balance_display_a, self.cash_count_input_a, \
            self.cash_result_display_a, self.variance_status_label_a, \
            self.cash_float_input_a = \
            self._build_brand_summary("A")

        col_b, self.ending_balance_display_b, self.cash_count_input_b, \
            self.cash_result_display_b, self.variance_status_label_b, \
            _ = \
            self._build_brand_summary("B")

 
        self.ending_balance_display  = self.ending_balance_display_a
        self.cash_count_input        = self.cash_count_input_a
        self.cash_result_display     = self.cash_result_display_a
        self.variance_status_label   = self.variance_status_label_a

        self.cash_count_input_a.textChanged.connect(self.recalculate_all)
        self.cash_count_input_a.textChanged.connect(self.update_cash_result)
        self.cash_count_input_b.textChanged.connect(self.recalculate_all)
        self.cash_count_input_b.textChanged.connect(self.update_cash_result)


        self.summary_container_a = QWidget()
        self.summary_container_a.setStyleSheet("background: transparent;")
        container_layout_a = QHBoxLayout(self.summary_container_a)
        container_layout_a.setContentsMargins(0, 0, 0, 0)
        container_layout_a.setSpacing(0)
        container_layout_a.addLayout(col_a, stretch=1)

        self.summary_container_b = QWidget()
        self.summary_container_b.setStyleSheet("background: transparent;")
        container_layout_b = QHBoxLayout(self.summary_container_b)
        container_layout_b.setContentsMargins(0, 0, 0, 0)
        container_layout_b.setSpacing(0)
        container_layout_b.addLayout(col_b, stretch=1)

        h.addWidget(self.summary_container_a, stretch=1)
        h.addSpacing(16)
        h.addWidget(self.summary_container_b, stretch=1)

   
        self.summary_container_b.hide()

        return strip

    def _build_brand_summary(self, brand: str):
        is_a   = brand == "A"
        accent = _INDIGO_400 if is_a else _EMERALD_400

        def _cap(t):
            l = QLabel(t)
            l.setAlignment(Qt.AlignCenter)
            l.setStyleSheet(
                f"font-size: {_sz(11)}px; font-weight: 700; color: {_SLATE_300}; "
                f"letter-spacing: 1px; background: transparent; text-transform: uppercase;"
            )
            return l

        def _val(init="0.00"):
            l = QLabel(init)
            l.setAlignment(Qt.AlignCenter)
            l.setStyleSheet(
                f"font-size: {_sz(18)}px; font-weight: 800; color: {_WHITE}; background: transparent;"
            )
            return l

        def _cc_inp():
            inp = QLineEdit()
            inp.setValidator(QDoubleValidator(0.0, 1e12, 2))
            inp.setPlaceholderText("0.00")
            inp.setAlignment(Qt.AlignCenter)
            inp.setFixedHeight(_sz(32))
            inp.setStyleSheet(f"""
                QLineEdit {{
                    background: rgba(255,255,255,0.08);
                    border: 1.5px solid {_SLATE_600};
                    border-radius: {_sz(6)}px;
                    color: {_WHITE};
                    font-size: {_sz(14)}px;
                    font-weight: 700;
                    padding: {_sz(3)}px {_sz(10)}px;
                }}
                QLineEdit:focus {{
                    border: 2px solid {accent};
                    background: rgba(255,255,255,0.12);
                }}
            """)
            return inp

        col = QVBoxLayout()
        col.setSpacing(_sz(4))
        col.setContentsMargins(_sz(8), 0, _sz(8), 0)

        pill_row = QHBoxLayout()
        pill_row.setAlignment(Qt.AlignCenter)
        pill = QLabel(f"BRAND {'A' if is_a else 'B'}")
        pill.setAlignment(Qt.AlignCenter)
        pill.setStyleSheet(
            f"font-size: {_sz(12)}px; font-weight: 800; color: {accent}; letter-spacing: 1.5px; "
            f"background: transparent; padding: 0px 0;"
        )
        pill_row.addWidget(pill)
        col.addLayout(pill_row)

  
        row = QHBoxLayout()
        row.setSpacing(_sz(8))

        def _metric(cap_w, val_w):
            mc = QVBoxLayout()
            mc.setSpacing(_sz(2))
            mc.addWidget(cap_w)
            mc.addWidget(val_w, alignment=Qt.AlignCenter)
            return mc

        eb_val  = _val()
        cc_inp  = _cc_inp()
        var_val = _val()

        status_lbl = QLabel("—")
        status_lbl.setAlignment(Qt.AlignCenter)
        status_lbl.setStyleSheet(
            f"font-size: {_sz(13)}px; font-weight: 700; color: {_SLATE_300}; "
            f"min-width: {_sz(80)}px; background: transparent;"
        )

        row.addLayout(_metric(_cap("Ending Bal"),  eb_val),  stretch=2)
        row.addWidget(_vline(dark=True))
        row.addLayout(_metric(_cap("Cash Count"),  cc_inp),  stretch=2)
        row.addWidget(_vline(dark=True))


        cf_inp = None
        if is_a:
            cf_inp = QLineEdit()
            cf_inp.setValidator(QDoubleValidator(0.0, 1e12, 2))
            cf_inp.setPlaceholderText("0.00")
            cf_inp.setAlignment(Qt.AlignCenter)
            cf_inp.setFixedHeight(_sz(32))
            cf_inp.setStyleSheet(f"""
                QLineEdit {{
                    background: rgba(255,255,255,0.08);
                    border: 1.5px solid {_SLATE_600};
                    border-radius: {_sz(6)}px;
                    color: {_WHITE};
                    font-size: {_sz(14)}px;
                    font-weight: 700;
                    padding: {_sz(3)}px {_sz(10)}px;
                }}
                QLineEdit:focus {{
                    border: 2px solid {_AMBER_400};
                    background: rgba(255,255,255,0.12);
                }}
            """)
            row.addLayout(_metric(_cap("Cash Float"), cf_inp), stretch=2)
            row.addWidget(_vline(dark=True))

        row.addLayout(_metric(_cap("Variance"),    var_val), stretch=2)
        row.addWidget(_vline(dark=True))
        row.addLayout(_metric(_cap("Status"),      status_lbl), stretch=3)

        col.addLayout(row)
        return col, eb_val, cc_inp, var_val, status_lbl, cf_inp

    def _build_footer(self):
        bar = QFrame()
        bar.setFixedHeight(_sz(48))
        row = QHBoxLayout(bar)
        row.setContentsMargins(0, _sz(6), 0, _sz(2))
        row.setSpacing(_sz(10))
        row.addStretch()

        self.draft_button = QPushButton("Save Draft")
        self.draft_button.setFixedSize(_sz(100), _sz(34))
        self.draft_button.setStyleSheet(f"""
            QPushButton {{
                background-color: transparent;
                color: {_SLATE_600};
                font-size: {_sz(12)}px;
                font-weight: 700;
                border: 1.5px solid {_SLATE_300};
                border-radius: {_sz(6)}px;
            }}
            QPushButton:hover {{
                background-color: {_SLATE_100};
                border-color: {_SLATE_400};
                color: {_TEXT_PRI};
            }}
            QPushButton:pressed {{ background-color: {_SLATE_200}; }}
        """)
        self.draft_button.clicked.connect(self.save_draft)
        self.draft_button.setToolTip(
            "Save all current fields as a draft.\n"
            "Your data will be restored next time you log in."
        )

        self.post_button = QPushButton("Post Both Brands")
        self.post_button.setFixedSize(_sz(158), _sz(34))
        self.post_button.setStyleSheet(f"""
            QPushButton {{
                background-color: {_SUCCESS};
                color: {_WHITE};
                font-size: {_sz(13)}px;
                font-weight: 800;
                border-radius: {_sz(6)}px;
                letter-spacing: 0.2px;
            }}
            QPushButton:hover   {{ background-color: {_SUCCESS_DK}; }}
            QPushButton:pressed {{ background-color: #047857; }}
            QPushButton:disabled {{
                background-color: {_SLATE_200};
                color: {_TEXT_MUTED};
            }}
        """)
        self.post_button.clicked.connect(self.handle_post)
        self.post_button.setEnabled(False)


        self.print_button = QPushButton("Print")
        self.print_button.setFixedSize(_sz(100), _sz(34))
        self.print_button.setStyleSheet(f"""
            QPushButton {{
                background-color: #3B82F6;
                color: {_WHITE};
                font-size: {_sz(12)}px;
                font-weight: 700;
                border: none;
                border-radius: {_sz(6)}px;
            }}
            QPushButton:hover {{ background-color: #2563EB; }}
            QPushButton:pressed {{ background-color: #1D4ED8; }}
        """)
        self.print_button.clicked.connect(self.print_nonzero_report)
        self.print_button.setToolTip(
            "Print only fields with non-zero values.\n"
            "Filters out all zero-value entries from the report."
        )

    
        self.export_excel_button = QPushButton("Export Excel")
        self.export_excel_button.setFixedSize(_sz(120), _sz(34))
        self.export_excel_button.setStyleSheet(f"""
            QPushButton {{
                background-color: #059669;
                color: {_WHITE};
                font-size: {_sz(12)}px;
                font-weight: 700;
                border: none;
                border-radius: {_sz(6)}px;
            }}
            QPushButton:hover {{ background-color: #047857; }}
            QPushButton:pressed {{ background-color: #065F46; }}
        """)
        self.export_excel_button.clicked.connect(self.export_to_excel)
        self.export_excel_button.setToolTip(
            "Export non-zero fields to Excel (.xlsx).\n"
            "Only includes fields with values greater or less than zero."
        )

  
        self.clear_table_button = QPushButton(" Clear Table")
        self.clear_table_button.setFixedSize(_sz(120), _sz(34))
        self.clear_table_button.setStyleSheet(f"""
            QPushButton {{
                background-color: transparent;
                color: {_RED_500};
                font-size: {_sz(12)}px;
                font-weight: 700;
                border: 1.5px solid {_RED_400};
                border-radius: {_sz(6)}px;
            }}
            QPushButton:hover {{
                background-color: #FEF2F2;
                border-color: {_RED_500};
                color: #B91C1C;
            }}
            QPushButton:pressed {{ background-color: #FEE2E2; }}
        """)
        self.clear_table_button.clicked.connect(self.clear_table)
        self.clear_table_button.setToolTip(
            "Clear all fields in both Brand A and Brand B.\n"
            "This will not delete any submitted data."
        )

        self.load_draft_button = QPushButton("Load Draft")
        self.load_draft_button.setFixedSize(_sz(100), _sz(34))
        self.load_draft_button.setStyleSheet(f"""
            QPushButton {{
                background-color: transparent;
                color: {_SLATE_600};
                font-size: {_sz(12)}px;
                font-weight: 700;
                border: 1.5px solid {_SLATE_300};
                border-radius: {_sz(6)}px;
            }}
            QPushButton:hover {{
                background-color: {_SLATE_100};
                border-color: {_SLATE_400};
                color: {_TEXT_PRI};
            }}
            QPushButton:pressed {{ background-color: {_SLATE_200}; }}
        """)
        self.load_draft_button.clicked.connect(self._load_draft)
        self.load_draft_button.setToolTip(
            "Load a previously saved draft.\n"
            "Select from available drafts to restore."
        )

        row.addWidget(self.print_button)
        row.addWidget(self.export_excel_button)
        row.addWidget(self.clear_table_button)
        row.addWidget(self.load_draft_button)
        row.addWidget(self.draft_button)
        row.addWidget(self.post_button)
        return bar

  
    def _money_input(self, placeholder=""):
        f = QLineEdit()
        f.setValidator(QDoubleValidator(0.0, 1e12, 2))
        f.setPlaceholderText(placeholder)
        f.textChanged.connect(self.recalculate_all)
        return f

    def create_money_input(self, placeholder=""):
        return self._money_input(placeholder)

    def create_display_field(self, placeholder=""):
        f = QLineEdit()
        f.setReadOnly(True)
        f.setPlaceholderText(placeholder)
        return f

    def create_separator(self):
        return _hline()

    def get_total_amount(self, inputs_dict):
        total = 0.0
        for field in inputs_dict.values():
            try:
               
                text = field.text().strip().replace(",", "")
                total += float(text or 0)
            except ValueError:
                pass
        return total


    def handle_logout(self):
        r = QMessageBox.question(
            self, "Confirm Sign Out", "Sign out of the current session?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        if r == QMessageBox.Yes:
            self.session.logout()
            self.logout_requested.emit()
            self.close()

    def _check_session_timeout(self):
        # Don't time out while a loading overlay is visible (network op in progress)
        if hasattr(self, 'loading_overlay') and self.loading_overlay.isVisible():
            self.session.update_activity()
            return
        if self.session.check_timeout():
            self._session_timer.stop()
            QMessageBox.warning(
                self, "Session Expired",
                "Your session has expired due to inactivity.\nPlease log in again.",
                QMessageBox.Ok
            )
            self.logout_requested.emit()
            self.close()

    def mousePressEvent(self, event):
        self.session.update_activity()
        super().mousePressEvent(event)

    def eventFilter(self, obj, event):
        """Handle application-level events for zoom and session activity tracking."""
        from PyQt5.QtCore import QEvent
        # Reset session timer on any real user interaction
        _activity_types = (
            QEvent.MouseButtonPress, QEvent.MouseButtonRelease,
            QEvent.MouseMove, QEvent.KeyPress, QEvent.Wheel,
            QEvent.FocusIn, QEvent.TouchBegin,
        )
        if event.type() in _activity_types:
            self.session.update_activity()

        # Capture wheel events at app level for zoom
        if event.type() == QEvent.Wheel:
            if event.modifiers() & Qt.ControlModifier:
                delta = event.angleDelta().y()
                if delta > 0:
                    self.zoom_in()
                else:
                    self.zoom_out()
                return True
        # Check for Ctrl+0 key press
        elif event.type() == QEvent.KeyPress:
            if event.modifiers() & Qt.ControlModifier and event.key() == Qt.Key_0:
                self.reset_zoom()
                return True
        return super().eventFilter(obj, event)

    def wheelEvent(self, event):
        """Handle Ctrl + mouse wheel for zoom"""
        if event.modifiers() & Qt.ControlModifier:
            # Zoom in (scroll up) or zoom out (scroll down)
            delta = event.angleDelta().y()
            if delta > 0:
                self.zoom_in()
            else:
                self.zoom_out()
            event.accept()
        else:
            super().wheelEvent(event)

    def keyPressEvent(self, event):
        # Handle Ctrl+0 to reset zoom
        if event.modifiers() & Qt.ControlModifier and event.key() == Qt.Key_0:
            self.reset_zoom()
            event.accept()
        else:
            self.session.update_activity()
            super().keyPressEvent(event)

    def zoom_in(self):
        """Increase zoom level by 20%"""
        self.set_zoom_level(self.zoom_level + 20)
    
    def zoom_out(self):
        """Decrease zoom level by 20%"""
        self.set_zoom_level(self.zoom_level - 20)
    
    def reset_zoom(self):
        """Reset zoom to 100%"""
        self.set_zoom_level(100)
    
    def set_zoom_level(self, level):
        """Set zoom level and apply to all widgets"""
        # Clamp zoom level between 50% and 200%
        level = max(50, min(200, level))
        if level == self.zoom_level:
            return  # No change
        self.zoom_level = level
        self._apply_zoom_to_all()
    
    def _capture_base_fonts(self):
        """Capture base font sizes for zoom-target widgets only."""
        for w in self._get_zoom_target_widgets():
            font = w.font()
            point_size = font.pointSize()
            pixel_size = font.pixelSize()

            # Preserve original stylesheet so repeated zoom uses a stable base.
            base_stylesheet = w.styleSheet() or ""
            w.setProperty('_base_stylesheet', base_stylesheet)
            w.setProperty('_base_zoom_height', max(w.minimumHeight(), w.sizeHint().height()))
            
            # Handle case where font size is inherited (pointSize == -1)
            if point_size < 0 and pixel_size < 0:
                # Get from application default
                app_font = QApplication.font()
                point_size = app_font.pointSize()
                if point_size < 0:
                    point_size = 10  # Default fallback
            
            # Store whichever is valid
            if point_size > 0:
                w.setProperty('_base_point_size', point_size)
            elif pixel_size > 0:
                w.setProperty('_base_pixel_size', pixel_size)

    def _get_zoom_target_widgets(self):
        """Return only debit/credit amount inputs from both cash-flow tabs."""
        targets = []
        seen = set()
        for tab in (getattr(self, 'cash_flow_tab_a', None), getattr(self, 'cash_flow_tab_b', None)):
            if tab is None:
                continue
            for inp in list(getattr(tab, 'debit_inputs', {}).values()) + list(getattr(tab, 'credit_inputs', {}).values()):
                if inp is None:
                    continue
                wid = id(inp)
                if wid in seen:
                    continue
                seen.add(wid)
                targets.append(inp)
        return targets

    def _scale_stylesheet_font_sizes(self, stylesheet, zoom_factor):
        """Scale font-size declarations in a stylesheet by zoom factor."""
        if not stylesheet or "font-size" not in stylesheet:
            return stylesheet

        def _replace(match):
            base_size = float(match.group(1))
            scaled = max(1, min(500, int(round(base_size * zoom_factor))))
            return f"font-size: {scaled}px"

        return re.sub(
            r"font-size\s*:\s*([0-9]*\.?[0-9]+)\s*px",
            _replace,
            stylesheet,
            flags=re.IGNORECASE,
        )

    def _apply_zoom_to_all(self):
        """Apply zoom only to debit/credit amount input widgets."""
        zoom_factor = self.zoom_level / 100.0
        for w in self._get_zoom_target_widgets():
            font = w.font()
            new_size = None
            
            # Try point size first
            base_point = w.property('_base_point_size')
            if base_point is not None:
                try:
                    base_point = int(base_point)
                    if base_point > 0:
                        new_size = max(1, min(500, int(base_point * zoom_factor)))
                        font.setPointSize(new_size)
                        w.setFont(font)
                except (ValueError, TypeError, OverflowError):
                    pass
            
            # Try pixel size
            base_pixel = w.property('_base_pixel_size')
            if base_pixel is not None:
                try:
                    base_pixel = int(base_pixel)
                    if base_pixel > 0:
                        new_size = max(1, min(500, int(base_pixel * zoom_factor)))
                        font.setPixelSize(new_size)
                        w.setFont(font)
                except (ValueError, TypeError, OverflowError):
                    pass

            base_stylesheet = w.property('_base_stylesheet')
            if isinstance(base_stylesheet, str):
                scaled_stylesheet = self._scale_stylesheet_font_sizes(base_stylesheet, zoom_factor)
                if new_size is not None and "font-size" not in scaled_stylesheet.lower():
                    scaled_stylesheet = (scaled_stylesheet + f"\nfont-size: {new_size}px;").strip()
                if scaled_stylesheet != w.styleSheet():
                    w.setStyleSheet(scaled_stylesheet)

            base_height = w.property('_base_zoom_height')
            if base_height is not None:
                try:
                    scaled_height = max(20, min(500, int(int(base_height) * zoom_factor)))
                    w.setMinimumHeight(scaled_height)
                    w.updateGeometry()
                except (ValueError, TypeError, OverflowError):
                    pass

        self.update()

    def check_for_updates(self):
        if AUTO_UPDATE_ENABLED:
            check_for_updates(parent=self, silent=False)
        else:
            QMessageBox.information(
                self, "Auto-Updater",
                "Auto-updater is not enabled.\n\n"
                "To enable it, install required dependencies:\n"
                "pip install requests packaging"
            )

    def closeEvent(self, event):
        if hasattr(self, '_update_checker_threads'):
            for thread in self._update_checker_threads[:]:
                if thread.isRunning():
                    thread.quit()
                    thread.wait(2000)
        event.accept()


    def get_current_brand(self):
        return "Both"

    def get_previous_day_ending_balance(self, selected_date, brand="Brand A"):
  

        if self.offline_mode and OFFLINE_SUPPORT and offline_manager:
            return self._get_offline_previous_balance(selected_date, brand)
        
   
        try:
            current    = datetime.datetime.strptime(selected_date, "%Y-%m-%d")
            table_name = "daily_reports_brand_a" if brand == "Brand A" else "daily_reports"
            for days_back in range(1, 11):
                prev_str = (current - datetime.timedelta(days=days_back)).strftime("%Y-%m-%d")
                try:
                    q = f"""SELECT ending_balance FROM {table_name}
                            WHERE date=%s AND branch=%s AND corporation=%s
                            ORDER BY id DESC LIMIT 1"""
                    result = self.db_manager.execute_query(
                        q, (prev_str, self.branch, self.corporation))
                    if result and len(result) > 0:
                        val = result[0].get('ending_balance')
                        if val is not None:
                            return float(val), prev_str
                except Exception as e:
                    logger.warning("Query error for %s: %s", prev_str, e)
        except Exception as e:
            logger.error("get_previous_day_ending_balance(%s): %s", brand, e)
        return None, None

    def _get_offline_previous_balance(self, selected_date, brand):
    
        try:
            pending_bal, pending_date = offline_manager.get_latest_pending_balance(
                self.user_email, self.branch, self.corporation, brand, selected_date
            )
            
            cached_bal, cached_date = offline_manager.get_cached_balance(
                self.user_email, self.branch, self.corporation, brand
            )
            
    
            best_bal, best_date = None, None
            
            if pending_date and pending_date < selected_date:
                best_bal, best_date = pending_bal, pending_date
            
            if cached_date and cached_date < selected_date:
                if best_date is None or cached_date > best_date:
                    best_bal, best_date = cached_bal, cached_date
   
            if pending_date and best_date:
                if pending_date > best_date:
                    best_bal, best_date = pending_bal, pending_date
            
            return best_bal, best_date
            
        except Exception as e:
            logger.error("_get_offline_previous_balance error: %s", e)
            return None, None


    _table_columns_cache = {}  
    _table_columns_cache_time = 0

    def _get_table_columns(self, table_name):
  
        import time as _time
        now = _time.time()
        if (table_name in self._table_columns_cache
                and (now - self._table_columns_cache_time) < 300):
            return self._table_columns_cache[table_name]
        try:
            result = self.db_manager.execute_query(
                "SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS "
                "WHERE TABLE_NAME = %s AND TABLE_SCHEMA = DATABASE()",
                (table_name,)
            )
            if result:
                cols = set()
                for row in result:
                    if isinstance(row, dict):
                        c = row.get('COLUMN_NAME') or row.get('column_name') or ''
                    elif isinstance(row, (list, tuple)) and row:
                        c = row[0]
                    else:
                        c = ''
                    if c:
                        cols.add(c)
                if cols:
                    self._table_columns_cache[table_name] = cols
                    self._table_columns_cache_time = now
                    return cols
        except Exception as e:
            logger.error("_get_table_columns(%s): %s", table_name, e)
        return self._table_columns_cache.get(table_name, set())

    def _filter_vals_for_table(self, all_vals, table_name):
        """Remove keys from all_vals that are not columns in the table."""
        existing = self._get_table_columns(table_name)
        if not existing:
            return all_vals 
        return {k: v for k, v in all_vals.items() if k in existing}

    def check_existing_entry(self, selected_date, brand="Brand A"):

        try:
            if brand == "Brand A":
                table_name = "daily_reports_brand_a"
                # Try with corporation first, then branch-only fallback
                for q, params in [
                    (f"SELECT * FROM {table_name} WHERE date=%s AND branch=%s AND corporation=%s LIMIT 1",
                     (selected_date, self.branch, self.corporation)),
                    (f"SELECT * FROM {table_name} WHERE date=%s AND branch=%s LIMIT 1",
                     (selected_date, self.branch)),
                ]:
                    result = self.db_manager.execute_query(q, params)
                    if result:
                        break
            else:
                # Brand B: daily_reports is exclusively the Brand B table.
                # Query by date+branch+corporation; no brand filter needed.
                table_name = "daily_reports"
                for q, params in [
                    (f"SELECT * FROM {table_name} WHERE date=%s AND branch=%s AND corporation=%s LIMIT 1",
                     (selected_date, self.branch, self.corporation)),
                    (f"SELECT * FROM {table_name} WHERE date=%s AND branch=%s LIMIT 1",
                     (selected_date, self.branch)),
                ]:
                    result = self.db_manager.execute_query(q, params)
                    if result:
                        break
            if result and len(result) > 0:
                row = result[0]
                is_locked = row.get('is_locked', 1)
                return "locked" if is_locked else "unlocked"
        except Exception as e:
            logger.error("check_existing_entry(%s): %s", brand, e)
        return None


    def _clear_all_dialog_caches(self):
        """Clear all cached dialog instances and breakdown data to prevent stale data from persisting."""
        # Clear Empeno Jewelry dialogs (prevents old jewelry entries from showing)
        if hasattr(self, '_jew_dialogs'):
            for dlg in self._jew_dialogs.values():
                if dlg and isinstance(dlg, QDialog):
                    try:
                        dlg.close()
                    except Exception:
                        pass
            self._jew_dialogs.clear()
        
        # Clear Empeno Jewelry computed totals
        if hasattr(self, '_jew_computed'):
            for k in self._jew_computed:
                self._jew_computed[k] = 0.0
        if hasattr(self, '_jew_computed_b'):
            for k in self._jew_computed_b:
                self._jew_computed_b[k] = 0.0
        
        # Clear Motor/Car dialog instance
        if hasattr(self, '_motor_dialog') and self._motor_dialog:
            if isinstance(self._motor_dialog, QDialog):
                try:
                    self._motor_dialog.close()
                except Exception:
                    pass
            self._motor_dialog = None
        
        # Clear Motor/Car breakdown data
        if hasattr(self, '_motor_car_breakdown'):
            self._motor_car_breakdown = {}
        
        # Clear Fund Transfer HO dialogs (prevents old FT data from showing)
        if hasattr(self, '_ft_ho_dialogs'):
            for dlg in self._ft_ho_dialogs.values():
                if dlg and isinstance(dlg, QDialog):
                    try:
                        dlg.close()
                    except Exception:
                        pass
            self._ft_ho_dialogs.clear()
        
        # Clear Fund Transfer HO breakdowns
        if hasattr(self, '_ft_ho_breakdowns'):
            self._ft_ho_breakdowns.clear()
        
        # Clear PC Salary dialog cache
        if hasattr(self, '_salary_dialogs'):
            for dlg in self._salary_dialogs.values():
                if dlg and isinstance(dlg, QDialog):
                    try:
                        dlg.close()
                    except Exception:
                        pass
            self._salary_dialogs.clear()
        
        # Clear PC Salary breakdown data
        if hasattr(self, '_pc_salary_breakdown'):
            self._pc_salary_breakdown = None
        
        # Clear MC (Multi-Currency) In/Out data for both brands
        for cf_tab in [self.cash_flow_tab_a, self.cash_flow_tab_b]:
            if hasattr(cf_tab, 'mc_currency_details'):
                cf_tab.mc_currency_details['MC In'] = []
                cf_tab.mc_currency_details['MC Out'] = []
        
        # Clear Fund Transfer to/from BRANCH data for both brands
        for cf_tab in [self.cash_flow_tab_a, self.cash_flow_tab_b]:
            if hasattr(cf_tab, 'branch_dest_inputs'):
                # Clear the display/cache for Fund Transfer to BRANCH
                ft_to_branch = cf_tab.branch_dest_inputs.get('Fund Transfer to BRANCH')
                if ft_to_branch:
                    ft_to_branch.blockSignals(True)
                    ft_to_branch.clear()
                    ft_to_branch.blockSignals(False)
                
                # Clear the display/cache for Fund Transfer from BRANCH
                ft_from_branch = cf_tab.branch_dest_inputs.get('Fund Transfer from BRANCH')
                if ft_from_branch:
                    ft_from_branch.blockSignals(True)
                    ft_from_branch.clear()
                    ft_from_branch.blockSignals(False)
        
        logger.debug("All dialog caches cleared (Jewelry, Motor, FT HO, FT Branch, MC In/Out, Salary)")

    def _check_internet(self):
        """Submit non-blocking socket check to QThreadPool every 5 seconds."""
        from connection_watcher import _PingWorker
        from PyQt5.QtCore import QThreadPool
        worker = _PingWorker()
        worker.signals.result.connect(self._on_internet_check)
        QThreadPool.globalInstance().start(worker)

    def _on_internet_check(self, ok: bool):
        """Called when socket check completes — show/hide banner based on result."""
        if not ok and self._is_connected:
            # Just went offline
            self._on_connection_lost()
        elif ok and not self._is_connected:
            # Just came back online
            self._on_connection_restored()

    def _hide_conn_banner_and_reset(self):
        self._conn_banner.hide_banner()
        self._conn_banner.setStyleSheet("""
            QWidget#ConnectionBanner { background-color: #c0392b; }
        """)
        self._conn_banner.set_message(
            "⚠  No internet connection — check your cable or try to restart the app."
        )

    def _on_connection_lost(self):
        if self._is_connected is False:
            return  # already in offline mode
        self._is_connected = False

        # Hide any stuck loading overlay
        if hasattr(self, 'loading_overlay'):
            self.loading_overlay.hide()

        # Show offline banner with clear message
        self._conn_banner.set_message(
            "⚠  No internet connection — check your cable or try to restart the app."
        )
        self._conn_banner.show_banner()

        # Pause session timeout
        self._session_timer.stop()

        # Disable ALL input fields when offline
        self._toggle_inputs(False)

        # Disable all action buttons
        self.post_button.setEnabled(False)
        self.post_button.setToolTip("Disabled — no internet connection.")
        self.refresh_button.setEnabled(False)
        self.refresh_button.setToolTip("Disabled — no internet connection.")
        for btn in (self.auto_fill_button_a, self.auto_fill_button_b):
            btn.setEnabled(False)
            btn.setToolTip("Disabled — no internet connection.")

    def _on_connection_restored(self):
        if self._is_connected is True:
            return  # already online
        self._is_connected = True

        # Re-enable all inputs
        self._toggle_inputs(True)

        # Show restored banner briefly
        self._conn_banner.set_message(
            "✓  Connection restored — click Refresh to reload your report."
        )
        self._conn_banner.setStyleSheet("QWidget#ConnectionBanner { background-color: #27ae60; }")
        self._conn_banner.show_banner()

        self._session_timer.start(60_000)

        # Re-enable DB-dependent buttons
        self.post_button.setToolTip("")
        self.refresh_button.setEnabled(True)
        self.refresh_button.setToolTip("Reload the report for the current date.")
        for btn in (self.auto_fill_button_a, self.auto_fill_button_b):
            btn.setEnabled(True)
            btn.setToolTip("")
        self.update_cash_result()

        # Hide restored banner after 3 seconds
        QTimer.singleShot(3000, self._hide_conn_banner_and_reset)

    def _is_network_error(self, exc):
        """Return True if the exception looks like a connectivity or timeout issue."""
        msg = str(exc).lower()
        return any(k in msg for k in (
            "timeout", "timed out", "connection", "network", "unreachable",
            "refused", "reset", "broken pipe", "eof", "errno",
            "requests.exceptions", "connectionerror", "readtimeout",
        ))

    def _show_network_error(self, action="load data"):
        """Show a friendly retry dialog and hide the loading overlay."""
        if hasattr(self, 'loading_overlay'):
            self.loading_overlay.hide()
        QMessageBox.warning(
            self,
            "Connection Problem",
            f"Unable to {action} — the connection is slow or unavailable.\n\n"
            "Please check your internet connection and try again.",
            QMessageBox.Ok
        )

    def on_date_changed(self):
        # Skip entirely when offline — no network calls, no lag
        if not self._is_connected:
            return

        sd = self.date_picker.date().toString("yyyy-MM-dd")

        # Show loading indicator while checking and loading report
        if hasattr(self, 'loading_overlay'):
            self.loading_overlay.set_status("Checking report...", f"Loading {sd}")
            self.loading_overlay.show()
            QApplication.processEvents()

        try:
            self._on_date_changed_inner(sd)
        except Exception as exc:
            if self._is_network_error(exc):
                # Immediately enter offline mode — don't wait for the watcher
                self._on_connection_lost()
            else:
                if hasattr(self, 'loading_overlay'):
                    self.loading_overlay.hide()
                raise

    def _on_date_changed_inner(self, sd):
        # Clear all cached dialogs and breakdowns FIRST to prevent stale data
        self._clear_all_dialog_caches()

        for bb in (self.beginning_balance_input_a, self.beginning_balance_input_b):
            bb.clear()
            bb.setReadOnly(True)
            bb.setStyleSheet(f"""
                QLineEdit {{
                    background-color: {_SLATE_100};
                    border: 1.5px solid {_SLATE_200};
                    border-radius: 6px;
                    padding: 7px 12px;
                    font-size: 14px;
                    font-weight: 700;
                    color: {_TEXT_PRI};
                }}
            """)
        for cc in (self.cash_count_input_a, self.cash_count_input_b):
            cc.clear()
        # Clear cash float input for Brand A
        if hasattr(self, 'cash_float_input_a') and self.cash_float_input_a:
            self.cash_float_input_a.clear()
            self.cash_float_input_a.setReadOnly(False)  # Make editable by default

        for tab in (self.cash_flow_tab_a, self.cash_flow_tab_b, self.palawan_tab):
            if hasattr(tab, 'clear_fields'):
                tab.clear_fields()

        # Load payable data (both sendout/payout/int and adjustments for locked reports)
        self._restore_palawan_payable(sd)

        self.beginning_balance_auto_filled_a = False
        self.beginning_balance_auto_filled_b = False
        self.previous_day_balance_a = self.previous_day_date_a = None
        self.previous_day_balance_b = self.previous_day_date_b = None

        status_a = self.check_existing_entry(sd, "Brand A")
        status_b = self.check_existing_entry(sd, "Brand B")
        logger.debug(f"on_date_changed: date={sd} branch={self.branch} corp={self.corporation} status_a={status_a} status_b={status_b}")

        if status_a == "locked" and status_b == "locked":

            try:
                if hasattr(self, 'loading_overlay'):
                    self.loading_overlay.set_status("Loading submitted reports...", "Please wait")
                    self.loading_overlay.show()
                    QApplication.processEvents()

                self._load_brand_report_data("Brand A", sd)
                self._load_brand_report_data("Brand B", sd)
                # Load Palawan tab data including adjustment fields from daily_reports
                self._restore_palawan_payable(sd)
                self._restore_palawan_tab(sd)
            finally:
                try:
                    if hasattr(self, 'loading_overlay'):
                        self.loading_overlay.hide()
                except Exception:
                    pass

            for brand in ("A", "B"):
                self._set_status_brand(brand, "Submitted", _RED_500, bold=True)
            self.auto_fill_button_a.setEnabled(False)
            self.auto_fill_button_b.setEnabled(False)
            self.post_button.setEnabled(False)
            # Prevent loading draft for submitted reports
            if hasattr(self, 'load_draft_button'):
                self.load_draft_button.setEnabled(False)
            self._toggle_inputs(False)
            # Make cash float read-only when viewing locked report
            if hasattr(self, 'cash_float_input_a') and self.cash_float_input_a:
                self.cash_float_input_a.setReadOnly(True)
            # Silently remove drafts for all submitted dates
            self._purge_submitted_drafts()
            return

        self._toggle_inputs(True)
        if hasattr(self, 'load_draft_button'):
            self.load_draft_button.setEnabled(True)

        any_unlocked = False
        for brand, status, bf in [("A", status_a, "Brand A"), ("B", status_b, "Brand B")]:
            bb_input = self.beginning_balance_input_a if brand == "A" else self.beginning_balance_input_b
            af_btn   = self.auto_fill_button_a        if brand == "A" else self.auto_fill_button_b
            cf_tab   = self.cash_flow_tab_a           if brand == "A" else self.cash_flow_tab_b
            cc_input = self.cash_count_input_a        if brand == "A" else self.cash_count_input_b

            if status == "locked":

                self._set_status_brand(brand, "Submitted", _RED_500, bold=True)
                af_btn.setEnabled(False)
                bb_input.setEnabled(False)
                cc_input.setEnabled(False)
                if hasattr(cf_tab, 'set_enabled'):
                    cf_tab.set_enabled(False)
            elif status == "unlocked":

                any_unlocked = True

                self._load_brand_report_data(bf, sd)
  
                self._set_status_brand(brand, "🔄 Reset by Admin — Edit & Resubmit", "#E67E22", bold=True)
                af_btn.setEnabled(True)
                bb_input.setEnabled(True)
                bb_input.setReadOnly(False)
                cc_input.setEnabled(True)
                if hasattr(cf_tab, 'set_enabled'):
                    cf_tab.set_enabled(True)
         
                if brand == "A":
                    self.beginning_balance_auto_filled_a = True
                else:
                    self.beginning_balance_auto_filled_b = True

                prev_bal, prev_date = self.get_previous_day_ending_balance(sd, bf)
                if prev_bal is not None:
                    if brand == "A":
                        self.previous_day_balance_a, self.previous_day_date_a = prev_bal, prev_date
                    else:
                        self.previous_day_balance_b, self.previous_day_date_b = prev_bal, prev_date
            else:
     
                af_btn.setEnabled(True)
                prev_bal, prev_date = self.get_previous_day_ending_balance(sd, bf)
                if prev_bal is not None:
                    if brand == "A":
                        self.previous_day_balance_a, self.previous_day_date_a = prev_bal, prev_date
                    else:
                        self.previous_day_balance_b, self.previous_day_date_b = prev_bal, prev_date
                    self._set_status_brand(brand,
                        f"Found: {prev_date}  ·  {prev_bal:,.2f}", _TEXT_MUTED)
                else:
                    self._set_status_brand(brand, "No previous record — enter manually", _AMBER_500)
                    bb_input.setReadOnly(False)
                    bb_input.setPlaceholderText("Enter opening balance")
                    bb_input.setStyleSheet(f"""
                        QLineEdit {{
                            background-color: {_BG_INPUT};
                            border: 1.5px solid {_BORDER};
                            border-radius: 6px;
                            padding: 7px 12px;
                            font-size: 14px;
                            font-weight: 700;
                            color: {_TEXT_PRI};
                        }}
                        QLineEdit:focus {{
                            border: 2px solid {_PRIMARY};
                            background-color: {_INDIGO_50};
                        }}
                    """)


        if any_unlocked or not status_a or not status_b:
            self.post_button.setEnabled(True)

        # Always restore palawan tab from Brand A table regardless of lock state
        self._restore_palawan_tab(sd)

        self.recalculate_all()

        # Hide loading overlay after report is fully loaded
        if hasattr(self, 'loading_overlay'):
            self.loading_overlay.hide()

    def _restore_palawan_tab(self, date_str):
        """Load palawan data from both Brand A and Brand B tables.
        Brand A is the canonical source; Brand B values take precedence for
        all palawan payable and lotes columns when non-zero (needed because
        daily_reports_brand_a may not have the palawan payable columns)."""
        _PALAWAN_COLS = (
            'palawan_sendout_principal', 'palawan_sendout_sc',
            'palawan_sendout_commission', 'palawan_sendout_regular_total',
            'palawan_sendout_lotes_total',
            'palawan_payout_principal', 'palawan_payout_sc',
            'palawan_payout_commission', 'palawan_payout_regular_total',
            'palawan_payout_lotes_total',
            'palawan_international_principal', 'palawan_international_sc',
            'palawan_international_commission', 'palawan_international_regular_total',
            'palawan_international_lotes_total',
        )
        try:
            params = (date_str, self.branch, self.corporation)
            result_a = self.db_manager.execute_query(
                "SELECT * FROM daily_reports_brand_a "
                "WHERE date=%s AND branch=%s AND corporation=%s LIMIT 1", params
            )
            result_b = self.db_manager.execute_query(
                "SELECT * FROM daily_reports "
                "WHERE date=%s AND branch=%s AND corporation=%s LIMIT 1", params
            )
            if not result_b:
                result_b = self.db_manager.execute_query(
                    "SELECT * FROM daily_reports "
                    "WHERE date=%s AND branch=%s LIMIT 1",
                    (date_str, self.branch)
                )

            if not result_a and not result_b:
                return

            data = dict(result_a[0]) if result_a else {}
            if result_b:
                data_b = result_b[0]
                for col in _PALAWAN_COLS:
                    b_val = data_b.get(col) or 0
                    try:
                        b_float = float(b_val)
                    except (TypeError, ValueError):
                        b_float = 0.0
                    if b_float != 0:
                        data[col] = b_val

            self.palawan_tab.load_data(data)
        except Exception as e:
            logger.error("[_restore_palawan_tab] %s", e)

    def _restore_palawan_payable(self, date_str):
        """Load Palawan adjustment fields from the daily_reports tables.
        Adjustment fields (Skid, Skir, Cancel, Inc) can be stored for BOTH brands.
        Brand B values take precedence when non-zero."""
        mapped = {}

        # Load adjustments from Brand A (daily_reports_brand_a)
        try:
            result_a = self.db_manager.execute_query(
                "SELECT palawan_suki_discounts, palawan_suki_rebates, "
                "palawan_cancel, palawan_pay_out_incentives "
                "FROM daily_reports_brand_a "
                "WHERE date=%s AND branch=%s AND corporation=%s LIMIT 1",
                (date_str, self.branch, self.corporation)
            )
            if result_a:
                row = dict(result_a[0])
                for col in ("palawan_suki_discounts", "palawan_suki_rebates",
                            "palawan_cancel", "palawan_pay_out_incentives"):
                    val = row.get(col) or 0
                    if float(val) != 0:
                        mapped[col] = val
        except Exception as e:
            logger.error("[_restore_palawan_payable] Brand A query error: %s", e)

        # Load adjustments from Brand B (daily_reports — no brand column, exclusively Brand B)
        try:
            result_b = self.db_manager.execute_query(
                "SELECT palawan_suki_discounts, palawan_suki_rebates, "
                "palawan_cancel, palawan_pay_out_incentives "
                "FROM daily_reports "
                "WHERE date=%s AND branch=%s AND corporation=%s LIMIT 1",
                (date_str, self.branch, self.corporation)
            )
            if result_b:
                row = dict(result_b[0])
                for col in ("palawan_suki_discounts", "palawan_suki_rebates",
                            "palawan_cancel", "palawan_pay_out_incentives"):
                    val = row.get(col) or 0
                    if float(val) != 0:
                        mapped[col] = val  # Brand B value takes precedence
        except Exception as e:
            logger.error("[_restore_palawan_payable] Brand B query error: %s", e)

        if mapped:
            try:
                self.palawan_tab.load_data(mapped)
            except Exception as e:
                logger.error("[_restore_palawan_payable] load_data error: %s", e)

    def _set_status(self, text, color, bold=False):
        self._set_status_brand("A", text, color, bold)

    def _set_status_brand(self, brand, text, color, bold=False):
        lbl = self.balance_status_label_a if brand == "A" else self.balance_status_label_b
        w   = "700" if bold else "500"
        lbl.setText(text)
        lbl.setStyleSheet(
            f"font-size: 12px; font-weight: {w}; color: {color}; background: transparent;"
        )


    def auto_fill_beginning_balance(self, brand="A"):
        if not self._is_connected:
            return  # Auto-fill button is already disabled offline — safety guard
        sd        = self.date_picker.date().toString("yyyy-MM-dd")
        brand_full = "Brand A" if brand == "A" else "Brand B"

        status = self.check_existing_entry(sd, brand_full)
        if status == "locked":
            self._msg("Entry Exists",
                      f"An entry already exists for {brand_full} on {sd} and it is locked.",
                      QMessageBox.Warning)
            return

        prev_bal  = self.previous_day_balance_a if brand == "A" else self.previous_day_balance_b
        prev_date = self.previous_day_date_a    if brand == "A" else self.previous_day_date_b
        if prev_bal is None:
            prev_bal, prev_date = self.get_previous_day_ending_balance(sd, brand_full)
        bb_input  = self.beginning_balance_input_a if brand == "A" else self.beginning_balance_input_b
        af_btn    = self.auto_fill_button_a        if brand == "A" else self.auto_fill_button_b

        if prev_bal is not None:
            bb_input.setText(f"{prev_bal:.2f}")
            bb_input.setReadOnly(True)
            bb_input.setStyleSheet(f"""
                QLineEdit {{
                    background-color: {_SLATE_50};
                    border: 2px solid {_INDIGO_500};
                    border-radius: 6px;
                    padding: 7px 12px;
                    font-size: 14px;
                    font-weight: 700;
                    color: #065F46;
                }}
            """)
            self._set_status_brand(brand,
                f"Loaded from {prev_date}",
                _EMERALD_500, bold=True)
            af_btn.setText("✓  Loaded")
            af_btn.setEnabled(False)
            af_btn.setStyleSheet(f"""
                QPushButton {{
                    background-color: transparent;
                    color: {_EMERALD_500};
                    border: 1.5px solid {_INDIGO_500};
                    border-radius: 5px;
                    font-size: 12px;
                    font-weight: 700;
                    padding: 4px 14px;
                    min-width: 0;
                }}
            """)
            if brand == "A":
                self.beginning_balance_auto_filled_a = True
            else:
                self.beginning_balance_auto_filled_b = True
            self._msg("Balance Loaded",
                      f"{brand_full} opening balance set to {prev_bal:,.2f}",
                      QMessageBox.Information)
        else:
            self._msg("No Previous Record",
                      f"No previous record for {brand_full}. Enter the opening balance manually.",
                      QMessageBox.Information)
            bb_input.setReadOnly(False)
            bb_input.setPlaceholderText("Enter opening balance")
            bb_input.setStyleSheet(f"""
                QLineEdit {{
                    background-color: {_BG_INPUT};
                    border: 1.5px solid {_BORDER};
                    border-radius: 6px;
                    padding: 7px 12px;
                    font-size: 14px;
                    font-weight: 700;
                    color: {_TEXT_PRI};
                }}
                QLineEdit:focus {{
                    border: 2px solid {_PRIMARY};
                    background-color: {_INDIGO_50};
                }}
            """)
            if brand == "A":
                self.beginning_balance_auto_filled_a = True
            else:
                self.beginning_balance_auto_filled_b = True

    def _toggle_inputs(self, enabled):
        for w in (self.beginning_balance_input_a, self.beginning_balance_input_b,
                  self.cash_count_input_a, self.cash_count_input_b):
            w.setEnabled(enabled)
        for tab in (self.cash_flow_tab_a, self.cash_flow_tab_b, self.palawan_tab):
            if hasattr(tab, 'set_enabled'):
                tab.set_enabled(enabled)

    def disable_all_inputs(self):
        self._toggle_inputs(False)

    def enable_all_inputs(self):
        self._toggle_inputs(True)

    def recalculate_all(self):
        for brand in ("a", "b"):
            bb  = getattr(self, f"beginning_balance_input_{brand}")
            cft = getattr(self, f"cash_flow_tab_{brand}")
            eb  = getattr(self, f"ending_balance_display_{brand}")
            try:
                beg = float(bb.text().strip().replace(",", "") or 0)
            except ValueError:
                beg = 0.0
            deb  = cft.get_debit_total()
            cred = cft.get_credit_total()
            cft.update_totals(beg, deb, cred)
            end = beg + deb - cred
            eb.setText(f"{end:,.2f}")
            c = _EMERALD_400 if end > 0 else (_RED_400 if end < 0 else _WHITE)
            eb.setStyleSheet(
                f"font-size: 15px; font-weight: 800; color: {c}; background: transparent;"
            )

        self.palawan_tab.calculate_palawan_totals()
        self.update_cash_result()

    def update_cash_result(self):
        ready_a = self._update_brand_variance("A")
        ready_b = self._update_brand_variance("B")
        # Never re-enable post button while offline
        if self._is_connected:
            self.post_button.setEnabled(ready_a and ready_b)

    def _update_brand_variance(self, brand):
        bb  = self.beginning_balance_input_a if brand == "A" else self.beginning_balance_input_b
        cc  = self.cash_count_input_a        if brand == "A" else self.cash_count_input_b
        eb  = self.ending_balance_display_a  if brand == "A" else self.ending_balance_display_b
        var = self.cash_result_display_a     if brand == "A" else self.cash_result_display_b
        stl = self.variance_status_label_a   if brand == "A" else self.variance_status_label_b
        try:
            cc_val  = float(cc.text().strip().replace(",", "") or 0)
            eb_val  = float((eb.text() or "0").replace(",", ""))
            diff    = cc_val - eb_val
            var.setText(f"{diff:,.2f}")
            has_req = bool(bb.text().strip()) and bool(cc.text().strip())
            font    = "font-size: 15px; font-weight: 800; background: transparent;"

            if abs(diff) < 0.01:
                var.setStyleSheet(font + f"color: {_EMERALD_400};")
                if has_req:
                    stl.setText("✓  Balanced")
                    stl.setStyleSheet(
                        f"font-size: 12px; font-weight: 700; color: {_EMERALD_400}; background: transparent;"
                    )
                else:
                    stl.setText("Fill required fields")
                    stl.setStyleSheet(
                        f"font-size: 12px; font-weight: 600; color: {_SLATE_500}; background: transparent;"
                    )
            elif diff > 0:
                var.setStyleSheet(font + f"color: {_AMBER_400};")
                stl.setText(f"Over  +{diff:,.2f}")
                stl.setStyleSheet(
                    f"font-size: 12px; font-weight: 700; color: {_AMBER_400}; background: transparent;"
                )
            else:
                var.setStyleSheet(font + f"color: {_RED_400};")
                stl.setText(f"Short  {diff:,.2f}")
                stl.setStyleSheet(
                    f"font-size: 12px; font-weight: 700; color: {_RED_400}; background: transparent;"
                )

            return has_req
        except ValueError:
            var.setText("—")
            stl.setText("Invalid input")
            stl.setStyleSheet(
                f"font-size: 12px; font-weight: 600; color: {_RED_500}; background: transparent;"
            )
            return False

    def _set_var_status(self, text, color):
        self.variance_status_label_a.setText(text)
        self.variance_status_label_a.setStyleSheet(
            f"font-size: 12px; font-weight: 700; color: {color}; background: transparent;"
        )

    def _check_optional_tabs_empty(self):
        empty_tabs = []
        try:
            if all(v == 0 or v == 0.0 for v in self.palawan_tab.get_data().values()):
                empty_tabs.append("Palawan Details B")
        except Exception:
            empty_tabs.append("Palawan Details B")

        if not empty_tabs:
            return True

        dlg = QMessageBox(self)
        dlg.setWindowTitle("Incomplete Report")
        dlg.setIcon(QMessageBox.Warning)
        dlg.setText(
            f"<b>The following tabs have no data:</b><br><br>"
            + "<br>".join(f"&nbsp;&nbsp;&#8226;&nbsp; <b>{t}</b>" for t in empty_tabs)
            + "<br><br>Are you sure you want to post without filling them in?"
        )
        dlg.setInformativeText(
            "Click <b>Go Back</b> to add missing details, "
            "or <b>Post Anyway</b> to submit with zeros."
        )
        btn_post = dlg.addButton("Post Anyway",  QMessageBox.AcceptRole)
        btn_back = dlg.addButton("Go Back",       QMessageBox.RejectRole)
        dlg.setDefaultButton(btn_back)
        dlg.setStyleSheet(self._MSG_QSS)
        btn_post.setStyleSheet(f"background-color:{_AMBER_500};color:{_WHITE};border:none;")
        btn_back.setStyleSheet(f"background-color:{_PRIMARY};color:{_WHITE};border:none;")
        dlg.exec_()
        return dlg.clickedButton() == btn_post


    def validate_calculations(self):
        """
        CRITICAL: Comprehensive validation to BLOCK any database save with calculation errors.
        
        CORRECT FORMULA:
        - Debit Total = Beginning Balance + Sum of all debit field values
        - Credit Total = Sum of all credit field values  
        - Ending Balance = Debit Total - Credit Total
        
        ⚠️ STRICT MODE: If ANY calculation is wrong, the post is BLOCKED and NO DATA reaches database.
        """
        errors = []
        
        for brand_index, (brand_char, bb_input, cf_tab) in enumerate([
            ("A", self.beginning_balance_input_a, self.cash_flow_tab_a),
            ("B", self.beginning_balance_input_b, self.cash_flow_tab_b)
        ]):
            try:
                # Get beginning balance
                beginning = float(bb_input.text().strip().replace(',', '') or 0)
                if beginning < 0:
                    errors.append(f"Brand {brand_char}: Beginning Balance cannot be negative ({beginning:.2f})")
                    continue
            except (ValueError, TypeError):
                errors.append(f"Brand {brand_char}: Invalid Beginning Balance value")
                continue
            
            cf_data = cf_tab.get_data()
            
            # Sum debit and credit field values (excluding _lotes) - MUST MATCH save code exactly!
            # This includes all values (including 0) to match what handle_post() does
            debit_field_sum = sum(v for k, v in cf_data.get('debit', {}).items() 
                                 if not k.endswith('_lotes'))
            credit_field_sum = sum(v for k, v in cf_data.get('credit', {}).items() 
                                  if not k.endswith('_lotes'))
            
            # Calculate what totals SHOULD be using CORRECT FORMULA
            calculated_debit_total = beginning + debit_field_sum
            calculated_credit_total = credit_field_sum
            calculated_ending = calculated_debit_total - calculated_credit_total
            
            # Get displayed totals from UI
            debit_fields_only = cf_tab.get_debit_total()  # This is just fields sum
            displayed_credit = cf_tab.get_credit_total()
            
            # STRICT VALIDATION #1: Verify debit fields sum matches
            if abs(debit_fields_only - debit_field_sum) > 0.01:
                errors.append(f"Brand {brand_char} DEBIT FIELDS SUM: Calculation Error!\n"
                            f"  Should be: {debit_field_sum:.2f}\n"
                            f"  Shows: {debit_fields_only:.2f}\n"
                            f"  Difference: {abs(debit_fields_only - debit_field_sum):.2f}")
            
            # STRICT VALIDATION #2: Verify credit fields sum matches
            if abs(displayed_credit - credit_field_sum) > 0.01:
                errors.append(f"Brand {brand_char} CREDIT FIELDS SUM: Calculation Error!\n"
                            f"  Should be: {credit_field_sum:.2f}\n"
                            f"  Shows: {displayed_credit:.2f}\n"
                            f"  Difference: {abs(displayed_credit - credit_field_sum):.2f}")
            
            # STRICT VALIDATION #3: Verify debit total = beginning + fields
            displayed_debit = beginning + debit_fields_only
            if abs(displayed_debit - calculated_debit_total) > 0.01:
                errors.append(f"Brand {brand_char} DEBIT TOTAL: Calculation Error!\n"
                            f"  Formula: Beginning ({beginning:.2f}) + Fields ({debit_field_sum:.2f})\n"
                            f"  Should be: {calculated_debit_total:.2f}\n"
                            f"  Error detected!")
            
            # STRICT VALIDATION #4: Verify ending balance calculation
            displayed_ending = displayed_debit - displayed_credit
            if abs(displayed_ending - calculated_ending) > 0.01:
                errors.append(f"Brand {brand_char} ENDING BALANCE: Calculation Error!\n"
                            f"  Formula: Debit ({displayed_debit:.2f}) - Credit ({displayed_credit:.2f})\n"
                            f"  Should be: {calculated_ending:.2f}\n"
                            f"  Shows: {displayed_ending:.2f}")
        
        if errors:
            error_text = "🚫 VALIDATION FAILED - DATABASE SAVE BLOCKED 🚫\n\n"
            error_text += "Calculation errors detected:\n\n"
            error_text += "\n\n".join(errors)
            error_text += "\n\n⚠️ NO DATA HAS BEEN SAVED TO DATABASE ⚠️\n\n"
            error_text += "Please verify your calculations and try again."
            
            QMessageBox.critical(self, "POSTING BLOCKED - Calculation Error", error_text)
            return False
        
        return True

    def verify_database_save(self, date_str, brand_table,
                              expected_debit=None, expected_credit=None, expected_ending=None,
                              brand=None):
        """
        POST-SAVE VERIFICATION: Read data back from database and verify stored totals
        match the values that were calculated before saving.

        Expected values are passed in from handle_post so we compare against the
        authoritative pre-save calculation rather than re-summing individual columns
        (which may not be stored when a credit/debit field has no DB column in the table).
        """
        try:
            query = f"SELECT beginning_balance, debit_total, credit_total, ending_balance FROM {brand_table} WHERE date = %s AND branch = %s AND corporation = %s"
            result = self.db_manager.execute_query(query, (date_str, self.branch, self.corporation))

            if not result:
                logger.warning("Could not verify %s save - record not found!", brand_table)
                return True

            saved_row = result[0]
            stored_debit   = float(saved_row.get('debit_total', 0))
            stored_credit  = float(saved_row.get('credit_total', 0))
            stored_ending  = float(saved_row.get('ending_balance', 0))

            if expected_debit is None or expected_credit is None or expected_ending is None:
                # Fallback: nothing to compare against, just log what was stored
                logger.info("%s stored: Debit=%.2f, Credit=%.2f, Ending=%.2f", brand_table, stored_debit, stored_credit, stored_ending)
                return True

            debit_match  = abs(stored_debit  - expected_debit)  < 0.01
            credit_match = abs(stored_credit - expected_credit) < 0.01
            ending_match = abs(stored_ending - expected_ending) < 0.01

            if not debit_match:
                logger.error("%s Debit Total mismatch! Expected=%.2f Stored=%.2f Diff=%+.2f",
                             brand_table, expected_debit, stored_debit, stored_debit - expected_debit)
                return False

            if not credit_match:
                logger.error("%s Credit Total mismatch! Expected=%.2f Stored=%.2f Diff=%+.2f",
                             brand_table, expected_credit, stored_credit, stored_credit - expected_credit)
                return False

            if not ending_match:
                logger.error("%s Ending Balance mismatch! Expected=%.2f Stored=%.2f Diff=%+.2f",
                             brand_table, expected_ending, stored_ending, stored_ending - expected_ending)
                return False

            logger.info("%s verified: Debit=%.2f, Credit=%.2f, Ending=%.2f",
                        brand_table, stored_debit, stored_credit, stored_ending)
            return True

        except Exception as e:
            logger.error("Could not verify database save: %s", e)
            return True  # Don't fail hard, just log it

    def _propagate_opening_balance_to_following_days(self, date_str, brand_full, ending_balance):
        table_name = "daily_reports_brand_a" if brand_full == "Brand A" else "daily_reports"
        try:
            current = datetime.datetime.strptime(date_str, "%Y-%m-%d")
        except Exception as e:
            logger.error("Invalid date for propagation: %s", e)
            return

        next_date = current + datetime.timedelta(days=1)
        while True:
            next_date_str = next_date.strftime("%Y-%m-%d")
            query = (
                f"SELECT beginning_balance, debit_total, credit_total, cash_count, ending_balance, is_locked "
                f"FROM {table_name} WHERE date = %s AND branch = %s AND corporation = %s LIMIT 1"
            )
            result = self.db_manager.execute_query(query, (next_date_str, self.branch, self.corporation))
            if not result:
                break

            row = result[0]
            if row.get("is_locked", 1):
                break

            old_begin = float(row.get("beginning_balance", 0) or 0)
            if abs(old_begin - ending_balance) < 0.01:
                ending_balance = float(row.get("ending_balance", 0) or 0)
                next_date += datetime.timedelta(days=1)
                continue

            debit_total = float(row.get("debit_total", 0) or 0)
            credit_total = float(row.get("credit_total", 0) or 0)
            cash_count = float(row.get("cash_count", 0) or 0)

            debit_amount = debit_total - old_begin
            new_debit_total = ending_balance + debit_amount
            new_ending = ending_balance + debit_amount - credit_total
            new_cash_result = cash_count - new_ending
            variance_status = (
                "balanced" if abs(new_cash_result) < 0.01
                else "over" if new_cash_result > 0
                else "short"
            )

            update_query = (
                f"UPDATE {table_name} SET beginning_balance = %s, debit_total = %s, "
                f"ending_balance = %s, cash_result = %s, variance_status = %s "
                f"WHERE date = %s AND branch = %s AND corporation = %s"
            )
            self.db_manager.execute_query(
                update_query,
                (ending_balance, new_debit_total, new_ending, new_cash_result,
                 variance_status, next_date_str, self.branch, self.corporation)
            )

            ending_balance = new_ending
            next_date += datetime.timedelta(days=1)

    def handle_post(self):
        if not self._is_connected:
            return  # Post button is already disabled offline — safety guard
        try:
            sd = self.date_picker.date().toString("yyyy-MM-dd")
            if not self._check_optional_tabs_empty():
                return
            if not self.validate_all_requirements():
                return
            
            # CRITICAL: Verify calculations before saving
            if not self.validate_calculations():
                return

            if self.offline_mode:
                self._handle_offline_post(sd)
                return

            self.loading_overlay.set_status("Preparing data...", "Validating report fields")
            self.loading_overlay.show()
            self.loading_overlay.raise_()
            QApplication.processEvents()

            pal = self.palawan_tab.get_data()

            brand_specs = [
                ("Brand A", self.cash_flow_tab_a,
                 self.beginning_balance_input_a, self.cash_count_input_a,
                 "daily_reports_brand_a"),
                ("Brand B", self.cash_flow_tab_b,
                 self.beginning_balance_input_b, self.cash_count_input_b,
                 "daily_reports"),
            ]

            results = []
            brand_all_vals = {}  
            for brand_full, cf_tab, bb_input, cc_input, table_name in brand_specs:
  
                self.loading_overlay.set_status(f"Posting {brand_full}...", f"Saving to {table_name}")
                QApplication.processEvents()
                
                entry_status = self.check_existing_entry(sd, brand_full)
                if entry_status == "locked":
                    results.append((brand_full, "skipped", None))
                    continue

                cf        = cf_tab.get_data()
                beginning = float(bb_input.text().strip().replace(',', '') or 0)
                deb = sum(v for k, v in cf['debit'].items()  if not k.endswith('_lotes'))
                cre = sum(v for k, v in cf['credit'].items() if not k.endswith('_lotes'))
                ending      = beginning + deb - cre
                cash_count  = float(cc_input.text().strip().replace(',', '') or 0)
                cash_result = cash_count - ending
                variance_status = (
                    "balanced" if abs(cash_result) < 0.01
                    else "over"  if cash_result > 0
                    else "short"
                )

                # Merge: palawan tab provides supplementary fields; credit tab takes
                # precedence for any key that appears in both (non-zero credit wins).
                all_vals = {**cf['debit'], **pal}
                for k, v in cf['credit'].items():
                    if v != 0 or k not in all_vals:
                        all_vals[k] = v

                if hasattr(cf_tab, 'selected_bank_account') and cf_tab.selected_bank_account:
                    all_vals['fund_transfer_bank_account'] = cf_tab.selected_bank_account
                
   
                if hasattr(self, '_ft_ho_breakdowns'):
                    bd = self._ft_ho_breakdowns.get(brand_full)
                    if bd:
                        all_vals['ft_ho_breakdown'] = json.dumps(bd)
                
                
                if hasattr(cf_tab, 'branch_dest_inputs'):
                    branch_dest = cf_tab.branch_dest_inputs.get('Fund Transfer to BRANCH')
                    if branch_dest and branch_dest.text().strip():
                        all_vals['fund_transfer_to_branch_dest'] = branch_dest.text().strip()
                
              
                if hasattr(cf_tab, 'branch_dest_inputs'):
                    branch_src = cf_tab.branch_dest_inputs.get('Fund Transfer from BRANCH')
                    if branch_src and branch_src.text().strip():
                        all_vals['fund_transfer_from_branch_dest'] = branch_src.text().strip()
                
     
                if brand_full == "Brand A" and hasattr(self, '_pc_salary_breakdown') and self._pc_salary_breakdown:
                    all_vals['pc_salary_breakdown'] = json.dumps(self._pc_salary_breakdown)
                
      
                if hasattr(cf_tab, 'mc_currency_details'):
                    mc_in_details = cf_tab.mc_currency_details.get('MC In', [])
                    mc_out_details = cf_tab.mc_currency_details.get('MC Out', [])
                    if mc_in_details:
                        all_vals['mc_in_details'] = json.dumps(mc_in_details)
                    if mc_out_details:
                        all_vals['mc_out_details'] = json.dumps(mc_out_details)
                

                if hasattr(self, '_motor_car_breakdown') and self._motor_car_breakdown:
                    all_vals['empeno_motor_car_breakdown'] = json.dumps(self._motor_car_breakdown)
                
                brand_all_vals[brand_full] = all_vals  

                # CRITICAL SAFETY CHECK: Verify this brand's calculations one more time before saving
                # This is a secondary check to ensure no corrupted data reaches database
                expected_debit = beginning + deb
                expected_credit = cre
                expected_ending = expected_debit - expected_credit
                
                if expected_debit < 0:
                    results.append((brand_full, "validation_error",
                                   f"BLOCKED: Debit total is negative (invalid)"))
                    continue
                if expected_credit < 0:
                    results.append((brand_full, "validation_error",
                                   f"BLOCKED: Credit total is negative (invalid)"))
                    continue
                if abs(ending - expected_ending) > 0.01:
                    results.append((brand_full, "validation_error",
                                   f"BLOCKED: Ending balance mismatch (expected {expected_ending:.2f}, got {ending:.2f})"))
                    continue

                if entry_status == "unlocked":
         
                    update_fields = {
                        'username': self.user_email,
                        'beginning_balance': beginning,
                        'debit_total': beginning + deb,
                        'credit_total': cre,
                        'ending_balance': ending,
                        'cash_count': cash_count,
                        'cash_result': cash_result,
                        'variance_status': variance_status,
                        'is_locked': 1,
                    }
                    update_fields.update(self._filter_vals_for_table(all_vals, table_name))
                    set_clause = ', '.join(f"{c} = %s" for c in update_fields.keys())
                    vals = list(update_fields.values()) + [sd, self.branch, self.corporation]
                    query = (
                        f"UPDATE {table_name} SET {set_clause} "
                        f"WHERE date = %s AND branch = %s AND corporation = %s"
                    )
                else:

                    filtered = self._filter_vals_for_table(all_vals, table_name)
                    # Include brand column for daily_reports so it doesn't default to 'Brand A'
                    _brand_cols = ['brand'] if table_name == 'daily_reports' else []
                    _brand_vals = [brand_full] if table_name == 'daily_reports' else []
                    cols = [
                        'date', 'username', 'branch', 'corporation',
                        'beginning_balance', 'debit_total', 'credit_total',
                        'ending_balance', 'cash_count', 'cash_result', 'variance_status',
                        'is_locked'
                    ] + _brand_cols + list(filtered.keys())
                    vals = [
                        sd, self.user_email, self.branch, self.corporation,
                        beginning, beginning + deb, cre,
                        ending, cash_count, cash_result, variance_status,
                        1
                    ] + _brand_vals + list(filtered.values())

                    ph    = ', '.join(['%s'] * len(cols))
                    # Build ON DUPLICATE KEY UPDATE clause for all non-key columns.
                    # Only updates when the existing row is unlocked (is_locked = 0),
                    # so admin-locked records are never silently overwritten.
                    _key_cols = {'date', 'branch', 'corporation'}
                    _upd_pairs = [
                        f"`{c}` = IF(is_locked = 0, VALUES(`{c}`), `{c}`)"
                        for c in cols if c not in _key_cols
                    ]
                    _upd_clause = ', '.join(_upd_pairs)
                    query = (
                        f"INSERT INTO {table_name} ({', '.join(cols)}) VALUES ({ph})"
                        f" ON DUPLICATE KEY UPDATE {_upd_clause}"
                    )
                    # Keep a copy so we can build a fallback UPDATE on duplicate key
                    # (used by the error-1062 handler for databases without the UNIQUE constraint)
                    _insert_cols = list(cols)
                    _insert_vals = list(vals)

                rows, last_err = None, None
                for attempt in range(1, 5):
                    res, err = self.db_manager.execute_query_with_exception(query, vals)
                    if err is None:
                        rows = res
                        break
                    last_err = err
                    _err_str = str(err)
                    is_dl  = (
                        (hasattr(err, 'args') and isinstance(err.args, tuple)
                         and len(err.args) > 0 and err.args[0] == 1213)
                        or '1213' in _err_str
                    )
                    is_dup = (
                        (hasattr(err, 'args') and isinstance(err.args, tuple)
                         and len(err.args) > 0 and err.args[0] == 1062)
                        or '1062' in _err_str
                    )
                    if is_dl and attempt < 4:
                        try:
                            self.db_manager.logger.error(f"DB attempt {attempt}: {err}")
                        except Exception:
                            pass
                        time.sleep(0.5 * attempt)
                    elif is_dup and entry_status is None:
                        # INSERT hit a duplicate key — the record already exists but
                        # check_existing_entry missed it. Re-check lock status first.
                        recheck = self.check_existing_entry(sd, brand_full)
                        if recheck == "locked":
                            # Locked by admin/another session — do NOT overwrite
                            last_err = None
                            rows = 0
                            results.append((brand_full, "skipped", None))
                            break
                        # Unlocked (or still undetected) — fall back to UPDATE
                        upd_pairs = [
                            (c, v) for c, v in zip(_insert_cols, _insert_vals)
                            if c not in ('date', 'branch', 'corporation', 'brand')
                        ]
                        set_clause_fb = ', '.join(f"`{c}` = %s" for c, _ in upd_pairs)
                        query = (
                            f"UPDATE {table_name} SET {set_clause_fb} "
                            f"WHERE date = %s AND branch = %s AND corporation = %s"
                        )
                        vals = [v for _, v in upd_pairs] + [sd, self.branch, self.corporation]
                        entry_status = "unlocked"  # treat as update from here
                        # continue loop to retry with UPDATE
                    else:
                        try:
                            self.db_manager.logger.error(f"DB attempt {attempt}: {err}")
                        except Exception:
                            pass
                        break

                if isinstance(rows, int) and rows > 0:
                    # POST-SAVE VERIFICATION: Check database for calculation errors
                    is_valid = self.verify_database_save(
                        sd, table_name,
                        expected_debit=beginning + deb,
                        expected_credit=cre,
                        expected_ending=ending,
                        brand=brand_full,
                    )
                    if not is_valid:
                        logger.critical("Database validation failed for %s! Removing record from database.", table_name)
                        try:
                            if entry_status is None:
                                # New record was just inserted — DELETE it entirely so no bad data remains
                                self.db_manager.execute_query(
                                    f"DELETE FROM {table_name} WHERE date = %s AND branch = %s AND corporation = %s",
                                    (sd, self.branch, self.corporation)
                                )
                                logger.info("Deleted invalid record from %s", table_name)
                            else:
                                # Existing record was updated — revert to unlocked so user can fix and resubmit
                                self.db_manager.execute_query(
                                    f"UPDATE {table_name} SET is_locked = 0 WHERE date = %s AND branch = %s AND corporation = %s",
                                    (sd, self.branch, self.corporation)
                                )
                                logger.info("Reverted %s record to unlocked state", table_name)
                        except Exception as revert_err:
                            logger.error("Could not revert record: %s", revert_err)
                        results.append((brand_full, "validation_error",
                                      "Calculation error detected. Report was NOT saved to the database.\n"
                                      "Please check your values and try again."))
                    else:
                        results.append((brand_full, "success", None))

                    if OFFLINE_SUPPORT and offline_manager:
                        offline_manager.cache_ending_balance(
                            self.user_email, self.branch, self.corporation,
                            brand_full, sd, ending
                        )
                    self._propagate_opening_balance_to_following_days(sd, brand_full, ending)
                elif rows is None:
                    results.append((brand_full, "error",
                                    str(last_err) if last_err else "Unknown error"))
                else:
                    results.append((brand_full, "no_rows", None))

            successes = [b for b, s, _ in results if s == "success"]
            errors    = [(b, e) for b, s, e in results if s == "error"]
            val_errors = [(b, e) for b, s, e in results if s == "validation_error"]
            skipped   = [b for b, s, _ in results if s == "skipped"]

            if errors or val_errors:
                self.loading_overlay.hide()
                error_msg = ""
                if errors:
                    error_msg += "Database Errors:\n\n" + "\n".join(f"{b}: {e}" for b, e in errors)
                if val_errors:
                    if error_msg:
                        error_msg += "\n\n"
                    error_msg += "⚠️ VALIDATION ERRORS (Amounts may be incorrect):\n\n"
                    error_msg += "\n".join(f"{b}: {e}" for b, e in val_errors)
                    error_msg += "\n\nPlease contact support immediately!"
                
                self._msg("Save Error - Validation Failed" if val_errors else "Database Error",
                          error_msg,
                          QMessageBox.Critical)
                return

            if successes:

                self.loading_overlay.set_status("Finalizing report...", "Complete")
                QApplication.processEvents()

                # Save cash float for Brand A
                if "Brand A" in successes:
                    self._save_cash_float(sd)

                # NOTE: Client only writes to daily_reports_brand_a and daily_reports.
                # Supplementary tables (service tables, palawan payable) are
                # now managed server-side from the canonical daily report tables.
                # This prevents duplicate/sync issues across multiple tables.

                self.loading_overlay.hide()

                parts = f"Posted: {', '.join(successes)}"
                if skipped:
                    parts += f"\nSkipped (already exists): {', '.join(skipped)}"
                self._msg_success(f"Report for {sd}\n\n{parts}")
                self._delete_draft(sd)

                # Immediately lock inputs and disable post button for successfully posted brands
                if "Brand A" in successes or "Brand B" in successes:
                    self._toggle_inputs(False)
                    self.post_button.setEnabled(False)

                # Brief delay to ensure database transaction is committed and visible
                time.sleep(0.5)

                self.on_date_changed()
                self.clear_all_fields()
            elif skipped and not successes:
                self.loading_overlay.hide()
                self._msg("Nothing to Post",
                          "All brands already have entries for this date.",
                          QMessageBox.Information)

        except Exception as e:
            self.loading_overlay.hide()
            if self._is_network_error(e):
                self._on_connection_lost()
            else:
                self._msg("Error", f"Failed to post: {e}", QMessageBox.Critical)

    def _handle_offline_post(self, selected_date):

        try:
            if not OFFLINE_SUPPORT or not offline_manager:
                self._msg("Error", "Offline support not available.", QMessageBox.Critical)
                return
            
            
            pal = self.palawan_tab.get_data()
            brand_data = {}
            # SKID/SKIR/CANCEL/INC will be injected after brand_data is built
            for brand_full, cf_tab, bb_input, cc_input, table_name in [
                ("Brand A", self.cash_flow_tab_a,
                 self.beginning_balance_input_a, self.cash_count_input_a,
                 "daily_reports_brand_a"),
                ("Brand B", self.cash_flow_tab_b,
                 self.beginning_balance_input_b, self.cash_count_input_b,
                 "daily_reports"),
            ]:
                try:
                    beginning = float(bb_input.text().strip().replace(',', '') or 0)
                    cash_count = float(cc_input.text().strip().replace(',', '') or 0)
                except ValueError:
                    beginning = 0
                    cash_count = 0
                
                cf = cf_tab.get_data()
                deb = sum(v for k, v in cf['debit'].items() if not k.endswith('_lotes'))
                cre = sum(v for k, v in cf['credit'].items() if not k.endswith('_lotes'))
                ending = beginning + deb - cre
                cash_result = cash_count - ending
                variance_status = (
                    "balanced" if abs(cash_result) < 0.01
                    else "over" if cash_result > 0
                    else "short"
                )
                
                # Merge: palawan tab provides supplementary fields; credit tab takes
                # precedence for any key that appears in both (non-zero credit wins).
                all_vals = {**cf['debit'], **pal}
                for k, v in cf['credit'].items():
                    if v != 0 or k not in all_vals:
                        all_vals[k] = v

                if hasattr(cf_tab, 'selected_bank_account') and cf_tab.selected_bank_account:
                    all_vals['fund_transfer_bank_account'] = cf_tab.selected_bank_account


                if hasattr(self, '_ft_ho_breakdowns'):
                    bd = self._ft_ho_breakdowns.get(brand_full)
                    if bd:
                        all_vals['ft_ho_breakdown'] = json.dumps(bd)
                
                
                if hasattr(cf_tab, 'branch_dest_inputs'):
                    branch_dest = cf_tab.branch_dest_inputs.get('Fund Transfer to BRANCH')
                    if branch_dest and branch_dest.text().strip():
                        all_vals['fund_transfer_to_branch_dest'] = branch_dest.text().strip()
                
                
                if hasattr(cf_tab, 'branch_dest_inputs'):
                    branch_src = cf_tab.branch_dest_inputs.get('Fund Transfer from BRANCH')
                    if branch_src and branch_src.text().strip():
                        all_vals['fund_transfer_from_branch_dest'] = branch_src.text().strip()
                
                
                if brand_full == "Brand A" and hasattr(self, '_pc_salary_breakdown') and self._pc_salary_breakdown:
                    all_vals['pc_salary_breakdown'] = json.dumps(self._pc_salary_breakdown)
                
               
                if hasattr(cf_tab, 'mc_currency_details'):
                    mc_in_details = cf_tab.mc_currency_details.get('MC In', [])
                    mc_out_details = cf_tab.mc_currency_details.get('MC Out', [])
                    if mc_in_details:
                        all_vals['mc_in_details'] = json.dumps(mc_in_details)
                    if mc_out_details:
                        all_vals['mc_out_details'] = json.dumps(mc_out_details)
                

                if hasattr(self, '_motor_car_breakdown') and self._motor_car_breakdown:
                    all_vals['empeno_motor_car_breakdown'] = json.dumps(self._motor_car_breakdown)
                
                brand_data[brand_full] = {
                    "table_name": table_name,
                    "beginning_balance": beginning,
                    "debit_total": deb,
                    "credit_total": cre,
                    "ending_balance": ending,
                    "cash_count": cash_count,
                    "cash_result": cash_result,
                    "variance_status": variance_status,
                    "all_values": all_vals
                }
            
            # Both brands use the same palawan data from payable_tbl_brand_a
            entry_data = {
                "date": selected_date,
                "username": self.user_email,
                "branch": self.branch,
                "corporation": self.corporation,
                "palawan_data": pal,
                "brand_data": brand_data
            }
            

            entry_id = offline_manager.save_pending_entry(
                self.user_email,
                self.branch,
                self.corporation,
                entry_data
            )
            
   
            for brand_full, data in brand_data.items():
                offline_manager.cache_ending_balance(
                    self.user_email, self.branch, self.corporation,
                    brand_full, selected_date, data.get('ending_balance', 0)
                )
            
            self._msg_success(
                f"📤 Entry Saved for Later Posting\n\n"
                f"Date: {selected_date}\n"
                f"Entry ID: {entry_id[:20]}...\n\n"
                f"This entry will be automatically posted when\n"
                f"you log in with internet connection.\n\n"
                f"You can continue entering more reports."
            )
            

            self._delete_draft(selected_date)
            self.clear_all_fields()
            
        except Exception as e:
            self._msg("Offline Save Error", f"Failed to save entry: {e}", QMessageBox.Critical)

    def _save_palawan_to_payable(self, selected_date, brand_full, palawan_data):
        """DEPRECATED: Client no longer writes to payable_tbl_brand_a.
        
        All palawan data is now stored in daily_reports_brand_a and daily_reports.
        Server-side processes will extract and manage payable table as needed.
        """
        logger.info("_save_palawan_to_payable() called but skipped - client writes only to daily report tables")
        pass

    def _save_palawan_to_payable_DISABLED_old(self, selected_date, brand_full, palawan_data):
        """DISABLED: Old version kept for reference."""
        try:
            payable_table = "payable_tbl_brand_a"

            def _to_int(val):
                try:
                    return int(float(val or 0))
                except (TypeError, ValueError):
                    return 0

            # palawan_data comes from palawan_tab.get_data() → so_*/po_*/int_* keys
            sendout_lotes      = _to_int(palawan_data.get('so_lotes', 0))
            sendout_capital    = palawan_data.get('so_principal', 0) or 0
            sendout_sc         = palawan_data.get('so_sc', 0) or 0
            sendout_commission = palawan_data.get('so_commission', 0) or 0
            sendout_total      = palawan_data.get('so_total', 0) or 0

            payout_lotes      = _to_int(palawan_data.get('po_lotes', 0))
            payout_capital    = palawan_data.get('po_principal', 0) or 0
            payout_sc         = palawan_data.get('po_sc', 0) or 0
            payout_commission = palawan_data.get('po_commission', 0) or 0
            payout_total      = palawan_data.get('po_total', 0) or 0

            int_lotes      = _to_int(palawan_data.get('int_lotes', 0))
            int_capital    = palawan_data.get('int_principal', 0) or 0
            int_sc         = palawan_data.get('int_sc', 0) or 0
            int_commission = palawan_data.get('int_commission', 0) or 0
            int_total      = palawan_data.get('int_total', 0) or 0

            skid         = palawan_data.get('palawan_suki_discounts', 0) or 0
            skir         = palawan_data.get('palawan_suki_rebates', 0) or 0
            cancellation = palawan_data.get('palawan_cancel', 0) or 0
            inc          = palawan_data.get('palawan_pay_out_incentives', 0) or 0

            query = f"""
                INSERT INTO {payable_table} (
                    corporation, branch, date,
                    sendout_lotes, sendout_capital, sendout_sc, sendout_commission, sendout_total,
                    payout_lotes, payout_capital, payout_sc, payout_commission, payout_total,
                    international_lotes, international_capital, international_sc,
                    international_commission, international_total,
                    skid, skir, cancellation, inc
                ) VALUES (
                    %s, %s, %s,
                    %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s,
                    %s, %s, %s, %s
                )
                ON DUPLICATE KEY UPDATE
                    sendout_lotes = VALUES(sendout_lotes),
                    sendout_capital = VALUES(sendout_capital),
                    sendout_sc = VALUES(sendout_sc),
                    sendout_commission = VALUES(sendout_commission),
                    sendout_total = VALUES(sendout_total),
                    payout_lotes = VALUES(payout_lotes),
                    payout_capital = VALUES(payout_capital),
                    payout_sc = VALUES(payout_sc),
                    payout_commission = VALUES(payout_commission),
                    payout_total = VALUES(payout_total),
                    international_lotes = VALUES(international_lotes),
                    international_capital = VALUES(international_capital),
                    international_sc = VALUES(international_sc),
                    international_commission = VALUES(international_commission),
                    international_total = VALUES(international_total),
                    skid = VALUES(skid),
                    skir = VALUES(skir),
                    cancellation = VALUES(cancellation),
                    inc = VALUES(inc),
                    updated_at = CURRENT_TIMESTAMP
            """
            params = (
                self.corporation, self.branch, selected_date,
                sendout_lotes, sendout_capital, sendout_sc, sendout_commission, sendout_total,
                payout_lotes, payout_capital, payout_sc, payout_commission, payout_total,
                int_lotes, int_capital, int_sc, int_commission, int_total,
                skid, skir, cancellation, inc
            )

            self.db_manager.execute_query(query, params)
            logger.info("Palawan details saved to %s for %s", payable_table, brand_full)

        except Exception as e:
            logger.error("Error saving Palawan to payable (%s): %s", brand_full, e)

    def _post_to_service_tables(self, selected_date, all_vals: dict):
        """DEPRECATED: Client no longer writes to supplementary service tables.
        
        All data is now written only to daily_reports_brand_a and daily_reports.
        Server-side processes will extract and denormalize data to service tables as needed.
        This eliminates duplicate/sync problems across multiple tables.
        """
        logger.info("_post_to_service_tables() called but skipped - client writes only to daily report tables")
        pass

    def _save_cash_float(self, selected_date):

        try:
            cash_float_val = 0.0
            if hasattr(self, 'cash_float_input_a') and self.cash_float_input_a:
                txt = self.cash_float_input_a.text().strip()
                if txt:
                    cash_float_val = float(txt)
            
            query = """
                INSERT INTO cash_float_tbl (date, branch, corporation, username, cash_float)
                VALUES (%s, %s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE
                    cash_float = VALUES(cash_float),
                    username = VALUES(username),
                    updated_at = CURRENT_TIMESTAMP
            """
            params = (selected_date, self.branch, self.corporation, self.user_email, cash_float_val)
            self.db_manager.execute_query(query, params)
            logger.info("Cash Float saved: %.2f", cash_float_val)
        except Exception as e:
            logger.error("Error saving Cash Float: %s", e)


    def validate_all_requirements(self):
        sd = self.date_picker.date().toString("yyyy-MM-dd")
        brand_specs = [
            ("A", "Brand A",
             self.beginning_balance_input_a, self.cash_count_input_a,
             self.beginning_balance_auto_filled_a,
             self.previous_day_balance_a, self.previous_day_date_a,
             self.ending_balance_display_a),
            ("B", "Brand B",
             self.beginning_balance_input_b, self.cash_count_input_b,
             self.beginning_balance_auto_filled_b,
             self.previous_day_balance_b, self.previous_day_date_b,
             self.ending_balance_display_b),
        ]

        for ltr, brand_full, bb_input, cc_input, bb_filled, prev_bal, prev_date, eb_disp in brand_specs:
            if self.check_existing_entry(sd, brand_full) == "locked":
                continue

            if not bb_filled:
                if prev_bal is not None:
                    self._msg("Opening Balance Required",
                              f"{brand_full}: Click 'Load Previous Day' to set the opening balance.\n"
                              f"Expected: {prev_bal:,.2f}  ({prev_date})",
                              QMessageBox.Critical)
                    return False
                elif not bb_input.text().strip():
                    self._msg("Opening Balance Required",
                              f"{brand_full}: First entry — please type the opening balance.",
                              QMessageBox.Critical)
                    bb_input.setFocus()
                    return False

            if not bb_input.text().strip():
                self._msg("Validation Error",
                          f"{brand_full}: Opening balance is empty.",
                          QMessageBox.Warning)
                return False

            if not cc_input.text().strip():
                self._msg("Validation Error",
                          f"{brand_full}: Enter the actual cash count.",
                          QMessageBox.Warning)
                cc_input.setFocus()
                return False

            if prev_bal is not None:
                try:
                    cur_beg = float(bb_input.text().strip().replace(',', '') or 0)
                    if abs(cur_beg - prev_bal) > 0.01:
                        self._msg("Balance Mismatch",
                                  f"{brand_full}: Opening balance does not match previous day.\n\n"
                                  f"Expected : {prev_bal:,.2f}  ({prev_date})\n"
                                  f"Current  : {cur_beg:,.2f}\n\n"
                                  f"Use 'Load Previous Day' to correct this.",
                                  QMessageBox.Critical)
                        return False
                except ValueError:
                    self._msg("Validation Error",
                              f"{brand_full}: Invalid opening balance.",
                              QMessageBox.Warning)
                    return False

            try:
                cc_val = float(cc_input.text().strip().replace(',', '') or 0)
                eb_val = float((eb_disp.text() or "0").replace(",", ""))
                diff   = cc_val - eb_val
                if abs(diff) >= 0.01:
                    vtype = "OVER" if diff > 0 else "SHORT"
                    reply = QMessageBox.question(
                        self,
                        f"Variance Warning ({brand_full})",
                        f"{brand_full} has a cash variance ({vtype} by {abs(diff):,.2f}).\n\n"
                        f"Ending Balance : {eb_val:,.2f}\n"
                        f"Cash Count     : {cc_val:,.2f}\n"
                        f"Variance       : {diff:,.2f}\n\n"
                        f"This entry will be tagged for admin review.\n\n"
                        f"Do you want to continue posting?",
                        QMessageBox.Yes | QMessageBox.No,
                        QMessageBox.No
                    )
                    if reply == QMessageBox.No:
                        return False
            except ValueError:
                self._msg("Validation Error",
                          f"{brand_full}: Invalid cash count.",
                          QMessageBox.Warning)
                return False

        return True

    def clear_table(self):
   
        r = QMessageBox.question(
            self, "Clear Table",
            "Are you sure you want to clear all fields?\n\nThis will not delete any submitted data.",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        if r == QMessageBox.Yes:
            for bb in (self.beginning_balance_input_a, self.beginning_balance_input_b):
                bb.clear()
            for cc in (self.cash_count_input_a, self.cash_count_input_b):
                cc.clear()
            for tab, meth in [
                (self.cash_flow_tab_a, 'clear_fields'),
                (self.cash_flow_tab_b, 'clear_fields'),
                (self.palawan_tab,     'clear_fields'),
            ]:
                if hasattr(tab, meth):
                    getattr(tab, meth)()

            if hasattr(self, '_jew_computed'):
                for k in self._jew_computed:
                    self._jew_computed[k] = 0.0
            self.recalculate_all()

    def clear_all_fields(self):
        r = QMessageBox.question(
            self, "Advance Date", "Report posted! Move to the next day?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes
        )
        # ALWAYS clear dialog caches, even if user doesn't move to next day
        self._clear_all_dialog_caches()

        if r == QMessageBox.Yes:
            for tab, meth in [
                (self.cash_flow_tab_a, 'clear_fields'),
                (self.cash_flow_tab_b, 'clear_fields'),
                (self.palawan_tab,     'clear_fields'),
            ]:
                if hasattr(tab, meth):
                    getattr(tab, meth)()
            self.date_picker.setDate(self.date_picker.date().addDays(1))
        else:
            # User clicked "No" — reload current day's report to show it's locked
            sd = self.date_picker.date().toString("yyyy-MM-dd")
            self.on_date_changed()

    def clear_all_fields_silent(self):
        # Clear all cached dialogs and breakdowns FIRST
        self._clear_all_dialog_caches()
        
        for bb in (self.beginning_balance_input_a, self.beginning_balance_input_b):
            bb.clear()
            bb.setStyleSheet(f"""
                QLineEdit {{
                    background-color: {_SLATE_100};
                    border: 1.5px solid {_SLATE_200};
                    border-radius: 6px;
                    padding: 7px 12px;
                    font-size: 14px;
                    font-weight: 700;
                    color: {_TEXT_PRI};
                }}
            """)
        for cc in (self.cash_count_input_a, self.cash_count_input_b):
            cc.clear()
        self.beginning_balance_auto_filled_a = False
        self.beginning_balance_auto_filled_b = False
        self.previous_day_balance_a = self.previous_day_date_a = None
        self.previous_day_balance_b = self.previous_day_date_b = None
        for btn in (self.auto_fill_button_a, self.auto_fill_button_b):
            btn.setText("↓  Load Previous Day")
            btn.setEnabled(True)
            btn.setStyleSheet(f"""
                QPushButton {{
                    background-color: transparent;
                    color: {_PRIMARY};
                    border: 1.5px solid {_INDIGO_100};
                    border-radius: 5px;
                    font-size: 12px;
                    font-weight: 700;
                    padding: 4px 14px;
                    min-width: 0;
                }}
                QPushButton:hover {{
                    background-color: {_INDIGO_50};
                    border-color: {_PRIMARY};
                }}
            """)
        for lbl in (self.balance_status_label_a, self.balance_status_label_b):
            lbl.setText("")
        for tab, meth in [
            (self.cash_flow_tab_a, 'clear_fields'),
            (self.cash_flow_tab_b, 'clear_fields'),
            (self.palawan_tab,     'clear_fields'),
        ]:
            if hasattr(tab, meth):
                getattr(tab, meth)()
        self.on_date_changed()


    def _connect_shared_fields(self):
        try:
            import json as _json, re as _re, os as _os, sys as _sys
            if getattr(_sys, 'frozen', False):
                _base = _os.path.dirname(_sys.executable)
            else:
                _base = _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__)))
            _cfg_path = _os.path.join(_base, "field_config.json")
            with open(_cfg_path, "r", encoding="utf-8") as _f:
                cfg = _json.load(_f)

            def _col(entry):
                return entry[2] if len(entry) >= 3 else \
                    _re.sub(r"[^a-z0-9]+", "_", entry[0].lower()).strip("_")

            a_col_widget, b_col_widget = {}, {}
            for section in ("debit", "credit"):
                for brand, col_w in [("Brand A", a_col_widget), ("Brand B", b_col_widget)]:
                    tab = self.cash_flow_tab_a if brand == "Brand A" else self.cash_flow_tab_b
                    tab_d   = tab.debit_inputs   if section == "debit" else tab.credit_inputs
                    lotes_d = (tab.debit_lotes_inputs if section == "debit"
                               else tab.credit_lotes_inputs)
                    for entry in cfg.get(brand, {}).get(section, []):
                        label = entry[0]
                        c     = _col(entry)
                        if label in tab_d:
                            col_w[c] = (tab_d[label], lotes_d.get(label))

            shared = set(a_col_widget) & set(b_col_widget)
            

            excluded_fields = {
                "fund_transfer_to_head_office",
                "fund_transfer_from_branch",
                "fund_transfer_to_branch",
                "palawan_pay_cash_out_sc",
                "palawan_send_out",
                "palawan_send_out_lotes",
                "palawan_sc",
                "palawan_suki_card",
                "palawan_pay_cash_in_sc",
                "palawan_pay_bills_sc",
                "palawan_change_receiver",
                "palawan_pay_out",
                "palawan_pay_out_lotes",
                "palawan_pay_out_incentives",
                "palawan_pay_cash_out",
                "palawan_cancel",
                "palawan_suki_discounts",
                "palawan_suki_rebates",
                "palawan_load_sc",  
                "palawan_load", 
                "handling_fee",
                "other_penalty",
                "cash_overage",
                "mc_in",
                "mc_out",

                "pc_transpo",
                "pc_salary",
                "pc_inc_motor",
                "pc_inc_emp",
                "pc_inc_suki_card",
                "pc_inc_insurance",
                "pc_inc_mc",
                "pc_supplies_xerox_maintenance",
                "pc_electric",
                "pc_water",
                "pc_internet",
                "pc_rental",
                "pc_permits_bir_payments",
                "pc_lbc_jrs_jnt",
                "empeno_motor_car",
                "motor_ai",
            }
            shared = shared - excluded_fields
            
            logger.debug("Shared fields after exclusions: %s", shared)
            
            self._shared_carry_map = []

            for col in shared:
                a_w, a_l = a_col_widget[col]
                b_w, b_l = b_col_widget[col]

                def _amt_carry(bw):
                    def _carry(text):
                        if not bw.hasFocus():
                            bw.blockSignals(True); bw.setText(text); bw.blockSignals(False)
                            self.recalculate_all()
                    return _carry

                def _lot_carry(bw):
                    def _carry(text):
                        if bw and not bw.hasFocus():
                            bw.blockSignals(True); bw.setText(text); bw.blockSignals(False)
                    return _carry

                fn = _amt_carry(b_w)
                a_w.textChanged.connect(fn)
                self._shared_carry_map.append(fn)

                if a_l and b_l:
                    fl = _lot_carry(b_l)
                    a_l.textChanged.connect(fl)
                    self._shared_carry_map.append(fl)

        except Exception as e:
            logger.warning("_connect_shared_fields error (non-fatal): %s", e)

    def _connect_palawan_adjustments_to_brand_b(self):

        try:
            brand_b_debits = getattr(self.cash_flow_tab_b, 'debit_inputs', {})
            brand_b_credits = getattr(self.cash_flow_tab_b, 'credit_inputs', {})

            def make_carry_fn(target_field):
                def carry(text):
                    target_field.blockSignals(True)
                    target_field.setText(text)
                    target_field.blockSignals(False)
                    self.recalculate_all()
                return carry

            def setup_readonly_field(field):

                field.setReadOnly(True)
                field.setStyleSheet(
                    field.styleSheet() + 
                    "background-color: #FEF3C7; color: #92400E;"
                )
                field.setPlaceholderText("Auto from Palawan Details")

            connected_count = 0

            sendout_inputs = getattr(self.palawan_tab, 'sendout_inputs', {})
            
            if "Principal" in sendout_inputs and "Palawan Send Out" in brand_b_debits:
                palawan_field = sendout_inputs["Principal"]
                cashflow_field = brand_b_debits["Palawan Send Out"]
                setup_readonly_field(cashflow_field)
                palawan_field.textChanged.connect(make_carry_fn(cashflow_field))
                connected_count += 1

            if "SC" in sendout_inputs and "Commission" in sendout_inputs and "Palawan S.C" in brand_b_debits:
                sc_field = sendout_inputs["SC"]
                commission_field = sendout_inputs["Commission"]
                cashflow_field = brand_b_debits["Palawan S.C"]
                setup_readonly_field(cashflow_field)
                
                # Create function to sum SC + Commission
                def make_sc_commission_sum_fn(sc_fld, comm_fld, target_fld):
                    def update_sc_commission():
                        try:
                            sc_val = float(sc_fld.text().strip() or 0)
                        except ValueError:
                            sc_val = 0.0
                        try:
                            comm_val = float(comm_fld.text().strip() or 0)
                        except ValueError:
                            comm_val = 0.0
                        
                        total = sc_val + comm_val
                        target_fld.blockSignals(True)
                        target_fld.setText(f"{total:.2f}")
                        target_fld.blockSignals(False)
                        self.recalculate_all()
                    return update_sc_commission
                
                # Connect both SC and Commission fields to the sum function
                update_fn = make_sc_commission_sum_fn(sc_field, commission_field, cashflow_field)
                sc_field.textChanged.connect(update_fn)
                commission_field.textChanged.connect(update_fn)
                connected_count += 1

            payout_inputs = getattr(self.palawan_tab, 'payout_inputs', {})
            
            if "Principal" in payout_inputs and "Palawan Pay Out" in brand_b_credits:
                palawan_field = payout_inputs["Principal"]
                cashflow_field = brand_b_credits["Palawan Pay Out"]
                setup_readonly_field(cashflow_field)
                palawan_field.textChanged.connect(make_carry_fn(cashflow_field))
                connected_count += 1

            adjustment_to_cashflow_map = {
                "Palawan Pay Out Incentives": "Palawan Pay Out (incentives)",
                "Palawan Suki Discounts": "Palawan Suki Discounts",
                "Palawan Suki Rebates": "Palawan Suki Rebates",
                "Palawan Cancel": "Palawan Cancel",
            }

            palawan_adjustments = getattr(self.palawan_tab, 'adjustments_inputs', {})

            for adj_label, cf_label in adjustment_to_cashflow_map.items():
                if adj_label in palawan_adjustments and cf_label in brand_b_credits:
                    palawan_field = palawan_adjustments[adj_label]
                    cashflow_field = brand_b_credits[cf_label]
                    setup_readonly_field(cashflow_field)
                    palawan_field.textChanged.connect(make_carry_fn(cashflow_field))
                    connected_count += 1

            logger.debug("Connected %d Palawan fields to Brand B Cash Flow", connected_count)

        except Exception as e:
            logger.warning("_connect_palawan_adjustments_to_brand_b error (non-fatal): %s", e)

    def _connect_brand_a_auto_calculations(self):

        try:
            from Client.cash_flow_tab import _load_field_config

            d_inp  = self.cash_flow_tab_a.debit_inputs
            c_inp  = self.cash_flow_tab_a.credit_inputs
            c_lots = self.cash_flow_tab_a.credit_lotes_inputs

            cfg = _load_field_config(getattr(self, 'db_manager', None))
            brand_cfg = cfg.get("Brand A", {}) if isinstance(cfg, dict) else {}

            def _sanitize(name):
                return re.sub(r"[^a-z0-9]+", "_", (name or "").lower()).strip("_")

            section_col_to_labels = {"debit": {}, "credit": {}}
            for section in ("debit", "credit"):
                for entry in brand_cfg.get(section, []):
                    if not entry:
                        continue
                    label = entry[0]
                    col = entry[2] if len(entry) >= 3 else _sanitize(label)
                    section_col_to_labels[section].setdefault(col, []).append(label)
            
            # Flag to prevent recalc during signal connection
            recalc_enabled = False

            def _v(widget_dict, key):
                """Safely parse float from a widget dict entry."""
                w = widget_dict.get(key)
                if w is None:
                    return 0.0
                try:
                    return float(w.text().replace(',', '').strip() or 0)
                except ValueError:
                    return 0.0

            def _set_readonly(field, tooltip="Auto-calculated"):
                field.setReadOnly(True)
                field.setStyleSheet(
                    field.styleSheet() +
                    "background-color: #EFF6FF; color: #1E40AF;"
                )
                field.setToolTip(tooltip)
                field.setPlaceholderText("Auto-calculated")

            def _set(field, value):
                field.blockSignals(True)
                field.setText(f"{value:.2f}")
                field.blockSignals(False)
                field.update()
                field.repaint()

            def _resolve_label(widget_dict, section, column, fallback_labels):
                candidates = []
                candidates.extend(section_col_to_labels.get(section, {}).get(column, []))
                candidates.extend(fallback_labels)

                seen = set()
                ordered = []
                for c in candidates:
                    if c and c not in seen:
                        ordered.append(c)
                        seen.add(c)

                for label in ordered:
                    if label in widget_dict:
                        return label

                # Last resort: normalized label match for punctuation/case differences.
                normalized_widget_labels = {
                    re.sub(r"[^a-z0-9]+", "", lbl.lower()): lbl
                    for lbl in widget_dict.keys()
                }
                for label in ordered:
                    normalized = re.sub(r"[^a-z0-9]+", "", label.lower())
                    resolved = normalized_widget_labels.get(normalized)
                    if resolved:
                        return resolved
                return None

            def _get_value(section, column, widget_dict, fallback_labels):
                label = _resolve_label(widget_dict, section, column, fallback_labels)
                return _v(widget_dict, label) if label else 0.0

            def _set_value(section, column, widget_dict, fallback_labels, value):
                label = _resolve_label(widget_dict, section, column, fallback_labels)
                if label:
                    _set(widget_dict[label], value)

            targets = [
                ("service_charge", ["S.C"], "(Lotes JEW NEW + Lotes JEW RENEW) × 5"),
                ("osf_storage", ["O.s.f Sto.", "OSF STO"], "(Empeno STO NEW + Empeno STO RENEW) × 0.75%"),
                ("storage_ai", ["Sto. A.I", "STO AI"], "(Empeno STO NEW + Empeno STO RENEW) × 20%"),
                ("silver_ai", ["Silver A.I", "SILVER AI"], "Empeno Silver × 20%"),
                ("osf_silver", ["O.s.f Silver", "OSF SILVER"], "Empeno Silver × 0.75%"),
                ("osf_motor", ["O.s.f Motor", "OSF MOTOR"], "Empeno Motor/Car × 0.75%"),
            ]
            for col, fallback_labels, tip in targets:
                label = _resolve_label(d_inp, "debit", col, fallback_labels)
                if label:
                    _set_readonly(d_inp[label], tip)

            def recalc(*_):
                nonlocal recalc_enabled
                if not recalc_enabled:
                    return

                lotes_jew_new = _get_value("credit", "empeno_jew_new", c_lots, ["Empeno JEW. (NEW)"])
                lotes_jew_renew = _get_value("credit", "empeno_jew_renew", c_lots, ["Empeno JEW (RENEW)"])
                sc_value = (lotes_jew_new + lotes_jew_renew) * 5
                _set_value("debit", "service_charge", d_inp, ["S.C"], sc_value)

                sto_new = _get_value("credit", "empeno_sto_new", c_inp, ["Empeno STO. (NEW)"])
                sto_renew = _get_value("credit", "fund_empeno_sto_renew", c_inp, ["Fund Empeno STO. (RENEW)"])
                sto_base  = sto_new + sto_renew
                _set_value("debit", "osf_storage", d_inp, ["O.s.f Sto.", "OSF STO"], sto_base * 0.0075)
                _set_value("debit", "storage_ai", d_inp, ["Sto. A.I", "STO AI"], sto_base * 0.20)

                silver = _get_value("credit", "empeno_silver", c_inp, ["Empeno silver", "Empeno Silver"])
                _set_value("debit", "silver_ai", d_inp, ["Silver A.I", "SILVER AI"], silver * 0.20)
                _set_value("debit", "osf_silver", d_inp, ["O.s.f Silver", "OSF SILVER"], silver * 0.0075)

                motor = _get_value("credit", "empeno_motor_car", c_inp, ["Empeno Motor/Car"])
                _set_value("debit", "motor_ai", d_inp, ["Motor A.I", "MOTOR AI"], motor * 0.10)
                _set_value("debit", "osf_motor", d_inp, ["O.s.f Motor", "OSF MOTOR"], motor * 0.0075)

                self.recalculate_all()

            source_specs = [
                ("credit", "empeno_jew_new", c_lots, ["Empeno JEW. (NEW)"]),
                ("credit", "empeno_jew_renew", c_lots, ["Empeno JEW (RENEW)"]),
                ("credit", "empeno_sto_new", c_inp, ["Empeno STO. (NEW)"]),
                ("credit", "fund_empeno_sto_renew", c_inp, ["Fund Empeno STO. (RENEW)"]),
                ("credit", "empeno_silver", c_inp, ["Empeno silver", "Empeno Silver"]),
                ("credit", "empeno_motor_car", c_inp, ["Empeno Motor/Car"]),
            ]
            for section, col, widget_dict, fallback_labels in source_specs:
                resolved_label = _resolve_label(widget_dict, section, col, fallback_labels)
                w = widget_dict.get(resolved_label) if resolved_label else None
                if w:
                    w.textChanged.connect(recalc)
                else:
                    logger.debug("Auto-calc source field not found: section=%s col=%s", section, col)
            
            # Enable recalc now that all signals are connected
            recalc_enabled = True
            # Trigger one initial recalc to populate fields
            recalc()

            logger.debug("Brand A auto-calculations connected.")

        except Exception as e:
            logger.warning("_connect_brand_a_auto_calculations error (non-fatal): %s", e)

    def _setup_empeno_jew_buttons(self):

        try:
            c_inp_a   = self.cash_flow_tab_a.credit_inputs
            d_inp_a   = self.cash_flow_tab_a.debit_inputs
            c_lotes_a = self.cash_flow_tab_a.credit_lotes_inputs
            
            c_inp_b   = self.cash_flow_tab_b.credit_inputs
            d_inp_b   = self.cash_flow_tab_b.debit_inputs
            c_lotes_b = self.cash_flow_tab_b.credit_lotes_inputs

            self._jew_dialogs  = {}
            self._jew_computed = {"Empeno JEW. (NEW)": 0.0, "Empeno JEW (RENEW)": 0.0}
            self._jew_computed_b = {"Empeno JEW. (NEW)": 0.0, "Empeno JEW (RENEW)": 0.0}

            def _set_auto_readonly(field, tip):
                """Set field as read-only (for Brand A)"""
                field.setReadOnly(True)
                field.setStyleSheet(
                    field.styleSheet() +
                    "background-color: #EFF6FF; color: #1D4ED8;"
                )
                field.setToolTip(tip)
                field.setPlaceholderText("Click + to enter detail")

            def _set_auto_editable(field, tip):
                """Set field as editable and styled for carry-over (for Brand B)"""
                field.setReadOnly(False)
                field.setStyleSheet(
                    f"""
                    QLineEdit {{
                        background-color: #FFFBEB;
                        border: 1.5px solid #FCD34D;
                        border-radius: 6px;
                        padding: 7px 12px;
                        font-size: 14px;
                        font-weight: 600;
                        color: #78350F;
                    }}
                    QLineEdit:focus {{
                        border: 2px solid #F59E0B;
                        background-color: #FEF3C7;
                    }}
                    """
                )
                field.setToolTip(tip)
                field.setPlaceholderText("Carries from Brand A")

            def _update_jew_ai_a():
                """Update Jew. A.I for Brand A"""
                total = sum(self._jew_computed.values())
                jew_ai = d_inp_a.get("Jew. A.I")
                if jew_ai:
                    jew_ai.blockSignals(True)
                    jew_ai.setText(f"{total:.2f}")
                    jew_ai.blockSignals(False)
                self.recalculate_all()

            def _update_jew_ai_b():
                """Update Jew. A.I for Brand B"""
                total = sum(self._jew_computed_b.values())
                jew_ai_b = d_inp_b.get("Jew. A.I")
                if jew_ai_b:
                    jew_ai_b.blockSignals(True)
                    jew_ai_b.setText(f"{total:.2f}")
                    jew_ai_b.blockSignals(False)
                self.recalculate_all()

            def _calculate_sc_b():
                """Calculate S.C. for Brand B: (Lotes JEW NEW + Lotes JEW RENEW) × 5"""
                try:
                    lotes_new = 0.0
                    lotes_renew = 0.0
                    
                    lotes_new_field = c_lotes_b.get("Empeno JEW. (NEW)")
                    if lotes_new_field:
                        try:
                            lotes_new = float(lotes_new_field.text().strip() or 0)
                        except ValueError:
                            lotes_new = 0.0
                    
                    lotes_renew_field = c_lotes_b.get("Empeno JEW (RENEW)")
                    if lotes_renew_field:
                        try:
                            lotes_renew = float(lotes_renew_field.text().strip() or 0)
                        except ValueError:
                            lotes_renew = 0.0
                    
                    sc_value = (lotes_new + lotes_renew) * 5
                    
                    sc_b = d_inp_b.get("S.C")
                    if sc_b:
                        sc_b.blockSignals(True)
                        sc_b.setText(f"{sc_value:.2f}")
                        sc_b.blockSignals(False)
                        self.recalculate_all()
                except Exception as e:
                    logger.error("Error calculating S.C. for Brand B: %s", e)

            def _make_open_dialog_a(field_label, target_field_a, target_field_b):
                """Create dialog function for Brand A"""
                def _open():
                    if field_label not in self._jew_dialogs:
                        self._jew_dialogs[field_label] = EmpenaDetailDialog(
                            field_label, parent=self
                        )
                    dlg = self._jew_dialogs[field_label]

                    if dlg.exec_() == QDialog.Accepted:
                        val        = f"{dlg.get_total_amount():.2f}"
                        row_count  = dlg.get_row_count()

                        target_field_a.blockSignals(True)
                        target_field_a.setText(val)
                        target_field_a.blockSignals(False)

                        lotes_a = c_lotes_a.get(field_label)
                        if lotes_a:
                            lotes_a.blockSignals(True)
                            lotes_a.setText(str(row_count))
                            lotes_a.blockSignals(False)
                            lotes_a.textChanged.emit(str(row_count))

                        self._jew_computed[field_label] = dlg.get_computed_total()
                        _update_jew_ai_a()
                        
                        # Carry over to Brand B after dialog acceptance
                        target_field_b.blockSignals(True)
                        target_field_b.setText(val)
                        target_field_b.blockSignals(False)
                        
                        lotes_b = c_lotes_b.get(field_label)
                        if lotes_b:
                            lotes_b.blockSignals(True)
                            lotes_b.setText(str(row_count))
                            lotes_b.blockSignals(False)
                            lotes_b.textChanged.emit(str(row_count))
                        
                        # Update S.C. calculation in Brand B
                        _calculate_sc_b()
                return _open

            def _make_open_dialog_b(field_label, target_field_b, target_field_a):
                """Create dialog function for Brand B"""
                def _open():
                    if field_label not in self._jew_dialogs:
                        self._jew_dialogs[field_label] = EmpenaDetailDialog(
                            field_label, parent=self
                        )
                    dlg = self._jew_dialogs[field_label]

                    if dlg.exec_() == QDialog.Accepted:
                        val        = f"{dlg.get_total_amount():.2f}"
                        row_count  = dlg.get_row_count()

                        target_field_b.blockSignals(True)
                        target_field_b.setText(val)
                        target_field_b.blockSignals(False)

                        lotes_b = c_lotes_b.get(field_label)
                        if lotes_b:
                            lotes_b.blockSignals(True)
                            lotes_b.setText(str(row_count))
                            lotes_b.blockSignals(False)
                            lotes_b.textChanged.emit(str(row_count))

                        self._jew_computed_b[field_label] = dlg.get_computed_total()
                        _update_jew_ai_b()
                        
                        # Update S.C. calculation in Brand B
                        _calculate_sc_b()
                return _open

            # Setup both Brand A and Brand B
            for label in ("Empeno JEW. (NEW)", "Empeno JEW (RENEW)"):
                # Brand A setup
                field_a = c_inp_a.get(label)
                if field_a is not None:
                    field_b = c_inp_b.get(label)
                    _set_auto_readonly(field_a, f"Click + to enter breakdown for {label}")

                    container_a = field_a.parentWidget()
                    layout_a    = container_a.layout() if container_a else None
                    if layout_a is not None:
                        plus_btn_a = QPushButton("+")
                        plus_btn_a.setFixedSize(28, 28)
                        plus_btn_a.setStyleSheet("""
                            QPushButton {
                                background-color: #2563EB; color: white;
                                border: none; border-radius: 5px;
                                font-size: 14px; font-weight: 900;
                            }
                            QPushButton:hover { background-color: #1D4ED8; }
                        """)
                        plus_btn_a.setToolTip(f"Open detail breakdown for {label}")
                        plus_btn_a.clicked.connect(_make_open_dialog_a(label, field_a, field_b))
                        layout_a.insertWidget(0, plus_btn_a)

                # Brand B setup
                field_b = c_inp_b.get(label)
                if field_b is not None:
                    _set_auto_editable(field_b, f"Click + to enter breakdown for {label} • or carries from Brand A")

                    container_b = field_b.parentWidget()
                    layout_b    = container_b.layout() if container_b else None
                    if layout_b is not None:
                        plus_btn_b = QPushButton("+")
                        plus_btn_b.setFixedSize(28, 28)
                        plus_btn_b.setStyleSheet("""
                            QPushButton {
                                background-color: #F59E0B; color: white;
                                border: none; border-radius: 5px;
                                font-size: 14px; font-weight: 900;
                            }
                            QPushButton:hover { background-color: #D97706; }
                        """)
                        plus_btn_b.setToolTip(f"Open detail breakdown for {label} (Brand B)")
                        field_a_ref = c_inp_a.get(label)
                        plus_btn_b.clicked.connect(_make_open_dialog_b(label, field_b, field_a_ref))
                        layout_b.insertWidget(0, plus_btn_b)


            # Setup Jew. A.I for Brand A (editable)
            jew_ai_a = d_inp_a.get("Jew. A.I")
            if jew_ai_a:
                jew_ai_a.setReadOnly(False)
                jew_ai_a.setStyleSheet(
                    f"""
                    QLineEdit {{
                        background-color: #FFFBEB;
                        border: 1.5px solid #FCD34D;
                        border-radius: 6px;
                        padding: 7px 12px;
                        font-size: 14px;
                        font-weight: 600;
                        color: #78350F;
                    }}
                    QLineEdit:focus {{
                        border: 2px solid #F59E0B;
                        background-color: #FEF3C7;
                    }}
                    """
                )
                jew_ai_a.setToolTip("Computed Total (JEW NEW) + Computed Total (JEW RENEW) • Editable")
                jew_ai_a.setPlaceholderText("Auto-calculated from dialogs")

            # Setup Jew. A.I for Brand B (editable, calculated from computed totals)
            jew_ai_b = d_inp_b.get("Jew. A.I")
            if jew_ai_b:
                jew_ai_b.setReadOnly(False)
                jew_ai_b.setStyleSheet(
                    f"""
                    QLineEdit {{
                        background-color: #FFFBEB;
                        border: 1.5px solid #FCD34D;
                        border-radius: 6px;
                        padding: 7px 12px;
                        font-size: 14px;
                        font-weight: 600;
                        color: #78350F;
                    }}
                    QLineEdit:focus {{
                        border: 2px solid #F59E0B;
                        background-color: #FEF3C7;
                    }}
                    """
                )
                jew_ai_b.setToolTip("Computed Total (JEW NEW) + Computed Total (JEW RENEW) • Editable")
                jew_ai_b.setPlaceholderText("Auto-calculated from dialogs")


            # Setup S.C. for both Brand A and Brand B (editable)
            sc_a = d_inp_a.get("S.C")
            sc_b = d_inp_b.get("S.C")
            if sc_a:
                sc_a.setReadOnly(False)
                sc_a.setStyleSheet(
                    f"""
                    QLineEdit {{
                        background-color: #FFFBEB;
                        border: 1.5px solid #FCD34D;
                        border-radius: 6px;
                        padding: 7px 12px;
                        font-size: 14px;
                        font-weight: 600;
                        color: #78350F;
                    }}
                    QLineEdit:focus {{
                        border: 2px solid #F59E0B;
                        background-color: #FEF3C7;
                    }}
                    """
                )
                sc_a.setToolTip("Calculated as (Lotes JEW NEW + Lotes JEW RENEW) × 5 • Editable")
                sc_a.setPlaceholderText("Auto or manual")

            if sc_b:
                sc_b.setReadOnly(False)
                sc_b.setStyleSheet(
                    f"""
                    QLineEdit {{
                        background-color: #FFFBEB;
                        border: 1.5px solid #FCD34D;
                        border-radius: 6px;
                        padding: 7px 12px;
                        font-size: 14px;
                        font-weight: 600;
                        color: #78350F;
                    }}
                    QLineEdit:focus {{
                        border: 2px solid #F59E0B;
                        background-color: #FEF3C7;
                    }}
                    """
                )
                sc_b.setToolTip("Calculated as (Lotes JEW NEW + Lotes JEW RENEW) × 5 • Editable")
                sc_b.setPlaceholderText("Auto or manual")

            # Connect lotes changes to recalculate S.C for Brand B only (if needed)
            lotes_new_b = c_lotes_b.get("Empeno JEW. (NEW)")
            lotes_renew_b = c_lotes_b.get("Empeno JEW (RENEW)")
            if lotes_new_b:
                lotes_new_b.textChanged.connect(_calculate_sc_b)
            if lotes_renew_b:
                lotes_renew_b.textChanged.connect(_calculate_sc_b)

            # Connect Empeno JEW amount fields in Brand B to auto-update Jew. A.I and S.C
            jewb_new = c_inp_b.get("Empeno JEW. (NEW)")
            jewb_renew = c_inp_b.get("Empeno JEW (RENEW)")
            
            if jewb_new:
                jewb_new.textChanged.connect(_update_jew_ai_b)
                jewb_new.textChanged.connect(_calculate_sc_b)
            if jewb_renew:
                jewb_renew.textChanged.connect(_update_jew_ai_b)
                jewb_renew.textChanged.connect(_calculate_sc_b)

            logger.debug("Empeno JEW detail buttons set up for both Brand A and Brand B.")

        except Exception as e:
            logger.warning("_setup_empeno_jew_buttons error (non-fatal): %s", e)

    def _setup_empeno_motor_button(self):
  
        try:
            c_inp  = self.cash_flow_tab_a.credit_inputs
            d_inp  = self.cash_flow_tab_a.debit_inputs
            cb_inp = self.cash_flow_tab_b.credit_inputs

            field_label = "Empeno Motor/Car"
            field_a = c_inp.get(field_label)
            if field_a is None:
                return

            field_b = cb_inp.get(field_label)

            field_a.setReadOnly(True)
            field_a.setStyleSheet(
                field_a.styleSheet() +
                "background-color: #EFF6FF; color: #1D4ED8;"
            )
            field_a.setToolTip(f"Click + to enter breakdown for {field_label}")
            field_a.setPlaceholderText("Click + to enter detail")

            self._motor_dialog = None

            def _open_motor_dialog():
                if self._motor_dialog is None:
                    self._motor_dialog = MotorCarDetailDialog(field_label, parent=self)

                dlg = self._motor_dialog
                if dlg.exec_() == QDialog.Accepted:
                    val = f"{dlg.get_total_amount():.2f}"
                    row_count = dlg.get_row_count()

                    field_a.blockSignals(True)
                    field_a.setText(val)
                    field_a.blockSignals(False)

                    lotes_a = self.cash_flow_tab_a.credit_lotes_inputs.get(field_label)
                    if lotes_a:
                        lotes_a.blockSignals(True)
                        lotes_a.setText(str(row_count))
                        lotes_a.blockSignals(False)
                        lotes_a.textChanged.emit(str(row_count))

                    self._motor_car_breakdown = dlg.get_breakdown_data()

                    computed = dlg.get_computed_total()
                    motor_ai_a = d_inp.get("Motor A.I")
                    if motor_ai_a:
                        motor_ai_a.blockSignals(True)
                        motor_ai_a.setText(f"{computed:.2f}")
                        motor_ai_a.blockSignals(False)
                    
                    # Trigger auto-calculations for O.s.f Motor and other fields
                    field_a.textChanged.emit(val)

                    self.recalculate_all()

            container = field_a.parentWidget()
            layout = container.layout() if container else None
            if layout is None:
                return

            plus_btn = QPushButton("+")
            plus_btn.setFixedSize(28, 28)
            plus_btn.setStyleSheet("""
                QPushButton {
                    background-color: #2563EB; color: white;
                    border: none; border-radius: 5px;
                    font-size: 14px; font-weight: 900;
                }
                QPushButton:hover { background-color: #1D4ED8; }
            """)
            plus_btn.setToolTip(f"Open detail breakdown for {field_label}")
            plus_btn.clicked.connect(_open_motor_dialog)
            layout.insertWidget(0, plus_btn)

            motor_ai = d_inp.get("Motor A.I")
            if motor_ai:
                motor_ai.setToolTip("Computed Total from Empeno Motor/Car breakdown; editable")
                motor_ai.setPlaceholderText("Computed from Empeno Motor/Car — editable")

            logger.debug("Empeno Motor/Car detail button set up.")

        except Exception as e:
            logger.warning("_setup_empeno_motor_button error (non-fatal): %s", e)

    def _setup_ft_ho_button(self):

        try:
            field_label = "Fund Transfer to HEAD OFFICE"
            self._ft_ho_dialogs = {} 
            self._ft_ho_breakdowns = {} 

            brand_tabs = [
                ("Brand A", self.cash_flow_tab_a),
                ("Brand B", self.cash_flow_tab_b),
            ]

            for brand_key, cf_tab in brand_tabs:
                field = cf_tab.credit_inputs.get(field_label)
                if field is None:
                    continue

                if brand_key == "Brand A":
                    field.setReadOnly(True)
                    field.setStyleSheet(
                        field.styleSheet() +
                        "background-color: #F5F3FF; color: #6D28D9;"
                    )
                    field.setToolTip(f"Click + to enter breakdown for {field_label}")
                    field.setPlaceholderText("Click + to enter detail")

                if hasattr(cf_tab, 'bank_account_btn') and cf_tab.bank_account_btn:
                    cf_tab.bank_account_btn.setVisible(False)

                def _make_open_dialog(bkey, target_field, tab):
                    def _open():
                        if bkey not in self._ft_ho_dialogs:
                            self._ft_ho_dialogs[bkey] = FundTransferHODialog(
                                f"{field_label} ({bkey})", parent=self
                            )
                        dlg = self._ft_ho_dialogs[bkey]
                        if dlg.exec_() == QDialog.Accepted:
                            val = f"{dlg.get_total_amount():.2f}"

                            target_field.blockSignals(True)
                            target_field.setText(val)
                            target_field.blockSignals(False)

                            lotes = tab.credit_lotes_inputs.get(field_label)
                            if lotes:
                                lotes.blockSignals(True)
                                lotes.setText(str(dlg.get_row_count()))
                                lotes.blockSignals(False)
                                lotes.textChanged.emit(str(dlg.get_row_count()))

                            self._ft_ho_breakdowns[bkey] = dlg.get_breakdown_data()
                            self.recalculate_all()
                    return _open

                container = field.parentWidget()
                layout = container.layout() if container else None
                if layout is None:
                    continue

                # Only Brand A gets the + button (Brand B has no bank account)
                if brand_key != "Brand A":
                    continue

                plus_btn = QPushButton("+")
                plus_btn.setFixedSize(28, 28)
                plus_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #8B5CF6; color: white;
                        border: none; border-radius: 5px;
                        font-size: 14px; font-weight: 900;
                    }
                    QPushButton:hover { background-color: #7C3AED; }
                """)
                plus_btn.setToolTip(f"Open detail breakdown for {field_label} ({brand_key})")
                plus_btn.clicked.connect(_make_open_dialog(brand_key, field, cf_tab))
                layout.insertWidget(0, plus_btn)

            logger.debug("Fund Transfer to HEAD OFFICE detail buttons set up (both brands).")

        except Exception as e:
            logger.warning("_setup_ft_ho_button error (non-fatal): %s", e)

    def _get_draft_path(self):
        # Always store drafts in %APPDATA%\OperationReportSystem\drafts\
        # Program Files requires admin rights — writing there causes WinError 5.
        # %APPDATA% is always writable by the current user and survives restarts.
        if getattr(sys, 'frozen', False):
            appdata = os.environ.get('APPDATA') or os.path.expanduser('~')
            base = os.path.join(appdata, 'OperationReportSystem')
        else:
            base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        drafts_dir = os.path.join(base, "drafts")
        os.makedirs(drafts_dir, exist_ok=True)
        def _safe(s):
            return "".join(c if c.isalnum() or c in "_-" else "_" for c in str(s))
        return os.path.join(
            drafts_dir,
            f"{_safe(self.user_email)}_{_safe(self.branch)}_{_safe(self.corporation)}.json"
        )

    def save_draft(self):
        try:
            sd = self.date_picker.date().toString("yyyy-MM-dd")

            status_a = self.check_existing_entry(sd, "Brand A")
            status_b = self.check_existing_entry(sd, "Brand B")
            if status_a == "locked" and status_b == "locked":
                self._msg("Draft Not Saved",
                          f"Report for {sd} has already been submitted.\n"
                          "There is no need to save a draft for this date.",
                          QMessageBox.Information)
                return
            draft = {
                "saved_at": datetime.datetime.now().isoformat(),
                "date": sd,
                "brand_a": {
                    "beginning_balance": self.beginning_balance_input_a.text(),
                    "beginning_balance_auto_filled": self.beginning_balance_auto_filled_a,
                    "cash_count": self.cash_count_input_a.text(),
                    "cash_flow": self.cash_flow_tab_a.get_raw_field_values(),
                    "bank_account": self.cash_flow_tab_a.selected_bank_account,
                },
                "brand_b": {
                    "beginning_balance": self.beginning_balance_input_b.text(),
                    "beginning_balance_auto_filled": self.beginning_balance_auto_filled_b,
                    "cash_count": self.cash_count_input_b.text(),
                    "cash_flow": self.cash_flow_tab_b.get_raw_field_values(),
                    "bank_account": self.cash_flow_tab_b.selected_bank_account,
                },
                "palawan":     self._collect_palawan_for_draft(),
                "mc_in": self._collect_mc_details_for_draft("MC In"),
                "mc_out": self._collect_mc_details_for_draft("MC Out"),
            }
            path = self._get_draft_path()
            all_drafts = {}
            if os.path.exists(path):
                with open(path, "r", encoding="utf-8") as f:
                    existing = json.load(f)
                if "date" in existing and "saved_at" in existing:
                    old_date = existing["date"]
                    all_drafts[old_date] = existing
                else:
                    all_drafts = existing
            all_drafts[sd] = draft
            with open(path, "w", encoding="utf-8") as f:
                json.dump(all_drafts, f, indent=2, ensure_ascii=False)
            total = len(all_drafts)
            self._msg_success(
                f"Draft saved for {sd}.\n\n"
                f"You have {total} saved draft(s).\n"
                "Your data will be automatically restored when you return."
            )
        except Exception as e:
            self._msg("Save Draft Error", f"Could not save draft:\n{e}", QMessageBox.Warning)

    def _collect_palawan_for_draft(self):
        d = {}
        for section in ("sendout", "payout", "international", "adjustments"):
            for k, w in getattr(self.palawan_tab, f"{section}_inputs", {}).items():
                d[f"{section}:{k}"] = w.text()
        for k, w in getattr(self.palawan_tab, "lotes_inputs", {}).items():
            d[f"lotes:{k}"] = w.text()
        return d

    def _collect_mc_details_for_draft(self, field_name):
        details = []
        for cf_tab in [self.cash_flow_tab_a, self.cash_flow_tab_b]:
            if hasattr(cf_tab, 'mc_currency_details') and field_name in cf_tab.mc_currency_details:
                details.extend(cf_tab.mc_currency_details[field_name])
        return details

    def _load_draft(self):
        try:
            # Purge first — DB is guaranteed available at this point
            self._purge_submitted_drafts()

            path = self._get_draft_path()
            if not os.path.exists(path):
                return
            with open(path, "r", encoding="utf-8") as f:
                raw = json.load(f)

            # Normalise old single-entry format to {date: draft} dict
            if "date" in raw and "saved_at" in raw:
                all_drafts = {raw["date"]: raw}
            else:
                all_drafts = dict(raw)

            if not all_drafts:
                return

            sorted_dates = sorted(all_drafts.keys(), reverse=True)
            if not sorted_dates:
                return

            if len(sorted_dates) == 1:
                chosen_date = sorted_dates[0]
                draft = all_drafts[chosen_date]
                saved_at = draft.get("saved_at", "")[:19].replace("T", " ")
                reply = QMessageBox.question(
                    self, "Restore Draft",
                    f"A saved draft was found:\n\n"
                    f"  Date : {chosen_date}\n"
                    f"  Saved: {saved_at}\n\n"
                    f"Do you want to restore it?",
                    QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes
                )
                if reply != QMessageBox.Yes:
                    return
            else:
                items = []
                for d in sorted_dates:
                    saved_at = all_drafts[d].get("saved_at", "")[:19].replace("T", " ")
                    items.append(f"{d}  (saved: {saved_at})")
                from PyQt5.QtWidgets import QInputDialog
                choice, ok = QInputDialog.getItem(
                    self, "Restore Draft",
                    f"You have {len(sorted_dates)} saved drafts.\n"
                    f"Select a date to restore:",
                    items, 0, False
                )
                if not ok:
                    return
                chosen_date = sorted_dates[items.index(choice)]

            draft = all_drafts[chosen_date]

            qdate = QDate.fromString(chosen_date, "yyyy-MM-dd")
            if qdate.isValid():
                self.date_picker.setDate(qdate)

            # CRITICAL: After changing date, check if report is already submitted
            # If so, don't overwrite with draft data
            status_a = self.check_existing_entry(chosen_date, "Brand A")
            status_b = self.check_existing_entry(chosen_date, "Brand B")
            if (status_a == "locked") or (status_b == "locked"):
                # Report already submitted - don't load draft data
                return

            for letter, key in [("a", "brand_a"), ("b", "brand_b")]:
                bd      = draft.get(key, {})
                bb_inp  = getattr(self, f"beginning_balance_input_{letter}")
                cc_inp  = getattr(self, f"cash_count_input_{letter}")
                cf_tab  = getattr(self, f"cash_flow_tab_{letter}")
                if bd.get("beginning_balance"):
                    bb_inp.setReadOnly(False)
                    bb_inp.setText(bd["beginning_balance"])
                    setattr(self, f"beginning_balance_auto_filled_{letter}",
                            bd.get("beginning_balance_auto_filled", True))
                if bd.get("cash_count"):
                    cc_inp.setText(bd["cash_count"])
                if bd.get("cash_flow"):
                    cf_tab.set_raw_field_values(bd["cash_flow"])

            for key, val in draft.get("palawan", {}).items():
                section, label = key.split(":", 1)
                w = (self.palawan_tab.lotes_inputs.get(label) if section == "lotes"
                     else getattr(self.palawan_tab, f"{section}_inputs", {}).get(label))
                if w:
                    w.blockSignals(True); w.setText(val); w.blockSignals(False)
            self.palawan_tab.calculate_palawan_totals()
            self.palawan_tab.calculate_lotes_total()
            self.palawan_tab.calculate_adjustments_total()

            mc_in_entries = draft.get("mc_in", [])
            mc_out_entries = draft.get("mc_out", [])
            for cf_tab in [self.cash_flow_tab_a, self.cash_flow_tab_b]:
                if hasattr(cf_tab, 'mc_currency_details'):
                    if mc_in_entries:
                        cf_tab.set_mc_currency_details('MC In', mc_in_entries)
                    if mc_out_entries:
                        cf_tab.set_mc_currency_details('MC Out', mc_out_entries)
            
            self.recalculate_all()

        except Exception as e:
            logger.warning("_load_draft error (non-fatal): %s", e)

    def _purge_submitted_drafts(self):
        """Remove all draft entries whose dates already have a submitted report.
        Safe to call at any time — silently skips if file doesn't exist."""
        try:
            path = self._get_draft_path()
            if not os.path.exists(path):
                return
            with open(path, "r", encoding="utf-8") as f:
                raw = json.load(f)

            # Normalise old single-entry format to {date: draft} dict
            if "date" in raw and "saved_at" in raw:
                all_drafts = {raw["date"]: raw}
            else:
                all_drafts = dict(raw)

            removed = []
            for d in list(all_drafts.keys()):
                sa = self.check_existing_entry(d, "Brand A")
                sb = self.check_existing_entry(d, "Brand B")
                if sa == "locked" or sb == "locked":
                    del all_drafts[d]
                    removed.append(d)

            if not removed:
                return

            logger.debug("Purged submitted drafts: %s", removed)
            if not all_drafts:
                os.remove(path)
            else:
                with open(path, "w", encoding="utf-8") as f:
                    json.dump(all_drafts, f, indent=2, ensure_ascii=False)
        except Exception as e:
            logger.warning("_purge_submitted_drafts error: %s", e)

    def _delete_draft(self, date=None):
        try:
            path = self._get_draft_path()
            if not os.path.exists(path):
                return
            with open(path, "r", encoding="utf-8") as f:
                raw = json.load(f)

            if "date" in raw and "saved_at" in raw:
                if date is None or raw.get("date") == date:
                    os.remove(path)
                return

            target = date or self.date_picker.date().toString("yyyy-MM-dd")
            if target in raw:
                del raw[target]
            if not raw:
                os.remove(path)
            else:
                with open(path, "w", encoding="utf-8") as f:
                    json.dump(raw, f, indent=2, ensure_ascii=False)
        except Exception:
            pass

    def get_submitted_dates(self):
  
        submitted_dates = {}
        
        try:

            query_a = """
                SELECT DISTINCT date FROM daily_reports_brand_a
                WHERE branch=%s AND corporation=%s
                ORDER BY date DESC
            """
            results_a = self.db_manager.execute_query(query_a, (self.branch, self.corporation))
            if results_a:
                for row in results_a:
                    date_str = str(row.get('date', ''))
                    if date_str:
                        if date_str not in submitted_dates:
                            submitted_dates[date_str] = {'date': date_str, 'brand_a': False, 'brand_b': False}
                        submitted_dates[date_str]['brand_a'] = True
            
            query_b = """
                SELECT DISTINCT date FROM daily_reports
                WHERE branch=%s AND corporation=%s
                ORDER BY date DESC
            """
            results_b = self.db_manager.execute_query(query_b, (self.branch, self.corporation))
            if results_b:
                for row in results_b:
                    date_str = str(row.get('date', ''))
                    if date_str:
                        if date_str not in submitted_dates:
                            submitted_dates[date_str] = {'date': date_str, 'brand_a': False, 'brand_b': False}
                        submitted_dates[date_str]['brand_b'] = True
                        
        except Exception as e:
            logger.error("get_submitted_dates error: %s", e)
        
        return sorted(submitted_dates.values(), key=lambda x: x['date'], reverse=True)

    def show_load_report_dialog(self):
        """
        Show a dialog with all submitted report dates for the user to select.
        """
        submitted = self.get_submitted_dates()
        
        if not submitted:
            QMessageBox.information(
                self, "No Submitted Reports",
                "No submitted reports found for your branch.\n"
                "Submit a report first to be able to load it later.",
                QMessageBox.Ok
            )
            return
        
        dialog = QDialog(self)
        dialog.setWindowTitle("Load Submitted Report")
        dialog.setMinimumSize(500, 400)
        dialog.setStyleSheet(f"""
            QDialog {{
                background-color: {_BG_CARD};
            }}
            QLabel {{
                color: {_TEXT_PRI};
                background: transparent;
            }}
        """)
        
        layout = QVBoxLayout(dialog)
        layout.setSpacing(12)
        layout.setContentsMargins(16, 16, 16, 16)
        
        # Header
        title = QLabel("Select a Submitted Report to Load")
        title.setStyleSheet(f"font-size: 14px; font-weight: 800; color: {_TEXT_PRI};")
        layout.addWidget(title)
        
        info = QLabel("This will load the report data into the form for viewing or printing.\nThe data is read-only.")
        info.setStyleSheet(f"font-size: 11px; color: {_TEXT_SEC};")
        layout.addWidget(info)

        table = QTableWidget()
        table.setColumnCount(4)
        table.setHorizontalHeaderLabels(["Date", "Brand A", "Brand B", "Action"])
        table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        table.setSelectionMode(QAbstractItemView.NoSelection)
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        table.setAlternatingRowColors(True)
        table.setStyleSheet(f"""
            QTableWidget {{
                border: 1px solid {_BORDER};
                border-radius: 6px;
                background-color: {_BG_CARD};
            }}
            QHeaderView::section {{
                background-color: {_SLATE_100};
                font-weight: 700;
                font-size: 11px;
                padding: 8px;
                border: none;
                border-bottom: 1px solid {_BORDER};
            }}
            QTableWidget::item {{
                padding: 6px;
            }}
        """)
        
        table.setRowCount(len(submitted))
        
        for row, entry in enumerate(submitted):

            date_item = QTableWidgetItem(entry['date'])
            date_item.setTextAlignment(Qt.AlignCenter)
            table.setItem(row, 0, date_item)
            
     
            a_status = "✅" if entry['brand_a'] else "—"
            a_item = QTableWidgetItem(a_status)
            a_item.setTextAlignment(Qt.AlignCenter)
            table.setItem(row, 1, a_item)
            

            b_status = "✅" if entry['brand_b'] else "—"
            b_item = QTableWidgetItem(b_status)
            b_item.setTextAlignment(Qt.AlignCenter)
            table.setItem(row, 2, b_item)
            
            load_btn = QPushButton("Load")
            load_btn.setStyleSheet(f"""
                QPushButton {{
                    background-color: {_PRIMARY};
                    color: {_WHITE};
                    border: none;
                    border-radius: 4px;
                    padding: 5px 15px;
                    font-size: 11px;
                    font-weight: 700;
                }}
                QPushButton:hover {{ background-color: {_PRIMARY_DK}; }}
            """)
            load_btn.clicked.connect(lambda checked, d=entry['date'], dlg=dialog: self._load_submitted_report(d, dlg))
            table.setCellWidget(row, 3, load_btn)
        
        table.resizeRowsToContents()
        layout.addWidget(table)
        
        close_btn = QPushButton("Close")
        close_btn.setFixedSize(100, 34)
        close_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {_SLATE_200};
                color: {_TEXT_PRI};
                border: none;
                border-radius: 6px;
                font-weight: 700;
            }}
            QPushButton:hover {{ background-color: {_SLATE_300}; }}
        """)
        close_btn.clicked.connect(dialog.close)
        layout.addWidget(close_btn, alignment=Qt.AlignRight)
        
        dialog.exec_()

    def _load_submitted_report(self, date_str, dialog=None):

        try:
            if dialog:
                dialog.close()

            # Clear cached dialogs/data to avoid showing stale values
            try:
                self._clear_all_dialog_caches()
            except Exception:
                pass

            # Show loading overlay while fetching and populating report
            try:
                self.loading_overlay.set_status("Loading report...", "Fetching latest report")
                self.loading_overlay.show()
                QApplication.processEvents()
            except Exception:
                pass
            
            for tab in (self.cash_flow_tab_a, self.cash_flow_tab_b):
                if hasattr(tab, 'clear_fields'):
                    tab.clear_fields()
            for bb in (self.beginning_balance_input_a, self.beginning_balance_input_b):
                bb.clear()
            for cc in (self.cash_count_input_a, self.cash_count_input_b):
                cc.clear()

            self._toggle_inputs(True)
            
            qdate = QDate.fromString(date_str, "yyyy-MM-dd")
            if qdate.isValid():
                self.date_picker.blockSignals(True)
                self.date_picker.setDate(qdate)
                self.date_picker.blockSignals(False)
            
            self._load_brand_report_data("Brand A", date_str)
            
            self._load_brand_report_data("Brand B", date_str)
            
            self.recalculate_all()

            self._toggle_inputs(False)
            self.post_button.setEnabled(False)
            self.auto_fill_button_a.setEnabled(False)
            self.auto_fill_button_b.setEnabled(False)

            for brand in ("A", "B"):
                self._set_status_brand(brand, "Viewing Submitted Report", _TEXT_MUTED, bold=True)
            
            # Ensure UI updated before notifying user
            QApplication.processEvents()

            QMessageBox.information(
                self, "Report Loaded",
                f"Report for {date_str} has been loaded.\n\n"
                "Note: This is a read-only view of the submitted report.",
                QMessageBox.Ok
            )
            
            try:
                self.loading_overlay.hide()
            except Exception:
                pass
            
        except Exception as e:
            QMessageBox.critical(
                self, "Load Error",
                f"Failed to load report:\n{str(e)}",
                QMessageBox.Ok
            )

    def _load_brand_report_data(self, brand, date_str):

        try:
            table_name = "daily_reports_brand_a" if brand == "Brand A" else "daily_reports"
            cf_tab = self.cash_flow_tab_a if brand == "Brand A" else self.cash_flow_tab_b
            bb_input = self.beginning_balance_input_a if brand == "Brand A" else self.beginning_balance_input_b
            cc_input = self.cash_count_input_a if brand == "Brand A" else self.cash_count_input_b

            # Clear stale widget state before populating from DB
            if hasattr(cf_tab, 'clear_fields'):
                cf_tab.clear_fields()

            if brand == "Brand B":
                # daily_reports is exclusively the Brand B table; no brand filter needed.
                results = None
                for q, params in [
                    (f"SELECT * FROM {table_name} WHERE date=%s AND branch=%s AND corporation=%s LIMIT 1",
                     (date_str, self.branch, self.corporation)),
                    (f"SELECT * FROM {table_name} WHERE date=%s AND branch=%s LIMIT 1",
                     (date_str, self.branch)),
                ]:
                    results = self.db_manager.execute_query(q, params)
                    if results:
                        break
            else:
                results = None
                for q, params in [
                    (f"SELECT * FROM {table_name} WHERE date=%s AND branch=%s AND corporation=%s LIMIT 1",
                     (date_str, self.branch, self.corporation)),
                    (f"SELECT * FROM {table_name} WHERE date=%s AND branch=%s LIMIT 1",
                     (date_str, self.branch)),
                ]:
                    results = self.db_manager.execute_query(q, params)
                    if results:
                        break
            
            if not results or len(results) == 0:
                logger.debug(
                    "_load_brand_report_data: No row found for brand=%s branch=%s date=%s corp=%s",
                    brand, self.branch, date_str, self.corporation
                )
                return
            
            data = results[0]
            
            beginning_balance = data.get('beginning_balance', 0) or 0
            bb_input.setReadOnly(False)
            bb_input.setText(f"{float(beginning_balance):.2f}")
            
            cash_count = data.get('cash_count', 0) or 0
            cc_input.setText(f"{float(cash_count):.2f}")

            # Load cash float for Brand A from cash_float_tbl
            if brand == "Brand A":
                try:
                    # Query cash_float_tbl for the stored cash float value
                    cf_result = self.db_manager.execute_query(
                        "SELECT cash_float FROM cash_float_tbl WHERE date=%s AND branch=%s AND corporation=%s LIMIT 1",
                        (date_str, self.branch, self.corporation)
                    )
                    cash_float = 0.0
                    if cf_result:
                        cash_float = cf_result[0].get('cash_float', 0) or 0

                    cash_float = float(cash_float)
                    if hasattr(self, 'cash_float_input_a') and self.cash_float_input_a:
                        self.cash_float_input_a.blockSignals(True)
                        self.cash_float_input_a.setText(f"{cash_float:.2f}")
                        self.cash_float_input_a.blockSignals(False)
                    else:
                        logger.warning("cash_float_input_a widget not found or is None")
                except Exception as e:
                    logger.error("Error loading cash_float for Brand A: %s", e)
            
   
            col_mapping = cf_tab._build_column_mapping()
            

            reverse_mapping = {v: k for k, v in col_mapping.items()}
            
            for label, widget in cf_tab.debit_inputs.items():
                db_col = col_mapping.get(label, cf_tab._sanitize_column(label))
                value = data.get(db_col, 0)
                widget.blockSignals(True)
                if value:
                    widget.setText(f"{float(value):.2f}")
                else:
                    widget.clear()  # Clear if value is 0 or missing
                widget.blockSignals(False)
                
                lotes_col = db_col + "_lotes"
                lotes_val = data.get(lotes_col, 0)
                lotes_widget = cf_tab.debit_lotes_inputs.get(label)
                if lotes_widget:
                    lotes_widget.blockSignals(True)
                    if lotes_val:
                        try:
                            lotes_widget.setText(str(int(float(lotes_val))))
                        except (ValueError, TypeError):
                            lotes_widget.setText(str(lotes_val))
                    else:
                        lotes_widget.clear()  # Clear if value is 0 or missing
                    lotes_widget.blockSignals(False)
            
            for label, widget in cf_tab.credit_inputs.items():
                db_col = col_mapping.get(label, cf_tab._sanitize_column(label))
                value = data.get(db_col, 0)
                widget.blockSignals(True)
                if value:
                    widget.setText(f"{float(value):.2f}")
                else:
                    widget.clear()  # Clear if value is 0 or missing
                widget.blockSignals(False)
                
                lotes_col = db_col + "_lotes"
                lotes_val = data.get(lotes_col, 0)
                lotes_widget = cf_tab.credit_lotes_inputs.get(label)
                if lotes_widget:
                    lotes_widget.blockSignals(True)
                    if lotes_val:
                        try:
                            lotes_widget.setText(str(int(float(lotes_val))))
                        except (ValueError, TypeError):
                            lotes_widget.setText(str(lotes_val))
                    else:
                        lotes_widget.clear()  # Clear if value is 0 or missing
                    lotes_widget.blockSignals(False)
            
            # Force a full UI recalculation now that all fields are populated
            self.recalculate_all()
            
        except Exception as e:
            logger.error("_load_brand_report_data error (%s): %s", brand, e)


    def _gather_nonzero_fields(self):

        return self._gather_nonzero_fields_brand("a")

    def _gather_nonzero_fields_brand(self, brand_letter):

        result = {"debit": [], "credit": []}
        letter = brand_letter.lower()
        
        cf_tab   = getattr(self, f"cash_flow_tab_{letter}")
        bb_input = getattr(self, f"beginning_balance_input_{letter}")
        eb_disp  = getattr(self, f"ending_balance_display_{letter}")
        cc_input = getattr(self, f"cash_count_input_{letter}")
        cr_disp  = getattr(self, f"cash_result_display_{letter}")
        
        for label, widget in cf_tab.debit_inputs.items():
            try:
                amount = float(widget.text().strip().replace(',', '') or 0)
            except ValueError:
                amount = 0.0
            if amount != 0:
                lotes_widget = cf_tab.debit_lotes_inputs.get(label)
                lotes = int(lotes_widget.text().strip() or 0) if lotes_widget else 0
                result["debit"].append((label, amount, lotes))
        
        for label, widget in cf_tab.credit_inputs.items():
            try:
                amount = float(widget.text().strip().replace(',', '') or 0)
            except ValueError:
                amount = 0.0
            if amount != 0:
                lotes_widget = cf_tab.credit_lotes_inputs.get(label)
                lotes = int(lotes_widget.text().strip() or 0) if lotes_widget else 0
                result["credit"].append((label, amount, lotes))
        
        try:
            bb_text = bb_input.text().replace(',', '').strip()
            result["beginning_balance"] = float(bb_text) if bb_text and bb_text != "—" else 0.0
        except ValueError:
            result["beginning_balance"] = 0.0
        
        try:
            eb_text = eb_disp.text().replace(',', '').strip()
            result["ending_balance"] = float(eb_text) if eb_text and eb_text != "—" else 0.0
        except ValueError:
            result["ending_balance"] = 0.0
        
        try:
            cc_text = cc_input.text().replace(',', '').strip()
            result["cash_count"] = float(cc_text) if cc_text and cc_text != "—" else 0.0
        except ValueError:
            result["cash_count"] = 0.0
        
        try:
            var_text = cr_disp.text().replace(',', '').strip()
            result["variance"] = float(var_text) if var_text and var_text != "—" else 0.0
        except ValueError:
            result["variance"] = 0.0
        
        return result

    def print_nonzero_report(self):

        try:
            data = self._gather_nonzero_fields()
            
            if not data["debit"] and not data["credit"]:
                QMessageBox.warning(
                    self, "No Data",
                    "No non-zero values found in the current report.\n"
                    "Please enter some values first.",
                    QMessageBox.Ok
                )
                return
            
            selected_date = self.date_picker.date().toString("yyyy-MM-dd")
            user_info = f"{self.user_email} · {self.branch} · {self.corporation}"
            
            debit_total = sum(amount for _, amount, _ in data["debit"])
            credit_total = sum(amount for _, amount, _ in data["credit"])
            
            doc = QTextDocument()
            # Set default document font to ensure proper scaling for print
            default_font = QFont("Segoe UI", 12)
            doc.setDefaultFont(default_font)

            html = f"""
            <html>
            <head>
                <style>
                    /* Page body uses a centered container sized for print */
                    body {{ font-family: 'Segoe UI', Arial, sans-serif; font-size: 8pt; color: black; margin: 0; padding: 0; }}
                    /* Container sized to approximately half A4 width (95mm) and 3/4 A4 height (~208mm) */
                    .page-container {{ width: 95mm; height: 208mm; margin: 10mm auto; padding: 6mm; box-sizing: border-box; }}
                    h1 {{ font-size: 8pt; margin-bottom: 8px; color: black; font-weight: bold; }}
                    h2 {{ font-size: 6pt; margin-top: 2px; margin-bottom: 2px; color: black; font-weight: bold; }}
                    h3 {{ font-size: 6pt; margin-top: 2px; margin-bottom: 2px; color: black; font-weight: bold; }}
                    .meta {{ font-size: 8pt; margin-bottom: 4px; color: black; line-height: 1.4; }}
                    table {{ width: 100%; margin-bottom: 12px; border-spacing: 0; }}
                    th {{ padding: 4px; text-align: left; font-size: 7pt; color: black; border: none; font-weight: bold; }}
                    td {{ padding: 2px 2px; font-size: 5pt; border: none; line-height: 1.3; }}
                    .amount {{ text-align: right; font-weight: 600; }}
                    .lotes {{ text-align: center; color: black; }}
                    .total-row {{ font-weight: 700; }}
                    .summary-table {{ margin-top: 8px; }}
                    .summary-label {{ font-weight: 700; color: black; }}
                    .variance-positive {{ color: black; font-weight: 700; }}
                    .variance-negative {{ color: black; font-weight: 700; }}
                    .variance-zero {{ color: black; font-weight: 700; }}
                </style>
            </head>
            <body>
                <div class="page-container">
                <h1 style="font-size:8pt;">Daily Cash Report - Brand A</h1>
                <div class="meta">
                    <b>Date:</b> {selected_date}<br>
                    <b>User:</b> {user_info}<br>
                    <b>Generated:</b> {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
                </div>
                
                <h2 style="font-size:8pt;">Summary</h2>
                <table class="summary-table">
                    <tr>
                        <td class="summary-label">Beginning Balance</td>
                        <td class="amount">{data['beginning_balance']:,.2f}</td>
                    </tr>
                    <tr>
                        <td class="summary-label">Ending Balance</td>
                        <td class="amount">{data['ending_balance']:,.2f}</td>
                    </tr>
                    <tr>
                        <td class="summary-label">Cash Count</td>
                        <td class="amount">{data['cash_count']:,.2f}</td>
                    </tr>
                    <tr>
                        <td class="summary-label">Variance</td>
                        <td class="amount {'variance-positive' if data['variance'] > 0 else 'variance-negative' if data['variance'] < 0 else 'variance-zero'}">
                            {data['variance']:,.2f} {'(Over)' if data['variance'] > 0 else '(Short)' if data['variance'] < 0 else '(Balanced)'}
                        </td>
                    </tr>
                </table>
            """
            
            if data["debit"]:
                html += '<h3 style="font-size:8pt;">Cash In (Debit)</h3>'
                html += '<table>'
                html += '<tr><th>Field</th><th>Amount</th><th>Lotes</th></tr>'
                
                for label, amount, lotes in data["debit"]:
                    html += f'<tr><td>{label}</td><td class="amount">{amount:,.2f}</td><td class="lotes">{lotes if lotes else "-"}</td></tr>'
                
                html += f'<tr class="total-row"><td><b>Total Cash Receipt</b></td><td class="amount"><b>{debit_total:,.2f}</b></td><td></td></tr>'
                html += '</table>'
            
            if data["credit"]:
                html += '<h3 style="font-size:8pt;">Cash Out (Credit)</h3>'
                html += '<table>'
                html += '<tr><th>Field</th><th>Amount</th><th>Lotes</th></tr>'
                
                for label, amount, lotes in data["credit"]:
                    html += f'<tr><td>{label}</td><td class="amount">{amount:,.2f}</td><td class="lotes">{lotes if lotes else "-"}</td></tr>'
                
                html += f'<tr class="total-row"><td><b>Total Cash Out</b></td><td class="amount"><b>{credit_total:,.2f}</b></td><td></td></tr>'
                html += '</table>'
            
            html += """
                </div>
                <br><br>
                <table style="width: 100%;">
                    <tr>
                        <td style="width: 50%;"><b>Prepared by:</b> ________________________</td>
                        <td style="width: 50%;"><b>Approved by:</b> ________________________</td>
                    </tr>
                </table>
            </body>
            </html>
            """
            
            doc.setHtml(html)
            
            printer = QPrinter(QPrinter.HighResolution)
            printer.setPageSize(QPrinter.A4)
            printer.setPageMargins(10, 10, 10, 10, QPrinter.Millimeter)
            printer.setResolution(300)
            dialog = QPrintDialog(printer, self)
            dialog.setWindowTitle("Print Non-Zero Report")
            
            if dialog.exec_() == QPrintDialog.Accepted:
                doc.print_(printer)
                
        except Exception as e:
            QMessageBox.critical(
                self, "Print Error",
                f"An error occurred while printing:\n{str(e)}",
                QMessageBox.Ok
            )

    def export_to_excel(self):

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            QMessageBox.critical(
                self,
                "Missing Dependency",
                "The openpyxl package is required to export to Excel.\n"
                "Install with: pip install openpyxl",
                QMessageBox.Ok
            )
            return
        
        try:
            data_a = self._gather_nonzero_fields_brand("a")
            data_b = self._gather_nonzero_fields_brand("b")
            
            has_data_a = data_a["debit"] or data_a["credit"]
            has_data_b = data_b["debit"] or data_b["credit"]
            
            if not has_data_a and not has_data_b:
                QMessageBox.warning(
                    self, "No Data",
                    "No non-zero values found in the current report.\n"
                    "Please enter some values first.",
                    QMessageBox.Ok
                )
                return
            
            selected_date = self.date_picker.date().toString("yyyy-MM-dd")
            
            default_filename = f"Daily_Report_{self.branch}_{selected_date}.xlsx"
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Save Excel File",
                default_filename,
                "Excel Files (*.xlsx);;All Files (*)"
            )
            
            if not file_path:
                return  
            

            wb = Workbook()
            

            # Use only black text; remove fills and borders for black-only printers
            title_font = Font(bold=True, size=16, color="000000")
            header_font = Font(bold=True, size=11, color="000000")
            # No fills (remove background colors)
            header_fill = PatternFill(fill_type=None)
            summary_fill = PatternFill(fill_type=None)
            total_fill = PatternFill(fill_type=None)
            # No visible borders
            border = Border()
            
            brands_to_export = []
            if has_data_a:
                brands_to_export.append(("Brand A", data_a))
            if has_data_b:
                brands_to_export.append(("Brand B", data_b))
            
            first_sheet = True
            for brand_name, data in brands_to_export:
                if first_sheet:
                    ws = wb.active
                    ws.title = f"{brand_name} Report"
                    first_sheet = False
                else:
                    ws = wb.create_sheet(title=f"{brand_name} Report")

                # Hide Excel gridlines for cleaner printouts
                try:
                    ws.sheet_view.showGridLines = False
                except Exception:
                    pass
                

                ws.merge_cells('A1:D1')
                ws['A1'] = f"Daily Cash Report - {brand_name}"
                ws['A1'].font = title_font
                ws['A1'].alignment = Alignment(horizontal='center')
                

                ws['A3'] = "Date:"
                ws['B3'] = selected_date
                ws['A4'] = "Branch:"
                ws['B4'] = self.branch
                ws['A5'] = "Corporation:"
                ws['B5'] = self.corporation
                ws['A6'] = "User:"
                ws['B6'] = self.user_email
                ws['A7'] = "Generated:"
                ws['B7'] = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                
                for row in range(3, 8):
                    ws.cell(row=row, column=1).font = Font(bold=True)
                
                current_row = 9
                
                ws.merge_cells(f'A{current_row}:D{current_row}')
                summary_header = ws.cell(row=current_row, column=1)
                summary_header.value = "Summary"
                summary_header.font = Font(bold=True, size=14, color="000000")
                summary_header.fill = summary_fill
                summary_header.alignment = Alignment(horizontal='center')
                current_row += 1
                
                summary_items = [
                    ("Beginning Balance", data["beginning_balance"]),
                    ("Ending Balance", data["ending_balance"]),
                    ("Cash Count", data["cash_count"]),
                    ("Variance", data["variance"]),
                ]
                
                for label, value in summary_items:
                    label_cell = ws.cell(row=current_row, column=1)
                    label_cell.value = label
                    label_cell.font = Font(bold=True, color="000000")
                    # no border for black-only print
                    label_cell.border = border
                    
                    value_cell = ws.cell(row=current_row, column=2)
                    value_cell.value = value
                    value_cell.number_format = '#,##0.00'
                    value_cell.alignment = Alignment(horizontal='right')
                    value_cell.border = border
                    
                    if label == "Variance":
                        status_cell = ws.cell(row=current_row, column=3)
                        if value > 0:
                            status_cell.value = "(Over)"
                            status_cell.font = Font(bold=True, color="000000")
                        elif value < 0:
                            status_cell.value = "(Short)"
                            status_cell.font = Font(bold=True, color="000000")
                        else:
                            status_cell.value = "(Balanced)"
                            status_cell.font = Font(bold=True, color="000000")
                        status_cell.border = border
                    
                    current_row += 1
                
                current_row += 1  
                

                if data["debit"]:
                    for col, hdr in enumerate(["Field", "Amount", "Lotes", ""], 1):
                        cell = ws.cell(row=current_row, column=col)
                        cell.value = hdr
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.border = border
                        cell.alignment = Alignment(horizontal='center')
                    current_row += 1
                    
                    debit_total = 0.0
                    for label, amount, lotes in data["debit"]:
                        debit_total += amount
                        ws.cell(row=current_row, column=1).value = label
                        ws.cell(row=current_row, column=1).border = border
                        
                        amount_cell = ws.cell(row=current_row, column=2)
                        amount_cell.value = amount
                        amount_cell.number_format = '#,##0.00'
                        amount_cell.alignment = Alignment(horizontal='right')
                        amount_cell.border = border
                        
                        lotes_cell = ws.cell(row=current_row, column=3)
                        lotes_cell.value = lotes if lotes else "-"
                        lotes_cell.alignment = Alignment(horizontal='center')
                        lotes_cell.border = border
                        
                        current_row += 1
                    
                    total_label = ws.cell(row=current_row, column=1)
                    total_label.value = "Total Cash Receipt"
                    total_label.font = Font(bold=True, color="000000")
                    total_label.fill = total_fill
                    total_label.border = border
                    
                    total_amount = ws.cell(row=current_row, column=2)
                    total_amount.value = debit_total
                    total_amount.number_format = '#,##0.00'
                    total_amount.font = Font(bold=True)
                    total_amount.fill = total_fill
                    total_amount.alignment = Alignment(horizontal='right')
                    total_amount.border = border
                    
                    ws.cell(row=current_row, column=3).fill = total_fill
                    ws.cell(row=current_row, column=3).border = border
                    
                    current_row += 2
                
      
                if data["credit"]:
                    for col, hdr in enumerate(["Field", "Amount", "Lotes", ""], 1):
                        cell = ws.cell(row=current_row, column=col)
                        cell.value = hdr
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.border = border
                        cell.alignment = Alignment(horizontal='center')
                    current_row += 1
                    
                    credit_total = 0.0
                    for label, amount, lotes in data["credit"]:
                        credit_total += amount
                        ws.cell(row=current_row, column=1).value = label
                        ws.cell(row=current_row, column=1).border = border
                        
                        amount_cell = ws.cell(row=current_row, column=2)
                        amount_cell.value = amount
                        amount_cell.number_format = '#,##0.00'
                        amount_cell.alignment = Alignment(horizontal='right')
                        amount_cell.border = border
                        
                        lotes_cell = ws.cell(row=current_row, column=3)
                        lotes_cell.value = lotes if lotes else "-"
                        lotes_cell.alignment = Alignment(horizontal='center')
                        lotes_cell.border = border
                        
                        current_row += 1
                    
                    total_label = ws.cell(row=current_row, column=1)
                    total_label.value = "Total Cash Out"
                    total_label.font = Font(bold=True, color="000000")
                    total_label.fill = total_fill
                    total_label.border = border
                    
                    total_amount = ws.cell(row=current_row, column=2)
                    total_amount.value = credit_total
                    total_amount.number_format = '#,##0.00'
                    total_amount.font = Font(bold=True)
                    total_amount.fill = total_fill
                    total_amount.alignment = Alignment(horizontal='right')
                    total_amount.border = border
                    
                    ws.cell(row=current_row, column=3).fill = total_fill
                    ws.cell(row=current_row, column=3).border = border
                    
                    current_row += 2
                

                ws.column_dimensions['A'].width = 35
                ws.column_dimensions['B'].width = 18
                ws.column_dimensions['C'].width = 12
                ws.column_dimensions['D'].width = 5
            

            wb.save(file_path)
            
            sheets_exported = " & ".join(b for b, _ in brands_to_export)
            QMessageBox.information(
                self, "Export Successful",
                f"{sheets_exported} exported to:\n{file_path}",
                QMessageBox.Ok
            )
            
        except Exception as e:
            QMessageBox.critical(
                self, "Export Error",
                f"An error occurred while exporting:\n{str(e)}",
                QMessageBox.Ok
            )


    _MSG_QSS = f"""
        QMessageBox {{
            background-color: {_BG_CARD};
            font-family: 'Segoe UI', 'SF Pro Text', Arial, sans-serif;
            font-size: 12px;
        }}
        QMessageBox QLabel {{ color: {_TEXT_PRI}; font-size: 12px; min-width: 0; }}
        QMessageBox QPushButton {{
            background-color: {_PRIMARY};
            color: {_WHITE};
            border-radius: 6px;
            padding: 8px 20px;
            font-weight: 700;
            font-size: 11px;
            min-width: 80px;
        }}
        QMessageBox QPushButton:hover {{ background-color: {_PRIMARY_DK}; }}
    """

    def _msg(self, title, text, icon):
        m = QMessageBox(icon, title, text, QMessageBox.Ok, self)
        m.setStyleSheet(self._MSG_QSS)
        m.exec_()

    def _msg_success(self, text):
        m = QMessageBox(self)
        m.setIcon(QMessageBox.Information)
        m.setWindowTitle("Success")
        m.setText(text)
        m.setStyleSheet(
            self._MSG_QSS
            .replace(_PRIMARY, _SUCCESS)
            .replace(_PRIMARY_DK, _SUCCESS_DK)
        )
        m.exec_()


    def show_message(self, title, message, icon):
        self._msg(title, message, icon)

    def show_success_message(self, message):
        self._msg_success(message)

    def validate_beginning_balance(self):
        return self.beginning_balance_auto_filled_a

    def validate_required_fields(self):
        return self.validate_all_requirements()

    def check_duplicate_entry(self, selected_date):
        return self.check_existing_entry(selected_date, "Brand A")

    def _sync_palawan_mc_to_other_brand(self, date, palawan_data, mc_data):
        pass 