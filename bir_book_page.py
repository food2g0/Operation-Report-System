"""BIR Book Tab - Display all Palawan transaction details from database."""

from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QComboBox, QPushButton,
    QTableWidget, QTableWidgetItem, QHeaderView, QAbstractItemView, QDateEdit, QMessageBox,
    QSpinBox, QFileDialog, QApplication, QDialog, QGridLayout
)
from PyQt5.QtCore import Qt, QDate
from PyQt5.QtGui import QFont
from api_db_manager import db_manager
import json
import logging
from datetime import datetime

logger = logging.getLogger(__name__)


class MonthlyReportDialog(QDialog):
    """Dialog to select parameters for monthly report generation."""

    def __init__(self, corporations, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Generate Monthly Report")
        self.setMinimumWidth(400)
        self.corporations = corporations
        self._init_ui()

    def _init_ui(self):
        """Setup dialog UI."""
        layout = QGridLayout(self)
        layout.setSpacing(12)

        # Corporation
        corp_label = QLabel("Corporation:")
        corp_label.setStyleSheet("font-weight: 600;")
        self.corp_combo = QComboBox()
        self.corp_combo.addItems(self.corporations)
        layout.addWidget(corp_label, 0, 0)
        layout.addWidget(self.corp_combo, 0, 1)

        # Month
        month_label = QLabel("Month:")
        month_label.setStyleSheet("font-weight: 600;")
        self.month_combo = QComboBox()
        self.month_combo.addItems([
            "January", "February", "March", "April", "May", "June",
            "July", "August", "September", "October", "November", "December"
        ])
        self.month_combo.setCurrentIndex(QDate.currentDate().month() - 1)
        layout.addWidget(month_label, 1, 0)
        layout.addWidget(self.month_combo, 1, 1)

        # Year
        year_label = QLabel("Year:")
        year_label.setStyleSheet("font-weight: 600;")
        self.year_spinner = QSpinBox()
        self.year_spinner.setMinimum(2020)
        self.year_spinner.setMaximum(2099)
        self.year_spinner.setValue(QDate.currentDate().year())
        layout.addWidget(year_label, 2, 0)
        layout.addWidget(self.year_spinner, 2, 1)

        # Branch Status
        status_label = QLabel("Branch Status:")
        status_label.setStyleSheet("font-weight: 600;")
        self.status_combo = QComboBox()
        self.status_combo.addItems(["All Branches", "Registered", "Not Registered"])
        layout.addWidget(status_label, 3, 0)
        layout.addWidget(self.status_combo, 3, 1)

        # Buttons
        button_layout = QHBoxLayout()
        generate_btn = QPushButton("Generate Report")
        generate_btn.setStyleSheet("""
            QPushButton {
                background: #F59E0B; color: white; border: none;
                border-radius: 5px; padding: 6px 16px;
                font-weight: 700;
            }
            QPushButton:hover { background: #D97706; }
        """)
        generate_btn.clicked.connect(self.accept)

        cancel_btn = QPushButton("Cancel")
        cancel_btn.setStyleSheet("""
            QPushButton {
                background: #6B7280; color: white; border: none;
                border-radius: 5px; padding: 6px 16px;
                font-weight: 700;
            }
            QPushButton:hover { background: #4B5563; }
        """)
        cancel_btn.clicked.connect(self.reject)

        button_layout.addWidget(generate_btn)
        button_layout.addWidget(cancel_btn)
        layout.addLayout(button_layout, 4, 0, 1, 2)

    def get_params(self):
        """Return selected parameters."""
        return {
            'corporation': self.corp_combo.currentText(),
            'month': self.month_combo.currentIndex() + 1,
            'year': self.year_spinner.value(),
            'status': self.status_combo.currentText()
        }


class BIRBookPage(QWidget):
    """Tab for viewing all Palawan transactions (BIR Book compliance)."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_dashboard = parent
        self.db = db_manager
        self.branch_status_map = {}
        self.all_transactions = []  # Store all transactions for pagination
        self.current_page = 1
        self.rows_per_page = 50
        self.expanded_groups = set()  # Track which groups are expanded (date, branch)
        self._setup_ui()
        self._load_corporations()
        self._find_available_dates()  # Auto-detect latest date with data

    def _setup_ui(self):
        """Setup the user interface."""
        root = QVBoxLayout(self)
        root.setContentsMargins(12, 12, 12, 12)
        root.setSpacing(12)

        # Title
        title = QLabel("BIR Book - Palawan Transaction Details")
        title.setStyleSheet("font-size: 16px; font-weight: 800; color: #1E293B;")
        root.addWidget(title)

        # Filter section - Row 1: Corporation, Date, and Transaction Type
        filter_row1 = QHBoxLayout()
        filter_row1.setSpacing(16)

        # Corporation filter (REQUIRED - no "All" option)
        corp_label = QLabel("Corporation:")
        corp_label.setStyleSheet("font-weight: 600; color: #334155;")
        filter_row1.addWidget(corp_label)

        self.corporation_combo = QComboBox()
        self.corporation_combo.setMinimumWidth(250)
        self.corporation_combo.currentTextChanged.connect(self._on_filter_changed)
        filter_row1.addWidget(self.corporation_combo)

        filter_row1.addSpacing(32)

        # Date filter
        date_label = QLabel("Date:")
        date_label.setStyleSheet("font-weight: 600; color: #334155;")
        filter_row1.addWidget(date_label)

        self.date_picker = QDateEdit()
        # Set to today's date (use QDate explicitly)
        today = QDate.currentDate()
        self.date_picker.setDate(today)
        self.date_picker.setCalendarPopup(True)
        # Allow selecting any date (past or future)
        self.date_picker.setDateRange(QDate(2020, 1, 1), QDate(2099, 12, 31))
        self.date_picker.dateChanged.connect(self._on_filter_changed)
        filter_row1.addWidget(self.date_picker)

        filter_row1.addStretch()

        load_btn = QPushButton("Load Report")
        load_btn.setStyleSheet("""
            QPushButton {
                background: #0EA5E9; color: white; border: none;
                border-radius: 5px; padding: 6px 16px;
                font-weight: 700;
            }
            QPushButton:hover { background: #0284C7; }
        """)
        load_btn.clicked.connect(self._load_transactions)
        filter_row1.addWidget(load_btn)

        monthly_report_btn = QPushButton("Generate Monthly Report")
        monthly_report_btn.setStyleSheet("""
            QPushButton {
                background: #F59E0B; color: white; border: none;
                border-radius: 5px; padding: 6px 16px;
                font-weight: 700;
            }
            QPushButton:hover { background: #D97706; }
        """)
        monthly_report_btn.clicked.connect(self._generate_monthly_report)
        filter_row1.addWidget(monthly_report_btn)

        export_btn = QPushButton("Export to Excel")
        export_btn.setStyleSheet("""
            QPushButton {
                background: #10B981; color: white; border: none;
                border-radius: 5px; padding: 6px 16px;
                font-weight: 700;
            }
            QPushButton:hover { background: #059669; }
        """)
        export_btn.clicked.connect(self._export_to_excel)
        filter_row1.addWidget(export_btn)

        filter_row1.addSpacing(20)

        # Transaction Type toggle
        txn_type_label = QLabel("Transaction Type:")
        txn_type_label.setStyleSheet("font-weight: 600; color: #334155;")
        filter_row1.addWidget(txn_type_label)

        self.txn_type_combo = QComboBox()
        self.txn_type_combo.setMinimumWidth(150)
        self.txn_type_combo.addItem("Send-Out (Sendout)", "sendout")
        self.txn_type_combo.addItem("Pay-Out (Payout)", "payout")
        self.txn_type_combo.currentIndexChanged.connect(self._on_filter_changed)
        filter_row1.addWidget(self.txn_type_combo)

        root.addLayout(filter_row1)

        # Info label
        self.info_label = QLabel("")
        self.info_label.setStyleSheet("font-size: 11px; color: #64748B;")
        root.addWidget(self.info_label)

        # Transactions table
        self.table = QTableWidget()
        self.table.setColumnCount(21)
        self.table.setHorizontalHeaderLabels([
            "Date", "Branch", "Code", "Receiver", "Sender", "Principal",
            "Commission", "SC", "Total SC", "Cash on Hand", "Income", "VAT (12%)", "A/P Palawan",
            "KYC Docs", "Business Name", "Position", "Relationship",
            "Source Funds", "Purpose", "Evaluation", ""
        ])

        hh = self.table.horizontalHeader()
        hh.setSectionResizeMode(QHeaderView.Stretch)
        # Fix last column to have more width for button
        hh.setSectionResizeMode(20, QHeaderView.ResizeToContents)
        self.table.setSelectionMode(QAbstractItemView.NoSelection)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self.table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #E2E8F0;
                border-radius: 6px;
                gridline-color: #E2E8F0;
            }
            QHeaderView::section {
                background: #F1F5F9;
                font-weight: 700;
                font-size: 11px;
                padding: 8px;
                border: 1px solid #CBD5E1;
            }
            QTableWidget::item {
                padding: 6px 8px;
            }
        """)

        root.addWidget(self.table)

        # Pagination controls
        pagination_row = QHBoxLayout()
        pagination_row.setSpacing(12)

        prev_btn = QPushButton("◀ Previous")
        prev_btn.setMaximumWidth(100)
        prev_btn.setStyleSheet("""
            QPushButton {
                background: #94A3B8; color: white; border: none;
                border-radius: 4px; padding: 5px 12px;
                font-weight: 600;
            }
            QPushButton:hover { background: #78909C; }
        """)
        prev_btn.clicked.connect(self._prev_page)
        pagination_row.addWidget(prev_btn)

        self.page_info = QLabel("Page 1")
        self.page_info.setStyleSheet("font-weight: 600; color: #334155;")
        pagination_row.addWidget(self.page_info)

        next_btn = QPushButton("Next ▶")
        next_btn.setMaximumWidth(100)
        next_btn.setStyleSheet("""
            QPushButton {
                background: #94A3B8; color: white; border: none;
                border-radius: 4px; padding: 5px 12px;
                font-weight: 600;
            }
            QPushButton:hover { background: #78909C; }
        """)
        next_btn.clicked.connect(self._next_page)
        pagination_row.addWidget(next_btn)

        pagination_row.addSpacing(20)

        rows_label = QLabel("Rows per page:")
        rows_label.setStyleSheet("font-weight: 600; color: #334155;")
        pagination_row.addWidget(rows_label)

        self.rows_spinner = QSpinBox()
        self.rows_spinner.setValue(50)
        self.rows_spinner.setMinimum(10)
        self.rows_spinner.setMaximum(500)
        self.rows_spinner.setSingleStep(10)
        self.rows_spinner.setMaximumWidth(80)
        self.rows_spinner.valueChanged.connect(self._on_rows_per_page_changed)
        pagination_row.addWidget(self.rows_spinner)

        pagination_row.addStretch()
        root.addLayout(pagination_row)

    def _find_available_dates(self):
        """Find the most recent date with data in the database (not exceeding today)."""
        try:
            result = self.db.execute_query(
                "SELECT DISTINCT DATE(date) as report_date FROM payable_tbl_brand_a ORDER BY DATE(date) DESC LIMIT 1"
            )

            if result:
                date_str = result[0].get("report_date")
                if date_str:
                    logger.info(f"[BIRBookPage] Found latest data date: {date_str}")
                    # Parse the date and set it
                    parts = str(date_str).split('-')
                    if len(parts) == 3:
                        year, month, day = int(parts[0]), int(parts[1]), int(parts[2])
                        latest_date = QDate(year, month, day)

                        # Only use the date if it's not in the future
                        today = QDate.currentDate()
                        if latest_date <= today:
                            self.date_picker.blockSignals(True)
                            self.date_picker.setDate(latest_date)
                            self.date_picker.blockSignals(False)
                            return True
                        else:
                            logger.warning(f"[BIRBookPage] Latest date {date_str} is in future, using today instead")
                            return False
            return False
        except Exception as e:
            logger.error("[BIRBookPage] Find available dates error: %s", e)
            return False

    def _load_corporations(self):
        """Load list of corporations into filter (no 'All' option to prevent app crash)."""
        try:
            self.corporation_combo.blockSignals(True)
            self.corporation_combo.clear()

            result = self.db.execute_query(
                "SELECT name FROM corporations ORDER BY name"
            )

            if result:
                for row in result:
                    name = row['name'] if isinstance(row, dict) else row[0]
                    if name:
                        self.corporation_combo.addItem(name, name)
                logger.info(f"[BIRBookPage] Loaded {len(result)} corporations")
                if self.corporation_combo.count() > 0:
                    self.corporation_combo.setCurrentIndex(0)
            else:
                logger.warning("[BIRBookPage] No corporations found in database")
                QMessageBox.warning(self, "No Data", "No corporations found in the database.")

            self.corporation_combo.blockSignals(False)
        except Exception as e:
            logger.error("[BIRBookPage] Load corporations error: %s", e)
            QMessageBox.warning(self, "Error", f"Failed to load corporations: {str(e)}")

    def _on_filter_changed(self):
        """Handle filter changes."""
        self.current_page = 1
        self._load_transactions()

    def _load_transactions(self):
        """Load all transactions and display current page."""
        try:
            self.table.setRowCount(0)
            self.info_label.setText("Loading transactions...")

            selected_corp = self.corporation_combo.currentData()
            selected_date = self.date_picker.date().toString("yyyy-MM-dd")

            if not selected_corp:
                self.info_label.setText("❌ Please select a corporation")
                logger.warning("[BIRBookPage] No corporation selected")
                return

            logger.info(f"[BIRBookPage] Loading transactions - Corp: {selected_corp}, Date: {selected_date}")

            # Query all payable records for the selected corporation and date
            query = """
                SELECT date, branch, corporation,
                       sendout_detailed_principal,
                       payout_detailed_principal,
                       international_detailed_principal
                FROM payable_tbl_brand_a
                WHERE (
                    corporation = %s
                    OR branch COLLATE utf8mb4_general_ci IN (
                        SELECT br.name FROM branches br
                        INNER JOIN corporations mc ON br.corporation_id = mc.id
                        WHERE mc.name = %s
                        UNION
                        SELECT br.name FROM branches br
                        INNER JOIN corporations sc ON br.sub_corporation_id = sc.id
                        WHERE sc.name = %s
                    )
                ) AND date = %s
                ORDER BY branch
            """
            result = self.db.execute_query(query, (selected_corp, selected_corp, selected_corp, selected_date))

            if not result:
                self.all_transactions = []
                self.info_label.setText(f"No records found for {selected_date} | Corporation: {selected_corp}")
                logger.info("[BIRBookPage] No records found")
                return

            logger.info(f"[BIRBookPage] Found {len(result)} total records for date {selected_date}")

            # Deduplicate DB records by (branch, date) — the OR clause in the query can
            # return the same physical row twice when a branch matches both the corporation
            # column and the branch-lookup subquery.  Keep the last occurrence (latest save).
            seen_br_date = {}
            for rec in result:
                k = (str(rec.get('branch', '') or '').strip(),
                     str(rec.get('date',   '') or ''))
                seen_br_date[k] = rec
            unique_result = list(seen_br_date.values())
            if len(unique_result) < len(result):
                logger.warning(
                    f"[BIRBookPage] Collapsed {len(result) - len(unique_result)} "
                    f"duplicate DB records for the same (branch, date)"
                )

            # Collect all transactions into a flat list.
            # Dedup is done INLINE using the record's own date/branch (not the spread
            # txn dict, which can override 'date'/'branch' if old saves stored them).
            self.all_transactions = []
            seen_keys = set()

            for record in unique_result:
                date   = record.get("date",   "")
                branch = record.get("branch", "")
                # Use string-normalised values for the dedup key so that
                # datetime.date objects and their string equivalents match.
                date_key   = str(date   or '').strip()
                branch_key = str(branch or '').strip()

                for section_key in ("sendout", "payout", "international"):
                    col_name = f"{section_key}_detailed_principal"
                    json_str = record.get(col_name)

                    if json_str:
                        try:
                            transactions = json.loads(json_str)
                            logger.debug(
                                f"[BIRBookPage] Parsed {len(transactions)} "
                                f"transactions from {col_name}"
                            )
                            for txn in transactions:
                                code_key = str(txn.get('code', '') or '').strip()
                                # Fallback discriminator for codeless rows:
                                # use receiver + principal so two identical
                                # no-code transactions still collapse to one key.
                                if not code_key:
                                    code_key = (
                                        str(txn.get('receiver',  '') or '').strip()
                                        + '|'
                                        + str(txn.get('principal', 0) or 0)
                                    )
                                key = (date_key, branch_key, section_key, code_key)
                                if key in seen_keys:
                                    logger.debug(
                                        f"[BIRBookPage] Skipping duplicate txn "
                                        f"code={code_key!r} branch={branch_key!r}"
                                    )
                                    continue
                                seen_keys.add(key)
                                # Put **txn first so the explicit metadata keys
                                # (date, branch, _type) always win and cannot be
                                # overridden by a stale 'date' field inside the JSON.
                                txn_with_meta = {
                                    **txn,
                                    'date':   date,
                                    'branch': branch,
                                    '_type':  section_key,
                                }
                                self.all_transactions.append(txn_with_meta)
                        except (json.JSONDecodeError, TypeError) as e:
                            logger.error(
                                f"[BIRBookPage] JSON parse error for {col_name}: {e}"
                            )

            logger.info(
                f"[BIRBookPage] Total unique transactions: {len(self.all_transactions)}"
            )

            self.current_page = 1
            self._display_page()

        except Exception as e:
            logger.error("[BIRBookPage] Load transactions error: %s", e)
            self.info_label.setText(f"Error loading transactions: {str(e)}")
            QMessageBox.critical(self, "Error", f"Failed to load transactions:\n{str(e)}\n\nMake sure the database is connected and available.")

    def _display_page(self):
        """Display the current page with expandable groups and per-group totals."""
        self.table.setRowCount(0)

        if not self.all_transactions:
            self.info_label.setText("No transactions to display")
            return

        selected_corp = self.corporation_combo.currentData()
        selected_date = self.date_picker.date().toString("yyyy-MM-dd")
        selected_txn_type = self.txn_type_combo.currentData()

        # Show/hide Cash on Hand column (col 9) and update A/P vs A/R header (col 12)
        is_sendout = selected_txn_type == "sendout"
        self.table.setColumnHidden(9, not is_sendout)
        ar_ap_header = "A/P Palawan" if is_sendout else "A/R Palawan"
        self.table.setHorizontalHeaderItem(12, QTableWidgetItem(ar_ap_header))

        # Filter transactions by type
        filtered_txns = [t for t in self.all_transactions if t.get('_type') == selected_txn_type]

        # Safety-net dedup: catch any duplicates that survived _load_transactions.
        # Uses the DB-record date/branch (guaranteed by metadata-first construction)
        # so it's immune to stale 'date' fields inside the JSON payload.
        _seen = set()
        _clean = []
        for _t in filtered_txns:
            _code = str(_t.get('code', '') or '').strip()
            if not _code:
                _code = (str(_t.get('receiver', '') or '').strip()
                         + '|' + str(_t.get('principal', 0) or 0))
            _k = (str(_t.get('date',   '') or '').strip(),
                  str(_t.get('branch', '') or '').strip(),
                  _code)
            if _k not in _seen:
                _seen.add(_k)
                _clean.append(_t)
        filtered_txns = _clean

        start_idx = (self.current_page - 1) * self.rows_per_page
        end_idx = start_idx + self.rows_per_page
        page_transactions = filtered_txns[start_idx:end_idx]

        total_pages = (len(filtered_txns) + self.rows_per_page - 1) // self.rows_per_page

        # Group transactions by date and branch
        groups = {}
        for txn in page_transactions:
            key = (txn.get('date'), txn.get('branch'))
            if key not in groups:
                groups[key] = []
            groups[key].append(txn)

        # Store group data for expansion
        self.group_data = {}  # Maps (row_idx) -> (date, branch, all_txns)
        row_idx = 0

        # Display only first transaction per group with expand button
        for (date, branch), txns in groups.items():
            if txns:
                # INSERT NEW ROW FIRST
                self.table.insertRow(row_idx)

                # Add first transaction with date/branch visible
                self._add_table_row(txns[0], show_date_branch=True, row_idx=row_idx)

                # Add expand button in last column if group has multiple transactions
                if len(txns) > 1:
                    expand_btn = QPushButton("▼")
                    expand_btn.setMaximumWidth(35)
                    expand_btn.setMaximumHeight(25)
                    expand_btn.setMinimumWidth(35)
                    expand_btn.setMinimumHeight(25)
                    expand_btn.setCursor(Qt.PointingHandCursor)
                    expand_btn.setStyleSheet("""
                        QPushButton {
                            background: #3B82F6; color: white;
                            border: none; border-radius: 3px;
                            padding: 0px 5px; font-weight: 600;
                            font-size: 12px;
                        }
                        QPushButton:hover { background: #2563EB; }
                    """)
                    expand_btn.clicked.connect(lambda checked, r=row_idx, d=(date, branch), t=txns[1:]:
                                               self._toggle_group_expansion(r, d, t))

                    # Create container widget to center the button
                    container = QWidget()
                    layout = QHBoxLayout(container)
                    layout.setContentsMargins(2, 2, 2, 2)
                    layout.setSpacing(0)
                    layout.addStretch()
                    layout.addWidget(expand_btn)
                    layout.addStretch()

                    self.table.setCellWidget(row_idx, 18, container)
                    self.group_data[row_idx] = (date, branch, txns[1:])  # Store hidden transactions

                row_idx += 1

        self.page_info.setText(f"Page {self.current_page} of {total_pages}")
        self.info_label.setText(
            f"✓ Showing {len([g for g in groups])} branches | "
            f"Total: {len(filtered_txns)} transactions | "
            f"Type: {self.txn_type_combo.currentText()} | "
            f"Date: {selected_date} | Corporation: {selected_corp}"
        )

    def _toggle_group_expansion(self, header_row, group_key, hidden_txns):
        """Toggle expansion of a transaction group."""
        date, branch = group_key
        group_id = (header_row, date, branch)

        if group_id in self.expanded_groups:
            # Collapse: remove hidden rows and totals
            self._collapse_group(header_row, group_id)
        else:
            # Expand: add hidden rows and totals
            self._expand_group(header_row, group_key, hidden_txns)

    def _expand_group(self, header_row, group_key, hidden_txns):
        """Expand a group to show hidden transactions and totals."""
        date, branch = group_key
        group_id = (header_row, date, branch)

        # Get all transactions in this group (including the visible one)
        all_txns = [self.table.item(header_row, col).text() if self.table.item(header_row, col) else ''
                    for col in range(19)]
        all_txns_data = hidden_txns

        # Find first transaction to get the visible one
        first_txn_code = self.table.item(header_row, 2).text()
        first_txn = None
        for t in self.all_transactions:
            if t.get('code') == first_txn_code and t.get('date') == date and t.get('branch') == branch:
                first_txn = t
                break

        insert_row = header_row + 1

        # Insert hidden transactions (NOT including first_txn, which is already shown)
        for hidden_txn in hidden_txns:
            self.table.insertRow(insert_row)
            self._add_table_row(hidden_txn, show_date_branch=False, row_idx=insert_row)
            insert_row += 1

        # Insert totals row for this group (includes first_txn + hidden_txns)
        if first_txn:
            all_group_txns = [first_txn] + hidden_txns
        else:
            all_group_txns = hidden_txns

        self.table.insertRow(insert_row)
        self._add_group_totals_row(all_group_txns, insert_row)

        # Update button to collapse
        container = self.table.cellWidget(header_row, 18)
        if container:
            button = container.findChild(QPushButton)
            if button:
                button.setText("▲")

        # Mark as expanded
        self.expanded_groups.add(group_id)

    def _collapse_group(self, header_row, group_id):
        """Collapse a group to hide transactions and totals."""
        date, branch = group_id[:2] if len(group_id) > 2 else (None, None)

        # Count rows to remove (hidden transactions + totals)
        rows_to_remove = 0
        check_row = header_row + 1

        while check_row < self.table.rowCount():
            # Check if this is an empty row (part of the group)
            first_col = self.table.item(check_row, 0)
            if first_col and first_col.text().strip() == "":
                rows_to_remove += 1
                check_row += 1
            else:
                break

        # Remove rows in reverse order
        for _ in range(rows_to_remove):
            self.table.removeRow(header_row + 1)

        # Update button to expand
        container = self.table.cellWidget(header_row, 18)
        if container:
            button = container.findChild(QPushButton)
            if button:
                button.setText("▼")

        # Mark as collapsed
        self.expanded_groups.discard(group_id)


    def _prev_page(self):
        """Go to previous page."""
        if self.current_page > 1:
            self.current_page -= 1
            self._display_page()

    def _next_page(self):
        """Go to next page."""
        total_pages = (len(self.all_transactions) + self.rows_per_page - 1) // self.rows_per_page
        if self.current_page < total_pages:
            self.current_page += 1
            self._display_page()

    def _on_rows_per_page_changed(self):
        """Handle rows per page change."""
        self.rows_per_page = self.rows_spinner.value()
        self.current_page = 1
        self._display_page()

    def _add_table_row(self, txn, show_date_branch=True, row_idx=None):
        """Add a transaction row to the table."""
        if row_idx is None:
            row_idx = self.table.rowCount()

        # Map transaction data to table columns
        # Cash on Hand (col 9) is only relevant for sendout; blank for payout/international
        _txn_type = txn.get("_type", "")
        _principal = float(txn.get("principal", 0))
        _total_sc = float(txn.get("total_sc", 0))
        _income = float(txn.get("income", 0))
        _vat = float(txn.get("vat", _income / 1.12 * 0.12))
        _cash_on_hand = f"{_principal + _total_sc:.2f}" if _txn_type == "sendout" else ""
        data = [
            str(txn.get("date", "")) if show_date_branch else "",
            txn.get("branch", "") if show_date_branch else "",
            txn.get("code", ""),
            txn.get("receiver", ""),
            txn.get("sender", ""),
            f"{_principal:.2f}",
            f"{txn.get('commission', 0):.2f}",
            f"{txn.get('sc', 0):.2f}",
            f"{_total_sc:.2f}",
            _cash_on_hand,                   # col 9
            f"{_income:.2f}",                # col 10
            f"{_vat:.2f}",                   # col 11 VAT (12%)
            f"{txn.get('ar_palawan', 0):.2f}", # col 12
            txn.get("kyc_docs", ""),
            txn.get("business_name", ""),
            txn.get("position", ""),
            txn.get("relationship", ""),
            txn.get("source_funds", ""),
            txn.get("purpose", ""),
            txn.get("evaluation", ""),
            ""
        ]

        for col, value in enumerate(data):
            item = QTableWidgetItem(str(value))
            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
            # Right-align numeric columns
            if col in [5, 6, 7, 8, 9, 10, 11, 12]:
                item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            # Add light indentation for hidden rows (show_date_branch=False)
            if not show_date_branch and col == 2:
                item.setText("  " + item.text())
            self.table.setItem(row_idx, col, item)

    def _add_group_totals_row(self, transactions, row_idx):
        """Add a totals row for a specific branch group."""
        txn_type = transactions[0].get("_type", "") if transactions else ""
        is_sendout = txn_type == "sendout"

        total_principal = sum(float(t.get("principal", 0)) for t in transactions)
        total_commission = sum(float(t.get("commission", 0)) for t in transactions)
        total_sc = sum(float(t.get("sc", 0)) for t in transactions)
        total_total_sc = sum(float(t.get("total_sc", 0)) for t in transactions)
        total_income = sum(float(t.get("income", 0)) for t in transactions)
        total_vat = sum(float(t.get("vat", float(t.get("income", 0)) / 1.12 * 0.12)) for t in transactions)
        total_ar = sum(float(t.get("ar_palawan", 0)) for t in transactions)
        total_cash_on_hand = total_principal + total_total_sc

        totals_font = QFont()
        totals_font.setBold(True)

        def _create_totals_item(text=""):
            item = QTableWidgetItem(text)
            item.setFont(totals_font)
            item.setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
            return item

        def _create_num_item(value):
            item = _create_totals_item(f"{value:,.2f}")
            item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            return item

        # Empty cols 0-1
        for col in range(0, 2):
            self.table.setItem(row_idx, col, _create_totals_item())

        # "Totals" label in Code column (2)
        label_item = _create_totals_item("Totals")
        label_item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        self.table.setItem(row_idx, 2, label_item)

        # Empty cols 3-4
        for col in range(3, 5):
            self.table.setItem(row_idx, col, _create_totals_item())

        self.table.setItem(row_idx, 5, _create_num_item(total_principal))   # Principal
        self.table.setItem(row_idx, 6, _create_num_item(total_commission))  # Commission
        self.table.setItem(row_idx, 7, _create_num_item(total_sc))          # SC
        self.table.setItem(row_idx, 8, _create_num_item(total_total_sc))    # Total SC

        # Cash on Hand (col 9) — sendout only
        if is_sendout:
            self.table.setItem(row_idx, 9, _create_num_item(total_cash_on_hand))
        else:
            self.table.setItem(row_idx, 9, _create_totals_item())

        self.table.setItem(row_idx, 10, _create_num_item(total_income))     # Income
        self.table.setItem(row_idx, 11, _create_num_item(total_vat))       # VAT (12%)

        # A/P Total (sendout) or A/R Total (payout/international) at col 12
        ar_ap_item = _create_num_item(total_ar)
        self.table.setItem(row_idx, 12, ar_ap_item)

        # Rest of columns
        for col in range(13, 20):
            self.table.setItem(row_idx, col, _create_totals_item())

    def _export_to_excel(self):
        """Export current transactions to Excel file."""
        if not self.all_transactions:
            QMessageBox.warning(self, "No Data", "No transactions to export. Load a report first.")
            return

        selected_txn_type = self.txn_type_combo.currentData()
        filtered_txns = [t for t in self.all_transactions if t.get('_type') == selected_txn_type]

        if not filtered_txns:
            QMessageBox.warning(self, "No Data", f"No {self.txn_type_combo.currentText()} transactions to export.")
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            QMessageBox.warning(self, "Missing Library", "openpyxl is required for Excel export. Please install it.")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self, "Export Transactions to Excel", "",
            "Excel Files (*.xlsx);;All Files (*)"
        )

        if not file_path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "BIR Book"

            # Header styling
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Columns — labels vary by transaction type
            _income_lbl = "Income (39%)" if selected_txn_type == 'sendout' else "Income (43%)"
            _ar_ap_lbl = "A/P Palawan" if selected_txn_type == 'sendout' else "A/R Palawan"
            columns = [
                "Date", "Branch", "Code", "Receiver", "Sender", "Principal",
                "Commission", "SC", "Total SC", _income_lbl, "VAT (12%)", _ar_ap_lbl,
                "KYC Docs", "Business Name", "Position", "Relationship",
                "Source Funds", "Purpose", "Evaluation"
            ]

            # Write headers
            for col_idx, col_name in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = col_name
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = border

            # Write data
            for row_idx, txn in enumerate(filtered_txns, 2):
                ws.cell(row=row_idx, column=1).value = str(txn.get("date", ""))
                ws.cell(row=row_idx, column=2).value = txn.get("branch", "")
                ws.cell(row=row_idx, column=3).value = txn.get("code", "")
                ws.cell(row=row_idx, column=4).value = txn.get("receiver", "")
                ws.cell(row=row_idx, column=5).value = txn.get("sender", "")
                ws.cell(row=row_idx, column=6).value = float(txn.get("principal", 0))
                ws.cell(row=row_idx, column=7).value = float(txn.get("commission", 0))
                ws.cell(row=row_idx, column=8).value = float(txn.get("sc", 0))
                ws.cell(row=row_idx, column=9).value = float(txn.get("total_sc", 0))
                _inc = float(txn.get("income", 0))
                ws.cell(row=row_idx, column=10).value = _inc
                ws.cell(row=row_idx, column=11).value = float(txn.get("vat", _inc / 1.12 * 0.12))
                ws.cell(row=row_idx, column=12).value = float(txn.get("ar_palawan", 0))
                ws.cell(row=row_idx, column=13).value = txn.get("kyc_docs", "")
                ws.cell(row=row_idx, column=14).value = txn.get("business_name", "")
                ws.cell(row=row_idx, column=15).value = txn.get("position", "")
                ws.cell(row=row_idx, column=16).value = txn.get("relationship", "")
                ws.cell(row=row_idx, column=17).value = txn.get("source_funds", "")
                ws.cell(row=row_idx, column=18).value = txn.get("purpose", "")
                ws.cell(row=row_idx, column=19).value = txn.get("evaluation", "")

                # Apply borders and formatting
                for col in range(1, 19):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.border = border
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    if col in [6, 7, 8, 9, 10, 11, 12]:  # Number columns
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal="right", vertical="top")

            # Adjust column widths
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 10
            ws.column_dimensions['I'].width = 12
            ws.column_dimensions['J'].width = 12
            ws.column_dimensions['K'].width = 12
            ws.column_dimensions['L'].width = 12
            ws.column_dimensions['M'].width = 15
            ws.column_dimensions['N'].width = 18
            ws.column_dimensions['O'].width = 15
            ws.column_dimensions['P'].width = 15
            ws.column_dimensions['Q'].width = 15
            ws.column_dimensions['R'].width = 15

            # Add summary at bottom
            summary_row = len(filtered_txns) + 3
            ws.cell(row=summary_row, column=1).value = "Summary"
            ws.cell(row=summary_row, column=1).font = Font(bold=True, size=11)

            total_principal = sum(float(txn.get("principal", 0)) for txn in filtered_txns)
            total_commission = sum(float(txn.get("commission", 0)) for txn in filtered_txns)
            total_sc = sum(float(txn.get("sc", 0)) for txn in filtered_txns)
            total_total_sc = sum(float(txn.get("total_sc", 0)) for txn in filtered_txns)
            total_income = sum(float(txn.get("income", 0)) for txn in filtered_txns)
            total_vat = sum(float(txn.get("vat", float(txn.get("income", 0)) / 1.12 * 0.12)) for txn in filtered_txns)
            total_ar = sum(float(txn.get("ar_palawan", 0)) for txn in filtered_txns)

            ws.cell(row=summary_row + 1, column=5).value = "Total Principal:"
            ws.cell(row=summary_row + 1, column=6).value = total_principal
            ws.cell(row=summary_row + 1, column=6).number_format = '#,##0.00'
            ws.cell(row=summary_row + 1, column=6).font = Font(bold=True)

            ws.cell(row=summary_row + 2, column=5).value = "Total Commission:"
            ws.cell(row=summary_row + 2, column=7).value = total_commission
            ws.cell(row=summary_row + 2, column=7).number_format = '#,##0.00'
            ws.cell(row=summary_row + 2, column=7).font = Font(bold=True)

            ws.cell(row=summary_row + 3, column=5).value = "Total SC:"
            ws.cell(row=summary_row + 3, column=8).value = total_sc
            ws.cell(row=summary_row + 3, column=8).number_format = '#,##0.00'
            ws.cell(row=summary_row + 3, column=8).font = Font(bold=True)

            ws.cell(row=summary_row + 4, column=5).value = "Total VAT (12%):"
            ws.cell(row=summary_row + 4, column=11).value = total_vat
            ws.cell(row=summary_row + 4, column=11).number_format = '#,##0.00'
            ws.cell(row=summary_row + 4, column=11).font = Font(bold=True)

            ws.cell(row=summary_row + 5, column=5).value = "Total A/R Palawan:"
            ws.cell(row=summary_row + 5, column=12).value = total_ar
            ws.cell(row=summary_row + 5, column=12).number_format = '#,##0.00'
            ws.cell(row=summary_row + 5, column=12).font = Font(bold=True)

            wb.save(file_path)
            QMessageBox.information(
                self, "Export Successful",
                f"✓ Exported {len(filtered_txns)} {self.txn_type_combo.currentText()} transactions to:\n{file_path}"
            )
            logger.info(f"[BIRBookPage] Exported {len(filtered_txns)} transactions to {file_path}")

        except Exception as e:
            logger.error(f"[BIRBookPage] Export error: {e}")
            QMessageBox.critical(self, "Export Error", f"Failed to export to Excel:\n{str(e)}")

    def _generate_monthly_report(self):
        """Generate monthly Excel report with separate sheets per branch."""
        # Get list of corporations
        corporations = []
        for i in range(self.corporation_combo.count()):
            corp = self.corporation_combo.itemText(i)
            if corp:
                corporations.append(corp)

        if not corporations:
            QMessageBox.warning(self, "No Corporations", "No corporations available.")
            return

        # Show dialog
        dialog = MonthlyReportDialog(corporations, self)
        if dialog.exec_() != QDialog.Accepted:
            return

        params = dialog.get_params()
        selected_corp = params['corporation']
        month = params['month']
        year = params['year']
        status_filter = params['status']

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            QMessageBox.warning(self, "Missing Library", "openpyxl is required. Please install it.")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save Monthly Report",
            f"BIR_Book_{selected_corp}_{year}-{month:02d}.xlsx",
            "Excel Files (*.xlsx);;All Files (*)"
        )

        if not file_path:
            return

        try:
            self.info_label.setText("⏳ Generating monthly report...")

            # Load branch status map if filtering
            branch_status_map = {}
            if status_filter != "All Branches":
                try:
                    status_result = self.db.execute_query(
                        "SELECT branch_name, status FROM branch_table WHERE status IN ('Registered', 'Not Registered')"
                    )
                    if status_result:
                        for row in status_result:
                            branch_status_map[row.get("branch_name", "")] = row.get("status", "")
                except Exception as e:
                    logger.warning(f"Could not load branch status: {e}")

            # Query all transactions for the selected month and corporation
            query = """
                SELECT date, branch, corporation,
                       sendout_detailed_principal,
                       payout_detailed_principal,
                       international_detailed_principal
                FROM payable_tbl_brand_a
                WHERE (
                    corporation = %s
                    OR branch COLLATE utf8mb4_general_ci IN (
                        SELECT br.name FROM branches br
                        INNER JOIN corporations mc ON br.corporation_id = mc.id
                        WHERE mc.name = %s
                        UNION
                        SELECT br.name FROM branches br
                        INNER JOIN corporations sc ON br.sub_corporation_id = sc.id
                        WHERE sc.name = %s
                    )
                )
                AND YEAR(date) = %s
                AND MONTH(date) = %s
                ORDER BY date, branch
            """

            result = self.db.execute_query(query, (selected_corp, selected_corp, selected_corp, year, month))

            if not result:
                QMessageBox.warning(self, "No Data", f"No transactions found for {selected_corp} in {year}-{month:02d}")
                self.info_label.setText("No data for selected period")
                return

            # Deduplicate DB records by (branch, date) before parsing JSON
            seen_br_date = {}
            for rec in result:
                k = (str(rec.get('branch', '') or '').strip(),
                     str(rec.get('date',   '') or ''))
                seen_br_date[k] = rec
            unique_result = list(seen_br_date.values())
            if len(unique_result) < len(result):
                logger.warning(
                    f"[BIRBookPage] Monthly report: collapsed "
                    f"{len(result) - len(unique_result)} duplicate DB records"
                )

            # Collect all transactions
            all_txns = []
            seen_txn_keys = set()
            for record in unique_result:
                date = record.get("date", "")
                branch = record.get("branch", "")

                for section_key in ("sendout", "payout", "international"):
                    col_name = f"{section_key}_detailed_principal"
                    json_str = record.get(col_name)

                    if json_str:
                        try:
                            transactions = json.loads(json_str)
                            for txn in transactions:
                                txn_key = (
                                    str(date or ''),
                                    str(branch or '').strip(),
                                    section_key,
                                    str(txn.get('code', '') or ''),
                                )
                                if txn_key in seen_txn_keys:
                                    continue
                                seen_txn_keys.add(txn_key)
                                txn_with_meta = {
                                    'date': date,
                                    'branch': branch,
                                    '_type': section_key,
                                    **txn
                                }
                                all_txns.append(txn_with_meta)
                        except (json.JSONDecodeError, TypeError) as e:
                            logger.error(f"JSON parse error: {e}")

            if not all_txns:
                QMessageBox.warning(self, "No Data", "No detailed transactions found for this period.")
                self.info_label.setText("No detailed transactions")
                return

            # Create Excel workbook with separate sheets per branch
            wb = openpyxl.Workbook()
            wb.remove(wb.active)

            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            totals_fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
            totals_font = Font(bold=True, color="3730A3")

            sendout_columns = [
                "Date", "Branch", "Code", "Receiver", "Sender", "Principal",
                "Commission", "SC", "Total SC", "Cash on Hand", "Income (39%)", "A/P Palawan",
                "KYC Docs", "Business Name", "Position", "Relationship",
                "Source Funds", "Purpose", "Evaluation"
            ]

            payout_columns = [
                "Date", "Branch", "Code", "Receiver", "Sender", "Principal",
                "Commission", "SC", "Total SC", "Income (43%)", "A/R Palawan",
                "KYC Docs", "Business Name", "Position", "Relationship",
                "Source Funds", "Purpose", "Evaluation"
            ]

            # Group by branch and apply status filter
            branches = {}
            for txn in all_txns:
                branch = txn.get('branch', 'Unknown')

                # Apply branch status filter
                if status_filter != "All Branches":
                    branch_status = branch_status_map.get(branch, "Not Registered")
                    if status_filter == "Registered" and branch_status != "Registered":
                        continue
                    if status_filter == "Not Registered" and branch_status != "Not Registered":
                        continue

                if branch not in branches:
                    branches[branch] = []
                branches[branch].append(txn)

            # Create sheet per branch
            for branch_name, branch_txns in sorted(branches.items()):
                ws = wb.create_sheet(title=branch_name[:31])

                # Row 1: Section labels
                send_label_cell = ws.cell(row=1, column=1)
                send_label_cell.value = "SEND-OUT (Income 39%)"
                send_label_cell.font = Font(bold=True, size=12, color="0284C7")

                payout_label_cell = ws.cell(row=1, column=27)
                payout_label_cell.value = "PAY-OUT (Income 43%)"
                payout_label_cell.font = Font(bold=True, size=12, color="DC2626")

                # Row 2: Headers for SEND-OUT (columns A-Q)
                for col_idx, col_name in enumerate(sendout_columns, 1):
                    cell = ws.cell(row=2, column=col_idx)
                    cell.value = col_name
                    cell.font = Font(bold=True, size=11)
                    cell.alignment = header_alignment
                    cell.border = border

                # Row 2: Headers for PAY-OUT (columns AA-AR)
                for col_idx, col_name in enumerate(payout_columns, 27):  # Column 27 = AA
                    cell = ws.cell(row=2, column=col_idx)
                    cell.value = col_name
                    cell.font = Font(bold=True, size=11)
                    cell.alignment = header_alignment
                    cell.border = border

                # Separate sendout and payout transactions
                sendout_txns = [t for t in branch_txns if t.get('_type') == 'sendout']
                payout_txns = [t for t in branch_txns if t.get('_type') == 'payout']

                # Group by date
                dates_sendout = {}
                for t in sendout_txns:
                    d = str(t.get('date', ''))
                    if d not in dates_sendout:
                        dates_sendout[d] = []
                    dates_sendout[d].append(t)

                dates_payout = {}
                for t in payout_txns:
                    d = str(t.get('date', ''))
                    if d not in dates_payout:
                        dates_payout[d] = []
                    dates_payout[d].append(t)

                # Write data by date with separators
                row_idx = 3  # Start from row 3 (row 1 = labels, row 2 = headers)
                all_dates = sorted(set(list(dates_sendout.keys()) + list(dates_payout.keys())))

                for day_date in all_dates:
                    day_sendout = dates_sendout.get(day_date, [])
                    day_payout = dates_payout.get(day_date, [])
                    max_txns = max(len(day_sendout), len(day_payout))

                    # Write sendout and payout side by side
                    for idx in range(max_txns):
                        # SEND-OUT (columns A-Q)
                        if idx < len(day_sendout):
                            txn = day_sendout[idx]
                            commission = float(txn.get("commission", 0))
                            income_39 = round(commission * 0.39, 2)
                            principal = float(txn.get("principal", 0))
                            total_sc_so = float(txn.get("total_sc", 0))
                            cash_on_hand_so = principal + total_sc_so
                            ar_palawan = cash_on_hand_so - income_39

                            # Show date/branch only on first transaction of the day
                            ws.cell(row=row_idx, column=1).value = day_date if idx == 0 else ""
                            ws.cell(row=row_idx, column=2).value = txn.get("branch", "") if idx == 0 else ""
                            ws.cell(row=row_idx, column=3).value = txn.get("code", "")
                            ws.cell(row=row_idx, column=4).value = txn.get("receiver", "")
                            ws.cell(row=row_idx, column=5).value = txn.get("sender", "")
                            ws.cell(row=row_idx, column=6).value = principal
                            ws.cell(row=row_idx, column=7).value = commission
                            ws.cell(row=row_idx, column=8).value = float(txn.get("sc", 0))
                            ws.cell(row=row_idx, column=9).value = float(txn.get("total_sc", 0))
                            ws.cell(row=row_idx, column=10).value = cash_on_hand_so
                            ws.cell(row=row_idx, column=11).value = income_39
                            ws.cell(row=row_idx, column=12).value = ar_palawan
                            ws.cell(row=row_idx, column=13).value = txn.get("kyc_docs", "")
                            ws.cell(row=row_idx, column=14).value = txn.get("business_name", "")
                            ws.cell(row=row_idx, column=15).value = txn.get("position", "")
                            ws.cell(row=row_idx, column=16).value = txn.get("relationship", "")
                            ws.cell(row=row_idx, column=17).value = txn.get("source_funds", "")
                            ws.cell(row=row_idx, column=18).value = txn.get("purpose", "")
                            ws.cell(row=row_idx, column=19).value = txn.get("evaluation", "")

                        # PAY-OUT (columns AA-AR)
                        if idx < len(day_payout):
                            txn = day_payout[idx]
                            commission = float(txn.get("commission", 0))
                            income_43 = round(commission * 0.43, 2)
                            principal = float(txn.get("principal", 0))
                            total_sc = float(txn.get("total_sc", 0))
                            ar_palawan_po = principal + total_sc - income_43

                            ws.cell(row=row_idx, column=27).value = day_date if idx == 0 else ""  # AA
                            ws.cell(row=row_idx, column=28).value = txn.get("branch", "") if idx == 0 else ""  # AB
                            ws.cell(row=row_idx, column=29).value = txn.get("code", "")  # AC
                            ws.cell(row=row_idx, column=30).value = txn.get("receiver", "")  # AD
                            ws.cell(row=row_idx, column=31).value = txn.get("sender", "")  # AE
                            ws.cell(row=row_idx, column=32).value = principal  # AF
                            ws.cell(row=row_idx, column=33).value = commission  # AG
                            ws.cell(row=row_idx, column=34).value = float(txn.get("sc", 0))  # AH
                            ws.cell(row=row_idx, column=35).value = total_sc  # AI
                            ws.cell(row=row_idx, column=36).value = income_43  # AJ
                            ws.cell(row=row_idx, column=37).value = ar_palawan_po  # AK
                            ws.cell(row=row_idx, column=38).value = txn.get("kyc_docs", "")  # AL
                            ws.cell(row=row_idx, column=39).value = txn.get("business_name", "")  # AM
                            ws.cell(row=row_idx, column=40).value = txn.get("position", "")  # AN
                            ws.cell(row=row_idx, column=41).value = txn.get("relationship", "")  # AO
                            ws.cell(row=row_idx, column=42).value = txn.get("source_funds", "")  # AP
                            ws.cell(row=row_idx, column=43).value = txn.get("purpose", "")  # AQ
                            ws.cell(row=row_idx, column=44).value = txn.get("evaluation", "")  # AR

                        # Format all cells
                        for col in list(range(1, 20)) + list(range(27, 45)):  # sendout A-S, payout AA-AR
                            cell = ws.cell(row=row_idx, column=col)
                            cell.border = border
                            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                            if col in [6, 7, 8, 9, 10, 11, 12, 32, 33, 34, 35, 36, 37]:  # Number columns
                                cell.number_format = '#,##0.00'
                                cell.alignment = Alignment(horizontal="right", vertical="top")

                        row_idx += 1

                    # ── Per-day totals row ─────────────────────────────────
                    day_total_row = row_idx
                    row_idx += 1

                    # SEND-OUT daily totals
                    if day_sendout:
                        d_principal_so = sum(float(t.get("principal", 0)) for t in day_sendout)
                        d_commission_so = sum(float(t.get("commission", 0)) for t in day_sendout)
                        d_sc_so = sum(float(t.get("sc", 0)) for t in day_sendout)
                        d_total_sc_so = sum(float(t.get("total_sc", 0)) for t in day_sendout)
                        d_income_so = round(d_commission_so * 0.39, 2)
                        d_cash_on_hand_so = d_principal_so + d_total_sc_so
                        d_ap_so = d_cash_on_hand_so - d_income_so

                        ws.cell(row=day_total_row, column=3).value = "Totals"
                        ws.cell(row=day_total_row, column=6).value = d_principal_so
                        ws.cell(row=day_total_row, column=7).value = d_commission_so
                        ws.cell(row=day_total_row, column=8).value = d_sc_so
                        ws.cell(row=day_total_row, column=9).value = d_total_sc_so
                        ws.cell(row=day_total_row, column=10).value = d_cash_on_hand_so
                        ws.cell(row=day_total_row, column=11).value = d_income_so
                        ws.cell(row=day_total_row, column=12).value = d_ap_so

                    # PAY-OUT daily totals
                    if day_payout:
                        d_principal_po = sum(float(t.get("principal", 0)) for t in day_payout)
                        d_commission_po = sum(float(t.get("commission", 0)) for t in day_payout)
                        d_sc_po = sum(float(t.get("sc", 0)) for t in day_payout)
                        d_total_sc_po = sum(float(t.get("total_sc", 0)) for t in day_payout)
                        d_income_po = round(d_commission_po * 0.43, 2)
                        d_ar_po = d_principal_po + d_total_sc_po - d_income_po

                        ws.cell(row=day_total_row, column=29).value = "Totals"
                        ws.cell(row=day_total_row, column=32).value = d_principal_po
                        ws.cell(row=day_total_row, column=33).value = d_commission_po
                        ws.cell(row=day_total_row, column=34).value = d_sc_po
                        ws.cell(row=day_total_row, column=35).value = d_total_sc_po
                        ws.cell(row=day_total_row, column=36).value = d_income_po
                        ws.cell(row=day_total_row, column=37).value = d_ar_po

                    # Style the daily totals row
                    for col in list(range(1, 20)) + list(range(27, 45)):
                        cell = ws.cell(row=day_total_row, column=col)
                        cell.fill = totals_fill
                        cell.font = totals_font
                        cell.border = border
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                        if col in [6, 7, 8, 9, 10, 11, 12, 32, 33, 34, 35, 36, 37]:
                            cell.number_format = '#,##0.00'
                            cell.alignment = Alignment(horizontal="right", vertical="center")

                    # Blank rows between days for visual separation
                    for _ in range(2):
                        row_idx += 1

                # Branch totals
                totals_row = row_idx + 1
                ws.cell(row=totals_row, column=3).value = "BRANCH TOTAL"
                ws.cell(row=totals_row, column=3).fill = totals_fill
                ws.cell(row=totals_row, column=3).font = totals_font

                # SEND-OUT totals
                for col in range(1, 20):
                    cell = ws.cell(row=totals_row, column=col)
                    cell.fill = totals_fill
                    cell.font = totals_font
                    cell.border = border

                # PAY-OUT totals
                for col in range(27, 45):
                    cell = ws.cell(row=totals_row, column=col)
                    cell.fill = totals_fill
                    cell.font = totals_font
                    cell.border = border

                # Calculate totals with correct income percentages
                total_principal_so = sum(float(t.get("principal", 0)) for t in sendout_txns)
                total_commission_so = sum(float(t.get("commission", 0)) for t in sendout_txns)
                total_sc_so = sum(float(t.get("sc", 0)) for t in sendout_txns)
                total_total_sc_so = sum(float(t.get("total_sc", 0)) for t in sendout_txns)
                total_cash_on_hand_so = total_principal_so + total_total_sc_so
                total_income_so = round(total_commission_so * 0.39, 2)
                total_ar_so = total_cash_on_hand_so - total_income_so

                total_principal_po = sum(float(t.get("principal", 0)) for t in payout_txns)
                total_commission_po = sum(float(t.get("commission", 0)) for t in payout_txns)
                total_sc_po = sum(float(t.get("sc", 0)) for t in payout_txns)
                total_total_sc_po = sum(float(t.get("total_sc", 0)) for t in payout_txns)
                total_income_po = round(total_commission_po * 0.43, 2)
                total_ar_po = (total_principal_po + total_total_sc_po) - total_income_po

                # Write SEND-OUT totals
                ws.cell(row=totals_row, column=6).value = total_principal_so
                ws.cell(row=totals_row, column=6).number_format = '#,##0.00'
                ws.cell(row=totals_row, column=7).value = total_commission_so
                ws.cell(row=totals_row, column=7).number_format = '#,##0.00'
                ws.cell(row=totals_row, column=8).value = total_sc_so
                ws.cell(row=totals_row, column=8).number_format = '#,##0.00'
                ws.cell(row=totals_row, column=9).value = total_total_sc_so
                ws.cell(row=totals_row, column=9).number_format = '#,##0.00'
                ws.cell(row=totals_row, column=10).value = total_cash_on_hand_so
                ws.cell(row=totals_row, column=10).number_format = '#,##0.00'
                ws.cell(row=totals_row, column=11).value = total_income_so
                ws.cell(row=totals_row, column=11).number_format = '#,##0.00'
                ws.cell(row=totals_row, column=12).value = total_ar_so
                ws.cell(row=totals_row, column=12).number_format = '#,##0.00'

                # Write PAY-OUT totals
                ws.cell(row=totals_row, column=32).value = total_principal_po
                ws.cell(row=totals_row, column=32).number_format = '#,##0.00'
                ws.cell(row=totals_row, column=33).value = total_commission_po
                ws.cell(row=totals_row, column=33).number_format = '#,##0.00'
                ws.cell(row=totals_row, column=34).value = total_sc_po
                ws.cell(row=totals_row, column=34).number_format = '#,##0.00'
                ws.cell(row=totals_row, column=35).value = total_total_sc_po
                ws.cell(row=totals_row, column=35).number_format = '#,##0.00'
                ws.cell(row=totals_row, column=36).value = total_income_po
                ws.cell(row=totals_row, column=36).number_format = '#,##0.00'
                ws.cell(row=totals_row, column=37).value = total_ar_po
                ws.cell(row=totals_row, column=37).number_format = '#,##0.00'

                # Column widths for SEND-OUT (A-S, 19 cols)
                for col, width in [('A', 12), ('B', 15), ('C', 12), ('D', 15), ('E', 15), ('F', 12), ('G', 12), ('H', 10), ('I', 12), ('J', 12), ('K', 12), ('L', 12), ('M', 15), ('N', 18), ('O', 15), ('P', 15), ('Q', 15), ('R', 15), ('S', 15)]:
                    ws.column_dimensions[col].width = width
                # Spacing columns (T-Z) between send-out and pay-out
                for col in ['T', 'U', 'V', 'W', 'X', 'Y', 'Z']:
                    ws.column_dimensions[col].width = 2
                # Pay-Out column widths (AA-AR)
                for col, width in [('AA', 12), ('AB', 15), ('AC', 12), ('AD', 15), ('AE', 15), ('AF', 12), ('AG', 12), ('AH', 10), ('AI', 12), ('AJ', 12), ('AK', 12), ('AL', 15), ('AM', 18), ('AN', 15), ('AO', 15), ('AP', 15), ('AQ', 15), ('AR', 15)]:
                    ws.column_dimensions[col].width = width

            wb.save(file_path)
            QMessageBox.information(self, "Report Generated", f"✓ Monthly report saved:\n{file_path}")
            self.info_label.setText(f"✓ Report generated for {year}-{month:02d} ({len(branches)} branches)")
            logger.info(f"[BIRBookPage] Generated monthly report for {selected_corp} {year}-{month:02d}")

        except Exception as e:
            logger.error(f"[BIRBookPage] Monthly report error: {e}")
            QMessageBox.critical(self, "Error", f"Failed to generate report:\n{str(e)}")
            self.info_label.setText("Report generation failed")

    def refresh(self):
        """Refresh data (called when tab is shown)."""
        self._load_corporations()
