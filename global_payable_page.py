"""
Global Payable Page
───────────────────
Shows branches grouped by OS (Group) with Global tagging.
Displays Palawan Send Out, Payout, International data along with SKID, SKIR, INC, CANCEL.
"""

from PyQt5.QtWidgets import (
    QWidget, QLabel, QVBoxLayout, QComboBox, QTableWidget, QTableWidgetItem,
    QMessageBox, QHeaderView, QSizePolicy, QPushButton, QHBoxLayout,
    QScrollArea, QFrame, QFileDialog, QTreeWidget, QTreeWidgetItem
)
from PyQt5.QtCore import Qt, QDate
from PyQt5.QtGui import QFont, QColor, QBrush
from db_connect_pooled import db_manager
from db_worker import run_query_async
from date_range_widget import DateRangeWidget


class GlobalPayablePage(QWidget):
    """Global Payable page - shows branches grouped by OS with Global tagging"""
    
    def __init__(self, account_type=2):
        super().__init__()
        self.account_type = account_type
        # Brand B uses daily_reports table
        self.daily_table = "daily_reports"
        self._is_loading = False
        
        self.main_layout = QVBoxLayout(self)
        self.main_layout.setContentsMargins(15, 15, 15, 15)
        self.main_layout.setSpacing(10)
        
        self._build_controls()
        self._build_table()
        self._build_buttons()
        
        self._load_os_options()
    
    def _build_controls(self):
        """Build control panel with date selector and filters"""
        controls_frame = QFrame()
        controls_frame.setStyleSheet("""
            QFrame { background-color: #f8f9fa; border: 1px solid #dee2e6;
                     border-radius: 5px; padding: 10px; }
        """)
        controls_layout = QHBoxLayout(controls_frame)
        controls_layout.setSpacing(20)
        
        # Date selector
        date_label = QLabel("Date:")
        date_label.setFont(QFont("Arial", 12, QFont.Bold))
        self.date_range_widget = DateRangeWidget()
        self.date_range_widget.dateRangeChanged.connect(self.populate_table)
        
        controls_layout.addWidget(date_label)
        controls_layout.addWidget(self.date_range_widget)
        
        # Group Filter (OS)
        controls_layout.addSpacing(30)
        os_label = QLabel("Group Filter:")
        os_label.setFont(QFont("Arial", 12, QFont.Bold))
        self.os_filter_selector = QComboBox()
        self.os_filter_selector.setMinimumWidth(180)
        self.os_filter_selector.setMinimumHeight(40)
        self.os_filter_selector.addItem("All Groups", None)
        self.os_filter_selector.currentIndexChanged.connect(self.populate_table)
        self.os_filter_selector.setStyleSheet("""
            QComboBox { padding:10px; border:2px solid #dee2e6; border-radius:6px;
                        background-color:white; font-size:13px; }
            QComboBox:focus { border-color:#007bff; }
            QComboBox::drop-down { width: 30px; }
        """)
        controls_layout.addWidget(os_label)
        controls_layout.addWidget(self.os_filter_selector)

        # Branch Status filter
        controls_layout.addSpacing(30)
        reg_label = QLabel("Branch Status:")
        reg_label.setFont(QFont("Arial", 12, QFont.Bold))
        self.reg_filter_selector = QComboBox()
        self.reg_filter_selector.setMinimumWidth(150)
        self.reg_filter_selector.setMinimumHeight(40)
        self.reg_filter_selector.addItem("All", "all")
        self.reg_filter_selector.addItem("Registered Only", "registered")
        self.reg_filter_selector.addItem("Unregistered", "not_registered")
        self.reg_filter_selector.setCurrentIndex(1)  # Default to Registered
        self.reg_filter_selector.currentIndexChanged.connect(self.populate_table)
        self.reg_filter_selector.setStyleSheet("""
            QComboBox { padding:10px; border:2px solid #dee2e6; border-radius:6px;
                        background-color:white; font-size:13px; }
            QComboBox:focus { border-color:#007bff; }
            QComboBox::drop-down { width: 30px; }
        """)
        controls_layout.addWidget(reg_label)
        controls_layout.addWidget(self.reg_filter_selector)
        
        controls_layout.addStretch()
        self.main_layout.addWidget(controls_frame, 0)
    
    def _build_table(self):
        """Build the tree-style table showing groups and branches"""
        table_scroll = QScrollArea()
        table_scroll.setWidgetResizable(True)
        table_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        table_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        table_scroll.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        
        self.tree = QTreeWidget()
        self.tree.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.tree.setColumnCount(20)
        self.tree.setHeaderLabels([
            "Group / Branch",
            "SO:Lotes", "SO:Cap", "SO:SC", "SO:Com", "SO:TOTAL",
            "PO:Lotes", "PO:Cap", "PO:SC", "PO:Com", "PO:TOTAL",
            "IN:Lotes", "IN:Cap", "IN:SC", "IN:Com", "IN:TOTAL",
            "SKID", "SKIR", "CANCEL", "INC"
        ])
        
        # Set column widths
        self.tree.setColumnWidth(0, 200)  # Group/Branch name
        for i in range(1, 20):
            self.tree.setColumnWidth(i, 80)
        
        self.tree.setMinimumHeight(400)
        self.tree.setAlternatingRowColors(True)
        self.tree.setStyleSheet("""
            QTreeWidget {
                gridline-color: #d0d0d0;
                border: 1px solid #c0c0c0;
                font-size: 11px;
                alternate-background-color: #f9f9f9;
            }
            QTreeWidget::item { padding: 4px; }
            QHeaderView::section {
                background-color: #495057;
                color: white;
                padding: 6px;
                border: 1px solid #666;
                font-weight: bold;
            }
        """)
        
        table_scroll.setWidget(self.tree)
        self.main_layout.addWidget(table_scroll, 1)
    
    def _build_buttons(self):
        """Build action buttons"""
        button_frame = QFrame()
        button_frame.setFixedHeight(60)
        button_layout = QHBoxLayout(button_frame)
        
        refresh_btn = QPushButton("🔄 Refresh")
        refresh_btn.setMinimumHeight(40)
        refresh_btn.setStyleSheet("""
            QPushButton { background-color: #007bff; color: white; border: none;
                          border-radius: 5px; padding: 10px 20px; font-size: 13px; font-weight: bold; }
            QPushButton:hover { background-color: #0056b3; }
        """)
        refresh_btn.clicked.connect(self.populate_table)
        
        export_btn = QPushButton("📊 Export to Excel")
        export_btn.setMinimumHeight(40)
        export_btn.setStyleSheet("""
            QPushButton { background-color: #217346; color: white; border: none;
                          border-radius: 5px; padding: 10px 20px; font-size: 13px; font-weight: bold; }
            QPushButton:hover { background-color: #1a5c38; }
        """)
        export_btn.clicked.connect(self.export_to_excel)
        
        button_layout.addWidget(refresh_btn)
        button_layout.addWidget(export_btn)
        button_layout.addStretch()
        
        self.main_layout.addWidget(button_frame, 0)
    
    def _load_os_options(self):
        """Load OS (Group) names that have branches with Global tagging"""
        self.os_filter_selector.blockSignals(True)
        current = self.os_filter_selector.currentData()
        self.os_filter_selector.clear()
        self.os_filter_selector.addItem("All Groups", None)
        
        try:
            # Get OS names that have at least one branch with global_tag = 'GLOBAL'
            rows = db_manager.execute_query("""
                SELECT DISTINCT b.os_name 
                FROM branches b 
                WHERE b.os_name IS NOT NULL 
                  AND b.os_name != '' 
                  AND b.global_tag = 'GLOBAL'
                ORDER BY b.os_name
            """)
            if rows:
                for r in rows:
                    os_name = r['os_name'] if isinstance(r, dict) else r[0]
                    self.os_filter_selector.addItem(os_name, os_name)
            
            # Restore selection
            if current:
                idx = self.os_filter_selector.findData(current)
                if idx >= 0:
                    self.os_filter_selector.setCurrentIndex(idx)
        except Exception as e:
            print(f"Error loading OS options: {e}")
        finally:
            self.os_filter_selector.blockSignals(False)
    
    def populate_table(self):
        """Populate the tree with groups and their global branches"""
        self._is_loading = True
        self.tree.clear()
        
        date_start, date_end = self.date_range_widget.get_date_range()
        is_range = self.date_range_widget.is_range_mode()
        os_filter = self.os_filter_selector.currentData()
        reg_filter = self.reg_filter_selector.currentData() if hasattr(self, 'reg_filter_selector') else "all"
        
        try:
            # Build query for branches with Global tagging
            if is_range:
                select_cols = """
                    SELECT b.os_name AS group_name,
                           dr.branch,
                           SUM(COALESCE(dr.palawan_sendout_lotes_total, 0))         AS so_lotes,
                           SUM(COALESCE(dr.palawan_sendout_principal, 0))           AS so_capital,
                           SUM(COALESCE(dr.palawan_sendout_sc, 0))                  AS so_sc,
                           SUM(COALESCE(dr.palawan_sendout_commission, 0))          AS so_commission,
                           SUM(COALESCE(dr.palawan_sendout_regular_total, 0))       AS so_total,
                           SUM(COALESCE(dr.palawan_payout_lotes_total, 0))          AS po_lotes,
                           SUM(COALESCE(dr.palawan_payout_principal, 0))            AS po_capital,
                           SUM(COALESCE(dr.palawan_payout_sc, 0))                   AS po_sc,
                           SUM(COALESCE(dr.palawan_payout_commission, 0))           AS po_commission,
                           SUM(COALESCE(dr.palawan_payout_regular_total, 0))        AS po_total,
                           SUM(COALESCE(dr.palawan_international_lotes_total, 0))   AS int_lotes,
                           SUM(COALESCE(dr.palawan_international_principal, 0))     AS int_capital,
                           SUM(COALESCE(dr.palawan_international_sc, 0))            AS int_sc,
                           SUM(COALESCE(dr.palawan_international_commission, 0))    AS int_commission,
                           SUM(COALESCE(dr.palawan_international_regular_total, 0)) AS int_total,
                           SUM(COALESCE(dr.palawan_suki_discounts, 0))              AS skid,
                           SUM(COALESCE(dr.palawan_suki_rebates, 0))                AS skir,
                           SUM(COALESCE(dr.palawan_cancel, 0))                      AS cancellation,
                           SUM(COALESCE(dr.palawan_pay_out_incentives, 0))          AS inc
                    FROM {table} dr
                    INNER JOIN branches b ON dr.branch COLLATE utf8mb4_general_ci = b.name COLLATE utf8mb4_general_ci
                """.format(table=self.daily_table)
            else:
                select_cols = """
                    SELECT b.os_name AS group_name,
                           dr.branch,
                           COALESCE(dr.palawan_sendout_lotes_total, 0)         AS so_lotes,
                           COALESCE(dr.palawan_sendout_principal, 0)           AS so_capital,
                           COALESCE(dr.palawan_sendout_sc, 0)                  AS so_sc,
                           COALESCE(dr.palawan_sendout_commission, 0)          AS so_commission,
                           COALESCE(dr.palawan_sendout_regular_total, 0)       AS so_total,
                           COALESCE(dr.palawan_payout_lotes_total, 0)          AS po_lotes,
                           COALESCE(dr.palawan_payout_principal, 0)            AS po_capital,
                           COALESCE(dr.palawan_payout_sc, 0)                   AS po_sc,
                           COALESCE(dr.palawan_payout_commission, 0)           AS po_commission,
                           COALESCE(dr.palawan_payout_regular_total, 0)        AS po_total,
                           COALESCE(dr.palawan_international_lotes_total, 0)   AS int_lotes,
                           COALESCE(dr.palawan_international_principal, 0)     AS int_capital,
                           COALESCE(dr.palawan_international_sc, 0)            AS int_sc,
                           COALESCE(dr.palawan_international_commission, 0)    AS int_commission,
                           COALESCE(dr.palawan_international_regular_total, 0) AS int_total,
                           COALESCE(dr.palawan_suki_discounts, 0)              AS skid,
                           COALESCE(dr.palawan_suki_rebates, 0)                AS skir,
                           COALESCE(dr.palawan_cancel, 0)                      AS cancellation,
                           COALESCE(dr.palawan_pay_out_incentives, 0)          AS inc
                    FROM {table} dr
                    INNER JOIN branches b ON dr.branch COLLATE utf8mb4_general_ci = b.name COLLATE utf8mb4_general_ci
                """.format(table=self.daily_table)
            
            # Build WHERE clauses
            if is_range:
                where_parts = ["dr.date >= %s", "dr.date <= %s"]
                params = [date_start, date_end]
            else:
                where_parts = ["dr.date = %s"]
                params = [date_start]
            
            # Only branches with Global tagging
            where_parts.append("b.global_tag = 'GLOBAL'")
            
            # Optional OS filter
            if os_filter:
                where_parts.append("b.os_name = %s")
                params.append(os_filter)
            
            # Registration filter
            if reg_filter == "registered":
                where_parts.append("b.is_registered = 1")
            elif reg_filter == "not_registered":
                where_parts.append("(b.is_registered = 0 OR b.is_registered IS NULL)")
            # "all" - no filter needed
            
            group_by = " GROUP BY b.os_name, dr.branch" if is_range else ""
            order_by = " ORDER BY b.os_name, dr.branch"
            
            query = select_cols + " WHERE " + " AND ".join(where_parts) + group_by + order_by
            
            run_query_async(
                parent=self,
                query=query,
                params=tuple(params),
                on_result=self._on_populate_result,
                on_error=self._on_populate_error,
                loading_message="\u23f3  Loading global payable data\u2026",
            )

        except Exception as e:
            QMessageBox.critical(self, "Database Error", f"Error loading data: {str(e)}")
            self._is_loading = False

    def _on_populate_error(self, err):
        QMessageBox.critical(self, "Database Error", f"Error loading data: {err}")
        self._is_loading = False

    def _on_populate_result(self, results):
        if not results:
            self._is_loading = False
            return

        # Group results by OS name
        groups = {}
        for row in results:
            group_name = row['group_name'] or 'Unassigned'
            if group_name not in groups:
                groups[group_name] = []
            groups[group_name].append(row)

        # Grand totals
        grand_totals = [0.0] * 19

        # Create tree items for each group
        for group_name in sorted(groups.keys()):
            branches = groups[group_name]

            # Create group item
            group_item = QTreeWidgetItem(self.tree)
            group_item.setText(0, f"📁 {group_name}")
            group_item.setFont(0, QFont("Arial", 10, QFont.Bold))
            group_item.setBackground(0, QBrush(QColor("#e3f2fd")))

            # Group totals
            group_totals = [0.0] * 19

            # Add branch items under group
            for branch_data in branches:
                branch_item = QTreeWidgetItem(group_item)
                branch_item.setText(0, f"  🏢 {branch_data['branch']}")

                # Add values
                values = [
                    float(branch_data['so_lotes']), float(branch_data['so_capital']),
                    float(branch_data['so_sc']), float(branch_data['so_commission']),
                    float(branch_data['so_total']),
                    float(branch_data['po_lotes']), float(branch_data['po_capital']),
                    float(branch_data['po_sc']), float(branch_data['po_commission']),
                    float(branch_data['po_total']),
                    float(branch_data['int_lotes']), float(branch_data['int_capital']),
                    float(branch_data['int_sc']), float(branch_data['int_commission']),
                    float(branch_data['int_total']),
                    float(branch_data['skid']), float(branch_data['skir']),
                    float(branch_data['cancellation']), float(branch_data['inc'])
                ]

                for i, val in enumerate(values):
                    branch_item.setText(i + 1, f"{val:,.2f}")
                    branch_item.setTextAlignment(i + 1, Qt.AlignRight | Qt.AlignVCenter)
                    group_totals[i] += val
                    grand_totals[i] += val

            # Set group totals
            for i, total in enumerate(group_totals):
                group_item.setText(i + 1, f"{total:,.2f}")
                group_item.setTextAlignment(i + 1, Qt.AlignRight | Qt.AlignVCenter)
                group_item.setBackground(i + 1, QBrush(QColor("#e3f2fd")))
                group_item.setFont(i + 1, QFont("Arial", 9, QFont.Bold))

            # Expand group by default
            group_item.setExpanded(True)

        # Add grand total row
        total_item = QTreeWidgetItem(self.tree)
        total_item.setText(0, "📊 GRAND TOTAL")
        total_item.setFont(0, QFont("Arial", 11, QFont.Bold))
        total_item.setBackground(0, QBrush(QColor("#ffd700")))

        for i, total in enumerate(grand_totals):
            total_item.setText(i + 1, f"{total:,.2f}")
            total_item.setTextAlignment(i + 1, Qt.AlignRight | Qt.AlignVCenter)
            total_item.setBackground(i + 1, QBrush(QColor("#ffd700")))
            total_item.setFont(i + 1, QFont("Arial", 10, QFont.Bold))

        self._is_loading = False
    
    def export_to_excel(self):
        """Export the tree data to Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            QMessageBox.warning(self, "Missing Package", "Install openpyxl:\npip install openpyxl")
            return
        
        if self.tree.topLevelItemCount() == 0:
            QMessageBox.warning(self, "No Data", "Please load data first.")
            return
        
        date_start, date_end = self.date_range_widget.get_date_range()
        date_label = date_start if date_start == date_end else f"{date_start}_to_{date_end}"
        
        path, _ = QFileDialog.getSaveFileName(
            self, "Save Excel", f"Global_Payable_{date_label}.xlsx",
            "Excel Files (*.xlsx)"
        )
        if not path:
            return
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Global Payable"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill("solid", fgColor="495057")
        group_fill = PatternFill("solid", fgColor="E3F2FD")
        total_fill = PatternFill("solid", fgColor="FFD700")
        thin = Side(style='thin')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        
        # Title
        ws.cell(row=1, column=1, value="Global Payable Report").font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=f"Date: {date_label}").font = Font(bold=True, size=11)
        
        # Headers
        headers = [
            "Group / Branch",
            "SO:Lotes", "SO:Cap", "SO:SC", "SO:Com", "SO:TOTAL",
            "PO:Lotes", "PO:Cap", "PO:SC", "PO:Com", "PO:TOTAL",
            "IN:Lotes", "IN:Cap", "IN:SC", "IN:Com", "IN:TOTAL",
            "SKID", "SKIR", "CANCEL", "INC"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data rows
        row_num = 5
        for i in range(self.tree.topLevelItemCount()):
            group_item = self.tree.topLevelItem(i)
            
            # Group row
            for col in range(20):
                cell = ws.cell(row=row_num, column=col + 1, value=group_item.text(col))
                cell.border = border
                cell.fill = group_fill if "GRAND TOTAL" not in group_item.text(0) else total_fill
                if col > 0:
                    cell.alignment = Alignment(horizontal='right')
                    try:
                        cell.value = float(group_item.text(col).replace(",", ""))
                        cell.number_format = '#,##0.00'
                    except:
                        pass
            row_num += 1
            
            # Child rows (branches)
            for j in range(group_item.childCount()):
                branch_item = group_item.child(j)
                for col in range(20):
                    cell = ws.cell(row=row_num, column=col + 1, value=branch_item.text(col))
                    cell.border = border
                    if col > 0:
                        cell.alignment = Alignment(horizontal='right')
                        try:
                            cell.value = float(branch_item.text(col).replace(",", ""))
                            cell.number_format = '#,##0.00'
                        except:
                            pass
                row_num += 1
        
        # Auto-width
        for col_cells in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 20)
        
        wb.save(path)
        QMessageBox.information(self, "Export Successful", f"Saved to:\n{path}")
