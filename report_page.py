from PyQt5.QtWidgets import (
    QWidget, QLabel, QVBoxLayout, QComboBox, QTableWidget, QTableWidgetItem,
    QDateEdit, QMessageBox, QHeaderView, QSizePolicy, QPushButton, QHBoxLayout,
    QApplication, QDesktopWidget, QScrollArea, QFrame, QLineEdit, QFileDialog
)
from PyQt5.QtCore import Qt, QDate, QTimer
from PyQt5.QtGui import QFont, QColor, QBrush, QTextDocument
from PyQt5.QtPrintSupport import QPrinter, QPrintDialog
from db_connect_pooled import db_manager
from date_range_widget import DateRangeWidget
from decimal import Decimal, ROUND_HALF_UP
import datetime

def _round2(value):
    """Round to 2 decimal places using round-half-up (traditional rounding)."""
    return float(Decimal(str(value)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))

def _mult_round2(a, b):
    """Multiply two numbers using Decimal arithmetic and round to 2 decimal places (ROUND_HALF_UP).
    This avoids floating point precision issues like 2443.50 * 0.61 = 1490.5349999... instead of 1490.535.
    """
    result = Decimal(str(a)) * Decimal(str(b))
    return float(result.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))


class ReportPage(QWidget):
    def __init__(self, account_type=2):
        super().__init__()
        self.account_type = account_type
        # Set correct tables based on brand
        self.daily_table = "daily_reports_brand_a" if account_type == 1 else "daily_reports"
        self.payable_table = "payable_tbl_brand_a" if account_type == 1 else "payable_tbl"
        self.setWindowTitle("PEPP Reconciliation Report")
        self.setup_window_size()

        self.main_layout = QVBoxLayout()
        self.main_layout.setContentsMargins(15, 15, 15, 15)
        self.main_layout.setSpacing(10)
        self.setLayout(self.main_layout)

        self.create_controls()
        self.create_report_table()
        # Buttons moved into the top controls so they're visible on small screens

        self.load_corporations()

    # ─────────────────────────────────────────────────────────────────────────

    def setup_window_size(self):
        desktop = QApplication.desktop()
        self.setMinimumSize(800, 600)

    # ─────────────────────────────────────────────────────────────────────────

    def create_controls(self):
        controls_frame = QFrame()
        controls_frame.setFrameStyle(QFrame.Box)
        controls_frame.setStyleSheet("""
            QFrame {
                background-color: #f8f9fa;
                border: 1px solid #dee2e6;
                border-radius: 5px;
                padding: 10px;
            }
        """)
        controls_layout = QHBoxLayout(controls_frame)
        controls_layout.setSpacing(20)

        corp_label = QLabel("Corporation:")
        corp_label.setFont(QFont("Arial", 10, QFont.Bold))
        self.corp_selector = QComboBox()
        self.corp_selector.setMinimumWidth(200)
        self.corp_selector.currentTextChanged.connect(self._on_corp_changed)

        date_label = QLabel("Date:")
        date_label.setFont(QFont("Arial", 10, QFont.Bold))
        self.date_range_widget = DateRangeWidget()
        self.date_range_widget.dateRangeChanged.connect(self.generate_report)
        # Backward-compat alias
        self.date_selector = self.date_range_widget

        registry_label = QLabel("Partner Registry No:")
        registry_label.setFont(QFont("Arial", 10, QFont.Bold))
        self.registry_input = QLineEdit()
        self.registry_input.setPlaceholderText("e.g., P250683A")
        self.registry_input.setMaximumWidth(150)
        self.registry_input.setReadOnly(True)
        # Re-generate report when registry number changes so header updates live
        self.registry_input.textChanged.connect(self.generate_report)

        # Registration status filter
        reg_filter_label = QLabel("Branch Status:")
        reg_filter_label.setFont(QFont("Arial", 10, QFont.Bold))
        self.reg_filter_selector = QComboBox()
        self.reg_filter_selector.setMinimumWidth(150)
        self.reg_filter_selector.addItem("Registered Only", "registered")
        self.reg_filter_selector.currentIndexChanged.connect(self.generate_report)

        controls_layout.addWidget(corp_label)
        controls_layout.addWidget(self.corp_selector)
        controls_layout.addSpacing(30)
        controls_layout.addWidget(date_label)
        controls_layout.addWidget(self.date_range_widget)
        controls_layout.addSpacing(30)
        controls_layout.addWidget(registry_label)
        controls_layout.addWidget(self.registry_input)
        controls_layout.addSpacing(30)
        controls_layout.addWidget(reg_filter_label)
        controls_layout.addWidget(self.reg_filter_selector)
        # Export / Print buttons moved here so they remain visible on small screens
        # Use a compact style to avoid overflowing when window is narrow
        button_style = """
            QPushButton {
                padding: 8px 16px;
                border: 2px solid #007bff;
                border-radius: 6px;
                background-color: #007bff;
                color: white;
                font-weight: bold;
                font-size: 11px;
                min-width: 90px;
            }
            QPushButton:hover { background-color: #0056b3; }
        """

        self.export_button = QPushButton("Export to Excel")
        self.export_button.setStyleSheet(button_style)
        self.export_button.clicked.connect(self.export_to_excel)
        self.export_button.setSizePolicy(QSizePolicy.Minimum, QSizePolicy.Fixed)

        self.print_button = QPushButton("Print Report")
        self.print_button.setStyleSheet(button_style)
        self.print_button.clicked.connect(self.print_report)
        self.print_button.setSizePolicy(QSizePolicy.Minimum, QSizePolicy.Fixed)

        controls_layout.addSpacing(20)
        controls_layout.addWidget(self.export_button)
        controls_layout.addSpacing(8)
        controls_layout.addWidget(self.print_button)
        controls_layout.addStretch()

        self.main_layout.addWidget(controls_frame)

    # ─────────────────────────────────────────────────────────────────────────

    def create_report_table(self):
        self.table = QTableWidget()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(["", "", "", ""])
        self.table.horizontalHeader().hide()
        self.table.verticalHeader().hide()
        self.table.setColumnWidth(0, 400)
        self.table.setColumnWidth(1, 50)
        self.table.setColumnWidth(2, 100)
        self.table.setColumnWidth(3, 150)
        self.table.setStyleSheet("""
            QTableWidget {
                gridline-color: #d0d0d0;
                border: 1px solid #c0c0c0;
                background-color: white;
                font-size: 11px;
            }
            QTableWidget::item {
                border: 1px solid #e0e0e0;
                padding: 5px;
            }
        """)
        self.main_layout.addWidget(self.table)

    # ─────────────────────────────────────────────────────────────────────────

    def create_buttons(self):
        button_frame = QFrame()
        button_frame.setMaximumHeight(60)
        button_layout = QHBoxLayout(button_frame)

        button_style = """
            QPushButton {
                padding: 12px 24px;
                border: 2px solid #007bff;
                border-radius: 6px;
                background-color: #007bff;
                color: white;
                font-weight: bold;
                font-size: 11px;
                min-width: 120px;
            }
            QPushButton:hover { background-color: #0056b3; }
        """

        self.export_button = QPushButton("Export to Excel")
        self.export_button.clicked.connect(self.export_to_excel)
        self.export_button.setStyleSheet(button_style)

        self.print_button = QPushButton("Print Report")
        self.print_button.clicked.connect(self.print_report)
        self.print_button.setStyleSheet(button_style)

        button_layout.addStretch()
        button_layout.addWidget(self.export_button)
        button_layout.addSpacing(15)
        button_layout.addWidget(self.print_button)
        button_layout.addStretch()

        self.main_layout.addWidget(button_frame)

    # ─────────────────────────────────────────────────────────────────────────

    # ── Corporation → Partner Registry No. mapping ─────────────────────────
    REGISTRY_MAP = {
        'SILVERSTAR JEWELRY PAWNSHOP INC':                  'P250682A',
        'ALEXITE JEWELRY PAWNSHOP INC':                     'P250683A',
        'SAN RAMON PLATINUM PAWNSHOP INC':                  'P250681A',
        'HOMENEEDS PAWNSHOP INC':                           'P250677A',
        'KRISTAL CLEAR DIAMOND AND GOLD PAWNSHOP INC':      'P250678A',
        'SAFELOCK PAWNSHOP INC':                            'P250680A',
        'MEGAWORLD DOMESTIC PAWNSHOP INC':                  'P250679A',
        'GLOBAL RELIANCE MANAGEMENT & HOLDINGS CORP.':      'P210021A',
    }

    def _on_corp_changed(self, corp_name):
        """Auto-fill partner registry number when corporation changes."""
        key = (corp_name or '').upper().strip().rstrip('.')
        registry = ''
        for k, v in self.REGISTRY_MAP.items():
            if k.upper().strip().rstrip('.') == key:
                registry = v
                break
        self.registry_input.setText(registry)
        self.generate_report()

    ALLOWED_CORPS = [
        'SILVERSTAR JEWELRY AND PAWNSHOP INC.',
        'ALLEXITE JEWELRY PAWNSHOP INC.',
        'SAN RAMON PLATINUM PAWNSHOP INC.',
        'HOMENEEDS PAWNSHOP INC.',
        'KRISTAL CLEAR DIAMOND & GOLD PAWNSHOP INC.',
        'SAFELOCK PAWNSHOP INC.',
        'MEGAWORLD DOMESTIC PAWNSHOP INC',
    ]

    def load_corporations(self):
        self.corp_selector.clear()
        allowed = {c.upper().strip().rstrip('.') for c in self.ALLOWED_CORPS}
        try:
            rows = db_manager.execute_query(
                "SELECT name FROM corporations ORDER BY name"
            )
            if rows:
                for row in rows:
                    name = row.get('name') or row[0]
                    if name.upper().strip().rstrip('.') in allowed:
                        self.corp_selector.addItem(name)
        except Exception as e:
            print(f"Error loading corporations: {e}")

        self.corp_selector.addItem('Global Reliance Management & Holdings Corp.')

    # ─────────────────────────────────────────────────────────────────────────

    def get_totals_from_database(self):
    
        corp          = self.corp_selector.currentText()
        date_start, date_end = self.date_range_widget.get_date_range()
        reg_filter    = self.reg_filter_selector.currentData()

        if not corp:
            return None

       
        is_global_reliance = 'GLOBAL RELIANCE' in corp.upper()

        try:
            # Determine date clause
            is_range = self.date_range_widget.is_range_mode()
            if is_range:
                date_clause = "dr.date >= %s AND dr.date <= %s"
                date_params = (date_start, date_end)
            else:
                date_clause = "dr.date = %s"
                date_params = (date_start,)

            # Build corp filter clause - Global Reliance uses global_tag, others use corporation name
            if is_global_reliance:
                corp_clause = "b.global_tag = 'GLOBAL'"
                corp_params = ()
            else:
                corp_clause = "dr.corporation = %s"
                corp_params = (corp,)

            # Build query based on registration filter
            if reg_filter == "registered":
                result = db_manager.execute_query(f"""
                    SELECT SUM(dr.palawan_sendout_principal)      AS total_sendout_capital,
                           SUM(dr.palawan_sendout_commission)     AS total_sendout_commission,
                           SUM(dr.palawan_sendout_sc)             AS total_sendout_sc,
                           SUM(dr.palawan_payout_principal)       AS total_payout_capital,
                           SUM(dr.palawan_payout_commission)      AS total_payout_commission,
                           SUM(dr.palawan_payout_sc)              AS total_payout_sc,
                           SUM(dr.palawan_international_commission) AS total_international_commission,
                           SUM(dr.palawan_suki_discounts)         AS total_skid,
                           SUM(dr.palawan_suki_rebates)           AS total_skir,
                           SUM(dr.palawan_cancel)                 AS total_cancellation,
                           SUM(COALESCE(p.inc, 0))                AS total_inc
                    FROM {self.daily_table} dr
                    INNER JOIN branches b ON dr.branch COLLATE utf8mb4_general_ci = b.name COLLATE utf8mb4_general_ci
                    INNER JOIN corporations c ON b.corporation_id = c.id AND dr.corporation COLLATE utf8mb4_general_ci = c.name COLLATE utf8mb4_general_ci
                    LEFT JOIN {self.payable_table} p ON dr.corporation COLLATE utf8mb4_general_ci = p.corporation COLLATE utf8mb4_general_ci AND dr.branch COLLATE utf8mb4_general_ci = p.branch COLLATE utf8mb4_general_ci AND dr.date = p.date
                    WHERE {corp_clause}
                      AND {date_clause}
                      AND b.is_registered = 1
                """, corp_params + date_params)
            elif reg_filter == "not_registered":
                result = db_manager.execute_query(f"""
                    SELECT SUM(dr.palawan_sendout_principal)      AS total_sendout_capital,
                           SUM(dr.palawan_sendout_commission)     AS total_sendout_commission,
                           SUM(dr.palawan_sendout_sc)             AS total_sendout_sc,
                           SUM(dr.palawan_payout_principal)       AS total_payout_capital,
                           SUM(dr.palawan_payout_commission)      AS total_payout_commission,
                           SUM(dr.palawan_payout_sc)              AS total_payout_sc,
                           SUM(dr.palawan_international_commission) AS total_international_commission,
                           SUM(dr.palawan_suki_discounts)         AS total_skid,
                           SUM(dr.palawan_suki_rebates)           AS total_skir,
                           SUM(dr.palawan_cancel)                 AS total_cancellation,
                           SUM(COALESCE(p.inc, 0))                AS total_inc
                    FROM {self.daily_table} dr
                    INNER JOIN branches b ON dr.branch COLLATE utf8mb4_general_ci = b.name COLLATE utf8mb4_general_ci
                    INNER JOIN corporations c ON b.corporation_id = c.id AND dr.corporation COLLATE utf8mb4_general_ci = c.name COLLATE utf8mb4_general_ci
                    LEFT JOIN {self.payable_table} p ON dr.corporation COLLATE utf8mb4_general_ci = p.corporation COLLATE utf8mb4_general_ci AND dr.branch COLLATE utf8mb4_general_ci = p.branch COLLATE utf8mb4_general_ci AND dr.date = p.date
                    WHERE {corp_clause}
                      AND {date_clause}
                      AND b.is_registered = 0
                """, corp_params + date_params)
            elif is_global_reliance:
                # Global Reliance with "All" registration filter - needs branches join for global_tag
                result = db_manager.execute_query(f"""
                    SELECT SUM(dr.palawan_sendout_principal)      AS total_sendout_capital,
                           SUM(dr.palawan_sendout_commission)     AS total_sendout_commission,
                           SUM(dr.palawan_sendout_sc)             AS total_sendout_sc,
                           SUM(dr.palawan_payout_principal)       AS total_payout_capital,
                           SUM(dr.palawan_payout_commission)      AS total_payout_commission,
                           SUM(dr.palawan_payout_sc)              AS total_payout_sc,
                           SUM(dr.palawan_international_commission) AS total_international_commission,
                           SUM(dr.palawan_suki_discounts)         AS total_skid,
                           SUM(dr.palawan_suki_rebates)           AS total_skir,
                           SUM(dr.palawan_cancel)                 AS total_cancellation,
                           SUM(COALESCE(p.inc, 0))                AS total_inc
                    FROM {self.daily_table} dr
                    INNER JOIN branches b ON dr.branch COLLATE utf8mb4_general_ci = b.name COLLATE utf8mb4_general_ci
                    LEFT JOIN {self.payable_table} p ON dr.corporation COLLATE utf8mb4_general_ci = p.corporation COLLATE utf8mb4_general_ci AND dr.branch COLLATE utf8mb4_general_ci = p.branch COLLATE utf8mb4_general_ci AND dr.date = p.date
                    WHERE {corp_clause}
                      AND {date_clause}
                """, corp_params + date_params)
            else:
                result = db_manager.execute_query(f"""
                    SELECT SUM(dr.palawan_sendout_principal)      AS total_sendout_capital,
                           SUM(dr.palawan_sendout_commission)     AS total_sendout_commission,
                           SUM(dr.palawan_sendout_sc)             AS total_sendout_sc,
                           SUM(dr.palawan_payout_principal)       AS total_payout_capital,
                           SUM(dr.palawan_payout_commission)      AS total_payout_commission,
                           SUM(dr.palawan_payout_sc)              AS total_payout_sc,
                           SUM(dr.palawan_international_commission) AS total_international_commission,
                           SUM(dr.palawan_suki_discounts)         AS total_skid,
                           SUM(dr.palawan_suki_rebates)           AS total_skir,
                           SUM(dr.palawan_cancel)                 AS total_cancellation,
                           SUM(COALESCE(p.inc, 0))                AS total_inc
                    FROM {self.daily_table} dr
                    LEFT JOIN {self.payable_table} p ON dr.corporation COLLATE utf8mb4_general_ci = p.corporation COLLATE utf8mb4_general_ci AND dr.branch COLLATE utf8mb4_general_ci = p.branch COLLATE utf8mb4_general_ci AND dr.date = p.date
                    WHERE {corp_clause}
                      AND {date_clause}
                """, corp_params + date_params)

            if not result or not result[0]:
                return None

            row = result[0]

            # If every SUM came back NULL it means no matching rows at all
            if isinstance(row, dict):
                if all(v is None for v in row.values()):
                    return None
                return row
            else:
                if all(v is None for v in row):
                    return None
                keys = [
                    'total_sendout_capital', 'total_sendout_commission', 'total_sendout_sc',
                    'total_payout_capital',  'total_payout_commission',  'total_payout_sc',
                    'total_international_commission',
                    'total_skid', 'total_skir', 'total_cancellation', 'total_inc',
                ]
                return dict(zip(keys, row))

        except Exception as e:
            QMessageBox.critical(self, "Database Error", f"Error retrieving totals: {str(e)}")
            return None

    # ─────────────────────────────────────────────────────────────────────────

    def generate_report(self):
        """
        Auto-generates the report whenever corp/date/registry changes.
        If no data exists yet, clears the table silently — no popup.
        """
        corp          = self.corp_selector.currentText()
        date_start, date_end = self.date_range_widget.get_date_range()
        date_display = date_start if date_start == date_end else f"{date_start} to {date_end}"

        if not corp:
            self.table.setRowCount(0)
            return

  
        corp_abbrev_map = {
            'SILVERSTAR JEWELRY PAWNSHOP INC': 'SJPI',
            'ALEXITE JEWELRY PAWNSHOP INC': 'AJPI',
            'SAN RAMON PLATINUM PAWNSHOP INC': 'SRPPI',
            'HOMENEEDS PAWNSHOP INC': 'HPI',
            'KRISTAL CLEAR DIAMOND AND GOLD PAWNSHOP INC': 'KCDGPI',
            'SAFELOCK PAWNSHOP INC': 'SPI',
            'MEGAWORLD DOMESTIC PAWNSHOP INC': 'MDPI',
            'GLOBAL RELIANCE MANAGEMENT & HOLDINGS CORP.': 'GRMHC',
        }
        corp_upper = corp.upper().strip().rstrip('.')
        corp_abbrev = next((v for k, v in corp_abbrev_map.items() if k.upper().strip().rstrip('.') == corp_upper), 'AJPI')

        totals = self.get_totals_from_database()

        # ── No data yet: clear table silently and exit ────────────────────────
        if not totals:
            self.table.setRowCount(0)
            return

        # ── Extract values ────────────────────────────────────────────────────
        sendout_capital          = float(totals.get('total_sendout_capital')          or 0)
        sendout_commission       = float(totals.get('total_sendout_commission')       or 0)
        sendout_sc               = float(totals.get('total_sendout_sc')               or 0)
        payout_capital           = float(totals.get('total_payout_capital')           or 0)
        payout_commission        = float(totals.get('total_payout_commission')        or 0)
        payout_sc                = float(totals.get('total_payout_sc')                or 0)
        international_commission = float(totals.get('total_international_commission') or 0)
        total_skid               = float(totals.get('total_skid')                     or 0)
        total_skir               = float(totals.get('total_skir')                     or 0)
        total_cancellation       = float(totals.get('total_cancellation')             or 0)
        total_inc                = float(totals.get('total_inc')                      or 0)

        # ── Derived calculations (round-half-up to 2dp) ───────────────────
        # Use _mult_round2 to avoid floating point precision issues
        pepp_commission_61       = _mult_round2(sendout_commission, 0.61)
        skid_61                  = _mult_round2(total_skid, 0.61)
        corp_commission_43       = _mult_round2(payout_commission, 0.43)
        corp_international_80    = _mult_round2(international_commission, 0.80)
        skir_57                  = _mult_round2(total_skir, 0.57)

        send_subtotal                = _round2(sendout_capital + pepp_commission_61 + sendout_sc)
        send_subtotal_after_discount = _round2(send_subtotal - skid_61)
        total_net_send               = _round2(send_subtotal_after_discount - total_cancellation)

        release_subtotal             = _round2(payout_capital + corp_commission_43 + corp_international_80)
        release_subtotal_with_inc    = _round2(release_subtotal + total_inc)
        total_net_released           = _round2(release_subtotal_with_inc + skir_57)

        net_receivable_payable       = _round2(total_net_send - total_net_released)

        registry = self.registry_input.text().strip()

        # ── Build report rows ─────────────────────────────────────────────────
        report_data = [
            # Header
            ("Palawan Express Pera Padala - " + corp, "",  "",            "",                      "header"),
            ("PEPP Reconciliation for",               "",  "Partner Registry No.", "",             "header"),
            (date_display,                           "",  registry,      "",                      "header"),
            ("", "", "", "", "blank"),
            ("", "", "", "", "blank"),

            # Send Transaction
            ("Send Transaction",                                           "", "",                          "",                              "section"),
            (f"    PEPP Remittance from {corp}",                           "", "P",                         f"{sendout_capital:,.2f}",        "indent"),
            ("    PEPP share: 61% of commission",                          "P", f"{sendout_commission:,.2f}", f"{pepp_commission_61:,.2f}",   "indent"),
            ("    PEPP share: Service Charge",                             "", f"{sendout_sc:,.2f}",         f"{sendout_sc:,.2f}",            "indent"),
            ("        Subtotal",                                           "", "",                          f"{send_subtotal:,.2f}",          "subtotal"),
            ("    Less: Discount (Suki Card)",                             "", f"({total_skid:,.2f})",       f"({skid_61:,.2f})",             "indent"),
            ("        Subtotal",                                           "", "",                          f"{send_subtotal_after_discount:,.2f}", "subtotal"),
            ("    Less: Cancellation",                                     "", f"({total_cancellation:,.2f})", f"({total_cancellation:,.2f})", "indent"),
            ("    Total Net Send",                                         "", "",                          f"{total_net_send:,.2f}",         "total"),
            ("", "", "", "", "blank"),
            ("", "", "", "", "blank"),

            # Release Transaction
            (f"    RELEASE Transaction (Payable to {corp_abbrev})",           "", "",                          "",                              "section"),
            (f"    PEPP Remittances released at {corp_abbrev}",               "", "P",                         f"{payout_capital:,.2f}",         "indent"),
            (f"    {corp_abbrev} share: 43% of commission",                   "P", f"{payout_commission:,.2f}", f"{corp_commission_43:,.2f}",    "indent"),
            (f"    {corp_abbrev} share: 50% of commission (LBC Domestic Payout)",    "", "",                          "",                              "indent"),
            (f"    {corp_abbrev} share: 80% of commission (International Payout)",   "", f"{international_commission:,.2f}", f"{corp_international_80:,.2f}", "indent"),
            ("    Service Charge",                                         "", f"{payout_sc:,.2f}",          "-",                             "indent"),
            ("        Subtotal",                                           "", "",                          f"{release_subtotal:,.2f}",       "subtotal"),
            (f"    Add: {corp_abbrev} Branch Incentives released",       "", "",                        f"{total_inc:,.2f}",              "indent"),
            ("        Subtotal",                                           "", "",                          f"{release_subtotal_with_inc:,.2f}", "subtotal"),
            ("    Add: Rebates (Suki Card)",                               "", f"{total_skir:,.2f}",         f"{skir_57:,.2f}",               "indent"),
            ("    Total Net Released",                                     "", "",                          f"{total_net_released:,.2f}",     "total"),
            ("", "", "", "", "blank"),
            ("", "", "", "", "blank"),

            # Summary
            ("    Net Send",                    "", "", f"{total_net_send:,.2f}",         "regular"),
            ("    Less : Net Released",         "", "", f"{total_net_released:,.2f}",     "regular"),
            ("    Net Receivable / (Payable)",  "", "", f"{net_receivable_payable:,.2f}", "total"),
        ]

        # ── Populate table ────────────────────────────────────────────────────
        self.table.setRowCount(len(report_data))

        bold_font    = QFont(); bold_font.setBold(True)
        regular_font = QFont()

        for row, (col1, col2, col3, col4, row_type) in enumerate(report_data):
            for col, text in enumerate([col1, col2, col3, col4]):
                item = QTableWidgetItem(text)

                # Font
                if row_type in ("header", "section", "total", "subtotal"):
                    item.setFont(bold_font)
                else:
                    item.setFont(regular_font)

                # Background
                if row_type == "total":
                    item.setBackground(QBrush(QColor("#f0f0f0")))
                elif row_type == "subtotal":
                    item.setBackground(QBrush(QColor("#f8f8f8")))

                # Alignment
                if col in (2, 3):
                    item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                else:
                    item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)

                item.setFlags(item.flags() & ~Qt.ItemIsEditable)  # read-only
                self.table.setItem(row, col, item)

            self.table.setRowHeight(row, 25)

        # Center the title row
        title_item = self.table.item(0, 0)
        if title_item:
            title_item.setTextAlignment(Qt.AlignCenter)

    # ─────────────────────────────────────────────────────────────────────────

    def export_to_excel(self):
        if self.table.rowCount() == 0:
            QMessageBox.warning(self, "No Data", "No report to export. Select a corporation and date first.")
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            corp          = self.corp_selector.currentText()
            date_start, date_end = self.date_range_widget.get_date_range()
            date          = date_start if date_start == date_end else f"{date_start}_to_{date_end}"
            date_display  = date_start if date_start == date_end else f"{date_start} to {date_end}"
            registry      = self.registry_input.text().strip()

            # Corporation abbreviation mapping
            corp_abbrev_map = {
                'SILVERSTAR JEWELRY PAWNSHOP INC': 'SJPI',
                'ALEXITE JEWELRY PAWNSHOP INC': 'AJPI',
                'SAN RAMON PLATINUM PAWNSHOP INC': 'SRPPI',
                'HOMENEEDS PAWNSHOP INC': 'HPI',
                'KRISTAL CLEAR DIAMOND AND GOLD PAWNSHOP INC': 'KCDGPI',
                'SAFELOCK PAWNSHOP INC': 'SPI',
                'MEGAWORLD DOMESTIC PAWNSHOP INC': 'MDPI',
                'GLOBAL RELIANCE MANAGEMENT & HOLDINGS CORP.': 'GRMHC',
            }
            corp_upper = corp.upper().strip().rstrip('.')
            corp_abbrev = next((v for k, v in corp_abbrev_map.items() if k.upper().strip().rstrip('.') == corp_upper), 'AJPI')

            totals = self.get_totals_from_database()
            if not totals:
                QMessageBox.warning(self, "No Data", "No data available for export.")
                return

            sendout_capital          = float(totals.get('total_sendout_capital')          or 0)
            sendout_commission       = float(totals.get('total_sendout_commission')       or 0)
            sendout_sc               = float(totals.get('total_sendout_sc')               or 0)
            payout_capital           = float(totals.get('total_payout_capital')           or 0)
            payout_commission        = float(totals.get('total_payout_commission')        or 0)
            payout_sc                = float(totals.get('total_payout_sc')                or 0)
            international_commission = float(totals.get('total_international_commission') or 0)
            total_skid               = float(totals.get('total_skid')                     or 0)
            total_skir               = float(totals.get('total_skir')                     or 0)
            total_cancellation       = float(totals.get('total_cancellation')             or 0)
            total_inc                = float(totals.get('total_inc')                      or 0)

            pepp_commission_61       = _mult_round2(sendout_commission, 0.61)
            skid_61                  = _mult_round2(total_skid, 0.61)
            corp_commission_43       = _mult_round2(payout_commission, 0.43)
            corp_international_80    = _mult_round2(international_commission, 0.80)
            skir_57                  = _mult_round2(total_skir, 0.57)

            send_subtotal                = _round2(sendout_capital + pepp_commission_61 + sendout_sc)
            send_subtotal_after_discount = _round2(send_subtotal - skid_61)
            total_net_send               = _round2(send_subtotal_after_discount - total_cancellation)
            release_subtotal             = _round2(payout_capital + corp_commission_43 + corp_international_80)
            release_subtotal_with_inc    = _round2(release_subtotal + total_inc)
            total_net_released           = _round2(release_subtotal_with_inc + skir_57)
            net_receivable_payable       = _round2(total_net_send - total_net_released)

            wb = Workbook()
            ws = wb.active
            ws.title = "PEPP Reconciliation"

            header_font   = Font(bold=True, size=12)
            section_font  = Font(bold=True, size=11)
            regular_font  = Font(size=10)
            total_font    = Font(bold=True, size=11)
            header_fill   = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
            subtotal_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            total_fill    = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            right_align   = Alignment(horizontal='right')
            center_align  = Alignment(horizontal='center')

            r = 1
            ws.merge_cells(f'A{r}:D{r}')
            ws[f'A{r}'] = f"Palawan Express Pera Padala - {corp}"
            ws[f'A{r}'].font = header_font
            ws[f'A{r}'].alignment = center_align
            ws[f'A{r}'].fill = header_fill
            r += 1

            ws[f'A{r}'] = f"PEPP Reconciliation for {date_display}"
            ws[f'C{r}'] = "Partner Registry No."
            ws[f'D{r}'] = registry
            for cell in [ws[f'A{r}'], ws[f'C{r}'], ws[f'D{r}']]:
                cell.font = header_font
            r += 2

            rows = [
                ("Send Transaction",                                           "", "",                              "",                                   "section"),
                (f"    PEPP Remittance from {corp}",                           "", "P",                             f"{sendout_capital:,.2f}",             "indent"),
                ("    PEPP share: 61% of commission",                          "P", f"{sendout_commission:,.2f}",   f"{pepp_commission_61:,.2f}",          "indent"),
                ("    PEPP share: Service Charge",                             "", f"{sendout_sc:,.2f}",            f"{sendout_sc:,.2f}",                  "indent"),
                ("        Subtotal",                                           "", "",                              f"{send_subtotal:,.2f}",               "subtotal"),
                ("    Less: Discount (Suki Card)",                             "", f"({total_skid:,.2f})",          f"({skid_61:,.2f})",                   "indent"),
                ("        Subtotal",                                           "", "",                              f"{send_subtotal_after_discount:,.2f}","subtotal"),
                ("    Less: Cancellation",                                     "", f"({total_cancellation:,.2f})",  f"({total_cancellation:,.2f})",        "indent"),
                ("    Total Net Send",                                         "", "",                              f"{total_net_send:,.2f}",              "total"),
                ("", "", "", "", "blank"),
                (f"    RELEASE Transaction (Payable to {corp_abbrev})",        "", "",                              "",                                   "section"),
                (f"    PEPP Remittances released at {corp_abbrev}",            "", "P",                             f"{payout_capital:,.2f}",              "indent"),
                (f"    {corp_abbrev} share: 43% of commission",                "P", f"{payout_commission:,.2f}",    f"{corp_commission_43:,.2f}",          "indent"),
                (f"    {corp_abbrev} share: 50% of commission (LBC Domestic Payout)",    "", "",                              "",                                   "indent"),
                (f"    {corp_abbrev} share: 80% of commission (International Payout)",   "", f"{international_commission:,.2f}", f"{corp_international_80:,.2f}",   "indent"),
                ("    Service Charge",                                         "", f"{payout_sc:,.2f}",             "-",                                  "indent"),
                ("        Subtotal",                                           "", "",                              f"{release_subtotal:,.2f}",            "subtotal"),
                (f"    Add: {corp_abbrev} Branch Incentives released",        "", "",                              f"{total_inc:,.2f}",                   "indent"),
                ("        Subtotal",                                           "", "",                              f"{release_subtotal_with_inc:,.2f}",   "subtotal"),
                ("    Add: Rebates (Suki Card)",                               "", f"{total_skir:,.2f}",            f"{skir_57:,.2f}",                    "indent"),
                ("    Total Net Released",                                     "", "",                              f"{total_net_released:,.2f}",          "total"),
                ("", "", "", "", "blank"),
                ("    Net Send",                   "", "", f"{total_net_send:,.2f}",         "regular"),
                ("    Less : Net Released",        "", "", f"{total_net_released:,.2f}",     "regular"),
                ("    Net Receivable / (Payable)", "", "", f"{net_receivable_payable:,.2f}", "total"),
            ]

            for col1, col2, col3, col4, row_type in rows:
                ws[f'A{r}'] = col1
                ws[f'B{r}'] = col2
                ws[f'C{r}'] = col3
                ws[f'D{r}'] = col4

                if row_type == "section":
                    for c in ['A','B','C','D']: ws[f'{c}{r}'].font = section_font
                elif row_type == "total":
                    for c in ['A','B','C','D']:
                        ws[f'{c}{r}'].font = total_font
                        ws[f'{c}{r}'].fill = total_fill
                elif row_type == "subtotal":
                    for c in ['A','B','C','D']:
                        ws[f'{c}{r}'].font = total_font
                        ws[f'{c}{r}'].fill = subtotal_fill
                else:
                    for c in ['A','B','C','D']: ws[f'{c}{r}'].font = regular_font

                ws[f'C{r}'].alignment = right_align
                ws[f'D{r}'].alignment = right_align
                r += 1

            # Signature
            r += 2
            ws[f'A{r}'] = "Prepared by:"
            ws[f'C{r}'] = "Noted by:"
            r += 2
            ws[f'A{r}'] = "Rochelle G. Serrano"
            ws[f'C{r}'] = "Aimee M. Martinez"

            ws.column_dimensions['A'].width = 50
            ws.column_dimensions['B'].width = 8
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 18

            # Generate default filename
            ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            default_filename = f"PEPP_Reconciliation_{corp}_{date}_{ts}.xlsx"
            
            # Show save dialog
            filename, _ = QFileDialog.getSaveFileName(
                self,
                "Save Excel Report",
                default_filename,
                "Excel Files (*.xlsx);;All Files (*)"
            )
            
            if not filename:
                # User cancelled
                return
            
            # Ensure .xlsx extension
            if not filename.lower().endswith('.xlsx'):
                filename += '.xlsx'
            
            wb.save(filename)
            QMessageBox.information(self, "Export Successful", f"Saved as:\n{filename}")

        except ImportError:
            QMessageBox.critical(self, "Missing Library",
                                 "openpyxl is required.\nInstall with: pip install openpyxl")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Error exporting: {str(e)}")

    # ─────────────────────────────────────────────────────────────────────────

    def print_report(self):
        if self.table.rowCount() == 0:
            QMessageBox.warning(self, "No Data", "No report to print. Select a corporation and date first.")
            return

        try:
            doc  = QTextDocument()
            html = "<html><body style='font-family:Arial; font-size:12px;'>"

            for row in range(self.table.rowCount()):
                parts = []
                for col in range(4):
                    item = self.table.item(row, col)
                    if item and item.text().strip():
                        parts.append(item.text() if col == 0 else f"&nbsp;&nbsp;&nbsp;{item.text()}")

                line = "".join(parts)
                if line.strip():
                    bold = any(kw in line for kw in [
                        "Palawan Express", "PEPP Reconciliation", "Transaction",
                        "Total Net", "Net Receivable", "Subtotal",
                    ])
                    tag = "b" if bold else "span"
                    html += f"<p><{tag}>{line}</{tag}></p>"
                else:
                    html += "<p>&nbsp;</p>"

            html += "</body></html>"
            doc.setHtml(html)

            printer = QPrinter()
            dialog  = QPrintDialog(printer, self)
            if dialog.exec_() == QPrintDialog.Accepted:
                doc.print_(printer)
                QMessageBox.information(self, "Print Successful", "Report printed successfully!")

        except Exception as e:
            QMessageBox.critical(self, "Print Error", f"Error printing: {str(e)}")