from PyQt5.QtWidgets import (
    QWidget, QLabel, QVBoxLayout, QComboBox, QTableWidget, QTableWidgetItem,
    QDateEdit, QMessageBox, QHeaderView, QSizePolicy, QPushButton, QHBoxLayout,
    QFileDialog
)
from PyQt5.QtCore import Qt, QDate
from PyQt5.QtGui import QFont, QColor, QBrush, QTextDocument
from PyQt5.QtPrintSupport import QPrinter, QPrintDialog
from db_connect_pooled import db_manager
from db_worker import run_query_async
from date_range_widget import DateRangeWidget
import datetime


class MCPage(QWidget):
    def __init__(self, account_type=2):
        super().__init__()
        self.account_type = account_type
        # Set correct table based on brand: Brand A -> daily_reports_brand_a, Brand B -> daily_reports
        self.daily_table = "daily_reports_brand_a" if account_type == 1 else "daily_reports"
        self.setWindowTitle("MC Transactions")
        self.resize(900, 600)

        self.layout = QVBoxLayout()
        self.setLayout(self.layout)

        # Corporation selector (hidden for Brand A)
        self.corp_label = QLabel("Select Corporation:")
        self.corp_selector = QComboBox()
        self.corp_selector.currentTextChanged.connect(self.populate_table)

        # Group selector (visible for Brand A)
        self.group_label = QLabel("Select Group:")
        self.group_selector = QComboBox()
        self.group_selector.currentTextChanged.connect(self.populate_table)
        self.group_label.setVisible(False)
        self.group_selector.setVisible(False)

        # Date selector
        self.date_range_widget = DateRangeWidget()
        self.date_range_widget.dateRangeChanged.connect(self.populate_table)
        self.date_selector = self.date_range_widget  # backward-compat

        # Registration status filter
        self.reg_filter_label = QLabel("Branch Status:")
        self.reg_filter_selector = QComboBox()
        self.reg_filter_selector.addItem("Registered Only", "registered")
        self.reg_filter_selector.addItem("Not Registered", "not_registered")
        self.reg_filter_selector.addItem("All Branches", "all")
        self.reg_filter_selector.currentIndexChanged.connect(self.populate_table)
        self.reg_filter_label.setVisible(False)
        self.reg_filter_selector.setVisible(False)

        # Add selectors to layout
        self.layout.addWidget(self.corp_label)
        self.layout.addWidget(self.corp_selector)
        self.layout.addWidget(self.group_label)
        self.layout.addWidget(self.group_selector)
        self.layout.addWidget(QLabel("Select Date:"))
        self.layout.addWidget(self.date_range_widget)
        self.layout.addWidget(self.reg_filter_label)
        self.layout.addWidget(self.reg_filter_selector)

        # Table - Fixed to 3 columns
        self.table = QTableWidget()
        self.table.setColumnCount(3)
        self.table.setHorizontalHeaderLabels(["Branch", "MC Buying", "MC Selling"])
        self.table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        header = self.table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Stretch)
        header.setStretchLastSection(True)

        self.layout.addWidget(self.table)

        # Export & Print Buttons
        self.export_button = QPushButton("📊 Export to Excel")
        self.export_button.clicked.connect(self.export_to_excel)
        self.export_button.setStyleSheet("""
            QPushButton {
                background-color: #217346;
                color: white;
                padding: 8px 16px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #1a5c38;
            }
        """)

        self.print_button = QPushButton("Print")
        self.print_button.clicked.connect(self.print_table)

        button_layout = QHBoxLayout()
        button_layout.addWidget(self.export_button)
        button_layout.addWidget(self.print_button)
        self.layout.addLayout(button_layout)

        self.load_corporations()
        self.load_groups()

        # Brand A: filter by Group only, hide Corporation
        if self.account_type == 1:
            self.corp_label.setVisible(False)
            self.corp_selector.setVisible(False)
            self.group_label.setVisible(True)
            self.group_selector.setVisible(True)
            self.reg_filter_label.setVisible(True)
            self.reg_filter_selector.setVisible(True)

    def load_corporations(self):
        """Load unique corporations from the daily reports table"""
        self.corp_selector.clear()
        try:
            # Get distinct corporations from appropriate table
            query = f"SELECT DISTINCT corporation FROM {self.daily_table} ORDER BY corporation"
            corporations = db_manager.execute_query(query)

            self.corp_selector.blockSignals(True)
            for corp in corporations:
                self.corp_selector.addItem(corp['corporation'])
            self.corp_selector.blockSignals(False)

        except Exception as e:
            self.corp_selector.blockSignals(False)
            QMessageBox.critical(self, "Database Error", f"Error loading corporations: {str(e)}")

    def load_groups(self):
        """Load OS/Group names from branches table"""
        self.group_selector.blockSignals(True)
        self.group_selector.clear()
        try:
            rows = db_manager.execute_query(
                "SELECT DISTINCT os_name FROM branches "
                "WHERE os_name IS NOT NULL AND os_name != '' "
                "ORDER BY os_name"
            )
            if rows:
                for r in rows:
                    name = r['os_name'] if isinstance(r, dict) else r[0]
                    self.group_selector.addItem(name)
        except Exception as e:
            print(f"Error loading groups: {e}")
        finally:
            self.group_selector.blockSignals(False)

    def populate_table(self):
        """Populate table with MC transaction data from daily_reports - shows ALL branches"""
        date_start, date_end = self.date_range_widget.get_date_range()
        is_range = self.date_range_widget.is_range_mode()

        # Brand A: filter by Group
        if self.account_type == 1:
            group = self.group_selector.currentText()
            if not group:
                return
            reg_filter = self.reg_filter_selector.currentData()

            # Build registration clause
            reg_clause = ""
            if reg_filter == "registered":
                reg_clause = "AND b.is_registered = 1"
            elif reg_filter == "not_registered":
                reg_clause = "AND (b.is_registered = 0 OR b.is_registered IS NULL)"

            if is_range:
                query = f"""
                    SELECT b.name as branch,
                           SUM(COALESCE(dr.mc_out, 0)) as mc_buying,
                           SUM(COALESCE(dr.mc_in, 0)) as mc_selling
                    FROM branches b
                    LEFT JOIN {self.daily_table} dr ON b.name COLLATE utf8mb4_general_ci = dr.branch COLLATE utf8mb4_general_ci
                        AND dr.date >= %s AND dr.date <= %s
                    WHERE b.os_name = %s
                      {reg_clause}
                    GROUP BY b.name
                    ORDER BY b.name
                """
                params = (date_start, date_end, group)
            else:
                query = f"""
                    SELECT b.name as branch,
                           COALESCE(dr.mc_out, 0) as mc_buying,
                           COALESCE(dr.mc_in, 0) as mc_selling
                    FROM branches b
                    LEFT JOIN {self.daily_table} dr ON b.name COLLATE utf8mb4_general_ci = dr.branch COLLATE utf8mb4_general_ci
                        AND dr.date = %s
                    WHERE b.os_name = %s
                      {reg_clause}
                    ORDER BY b.name
                """
                params = (date_start, group)
        else:
            # Brand B: filter by Corporation - show all branches in corp
            corp = self.corp_selector.currentText()
            if not corp:
                return

            if is_range:
                query = f"""
                    SELECT b.name as branch,
                           SUM(COALESCE(dr.mc_out, 0)) as mc_buying,
                           SUM(COALESCE(dr.mc_in, 0)) as mc_selling
                    FROM branches b
                    LEFT JOIN {self.daily_table} dr ON b.name COLLATE utf8mb4_general_ci = dr.branch COLLATE utf8mb4_general_ci
                        AND dr.corporation = %s
                        AND dr.date >= %s AND dr.date <= %s
                    WHERE b.corporation_id = (SELECT id FROM corporations WHERE name = %s)
                       OR b.sub_corporation_id = (SELECT id FROM corporations WHERE name = %s)
                    GROUP BY b.name
                    ORDER BY b.name
                """
                params = (corp, date_start, date_end, corp, corp)
            else:
                query = f"""
                    SELECT b.name as branch,
                           COALESCE(dr.mc_out, 0) as mc_buying,
                           COALESCE(dr.mc_in, 0) as mc_selling
                    FROM branches b
                    LEFT JOIN {self.daily_table} dr ON b.name COLLATE utf8mb4_general_ci = dr.branch COLLATE utf8mb4_general_ci
                        AND dr.corporation = %s
                        AND dr.date = %s
                    WHERE b.corporation_id = (SELECT id FROM corporations WHERE name = %s)
                       OR b.sub_corporation_id = (SELECT id FROM corporations WHERE name = %s)
                    ORDER BY b.name
                """
                params = (corp, date_start, corp, corp)

        try:
            run_query_async(
                parent=self,
                query=query,
                params=params,
                on_result=self._on_populate_result,
                on_error=self._on_populate_error,
                loading_message="\u23f3  Loading MC data\u2026",
            )

        except Exception as e:
            QMessageBox.critical(self, "Database Error", f"Error loading data: {str(e)}")

    def _on_populate_error(self, err):
        QMessageBox.critical(self, "Database Error", f"Error loading data: {err}")

    def _on_populate_result(self, results):
        self.table.setRowCount(0)

        if not results:
            return

        total_buying = 0.0
        total_selling = 0.0
        row_count = 0

        for row_data in results:
            branch_name = row_data['branch']
            mc_buying = float(row_data['mc_buying'] or 0)
            mc_selling = float(row_data['mc_selling'] or 0)

            # Add row to table
            self.table.insertRow(row_count)
            self.table.setItem(row_count, 0, QTableWidgetItem(branch_name))
            self.table.setItem(row_count, 1, QTableWidgetItem(f"{mc_buying:.2f}"))
            self.table.setItem(row_count, 2, QTableWidgetItem(f"{mc_selling:.2f}"))

            total_buying += mc_buying
            total_selling += mc_selling
            row_count += 1

        # Add totals row at the bottom
        self.table.insertRow(row_count)
        total_font = QFont()
        total_font.setBold(True)

        total_brush = QBrush(QColor("#f0f0f0"))

        items = [
            QTableWidgetItem("TOTAL"),
            QTableWidgetItem(f"{total_buying:.2f}"),
            QTableWidgetItem(f"{total_selling:.2f}")
        ]

        for col, item in enumerate(items):
            item.setFont(total_font)
            item.setBackground(total_brush)
            item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(row_count, col, item)

    def export_to_excel(self):
        """Export table data to Excel with file dialog"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            QMessageBox.critical(
                self,
                "Missing Dependency",
                "The openpyxl package is required to export to Excel.\nInstall with: pip install openpyxl"
            )
            return
        
        corp = self.corp_selector.currentText() if self.account_type != 1 else self.group_selector.currentText()
        date_start, date_end = self.date_range_widget.get_date_range()
        date = date_start if date_start == date_end else f"{date_start}_to_{date_end}"
        
        rows = self.table.rowCount()
        cols = self.table.columnCount()
        
        if rows == 0:
            QMessageBox.warning(self, "No Data", "No data to export.")
            return
        
        # File dialog for save location
        default_filename = f"MC_Report_{corp}_{date}.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Excel File",
            default_filename,
            "Excel Files (*.xlsx);;All Files (*)"
        )
        
        if not file_path:
            return  # User cancelled
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "MC Report"
            
            # Styles
            title_font = Font(bold=True, size=16)
            header_font = Font(bold=True, size=11, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            total_fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Title
            ws.merge_cells('A1:C1')
            ws['A1'] = "MC Transaction Report"
            ws['A1'].font = title_font
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Info
            ws['A3'] = "Corporation:"
            ws['B3'] = corp
            ws['A4'] = "Date:"
            ws['B4'] = date
            ws['A3'].font = Font(bold=True)
            ws['A4'].font = Font(bold=True)
            
            # Headers (row 6)
            header_row = 6
            for col in range(cols):
                cell = ws.cell(row=header_row, column=col+1)
                header_item = self.table.horizontalHeaderItem(col)
                cell.value = header_item.text() if header_item else ""
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            # Data rows
            for row in range(rows):
                excel_row = header_row + 1 + row
                is_total_row = (row == rows - 1)
                
                for col in range(cols):
                    cell = ws.cell(row=excel_row, column=col+1)
                    item = self.table.item(row, col)
                    
                    if item:
                        text = item.text()
                        if col > 0:  # Numeric columns
                            try:
                                cell.value = float(text) if text else 0
                                cell.number_format = '#,##0.00'
                            except ValueError:
                                cell.value = text
                        else:
                            cell.value = text
                    
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')
                    
                    if is_total_row:
                        cell.fill = total_fill
                        cell.font = Font(bold=True)
            
            # Auto-adjust column widths
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 15
            
            wb.save(file_path)
            QMessageBox.information(self, "Export Successful", f"Report exported to:\n{file_path}")
            
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Error exporting to Excel: {str(e)}")

    def print_table(self):
        """Print the table"""
        try:
            if self.table.rowCount() == 0:
                QMessageBox.warning(self, "No Data", "No data to print.")
                return

            doc = QTextDocument()
            html = "<h2>MC Transaction Report</h2>"
            date_start, date_end = self.date_range_widget.get_date_range()
            date_display = date_start if date_start == date_end else f"{date_start} to {date_end}"
            filter_label = "Group" if self.account_type == 1 else "Corporation"
            filter_value = self.group_selector.currentText() if self.account_type == 1 else self.corp_selector.currentText()
            html += f"<p><b>{filter_label}:</b> {filter_value}<br>"
            html += f"<b>Date:</b> {date_display}</p>"
            html += "<table border='1' cellspacing='0' cellpadding='4' style='border-collapse: collapse;'>"

            # Add headers
            html += "<tr style='background-color: #f0f0f0;'>"
            for col in range(self.table.columnCount()):
                header_item = self.table.horizontalHeaderItem(col)
                header_text = header_item.text() if header_item else ""
                html += f"<th style='padding: 8px; text-align: center;'>{header_text}</th>"
            html += "</tr>"

            # Add data rows
            for row in range(self.table.rowCount()):
                html += "<tr>"
                for col in range(self.table.columnCount()):
                    item = self.table.item(row, col)
                    cell_text = item.text() if item else ""

                    # Check if this is the totals row (last row)
                    if row == self.table.rowCount() - 1:
                        html += f"<td style='padding: 8px; text-align: center; font-weight: bold; background-color: #f0f0f0;'>{cell_text}</td>"
                    else:
                        html += f"<td style='padding: 8px; text-align: center;'>{cell_text}</td>"
                html += "</tr>"

            html += "</table>"
            doc.setHtml(html)

            printer = QPrinter()
            dialog = QPrintDialog(printer, self)
            if dialog.exec_() == QPrintDialog.Accepted:
                doc.print_(printer)

        except Exception as e:
            QMessageBox.critical(self, "Print Error", f"Error printing: {str(e)}")