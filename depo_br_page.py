from PyQt5.QtWidgets import (
    QWidget, QLabel, QVBoxLayout, QHBoxLayout, QComboBox, QTableWidget,
    QTableWidgetItem, QHeaderView, QSizePolicy, QPushButton, QFileDialog,
    QMessageBox, QAbstractItemView,
)
from PyQt5.QtGui import QFont, QColor, QBrush
from PyQt5.QtCore import Qt
import json
from api_db_manager import db_manager
from db_worker import run_func_async
from date_range_widget import DateRangeWidget

BANK_ACCOUNTS = [
    {"id": 1, "bank_name": "CIB-BDO",        "account_name": "Global Reliance",                              "account_number": "0077-9002-3923"},
    {"id": 2, "bank_name": "CIB-BPI",        "account_name": "Kristal Clear Diamond and Gold Pawnshop",     "account_number": "0091-0692-29"},
    {"id": 3, "bank_name": "CIB-BDO",        "account_name": "Kristal Clear",                               "account_number": "0077-9001-8784"},
    {"id": 4, "bank_name": "CIB-Union Bank",  "account_name": "Global Reliance Mgmt and Holdings Corp",     "account_number": "0015-6000-5790"},
    {"id": 5, "bank_name": "CIB-BDO",        "account_name": "Europacific Management & Holdings Corp",     "account_number": "0038-1801-5838"},
    {"id": 6, "bank_name": "CIB-BPI",        "account_name": "Europacific Management & Holdings Corp",     "account_number": "3541-0035-67"},
    {"id": 7, "bank_name": "CIB-UB",         "account_name": "Europacific Management & Holdings Corp",     "account_number": "0021-7001-7921"},
]
_BANK_DETAIL = {b['id']: b for b in BANK_ACCOUNTS}


def _resolve_remarks_depo(bank_account_id, ft_ho_breakdown):
    """Return human-readable REMARKS DEPO string.
    ft_ho_breakdown (JSON) takes priority; falls back to single bank_account_id.
    Items may be dicts {bank_account_id, amount} or lists [name, id, amount]."""
    _bd = (ft_ho_breakdown or '').strip()
    if _bd:
        try:
            items = json.loads(_bd)
            parts = []
            for item in items:
                if isinstance(item, dict):
                    bid = item.get('bank_account_id') or item.get('id')
                    bdt = _BANK_DETAIL.get(int(bid)) if bid else None
                    if bdt:
                        parts.append(
                            f"{bdt['bank_name']} - {bdt['account_name']}"
                            f" ({bdt.get('account_number', '')})"
                            f": {item.get('amount', '')}"
                        )
                    else:
                        parts.append(str(item))
                elif isinstance(item, (list, tuple)) and len(item) >= 2:
                    # Format: [display_name, bank_id, amount]  or  [display_name, bank_id]
                    display_name = str(item[0])
                    amount = item[2] if len(item) >= 3 else ''
                    parts.append(f"{display_name}: {amount}" if amount != '' else display_name)
            if parts:
                return '; '.join(parts)
        except Exception:
            pass
    if bank_account_id:
        try:
            bdt = _BANK_DETAIL.get(int(bank_account_id))
        except (ValueError, TypeError):
            bdt = None
        if bdt:
            return f"{bdt['bank_name']} - {bdt['account_name']} ({bdt.get('account_number', '')})"
        return str(bank_account_id)
    return ''


HEADERS = [
    "BRANCHES",
    "FT FROM BRANCH",
    "CR BRANCH NAME",
    "FT TO HEAD OFFICE",
    "REMARKS DEPO",
    "FT TO BRANCH",
    "CT BRANCH NAME",
]


class DepoBRPage(QWidget):
    """DEPO BR page — Brand A only.

    Displays per-branch Fund Transfer columns from daily_reports_brand_a
    including the note/destination fields for each transfer type.

    Columns:
      BRANCHES          - branch name
      FT FROM BRANCH    - fund_transfer_from_branch
      CR BRANCH NAME    - fund_transfer_from_branch_dest (note)
      FT TO HEAD OFFICE - fund_transfer_to_head_office
      REMARKS DEPO      - fund_transfer_bank_account (bank note)
      FT TO BRANCH      - fund_transfer_to_branch
      CT BRANCH NAME    - fund_transfer_to_branch_dest (note)
    """

    DAILY_TABLE = "daily_reports_brand_a"

    def __init__(self, account_type=1):
        super().__init__()
        self.account_type = account_type
        self._building = False
        self._build_ui()
        self.load_groups()

    # ── UI ────────────────────────────────────────────────────────────────

    def _build_ui(self):
        layout = QVBoxLayout()
        layout.setContentsMargins(5, 5, 5, 5)
        layout.setSpacing(4)
        self.setLayout(layout)

        # Filter row
        filter_row = QHBoxLayout()
        filter_row.setSpacing(8)

        filter_row.addWidget(QLabel("Group:"))
        self.group_selector = QComboBox()
        filter_row.addWidget(self.group_selector, 1)

        self.date_range_widget = DateRangeWidget()
        filter_row.addWidget(self.date_range_widget)

        filter_row.addWidget(QLabel("Status:"))
        self.reg_filter_selector = QComboBox()
        self.reg_filter_selector.addItem("Registered Only", "registered")
        self.reg_filter_selector.addItem("Not Registered", "not_registered")
        self.reg_filter_selector.addItem("All Branches", "all")
        filter_row.addWidget(self.reg_filter_selector)

        self.load_btn = QPushButton("🔍 Load Report")
        self.load_btn.setStyleSheet(
            "QPushButton{background:#27AE60;color:white;padding:6px 16px;"
            "border:none;border-radius:4px;font-weight:bold;}"
            "QPushButton:hover{background:#219A52;}"
        )
        self.load_btn.clicked.connect(self.populate_table)
        filter_row.addWidget(self.load_btn)

        layout.addLayout(filter_row)

        # Table
        self.table = QTableWidget()
        self.table.setColumnCount(len(HEADERS))
        self.table.setHorizontalHeaderLabels(HEADERS)
        self.table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        hdr = self.table.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(0, QHeaderView.Stretch)
        hdr.setSectionResizeMode(2, QHeaderView.Stretch)
        hdr.setSectionResizeMode(4, QHeaderView.Stretch)
        hdr.setSectionResizeMode(6, QHeaderView.Stretch)
        self.table.setStyleSheet("""
            QTableWidget {
                gridline-color: #000;
                font-size: 11px;
            }
            QHeaderView::section {
                background-color: #2c3e50;
                color: white;
                font-weight: bold;
                border: 1px solid #1a252f;
                padding: 6px;
            }
        """)
        layout.addWidget(self.table)

        # Export button
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        self.export_btn = QPushButton("📊 Export to Excel")
        self.export_btn.clicked.connect(self.export_to_excel)
        self.export_btn.setStyleSheet("""
            QPushButton {
                background-color: #217346; color: white;
                padding: 8px 16px; border-radius: 4px;
                font-weight: bold; font-size: 12px;
            }
            QPushButton:hover { background-color: #1a5c37; }
        """)
        btn_row.addWidget(self.export_btn)
        layout.addLayout(btn_row)

    # ── Data loading ──────────────────────────────────────────────────────

    def load_groups(self):
        self.group_selector.blockSignals(True)
        self.group_selector.clear()
        try:
            rows = db_manager.execute_query(
                "SELECT DISTINCT os_name FROM branches "
                "WHERE os_name IS NOT NULL AND os_name != '' "
                "ORDER BY os_name"
            )
            if rows:
                for r in rows:
                    name = r['os_name'] if isinstance(r, dict) else r[0]
                    self.group_selector.addItem(name)
        except Exception as e:
            print(f"DepoBR: error loading groups: {e}")
        self.group_selector.blockSignals(False)
        self.populate_table()

    def populate_table(self):
        group = self.group_selector.currentText()
        if not group:
            return

        date_start, date_end = self.date_range_widget.get_date_range()
        is_range = self.date_range_widget.is_range_mode()
        reg_filter = self.reg_filter_selector.currentData()

        reg_clause = ""
        if reg_filter == "registered":
            reg_clause = "AND b.is_registered = 1"
        elif reg_filter == "not_registered":
            reg_clause = "AND (b.is_registered = 0 OR b.is_registered IS NULL)"

        if is_range and date_start != date_end:
            # For ranges: SUM amounts, collect raw bank IDs + breakdowns per branch
            data_query = f"""
                SELECT
                    b.name AS branch,
                    SUM(COALESCE(dr.fund_transfer_from_branch, 0))    AS ft_from_branch,
                    GROUP_CONCAT(DISTINCT NULLIF(TRIM(dr.fund_transfer_from_branch_dest), '')
                                 ORDER BY dr.date SEPARATOR '; ')     AS cr_branch_name,
                    SUM(COALESCE(dr.fund_transfer_to_head_office, 0)) AS ft_to_ho,
                    GROUP_CONCAT(NULLIF(TRIM(dr.fund_transfer_bank_account), '')
                                 ORDER BY dr.date SEPARATOR '|')      AS raw_bank_ids,
                    GROUP_CONCAT(NULLIF(TRIM(dr.ft_ho_breakdown), '')
                                 ORDER BY dr.date SEPARATOR '|')      AS raw_ft_breakdowns,
                    SUM(COALESCE(dr.fund_transfer_to_branch, 0))      AS ft_to_branch,
                    GROUP_CONCAT(DISTINCT NULLIF(TRIM(dr.fund_transfer_to_branch_dest), '')
                                 ORDER BY dr.date SEPARATOR '; ')     AS ct_branch_name
                FROM branches b
                LEFT JOIN {self.DAILY_TABLE} dr
                    ON b.name COLLATE utf8mb4_general_ci
                     = dr.branch COLLATE utf8mb4_general_ci
                    AND dr.date >= %s AND dr.date <= %s
                WHERE b.os_name = %s {reg_clause}
                GROUP BY b.name
                ORDER BY b.name
            """
            data_params = (date_start, date_end, group)
        else:
            data_query = f"""
                SELECT
                    b.name AS branch,
                    COALESCE(dr.fund_transfer_from_branch, 0)         AS ft_from_branch,
                    COALESCE(dr.fund_transfer_from_branch_dest, '')    AS cr_branch_name,
                    COALESCE(dr.fund_transfer_to_head_office, 0)       AS ft_to_ho,
                    dr.fund_transfer_bank_account                      AS raw_bank_id,
                    dr.ft_ho_breakdown                                 AS raw_ft_breakdown,
                    COALESCE(dr.fund_transfer_to_branch, 0)            AS ft_to_branch,
                    COALESCE(dr.fund_transfer_to_branch_dest, '')      AS ct_branch_name
                FROM branches b
                LEFT JOIN {self.DAILY_TABLE} dr
                    ON b.name COLLATE utf8mb4_general_ci
                     = dr.branch COLLATE utf8mb4_general_ci
                    AND dr.date = %s
                WHERE b.os_name = %s {reg_clause}
                ORDER BY b.name
            """
            data_params = (date_start, group)

        _dq, _dp = data_query, data_params

        def _fetch():
            return db_manager.execute_query(_dq, _dp) or []

        run_func_async(
            parent=self,
            func=_fetch,
            on_result=self._on_data_loaded,
            on_error=lambda e: print(f"DepoBR load error: {e}"),
            loading_message="⏳  Loading DEPO BR data…",
        )

    def _on_data_loaded(self, rows):
        self._building = True
        self.table.setRowCount(0)

        total_from = total_ho = total_to = 0.0

        for row in rows:
            idx = self.table.rowCount()
            self.table.insertRow(idx)

            branch     = row.get("branch", "")
            ft_from    = float(row.get("ft_from_branch", 0) or 0)
            cr_name    = row.get("cr_branch_name", "") or ""
            ft_ho      = float(row.get("ft_to_ho", 0) or 0)
            # Resolve REMARKS DEPO: may be a single bank-id or a JSON breakdown
            # For range queries the columns are named raw_bank_ids / raw_ft_breakdowns
            raw_bid = row.get("raw_bank_id") or row.get("raw_bank_ids")
            raw_bd  = row.get("raw_ft_breakdown") or row.get("raw_ft_breakdowns")
            rem_depo = _resolve_remarks_depo(raw_bid, raw_bd)
            ft_to      = float(row.get("ft_to_branch", 0) or 0)
            ct_name    = row.get("ct_branch_name", "") or ""

            total_from += ft_from
            total_ho   += ft_ho
            total_to   += ft_to

            def _ro(text, align=Qt.AlignLeft | Qt.AlignVCenter):
                item = QTableWidgetItem(text)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                item.setTextAlignment(align)
                return item

            def _num(val):
                return _ro(f"{val:,.2f}", Qt.AlignRight | Qt.AlignVCenter)

            self.table.setItem(idx, 0, _ro(branch))
            self.table.setItem(idx, 1, _num(ft_from))
            self.table.setItem(idx, 2, _ro(cr_name))
            self.table.setItem(idx, 3, _num(ft_ho))
            self.table.setItem(idx, 4, _ro(rem_depo))
            self.table.setItem(idx, 5, _num(ft_to))
            self.table.setItem(idx, 6, _ro(ct_name))

        # Totals row
        total_idx = self.table.rowCount()
        self.table.insertRow(total_idx)
        bold = QFont(); bold.setBold(True)
        bg = QBrush(QColor("#ecf0f1"))

        for col in range(len(HEADERS)):
            item = QTableWidgetItem("")
            item.setFont(bold)
            item.setBackground(bg)
            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
            self.table.setItem(total_idx, col, item)

        lbl = QTableWidgetItem("TOTAL")
        lbl.setFont(bold); lbl.setBackground(bg)
        lbl.setFlags(lbl.flags() & ~Qt.ItemIsEditable)
        self.table.setItem(total_idx, 0, lbl)

        for col, val in [(1, total_from), (3, total_ho), (5, total_to)]:
            item = QTableWidgetItem(f"{val:,.2f}")
            item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            item.setFont(bold); item.setBackground(bg)
            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
            self.table.setItem(total_idx, col, item)

        self._building = False

    # ── Export ────────────────────────────────────────────────────────────

    def export_to_excel(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        except ImportError:
            QMessageBox.critical(self, "Error", "openpyxl is required.")
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "Save DEPO BR Report", "depo_br_report.xlsx",
            "Excel Files (*.xlsx)"
        )
        if not path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "DEPO BR"

        thin = Side(style='thin')
        bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
        hdr_font  = Font(bold=True, color="FFFFFF", size=11)
        hdr_fill  = PatternFill("solid", fgColor="2C3E50")
        tot_font  = Font(bold=True)
        tot_fill  = PatternFill("solid", fgColor="ECF0F1")

        # Write headers
        for col_idx, h in enumerate(HEADERS, 1):
            c = ws.cell(row=1, column=col_idx, value=h)
            c.font = hdr_font; c.fill = hdr_fill; c.border = bdr
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Write data from the table
        total_row_idx = self.table.rowCount() - 1  # last row is totals
        for r in range(self.table.rowCount()):
            is_total = (r == total_row_idx)
            for c_idx in range(len(HEADERS)):
                item = self.table.item(r, c_idx)
                text = item.text() if item else ""
                cell = ws.cell(row=r + 2, column=c_idx + 1, value=text)
                cell.border = bdr
                if is_total:
                    cell.font = tot_font; cell.fill = tot_fill
                if c_idx in (1, 3, 5) and text:
                    try:
                        cell.value = float(text.replace(",", ""))
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal='right')
                    except ValueError:
                        pass

        # Column widths
        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 22
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 22
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 22
        ws.freeze_panes = 'B2'

        try:
            wb.save(path)
            QMessageBox.information(self, "Exported", f"DEPO BR report saved to:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save: {e}")
