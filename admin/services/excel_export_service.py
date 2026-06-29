"""
Daily Cash Count Excel export service.

Extracted from AdminDashboard.export_daily_cash_to_excel
(admin/dashboard.py ~L2702).
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

_DATE_COL_W = 18


def export_daily_cash(debit_rows, credit_rows, meta, output_path):
    """
    Write a single-sheet Daily Cash Count workbook.

    Args:
        debit_rows:  list of {'label': str, 'amount': float, 'lotes': int}
        credit_rows: list of {'label': str, 'amount': float, 'lotes': int}
        meta:        dict with keys:
                       filter_label, filter_value, branch_name, selected_date,
                       beginning_balance, debit_total, credit_total,
                       ending_balance, cash_count, cash_result
        output_path: str file path for the .xlsx file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Cash Count"

    # Styles
    title_font = Font(bold=True, size=16)
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    money_font = Font(size=10)
    total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title
    ws.merge_cells('A1:D1')
    ws['A1'] = "Daily Cash Count Report"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')

    # Info section
    ws['A3'] = f"{meta['filter_label']}:"
    ws['B3'] = meta['filter_value']
    ws['A4'] = "Branch:"
    ws['B4'] = meta['branch_name']
    ws['A5'] = "Date:"
    ws['B5'] = meta['selected_date']
    ws['A6'] = "Beginning Balance:"
    ws['B6'] = meta['beginning_balance']
    ws['B6'].number_format = '#,##0.00'

    for row in range(3, 7):
        ws[f'A{row}'].font = header_font

    # Debit Section
    row = 8
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "\U0001f4b8 DEBIT TRANSACTIONS"
    ws[f'A{row}'].font = header_font_white
    ws[f'A{row}'].fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
    ws[f'A{row}'].alignment = Alignment(horizontal='center')

    row += 1
    ws[f'A{row}'] = "Description"
    ws[f'B{row}'] = "Amount"
    ws[f'C{row}'] = "Lotes"
    for col in ['A', 'B', 'C']:
        ws[f'{col}{row}'].font = header_font
        ws[f'{col}{row}'].border = border

    row += 1
    for entry in debit_rows:
        ws[f'A{row}'] = entry['label']
        ws[f'B{row}'] = entry['amount']
        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'C{row}'] = entry['lotes']
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].border = border
        row += 1

    # Debit Total
    ws[f'A{row}'] = "Total Debit"
    ws[f'A{row}'].font = header_font
    ws[f'B{row}'] = meta['debit_total']
    ws[f'B{row}'].number_format = '#,##0.00'
    ws[f'B{row}'].font = header_font
    for col in ['A', 'B', 'C']:
        ws[f'{col}{row}'].fill = total_fill
        ws[f'{col}{row}'].border = border

    row += 2

    # Credit Section
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "\U0001f4b3 CREDIT TRANSACTIONS"
    ws[f'A{row}'].font = header_font_white
    ws[f'A{row}'].fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
    ws[f'A{row}'].alignment = Alignment(horizontal='center')

    row += 1
    ws[f'A{row}'] = "Description"
    ws[f'B{row}'] = "Amount"
    ws[f'C{row}'] = "Lotes"
    for col in ['A', 'B', 'C']:
        ws[f'{col}{row}'].font = header_font
        ws[f'{col}{row}'].border = border

    row += 1
    for entry in credit_rows:
        ws[f'A{row}'] = entry['label']
        ws[f'B{row}'] = entry['amount']
        ws[f'B{row}'].number_format = '#,##0.00'
        ws[f'C{row}'] = entry['lotes']
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].border = border
        row += 1

    # Credit Total
    ws[f'A{row}'] = "Total Credit"
    ws[f'A{row}'].font = header_font
    ws[f'B{row}'] = meta['credit_total']
    ws[f'B{row}'].number_format = '#,##0.00'
    ws[f'B{row}'].font = header_font
    for col in ['A', 'B', 'C']:
        ws[f'{col}{row}'].fill = total_fill
        ws[f'{col}{row}'].border = border

    row += 2

    # Results Section
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "\U0001f4ca SUMMARY"
    ws[f'A{row}'].font = header_font_white
    ws[f'A{row}'].fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
    ws[f'A{row}'].alignment = Alignment(horizontal='center')

    row += 1
    ws[f'A{row}'] = "Ending Balance:"
    ws[f'B{row}'] = meta['ending_balance']
    ws[f'B{row}'].number_format = '#,##0.00'
    ws[f'A{row}'].font = header_font

    row += 1
    ws[f'A{row}'] = "Cash Count:"
    ws[f'B{row}'] = meta['cash_count']
    ws[f'B{row}'].number_format = '#,##0.00'
    ws[f'A{row}'].font = header_font

    row += 1
    ws[f'A{row}'] = "Short/Over:"
    ws[f'B{row}'] = meta['cash_result']
    ws[f'B{row}'].number_format = '#,##0.00'
    ws[f'A{row}'].font = header_font

    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = _DATE_COL_W
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10

    wb.save(output_path)
