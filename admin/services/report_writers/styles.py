from openpyxl.styles import Font, PatternFill, Border, Side


class ReportStyles:
    def __init__(self):
        thin = Side(style='thin')
        self.border = Border(left=thin, right=thin, top=thin, bottom=thin)

        self.TITLE_FILL   = self.fill("1F3864")
        self.INFO_FILL    = self.fill("E8F4F8")
        self.HDR_FILL     = self.fill("4472C4")
        self.TOTAL_FILL   = self.fill("E2EFDA")
        self.GRAND_FILL   = self.fill("D5A6BD")
        self.LOTES_FILL   = self.fill("EBF5FB")
        self.DEBIT_FILL   = self.fill("27AE60")
        self.CREDIT_FILL  = self.fill("E74C3C")
        self.SUMMARY_FILL = self.fill("F39C12")
        self.GLOBAL_FILL  = self.fill("FFFF99")

        self.HDR_FONT   = Font(bold=True, size=9,  color="FFFFFF")
        self.TOTAL_FONT = Font(bold=True, size=10)
        self.BOLD_FONT  = Font(bold=True, size=10)

        self.GROUP_FILL_MAP = {
            "JEWELRY":           "DC3545", "STORAGE":            "E67E22",
            "MOTOR/CAR":         "8E44AD", "MC":                 "2980B9",
            "SILVER":            "7F8C8D", "PALAWAN":            "16A085",
            "PALAWAN SEND OUT":  "16A085", "PALAWAN PAY OUT":    "1ABC9C",
            "INSURANCE":         "C0392B", "O.S.F":              "27AE60",
            "RESCATE JEW.":      "E74C3C", "RESCATE STO.":       "F39C12",
            "GCASH IN":          "1ABC9C", "GCASH OUT":          "1ABC9C",
            "MONEYGRAM":         "2C3E50", "TRANSFAST":          "34495E",
            "RIA":               "E74C3C", "I2I REM. IN":        "2980B9",
            "I2I BILLS":         "2980B9", "I2I INSTAPAY":       "2980B9",
            "SENDAH LOAD":       "8E44AD", "SENDAH BILLS":       "8E44AD",
            "PAYMAYA":           "27AE60", "SMART $ IN":         "117864",
            "SMART $ OUT":       "117864", "GCASH PADALA":       "148F77",
            "PAL PAY IN":        "16A085", "PAL PAY OUT":        "16A085",
            "REMITLY":           "7D3C98", "SEND OUT":           "27AE60",
            "PAY OUT":           "E74C3C", "INTERNATIONAL":      "8E44AD",
            "OTHER":             "F39C12", "FUND TRANSFER":      "2980B9",
            "FUND TRANSFER HO":  "2980B9", "SMART MONEY OUT":    "117864",
            "ABRA OUT":          "1A5276", "PAL PAY CASH OUT":   "196F3D",
            "MC OUT":            "2980B9", "MC IN (SELLING)":    "27AE60",
            "MC OUT (BUYING)":   "E74C3C", "DEBIT":              "27AE60",
            "CREDIT":            "E74C3C", "SUMMARY":            "F39C12",
            "JEWELRY EMPENO":    "DC3545", "STORAGE EMPENO":     "E67E22",
        }

    @staticmethod
    def fill(hex_color: str) -> PatternFill:
        return PatternFill("solid", fgColor=hex_color.upper())

    def group_fill(self, group_name: str) -> PatternFill:
        hex_c = self.GROUP_FILL_MAP.get(group_name, "4472C4")
        return self.fill(hex_c)
