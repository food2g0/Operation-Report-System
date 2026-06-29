import json
import logging
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from .shared import get_table_cols

logger = logging.getLogger(__name__)


def write_dcc_sheet(ws, ctx, styles):
    """Write the Daily Cash Count sheet in the same vertical layout as the
    date-range Daily Cash Count report (fields in rows, branches in columns)."""

    # ── Styles matching date range report ────────────────────────
    title_font_dcc   = Font(bold=True, size=14)
    header_font_dcc  = Font(bold=True, size=9, color="FFFFFF")
    date_fill_dcc    = PatternFill("solid", fgColor="9B59B6")
    debit_fill_c     = PatternFill("solid", fgColor="27AE60")
    credit_fill_c    = PatternFill("solid", fgColor="E74C3C")
    summary_fill_c   = PatternFill("solid", fgColor="F39C12")
    total_fill_c     = PatternFill("solid", fgColor="E2EFDA")
    grand_total_fill = PatternFill("solid", fgColor="D5A6BD")
    info_fill_dcc    = PatternFill("solid", fgColor="E8F4F8")
    thin_s           = Side(style='thin')
    bdr              = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)

    # PC field lists (same as date range report)
    salary_fields     = ['pc_salary']
    pc_fields_no_sal  = [
        'pc_inc_emp', 'pc_inc_motor', 'pc_inc_suki_card', 'pc_inc_insurance',
        'pc_inc_mc', 'pc_rental', 'pc_electric', 'pc_water', 'pc_internet',
        'pc_lbc_jrs_jnt', 'pc_permits_bir_payments',
        'pc_supplies_xerox_maintenance', 'pc_transpo'
    ]

    # ── Title & info ─────────────────────────────────────────────
    ws['A1'] = f"{ctx.brand_label} - Daily Cash Count Report"
    ws['A1'].font = title_font_dcc

    ws['A3'] = f"{ctx.filter_label}:";  ws['B3'] = ctx.filter_value
    ws['A4'] = "Date:";             ws['B4'] = ctx.selected_date
    ws['A3'].font = Font(bold=True)
    ws['A4'].font = Font(bold=True)

    # ── Determine available columns ──────────────────────────────
    tbl_cols = get_table_cols(ctx.db, ctx.daily_table, ctx.col_cache)
    d_cols = [c for c in ctx.debit_fields.values()
              if not tbl_cols or c in tbl_cols]
    c_cols = [c for c in ctx.credit_fields.values()
              if not tbl_cols or c in tbl_cols]

    # ── Build SELECT with GROUP BY aggregation ───────────────────
    # For Brand A: daily_reports_brand_a has N rows per branch/date (one per user).
    # Palawan adjustment fields are BRANCH-LEVEL singletons (same value in every
    # user row) → use MAX to avoid fan-out multiplication.
    # Regular transaction fields are per-user → use SUM.
    # For Brand B: 1 row per branch, so SUM/MAX behave identically.
    BRAND_A_SINGLETON_FIELDS = {
        'palawan_cancel', 'palawan_suki_discounts', 'palawan_suki_rebates',
        'palawan_pay_out_incentives', 'palawan_suki_card',
    }
    def _agg_fn(col):
        """Return MAX for branch-level singletons, SUM for transaction fields."""
        if ctx.is_brand_a and col in BRAND_A_SINGLETON_FIELDS:
            return "MAX"
        return "SUM"

    sel_parts = [
        "dr.branch",
        "MAX(COALESCE(dr.`beginning_balance`, 0)) AS `beginning_balance`",
    ]
    for col in d_cols:
        fn = _agg_fn(col)
        sel_parts.append(f"{fn}(COALESCE(dr.`{col}`, 0)) AS `{col}`")
        lc = f"{col}_lotes"
        if not tbl_cols or lc in tbl_cols:
            sel_parts.append(f"{fn}(COALESCE(dr.`{lc}`, 0)) AS `{lc}`")
            # Capture text notes clients type in the lotes field
            sel_parts.append(
                f"GROUP_CONCAT(DISTINCT IF(dr.`{lc}` REGEXP '[a-zA-Z]',"
                f" dr.`{lc}`, NULL) SEPARATOR ' | ') AS `{lc}_note`"
            )
        else:
            sel_parts.append(f"0 AS `{lc}`")
    sel_parts.append("MAX(COALESCE(dr.`debit_total`, 0)) AS `debit_total`")
    for col in c_cols:
        fn = _agg_fn(col)
        sel_parts.append(f"{fn}(COALESCE(dr.`{col}`, 0)) AS `{col}`")
        lc = f"{col}_lotes"
        if not tbl_cols or lc in tbl_cols:
            sel_parts.append(f"{fn}(COALESCE(dr.`{lc}`, 0)) AS `{lc}`")
            # Capture text notes clients type in the lotes field
            sel_parts.append(
                f"GROUP_CONCAT(DISTINCT IF(dr.`{lc}` REGEXP '[a-zA-Z]',"
                f" dr.`{lc}`, NULL) SEPARATOR ' | ') AS `{lc}_note`"
            )
        else:
            sel_parts.append(f"0 AS `{lc}`")
    sel_parts.append("MAX(COALESCE(dr.`credit_total`, 0)) AS `credit_total`")
    sel_parts += [
        "MAX(COALESCE(dr.`ending_balance`, 0)) AS `ending_balance`",
        "MAX(COALESCE(dr.`cash_count`, 0)) AS `cash_count`",
        "MAX(COALESCE(dr.`cash_result`, 0)) AS `cash_result`",
    ]
    for sf in salary_fields + pc_fields_no_sal:
        sel_parts.append(
            f"SUM(COALESCE(dr.`{sf}`, 0)) AS `{sf}`"
            if (not tbl_cols or sf in tbl_cols) else f"0 AS `{sf}`"
        )
    # Extra note columns for Fund Transfer and PC-Salary
    for _enc in ('fund_transfer_bank_account', 'fund_transfer_to_branch_dest',
                 'fund_transfer_from_branch_dest', 'pc_salary_breakdown',
                 'ft_ho_breakdown'):
        if not tbl_cols or _enc in tbl_cols:
            sel_parts.append(f"MAX(dr.`{_enc}`) AS `{_enc}`")
    sel_clause = ", ".join(sel_parts)

    try:
        if ctx.filter_type == "os":
            if ctx.reg_filter == "registered":
                reg_clause_dcc = "AND b.is_registered = 1"
            elif ctx.reg_filter == "not_registered":
                reg_clause_dcc = "AND (b.is_registered = 0 OR b.is_registered IS NULL)"
            else:
                reg_clause_dcc = ""
            dcc_sel_clause = sel_clause.replace("dr.branch", "b.name AS branch", 1)
            sql = (
                f"SELECT {dcc_sel_clause} FROM branches b "
                f"LEFT JOIN `{ctx.daily_table}` dr "
                "ON b.name COLLATE utf8mb4_general_ci = dr.branch COLLATE utf8mb4_general_ci "
                "AND dr.date = %s "
                f"WHERE b.os_name = %s {reg_clause_dcc} "
                "GROUP BY b.name ORDER BY b.name"
            )
            rows = ctx.db.execute_query(sql, (ctx.selected_date, ctx.filter_value)) or []
        else:
            # Corporation filter – use branches-first approach to include all branches
            if ctx.reg_filter == "registered":
                reg_clause_dcc = "AND b.is_registered = 1"
            elif ctx.reg_filter == "not_registered":
                reg_clause_dcc = "AND (b.is_registered = 0 OR b.is_registered IS NULL)"
            else:
                reg_clause_dcc = ""
            dcc_sel_clause = sel_clause.replace("dr.branch", "b.name AS branch", 1)
            sql = (
                f"SELECT {dcc_sel_clause} FROM branches b "
                f"INNER JOIN corporations c "
                f"  ON (c.id = b.corporation_id OR c.id = b.sub_corporation_id) "
                f"LEFT JOIN `{ctx.daily_table}` dr "
                "ON b.name COLLATE utf8mb4_general_ci = dr.branch COLLATE utf8mb4_general_ci "
                "AND dr.date = %s "
                f"WHERE c.name COLLATE utf8mb4_general_ci = %s {reg_clause_dcc} "
                "GROUP BY b.name ORDER BY b.name"
            )
            rows = ctx.db.execute_query(sql, (ctx.selected_date, ctx.filter_value)) or []
    except Exception as ex:
        ws.cell(row=8, column=1, value=f"Error loading data: {ex}")
        return

    # ── Aggregate by branch ──────────────────────────────────────
    branch_totals = {}
    branch_notes  = {}  # text notes from lotes fields
    branches_list = []
    for row_data in rows:
        bn = row_data.get('branch', 'Unknown')
        if bn not in branch_totals:
            branch_totals[bn] = {}
            branch_notes[bn]  = {}
            branches_list.append(bn)
        for col in (['beginning_balance'] + d_cols +
                    ['debit_total'] + c_cols +
                    ['credit_total', 'ending_balance', 'cash_count', 'cash_result']):
            branch_totals[bn][col] = (
                branch_totals[bn].get(col, 0.0) + float(row_data.get(col, 0) or 0)
            )
        for col in d_cols + c_cols:
            lc = f"{col}_lotes"
            branch_totals[bn][lc] = (
                branch_totals[bn].get(lc, 0) + int(float(row_data.get(lc, 0) or 0))
            )
            # Collect text note if client typed in lotes field
            note_val = row_data.get(f"{lc}_note", '') or ''
            if note_val:
                branch_notes[bn][lc] = note_val
        for sf in salary_fields:
            branch_totals[bn]['salary'] = (
                branch_totals[bn].get('salary', 0.0) + float(row_data.get(sf, 0) or 0)
            )
        for sf in pc_fields_no_sal:
            branch_totals[bn]['total_pc'] = (
                branch_totals[bn].get('total_pc', 0.0) + float(row_data.get(sf, 0) or 0)
            )
        # FT and PC-Salary notes
        if True:
            # Build a full bank-account detail map once per row
            _BANK_DETAIL = {b['id']: b for b in ctx.bank_accounts}
            # ft_ho_breakdown takes priority (multiple FT entries)
            _ft_ho_bd = (row_data.get('ft_ho_breakdown') or '').strip()
            if _ft_ho_bd:
                try:
                    _bd_items = json.loads(_ft_ho_bd)
                    _ft_parts = []
                    for _item in _bd_items:
                        if isinstance(_item, dict):
                            # dict format: {bank_account_id, amount, ...}
                            _bid = _item.get('bank_account_id') or _item.get('id')
                            _bdt = _BANK_DETAIL.get(int(_bid)) if _bid else None
                            if _bdt:
                                _ft_parts.append(
                                    f"{_bdt['bank_name']} - {_bdt['account_name']}"
                                    f" ({_bdt.get('account_number', '')})"
                                    f": {_item.get('amount', '')}"
                                )
                            else:
                                _ft_parts.append(str(_item))
                        elif isinstance(_item, (list, tuple)) and len(_item) >= 3:
                            # list format: [bank_display, bank_id, amount]
                            _ft_parts.append(f"{_item[0]}: {_item[2]}")
                        elif isinstance(_item, (list, tuple)) and len(_item) >= 2:
                            _ft_parts.append(f"{_item[0]}: {_item[1]}")
                    if _ft_parts:
                        branch_notes[bn]['fund_transfer_to_head_office_lotes'] = (
                            'Fund Transfer to HO:\n' + '\n'.join(_ft_parts)
                        )
                except Exception:
                    pass
            else:
                _bank_id = row_data.get('fund_transfer_bank_account')
                if _bank_id:
                    _bdt = _BANK_DETAIL.get(int(_bank_id))
                    if _bdt:
                        branch_notes[bn]['fund_transfer_to_head_office_lotes'] = (
                            f"Bank: {_bdt['bank_name']}\n"
                            f"Account: {_bdt['account_name']}\n"
                            f"No: {_bdt.get('account_number', 'N/A')}"
                        )
                    else:
                        branch_notes[bn]['fund_transfer_to_head_office_lotes'] = f"Bank ID: {_bank_id}"
            _ft_to_dest = (row_data.get('fund_transfer_to_branch_dest') or '').strip()
            if _ft_to_dest:
                branch_notes[bn]['fund_transfer_to_branch_lotes'] = f"To Branch: {_ft_to_dest}"
            _ft_from_src = (row_data.get('fund_transfer_from_branch_dest') or '').strip()
            if _ft_from_src:
                branch_notes[bn]['fund_transfer_from_branch_lotes'] = f"From Branch: {_ft_from_src}"
            _pc_sal_bd = (row_data.get('pc_salary_breakdown') or '').strip()
            if _pc_sal_bd:
                try:
                    _bd_items = json.loads(_pc_sal_bd)
                    _parts = [f"{_it[0]}: {_it[1]}" for _it in _bd_items
                              if isinstance(_it, (list, tuple)) and len(_it) >= 2]
                    if _parts:
                        branch_notes[bn]['pc_salary_lotes'] = '\n'.join(_parts)
                except Exception:
                    pass
    branches_sorted = sorted(branches_list)

    # Recalculate derived totals from actual field values (fixes stale DB values)
    for bn in branches_list:
        bt = branch_totals[bn]
        _beginning  = bt.get('beginning_balance', 0.0)
        _debit_sum  = sum(bt.get(col, 0.0) for col in d_cols)
        _credit_sum = sum(bt.get(col, 0.0) for col in c_cols)
        bt['debit_total']    = _beginning + _debit_sum
        bt['credit_total']   = _credit_sum
        bt['ending_balance'] = bt['debit_total'] - bt['credit_total']
        bt['cash_result']    = bt.get('cash_count', 0.0) - bt['ending_balance']

    # ── Header row 6: "Field" | Branch names (merged 2 cols) | TOTAL ──
    HDR = 6
    cell = ws.cell(row=HDR, column=1, value="Field")
    cell.font = header_font_dcc; cell.fill = date_fill_dcc
    cell.border = bdr; cell.alignment = Alignment(horizontal='center', wrap_text=True)

    col_idx = 2
    for bn in branches_sorted:
        ws.merge_cells(start_row=HDR, start_column=col_idx,
                       end_row=HDR, end_column=col_idx + 1)
        cell = ws.cell(row=HDR, column=col_idx, value=bn)
        cell.font = header_font_dcc
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.border = bdr
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        col_idx += 2

    ws.merge_cells(start_row=HDR, start_column=col_idx,
                   end_row=HDR, end_column=col_idx + 1)
    cell = ws.cell(row=HDR, column=col_idx, value="TOTAL")
    cell.font = header_font_dcc; cell.fill = grand_total_fill
    cell.border = bdr; cell.alignment = Alignment(horizontal='center', wrap_text=True)
    total_col_start = col_idx

    # ── Sub-header row 7: Lotes / Amount per branch ──────────────
    SHDR = HDR + 1
    cell = ws.cell(row=SHDR, column=1, value="")
    cell.font = header_font_dcc; cell.fill = date_fill_dcc; cell.border = bdr

    col_idx = 2
    for bn in branches_sorted:
        for sub_lbl in ("Lotes", "Amount"):
            cell = ws.cell(row=SHDR, column=col_idx, value=sub_lbl)
            cell.font = Font(bold=True, size=8, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="4472C4")
            cell.border = bdr
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            col_idx += 1

    for sub_lbl in ("Lotes", "Amount"):
        cell = ws.cell(row=SHDR, column=col_idx, value=sub_lbl)
        cell.font = Font(bold=True, size=8, color="FFFFFF")
        cell.fill = grand_total_fill
        cell.border = bdr
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        col_idx += 1

    # ── Helper: write one data row ───────────────────────────────
    current_row = SHDR + 1

    def _dcc_row(label, db_col, show_lotes=False,
                 row_font=None, row_fill=None, row_val_font=None):
        """Write a single data row; returns (total_lotes, total_amount)."""
        nonlocal current_row
        c = ws.cell(row=current_row, column=1, value=label)
        c.border = bdr
        if row_font:
            c.font = row_font
        if row_fill:
            c.fill = row_fill

        tot_lotes, tot_amt = 0, 0.0
        ci = 2
        for bn in branches_sorted:
            bt = branch_totals.get(bn, {})
            lotes_val = int(bt.get(f"{db_col}_lotes", 0)) if show_lotes else ""
            amt_val   = float(bt.get(db_col, 0))

            lc = ws.cell(row=current_row, column=ci, value=lotes_val)
            lc.border = bdr; lc.alignment = Alignment(horizontal='center')
            if show_lotes:
                _note = branch_notes.get(bn, {}).get(f"{db_col}_lotes", '')
                if _note:
                    try:
                        from openpyxl.comments import Comment as _OXLComment
                        lc.comment = _OXLComment(_note, "Client Note")
                        lc.comment.width = 300
                        lc.comment.height = 80
                    except Exception:
                        pass
            ci += 1

            ac = ws.cell(row=current_row, column=ci, value=amt_val)
            ac.number_format = '#,##0.00'
            ac.border = bdr; ac.alignment = Alignment(horizontal='right')
            if row_val_font:
                ac.font = row_val_font
            ci += 1

            if show_lotes:
                tot_lotes += lotes_val
            tot_amt += amt_val

        # Total Lotes
        tlc = ws.cell(row=current_row, column=total_col_start,
                      value=tot_lotes if show_lotes else "")
        tlc.border = bdr; tlc.alignment = Alignment(horizontal='center')
        tlc.fill = total_fill_c

        # Total Amount
        tac = ws.cell(row=current_row, column=total_col_start + 1, value=tot_amt)
        tac.number_format = '#,##0.00'
        tac.border = bdr; tac.alignment = Alignment(horizontal='right')
        tac.fill = total_fill_c
        if row_val_font:
            tac.font = row_val_font
        elif row_font:
            tac.font = row_font

        current_row += 1
        return tot_lotes, tot_amt

    def _dcc_section_header(label, fill):
        nonlocal current_row
        c = ws.cell(row=current_row, column=1, value=label)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill; c.border = bdr
        current_row += 1

    # ── Beginning Balance ────────────────────────────────────────
    _dcc_row("Beginning Balance", "beginning_balance",
             row_font=Font(bold=True), row_fill=info_fill_dcc)

    # ── CASH RECEIPT (DEBIT) ─────────────────────────────────────
    _dcc_section_header("CASH RECEIPT (DEBIT)", debit_fill_c)
    for lbl, db_col in ctx.debit_fields.items():
        if db_col in d_cols:
            _dcc_row(lbl, db_col, show_lotes=True)
    _dcc_row("Total Cash Receipt", "debit_total",
             row_font=Font(bold=True), row_fill=info_fill_dcc)
    current_row += 1   # blank separator row

    # ── CASH OUT (CREDIT) ────────────────────────────────────────
    _dcc_section_header("CASH OUT (CREDIT)", credit_fill_c)
    for lbl, db_col in ctx.credit_fields.items():
        _dcc_row(lbl, db_col, show_lotes=True)
    _dcc_row("Total Cash Out", "credit_total",
             row_font=Font(bold=True), row_fill=info_fill_dcc)
    current_row += 1   # blank separator row

    # ── SUMMARY ──────────────────────────────────────────────────
    _dcc_section_header("SUMMARY", summary_fill_c)
    for lbl, db_col in [("Ending Balance", "ending_balance"),
                         ("Cash Count",     "cash_count"),
                         ("Variance",       "cash_result")]:
        _dcc_row(lbl, db_col)

    # ── SALARY & Total PC ────────────────────────────────────────
    _dcc_row("SALARY",   "salary",   row_font=Font(bold=True), row_fill=info_fill_dcc,
             row_val_font=Font(bold=True, color="9B59B6"))
    _dcc_row("Total PC", "total_pc", row_font=Font(bold=True), row_fill=info_fill_dcc,
             row_val_font=Font(bold=True, color="9B59B6"))

    # ── Column widths ────────────────────────────────────────────
    ws.column_dimensions['A'].width = 25
    ci = 2
    for _ in branches_sorted:
        ws.column_dimensions[get_column_letter(ci)].width = 10      # Lotes
        ws.column_dimensions[get_column_letter(ci + 1)].width = 12  # Amount
        ci += 2
    ws.column_dimensions[get_column_letter(total_col_start)].width = 10
    ws.column_dimensions[get_column_letter(total_col_start + 1)].width = 12
    ws.freeze_panes = 'B8'
