import logging
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from .shared import get_table_cols

logger = logging.getLogger(__name__)


def write_ft_sheet(ws, ctx, styles):
  
    title_font_ft    = Font(bold=True, size=16)
    subtitle_font_ft = Font(bold=True, size=12)
    date_font_ft     = Font(size=11)
    header_font_ft   = Font(bold=True, size=11, color="FFFFFF")
    header_fill_ft   = PatternFill("solid", fgColor="4472C4")
    area_hdr_fill    = PatternFill("solid", fgColor="D4EDDA")
    area_hdr_font    = Font(bold=True, size=11)
    total_fill_ft    = PatternFill("solid", fgColor="E9ECEF")
    total_font_ft    = Font(bold=True)
    es_fill_ft       = PatternFill("solid", fgColor="FFF8E1")
    es_font_ft       = Font(bold=True)
    gt_fill_ft       = PatternFill("solid", fgColor="4472C4")
    gt_font_ft       = Font(bold=True, size=12, color="FFFFFF")
    thin_ft          = Side(style='thin')
    bdr_ft           = Border(left=thin_ft, right=thin_ft,
                              top=thin_ft,  bottom=thin_ft)

    COL_LETTERS = ['A','B','C','D','E','F','G','H','I','J','K','L']

    # ── Title / info rows ────────────────────────────────────────
    ws.merge_cells('A1:L1')
    ws['A1'] = "FUND TRANSFER"
    ws['A1'].font = title_font_ft
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 20

    ws.merge_cells('A2:L2')
    ws['A2'] = f"{ctx.filter_label.upper()} {ctx.filter_value.upper()}"
    ws['A2'].font = subtitle_font_ft
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center')

    ws.merge_cells('A3:L3')
    try:
        import datetime as _dtft2
        _dobj = _dtft2.datetime.strptime(str(ctx.selected_date), '%Y-%m-%d')
        _fdate = _dobj.strftime('%A, %B %d, %Y')
    except Exception:
        _fdate = str(ctx.selected_date)
    ws['A3'] = _fdate
    ws['A3'].font = date_font_ft
    ws['A3'].alignment = Alignment(horizontal='left', vertical='center')

    # Row 5: Column headers
    hdr_vals = ['AREA', '#', 'CORPORATION', 'LOB', 'GLOBAL', 'SUNDAY',
                'Branch Name', 'Invty', 'CASH FLOAT', 'CASH COUNT',
                'BR to HO', 'BR to BR']
    for cl, hval in zip(COL_LETTERS, hdr_vals):
        c = ws[f'{cl}5']
        c.value = hval
        c.font = header_font_ft; c.fill = header_fill_ft
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = bdr_ft

    # Row 6: Sub-headers (blue; only K/L have text)
    for cl in COL_LETTERS:
        c = ws[f'{cl}6']
        c.value = ('BR to HO' if cl == 'K' else
                   'BR to BR' if cl == 'L' else '')
        c.font = header_font_ft; c.fill = header_fill_ft
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = bdr_ft

    # ── Column availability ──────────────────────────────────────
    tbl_cols = get_table_cols(ctx.db, ctx.daily_table, ctx.col_cache)
    def _col(name):
        return name if (not tbl_cols or name in tbl_cols) else None

    cc_col  = _col('cash_count')

    def _expr(col):
        return (f"COALESCE(SUM(dr.`{col}`), 0)" if col else "0")

    # Brand B has no cash float — skip the cash_float_tbl JOIN entirely
    _is_brand_b = (not ctx.is_brand_a)
    _cf_select  = "0 AS cash_float" if _is_brand_b else "COALESCE(SUM(cf.cash_float), 0) AS cash_float"
    sel_ft = (
        "dr.branch AS branch, "
        "COALESCE(b.area, 'UNASSIGNED') AS area, "
        "COALESCE(MAX(c.name), '') AS corporation_name, "
        "COALESCE(b.line_of_business, '') AS line_of_business, "
        "COALESCE(b.global_tag, '') AS global_tag, "
        "COALESCE(b.sunday, '') AS sunday, "
        f"{_expr(cc_col)}  AS cash_count, "
        f"{_cf_select}"
    )
    join_c  = ("LEFT JOIN corporations c "
               "ON (c.id = b.corporation_id OR c.id = b.sub_corporation_id)")
    join_cf = ("" if _is_brand_b else
               "LEFT JOIN cash_float_tbl cf "
               "ON cf.branch COLLATE utf8mb4_general_ci = dr.branch COLLATE utf8mb4_general_ci "
               "AND cf.date = dr.date")
    grp     = ("GROUP BY dr.branch, b.area, b.line_of_business, b.global_tag, b.sunday "
               "ORDER BY COALESCE(b.area, 'ZZZZZ'), dr.branch")

    try:
        if ctx.filter_type == "os":
            if ctx.reg_filter == "registered":
                reg_clause_ft = "AND b.is_registered = 1"
            elif ctx.reg_filter == "not_registered":
                reg_clause_ft = "AND (b.is_registered = 0 OR b.is_registered IS NULL)"
            else:
                reg_clause_ft = ""
            sel_ft_os = sel_ft.replace("dr.branch AS branch", "b.name AS branch", 1)
            grp_os    = grp.replace("dr.branch", "b.name")
            sql_ft = (
                f"SELECT {sel_ft_os} FROM branches b "
                f"LEFT JOIN `{ctx.daily_table}` dr "
                "ON b.name COLLATE utf8mb4_general_ci = dr.branch COLLATE utf8mb4_general_ci "
                "AND dr.date = %s "
                f"{join_c} {join_cf} "
                f"WHERE b.os_name = %s {reg_clause_ft} {grp_os}"
            )
            ft_rows = ctx.db.execute_query(
                sql_ft, (ctx.selected_date, ctx.filter_value)) or []
        else:
            # Corporation filter – use branches-first approach to include all branches
            if ctx.reg_filter == "registered":
                reg_clause_ft = "AND b.is_registered = 1"
            elif ctx.reg_filter == "not_registered":
                reg_clause_ft = "AND (b.is_registered = 0 OR b.is_registered IS NULL)"
            else:
                reg_clause_ft = ""
            sel_ft_corp = sel_ft.replace("dr.branch AS branch", "b.name AS branch", 1)
            grp_corp    = grp.replace("dr.branch", "b.name")
            sql_ft = (
                f"SELECT {sel_ft_corp} FROM branches b "
                f"INNER JOIN corporations c "
                f"  ON (c.id = b.corporation_id OR c.id = b.sub_corporation_id) "
                f"LEFT JOIN `{ctx.daily_table}` dr "
                "ON b.name COLLATE utf8mb4_general_ci = dr.branch COLLATE utf8mb4_general_ci "
                "AND dr.date = %s "
                f"{join_cf} "
                f"WHERE c.name COLLATE utf8mb4_general_ci = %s {reg_clause_ft} {grp_corp}"
            )
            ft_rows = ctx.db.execute_query(
                sql_ft, (ctx.selected_date, ctx.filter_value)) or []
    except Exception as ex:
        ws['A7'] = f"Error loading fund transfer data: {ex}"
        return

    # Extra space value
    try:
        _es = ctx.db.execute_query(
            "SELECT amount FROM extra_space_fund_transfer WHERE report_date = %s",
            (ctx.selected_date,))
        extra_space_val = float(_es[0].get('amount', 0) or 0) if _es else 0.0
    except Exception:
        extra_space_val = 0.0

    # ── Group by area ────────────────────────────────────────────
    from collections import OrderedDict as _OD2
    area_groups = _OD2()
    for rd in ft_rows:
        area_groups.setdefault(rd.get('area') or 'UNASSIGNED', []).append(rd)

    # ── Write data rows ──────────────────────────────────────────
    excel_row   = 7
    grand_total = 0.0
    branch_num  = 1

    for area_name, area_branches in area_groups.items():
        # Area header (merged, green)
        ws.merge_cells(f'A{excel_row}:L{excel_row}')
        c = ws[f'A{excel_row}']
        c.value = f"{area_name} AREA"
        c.fill = area_hdr_fill; c.font = area_hdr_font
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = bdr_ft
        for cl in ['B','C','D','E','F','G','H','I','J','K','L']:
            ws[f'{cl}{excel_row}'].fill  = area_hdr_fill
            ws[f'{cl}{excel_row}'].border = bdr_ft
        excel_row += 1

        area_cc = 0.0
        for rd in area_branches:
            sunday = rd.get('sunday', '') or ''
            if sunday == 'NO':
                sunday = 'NO SUNDAY'
            cash_count = float(rd.get('cash_count', 0) or 0)
            cash_float = 0.0 if _is_brand_b else float(rd.get('cash_float', 0) or 0)

            row_vals = [
                (rd.get('area', '') or '',        'center', None),
                (branch_num,                       'center', None),
                (rd.get('corporation_name','')or '','center', None),
                (rd.get('line_of_business','')or '','center', None),
                (rd.get('global_tag', '') or '',   'center', None),
                (sunday,                           'center', None),
                (rd.get('branch', ''),             'left',   None),
                ('',                               'center', None),
                ('' if _is_brand_b else (cash_float if cash_float else ''), 'center', '#,##0.00'),
                (cash_count,                       'right',  '#,##0.00'),
                ('',                               'right',  None),
                ('',                               'right',  None),
            ]
            for ci, (val, align, nfmt) in enumerate(row_vals, 1):
                c = ws.cell(row=excel_row, column=ci, value=val)
                c.alignment = Alignment(horizontal=align)
                c.border = bdr_ft
                if nfmt and isinstance(val, float):
                    c.number_format = nfmt

            area_cc     += cash_count
            grand_total += cash_count
            branch_num  += 1
            excel_row   += 1

        # Area total row (grey)
        for ci in range(1, 13):
            c = ws.cell(row=excel_row, column=ci)
            c.fill = total_fill_ft; c.font = total_font_ft; c.border = bdr_ft
            if ci == 7:
                c.value = f"TOTAL {area_name}"
                c.alignment = Alignment(horizontal='right')
            elif ci == 10:
                c.value = area_cc
                c.number_format = '#,##0.00'
                c.alignment = Alignment(horizontal='right')
            else:
                c.value = ''; c.alignment = Alignment(horizontal='center')
        excel_row += 1

    # Extra space row (yellow)
    for ci in range(1, 13):
        c = ws.cell(row=excel_row, column=ci)
        c.fill = es_fill_ft; c.font = es_font_ft; c.border = bdr_ft
        if ci == 7:
            c.value = "EXTRA SPACE"
            c.alignment = Alignment(horizontal='center')
        elif ci == 10:
            if extra_space_val:
                c.value = extra_space_val
                c.number_format = '#,##0.00'
            c.alignment = Alignment(horizontal='right')
        else:
            c.value = ''; c.alignment = Alignment(horizontal='center')
    excel_row += 1

    # Spacer row
    for ci in range(1, 13):
        ws.cell(row=excel_row, column=ci).border = bdr_ft
    excel_row += 1

    # Grand total row (blue, white font)
    for ci in range(1, 13):
        c = ws.cell(row=excel_row, column=ci)
        c.fill = gt_fill_ft; c.font = gt_font_ft; c.border = bdr_ft
        if ci == 7:
            c.value = "GRAND TOTAL"
            c.alignment = Alignment(horizontal='right')
        elif ci == 10:
            c.value = grand_total
            c.number_format = '#,##0.00'
            c.alignment = Alignment(horizontal='right')
        else:
            c.value = ''; c.alignment = Alignment(horizontal='center')

    # ── Column widths ────────────────────────────────────────────
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 12
    ws.column_dimensions['L'].width = 12
    ws.freeze_panes = 'A7'

