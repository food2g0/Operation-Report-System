import json
import logging
from openpyxl.styles import Alignment

from .shared import write_info

logger = logging.getLogger(__name__)


def write_depo_br_sheet(ws, ctx, styles):
    write_info(ws, ctx, styles)

    depo_headers = [
        "BRANCHES",
        "FT FROM BRANCH",
        "CR BRANCH NAME",
        "FT TO HEAD OFFICE",
        "REMARKS DEPO",
        "FT TO BRANCH",
        "CT BRANCH NAME",
    ]

    HDR_ROW  = 7
    DATA_START = 8

    for c_idx, h in enumerate(depo_headers, 1):
        c = ws.cell(row=HDR_ROW, column=c_idx, value=h)
        c.font = styles.HDR_FONT; c.fill = styles.HDR_FILL; c.border = styles.border
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Build query — branches-first so every branch appears
    reg_where = ""
    if ctx.reg_filter == "registered":
        reg_where = "AND b.is_registered = 1"
    elif ctx.reg_filter == "not_registered":
        reg_where = "AND (b.is_registered = 0 OR b.is_registered IS NULL)"

    try:
        if ctx.filter_type == "os":
            sql = f"""
                SELECT b.name AS branch,
                       COALESCE(dr.fund_transfer_from_branch, 0)        AS ft_from_branch,
                       COALESCE(dr.fund_transfer_from_branch_dest, '')   AS cr_branch_name,
                       COALESCE(dr.fund_transfer_to_head_office, 0)      AS ft_to_ho,
                       dr.fund_transfer_bank_account                     AS raw_bank_id,
                       dr.ft_ho_breakdown                                AS raw_ft_breakdown,
                       COALESCE(dr.fund_transfer_to_branch, 0)           AS ft_to_branch,
                       COALESCE(dr.fund_transfer_to_branch_dest, '')     AS ct_branch_name
                FROM branches b
                LEFT JOIN `{ctx.daily_table}` dr
                    ON b.name COLLATE utf8mb4_general_ci
                     = dr.branch COLLATE utf8mb4_general_ci
                    AND dr.date = %s
                WHERE b.os_name = %s {reg_where}
                ORDER BY b.name
            """
            sql_params = (ctx.selected_date, ctx.filter_value)
        else:
            sql = f"""
                SELECT b.name AS branch,
                       COALESCE(dr.fund_transfer_from_branch, 0)        AS ft_from_branch,
                       COALESCE(dr.fund_transfer_from_branch_dest, '')   AS cr_branch_name,
                       COALESCE(dr.fund_transfer_to_head_office, 0)      AS ft_to_ho,
                       dr.fund_transfer_bank_account                     AS raw_bank_id,
                       dr.ft_ho_breakdown                                AS raw_ft_breakdown,
                       COALESCE(dr.fund_transfer_to_branch, 0)           AS ft_to_branch,
                       COALESCE(dr.fund_transfer_to_branch_dest, '')     AS ct_branch_name
                FROM branches b
                INNER JOIN corporations c
                    ON (c.id = b.corporation_id OR c.id = b.sub_corporation_id)
                    AND c.name = %s
                LEFT JOIN `{ctx.daily_table}` dr
                    ON b.name COLLATE utf8mb4_general_ci
                     = dr.branch COLLATE utf8mb4_general_ci
                    AND dr.date = %s
                WHERE 1=1 {reg_where}
                ORDER BY b.name
            """
            sql_params = (ctx.filter_value, ctx.selected_date)

        results = ctx.db.execute_query(sql, sql_params) or []
    except Exception as ex:
        ws.cell(row=DATA_START, column=1, value=f"Error loading data: {ex}")
        return

    # Bank account ID lookup (same list as BANK_ACCOUNTS on AdminDashboard)
    _bank_detail = {b['id']: b for b in ctx.bank_accounts}

    def _resolve_depo(raw_bank_id, raw_ft_breakdown):
        _bd = (raw_ft_breakdown or '').strip()
        if _bd:
            try:
                items = json.loads(_bd)
                parts = []
                for item in items:
                    if isinstance(item, dict):
                        bid = item.get('bank_account_id') or item.get('id')
                        bdt = _bank_detail.get(int(bid)) if bid else None
                        if bdt:
                            parts.append(
                                f"{bdt['bank_name']} - {bdt['account_name']}"
                                f" ({bdt.get('account_number', '')})"
                                f": {item.get('amount', '')}"
                            )
                        else:
                            parts.append(str(item))
                    elif isinstance(item, (list, tuple)) and len(item) >= 2:
                        # Format: [display_name, bank_id, amount]
                        display_name = str(item[0])
                        amount = item[2] if len(item) >= 3 else ''
                        parts.append(f"{display_name}: {amount}" if amount != '' else display_name)
                if parts:
                    return '; '.join(parts)
            except Exception:
                pass
        if raw_bank_id:
            try:
                bdt = _bank_detail.get(int(raw_bank_id))
            except (ValueError, TypeError):
                bdt = None
            if bdt:
                return f"{bdt['bank_name']} - {bdt['account_name']} ({bdt.get('account_number', '')})"
            return str(raw_bank_id)
        return ''

    total_from = total_ho = total_to = 0.0

    for r_idx, row in enumerate(results, DATA_START):
        ft_from  = float(row.get('ft_from_branch', 0) or 0)
        cr_name  = str(row.get('cr_branch_name', '') or '')
        ft_ho    = float(row.get('ft_to_ho', 0) or 0)
        rem_depo = _resolve_depo(row.get('raw_bank_id'), row.get('raw_ft_breakdown'))
        ft_to    = float(row.get('ft_to_branch', 0) or 0)
        ct_name  = str(row.get('ct_branch_name', '') or '')

        total_from += ft_from
        total_ho   += ft_ho
        total_to   += ft_to

        values = [
            row.get('branch', ''),
            ft_from,
            cr_name,
            ft_ho,
            rem_depo,
            ft_to,
            ct_name,
        ]
        for c_idx, val in enumerate(values, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = styles.border
            if c_idx in (2, 4, 6):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='left')

    # Totals row
    tot_row = DATA_START + len(results)
    tot_vals = ["TOTAL", total_from, "", total_ho, "", total_to, ""]
    for c_idx, val in enumerate(tot_vals, 1):
        cell = ws.cell(row=tot_row, column=c_idx, value=val)
        cell.font = styles.TOTAL_FONT; cell.fill = styles.TOTAL_FILL; cell.border = styles.border
        if c_idx in (2, 4, 6):
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal='right')

    # Column widths
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 22
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 22
    ws.freeze_panes = 'B8'

