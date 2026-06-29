import logging
from openpyxl.styles import Font, PatternFill, Alignment

logger = logging.getLogger(__name__)


def write_pepp_report_sheet(ws, ctx, styles):
    from decimal import Decimal, ROUND_HALF_UP

    def _r2(v):
        return float(Decimal(str(v)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))

    def _m2(a, b):
        return float((Decimal(str(a)) * Decimal(str(b))).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))

    _corp_abbrev_map = {
        'SILVERSTAR JEWELRY PAWNSHOP INC':                'SJPI',
        'ALEXITE JEWELRY PAWNSHOP INC':                   'AJPI',
        'SAN RAMON PLATINUM PAWNSHOP INC':                'SRPPI',
        'HOMENEEDS PAWNSHOP INC':                         'HPI',
        'KRISTAL CLEAR DIAMOND AND GOLD PAWNSHOP INC':    'KCDGPI',
        'SAFELOCK PAWNSHOP INC':                          'SPI',
        'MEGAWORLD DOMESTIC PAWNSHOP INC':                'MDPI',
        'GLOBAL RELIANCE MANAGEMENT & HOLDINGS CORP.':    'GRMHC',
    }
    _registry_map = {
        'SILVERSTAR JEWELRY PAWNSHOP INC':                'P250682A',
        'ALEXITE JEWELRY PAWNSHOP INC':                   'P250683A',
        'SAN RAMON PLATINUM PAWNSHOP INC':                'P250681A',
        'HOMENEEDS PAWNSHOP INC':                         'P250677A',
        'KRISTAL CLEAR DIAMOND AND GOLD PAWNSHOP INC':    'P250678A',
        'SAFELOCK PAWNSHOP INC':                          'P250680A',
        'MEGAWORLD DOMESTIC PAWNSHOP INC':                'P250679A',
        'GLOBAL RELIANCE MANAGEMENT & HOLDINGS CORP.':    'P210021A',
    }

    corp_name   = ctx.filter_value
    corp_upper  = corp_name.upper().strip().rstrip('.')
    corp_abbrev = next((v for k, v in _corp_abbrev_map.items() if k.upper().strip().rstrip('.') == corp_upper), corp_name[:4].upper())
    registry    = next((v for k, v in _registry_map.items()    if k.upper().strip().rstrip('.') == corp_upper), '')
    is_global   = 'GLOBAL RELIANCE' in corp_upper
    payable_tbl = "payable_tbl_brand_a"

    if ctx.reg_filter == "registered":
        reg_clause_payable = "AND b.is_registered = 1"
    elif ctx.reg_filter == "not_registered":
        reg_clause_payable = "AND (b.is_registered = 0 OR b.is_registered IS NULL)"
    else:
        reg_clause_payable = ""

    try:
        if ctx.filter_type == "os":
            result = ctx.db.execute_query(f"""
                SELECT SUM(COALESCE(p.sendout_capital, 0))        AS total_sendout_capital,
                       SUM(COALESCE(p.sendout_commission, 0))     AS total_sendout_commission,
                       SUM(COALESCE(p.sendout_sc, 0))             AS total_sendout_sc,
                       SUM(COALESCE(p.payout_capital, 0))         AS total_payout_capital,
                       SUM(COALESCE(p.payout_commission, 0))      AS total_payout_commission,
                       SUM(COALESCE(p.payout_sc, 0))              AS total_payout_sc,
                       SUM(COALESCE(p.international_commission, 0)) AS total_international_commission,
                       SUM(COALESCE(p.skid, 0))                   AS total_skid,
                       SUM(COALESCE(p.skir, 0))                   AS total_skir,
                       SUM(COALESCE(p.cancellation, 0))           AS total_cancellation,
                       SUM(COALESCE(p.inc, 0))                    AS total_inc
                FROM {payable_tbl} p
                INNER JOIN branches b ON p.branch COLLATE utf8mb4_general_ci = b.name COLLATE utf8mb4_general_ci
                INNER JOIN corporations c ON (c.id = b.corporation_id OR c.id = b.sub_corporation_id)
                                          AND p.corporation COLLATE utf8mb4_general_ci = c.name COLLATE utf8mb4_general_ci
                WHERE b.os_name = %s AND p.date = %s {reg_clause_payable}
            """, (ctx.filter_value, ctx.selected_date))
        elif is_global:
            result = ctx.db.execute_query(f"""
                SELECT SUM(COALESCE(p.sendout_capital, 0))        AS total_sendout_capital,
                       SUM(COALESCE(p.sendout_commission, 0))     AS total_sendout_commission,
                       SUM(COALESCE(p.sendout_sc, 0))             AS total_sendout_sc,
                       SUM(COALESCE(p.payout_capital, 0))         AS total_payout_capital,
                       SUM(COALESCE(p.payout_commission, 0))      AS total_payout_commission,
                       SUM(COALESCE(p.payout_sc, 0))              AS total_payout_sc,
                       SUM(COALESCE(p.international_commission, 0)) AS total_international_commission,
                       SUM(COALESCE(p.skid, 0))                   AS total_skid,
                       SUM(COALESCE(p.skir, 0))                   AS total_skir,
                       SUM(COALESCE(p.cancellation, 0))           AS total_cancellation,
                       SUM(COALESCE(p.inc, 0))                    AS total_inc
                FROM {payable_tbl} p
                INNER JOIN branches b ON p.branch COLLATE utf8mb4_general_ci = b.name COLLATE utf8mb4_general_ci
                INNER JOIN corporations c ON (c.id = b.corporation_id OR c.id = b.sub_corporation_id)
                                          AND p.corporation COLLATE utf8mb4_general_ci = c.name COLLATE utf8mb4_general_ci
                WHERE b.global_tag = 'GLOBAL' AND p.date = %s {reg_clause_payable}
            """, (ctx.selected_date,))
        else:
            result = ctx.db.execute_query(f"""
                SELECT SUM(COALESCE(p.sendout_capital, 0))        AS total_sendout_capital,
                       SUM(COALESCE(p.sendout_commission, 0))     AS total_sendout_commission,
                       SUM(COALESCE(p.sendout_sc, 0))             AS total_sendout_sc,
                       SUM(COALESCE(p.payout_capital, 0))         AS total_payout_capital,
                       SUM(COALESCE(p.payout_commission, 0))      AS total_payout_commission,
                       SUM(COALESCE(p.payout_sc, 0))              AS total_payout_sc,
                       SUM(COALESCE(p.international_commission, 0)) AS total_international_commission,
                       SUM(COALESCE(p.skid, 0))                   AS total_skid,
                       SUM(COALESCE(p.skir, 0))                   AS total_skir,
                       SUM(COALESCE(p.cancellation, 0))           AS total_cancellation,
                       SUM(COALESCE(p.inc, 0))                    AS total_inc
                FROM {payable_tbl} p
                INNER JOIN branches b ON p.branch COLLATE utf8mb4_general_ci = b.name COLLATE utf8mb4_general_ci
                INNER JOIN corporations c ON (c.id = b.corporation_id OR c.id = b.sub_corporation_id)
                                          AND p.corporation COLLATE utf8mb4_general_ci = c.name COLLATE utf8mb4_general_ci
                WHERE p.corporation = %s AND p.date = %s {reg_clause_payable}
            """, (ctx.filter_value, ctx.selected_date))
    except Exception as ex:
        ws['A1'] = f"Error loading data: {ex}"
        return

    if not result or not result[0]:
        ws['A1'] = "No data found."
        return

    row_data = result[0]
    if not isinstance(row_data, dict):
        _keys = [
            'total_sendout_capital', 'total_sendout_commission', 'total_sendout_sc',
            'total_payout_capital',  'total_payout_commission',  'total_payout_sc',
            'total_international_commission',
            'total_skid', 'total_skir', 'total_cancellation', 'total_inc',
        ]
        row_data = dict(zip(_keys, row_data))

    if all(v is None for v in row_data.values()):
        ws['A1'] = "No data found."
        return

    sendout_capital          = float(row_data.get('total_sendout_capital')          or 0)
    sendout_commission       = float(row_data.get('total_sendout_commission')       or 0)
    sendout_sc               = float(row_data.get('total_sendout_sc')               or 0)
    payout_capital           = float(row_data.get('total_payout_capital')           or 0)
    payout_commission        = float(row_data.get('total_payout_commission')        or 0)
    payout_sc                = float(row_data.get('total_payout_sc')                or 0)
    international_commission = float(row_data.get('total_international_commission') or 0)
    total_skid               = float(row_data.get('total_skid')                     or 0)
    total_skir               = float(row_data.get('total_skir')                     or 0)
    total_cancellation       = float(row_data.get('total_cancellation')             or 0)
    total_inc                = float(row_data.get('total_inc')                      or 0)

    pepp_commission_61       = _m2(sendout_commission, 0.61)
    skid_61                  = _m2(total_skid, 0.61)
    corp_commission_43       = _m2(payout_commission, 0.43)
    corp_international_80    = _m2(international_commission, 0.80)
    skir_57                  = _m2(total_skir, 0.57)

    send_subtotal                = _r2(sendout_capital + pepp_commission_61 + sendout_sc)
    send_subtotal_after_discount = _r2(send_subtotal - skid_61)
    total_net_send               = _r2(send_subtotal_after_discount - total_cancellation)
    release_subtotal             = _r2(payout_capital + corp_commission_43 + corp_international_80)
    release_subtotal_with_inc    = _r2(release_subtotal + total_inc)
    total_net_released           = _r2(release_subtotal_with_inc + skir_57)
    net_receivable_payable       = _r2(total_net_send - total_net_released)

    hdr_font_pr = Font(bold=True, size=12)
    sec_font_pr = Font(bold=True, size=11)
    reg_font_pr = Font(size=10)
    tot_font_pr = Font(bold=True, size=11)
    hdr_fill_pr = PatternFill("solid", fgColor="E6F3FF")
    sub_fill_pr = PatternFill("solid", fgColor="F0F0F0")
    tot_fill_pr = PatternFill("solid", fgColor="D9D9D9")
    right_al    = Alignment(horizontal='right')
    center_al   = Alignment(horizontal='center')

    r = 1
    ws.merge_cells(f'A{r}:D{r}')
    ws[f'A{r}'] = f"Palawan Express Pera Padala - {corp_name}"
    ws[f'A{r}'].font      = hdr_font_pr
    ws[f'A{r}'].alignment = center_al
    ws[f'A{r}'].fill      = hdr_fill_pr
    r += 1

    ws[f'A{r}'] = f"PEPP Reconciliation for {ctx.selected_date}"
    ws[f'C{r}'] = "Partner Registry No."
    ws[f'D{r}'] = registry
    for cell in [ws[f'A{r}'], ws[f'C{r}'], ws[f'D{r}']]:
        cell.font = hdr_font_pr
    ws[f'C{r}'].alignment = right_al
    ws[f'D{r}'].alignment = right_al
    r += 2

    rows_pr = [
        ("Send Transaction",                                                   "", "",                               "",                                    "section"),
        (f"    PEPP Remittance from {corp_name}",                              "", "P",                              f"{sendout_capital:,.2f}",              "indent"),
        ("    PEPP share: 61% of commission",                                  "P", f"{sendout_commission:,.2f}",    f"{pepp_commission_61:,.2f}",           "indent"),
        ("    PEPP share: Service Charge",                                     "", f"{sendout_sc:,.2f}",             f"{sendout_sc:,.2f}",                   "indent"),
        ("        Subtotal",                                                   "", "",                               f"{send_subtotal:,.2f}",                "subtotal"),
        ("    Less: Discount (Suki Card)",                                     "", f"({total_skid:,.2f})",           f"({skid_61:,.2f})",                    "indent"),
        ("        Subtotal",                                                   "", "",                               f"{send_subtotal_after_discount:,.2f}", "subtotal"),
        ("    Less: Cancellation",                                             "", f"({total_cancellation:,.2f})",   f"({total_cancellation:,.2f})",         "indent"),
        ("    Total Net Send",                                                 "", "",                               f"{total_net_send:,.2f}",               "total"),
        ("", "", "", "", "blank"),
        (f"    RELEASE Transaction (Payable to {corp_abbrev})",                "", "",                               "",                                    "section"),
        (f"    PEPP Remittances released at {corp_abbrev}",                    "", "P",                              f"{payout_capital:,.2f}",               "indent"),
        (f"    {corp_abbrev} share: 43% of commission",                        "P", f"{payout_commission:,.2f}",     f"{corp_commission_43:,.2f}",           "indent"),
        (f"    {corp_abbrev} share: 50% of commission (LBC Domestic Payout)",  "", "",                               "",                                    "indent"),
        (f"    {corp_abbrev} share: 80% of commission (International Payout)", "", f"{international_commission:,.2f}", f"{corp_international_80:,.2f}",    "indent"),
        ("    Service Charge",                                                 "", f"{payout_sc:,.2f}",              "-",                                   "indent"),
        ("        Subtotal",                                                   "", "",                               f"{release_subtotal:,.2f}",             "subtotal"),
        (f"    Add: {corp_abbrev} Branch Incentives released",                 "", "",                               f"{total_inc:,.2f}",                   "indent"),
        ("        Subtotal",                                                   "", "",                               f"{release_subtotal_with_inc:,.2f}",    "subtotal"),
        ("    Add: Rebates (Suki Card)",                                       "", f"{total_skir:,.2f}",             f"{skir_57:,.2f}",                     "indent"),
        ("    Total Net Released",                                             "", "",                               f"{total_net_released:,.2f}",           "total"),
        ("", "", "", "", "blank"),
        ("    Net Send",                   "", "", f"{total_net_send:,.2f}",         "regular"),
        ("    Less : Net Released",        "", "", f"{total_net_released:,.2f}",     "regular"),
        ("    Net Receivable / (Payable)", "", "", f"{net_receivable_payable:,.2f}", "total"),
    ]

    for col1, col2, col3, col4, row_type in rows_pr:
        ws[f'A{r}'] = col1
        ws[f'B{r}'] = col2
        ws[f'C{r}'] = col3
        ws[f'D{r}'] = col4

        if row_type == "section":
            for c in ['A', 'B', 'C', 'D']:
                ws[f'{c}{r}'].font = sec_font_pr
        elif row_type == "total":
            for c in ['A', 'B', 'C', 'D']:
                ws[f'{c}{r}'].font = tot_font_pr
                ws[f'{c}{r}'].fill = tot_fill_pr
        elif row_type == "subtotal":
            for c in ['A', 'B', 'C', 'D']:
                ws[f'{c}{r}'].font = tot_font_pr
                ws[f'{c}{r}'].fill = sub_fill_pr
        else:
            for c in ['A', 'B', 'C', 'D']:
                ws[f'{c}{r}'].font = reg_font_pr

        ws[f'C{r}'].alignment = right_al
        ws[f'D{r}'].alignment = right_al
        r += 1

    r += 2
    ws[f'A{r}'] = "Prepared by:"
    ws[f'C{r}'] = "Noted by:"
    r += 2
    ws[f'A{r}'] = "Rochelle G. Serrano"
    ws[f'C{r}'] = "Aimee M. Martinez"

    ws.column_dimensions['A'].width = 55
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18

