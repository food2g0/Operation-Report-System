import logging
from openpyxl.styles import Alignment

from .shared import write_info

logger = logging.getLogger(__name__)


def write_palawan_sheet(ws, ctx, styles):
    write_info(ws, ctx, styles)

    # ── Build query identical to palawan_page.py ──────────────
    select_cols = """
        SELECT DISTINCT b.name AS branch,
               COALESCE(dr.palawan_send_out, 0)                    AS palawan_send_out,
               COALESCE(dr.palawan_sc, 0)                          AS palawan_sc,
               COALESCE(dr.palawan_pay_out, 0)                     AS palawan_pay_out,
               COALESCE(dr.palawan_pay_out_incentives, 0)          AS palawan_pay_out_incentives,
               COALESCE(dr.palawan_send_out_lotes, 0)              AS palawan_send_out_lotes,
               COALESCE(dr.palawan_sc_lotes, 0)                    AS palawan_sc_lotes,
               COALESCE(dr.palawan_pay_out_lotes, 0)               AS palawan_pay_out_lotes,
               COALESCE(dr.palawan_pay_out_incentives_lotes, 0)    AS palawan_pay_out_incentives_lotes
    """

    if ctx.reg_filter == "registered":
        reg_clause = "AND b.is_registered = 1"
    elif ctx.reg_filter == "not_registered":
        reg_clause = "AND (b.is_registered = 0 OR b.is_registered IS NULL)"
    else:
        reg_clause = ""

    if ctx.filter_type == "corporation":
        query = f"""
            {select_cols}
            FROM branches b
            LEFT JOIN corporations c
                ON (b.corporation_id = c.id OR b.sub_corporation_id = c.id)
            LEFT JOIN `{ctx.daily_table}` dr
                ON b.name COLLATE utf8mb4_general_ci = dr.branch COLLATE utf8mb4_general_ci
               AND dr.corporation = %s
               AND dr.date = %s
            WHERE (b.corporation_id  = (SELECT id FROM corporations WHERE name = %s)
                OR b.sub_corporation_id = (SELECT id FROM corporations WHERE name = %s))
            {reg_clause}
            ORDER BY b.name
        """
        params = (ctx.filter_value, ctx.selected_date, ctx.filter_value, ctx.filter_value)
    else:  # os / group
        query = f"""
            {select_cols}
            FROM branches b
            LEFT JOIN `{ctx.daily_table}` dr
                ON b.name COLLATE utf8mb4_general_ci = dr.branch COLLATE utf8mb4_general_ci
               AND dr.date = %s
            WHERE b.os_name = %s
            {reg_clause}
            ORDER BY b.name
        """
        params = (ctx.selected_date, ctx.filter_value)

    try:
        results = ctx.db.execute_query(query, params) or []
    except Exception as ex:
        ws['A7'] = f"Error loading data: {ex}"
        return

    # ── Column headers (row 7) ────────────────────────────────
    headers = ["Branch", "Palawan In", "Lotes In", "Palawan Out", "Lotes Out"]
    hdr_fills = [styles.HDR_FILL, styles.fill("16A085"), styles.LOTES_FILL, styles.fill("E74C3C"), styles.LOTES_FILL]
    for col_idx, (hdr, fill) in enumerate(zip(headers, hdr_fills), start=1):
        c = ws.cell(row=7, column=col_idx, value=hdr)
        c.font      = styles.HDR_FONT
        c.fill      = fill
        c.border    = styles.border
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[7].height = 20

    # ── Data rows (row 8 onwards) ─────────────────────────────
    total_in      = 0.0
    total_out     = 0.0
    total_lotes_in  = 0
    total_lotes_out = 0
    data_row = 8

    for row_data in results:
        branch_name  = row_data['branch']
        send_out     = float(row_data['palawan_send_out']              or 0)
        sc           = float(row_data['palawan_sc']                    or 0)
        pay_out      = float(row_data['palawan_pay_out']               or 0)
        incentives   = float(row_data['palawan_pay_out_incentives']    or 0)
        so_lotes     = int(row_data['palawan_send_out_lotes']          or 0)
        sc_lotes     = int(row_data['palawan_sc_lotes']                or 0)
        po_lotes     = int(row_data['palawan_pay_out_lotes']           or 0)
        inc_lotes    = int(row_data['palawan_pay_out_incentives_lotes']or 0)

        palawan_in  = send_out + sc
        palawan_out = pay_out + incentives
        lotes_in    = so_lotes + sc_lotes
        lotes_out   = po_lotes + inc_lotes

        values = [branch_name, palawan_in, lotes_in, palawan_out, lotes_out]
        for col_idx, val in enumerate(values, start=1):
            c = ws.cell(row=data_row, column=col_idx, value=val)
            c.border = styles.border
            if col_idx > 1:
                if col_idx in (3, 5):  # lotes columns → integer
                    c.number_format = '0'
                else:
                    c.number_format = '#,##0.00'
                c.alignment = Alignment(horizontal='right')
            else:
                c.alignment = Alignment(horizontal='left')

        total_in      += palawan_in
        total_out     += palawan_out
        total_lotes_in  += lotes_in
        total_lotes_out += lotes_out
        data_row += 1

    # ── Totals row ────────────────────────────────────────────
    totals = ["TOTAL", total_in, total_lotes_in, total_out, total_lotes_out]
    for col_idx, val in enumerate(totals, start=1):
        c = ws.cell(row=data_row, column=col_idx, value=val)
        c.font   = styles.TOTAL_FONT
        c.fill   = styles.TOTAL_FILL
        c.border = styles.border
        if col_idx > 1:
            if col_idx in (3, 5):
                c.number_format = '0'
            else:
                c.number_format = '#,##0.00'
            c.alignment = Alignment(horizontal='right')
        else:
            c.alignment = Alignment(horizontal='left')

    # ── Column widths ─────────────────────────────────────────
    ws.column_dimensions['A'].width = 40
    for col_letter in ['B', 'C', 'D', 'E']:
        ws.column_dimensions[col_letter].width = 16

