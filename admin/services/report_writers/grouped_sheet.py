import logging
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from .shared import write_info, get_table_cols

logger = logging.getLogger(__name__)


def write_grouped_sheet(ws, ctx, styles, col_groups, table,
                        use_branch_join=False, category_filter=None,
                        show_lotes_total=False, show_amt_total=True):
    write_info(ws, ctx, styles)

    table_cols = get_table_cols(ctx.db, table, ctx.col_cache)

    flat = []
    needed = set()
    for grp, subs in col_groups:
        for sub_label, db_cols, is_lotes in subs:
            valid = [c for c in db_cols if c and (not table_cols or c in table_cols)]
            if not valid:
                continue
            needed.update(valid)
            flat.append((sub_label, valid[0], is_lotes, grp, valid))

    if not flat:
        ws['A7'] = "No columns defined."
        return

    # Column mapping between payable_tbl_brand_a and daily_reports_brand_a
    column_mapping = {
        "sendout_capital": "palawan_sendout_principal",
        "sendout_sc": "palawan_sendout_sc",
        "sendout_commission": "palawan_sendout_commission",
        "sendout_total": "palawan_sendout_regular_total",
        "sendout_lotes": "palawan_sendout_lotes_total",
        "payout_capital": "palawan_payout_principal",
        "payout_sc": "palawan_payout_sc",
        "payout_commission": "palawan_payout_commission",
        "payout_total": "palawan_payout_regular_total",
        "payout_lotes": "palawan_payout_lotes_total",
        "international_capital": "palawan_international_principal",
        "international_sc": "palawan_international_sc",
        "international_commission": "palawan_international_commission",
        "international_total": "palawan_international_regular_total",
        "international_lotes": "palawan_international_lotes_total",
        "inc": "palawan_pay_out_incentives",
        "skid": "palawan_suki_discounts",
        "skir": "palawan_suki_rebates",
        "cancellation": "palawan_cancel",
    }

    # Build SELECT: for multi-col entries, wrap in SUM(COALESCE(...)+COALESCE(...))
    seen_keys    = set()
    select_parts = []
    select_parts_daily = []  # Separate select parts for daily_reports tables
    for _, primary_key, _, _, all_cols in flat:
        if primary_key in seen_keys:
            continue
        seen_keys.add(primary_key)
        if len(all_cols) == 1:
            col_name = all_cols[0]
            daily_col = column_mapping.get(col_name, col_name)

            if table == "payable_tbl_brand_a" and col_name == "inc":
                select_parts.append(f"MAX(COALESCE(dr.`{col_name}`, 0)) AS `{col_name}`")
                select_parts_daily.append(f"MAX(COALESCE(dr.`{daily_col}`, 0)) AS `{col_name}`")
            else:
                select_parts.append(f"SUM(COALESCE(dr.`{col_name}`, 0)) AS `{col_name}`")
                select_parts_daily.append(f"SUM(COALESCE(dr.`{daily_col}`, 0)) AS `{col_name}`")
        else:
            # Sum multiple columns into the first key
            inner = " + ".join(f"COALESCE(dr.`{c}`, 0)" for c in all_cols)
            inner_daily = " + ".join(f"COALESCE(dr.`{column_mapping.get(c, c)}`, 0)" for c in all_cols)
            select_parts.append(f"SUM({inner}) AS `{all_cols[0]}`")
            select_parts_daily.append(f"SUM({inner_daily}) AS `{all_cols[0]}`")

    # ── Build WHERE clause with corp/os filter ────────────────
    cat_join = ""  # extra JOIN for 30%/60% category filter

    if ctx.filter_type == "corporation" and not use_branch_join:
        if category_filter == "30":
            cat_join = ("INNER JOIN branches b ON b.name COLLATE utf8mb4_general_ci "
                        "= dr.branch COLLATE utf8mb4_general_ci")
            where_clause = "AND dr.corporation = %s AND b.global_tag = 'GLOBAL'"
            sql_params   = (ctx.selected_date, ctx.filter_value)
        elif category_filter == "60":
            # 60% = non-global-tagged branches only
            cat_join = (
                "INNER JOIN branches b ON b.name COLLATE utf8mb4_general_ci "
                "= dr.branch COLLATE utf8mb4_general_ci"
            )
            where_clause = "AND dr.corporation = %s AND (b.global_tag IS NULL OR UPPER(TRIM(b.global_tag)) <> 'GLOBAL')"
            sql_params   = (ctx.selected_date, ctx.filter_value)
        else:
            where_clause = "AND dr.corporation = %s"
            sql_params   = (ctx.selected_date, ctx.filter_value)
    elif ctx.filter_type == "corporation" and use_branch_join:
        where_clause = "AND b.global_tag = 'GLOBAL'"
        sql_params   = (ctx.selected_date,)
    else:  # os/group – always need branches join
        if category_filter == "30":
            where_clause = "AND b.global_tag = 'GLOBAL'"
        elif category_filter == "60":
            # 60% = non-global-tagged branches only
            where_clause = "AND (b.global_tag IS NULL OR UPPER(TRIM(b.global_tag)) <> 'GLOBAL')"
        else:
            # For global-tagged tables (e.g. global_other_services_tbl),
            # restrict to branches with a global tag
            where_clause = "AND b.global_tag = 'GLOBAL'" if use_branch_join else ""
        sql_params = (ctx.selected_date,)

    try:
        if category_filter:
            # ── Branches-first approach with fallback for payable 60%/30% sheets ──────────
            # Query payable_tbl_brand_a first (primary source with backfilled data)
            # Then fallback to daily_reports/daily_reports_brand_a for any missing data
            # This ensures we use the clean consolidated table while maintaining backward compatibility
            global_where = ""
            if category_filter == "30":
                global_where = "AND b.global_tag = 'GLOBAL'"
            # 60%: show all branches (including global-tagged), highlight in yellow

            # Fallback approach: prioritize payable_tbl_brand_a (new clients) but fall back to daily_reports (old clients)
            # This supports mixed migration where both old and new clients coexist
            if ctx.filter_type == "os":
                if table == "payable_tbl_brand_a":
                    # UNION: payable_tbl_brand_a (primary) + daily_reports (fallback)
                    # Use DISTINCT to deduplicate identical branches
                    sql = (
                        f"SELECT DISTINCT b.name AS branch, MAX(b.global_tag) AS global_tag, {', '.join(select_parts)} "
                        f"FROM branches b "
                        f"LEFT JOIN payable_tbl_brand_a dr "
                        f"  ON b.name COLLATE utf8mb4_general_ci = dr.branch COLLATE utf8mb4_general_ci "
                        f"  AND dr.date = %s "
                        f"WHERE b.os_name = %s {global_where} "
                        f"GROUP BY b.name "
                        f"UNION "
                        f"SELECT DISTINCT b.name AS branch, MAX(b.global_tag) AS global_tag, {', '.join(select_parts_daily)} "
                        f"FROM branches b "
                        f"LEFT JOIN daily_reports_brand_a dr "
                        f"  ON b.name COLLATE utf8mb4_general_ci = dr.branch COLLATE utf8mb4_general_ci "
                        f"  AND dr.date = %s "
                        f"WHERE b.os_name = %s {global_where} "
                        f"AND NOT EXISTS (SELECT 1 FROM payable_tbl_brand_a p INNER JOIN corporations pc ON p.corporation COLLATE utf8mb4_general_ci = pc.name COLLATE utf8mb4_general_ci WHERE p.branch COLLATE utf8mb4_general_ci = b.name COLLATE utf8mb4_general_ci AND pc.id = COALESCE(b.sub_corporation_id, b.corporation_id) AND p.date = %s) "
                        f"GROUP BY b.name "
                        f"ORDER BY branch"
                    )
                    sql_params = (ctx.selected_date, ctx.filter_value, ctx.selected_date, ctx.filter_value, ctx.selected_date)
                else:
                    sql = (
                        f"SELECT b.name AS branch, MAX(b.global_tag) AS global_tag, {', '.join(select_parts)} "
                        f"FROM branches b "
                        f"INNER JOIN corporations c "
                        f"  ON c.id = COALESCE(b.sub_corporation_id, b.corporation_id) "
                        f"LEFT JOIN `{table}` dr "
                        f"  ON b.name COLLATE utf8mb4_general_ci = dr.branch COLLATE utf8mb4_general_ci "
                        f"  AND dr.corporation COLLATE utf8mb4_general_ci = c.name COLLATE utf8mb4_general_ci "
                        f"  AND dr.date = %s "
                        f"WHERE b.os_name = %s {global_where} "
                        f"GROUP BY b.name ORDER BY b.name"
                    )
                    sql_params = (ctx.selected_date, ctx.filter_value)
            else:
                # corporation filter – join corps table to filter by name
                # Also apply registration filter if specified
                if ctx.reg_filter == "registered":
                    reg_clause_corp = "AND b.is_registered = 1"
                elif ctx.reg_filter == "not_registered":
                    reg_clause_corp = "AND (b.is_registered = 0 OR b.is_registered IS NULL)"
                else:
                    reg_clause_corp = ""

                if table == "payable_tbl_brand_a":
                    # UNION: payable_tbl_brand_a (primary) + daily_reports (fallback)
                    # Use DISTINCT to deduplicate identical branches
                    sql = (
                        f"SELECT DISTINCT b.name AS branch, MAX(b.global_tag) AS global_tag, {', '.join(select_parts)} "
                        f"FROM branches b "
                        f"INNER JOIN corporations c "
                        f"  ON (c.id = b.corporation_id OR c.id = b.sub_corporation_id) "
                        f"LEFT JOIN payable_tbl_brand_a dr "
                        f"  ON b.name COLLATE utf8mb4_general_ci = dr.branch COLLATE utf8mb4_general_ci "
                        f"  AND dr.date = %s "
                        f"WHERE c.name COLLATE utf8mb4_general_ci = %s {global_where} {reg_clause_corp} "
                        f"GROUP BY b.name "
                        f"UNION "
                        f"SELECT DISTINCT b.name AS branch, MAX(b.global_tag) AS global_tag, {', '.join(select_parts_daily)} "
                        f"FROM branches b "
                        f"INNER JOIN corporations c "
                        f"  ON (c.id = b.corporation_id OR c.id = b.sub_corporation_id) "
                        f"LEFT JOIN daily_reports_brand_a dr "
                        f"  ON b.name COLLATE utf8mb4_general_ci = dr.branch COLLATE utf8mb4_general_ci "
                        f"  AND dr.date = %s "
                        f"WHERE c.name COLLATE utf8mb4_general_ci = %s {global_where} {reg_clause_corp} "
                        f"AND NOT EXISTS (SELECT 1 FROM payable_tbl_brand_a p WHERE p.branch COLLATE utf8mb4_general_ci = b.name COLLATE utf8mb4_general_ci AND p.corporation COLLATE utf8mb4_general_ci = c.name COLLATE utf8mb4_general_ci AND p.date = %s) "
                        f"GROUP BY b.name "
                        f"ORDER BY branch"
                    )
                    sql_params = (ctx.selected_date, ctx.filter_value, ctx.selected_date, ctx.filter_value, ctx.selected_date)
                else:
                    sql = (
                        f"SELECT b.name AS branch, MAX(b.global_tag) AS global_tag, {', '.join(select_parts)} "
                        f"FROM branches b "
                        f"INNER JOIN corporations c "
                        f"  ON (c.id = b.corporation_id OR c.id = b.sub_corporation_id) "
                        f"LEFT JOIN `{table}` dr "
                        f"  ON b.name COLLATE utf8mb4_general_ci = dr.branch COLLATE utf8mb4_general_ci "
                        f"  AND dr.corporation COLLATE utf8mb4_general_ci = c.name COLLATE utf8mb4_general_ci "
                        f"  AND dr.date = %s "
                        f"WHERE c.name COLLATE utf8mb4_general_ci = %s {global_where} {reg_clause_corp} "
                        f"GROUP BY b.name ORDER BY b.name"
                    )
                    sql_params = (ctx.selected_date, ctx.filter_value)

        elif ctx.filter_type == "os":
            if ctx.reg_filter == "registered":
                reg_clause_gs = "AND b.is_registered = 1"
            elif ctx.reg_filter == "not_registered":
                reg_clause_gs = "AND (b.is_registered = 0 OR b.is_registered IS NULL)"
            else:
                reg_clause_gs = ""
            sql_params = (ctx.selected_date, ctx.filter_value)
            sql = (
                f"SELECT b.name AS branch, {', '.join(select_parts)} "
                f"FROM branches b "
                f"LEFT JOIN `{table}` dr "
                f"  ON b.name COLLATE utf8mb4_general_ci = dr.branch COLLATE utf8mb4_general_ci "
                f"  AND dr.date = %s "
                f"WHERE b.os_name = %s {reg_clause_gs} "
                f"GROUP BY b.name ORDER BY b.name"
            )
        elif use_branch_join:
            # Corporation + use_branch_join (e.g. Global Other Services).
            # Must still restrict to the selected corporation — add corps join.
            if ctx.reg_filter == "registered":
                reg_clause_bj = "AND b.is_registered = 1"
            elif ctx.reg_filter == "not_registered":
                reg_clause_bj = "AND (b.is_registered = 0 OR b.is_registered IS NULL)"
            else:
                reg_clause_bj = ""
            sql = (
                f"SELECT b.name AS branch, {', '.join(select_parts)} "
                f"FROM branches b "
                f"INNER JOIN corporations c "
                f"  ON (c.id = b.corporation_id OR c.id = b.sub_corporation_id) "
                f"LEFT JOIN `{table}` dr "
                f"  ON b.name COLLATE utf8mb4_general_ci = dr.branch COLLATE utf8mb4_general_ci "
                f"  AND dr.date = %s "
                f"WHERE c.name COLLATE utf8mb4_general_ci = %s {where_clause} {reg_clause_bj} "
                f"GROUP BY b.name ORDER BY b.name"
            )
            sql_params = (ctx.selected_date, ctx.filter_value)
        else:
            # ── Branches-first approach for corporation filter without category ──
            # Include branches even if they have no entries for the selected date.
            if ctx.filter_type == "corporation":
                # For corporation filter, join corporations to filter by name
                # Also apply registration filter if specified
                if ctx.reg_filter == "registered":
                    reg_clause_corp = "AND b.is_registered = 1"
                elif ctx.reg_filter == "not_registered":
                    reg_clause_corp = "AND (b.is_registered = 0 OR b.is_registered IS NULL)"
                else:
                    reg_clause_corp = ""

                if table == "payable_tbl_brand_a":
                    sql = (
                        f"SELECT b.name AS branch, {', '.join(select_parts)} "
                        f"FROM branches b "
                        f"INNER JOIN corporations c "
                        f"  ON (c.id = b.corporation_id OR c.id = b.sub_corporation_id) "
                        f"LEFT JOIN `{table}` dr "
                        f"  ON b.name COLLATE utf8mb4_general_ci = dr.branch COLLATE utf8mb4_general_ci "
                        f"  AND dr.date = %s "
                        f"WHERE c.name COLLATE utf8mb4_general_ci = %s {reg_clause_corp} "
                        f"GROUP BY b.name ORDER BY b.name"
                    )
                else:
                    sql = (
                        f"SELECT b.name AS branch, {', '.join(select_parts)} "
                        f"FROM branches b "
                        f"INNER JOIN corporations c "
                        f"  ON (c.id = b.corporation_id OR c.id = b.sub_corporation_id) "
                        f"LEFT JOIN `{table}` dr "
                        f"  ON b.name COLLATE utf8mb4_general_ci = dr.branch COLLATE utf8mb4_general_ci "
                        f"  AND dr.corporation COLLATE utf8mb4_general_ci = c.name COLLATE utf8mb4_general_ci "
                        f"  AND dr.date = %s "
                        f"WHERE c.name COLLATE utf8mb4_general_ci = %s {reg_clause_corp} "
                        f"GROUP BY b.name ORDER BY b.name"
                    )
                sql_params = (ctx.selected_date, ctx.filter_value)
            else:
                # Fallback for non-corporation filters (shouldn't normally reach here)
                sql = (
                    f"SELECT dr.branch, {', '.join(select_parts)} "
                    f"FROM `{table}` dr {cat_join} "
                    f"WHERE dr.date = %s {where_clause} "
                    f"GROUP BY dr.branch ORDER BY dr.branch"
                )
                # sql_params already set from earlier
        logger.debug(f"Report sheet: {ws.title} | Table: {table} | Filter: {ctx.filter_type}={ctx.filter_value}")
        logger.debug(f"  Reg Filter: {ctx.reg_filter} | Category: {category_filter} | Use Branch Join: {use_branch_join}")
        logger.debug(f"Query: {sql[:400]}...")
        logger.debug(f"Params: {sql_params}")
        results = ctx.db.execute_query(sql, sql_params) or []
        logger.debug(f"Results: {len(results)} rows returned for {ws.title}")

        # If no results for corporation filter, debug why
        if not results and ctx.filter_type == "corporation":
            logger.warning(f"No branches found for corporation '{ctx.filter_value}'. Debugging...")
            corp_check = ctx.db.execute_query(
                "SELECT COUNT(*) as count FROM corporations WHERE name COLLATE utf8mb4_general_ci = %s",
                (ctx.filter_value,)
            )
            if corp_check and corp_check[0]['count'] > 0:
                logger.warning(f"  → Corporation '{ctx.filter_value}' exists")
                branch_check = ctx.db.execute_query(
                    "SELECT COUNT(*) as count FROM branches b WHERE "
                    "COALESCE(b.sub_corporation_id, b.corporation_id) = "
                    "(SELECT id FROM corporations WHERE name COLLATE utf8mb4_general_ci = %s)",
                    (ctx.filter_value,)
                )
                if branch_check:
                    logger.warning(f"  → {branch_check[0]['count']} branches found for this corporation")
            else:
                logger.warning(f"  → Corporation '{ctx.filter_value}' NOT FOUND in database")
    except Exception as ex:
        logger.error(f"Report query failed for sheet {ws.title}: {ex}")
        logger.error(f"SQL: {sql}")
        logger.error(f"Params: {sql_params}")
        ws['A7'] = f"Error loading data: {ex}"
        return

    GRP_HDR_ROW = 7   # rows 1-6 are title/info
    SUB_HDR_ROW = 8   # sub-label row
    HDR_ROW     = SUB_HDR_ROW  # data starts after this

    ws.merge_cells(start_row=GRP_HDR_ROW, start_column=1,
                   end_row=SUB_HDR_ROW,   end_column=1)
    c = ws.cell(row=GRP_HDR_ROW, column=1, value="Branch")
    c.font = styles.HDR_FONT; c.fill = styles.HDR_FILL
    c.border = styles.border
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Build group → list of column indices
    col_idx = 2
    grp_spans = []   # (group_name, start_col, end_col)
    for label, _, _, grp, _ in flat:
        grp_spans.append((grp, col_idx))
        col_idx += 1

    # Merge consecutive columns that share the same group name
    merged_groups = []
    i = 0
    while i < len(grp_spans):
        grp_name, start = grp_spans[i]
        end = start
        while i + 1 < len(grp_spans) and grp_spans[i + 1][0] == grp_name:
            i += 1
            end = grp_spans[i][1]
        merged_groups.append((grp_name, start, end))
        i += 1

    for grp_name, start, end in merged_groups:
        if start != end:
            ws.merge_cells(start_row=GRP_HDR_ROW, start_column=start,
                           end_row=GRP_HDR_ROW,   end_column=end)
        c = ws.cell(row=GRP_HDR_ROW, column=start, value=grp_name.upper())
        c.font  = styles.HDR_FONT
        c.fill  = styles.group_fill(grp_name)
        c.border = styles.border
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        # Apply border to every cell in the merged range
        for ci in range(start, end + 1):
            ws.cell(row=GRP_HDR_ROW, column=ci).border = styles.border

    # "TOTAL LOTES" column (optional)
    lotes_total_col = None
    if show_lotes_total:
        lotes_total_col = col_idx
        ws.merge_cells(start_row=GRP_HDR_ROW, start_column=lotes_total_col,
                       end_row=SUB_HDR_ROW,   end_column=lotes_total_col)
        c = ws.cell(row=GRP_HDR_ROW, column=lotes_total_col, value="TOTAL LOTES")
        c.font = styles.HDR_FONT; c.fill = styles.LOTES_FILL
        c.border = styles.border
        c.alignment = Alignment(horizontal='center', vertical='center')
        col_idx += 1

    # "AMT TOTAL" column (optional)
    total_col = None
    if show_amt_total:
        total_col = col_idx
        ws.merge_cells(start_row=GRP_HDR_ROW, start_column=total_col,
                       end_row=SUB_HDR_ROW,   end_column=total_col)
        c = ws.cell(row=GRP_HDR_ROW, column=total_col, value="AMT TOTAL")
        c.font = styles.HDR_FONT; c.fill = styles.GRAND_FILL
        c.border = styles.border
        c.alignment = Alignment(horizontal='center', vertical='center')

    # ── Row 8: Sub-label headers ──────────────────────────────────
    ws.cell(row=SUB_HDR_ROW, column=1).border = styles.border  # already merged

    col_idx = 2
    for label, _, _, grp, _ in flat:
        c = ws.cell(row=SUB_HDR_ROW, column=col_idx, value=label.upper())
        c.font  = styles.HDR_FONT
        c.fill  = styles.group_fill(grp)
        c.border = styles.border
        c.alignment = Alignment(horizontal='center', wrap_text=True)
        col_idx += 1

    if total_col is not None:
        ws.cell(row=SUB_HDR_ROW, column=total_col).border = styles.border  # already merged

    # Data rows
    grand_totals = {primary_key: 0.0 for _, primary_key, _, _, _ in flat}
    grand_lotes_total = 0
    for ri, row_data in enumerate(results, HDR_ROW + 1):
        branch_cell = ws.cell(row=ri, column=1, value=row_data.get('branch', ''))
        branch_cell.border = styles.border
        # Yellow fill for globally-tagged branches in 60% sheet
        if category_filter == "60" and (row_data.get('global_tag') or '').upper() == 'GLOBAL':
            branch_cell.fill = styles.GLOBAL_FILL
        row_amt = 0.0
        row_lotes = 0
        for ci, (_, primary_key, is_lotes, _, all_cols) in enumerate(flat, 2):
            raw = row_data.get(primary_key, 0) or 0
            val = int(float(raw)) if is_lotes else float(raw)
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = styles.border
            if is_lotes:
                cell.alignment = Alignment(horizontal='center')
                cell.fill = styles.LOTES_FILL
                row_lotes += int(float(raw))
                grand_totals[primary_key] = grand_totals.get(primary_key, 0.0) + int(float(raw))
            else:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
                row_amt += float(val)
                grand_totals[primary_key] = grand_totals.get(primary_key, 0.0) + float(val)

        if lotes_total_col is not None:
            c = ws.cell(row=ri, column=lotes_total_col, value=row_lotes)
            c.alignment = Alignment(horizontal='center')
            c.fill = styles.LOTES_FILL; c.border = styles.border
            grand_lotes_total += row_lotes

        if total_col is not None:
            c = ws.cell(row=ri, column=total_col, value=row_amt)
            c.number_format = '#,##0.00'; c.font = styles.TOTAL_FONT
            c.border = styles.border; c.fill = styles.TOTAL_FILL
            c.alignment = Alignment(horizontal='right')

    # Total row
    tr = HDR_ROW + len(results) + 1
    c = ws.cell(row=tr, column=1, value="TOTAL")
    c.font = styles.TOTAL_FONT; c.fill = styles.TOTAL_FILL; c.border = styles.border

    grand_sum = 0.0
    for ci, (_, primary_key, is_lotes, _, _) in enumerate(flat, 2):
        val = grand_totals.get(primary_key, 0.0)
        cell = ws.cell(row=tr, column=ci, value=int(val) if is_lotes else val)
        cell.font = styles.TOTAL_FONT; cell.fill = styles.TOTAL_FILL; cell.border = styles.border
        if is_lotes:
            cell.alignment = Alignment(horizontal='center')
        else:
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal='right')
            grand_sum += val

    if lotes_total_col is not None:
        lotes_grand = sum(int(grand_totals.get(pk, 0)) for _, pk, is_l, _, _ in flat if is_l)
        c = ws.cell(row=tr, column=lotes_total_col, value=lotes_grand)
        c.alignment = Alignment(horizontal='center')
        c.font = styles.TOTAL_FONT; c.fill = styles.GRAND_FILL; c.border = styles.border

    if total_col is not None:
        c = ws.cell(row=tr, column=total_col, value=grand_sum)
        c.number_format = '#,##0.00'; c.font = styles.TOTAL_FONT
        c.fill = styles.GRAND_FILL; c.border = styles.border
        c.alignment = Alignment(horizontal='right')

    # Column widths
    ws.column_dimensions['A'].width = 25
    for ci, (_, _, is_lotes, _, _) in enumerate(flat, 2):
        ws.column_dimensions[get_column_letter(ci)].width = 8 if is_lotes else 13
    if lotes_total_col is not None:
        ws.column_dimensions[get_column_letter(lotes_total_col)].width = 12
    if total_col is not None:
        ws.column_dimensions[get_column_letter(total_col)].width = 14
    ws.freeze_panes = f'B{HDR_ROW + 1}'
