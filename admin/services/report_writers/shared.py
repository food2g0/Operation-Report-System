import logging
from openpyxl.styles import Font, Alignment

logger = logging.getLogger(__name__)

BRANCH_JOIN_CLAUSE  = "b.name COLLATE utf8mb4_general_ci = dr.branch COLLATE utf8mb4_general_ci"
EXCEL_DATE_COL_WIDTH = 18


def write_info(ws, ctx, styles):
    """Write the standard 6-row info header used by most sheets."""
    import datetime as _dt
    ws['A1'] = f"{ctx.brand_label} – {ws.title}"
    ws['A1'].font      = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill      = styles.TITLE_FILL
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 22

    ws['A3'] = "Brand:";              ws['B3'] = ctx.brand_label
    ws['A4'] = f"{ctx.filter_label}:"; ws['B4'] = ctx.filter_value
    ws['A5'] = "Date:";               ws['B5'] = ctx.selected_date
    ws['A6'] = "Generated:";          ws['B6'] = _dt.datetime.now().strftime("%Y-%m-%d %H:%M")
    for r in range(3, 7):
        ws.cell(row=r, column=1).font = styles.BOLD_FONT
        ws.cell(row=r, column=2).font = Font(size=10)
        ws.cell(row=r, column=1).fill = styles.INFO_FILL
        ws.cell(row=r, column=2).fill = styles.INFO_FILL


def get_table_cols(db, tbl, cache):
    """Return set of column names for `tbl`, cached in `cache` dict."""
    if tbl not in cache:
        try:
            rows = db.execute_query(
                "SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS "
                "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = %s",
                (tbl,)
            )
            cache[tbl] = {r['COLUMN_NAME'] for r in rows} if rows else set()
        except Exception:
            cache[tbl] = set()
    return cache[tbl]
