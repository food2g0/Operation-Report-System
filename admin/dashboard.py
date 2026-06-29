from PyQt5.QtWidgets import (
    QWidget, QLabel, QVBoxLayout, QHBoxLayout, QGroupBox, QFormLayout,
    QLineEdit, QComboBox, QPushButton, QMessageBox, QDateEdit, QStackedWidget,
    QScrollArea, QFrame, QFileDialog, QDialog, QTableWidget, QTableWidgetItem, QHeaderView,
    QApplication, QSizePolicy, QCheckBox, QGridLayout, QRadioButton,
)
from PyQt5.QtGui import QFont
from PyQt5.QtCore import Qt, QDate, QEvent, pyqtSignal, QTimer
from api_db_manager import db_manager
from security import SessionManager
import json
import logging
import os
import re

logger = logging.getLogger(__name__)
import sys

from admin.constants import COLORS, FONT_SIZES, CONFIG, MESSAGES
from admin.widgets_util import MoneyInput, LotesInput, DisplayField

from admin.pages.palawan import PalawanPage
from admin.pages.mc import MCPage
from admin.pages.fund_transfer import FundTransferPage
from admin.pages.payable import PayablesPage
from admin.pages.global_payable import GlobalPayablePage
from admin.pages.report import ReportPage
from admin.manage import (
    create_corporation, create_branch, create_client, get_all_supervisors,
    ensure_payroll_column,
)
from admin.pages.variance_review import VarianceReviewPage
from admin.user_management import UserManagementPage
from admin.pages.daily_transaction import DailyTransactionPage
from admin.pages.new_sanla import NewSanlaPage
from admin.pages.new_renew import NewRenewPage
from admin.pages.global_other_services import GlobalOtherServicesPage
from admin.pages.ft_ho import FTHOPage
from admin.pages.depo_br import DepoBRPage
from admin.pages.review_summary import ReviewSummaryPage
from admin.pages.bir_book import BIRBookPage
from connection_watcher import ConnectionWatcher, ConnectionBanner
from admin.styles import ADMIN_STYLESHEET

from admin.services.balance_service import (
    parse_money      as _parse_money_svc,
    calculate_balances as _calculate_balances_svc,
)
from admin.services.reset_service import unlock_entry, clear_supplement_tables
from admin.services.entry_service import (
    load_entry_data, patch_brand_b_palawan,
    build_save_payload, persist_entry, upsert_palawan_payable,
)
from admin.services.excel_export_service import export_daily_cash as _export_daily_cash_svc

from Client.salary_detail_dialog import SalaryDetailDialog
from Client.dashboard.dialogs import FundTransferHODialog, MotorCarDetailDialog, EmpenaDetailDialog

try:
    import requests
    from api_config import API_URL, API_KEY
    _REQUESTS_AVAILABLE = True
except ImportError:
    _REQUESTS_AVAILABLE = False
    API_URL = API_KEY = ""

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

try:
    from auto_updater import check_for_updates, check_update_success
    from version import __version__
    AUTO_UPDATE_ENABLED = True
except ImportError:
    AUTO_UPDATE_ENABLED = False
    __version__ = "1.0.0"
    check_update_success = None

# ── Brand identity ────────────────────────────────────────────────────────────
BRAND_A_TYPE = 1

# ── SQL helpers ───────────────────────────────────────────────────────────────
_BRANCH_JOIN = "b.name COLLATE utf8mb4_general_ci = dr.branch COLLATE utf8mb4_general_ci"
BRAND_B_TYPE = 2

# ── Session / timers ──────────────────────────────────────────────────────────
_SESSION_TIMEOUT_SECONDS = 1800
_SESSION_TICK_MS         = 60_000

# ── Variance ──────────────────────────────────────────────────────────────────
_BALANCED_THRESHOLD = 0.01

# ── Zoom ──────────────────────────────────────────────────────────────────────
_ZOOM_MIN      = 50
_ZOOM_MAX      = 200
_FONT_SIZE_MIN = 1
_FONT_SIZE_MAX = 500

# ── UI dimensions ─────────────────────────────────────────────────────────────
_INFO_DIALOG_MIN_WIDTH    = 400   # bank / branch / from-branch dialogs
_FULL_REPORT_DIALOG_WIDTH = 460
_EXCEL_DATE_COL_WIDTH     = 18    # standard date-column width in generated sheets


class AdminDashboard(QWidget):
    logout_requested = pyqtSignal()
    
    def __init__(self, account_type=2, os_group=""):
        super().__init__()

        self.account_type = account_type
        self.os_group = os_group or ""   # assigned group; "" = no restriction
        brand_label = self._brand_label
        group_label = f" — {self.os_group}" if self.os_group else ""
        self.setWindowTitle(f"Operation Report System - Admin Dashboard ({brand_label}){group_label}")
        self.db = db_manager
        self._update_checker_threads = []
        self._loading_report = False  # Flag to prevent recalculation during load

        ensure_payroll_column()

        # Zoom functionality
        self.zoom_level = 100
        self.setFocusPolicy(Qt.StrongFocus)

        self.session = SessionManager(inactivity_timeout=_SESSION_TIMEOUT_SECONDS)
        self._session_timer = QTimer(self)
        self._session_timer.timeout.connect(self._check_session_timeout)
        self._session_timer.start(_SESSION_TICK_MS)

        # Network connectivity monitor
        self._is_connected = True
        self._conn_watcher = ConnectionWatcher(self.db)
        self._conn_watcher.connection_lost.connect(self._on_connection_lost)
        self._conn_watcher.connection_restored.connect(self._on_connection_restored)
        self._conn_watcher.start()

        self.debit_inputs = {}
        self.credit_inputs = {}
        self.debit_lotes_inputs = {}
        self.credit_lotes_inputs = {}
        self.selected_bank_account = None  
        self.bank_account_btn = None 
        self.selected_branch_dest = None
        self.branch_dest_btn = None
        self.selected_from_branch_dest = None 
        self.from_branch_dest_btn = None

      
        brand_key = self._brand_label
        field_config = self._load_field_config()
        
        if field_config and brand_key in field_config:
            brand_config = field_config[brand_key]

            self.debit_fields = {item[0]: item[2] for item in brand_config.get("debit", [])}
            self.credit_fields = {item[0]: item[2] for item in brand_config.get("credit", [])}
        else:
  
            self.debit_fields = {
                "Rescate Jewelry": "rescate_jewelry",
                "Interest": "interest",
                "Penalty": "penalty",
                "Stamp": "stamp",
                "Resguardo/Affidavit": "resguardo_affidavit",
                "HABOL Renew/Tubos": "habol_renew_tubos",
                "Habol R/T Interest&Stamp": "habol_rt_interest_stamp",
                "Jew. A.I": "jew_ai",
                "S.C": "sc",
                "Fund Transfer": "fund_transfer_from_branch",
                "Sendah Load + SC": "sendah_load_sc",
                "PPAY CO SC": "ppay_co_sc",
                "Palawan Send Out": "palawan_send_out",
                "Palawan S.C": "palawan_sc",
                "Palawan Suki Card": "palawan_suki_card",
                "Palawan Pay Cash-In + SC": "palawan_pay_cash_in_sc",
                "Palawan Pay Bills + SC": "palawan_pay_bills_sc",
                "Palawan Load": "palawan_load",
                "Palawan Change Receiver": "palawan_change_receiver",
                "MC In": "mc_in",
                "Handling fee": "handling_fee",
                "Other Penalty": "other_penalty",
                "Cash Overage": "cash_overage"
            }
            self.credit_fields = {
                "Empeno JEW. (NEW)": "empeno_jew_new",
                "Empeno JEW (RENEW)": "empeno_jew_renew",
                "Empeno Motor/Car": "empeno_motor_car",
                "Fund Transfer to HEAD OFFICE": "fund_transfer_to_head_office",
                "Fund Transfer to BRANCH": "fund_transfer_to_branch",
                "Palawan Pay Out": "palawan_pay_out",
                "Palawan Pay Out (incentives)": "palawan_pay_out_incentives",
                "Palawan Pay Cash Out": "palawan_pay_cash_out",
                "MC Out": "mc_out",
                "PC-Salary": "pc_salary",
                "PC-Rental": "pc_rental",
                "PC-Electric": "pc_electric",
                "PC-Water": "pc_water",
                "PC-Internet": "pc_internet",
                "PC-Lbc/Jrs/Jnt": "pc_lbc_jrs_jnt",
                "PC-Permits/BIR Payments": "pc_permits_bir_payments",
                "PC-Supplies/Xerox/Maintenance": "pc_supplies_xerox_maintenance",
                "PC-Transpo": "pc_transpo",
                "Palawan Cancel": "palawan_cancel",
                "Palawan Suki Discounts": "palawan_suki_discounts",
                "Palawan Suki Rebates": "palawan_suki_rebates",
                "OTHERS": "others",
                "Cash Shortage": "cash_shortage"
            }


        self.daily_table = "daily_reports_brand_a" if self._is_brand_a else "daily_reports"

        self.setup_styles()
        self.build_ui()


        self._capture_base_fonts()

        QTimer.singleShot(0,   self.load_corporations)     # populates corp/group selectors
        QTimer.singleShot(200, self._ensure_review_table)  # DDL check; not time-critical
        
        # Install event filter on application for zoom
        QApplication.instance().installEventFilter(self)

        if AUTO_UPDATE_ENABLED and check_update_success:
            check_update_success(parent=self)
    
    def _capture_base_fonts(self):

        for w in self._get_zoom_target_widgets():
            font = w.font()
            point_size = font.pointSize()
            pixel_size = font.pixelSize()

            base_stylesheet = w.styleSheet() or ""
            w.setProperty('_base_stylesheet', base_stylesheet)
            w.setProperty('_base_zoom_height', max(w.minimumHeight(), w.sizeHint().height()))
            

            if point_size < 0 and pixel_size < 0:
                # Get from application default
                app_font = QApplication.font()
                point_size = app_font.pointSize()
                if point_size < 0:
                    point_size = 10  # Default fallback
            
            # Store whichever is valid
            if point_size > 0:
                w.setProperty('_base_point_size', point_size)
            elif pixel_size > 0:
                w.setProperty('_base_pixel_size', pixel_size)

    def _get_zoom_target_widgets(self):
        """Return only debit/credit amount inputs that should respond to zoom."""
        targets = []
        seen = set()
        for inp in list(self.debit_inputs.values()) + list(self.credit_inputs.values()):
            if inp is None:
                continue
            wid = id(inp)
            if wid in seen:
                continue
            seen.add(wid)
            targets.append(inp)
        return targets

    def _scale_stylesheet_font_sizes(self, stylesheet, zoom_factor):
        """Scale font-size declarations in a stylesheet by zoom factor."""
        if not stylesheet or "font-size" not in stylesheet:
            return stylesheet

        def _replace(match):
            base_size = float(match.group(1))
            scaled = max(_FONT_SIZE_MIN, min(_FONT_SIZE_MAX, int(round(base_size * zoom_factor))))
            return f"font-size: {scaled}px"

        return re.sub(
            r"font-size\s*:\s*([0-9]*\.?[0-9]+)\s*px",
            _replace,
            stylesheet,
            flags=re.IGNORECASE,
        )

    def _apply_zoom_to_all(self):
        """Apply zoom only to debit/credit amount input widgets."""
        zoom_factor = self.zoom_level / 100.0
        for w in self._get_zoom_target_widgets():
            font = w.font()
            new_size = None
            
            # Try point size first
            base_point = w.property('_base_point_size')
            if base_point is not None:
                try:
                    base_point = int(base_point)
                    if base_point > 0:
                        new_size = max(_FONT_SIZE_MIN, min(_FONT_SIZE_MAX, int(base_point * zoom_factor)))
                        font.setPointSize(new_size)
                        w.setFont(font)
                except (ValueError, TypeError, OverflowError):
                    pass
            
            # Try pixel size
            base_pixel = w.property('_base_pixel_size')
            if base_pixel is not None:
                try:
                    base_pixel = int(base_pixel)
                    if base_pixel > 0:
                        new_size = max(_FONT_SIZE_MIN, min(_FONT_SIZE_MAX, int(base_pixel * zoom_factor)))
                        font.setPixelSize(new_size)
                        w.setFont(font)
                except (ValueError, TypeError, OverflowError):
                    pass

            base_stylesheet = w.property('_base_stylesheet')
            if isinstance(base_stylesheet, str):
                scaled_stylesheet = self._scale_stylesheet_font_sizes(base_stylesheet, zoom_factor)
                if new_size is not None and "font-size" not in scaled_stylesheet.lower():
                    scaled_stylesheet = (scaled_stylesheet + f"\nfont-size: {new_size}px;").strip()
                if scaled_stylesheet != w.styleSheet():
                    w.setStyleSheet(scaled_stylesheet)

            base_height = w.property('_base_zoom_height')
            if base_height is not None:
                try:
                    scaled_height = max(20, min(_FONT_SIZE_MAX, int(int(base_height) * zoom_factor)))
                    w.setMinimumHeight(scaled_height)
                    w.updateGeometry()
                except (ValueError, TypeError, OverflowError):
                    pass

        self.update()

    def _ensure_review_table(self):

        try:
            self.db.execute_query("""
                CREATE TABLE IF NOT EXISTS admin_review_marks (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    brand VARCHAR(10) NOT NULL,
                    branch VARCHAR(255) NOT NULL,
                    report_date DATE NOT NULL,
                    reviewed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE KEY uq_review (brand, branch, report_date)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
            """)
        except Exception as e:
            logger.error("[AdminDashboard] review table create: %s", e)

        # Add any columns to daily_reports_brand_a that may be missing on older DB installs
        _migrations = [
            ("daily_reports_brand_a", "pc_inc_insurance",              "DECIMAL(15,2) DEFAULT 0.00"),
            ("daily_reports_brand_a", "pc_inc_insurance_lotes",        "INT DEFAULT 0"),
            ("daily_reports_brand_a", "habol_renew_tubos",             "DECIMAL(15,2) DEFAULT 0.00"),
            ("daily_reports_brand_a", "habol_renew_tubos_lotes",       "INT DEFAULT 0"),
            ("daily_reports_brand_a", "habol_rt_interest_stamp",       "DECIMAL(15,2) DEFAULT 0.00"),
            ("daily_reports_brand_a", "habol_rt_interest_stamp_lotes", "INT DEFAULT 0"),
            ("daily_reports_brand_a", "transfast",                     "DECIMAL(15,2) DEFAULT 0.00"),
            ("daily_reports_brand_a", "transfast_lotes",               "INT DEFAULT 0"),
            # lotes columns for payable_tbl_brand_a (added for PalawanPayableContainer)
            ("payable_tbl_brand_a",   "sendout_lotes",                 "INT DEFAULT 0"),
            ("payable_tbl_brand_a",   "payout_lotes",                  "INT DEFAULT 0"),
            ("payable_tbl_brand_a",   "international_lotes",           "INT DEFAULT 0"),
        ]
        for table, col, typedef in _migrations:
            try:
                exists = self.db.execute_query(
                    "SELECT COUNT(*) as cnt FROM information_schema.COLUMNS "
                    "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = %s AND COLUMN_NAME = %s",
                    (table, col)
                )
                if not exists or exists[0].get('cnt', 0) == 0:
                    self.db.execute_query(
                        f"ALTER TABLE {table} ADD COLUMN {col} {typedef}"
                    )
            except Exception:
                pass  # Safe to ignore

    def _load_field_config(self):

        try:
            result = self.db.execute_query(
                "SELECT config_value FROM field_config WHERE config_key = 'field_definitions'"
            )
            if result and result[0].get('config_value'):
                cfg = json.loads(result[0]['config_value'])
                for brand in ("Brand A", "Brand B"):
                    cfg.setdefault(brand, {})
                    cfg[brand].setdefault("debit", [])
                    cfg[brand].setdefault("credit", [])
                return cfg
        except Exception as e:
            logger.error("[AdminDashboard] Failed to load config from DB: %s", e)

        try:
            if getattr(sys, 'frozen', False):
                config_dir = os.path.dirname(sys.executable)
            else:
                config_dir = os.path.dirname(os.path.abspath(__file__))
            
            config_path = os.path.join(config_dir, 'field_config.json')
            
            if os.path.exists(config_path):
                with open(config_path, 'r', encoding='utf-8') as f:
                    return json.load(f)
            else:
                logger.warning("field_config.json not found at %s", config_path)
                return None
        except Exception as e:
            logger.error("Error loading field_config.json: %s", e)
            return None

    # ── Brand identity helpers ────────────────────────────────────────────────
    @property
    def _is_brand_a(self) -> bool:
        return self.account_type == BRAND_A_TYPE

    @property
    def _brand_label(self) -> str:
        return "Brand A" if self._is_brand_a else "Brand B"

    @staticmethod
    def _calculate_balances(beginning: float, debit_sum: float, credit_sum: float, cash_count: float) -> dict:
        """Return debit_total/credit_total/ending_balance/cash_result/variance_status from raw sums."""
        return _calculate_balances_svc(beginning, debit_sum, credit_sum, cash_count)

    @staticmethod
    def _parse_money(text: str) -> float:
        """Parse a money input string to float, stripping commas and whitespace."""
        return _parse_money_svc(text)

    def _set_variance_display(self, status: str) -> None:
        """Apply text + colour to the variance status badge. status: 'balanced'|'over'|'short'."""
        if status == "short":
            text, bg, fg = "SHORT", "#ffcdd2", "#c62828"
        elif status == "over":
            text, bg, fg = "OVER", "#fff3cd", "#856404"
        else:
            text, bg, fg = "✓ Balanced", "#c8e6c9", "#2e7d32"
        self.variance_status_display.setText(text)
        self.variance_status_display.setStyleSheet(
            f"font-weight: bold; font-size: 12px; padding: 5px 10px; border-radius: 4px;"
            f" background-color: {bg}; color: {fg};"
        )

    def setup_styles(self):
        self.setStyleSheet(ADMIN_STYLESHEET)
        # Inject calendar PNG icon into the QDateEdit dropdown button.
        # Appending overrides the earlier ::down-arrow rule via CSS cascade.
        _cal = os.path.join(
            os.path.dirname(os.path.abspath(__file__)), 'assets', 'calendar.png'
        ).replace('\\', '/')
        self.setStyleSheet(self.styleSheet() + f"""
            QDateEdit::down-arrow {{
                image: url({_cal});
                width: 14px;
                height: 14px;
            }}
        """)

    def build_ui(self):
        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(10, 8, 10, 8)
        main_layout.setSpacing(6)

        # Connectivity banner (hidden until connection drops)
        self._conn_banner = ConnectionBanner()
        main_layout.addWidget(self._conn_banner)

        header_frame = QFrame()
        header_frame.setStyleSheet("""
            QFrame {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #3498db, stop:1 #2980b9);
                border-radius: 8px;
                padding: 8px;
                margin-bottom: 4px;
            }
        """)
        header_layout = QHBoxLayout(header_frame)
        header_layout.setContentsMargins(10, 4, 10, 4)

        title_label = QLabel("Admin Dashboard")
        title_label.setStyleSheet("font-size: 20px; font-weight: bold; color: white;")
        
        if AUTO_UPDATE_ENABLED:
            update_button = QPushButton(f"ℹ️ v{__version__}")
            update_button.setStyleSheet("""
                QPushButton {
                    background-color: #16a085;
                    color: white;
                    border: none;
                    padding: 10px 20px;
                    border-radius: 6px;
                    font-weight: bold;
                    font-size: 11px;
                    min-width: 80px;
                    max-height: 40px;
                }
                QPushButton:hover {
                    background-color: #1abc9c;
                }
                QPushButton:pressed {
                    background-color: #138d75;
                }
            """)
            update_button.clicked.connect(self.check_for_updates)
            update_button.setToolTip("Check for updates")
        
        logout_button = QPushButton("Logout")
        logout_button.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 6px;
                font-weight: bold;
                font-size: 12px;
                min-width: 100px;
                max-height: 40px;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
            QPushButton:pressed {
                background-color: #a93226;
            }
        """)
        logout_button.clicked.connect(self.handle_logout)
        
        ver_lbl = QLabel(f"v{__version__}")
        ver_lbl.setStyleSheet("""
            QLabel {
                color: rgba(255,255,255,0.6);
                font-size: 11px;
                font-weight: 600;
                padding: 0 8px;
                background: transparent;
            }
        """)
        ver_lbl.setToolTip(f"Operation Report System v{__version__}")

        header_layout.addWidget(title_label)
        header_layout.addStretch()
        header_layout.addWidget(ver_lbl)
        if AUTO_UPDATE_ENABLED:
            header_layout.addWidget(update_button)
        header_layout.addWidget(logout_button)
        
        main_layout.addWidget(header_frame)

        nav_frame = QFrame()
        nav_frame.setObjectName("navBar")
        nav_layout = QHBoxLayout(nav_frame)
        nav_layout.setContentsMargins(0, 0, 0, 0)
        nav_layout.setSpacing(0)

        self.daily_btn = QPushButton("Daily Cash Count")
        self.variance_btn = QPushButton("Variance Review")
        self.palawan_btn = QPushButton("Palawan")
        self.mc_btn = QPushButton("MC")
        self.fund_btn = QPushButton("Fund Transfer")
        self.payable_btn = QPushButton("Payable")
        self.global_payable_btn = QPushButton("Global Payable")
        self.report_btn = QPushButton("Payable Reports")
        self.daily_txn_btn = QPushButton("Daily Transaction")
        self.new_sanla_btn = QPushButton("New Sanla")
        self.new_renew_btn = QPushButton("New & Renew")
        self.global_os_btn = QPushButton("Global Other Services")
        self.ft_ho_btn = QPushButton("FT HO")
        self.depo_br_btn = QPushButton("DEPO BR")
        self.admin_btn = QPushButton("User Management")
        self.review_summary_btn = QPushButton("Review Summary")
        self.bir_book_btn = QPushButton("BIR Book")


        for btn in [self.daily_btn, self.variance_btn, self.palawan_btn, self.mc_btn,
                    self.fund_btn, self.payable_btn, self.global_payable_btn, self.report_btn, self.daily_txn_btn,
                    self.new_sanla_btn, self.new_renew_btn, self.global_os_btn, self.ft_ho_btn, self.depo_br_btn,
                    self.admin_btn, self.review_summary_btn, self.bir_book_btn]:
            btn.setCheckable(True)
        self.daily_btn.setChecked(True) 

        if self._is_brand_a:
  
            self.nav_buttons = [
                self.daily_btn, self.variance_btn, self.palawan_btn, self.mc_btn,
                self.fund_btn, self.payable_btn, self.daily_txn_btn, self.new_sanla_btn,
                self.new_renew_btn, self.global_os_btn, self.ft_ho_btn, self.depo_br_btn,
                self.review_summary_btn, self.admin_btn
            ]
        else:

            self.nav_buttons = [
                self.daily_btn, self.variance_btn, self.palawan_btn, self.mc_btn,
                self.fund_btn, self.payable_btn, self.global_payable_btn, self.report_btn, self.review_summary_btn, self.admin_btn, self.bir_book_btn
            ]

        for btn in self.nav_buttons:
            btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
            nav_layout.addWidget(btn)

        main_layout.addWidget(nav_frame)


        self.stack = QStackedWidget()

        self._lazy_factories = {} 
        self.daily_cash_widget = self.build_daily_cash_widget()

        if self._is_brand_a:

            self.stack.addWidget(self.daily_cash_widget)    
            self._add_lazy(1, lambda: VarianceReviewPage(account_type=self.account_type), 'variance_widget')
            self._add_lazy(2, lambda: PalawanPage(account_type=self.account_type), 'palawan_widget')
            self._add_lazy(3, lambda: MCPage(account_type=self.account_type), 'mc_widget')
            self._add_lazy(4, lambda: FundTransferPage(account_type=self.account_type), 'fund_widget')
            self._add_lazy(5, lambda: PayablesPage(account_type=self.account_type), 'payable_widget')
            self._add_lazy(6, lambda: DailyTransactionPage(), 'daily_txn_widget')
            self._add_lazy(7, lambda: NewSanlaPage(), 'new_sanla_widget')
            self._add_lazy(8, lambda: NewRenewPage(), 'new_renew_widget')
            self._add_lazy(9, lambda: GlobalOtherServicesPage(), 'global_os_widget')
            self._add_lazy(10, lambda: FTHOPage(account_type=self.account_type), 'ft_ho_widget')
            self._add_lazy(11, lambda: DepoBRPage(account_type=self.account_type), 'depo_br_widget')
            self._add_lazy(12, lambda: ReviewSummaryPage(account_type=self.account_type), 'review_summary_widget')
            self._add_lazy(13, lambda: UserManagementPage(), 'admin_widget')
        else:
      
            self.stack.addWidget(self.daily_cash_widget)         
            self._add_lazy(1, lambda: VarianceReviewPage(account_type=self.account_type), 'variance_widget')
            self._add_lazy(2, lambda: PalawanPage(account_type=self.account_type), 'palawan_widget')
            self._add_lazy(3, lambda: MCPage(account_type=self.account_type), 'mc_widget')
            self._add_lazy(4, lambda: FundTransferPage(account_type=self.account_type), 'fund_widget')
            self._add_lazy(5, lambda: PayablesPage(account_type=self.account_type), 'payable_widget')
            self._add_lazy(6, lambda: GlobalPayablePage(account_type=self.account_type), 'global_payable_widget')
            self._add_lazy(7, lambda: ReportPage(), 'report_widget')
            self._add_lazy(8, lambda: ReviewSummaryPage(account_type=self.account_type), 'review_summary_widget')
            self._add_lazy(9, lambda: UserManagementPage(), 'admin_widget')
            self._add_lazy(10, lambda: BIRBookPage(parent=self), 'bir_book_widget')
        
        main_layout.addWidget(self.stack)


        if self._is_brand_a:
         
            self.daily_btn.clicked.connect(lambda: self.switch_view(0, self.daily_btn))
            self.variance_btn.clicked.connect(lambda: self.switch_view(1, self.variance_btn))
            self.palawan_btn.clicked.connect(lambda: self.switch_view(2, self.palawan_btn))
            self.mc_btn.clicked.connect(lambda: self.switch_view(3, self.mc_btn))
            self.fund_btn.clicked.connect(lambda: self.switch_view(4, self.fund_btn))
            self.payable_btn.clicked.connect(lambda: self.switch_view(5, self.payable_btn))
            self.daily_txn_btn.clicked.connect(lambda: self.switch_view(6, self.daily_txn_btn))
            self.new_sanla_btn.clicked.connect(lambda: self.switch_view(7, self.new_sanla_btn))
            self.new_renew_btn.clicked.connect(lambda: self.switch_view(8, self.new_renew_btn))
            self.global_os_btn.clicked.connect(lambda: self.switch_view(9, self.global_os_btn))
            self.ft_ho_btn.clicked.connect(lambda: self.switch_view(10, self.ft_ho_btn))
            self.depo_br_btn.clicked.connect(lambda: self.switch_view(11, self.depo_br_btn))
            self.review_summary_btn.clicked.connect(lambda: self.switch_view(12, self.review_summary_btn))
            self.admin_btn.clicked.connect(lambda: self.switch_view(13, self.admin_btn))
        else:

            self.daily_btn.clicked.connect(lambda: self.switch_view(0, self.daily_btn))
            self.variance_btn.clicked.connect(lambda: self.switch_view(1, self.variance_btn))
            self.palawan_btn.clicked.connect(lambda: self.switch_view(2, self.palawan_btn))
            self.mc_btn.clicked.connect(lambda: self.switch_view(3, self.mc_btn))
            self.fund_btn.clicked.connect(lambda: self.switch_view(4, self.fund_btn))
            self.payable_btn.clicked.connect(lambda: self.switch_view(5, self.payable_btn))
            self.global_payable_btn.clicked.connect(lambda: self.switch_view(6, self.global_payable_btn))
            self.report_btn.clicked.connect(lambda: self.switch_view(7, self.report_btn))
            self.review_summary_btn.clicked.connect(lambda: self.switch_view(8, self.review_summary_btn))
            self.admin_btn.clicked.connect(lambda: self.switch_view(9, self.admin_btn))
            self.bir_book_btn.clicked.connect(lambda: self.switch_view(10, self.bir_book_btn))

        self.setLayout(main_layout)

    def handle_logout(self):
    
        reply = QMessageBox.question(
            self,
            "Confirm Logout",
            "Are you sure you want to logout?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            self.session.logout()
            self.logout_requested.emit()
            self.close()
    
    def _on_connection_lost(self):
        self._is_connected = False
        if hasattr(self, '_conn_banner'):
            self._conn_banner.show_banner()
        self._session_timer.stop()

    def _on_connection_restored(self):
        self._is_connected = True
        if hasattr(self, '_conn_banner'):
            self._conn_banner.hide_banner()
        self._session_timer.start(_SESSION_TICK_MS)

    def _check_session_timeout(self):

        if self.session.check_timeout():
            self._session_timer.stop()
            QMessageBox.warning(
                self,
                "Session Expired",
                "Your session has expired due to inactivity.\nPlease log in again.",
                QMessageBox.Ok
            )
            self.logout_requested.emit()
            self.close()
    
    def mousePressEvent(self, event):
  
        self.session.update_activity()
        super().mousePressEvent(event)
    
    def eventFilter(self, obj, event):
        """Handle application-level events for zoom"""
        # Capture wheel events at app level
        if event.type() == QEvent.Wheel:
            # Check if Ctrl is pressed
            if event.modifiers() & Qt.ControlModifier:
                delta = event.angleDelta().y()
                if delta > 0:
                    self.zoom_in()
                else:
                    self.zoom_out()
                return True  # Consume the event
        # Check for Ctrl+0 key press
        elif event.type() == QEvent.KeyPress:
            if event.modifiers() & Qt.ControlModifier and event.key() == Qt.Key_0:
                self.reset_zoom()
                return True  # Consume the event
        return super().eventFilter(obj, event)

    def wheelEvent(self, event):
        """Handle Ctrl + mouse wheel for zoom"""
        if event.modifiers() & Qt.ControlModifier:
            # Zoom in (scroll up) or zoom out (scroll down)
            delta = event.angleDelta().y()
            if delta > 0:
                self.zoom_in()
            else:
                self.zoom_out()
            event.accept()
        else:
            super().wheelEvent(event)
    
    def keyPressEvent(self, event):
        # Handle Ctrl+0 to reset zoom
        if event.modifiers() & Qt.ControlModifier and event.key() == Qt.Key_0:
            self.reset_zoom()
            event.accept()
        else:
            self.session.update_activity()
            super().keyPressEvent(event)
    
    def zoom_in(self):
        """Increase zoom level by 20%"""
        self.set_zoom_level(self.zoom_level + 20)
    
    def zoom_out(self):
        """Decrease zoom level by 20%"""
        self.set_zoom_level(self.zoom_level - 20)
    
    def reset_zoom(self):
        """Reset zoom to 100%"""
        self.set_zoom_level(100)
    
    def set_zoom_level(self, level):
        """Set zoom level and apply to all widgets"""
        # Clamp zoom level between 50% and 200%
        level = max(_ZOOM_MIN, min(_ZOOM_MAX, level))
        if level == self.zoom_level:
            return  # No change
        self.zoom_level = level
        self._apply_zoom_to_all()
    
    def check_for_updates(self):
       
        if AUTO_UPDATE_ENABLED:
            check_for_updates(parent=self, silent=False)
        else:
            QMessageBox.information(
                self,
                "Auto-Updater",
                "Auto-updater is not enabled.\n\n"
                "To enable it, install required dependencies:\n"
                "pip install requests packaging"
            )
    
    def closeEvent(self, event):

        if hasattr(self, '_update_checker_threads'):
            for thread in self._update_checker_threads[:]:
                if thread.isRunning():
                    thread.quit()
                    thread.wait(2000) 
        event.accept()

    def _add_lazy(self, index, factory, attr_name):

        self._lazy_factories[index] = (factory, attr_name)
        self.stack.addWidget(QWidget())

    def switch_view(self, index, active_button):

        if index in self._lazy_factories:
            factory, attr_name = self._lazy_factories.pop(index)
            QApplication.setOverrideCursor(Qt.WaitCursor)
            try:
                widget = factory()
                setattr(self, attr_name, widget)
                old = self.stack.widget(index)
                self.stack.removeWidget(old)
                old.deleteLater()
                self.stack.insertWidget(index, widget)
            finally:
                QApplication.restoreOverrideCursor()

        self.stack.setCurrentIndex(index)

        for btn in self.nav_buttons:
            if btn:
                btn.setChecked(False)

        if active_button:
            active_button.setChecked(True)

    def create_money_input(self):
        """Create a currency input using reusable MoneyInput widget."""
        field = MoneyInput(placeholder="0.00")
        field.setProperty("class", "money")
        return field

    def create_display_field(self):
        """Create a read-only display field using reusable DisplayField widget."""
        field = DisplayField()
        field.setProperty("class", "result")
        return field

    def create_lotes_input(self, read_only=False):
        """Create an integer lotes input using reusable LotesInput widget."""
        field = LotesInput(read_only=read_only)
        field.setMaximumWidth(70)
        return field

    # ── Palawan Details collapsible ───────────────────────────────────────────
    def _build_palawan_collapsible(self):
        """Build a collapsible Palawan Details section for the Daily Cash Count tab."""
        self.palawan_inputs = {}   # {db_col: QLineEdit}
        self.palawan_total_displays = {}  # {section: QLineEdit (read-only)}

        # Outer wrapper
        wrapper = QFrame()
        wrapper.setStyleSheet(
            "QFrame { border: 1px solid #BAE6FD; border-radius: 8px; "
            "background-color: #F0F9FF; margin-top: 4px; }"
        )
        wrapper_layout = QVBoxLayout(wrapper)
        wrapper_layout.setContentsMargins(0, 0, 0, 0)
        wrapper_layout.setSpacing(0)

        # Toggle button
        toggle_btn = QPushButton("▶  Palawan Details")
        toggle_btn.setCheckable(True)
        toggle_btn.setChecked(False)
        toggle_btn.setStyleSheet("""
            QPushButton {
                text-align: left; padding: 10px 14px;
                font-weight: 700; font-size: 13px;
                color: #0369A1; background-color: #E0F2FE;
                border: none; border-radius: 8px;
            }
            QPushButton:checked {
                background-color: #BAE6FD; border-bottom-left-radius: 0; border-bottom-right-radius: 0;
            }
            QPushButton:hover { background-color: #BAE6FD; }
        """)
        wrapper_layout.addWidget(toggle_btn)

        # Content area (hidden by default)
        content = QFrame()
        content.setVisible(False)
        content.setStyleSheet(
            "QFrame { background-color: #FFFFFF; border: none; "
            "border-top: 1px solid #BAE6FD; border-radius: 0; }"
        )
        content_layout = QVBoxLayout(content)
        content_layout.setContentsMargins(12, 12, 12, 12)
        content_layout.setSpacing(10)

        # ── Section builder ──
        def _make_section(title, section_key, color):
            box = QGroupBox(title)
            box.setStyleSheet(f"""
                QGroupBox {{
                    border: 1px solid #E2E8F0; border-radius: 6px;
                    margin-top: 20px; padding: 16px 14px 14px 14px;
                    background-color: #FFFFFF;
                }}
                QGroupBox::title {{
                    color: {color}; font-weight: 800; font-size: 12px;
                    padding: 1px 8px; background-color: #FFFFFF;
                }}
            """)
            form = QFormLayout()
            form.setSpacing(8)
            form.setContentsMargins(12, 20, 12, 12)

            for sub_label, db_col in [
                ("Principal", f"palawan_{section_key}_principal"),
                ("SC",        f"palawan_{section_key}_sc"),
                ("Commission",f"palawan_{section_key}_commission"),
            ]:
                inp = MoneyInput(placeholder="0.00")
                self.palawan_inputs[db_col] = inp
                lbl = QLabel(sub_label + ":")
                lbl.setStyleSheet("font-size: 13px; font-weight: 600; color: #334155;")
                form.addRow(lbl, inp)

            # Lotes
            lotes_col = f"palawan_{section_key}_lotes_total"
            lotes_inp = LotesInput(read_only=False)
            self.palawan_inputs[lotes_col] = lotes_inp
            lotes_lbl = QLabel("Lotes:")
            lotes_lbl.setStyleSheet("font-size: 13px; font-weight: 600; color: #334155;")
            form.addRow(lotes_lbl, lotes_inp)

            # Total (auto-calc)
            total_col = f"palawan_{section_key}_regular_total"
            total_disp = DisplayField()
            total_disp.setStyleSheet(
                f"font-weight: 800; font-size: 13px; color: {color}; "
                "background-color: #F0F9FF; border: 1px solid #BAE6FD; "
                "border-radius: 5px; padding: 5px 10px;"
            )
            self.palawan_inputs[total_col] = total_disp
            self.palawan_total_displays[section_key] = total_disp
            total_lbl = QLabel("TOTAL:")
            total_lbl.setStyleSheet(f"font-weight: 700; color: {color}; font-size: 13px;")
            form.addRow(total_lbl, total_disp)

            # Wire principal/sc/commission to auto-calc total
            def _recalc(_, sk=section_key):
                p  = float(self.palawan_inputs.get(f"palawan_{sk}_principal",  QLineEdit()).text() or 0)
                sc = float(self.palawan_inputs.get(f"palawan_{sk}_sc",         QLineEdit()).text() or 0)
                cm = float(self.palawan_inputs.get(f"palawan_{sk}_commission",  QLineEdit()).text() or 0)
                self.palawan_total_displays[sk].setText(f"{p + sc + cm:.2f}")
                self.palawan_inputs[f"palawan_{sk}_regular_total"].setText(f"{p + sc + cm:.2f}")

            for sub in ("principal", "sc", "commission"):
                self.palawan_inputs[f"palawan_{section_key}_{sub}"].textChanged.connect(_recalc)

            box.setLayout(form)
            return box

        # Three section groups in a grid
        grid_frame = QFrame()
        grid_layout = QGridLayout(grid_frame)
        grid_layout.setSpacing(10)
        grid_layout.addWidget(_make_section("PALAWAN SEND-OUT",      "sendout",       "#1b75bc"), 0, 0)
        grid_layout.addWidget(_make_section("PALAWAN PAY-OUT",       "payout",        "#1b75bc"), 0, 1)
        grid_layout.addWidget(_make_section("PALAWAN INTERNATIONAL", "international", "#1b75bc"), 1, 0, 1, 2)
        content_layout.addWidget(grid_frame)

        # Auto-sync: sendout principal/sc → daily cash count debit fields
        def _sync_sendout_principal(text):
            if "Palawan Send Out" in self.debit_inputs:
                self.debit_inputs["Palawan Send Out"].setText(text)

        def _sync_sendout_sc(text):
            if "Palawan S.C" in self.debit_inputs:
                self.debit_inputs["Palawan S.C"].setText(text)

        def _sync_payout_principal(text):
            if "Palawan Pay Out" in self.credit_inputs:
                self.credit_inputs["Palawan Pay Out"].setText(text)

        self.palawan_inputs["palawan_sendout_principal"].textChanged.connect(_sync_sendout_principal)
        self.palawan_inputs["palawan_sendout_sc"].textChanged.connect(_sync_sendout_sc)
        self.palawan_inputs["palawan_payout_principal"].textChanged.connect(_sync_payout_principal)

        # Adjustments section
        adj_box = QGroupBox("PALAWAN ADJUSTMENTS")
        adj_box.setStyleSheet("""
            QGroupBox {
                border: 2px solid #F59E0B; border-radius: 6px;
                margin-top: 20px; padding: 16px 14px 14px 14px;
                background-color: #FFFBEB;
            }
            QGroupBox::title {
                color: #D97706; font-weight: 800; font-size: 12px;
                padding: 1px 8px; background-color: #FFFBEB;
            }
        """)
        adj_form = QFormLayout()
        adj_form.setSpacing(8)
        adj_form.setContentsMargins(12, 20, 12, 12)

        # Mapping: adj_col (db key) → credit_fields UI label
        _adj_credit_label = {
            "palawan_pay_out_incentives": "Palawan Pay Out (incentives)",
            "palawan_suki_discounts":     "Palawan Suki Discounts",
            "palawan_suki_rebates":       "Palawan Suki Rebates",
            "palawan_cancel":             "Palawan Cancel",
        }

        def _make_adj_sync(db_col):
            cf_label = _adj_credit_label.get(db_col)
            def _sync(value):
                # value is float from MoneyInput.valueChanged signal
                if cf_label and cf_label in self.credit_inputs:
                    self.credit_inputs[cf_label].setText(f"{value:.2f}")
            return _sync

        for adj_label, adj_col in [
            ("Pay Out Incentives", "palawan_pay_out_incentives"),
            ("Suki Discounts",     "palawan_suki_discounts"),
            ("Suki Rebates",       "palawan_suki_rebates"),
            ("Cancel",             "palawan_cancel"),
        ]:
            inp = MoneyInput(placeholder="0.00")
            self.palawan_inputs[adj_col] = inp
            inp.valueChanged.connect(_make_adj_sync(adj_col))
            lbl = QLabel(adj_label + ":")
            lbl.setStyleSheet("font-size: 13px; font-weight: 600; color: #92400E;")
            adj_form.addRow(lbl, inp)
        adj_box.setLayout(adj_form)
        content_layout.addWidget(adj_box)

        # Save button
        save_btn = QPushButton("💾 Save Palawan Details")
        save_btn.setStyleSheet("""
            QPushButton {
                font-weight: 700; font-size: 13px; padding: 10px 16px;
                color: white; background-color: #059669;
                border: none; border-radius: 6px;
            }
            QPushButton:hover { background-color: #047857; }
            QPushButton:pressed { background-color: #065f46; }
        """)
        save_btn.clicked.connect(self._save_palawan_details)
        content_layout.addWidget(save_btn)

        wrapper_layout.addWidget(content)

        def _on_toggle(checked):
            content.setVisible(checked)
            toggle_btn.setText(
                "▼  Palawan Details" if checked
                else "▶  Palawan Details"
            )

        toggle_btn.toggled.connect(_on_toggle)
        return wrapper

    def _load_palawan_details(self, data: dict):
        """Populate palawan detail inputs from a loaded DB row."""
        if self._is_brand_a:
            # Brand A: sendout/payout/international data lives in payable_tbl_brand_a
            # (daily_reports_brand_a does NOT have palawan_sendout_principal etc.)
            corporation  = (data or {}).get('corporation') or self.corp_selector.currentText().strip()
            branch_name   = self.branch_selector.currentText()
            selected_date = self.date_picker.date().toString("yyyy-MM-dd")
            try:
                result = []
                if corporation:
                    result = self.db.execute_query(
                        "SELECT * FROM payable_tbl_brand_a WHERE corporation=%s AND branch=%s AND date=%s LIMIT 1",
                        (corporation, branch_name, selected_date)
                    )
                if not result:
                    # Compatibility fallback for legacy rows without corporation.
                    result = self.db.execute_query(
                        "SELECT * FROM payable_tbl_brand_a WHERE branch=%s AND date=%s LIMIT 1",
                        (branch_name, selected_date)
                    )
                if not result:
                    # New-structure fallback: branch posted to daily_reports_brand_a only.
                    # Map daily_reports fields to the payable_tbl schema best-effort.
                    try:
                        dr = self.db.execute_query(
                            "SELECT * FROM daily_reports_brand_a WHERE branch=%s AND date=%s LIMIT 1",
                            (branch_name, selected_date)
                        )
                        if dr:
                            d = dr[0]
                            so     = float(d.get('palawan_send_out', 0) or 0)
                            so_sc  = float(d.get('palawan_sc', 0) or 0)
                            po     = float(d.get('palawan_pay_out', 0) or 0)
                            po_inc = float(d.get('palawan_pay_out_incentives', 0) or 0)
                            result = [{
                                'sendout_capital':             so,
                                'sendout_sc':                  so_sc,
                                'sendout_commission':          0,
                                'sendout_lotes':               int(d.get('palawan_send_out_lotes', 0) or 0),
                                'sendout_total':               so + so_sc,
                                'payout_capital':              po,
                                'payout_sc':                   0,
                                'payout_commission':           0,
                                'payout_lotes':                int(d.get('palawan_pay_out_lotes', 0) or 0),
                                'payout_total':                po + po_inc,
                                'international_capital':       0,
                                'international_sc':            0,
                                'international_commission':    0,
                                'international_lotes':         0,
                                'international_total':         0,
                                'skid':        float(d.get('palawan_suki_discounts', 0) or 0),
                                'skir':        float(d.get('palawan_suki_rebates', 0) or 0),
                                'cancellation':float(d.get('palawan_cancel', 0) or 0),
                                'inc':         po_inc,
                            }]
                    except Exception as _fe:
                        logger.error("_load_palawan_details new-structure fallback: %s", _fe)
                if result:
                    r = result[0]
                    payable_map = {
                        'palawan_sendout_principal':            r.get('sendout_capital', 0) or 0,
                        'palawan_sendout_sc':                   r.get('sendout_sc', 0) or 0,
                        'palawan_sendout_commission':           r.get('sendout_commission', 0) or 0,
                        'palawan_sendout_lotes_total':          r.get('sendout_lotes', 0) or 0,
                        'palawan_sendout_regular_total':        r.get('sendout_total', 0) or 0,
                        'palawan_payout_principal':             r.get('payout_capital', 0) or 0,
                        'palawan_payout_sc':                    r.get('payout_sc', 0) or 0,
                        'palawan_payout_commission':            r.get('payout_commission', 0) or 0,
                        'palawan_payout_lotes_total':           r.get('payout_lotes', 0) or 0,
                        'palawan_payout_regular_total':         r.get('payout_total', 0) or 0,
                        'palawan_international_principal':      r.get('international_capital', 0) or 0,
                        'palawan_international_sc':             r.get('international_sc', 0) or 0,
                        'palawan_international_commission':     r.get('international_commission', 0) or 0,
                        'palawan_international_lotes_total':    r.get('international_lotes', 0) or 0,
                        'palawan_international_regular_total':  r.get('international_total', 0) or 0,
                    }
                    # CRITICAL FIX: Don't merge payable_map into data for variance calculation
                    # Palawan data should be loaded separately for DISPLAY only, not for debit/credit sums
                    # data = {**data, **payable_map}  # REMOVED - causes double-counting in variance
                    # Instead, load payable data directly into the palawan display fields
                    if hasattr(self, '_palawan_display_map'):
                        self._palawan_display_map = payable_map
            except Exception as e:
                logger.error("_load_palawan_details Brand A payable query: %s", e)

        else:
            # Brand B: daily_reports palawan columns are zeroed after migration;
            # authoritative values now live in payable_tbl_brand_a.
            branch_name   = self.branch_selector.currentText()
            selected_date = self.date_picker.date().toString("yyyy-MM-dd")
            corporation   = (data or {}).get('corporation') or self.corp_selector.currentText().strip()
            try:
                _pr = []
                if corporation:
                    _pr = self.db.execute_query(
                        "SELECT * FROM payable_tbl_brand_a WHERE corporation=%s AND branch=%s AND date=%s LIMIT 1",
                        (corporation, branch_name, selected_date)
                    )
                if not _pr:
                    _pr = self.db.execute_query(
                        "SELECT * FROM payable_tbl_brand_a WHERE branch=%s AND date=%s LIMIT 1",
                        (branch_name, selected_date)
                    )
                if _pr:
                    r = _pr[0]
                    _payable_to_daily = {
                        'palawan_sendout_principal':           r.get('sendout_capital', 0) or 0,
                        'palawan_sendout_sc':                  r.get('sendout_sc', 0) or 0,
                        'palawan_sendout_commission':          r.get('sendout_commission', 0) or 0,
                        'palawan_sendout_lotes_total':         r.get('sendout_lotes', 0) or 0,
                        'palawan_payout_principal':            r.get('payout_capital', 0) or 0,
                        'palawan_payout_sc':                   r.get('payout_sc', 0) or 0,
                        'palawan_payout_commission':           r.get('payout_commission', 0) or 0,
                        'palawan_payout_lotes_total':          r.get('payout_lotes', 0) or 0,
                        'palawan_international_principal':     r.get('international_capital', 0) or 0,
                        'palawan_international_sc':            r.get('international_sc', 0) or 0,
                        'palawan_international_commission':    r.get('international_commission', 0) or 0,
                        'palawan_international_lotes_total':   r.get('international_lotes', 0) or 0,
                        'palawan_suki_discounts':              r.get('skid', 0) or 0,
                        'palawan_suki_rebates':                r.get('skir', 0) or 0,
                        'palawan_cancel':                      r.get('cancellation', 0) or 0,
                        'palawan_pay_out_incentives':          r.get('inc', 0) or 0,
                    }
                    # Override daily_reports row with payable values where daily was zeroed
                    for col, val in _payable_to_daily.items():
                        if float(data.get(col, 0) or 0) == 0 and val != 0:
                            data[col] = val
            except Exception as e:
                logger.error("_load_palawan_details Brand B payable query: %s", e)

        for db_col, widget in getattr(self, 'palawan_inputs', {}).items():
            val = data.get(db_col, 0) or 0
            widget.blockSignals(True)
            try:
                if widget.isReadOnly():
                    widget.setText(f"{float(val):.2f}" if float(val) else "")
                elif widget.validator() and hasattr(widget.validator(), 'decimals'):
                    widget.setText(f"{float(val):.2f}" if float(val) else "")
                else:
                    widget.setText(str(int(val)) if int(float(val)) else "")
            except (TypeError, ValueError):
                widget.setText("")
            widget.blockSignals(False)
        # Recalc totals
        for section in ("sendout", "payout", "international"):
            p  = float(getattr(self, 'palawan_inputs', {}).get(f"palawan_{section}_principal",  QLineEdit()).text() or 0)
            sc = float(getattr(self, 'palawan_inputs', {}).get(f"palawan_{section}_sc",         QLineEdit()).text() or 0)
            cm = float(getattr(self, 'palawan_inputs', {}).get(f"palawan_{section}_commission",  QLineEdit()).text() or 0)
            total_disp = getattr(self, 'palawan_total_displays', {}).get(section)
            if total_disp:
                total_disp.setText(f"{p + sc + cm:.2f}")

    BANK_ACCOUNTS = [
                {"id": 1, "bank_name": "CIB-BDO", "account_name": "Global Reliance", "account_number": "0077-9002-3923"},
        {"id": 2, "bank_name": "CIB-BPI", "account_name": "Kristal Clear Diamond and Gold Pawnshop", "account_number": "0091-0692-29"},
        {"id": 3, "bank_name": "CIB-BDO", "account_name": "Kristal Clear", "account_number": "0077-9001-8784"},
        {"id": 4, "bank_name": "CIB-Union Bank", "account_name": "Golbal Reliance Mgmt and Holdings Corp", "account_number": "0015-6000-5790"},
        {"id": 5, "bank_name": "CIB-BDO", "account_name": "Europacific Management & Holdings Corp", "account_number": "0038-1801-5838"},
        {"id": 6, "bank_name": "CIB-BPI", "account_name": "Europacific Management & Holdings Corp", "account_number": "3541-0035-67"},
        {"id": 7, "bank_name": "CIB-UB", "account_name": "Europacific Management & Holdings Corp", "account_number": "0021-7001-7921"},
    ]
    def _show_info_dialog(
        self, title: str,
        header_text: str, header_color: str,
        body_rows,         # list of (html_text, fg_color) or None → show empty state
        btn_color: str, btn_hover: str,
        empty_header: str, empty_note: str,
    ) -> None:
        """Generic info dialog: coloured header, optional green body frame, Close button."""
        dialog = QDialog(self)
        dialog.setWindowTitle(title)
        dialog.setMinimumWidth(_INFO_DIALOG_MIN_WIDTH)
        layout = QVBoxLayout(dialog)

        if body_rows:
            hdr = QLabel(header_text)
            hdr.setStyleSheet(
                f"font-size: 14px; font-weight: bold; color: {header_color}; padding: 10px;"
            )
            layout.addWidget(hdr)

            frame = QFrame()
            frame.setStyleSheet(f"""
                QFrame {{
                    background-color: #ECFDF5;
                    border: 2px solid {header_color};
                    border-radius: 8px;
                    padding: 15px;
                }}
            """)
            frame_layout = QVBoxLayout(frame)
            frame_layout.setSpacing(8)
            for html_text, fg in body_rows:
                lbl = QLabel(html_text)
                lbl.setStyleSheet(f"font-size: 13px; color: {fg};")
                frame_layout.addWidget(lbl)
            layout.addWidget(frame)
        else:
            h = QLabel(empty_header)
            h.setStyleSheet(
                "font-size: 14px; font-weight: bold; color: #64748B; padding: 10px;"
            )
            layout.addWidget(h)
            n = QLabel(empty_note)
            n.setStyleSheet("font-size: 11px; color: #64748B; padding: 5px 10px;")
            layout.addWidget(n)

        layout.addStretch()
        btn = QPushButton("Close")
        btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {btn_color}; color: white;
                border: none; border-radius: 5px;
                font-size: 12px; font-weight: 700;
                padding: 8px 20px;
            }}
            QPushButton:hover {{ background-color: {btn_hover}; }}
        """)
        btn.clicked.connect(dialog.accept)
        layout.addWidget(btn)
        dialog.exec_()

    def show_bank_account_info(self):
        if self.selected_bank_account:
            bank = next((b for b in self.BANK_ACCOUNTS if b['id'] == self.selected_bank_account), None)
            if bank:
                rows = [
                    (f"<b>Bank:</b> {bank['bank_name']}", "#065F46"),
                    (f"<b>Account Name:</b> {bank['account_name']}", "#065F46"),
                ]
                if bank.get('account_number'):
                    rows.append((f"<b>Account #:</b> {bank['account_number']}", "#065F46"))
                self._show_info_dialog(
                    "Bank Account", "Client Selected Bank Account", "#10B981",
                    rows, "#8B5CF6", "#7C3AED", "", "",
                )
            else:
                self._show_info_dialog(
                    "Bank Account", "Bank Account Not Found", "#F59E0B",
                    [(f"Bank account ID {self.selected_bank_account} not found in system.", "#64748B")],
                    "#8B5CF6", "#7C3AED", "", "",
                )
        else:
            self._show_info_dialog(
                "Bank Account", "", "",
                None, "#8B5CF6", "#7C3AED",
                "No Bank Account Selected",
                "The client has not selected a bank account for this fund transfer.",
            )

    def show_branch_dest_info(self):
        """Show destination branch info for Fund Transfer to BRANCH (view only for admin)"""
        rows = (
            [(f"<b>Destination Branch:</b> {self.selected_branch_dest}", "#065F46")]
            if self.selected_branch_dest else None
        )
        self._show_info_dialog(
            "Destination Branch", "Fund Transfer Destination Branch", "#059669",
            rows, "#059669", "#047857",
            "No Destination Branch",
            "The client has not specified a destination branch for this fund transfer.",
        )

    def show_from_branch_dest_info(self):
        rows = (
            [(f"<b>Source Branch:</b> {self.selected_from_branch_dest}", "#065F46")]
            if self.selected_from_branch_dest else None
        )
        self._show_info_dialog(
            "Source Branch", "Fund Transfer Source Branch", "#059669",
            rows, "#059669", "#047857",
            "No Source Branch",
            "The client has not specified a source branch for this fund transfer.",
        )

    def build_admin_widget(self):
        """Build Admin Manage UI for corporations, branches, and clients"""
        widget = QWidget()
        
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_content = QWidget()
        layout = QVBoxLayout(scroll_content)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)

        corp_box = QGroupBox("Manage Corporations")
        corp_box.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                padding-top: 20px;
            }
        """)
        corp_layout = QVBoxLayout()
        
        corp_form = QFormLayout()
        corp_form.setSpacing(10)
        self.corp_name_input = QLineEdit()
        self.corp_name_input.setPlaceholderText("Enter corporation name")
        
        corp_add_btn = QPushButton("Add Corporation")
        corp_add_btn.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                padding: 8px 16px;
            }
            QPushButton:hover {
                background-color: #229954;
            }
        """)
        corp_add_btn.clicked.connect(self._on_add_corporation)
        
        corp_form.addRow(QLabel("Corporation Name:"), self.corp_name_input)
        corp_form.addRow(corp_add_btn)
        

        self.corp_list_display = QLabel("No corporations loaded")
        self.corp_list_display.setStyleSheet("""
            QLabel {
                background-color: #f8f9fa;
                border: 1px solid #dee2e6;
                border-radius: 4px;
                padding: 10px;
                font-family: 'Courier New', monospace;
                font-size: 10px;
                min-height: 100px;
            }
        """)
        self.corp_list_display.setWordWrap(True)
        self.corp_list_display.setTextInteractionFlags(Qt.TextSelectableByMouse)
        
        corp_refresh_btn = QPushButton("Refresh List")
        corp_refresh_btn.clicked.connect(self._refresh_corporation_display)
        
        corp_layout.addLayout(corp_form)
        corp_layout.addWidget(QLabel("Existing Corporations:"))
        corp_layout.addWidget(self.corp_list_display)
        corp_layout.addWidget(corp_refresh_btn)
        corp_box.setLayout(corp_layout)


        branch_box = QGroupBox("Manage Branches")
        branch_box.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                padding-top: 20px;
            }
        """)
        branch_layout = QVBoxLayout()

        branch_form = QFormLayout()
        branch_form.setSpacing(10)
        self.branch_corp_selector = QComboBox()
        self.branch_corp_selector.setMinimumWidth(200)
        self.branch_name_input = QLineEdit()
        self.branch_name_input.setPlaceholderText("Enter branch name")

        self.branch_os_selector = QComboBox()
        self.branch_os_selector.setMinimumWidth(200)
        self.branch_os_selector.addItem("-- Select OS (optional) --", None)
        self._load_os_options_for_branch()
        
        branch_add_btn = QPushButton("Add Branch")
        branch_add_btn.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                padding: 8px 16px;
            }
            QPushButton:hover {
                background-color: #229954;
            }
        """)
        branch_add_btn.clicked.connect(self._on_add_branch)
        
        branch_form.addRow(QLabel("Select Corporation:"), self.branch_corp_selector)
        branch_form.addRow(QLabel("Branch Name:"), self.branch_name_input)
        branch_form.addRow(QLabel("Operation Supervisor:"), self.branch_os_selector)
        branch_form.addRow(branch_add_btn)
        
        self.branch_list_display = QLabel("No branches loaded")
        self.branch_list_display.setStyleSheet("""
            QLabel {
                background-color: #f8f9fa;
                border: 1px solid #dee2e6;
                border-radius: 4px;
                padding: 10px;
                font-family: 'Courier New', monospace;
                font-size: 10px;
                min-height: 100px;
            }
        """)
        self.branch_list_display.setWordWrap(True)
        self.branch_list_display.setTextInteractionFlags(Qt.TextSelectableByMouse)
        
        branch_refresh_btn = QPushButton("Refresh List")
        branch_refresh_btn.clicked.connect(self._refresh_branch_display)
        
        branch_layout.addLayout(branch_form)
        branch_layout.addWidget(QLabel("Existing Branches:"))
        branch_layout.addWidget(self.branch_list_display)
        branch_layout.addWidget(branch_refresh_btn)
        branch_box.setLayout(branch_layout)


        client_box = QGroupBox("Manage Clients")
        client_box.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                padding-top: 20px;
            }
        """)
        client_layout = QVBoxLayout()
        

        client_form = QFormLayout()
        client_form.setSpacing(10)
        

        self.client_username_display = QLineEdit()
        self.client_username_display.setReadOnly(True)
        self.client_username_display.setPlaceholderText("Auto-generated (e.g., CL-0001)")
        self.client_username_display.setStyleSheet("""
            QLineEdit {
                background-color: #e9ecef;
                font-weight: bold;
                color: #495057;
                border: 2px solid #ced4da;
            }
        """)
        
        self.client_first_input = QLineEdit()
        self.client_first_input.setPlaceholderText("Enter first name")
        self.client_last_input = QLineEdit()
        self.client_last_input.setPlaceholderText("Enter last name")
        self.client_corp_selector = QComboBox()
        self.client_corp_selector.setMinimumWidth(200)
        self.client_branch_selector = QComboBox()
        self.client_branch_selector.setMinimumWidth(200)
        self.client_password_input = QLineEdit()
        self.client_password_input.setPlaceholderText("Enter password")
        self.client_password_input.setEchoMode(QLineEdit.Password)
        
        preview_btn = QPushButton("Preview Username")
        preview_btn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                padding: 6px 12px;
                font-size: 10px;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        preview_btn.clicked.connect(self._preview_username)
        
        client_add_btn = QPushButton("Add Client")
        client_add_btn.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                padding: 8px 16px;
            }
            QPushButton:hover {
                background-color: #229954;
            }
        """)
        client_add_btn.clicked.connect(self._on_add_client)
        
        client_form.addRow(QLabel("Username (Auto):"), self.client_username_display)
        client_form.addRow(preview_btn)
        client_form.addRow(QLabel("First Name:"), self.client_first_input)
        client_form.addRow(QLabel("Last Name:"), self.client_last_input)
        client_form.addRow(QLabel("Corporation:"), self.client_corp_selector)
        client_form.addRow(QLabel("Branch:"), self.client_branch_selector)
        client_form.addRow(QLabel("Password:"), self.client_password_input)
        client_form.addRow(client_add_btn)
        

        self.client_list_display = QLabel("No clients loaded")
        self.client_list_display.setStyleSheet("""
            QLabel {
                background-color: #f8f9fa;
                border: 1px solid #dee2e6;
                border-radius: 4px;
                padding: 10px;
                font-family: 'Courier New', monospace;
                font-size: 10px;
                min-height: 150px;
            }
        """)
        self.client_list_display.setWordWrap(True)
        self.client_list_display.setTextInteractionFlags(Qt.TextSelectableByMouse)
        
        client_refresh_btn = QPushButton("Refresh List")
        client_refresh_btn.clicked.connect(self._refresh_client_display)
        
        client_layout.addLayout(client_form)
        client_layout.addWidget(QLabel("Existing Clients:"))
        client_layout.addWidget(self.client_list_display)
        client_layout.addWidget(client_refresh_btn)
        client_box.setLayout(client_layout)

        layout.addWidget(corp_box)
        layout.addWidget(branch_box)
        layout.addWidget(client_box)
        layout.addStretch()


        scroll_area.setWidget(scroll_content)
        
        main_layout = QVBoxLayout(widget)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.addWidget(scroll_area)

        self._refresh_admin_corporations()
        self._refresh_corporation_display()
        self._refresh_branch_display()
        self._refresh_client_display()

        self.branch_corp_selector.currentIndexChanged.connect(
            lambda: self._refresh_admin_branches(self.branch_corp_selector.currentData())
        )
        self.client_corp_selector.currentIndexChanged.connect(
            lambda: self._refresh_admin_branches(self.client_corp_selector.currentData(), target='client')
        )

        return widget

    def _preview_username(self):
       
        try:
            row = self.db.execute_query(
                "SELECT MAX(CAST(SUBSTRING(username,4) AS UNSIGNED)) AS maxnum FROM users WHERE username LIKE 'CL-%'"
            )
            maxnum = 0
            if row and row[0] and row[0].get('maxnum') is not None:
                try:
                    maxnum = int(row[0]['maxnum'])
                except Exception:
                    maxnum = 0
            
            next_num = maxnum + 1
            username = f"CL-{str(next_num).zfill(4)}"
            self.client_username_display.setText(username)
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to preview username: {e}")

    def _refresh_corporation_display(self):
      
        try:
            rows = self.db.execute_query("SELECT id, name, created_at FROM corporations ORDER BY name")
            if not rows:
                self.corp_list_display.setText("No corporations found")
                return
            
   
            display_text = "┌─────┬────────────────────────────────┬─────────────────────┐\n"
            display_text += "│ ID  │ Corporation Name               │ Created At          │\n"
            display_text += "├─────┼────────────────────────────────┼─────────────────────┤\n"
            
            for r in rows:
                corp_id = str(r['id']).ljust(3)
                name = str(r['name'])[:30].ljust(30)
                created = str(r.get('created_at', 'N/A'))[:19].ljust(19)
                display_text += f"│ {corp_id} │ {name} │ {created} │\n"
            
            display_text += "└─────┴────────────────────────────────┴─────────────────────┘"
            self.corp_list_display.setText(display_text)
            
        except Exception as e:
            self.corp_list_display.setText(f"Error loading corporations: {e}")

    def _refresh_branch_display(self):

        try:
            query = """
                SELECT b.id, b.name, b.corporation_id, c.name as corp_name, b.created_at,
                       b.sub_corporation_id, sc.name as sub_corp_name,
                       COALESCE(b.payroll, '') as payroll
                FROM branches b
                LEFT JOIN corporations c ON b.corporation_id = c.id
                LEFT JOIN corporations sc ON b.sub_corporation_id = sc.id
                ORDER BY c.name, b.name
            """
            rows = self.db.execute_query(query)
            if not rows:
                self.branch_list_display.setText("No branches found")
                return

            display_text = "┌─────┬──────────────────────┬─┬────────────────────────────────────┬─────────────────────┐\n"
            display_text += "│ ID  │ Branch Name          │P│ Corporation                        │ Created At          │\n"
            display_text += "├─────┼──────────────────────┼─┼────────────────────────────────────┼─────────────────────┤\n"

            for r in rows:
                branch_id = str(r['id']).ljust(3)
                name = str(r['name'])[:20].ljust(20)
                payroll_flag = "★" if str(r.get('payroll') or '').upper() == 'YES' else " "
                corp_display = str(r.get('corp_name', 'N/A'))
                if r.get('sub_corp_name'):
                    corp_display += f" + {r.get('sub_corp_name')}"
                corp_display = corp_display[:34].ljust(34)
                created = str(r.get('created_at', 'N/A'))[:19].ljust(19)
                display_text += f"│ {branch_id} │ {name} │{payroll_flag}│ {corp_display} │ {created} │\n"

            display_text += "└─────┴──────────────────────┴─┴────────────────────────────────────┴─────────────────────┘"
            display_text += "\n★ = Payroll branch (Import from Excel enabled)"
            self.branch_list_display.setText(display_text)

        except Exception as e:
            self.branch_list_display.setText(f"Error loading branches: {e}")


    def _refresh_client_display(self):
        """Refresh the client list display"""
        try:
            query = """
                SELECT u.id, u.username, u.first_name, u.last_name, 
                       u.corporation as corp_name, u.branch as branch_name, u.created_at
                FROM users u
                WHERE u.role = 'user'
                ORDER BY u.id DESC
                LIMIT 50
            """
            rows = self.db.execute_query(query)
            if not rows:
                self.client_list_display.setText("No clients found")
                return
            

            display_text = "┌─────┬──────────┬──────────────────────┬──────────────────┬──────────────────┬─────────────────────┐\n"
            display_text += "│ ID  │ Username │ Name                 │ Corporation      │ Branch           │ Created At          │\n"
            display_text += "├─────┼──────────┼──────────────────────┼──────────────────┼──────────────────┼─────────────────────┤\n"
            
            for r in rows:
                client_id = str(r['id']).ljust(3)
                username = str(r['username']).ljust(8)
                full_name = f"{r.get('first_name', '')} {r.get('last_name', '')}"[:20].ljust(20)
                corp = str(r.get('corp_name', 'N/A'))[:16].ljust(16)
                branch = str(r.get('branch_name', 'N/A'))[:16].ljust(16)
                created = str(r.get('created_at', 'N/A'))[:19].ljust(19)
                display_text += f"│ {client_id} │ {username} │ {full_name} │ {corp} │ {branch} │ {created} │\n"
            
            display_text += "└─────┴──────────┴──────────────────────┴──────────────────┴──────────────────┴─────────────────────┘"
            display_text += f"\n\nShowing last 50 clients (Total in database may be more)"
            self.client_list_display.setText(display_text)
            
        except Exception as e:
            self.client_list_display.setText(f"Error loading clients: {e}")

    def load_corporations(self):
        try:
            self.corp_selector.clear()
            # Query all corporations from corporations table
            result = self.db.execute_query("SELECT name as corporation FROM corporations ORDER BY name")
            if result:
                for row in result:
                    if row['corporation']:
                        self.corp_selector.addItem(row['corporation'])
            # Also load OS options
            self.load_os_options()

        except Exception as e:
            logger.error("Error loading corporations: %s", e)
            QMessageBox.critical(self, "Database Error", f"Failed to load corporations: {e}")

    def load_os_options(self):
        try:
            self.os_selector.clear()
            if self.os_group:
                # Restricted admin: only show their assigned group
                self.os_selector.addItem(self.os_group)
                self.os_selector.setCurrentIndex(0)
                self.os_selector.setEnabled(False)
            else:
                self.os_selector.setEnabled(True)
                result = self.db.execute_query("""
                    SELECT DISTINCT os_name FROM branches
                    WHERE os_name IS NOT NULL AND os_name != ''
                    ORDER BY os_name
                """)
                if result:
                    for row in result:
                        os_name = row['os_name'] if isinstance(row, dict) else row[0]
                        if os_name:
                            self.os_selector.addItem(os_name)
        except Exception as e:
            logger.error("Error loading OS options: %s", e)

    def on_filter_type_changed(self):
        filter_type = self.filter_type_selector.currentData()
        if filter_type == "corporation":
            self.corp_label.setVisible(True)
            self.corp_selector.setVisible(True)
            self.os_label.setVisible(False)
            self.os_selector.setVisible(False)
            self.load_branches()
        else:
            self.corp_label.setVisible(False)
            self.corp_selector.setVisible(False)
            self.os_label.setVisible(True)
            self.os_selector.setVisible(True)
            self.load_branches_by_os()

    def load_branches(self):
    
        try:
            self.branch_selector.clear()
            corp_name = self.corp_selector.currentText()
            if corp_name:

                query = """
                    SELECT b.name as branch
                    FROM branches b
                    LEFT JOIN corporations c ON b.corporation_id = c.id
                    LEFT JOIN corporations sc ON b.sub_corporation_id = sc.id
                    WHERE c.name = %s OR sc.name = %s
                    ORDER BY b.name
                """
                result = self.db.execute_query(query, [corp_name, corp_name])
                if result:
                    for row in result:
                        if row['branch']:
                            self.branch_selector.addItem(row['branch'])
        except Exception as e:
            logger.error("Error loading branches: %s", e)
            QMessageBox.critical(self, "Database Error", f"Failed to load branches: {e}")

    def load_branches_by_os(self):
        """Load branches filtered by OS name"""
        try:
            self.branch_selector.clear()
            os_name = self.os_selector.currentText()
            if os_name:

                query = """
                    SELECT name as branch
                    FROM branches
                    WHERE os_name = %s
                    ORDER BY name
                """
                result = self.db.execute_query(query, [os_name])
                if result:
                    for row in result:
                        if row['branch']:
                            self.branch_selector.addItem(row['branch'])
        except Exception as e:
            logger.error("Error loading branches by OS: %s", e)
            QMessageBox.critical(self, "Database Error", f"Failed to load branches: {e}")

    def _refresh_admin_corporations(self):
        try:
            rows = self.db.execute_query("SELECT id, name FROM corporations ORDER BY name")
            self.branch_corp_selector.clear()
            self.client_corp_selector.clear()
            if rows:
                for r in rows:
                    self.branch_corp_selector.addItem(r['name'], r['id'])
                    self.client_corp_selector.addItem(r['name'], r['id'])
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load corporations: {e}")

    def _refresh_admin_branches(self, corp_id=None, target='both'):
        try:
            if corp_id is None:
                corp_id = self.branch_corp_selector.currentData()
            if not corp_id:
                self.client_branch_selector.clear()
                return


            rows = self.db.execute_query("SELECT id, name FROM branches WHERE corporation_id=%s OR sub_corporation_id=%s ORDER BY name", (corp_id, corp_id))
            self.client_branch_selector.clear()
            if rows:
                for r in rows:
                    self.client_branch_selector.addItem(r['name'], r['id'])
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load branches: {e}")

    def _on_add_corporation(self):
        name = self.corp_name_input.text().strip()
        if not name:
            QMessageBox.warning(self, "Input Required", "Please enter a corporation name.")
            return
        try:
            cid = create_corporation(name)
            if cid:
                QMessageBox.information(self, "Created", f"Corporation '{name}' created successfully (ID: {cid}).")
                self.corp_name_input.clear()
                self._refresh_admin_corporations()
                self._refresh_corporation_display()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to create corporation: {e}")

    def _load_os_options_for_branch(self):

        try:
            supervisors = get_all_supervisors()
            for sup in supervisors:
                self.branch_os_selector.addItem(sup['name'], sup['name'])
        except Exception as e:
            logger.error("Error loading OS options: %s", e)

    def _on_add_branch(self):
        name = self.branch_name_input.text().strip()
        corp_id = self.branch_corp_selector.currentData()
        os_name = self.branch_os_selector.currentData()
        if not corp_id:
            QMessageBox.warning(self, "Selection Required", "Please select a corporation for this branch.")
            return
        if not name:
            QMessageBox.warning(self, "Input Required", "Please enter a branch name.")
            return
        try:
            bid = create_branch(name, corp_id, os_name)
            if bid:
                QMessageBox.information(self, "Created", f"Branch '{name}' created successfully (ID: {bid}).")
                self.branch_name_input.clear()
                self.branch_os_selector.setCurrentIndex(0)
                self._refresh_admin_branches(corp_id)
                self._refresh_branch_display()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to create branch: {e}")

    def _on_add_client(self):
        first = self.client_first_input.text().strip()
        last = self.client_last_input.text().strip()
        corp_id = self.client_corp_selector.currentData()
        branch_id = self.client_branch_selector.currentData()
        password = self.client_password_input.text() or None

        if not (first and last):
            QMessageBox.warning(self, "Input Required", "Please enter client's first and last names.")
            return
        if not corp_id or not branch_id:
            QMessageBox.warning(self, "Selection Required", "Please select corporation and branch for the client.")
            return

        try:
            row = create_client(first, last, corp_id, branch_id, password)
            if row:
                QMessageBox.information(
                    self, 
                    "✅ Created", 
                    f"Client created successfully!\n\n"
                    f"Username: {row['username']}\n"
                    f"ID: {row['id']}\n"
                    f"Name: {first} {last}"
                )
                # Clear inputs
                self.client_first_input.clear()
                self.client_last_input.clear()
                self.client_password_input.clear()
                self.client_username_display.clear()
                self._refresh_client_display()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to create client: {e}")

    def reset_entry(self):

        try:
            branch_name = self.branch_selector.currentText()
            selected_date = self.date_picker.date().toString("yyyy-MM-dd")

            if not branch_name:
                QMessageBox.warning(self, "Selection Required", "Please select a branch.")
                return

            reply = QMessageBox.question(
                self,
                "Confirm Reset",
                f"Are you sure you want to reset the entry for:\n\n"
                f"Branch: {branch_name}\n"
                f"Date: {selected_date}\n\n"
                f"This will reset BOTH Brand A and Brand B,\n"
                f"allowing the branch to edit and resubmit their report.",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )

            if reply == QMessageBox.No:
                return


            found = unlock_entry(db_manager, branch_name, selected_date)
            clear_supplement_tables(db_manager, branch_name, selected_date)

            if found:
                # Send notification to clients
                self._notify_entry_reset(branch_name, selected_date)

                QMessageBox.information(
                    self,
                    "Entry Reset",
                    f"Entry for {branch_name} on {selected_date} has been reset.\n\n"
                    f"Both Brand A and Brand B are now unlocked.\n"
                    f"The branch can edit and resubmit their report."
                )
            else:
                QMessageBox.information(
                    self,
                    "No Entry Found",
                    f"No entry found for {branch_name} on {selected_date}."
                )

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to reset entry: {e}")

    def _notify_entry_reset(self, branch_name: str, selected_date: str):
        """Send notification to clients that entry has been reset."""
        try:
            # Get auth token
            response = requests.post(
                f"{API_URL}/api/token",
                json={"api_key": API_KEY},
                timeout=5
            )
            if response.status_code != 200:
                logger.warning(f"Failed to get token for notification: {response.status_code}")
                return

            token = response.json().get("token")
            if not token:
                logger.warning("No token in notification auth response")
                return

            # Send notification
            headers = {"Authorization": f"Bearer {token}"}
            response = requests.post(
                f"{API_URL}/api/notify/reset_entry",
                json={
                    "branch": branch_name,
                    "date": selected_date,
                    "admin_name": "Administrator"
                },
                headers=headers,
                timeout=5
            )

            if response.status_code == 200:
                result = response.json()
                clients_notified = result.get("clients_notified", 0)
                logger.info(f"Entry reset notification sent to {clients_notified} clients for branch {branch_name}")
            else:
                logger.warning(f"Failed to send entry reset notification: {response.status_code}")

        except Exception as e:
            logger.warning(f"Error sending entry reset notification: {e}")

    def export_daily_cash_to_excel(self):
        if not _OPENPYXL_AVAILABLE:
            QMessageBox.critical(
                self,
                "Missing Dependency",
                "The openpyxl package is required to export to Excel.\nInstall with: pip install openpyxl"
            )
            return
        
        filter_type = self.filter_type_selector.currentData()
        if filter_type == "corporation":
            filter_label = "Corporation"
            filter_value = self.corp_selector.currentText()
        else:
            filter_label = "Group"
            filter_value = self.os_selector.currentText()
        
        branch_name = self.branch_selector.currentText()
        selected_date = self.date_picker.date().toString("yyyy-MM-dd")
        
        if not branch_name:
            QMessageBox.warning(self, "Selection Required", "Please select a branch.")
            return
        
        # File dialog for save location
        default_filename = f"DailyCashCount_{filter_value}_{branch_name}_{selected_date}.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Excel File",
            default_filename,
            "Excel Files (*.xlsx);;All Files (*)"
        )
        
        if not file_path:
            return  # User cancelled
        
        try:
            debit_rows = []
            for label, db_col in self.debit_fields.items():
                amount = float(self.debit_inputs[label].text() or 0) if label in self.debit_inputs else 0
                lotes  = int(self.debit_lotes_inputs[label].text() or 0) if label in self.debit_lotes_inputs else 0
                debit_rows.append({'label': label, 'amount': amount, 'lotes': lotes})

            credit_rows = []
            for label, db_col in self.credit_fields.items():
                amount = float(self.credit_inputs[label].text() or 0) if label in self.credit_inputs else 0
                lotes  = int(self.credit_lotes_inputs[label].text() or 0) if label in self.credit_lotes_inputs else 0
                credit_rows.append({'label': label, 'amount': amount, 'lotes': lotes})

            meta = {
                'filter_label':      filter_label,
                'filter_value':      filter_value,
                'branch_name':       branch_name,
                'selected_date':     selected_date,
                'beginning_balance': float(self.beginning_balance_input.text() or 0),
                'debit_total':       float(self.debit_total_display.text() or 0),
                'credit_total':      float(self.credit_total_display.text() or 0),
                'ending_balance':    float(self.ending_balance_display.text() or 0),
                'cash_count':        float(self.cash_count_input.text() or 0),
                'cash_result':       float(self.cash_result_display.text() or 0),
            }
            _export_daily_cash_svc(debit_rows, credit_rows, meta, file_path)
            QMessageBox.information(self, "Export Successful", f"Daily Cash Count exported to:\n{file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Error exporting to Excel: {str(e)}")

    def show_full_brand_report_dialog(self):
        """Dialog to generate a multi-sheet Full Brand Report for all modules."""
        brand_label = self._brand_label
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Full Brand Report – {brand_label}")
        dialog.setMinimumWidth(_FULL_REPORT_DIALOG_WIDTH)

        layout = QVBoxLayout(dialog)
        layout.setSpacing(12)

        if self._is_brand_a:
            sheets_info = (
                "Daily Cash Count &nbsp;·&nbsp; Palawan &nbsp;·&nbsp; "
                "MC &nbsp;·&nbsp; Fund Transfer &nbsp;·&nbsp; Payable &nbsp;·&nbsp; "
                "Daily Transaction &nbsp;·&nbsp; Other Services &nbsp;·&nbsp; "
                "P&amp;L &nbsp;·&nbsp; New Sanla &nbsp;·&nbsp; New Renew &nbsp;·&nbsp; "
                "Global Other Services &nbsp;·&nbsp; FT HO"
            )
        else:
            sheets_info = (
                "Daily Cash Count &nbsp;·&nbsp; Palawan &nbsp;·&nbsp; "
                "MC &nbsp;·&nbsp; Fund Transfer &nbsp;·&nbsp; Payable &nbsp;·&nbsp; "
                "Global Payable &nbsp;·&nbsp; Payable Reports"
            )
        info_lbl = QLabel(
            f"Generates a comprehensive Excel workbook for <b>{brand_label}</b>.<br><br>"
            f"<b>Sheets:</b> {sheets_info}"
        )
        info_lbl.setWordWrap(True)
        info_lbl.setTextFormat(Qt.RichText)
        info_lbl.setStyleSheet("color:#2c3e50;font-size:11px;padding:6px;")
        layout.addWidget(info_lbl)

        # ── Filter type (Corporation / Group) ──────────────────────────────
        filter_grp = QGroupBox("Filter")
        filter_lay = QVBoxLayout(filter_grp)

        type_row = QHBoxLayout()
        self._fbr_corp_radio = QRadioButton("By Corporation")
        self._fbr_os_radio   = QRadioButton("By Group")
        self._fbr_corp_radio.setChecked(True)
        type_row.addWidget(self._fbr_corp_radio)
        type_row.addWidget(self._fbr_os_radio)
        type_row.addStretch()
        filter_lay.addLayout(type_row)

        sel_row = QHBoxLayout()
        self._fbr_corp_lbl  = QLabel("Corporation:")
        self._fbr_corp_lbl.setMinimumWidth(90)
        self._fbr_corp_combo = QComboBox()
        self._fbr_corp_combo.setMinimumWidth(260)
        try:
            rows = db_manager.execute_query(
                "SELECT name FROM corporations ORDER BY name"
            ) or []
            for r in rows:
                self._fbr_corp_combo.addItem(r['name'] if isinstance(r, dict) else r[0])
        except Exception:
            pass

        self._fbr_os_lbl  = QLabel("Group:")
        self._fbr_os_lbl.setMinimumWidth(90)
        self._fbr_os_combo = QComboBox()
        self._fbr_os_combo.setMinimumWidth(260)
        try:
            rows = db_manager.execute_query(
                "SELECT DISTINCT os_name FROM branches "
                "WHERE os_name IS NOT NULL AND os_name != '' ORDER BY os_name"
            ) or []
            for r in rows:
                self._fbr_os_combo.addItem(r['os_name'] if isinstance(r, dict) else r[0])
        except Exception:
            pass

        self._fbr_os_lbl.setVisible(False)
        self._fbr_os_combo.setVisible(False)

        sel_row.addWidget(self._fbr_corp_lbl)
        sel_row.addWidget(self._fbr_corp_combo)
        sel_row.addWidget(self._fbr_os_lbl)
        sel_row.addWidget(self._fbr_os_combo)
        sel_row.addStretch()
        filter_lay.addLayout(sel_row)
        layout.addWidget(filter_grp)

        def _toggle_fbr_filter():
            by_corp = self._fbr_corp_radio.isChecked()
            self._fbr_corp_lbl.setVisible(by_corp)
            self._fbr_corp_combo.setVisible(by_corp)
            self._fbr_os_lbl.setVisible(not by_corp)
            self._fbr_os_combo.setVisible(not by_corp)

        self._fbr_corp_radio.toggled.connect(_toggle_fbr_filter)

        # ── Branch Status filter ───────────────────────────────────────────
        reg_grp = QGroupBox("Branch Status")
        reg_lay = QHBoxLayout(reg_grp)
        reg_lay.addWidget(QLabel("Show:"))
        self._fbr_reg_filter = QComboBox()
        self._fbr_reg_filter.addItem("Registered Only", "registered")
        self._fbr_reg_filter.addItem("Not Registered",  "not_registered")
        self._fbr_reg_filter.addItem("All Branches",    "all")
        reg_lay.addWidget(self._fbr_reg_filter)
        reg_lay.addStretch()
        layout.addWidget(reg_grp)

        # ── Single date picker ─────────────────────────────────────────────
        date_grp = QGroupBox("Date")
        date_lay = QHBoxLayout(date_grp)
        date_lay.setSpacing(8)
        date_lay.addWidget(QLabel("Date:"))
        self._fbr_date = QDateEdit()
        self._fbr_date.setDisplayFormat("dd MMM yyyy")
        self._fbr_date.setCalendarPopup(True)
        self._fbr_date.setDate(QDate.currentDate())
        self._fbr_date.setMinimumWidth(150)
        date_lay.addWidget(self._fbr_date)
        date_lay.addStretch()
        layout.addWidget(date_grp)

        btn_lay = QHBoxLayout()
        gen_btn = QPushButton("Generate Report")
        gen_btn.setStyleSheet(
            "QPushButton{background-color:#27AE60;color:white;padding:8px 20px;"
            "font-weight:bold;border-radius:4px;}"
            "QPushButton:hover{background-color:#1E8449;}"
        )
        cancel_btn = QPushButton("Cancel")
        gen_btn.clicked.connect(lambda: self._generate_full_brand_report(dialog))
        cancel_btn.clicked.connect(dialog.reject)
        btn_lay.addStretch()
        btn_lay.addWidget(cancel_btn)
        btn_lay.addWidget(gen_btn)
        layout.addLayout(btn_lay)

        dialog.exec_()

    def _generate_full_brand_report(self, dialog):
        """Generate a multi-sheet Excel workbook with all module reports for the brand."""
        if not _OPENPYXL_AVAILABLE:
            QMessageBox.critical(self, "Missing Dependency",
                "The openpyxl package is required.\nInstall with: pip install openpyxl")
            return

        from admin.services.report_writers.context       import ReportContext
        from admin.services.report_writers.styles        import ReportStyles
        from admin.services.report_writers.grouped_sheet import write_grouped_sheet
        from admin.services.report_writers.dcc_sheet     import write_dcc_sheet
        from admin.services.report_writers.palawan_sheet import write_palawan_sheet
        from admin.services.report_writers.ft_sheet      import write_ft_sheet
        from admin.services.report_writers.depo_br_sheet import write_depo_br_sheet
        from admin.services.report_writers.pepp_sheet    import write_pepp_report_sheet

        selected_date = self._fbr_date.date().toString("yyyy-MM-dd")

        by_corp = self._fbr_corp_radio.isChecked()
        if by_corp:
            filter_type  = "corporation"
            filter_value = self._fbr_corp_combo.currentText().strip()
            filter_label = "Corporation"
        else:
            filter_type  = "os"
            filter_value = self._fbr_os_combo.currentText().strip()
            filter_label = "Group"

        if not filter_value:
            QMessageBox.warning(self, "Selection Required",
                f"Please select a {filter_label}.")
            return

        reg_filter  = self._fbr_reg_filter.currentData()   # "registered" | "not_registered" | "all"
        brand_label = self._brand_label
        safe_brand  = brand_label.replace(" ", "_")
        safe_filter = filter_value.replace(" ", "_").replace("/", "_")[:40]
        default_fn  = f"FullBrandReport_{safe_brand}_{safe_filter}_{selected_date}.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save Excel File", default_fn, "Excel Files (*.xlsx);;All Files (*)"
        )
        if not file_path:
            return

        from admin.pages.daily_transaction import (
            COLUMN_GROUPS          as DT_COLUMN_GROUPS,
            OTHER_SERVICES_COLUMN_GROUPS,
            PL_COLUMN_GROUPS,
        )

        try:
            try:
                col_rows = db_manager.execute_query(
                    "SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS "
                    "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = %s",
                    (self.daily_table,)
                )
                existing_cols = {r["COLUMN_NAME"] for r in col_rows} if col_rows else set()
            except Exception:
                existing_cols = set()

            ctx = ReportContext(
                db            = db_manager,
                brand_label   = brand_label,
                is_brand_a    = self._is_brand_a,
                filter_type   = filter_type,
                filter_value  = filter_value,
                filter_label  = filter_label,
                selected_date = selected_date,
                reg_filter    = reg_filter,
                daily_table   = self.daily_table,
                debit_fields  = self.debit_fields,
                credit_fields = self.credit_fields,
                bank_accounts = getattr(self, "BANK_ACCOUNTS", []),
            )
            styles = ReportStyles()

            wb = Workbook()

            # ════ Sheet 1 – Daily Cash Count ═══════════════════════════════════════════════
            ws1 = wb.active
            ws1.title = "Daily Cash Count"
            write_dcc_sheet(ws1, ctx, styles)

            # ════ Sheet 2 – Palawan ══════════════════════════════════════════════════
            ws2 = wb.create_sheet(title="Palawan")
            write_palawan_sheet(ws2, ctx, styles)

            # ════ Sheet 3 – MC ════════════════════════════════════════════════════════
            ws3 = wb.create_sheet(title="MC")
            mc_groups = [
                ("MC IN (SELLING)", [
                    ("Lotes",  ["mc_in_lotes"], True),
                    ("Amount", ["mc_in"],       False),
                ]),
                ("MC OUT (BUYING)", [
                    ("Lotes",  ["mc_out_lotes"], True),
                    ("Amount", ["mc_out"],       False),
                ]),
            ]
            write_grouped_sheet(ws3, ctx, styles, mc_groups, ctx.daily_table, show_amt_total=False)

            # ════ Sheet 4 – Fund Transfer ═══════════════════════════════════════════════
            ws4 = wb.create_sheet(title="Fund Transfer")
            write_ft_sheet(ws4, ctx, styles)

            # ════ Sheet 5 – Payable (Palawan Reconciliation) ═════════════════════════
            _payable_groups_a = [
                ("SEND OUT", [
                    ("S.O. Lotes",   ["sendout_lotes"],        True),
                    ("S.O. Capital", ["sendout_capital"],      False),
                    ("S.O. SC",      ["sendout_sc"],           False),
                    ("S.O. Comm.",   ["sendout_commission"],   False),
                    ("S.O. Total",   ["sendout_total"],        False),
                ]),
                ("PAY OUT", [
                    ("P.O. Lotes",   ["payout_lotes"],         True),
                    ("P.O. Capital", ["payout_capital"],       False),
                    ("P.O. SC",      ["payout_sc"],            False),
                    ("P.O. Comm.",   ["payout_commission"],    False),
                    ("P.O. Total",   ["payout_total"],         False),
                ]),
                ("INTERNATIONAL", [
                    ("Int. Lotes",   ["international_lotes"],        True),
                    ("Int. Capital", ["international_capital"],      False),
                    ("Int. SC",      ["international_sc"],           False),
                    ("Int. Comm.",   ["international_commission"],   False),
                    ("Int. Total",   ["international_total"],        False),
                ]),
                ("OTHER", [
                    ("SKID",      ["skid"],         False),
                    ("SKIR",      ["skir"],         False),
                    ("Cancel",    ["cancellation"], False),
                    ("P.O. Inc.", ["inc"],          False),
                ]),
            ]
            if ctx.is_brand_a:
                ws5a = wb.create_sheet(title="Palawan 60%")
                write_grouped_sheet(ws5a, ctx, styles, _payable_groups_a, "payable_tbl_brand_a",
                                    category_filter="60", show_amt_total=False)
                ws5b = wb.create_sheet(title="Palawan 30%")
                write_grouped_sheet(ws5b, ctx, styles, _payable_groups_a, "payable_tbl_brand_a",
                                    category_filter="30", show_amt_total=False)
            else:
                ws5 = wb.create_sheet(title="Payable")
                write_grouped_sheet(ws5, ctx, styles, _payable_groups_a, "payable_tbl_brand_a",
                                    category_filter="60", show_amt_total=False)

            # ════ Sheet 6+ – Brand-specific sheets ════════════════════════════════════════
            if ctx.is_brand_a:
                ws6 = wb.create_sheet(title="Daily Transaction")
                write_grouped_sheet(ws6, ctx, styles, DT_COLUMN_GROUPS,
                                    "daily_reports_brand_a", show_amt_total=False)

                ws7 = wb.create_sheet(title="Other Services")
                write_grouped_sheet(ws7, ctx, styles, OTHER_SERVICES_COLUMN_GROUPS,
                                    "daily_reports_brand_a", show_amt_total=False)

                ws8 = wb.create_sheet(title="P&L")
                write_grouped_sheet(ws8, ctx, styles, PL_COLUMN_GROUPS,
                                    "daily_reports_brand_a", show_amt_total=False)

                ws9 = wb.create_sheet(title="New Sanla")
                sanla_groups = [
                    ("JEWELRY EMPENO", [
                        ("Lotes",   ["empeno_jew_new_lotes"], True),
                        ("Capital", ["empeno_jew_new"],       False),
                    ]),
                    ("STORAGE EMPENO", [
                        ("Lotes",   ["empeno_sto_new_lotes"], True),
                        ("Capital", ["empeno_sto_new"],       False),
                    ]),
                ]
                write_grouped_sheet(ws9, ctx, styles, sanla_groups, ctx.daily_table,
                                    show_amt_total=False)

                ws10 = wb.create_sheet(title="New Renew")
                renew_groups = [
                    ("JEWELRY", [
                        ("JEW NEW Lotes",   ["empeno_jew_new_lotes"],           True),
                        ("JEW NEW Capital", ["empeno_jew_new"],                 False),
                        ("JEW RENEW Lotes", ["empeno_jew_renew_lotes"],         True),
                        ("JEW RENEW Cap.",  ["empeno_jew_renew"],               False),
                    ]),
                    ("STORAGE", [
                        ("STO NEW Lotes",   ["empeno_sto_new_lotes"],           True),
                        ("STO NEW Capital", ["empeno_sto_new"],                 False),
                        ("STO RENEW Lotes", ["fund_empeno_sto_renew_lotes"],    True),
                        ("STO RENEW Cap.",  ["fund_empeno_sto_renew"],          False),
                    ]),
                ]
                write_grouped_sheet(ws10, ctx, styles, renew_groups, ctx.daily_table,
                                    show_amt_total=False)

                ws11 = wb.create_sheet(title="Global Other Services")
                gos_groups = [
                    ("GCASH OUT",        [("Lotes", ["gcash_out_lotes"],        True),  ("Capital", ["gcash_out"],        False)]),
                    ("MONEYGRAM",        [("Lotes", ["moneygram_lotes"],        True),  ("Capital", ["moneygram"],        False)]),
                    ("TRANSFAST",        [("Lotes", ["transfast_lotes"],        True),  ("Capital", ["transfast"],        False)]),
                    ("RIA",              [("Lotes", ["ria_lotes"],              True),  ("Capital", ["ria"],              False)]),
                    ("SMART MONEY OUT",  [("Lotes", ["smart_money_out_lotes"],  True),  ("Capital", ["smart_money_out"],  False)]),
                    ("GCASH PADALA",     [("Lotes", ["gcash_padala_lotes"],     True),  ("Capital", ["gcash_padala"],     False)]),
                    ("ABRA OUT",         [("Lotes", ["abra_out_lotes"],         True),  ("Capital", ["abra_out"],         False)]),
                    ("REMITLY",          [("Lotes", ["remitly_lotes"],          True),  ("Capital", ["remitly"],          False)]),
                    ("PAL PAY CASH OUT", [("Lotes", ["pal_pay_cash_out_lotes"], True),  ("Capital", ["pal_pay_cash_out"], False)]),
                    ("MC OUT",           [("Lotes", ["mc_out_lotes"],           True),  ("Capital", ["mc_out"],           False)]),
                    ("EC PAY OUT",       [("Capital", ["ec_pay_out"],           False)]),
                ]
                write_grouped_sheet(ws11, ctx, styles, gos_groups, "daily_reports_brand_a",
                                    use_branch_join=True, show_amt_total=False)

                ws12 = wb.create_sheet(title="FT HO")
                ft_ho_groups = [
                    ("FUND TRANSFER HO", [
                        ("FT From Branch", ["fund_transfer_from_branch"],    False),
                        ("FT To HO",       ["fund_transfer_to_head_office"], False),
                        ("FT To Branch",   ["fund_transfer_to_branch"],      False),
                    ]),
                ]
                write_grouped_sheet(ws12, ctx, styles, ft_ho_groups, ctx.daily_table,
                                    show_amt_total=False)

                ws13 = wb.create_sheet(title="DEPO BR")
                write_depo_br_sheet(ws13, ctx, styles)

            else:
                ws6 = wb.create_sheet(title="Global Payable")
                global_payable_groups = [
                    ("SEND OUT", [
                        ("S.O. Lotes",   ["sendout_lotes"],    True),
                        ("S.O. Capital", ["sendout_capital"],  False),
                        ("S.O. SC",      ["sendout_sc"],       False),
                        ("S.O. Comm.",   ["sendout_commission"],False),
                        ("S.O. Total",   ["sendout_total"],    False),
                    ]),
                    ("PAY OUT", [
                        ("P.O. Lotes",   ["payout_lotes"],   True),
                        ("P.O. Capital", ["payout_capital"],  False),
                        ("P.O. SC",      ["payout_sc"],       False),
                        ("P.O. Comm.",   ["payout_commission"],False),
                        ("P.O. Total",   ["payout_total"],    False),
                    ]),
                    ("INTERNATIONAL", [
                        ("Int. Lotes",   ["international_lotes"],        True),
                        ("Int. Capital", ["international_capital"],      False),
                        ("Int. SC",      ["international_sc"],           False),
                        ("Int. Comm.",   ["international_commission"],   False),
                        ("Int. Total",   ["international_total"],        False),
                    ]),
                    ("OTHER", [
                        ("SKID",       ["skid"],         False),
                        ("SKIR",       ["skir"],         False),
                        ("Cancel",     ["cancellation"], False),
                        ("P.O. Inc.",  ["inc"],          False),
                    ]),
                ]
                write_grouped_sheet(ws6, ctx, styles, global_payable_groups,
                                    "payable_tbl_brand_a", use_branch_join=True, show_amt_total=False)

                ws7 = wb.create_sheet(title="Payable Reports")
                write_pepp_report_sheet(ws7, ctx, styles)

            # ════ Save ═══════════════════════════════════════════════════════════════════
            wb.save(file_path)
            dialog.accept()
            QMessageBox.information(
                self, "Export Successful",
                f"Full Brand Report exported to:\n{file_path}\n\n"
                f"{filter_label}: {filter_value}\n"
                f"Date: {selected_date}\n"
                f"Sheets generated: {len(wb.sheetnames)}"
            )

        except Exception as e:
            import traceback
            QMessageBox.critical(
                self, "Export Error",
                f"Error exporting: {str(e)}\n\n{traceback.format_exc()[:600]}"
            )

    def load_entry_by_date(self):
        filter_type = self.filter_type_selector.currentData()
        branch_name = self.branch_selector.currentText()
        selected_date = self.date_picker.date().toString("yyyy-MM-dd")

        if not branch_name:
            QMessageBox.warning(self, "Missing Selection", "Please select a branch.")
            return

        try:
            data = load_entry_data(db_manager, self.daily_table, branch_name, selected_date)
            if not data:
                QMessageBox.information(self, "No Entry", f"No entry found for {selected_date}.")
                self.clear_all_fields()
                return
            if not self._is_brand_a:
                data = patch_brand_b_palawan(data, db_manager, branch_name, selected_date)

            self._current_entry_data = data

            # CRITICAL FIX: Set flag FIRST to prevent _recalc_totals from running
            self._loading_report = True

            # Block all signals during load to prevent recalculation
            # while we're populating fields
            for inp in list(self.debit_inputs.values()) + list(self.credit_inputs.values()) + \
                       list(self.debit_lotes_inputs.values()) + list(self.credit_lotes_inputs.values()) + \
                       [self.beginning_balance_input, self.cash_count_input]:
                inp.blockSignals(True)

            beginning = float(data.get('beginning_balance') or 0)
            self.beginning_balance_input.setText(f"{beginning:.2f}")
            self.cash_count_input.setText(f"{float(data.get('cash_count') or 0):.2f}")

            for ui_label, db_column in self.debit_fields.items():
                if db_column in data and data[db_column] is not None:
                    self.debit_inputs[ui_label].setText(str(data[db_column]))
                else:
                    self.debit_inputs[ui_label].setText("0.00")

                lotes_col = db_column + "_lotes"
                if lotes_col in data and data[lotes_col] is not None:
                    self.debit_lotes_inputs[ui_label].setText(str(data[lotes_col]))
                else:
                    self.debit_lotes_inputs[ui_label].setText("0")

            for ui_label, db_column in self.credit_fields.items():
                if db_column in data and data[db_column] is not None:
                    self.credit_inputs[ui_label].setText(str(data[db_column]))
                else:
                    self.credit_inputs[ui_label].setText("0.00")

                lotes_col = db_column + "_lotes"
                if lotes_col in data and data[lotes_col] is not None:
                    self.credit_lotes_inputs[ui_label].setText(str(data[lotes_col]))
                else:
                    self.credit_lotes_inputs[ui_label].setText("0")

            # CRITICAL FIX: Display stored values from database, don't recalculate
            # This prevents variance discrepancies caused by loading/merging palawan data incorrectly
            debit_total = float(data.get('debit_total') or 0)
            credit_total = float(data.get('credit_total') or 0)
            ending_balance = float(data.get('ending_balance') or 0)
            cash_count = float(data.get('cash_count') or 0)
            cash_result = float(data.get('cash_result') or 0)
            variance_status = data.get('variance_status', 'balanced')

            self.debit_total_display.setText(f"{debit_total:.2f}")
            self.credit_total_display.setText(f"{credit_total:.2f}")
            self.ending_balance_display.setText(f"{ending_balance:.2f}")
            self.cash_result_display.setText(f"{cash_result:.2f}")

            # Unblock signals after displaying stored values
            for inp in list(self.debit_inputs.values()) + list(self.credit_inputs.values()) + \
                       list(self.debit_lotes_inputs.values()) + list(self.credit_lotes_inputs.values()) + \
                       [self.beginning_balance_input, self.cash_count_input]:
                inp.blockSignals(False)

            # Allow editing after load is FULLY complete (including palawan loading at 500ms)
            QTimer.singleShot(1000, lambda: setattr(self, '_loading_report', False))
            self._set_variance_display(variance_status)

            self.selected_bank_account = data.get('fund_transfer_bank_account')
            if self.bank_account_btn:

                ft_ho_bd = data.get('ft_ho_breakdown')
                if ft_ho_bd:
                    try:
                        bd_list = json.loads(ft_ho_bd)
                        self.bank_account_btn.setText(f"🏦 View ({len(bd_list)})")
                        self.bank_account_btn.setToolTip(f"View {len(bd_list)} fund transfer(s) breakdown")
                    except Exception:
                        self.bank_account_btn.setText("View")
                        self.bank_account_btn.setToolTip("View fund transfer breakdown")
                elif self.selected_bank_account:

                    bank_name = "View"
                    for bank in self.BANK_ACCOUNTS:
                        if bank['id'] == self.selected_bank_account:
                            bank_name = bank['bank_name']
                            break
                    self.bank_account_btn.setText(f"🏦 {bank_name}")
                    self.bank_account_btn.setToolTip("View selected bank account")
                else:
                    self.bank_account_btn.setText("View")
                    self.bank_account_btn.setToolTip("No bank account selected")


            self.selected_branch_dest = data.get('fund_transfer_to_branch_dest')
            if self.branch_dest_btn:
                self.branch_dest_btn.setText("View")
                if self.selected_branch_dest:
                    self.branch_dest_btn.setToolTip(f"Destination: {self.selected_branch_dest}")
                else:
                    self.branch_dest_btn.setToolTip("No destination branch specified")

       
            self.selected_from_branch_dest = data.get('fund_transfer_from_branch_dest')
            if self.from_branch_dest_btn:
                self.from_branch_dest_btn.setText("View")
                if self.selected_from_branch_dest:
                    self.from_branch_dest_btn.setToolTip(f"Source: {self.selected_from_branch_dest}")
                else:
                    self.from_branch_dest_btn.setToolTip("No source branch specified")

            self.current_record_id = data.get('id')

            # Check review status
            brand_key = "A" if self._is_brand_a else "B"
            try:
                review_row = self.db.execute_query(
                    "SELECT id FROM admin_review_marks WHERE brand = %s AND branch = %s AND report_date = %s",
                    [brand_key, branch_name, selected_date]
                )
                is_reviewed = bool(review_row)
                self.reviewed_checkbox.blockSignals(True)
                self.reviewed_checkbox.setChecked(is_reviewed)
                self.reviewed_checkbox.setText("✅ Reviewed" if is_reviewed else "Pending review")
                self.reviewed_checkbox.setStyleSheet(self.reviewed_checkbox.styleSheet().replace(
                    "color: #c0392b;" if is_reviewed else "color: #2e7d32;",
                    "color: #2e7d32;" if is_reviewed else "color: #c0392b;"
                ))
                self.reviewed_checkbox.blockSignals(False)
                self.reviewed_checkbox.setEnabled(True)
            except Exception:
                self.reviewed_checkbox.setEnabled(False)

            # Load palawan details into collapsible
            # CRITICAL FIX: Delay palawan loading to avoid interfering with display
            # Schedule it to run after all display updates are complete
            QTimer.singleShot(500, lambda: self._load_palawan_details(data))

            QMessageBox.information(self, "✅ Loaded", f"Entry for {selected_date} loaded successfully!")

        except Exception as e:
            logger.error("Error loading entry: %s", e)
            QMessageBox.critical(self, "Error", f"Failed to load entry: {e}")

    def _on_review_toggled(self, checked):
        """Save or remove the review mark for the current entry."""
        branch_name = self.branch_selector.currentText()
        selected_date = self.date_picker.date().toString("yyyy-MM-dd")
        brand_key = "A" if self._is_brand_a else "B"
        if not branch_name or not selected_date:
            return
        try:
            if checked:
                self.db.execute_query(
                    "INSERT IGNORE INTO admin_review_marks (brand, branch, report_date) VALUES (%s, %s, %s)",
                    [brand_key, branch_name, selected_date]
                )
                self.reviewed_checkbox.setText("Reviewed")
                self.reviewed_checkbox.setStyleSheet(self.reviewed_checkbox.styleSheet().replace(
                    "color: #c0392b;", "color: #2e7d32;"
                ))
            else:
                self.db.execute_query(
                    "DELETE FROM admin_review_marks WHERE brand = %s AND branch = %s AND report_date = %s",
                    [brand_key, branch_name, selected_date]
                )
                self.reviewed_checkbox.setText("Pending review")
                self.reviewed_checkbox.setStyleSheet(self.reviewed_checkbox.styleSheet().replace(
                    "color: #2e7d32;", "color: #c0392b;"
                ))
        except Exception as e:
            logger.error("Error toggling review mark: %s", e)

    def get_current_entry_data(self):
        """Return the current loaded entry data for breakdown views"""
        return getattr(self, '_current_entry_data', None)

    def show_mc_breakdown(self, field_type):
        """Show MC currency breakdown in a dialog"""
        entry = self.get_current_entry_data()
        if not entry:
            QMessageBox.information(self, "No Entry Loaded", 
                "Please load an entry first to view MC breakdown.")
            return
        
        # Determine which details field to use
        details_key = "mc_in_details" if field_type == "MC In" else "mc_out_details"
        breakdown = []
        
        if entry.get(details_key):
            try:
                breakdown = json.loads(entry[details_key])
            except Exception:
                breakdown = []
        
        if not breakdown:
            QMessageBox.information(self, "No Breakdown", 
                f"No {field_type} currency breakdown available for this entry.")
            return
        
        # Create breakdown dialog
        dialog = QDialog(self)
        dialog.setWindowTitle(f"{field_type} Currency Breakdown (View Only)")
        dialog.setMinimumWidth(600)
        dialog.setMinimumHeight(300)
        
        layout = QVBoxLayout(dialog)
        
        # Header
        header_text = "SELLING Currency (Money In)" if field_type == "MC In" else "BUYING Currency (Money Out)"
        header_color = "#22C55E" if field_type == "MC In" else "#DC2626"
        
        header = QLabel(f"💱 {header_text}")
        header.setStyleSheet(f"font-size: 14px; font-weight: 800; color: {header_color}; padding: 10px;")
        layout.addWidget(header)
        
        # Table for breakdown
        table = QTableWidget()
        table.setColumnCount(5)
        table.setHorizontalHeaderLabels(["Currency", "Pcs", "Denomination", "Rate", "Total"])
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        table.setEditTriggers(QTableWidget.NoEditTriggers)
        
        total = 0.0
        for entry_data in breakdown:
            row = table.rowCount()
            table.insertRow(row)
            
            currency = entry_data.get('currency', 'Unknown')
            qty = entry_data.get('quantity', 0)
            denom = entry_data.get('denomination', 0.0)
            rate = entry_data.get('rate', 0.0)
            total_php = entry_data.get('total_php', 0)
            if total_php == 0 and qty > 0 and rate > 0:
                if denom > 0:
                    total_php = qty * denom * rate
                else:
                    total_php = qty * rate
            total += total_php
            
            table.setItem(row, 0, QTableWidgetItem(str(currency)))
            table.setItem(row, 1, QTableWidgetItem(str(qty)))
            table.setItem(row, 2, QTableWidgetItem(f"{denom:,.2f}" if denom else "-"))
            table.setItem(row, 3, QTableWidgetItem(f"{rate:,.2f}"))
            table.setItem(row, 4, QTableWidgetItem(f"₱{total_php:,.2f}"))
        
        layout.addWidget(table)
        

        total_label = QLabel(f"TOTAL: ₱{total:,.2f}")
        total_label.setStyleSheet(f"font-size: 16px; font-weight: 800; color: {header_color}; padding: 10px;")
        total_label.setAlignment(Qt.AlignRight)
        layout.addWidget(total_label)
        

        close_btn = QPushButton("Close")
        close_btn.setStyleSheet("""
            QPushButton {
                background-color: #6B7280; color: white;
                padding: 10px 24px; border-radius: 6px;
                font-weight: 600;
            }
            QPushButton:hover { background-color: #4B5563; }
        """)
        close_btn.clicked.connect(dialog.accept)
        layout.addWidget(close_btn)
        
        dialog.exec_()

    def clear_all_fields(self):
        self._current_entry_data = None
        self.beginning_balance_input.setText("0.00")
        self.cash_count_input.setText("0.00")
        self.ending_balance_display.setText("0.00")
        self.cash_result_display.setText("0.00")
        self.debit_total_display.setText("0.00")
        self.credit_total_display.setText("0.00")

        for input_field in self.debit_inputs.values():
            input_field.setText("0.00")
        for input_field in self.credit_inputs.values():
            input_field.setText("0.00")
        for input_field in self.debit_lotes_inputs.values():
            input_field.setText("0")
        for input_field in self.credit_lotes_inputs.values():
            input_field.setText("0")
        

        self.selected_bank_account = None
        if self.bank_account_btn:
            self.bank_account_btn.setText("View")
            self.bank_account_btn.setToolTip("No bank account selected")
        

        self.selected_branch_dest = None
        if self.branch_dest_btn:
            self.branch_dest_btn.setText("View")
            self.branch_dest_btn.setToolTip("No destination branch specified")
        

        self.selected_from_branch_dest = None
        if self.from_branch_dest_btn:
            self.from_branch_dest_btn.setText("View")
            self.from_branch_dest_btn.setToolTip("No source branch specified")
        

        self.variance_status_display.setText("—")
        self.variance_status_display.setStyleSheet(
            "font-weight: bold; font-size: 12px; padding: 5px 10px; border-radius: 4px;"
        )

    
        self.reviewed_checkbox.blockSignals(True)
        self.reviewed_checkbox.setChecked(False)
        self.reviewed_checkbox.setText("Pending review")
        self.reviewed_checkbox.setStyleSheet(self.reviewed_checkbox.styleSheet().replace(
            "color: #2e7d32;", "color: #c0392b;"
        ))
        self.reviewed_checkbox.blockSignals(False)
        self.reviewed_checkbox.setEnabled(False)

    def _recalc_totals(self):
       
        if hasattr(self, '_loading_report') and self._loading_report:
            return

        try:
            beginning = self._parse_money(self.beginning_balance_input.text())
        except ValueError:
            beginning = 0.0
        try:
            cash_count = self._parse_money(self.cash_count_input.text())
        except ValueError:
            cash_count = 0.0

        debit_sum = 0.0
        for inp in self.debit_inputs.values():
            try:
                debit_sum += self._parse_money(inp.text())
            except ValueError:
                pass

        credit_sum = 0.0
        for inp in self.credit_inputs.values():
            try:
                credit_sum += self._parse_money(inp.text())
            except ValueError:
                pass

        b = self._calculate_balances(beginning, debit_sum, credit_sum, cash_count)

        self.debit_total_display.setText(f"{b['debit_total']:.2f}")
        self.credit_total_display.setText(f"{b['credit_total']:.2f}")
        self.ending_balance_display.setText(f"{b['ending_balance']:.2f}")
        self.cash_result_display.setText(f"{b['cash_result']:.2f}")
        self._set_variance_display(b['variance_status'])

    def _save_palawan_details(self):
        """Save palawan details to payable_tbl_brand_a (insert or update)."""
        branch_name = self.branch_selector.currentText()
        selected_date = self.date_picker.date().toString("yyyy-MM-dd")

        if not branch_name:
            QMessageBox.warning(self, "Selection Required", "Please select a branch first.")
            return

        # Column mapping: admin_dashboard field names → payable_tbl_brand_a column names
        column_mapping = {
            "palawan_sendout_principal": "sendout_capital",
            "palawan_sendout_sc": "sendout_sc",
            "palawan_sendout_commission": "sendout_commission",
            "palawan_sendout_lotes_total": "sendout_lotes",
            "palawan_sendout_regular_total": "sendout_total",
            "palawan_payout_principal": "payout_capital",
            "palawan_payout_sc": "payout_sc",
            "palawan_payout_commission": "payout_commission",
            "palawan_payout_lotes_total": "payout_lotes",
            "palawan_payout_regular_total": "payout_total",
            "palawan_international_principal": "international_capital",
            "palawan_international_sc": "international_sc",
            "palawan_international_commission": "international_commission",
            "palawan_international_lotes_total": "international_lotes",
            "palawan_international_regular_total": "international_total",
            "palawan_pay_out_incentives": "inc",
            "palawan_suki_discounts": "skid",
            "palawan_suki_rebates": "skir",
            "palawan_cancel": "cancellation",
        }

        try:
            # Extract values from palawan inputs
            values = {}
            for field_name, col_name in column_mapping.items():
                if field_name in self.palawan_inputs:
                    try:
                        text = self.palawan_inputs[field_name].text().strip()
                        values[col_name] = float(text) if text else 0
                    except (ValueError, AttributeError):
                        values[col_name] = 0
                else:
                    values[col_name] = 0

            # Get corporation from currently filtered/selected data
            corporation = self.corporation_selector.currentText() if hasattr(self, 'corporation_selector') else ""
            if not corporation:
                QMessageBox.warning(self, "Selection Required", "Please select a corporation/group first.")
                return

            # Build INSERT ... ON DUPLICATE KEY UPDATE query
            sql = """
                INSERT INTO payable_tbl_brand_a
                (corporation, branch, date,
                 sendout_lotes, sendout_capital, sendout_sc, sendout_commission, sendout_total,
                 payout_lotes, payout_capital, payout_sc, payout_commission, payout_total,
                 international_lotes, international_capital, international_sc, international_commission, international_total,
                 inc, skid, skir, cancellation)
                VALUES
                (%s, %s, %s,
                 %s, %s, %s, %s, %s,
                 %s, %s, %s, %s, %s,
                 %s, %s, %s, %s, %s,
                 %s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE
                sendout_lotes = VALUES(sendout_lotes),
                sendout_capital = VALUES(sendout_capital),
                sendout_sc = VALUES(sendout_sc),
                sendout_commission = VALUES(sendout_commission),
                sendout_total = VALUES(sendout_total),
                payout_lotes = VALUES(payout_lotes),
                payout_capital = VALUES(payout_capital),
                payout_sc = VALUES(payout_sc),
                payout_commission = VALUES(payout_commission),
                payout_total = VALUES(payout_total),
                international_lotes = VALUES(international_lotes),
                international_capital = VALUES(international_capital),
                international_sc = VALUES(international_sc),
                international_commission = VALUES(international_commission),
                international_total = VALUES(international_total),
                inc = VALUES(inc),
                skid = VALUES(skid),
                skir = VALUES(skir),
                cancellation = VALUES(cancellation),
                updated_at = CURRENT_TIMESTAMP
            """

            params = (
                corporation, branch_name, selected_date,
                values["sendout_lotes"], values["sendout_capital"], values["sendout_sc"], values["sendout_commission"], values["sendout_total"],
                values["payout_lotes"], values["payout_capital"], values["payout_sc"], values["payout_commission"], values["payout_total"],
                values["international_lotes"], values["international_capital"], values["international_sc"], values["international_commission"], values["international_total"],
                values["inc"], values["skid"], values["skir"], values["cancellation"],
            )

            db_manager.execute_query(sql, params)

            QMessageBox.information(
                self,
                "Success",
                f"✓ Palawan details saved for:\n\n"
                f"Branch: {branch_name}\n"
                f"Date: {selected_date}\n"
                f"Corporation: {corporation}\n\n"
                f"Data saved to payable_tbl_brand_a"
            )

        except Exception as e:
            logger.error(f"Error saving palawan details: {e}")
            QMessageBox.critical(self, "Error", f"Failed to save palawan details:\n\n{str(e)}")

    def save_entry(self):
        """Save edited entry to the database"""
        if not hasattr(self, 'current_record_id') or not self.current_record_id:
            QMessageBox.warning(self, "No Entry Loaded", "Please load an entry first before saving.")
            return

        branch_name = self.branch_selector.currentText()
        selected_date = self.date_picker.date().toString("yyyy-MM-dd")

        if not branch_name:
            QMessageBox.warning(self, "Selection Required", "Please select a branch.")
            return

        reply = QMessageBox.question(
            self,
            "Confirm Save",
            f"Are you sure you want to save changes for:\n\n"
            f"Branch: {branch_name}\n"
            f"Date: {selected_date}\n\n"
            f"This will update the existing entry.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )

        if reply == QMessageBox.No:
            return

        try:

            # Collect widget values
            debit_amounts = {}
            debit_lotes_vals = {}
            for ui_label in self.debit_fields:
                try:
                    debit_amounts[ui_label] = self._parse_money(self.debit_inputs[ui_label].text())
                except (ValueError, KeyError):
                    debit_amounts[ui_label] = 0
                try:
                    debit_lotes_vals[ui_label] = int(self.debit_lotes_inputs[ui_label].text().strip() or 0)
                except (ValueError, KeyError):
                    debit_lotes_vals[ui_label] = 0

            credit_amounts = {}
            credit_lotes_vals = {}
            for ui_label in self.credit_fields:
                try:
                    credit_amounts[ui_label] = self._parse_money(self.credit_inputs[ui_label].text())
                except (ValueError, KeyError):
                    credit_amounts[ui_label] = 0
                try:
                    credit_lotes_vals[ui_label] = int(self.credit_lotes_inputs[ui_label].text().strip() or 0)
                except (ValueError, KeyError):
                    credit_lotes_vals[ui_label] = 0

            try:
                beginning = self._parse_money(self.beginning_balance_input.text())
            except ValueError:
                beginning = 0
            try:
                cash_count = self._parse_money(self.cash_count_input.text())
            except ValueError:
                cash_count = 0

            update_data, b = build_save_payload(
                self.debit_fields, debit_amounts, debit_lotes_vals,
                self.credit_fields, credit_amounts, credit_lotes_vals,
                beginning, cash_count,
            )
            variance_status = b['variance_status']
            rows_affected = persist_entry(db_manager, self.daily_table, self.current_record_id, update_data)

            _PAYABLE_SECTION_COLS = {
                'palawan_sendout_principal', 'palawan_sendout_sc', 'palawan_sendout_commission',
                'palawan_sendout_lotes_total', 'palawan_sendout_regular_total',
                'palawan_payout_principal', 'palawan_payout_sc', 'palawan_payout_commission',
                'palawan_payout_lotes_total', 'palawan_payout_regular_total',
                'palawan_international_principal', 'palawan_international_sc',
                'palawan_international_commission', 'palawan_international_lotes_total',
                'palawan_international_regular_total',
                # adjustments — stored in payable_tbl_brand_a as skid/skir/cancellation/inc
                'palawan_suki_discounts', 'palawan_suki_rebates',
                'palawan_cancel', 'palawan_pay_out_incentives',
            }
            payable_a_vals = {}
            for db_col, widget in getattr(self, 'palawan_inputs', {}).items():
                if widget.isReadOnly():
                    continue  
                raw_text = widget.text().strip()
                if self._is_brand_a and db_col in _PAYABLE_SECTION_COLS and raw_text == "":
                   
                    continue
                try:
                    val = float(raw_text or 0)
                except ValueError:
                    val = 0
                if self._is_brand_a and db_col in _PAYABLE_SECTION_COLS:
                    payable_a_vals[db_col] = val  
                else:
                    update_data[db_col] = val
            # Also collect auto-calc totals
            for section in ("sendout", "payout", "international"):
                disp = getattr(self, 'palawan_total_displays', {}).get(section)
                if disp:
                    section_input_keys = (
                        f"palawan_{section}_principal",
                        f"palawan_{section}_sc",
                        f"palawan_{section}_commission",
                        f"palawan_{section}_lotes_total",
                    )
                    has_section_input = any(
                        getattr(self, 'palawan_inputs', {}).get(k, QLineEdit()).text().strip() != ""
                        for k in section_input_keys
                    )
                    try:
                        total_val = float(disp.text() or 0)
                    except ValueError:
                        total_val = 0
                    col = f"palawan_{section}_regular_total"
                    if self._is_brand_a and has_section_input:
                        payable_a_vals[col] = total_val
                    elif not self._is_brand_a:
                        update_data[col] = total_val

            if self._is_brand_a and payable_a_vals:
                _col_map = {
                    'palawan_sendout_principal':           'sendout_capital',
                    'palawan_sendout_sc':                  'sendout_sc',
                    'palawan_sendout_commission':          'sendout_commission',
                    'palawan_sendout_lotes_total':         'sendout_lotes',
                    'palawan_sendout_regular_total':       'sendout_total',
                    'palawan_payout_principal':            'payout_capital',
                    'palawan_payout_sc':                   'payout_sc',
                    'palawan_payout_commission':           'payout_commission',
                    'palawan_payout_lotes_total':          'payout_lotes',
                    'palawan_payout_regular_total':        'payout_total',
                    'palawan_international_principal':     'international_capital',
                    'palawan_international_sc':            'international_sc',
                    'palawan_international_commission':    'international_commission',
                    'palawan_international_lotes_total':   'international_lotes',
                    'palawan_international_regular_total': 'international_total',
                    'palawan_suki_discounts':              'skid',
                    'palawan_suki_rebates':                'skir',
                    'palawan_cancel':                      'cancellation',
                    'palawan_pay_out_incentives':          'inc',
                }
                entry_data   = getattr(self, '_current_entry_data', {}) or {}
                corporation  = (entry_data.get('corporation') or "").strip()
                if not corporation and getattr(self, 'current_record_id', None):
                    try:
                        rec = self.db.execute_query(
                            f"SELECT corporation FROM {self.daily_table} WHERE id=%s LIMIT 1",
                            [self.current_record_id]
                        )
                        if rec and rec[0].get('corporation'):
                            corporation = str(rec[0].get('corporation')).strip()
                    except Exception as ce:
                        logger.error("_save resolve corporation by id: %s", ce)
                if not corporation:
                    corporation = self.corp_selector.currentText().strip()
                branch_name  = self.branch_selector.currentText()
                p_cols        = {_col_map[k]: v for k, v in payable_a_vals.items() if k in _col_map}
                if p_cols:
                    upsert_palawan_payable(db_manager, corporation, branch_name, selected_date, p_cols)

            if rows_affected is not None and rows_affected > 0:

                self.debit_total_display.setText(f"{b['debit_total']:.2f}")
                self.credit_total_display.setText(f"{b['credit_total']:.2f}")
                self.ending_balance_display.setText(f"{b['ending_balance']:.2f}")
                self.cash_result_display.setText(f"{b['cash_result']:.2f}")
                
                self._set_variance_display(variance_status)
                
        
                QMessageBox.information(
                    self,
                    "✅ Saved",
                    f"Entry for {selected_date} has been updated successfully!"
                )
            else:
                QMessageBox.warning(
                    self,
                    "No Changes",
                    f"No entry was updated. The entry may not exist."
                )

        except Exception as e:
            logger.error("Error saving entry: %s", e)
            QMessageBox.critical(self, "Error", f"Failed to save entry: {e}")


# ── Daily-cash builder methods (defined in admin/builders/daily_cash_builder.py)
from admin.builders import daily_cash_builder as _dcb

AdminDashboard.build_daily_cash_widget    = _dcb.build_daily_cash_widget
AdminDashboard._build_header_frame        = _dcb._build_header_frame
AdminDashboard._build_debit_box           = _dcb._build_debit_box
AdminDashboard._build_credit_box          = _dcb._build_credit_box
AdminDashboard._build_totals_and_results  = _dcb._build_totals_and_results
AdminDashboard._build_actions_and_palawan = _dcb._build_actions_and_palawan
AdminDashboard._show_ft_ho_breakdown      = _dcb._show_ft_ho_breakdown
AdminDashboard._show_salary_breakdown     = _dcb._show_salary_breakdown
AdminDashboard._show_motor_breakdown      = _dcb._show_motor_breakdown
AdminDashboard._show_jew_breakdown        = _dcb._show_jew_breakdown
