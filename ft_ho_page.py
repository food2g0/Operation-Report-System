from PyQt5.QtWidgets import (
    QWidget, QLabel, QVBoxLayout, QHBoxLayout, QComboBox, QTableWidget,
    QTableWidgetItem, QHeaderView, QSizePolicy, QPushButton, QFileDialog,
    QMessageBox, QAbstractItemView,
)
from PyQt5.QtGui import QFont, QColor, QBrush
from PyQt5.QtCore import Qt
from db_connect_pooled import db_manager
from db_worker import run_func_async
from date_range_widget import DateRangeWidget


class FTHOPage(QWidget):
    """Fund Transfer Head Office page — Brand A only.

    Displays per-branch fund-transfer columns from daily_reports_brand_a
    with an editable Remarks column persisted in ft_ho_remarks table.
    """

    DAILY_TABLE = "daily_reports_brand_a"

    def __init__(self, account_type=1):
        super().__init__()
        self.account_type = account_type
        self._building = False
        self._build_ui()
        self._ensure_remarks_table()
        self.load_groups()

    # ── UI ────────────────────────────────────────────────────────────────

    def _build_ui(self):
        layout = QVBoxLayout()
        layout.setContentsMargins(5, 5, 5, 5)
        layout.setSpacing(4)
        self.setLayout(layout)

        # Filter row
        filter_row = QHBoxLayout()
        filter_row.setSpacing(8)

        filter_row.addWidget(QLabel("Group:"))
        self.group_selector = QComboBox()
        self.group_selector.currentTextChanged.connect(self.populate_table)
        filter_row.addWidget(self.group_selector, 1)

        self.date_range_widget = DateRangeWidget()
        self.date_range_widget.dateRangeChanged.connect(self.populate_table)
        filter_row.addWidget(self.date_range_widget)

        filter_row.addWidget(QLabel("Status:"))
        self.reg_filter_selector = QComboBox()
        self.reg_filter_selector.addItem("Registered Only", "registered")
        self.reg_filter_selector.addItem("Not Registered", "not_registered")
        self.reg_filter_selector.addItem("All Branches", "all")
        self.reg_filter_selector.currentIndexChanged.connect(self.populate_table)
        filter_row.addWidget(self.reg_filter_selector)

        layout.addLayout(filter_row)

        # Table
        self.table = QTableWidget()
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels([
            "Branches", "FT From Branch", "FT To Head Office",
            "FT To Branch", "Remarks",
        ])
        self.table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.table.setEditTriggers(
            QAbstractItemView.DoubleClicked | QAbstractItemView.SelectedClicked
        )
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(4, QHeaderView.Stretch)
        self.table.setStyleSheet("""
            QTableWidget {
                gridline-color: #000;
                font-size: 11px;
            }
            QHeaderView::section {
                background-color: #34495e;
                color: white;
                font-weight: bold;
                border: 1px solid #2c3e50;
                padding: 6px;
            }
        """)
        self.table.itemChanged.connect(self._on_item_changed)
        layout.addWidget(self.table)

        # Export button
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        self.export_btn = QPushButton("📊 Export to Excel")
        self.export_btn.clicked.connect(self.export_to_excel)
        self.export_btn.setStyleSheet("""
            QPushButton {
                background-color: #217346; color: white;
                padding: 8px 16px; border-radius: 4px;
                font-weight: bold; font-size: 12px;
            }
            QPushButton:hover { background-color: #1a5c37; }
        """)
        btn_row.addWidget(self.export_btn)
        layout.addLayout(btn_row)

    # ── DB helpers ────────────────────────────────────────────────────────

    def _ensure_remarks_table(self):
        try:
            db_manager.execute_query("""
                CREATE TABLE IF NOT EXISTS ft_ho_remarks (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    branch VARCHAR(255) NOT NULL,
                    report_date DATE NOT NULL,
                    remark TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    UNIQUE KEY uq_ft_ho_remark (branch, report_date)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
            """)
        except Exception as e:
            print(f"ft_ho_remarks table create: {e}")

    def load_groups(self):
        self.group_selector.blockSignals(True)
        self.group_selector.clear()
        try:
            rows = db_manager.execute_query(
                "SELECT DISTINCT os_name FROM branches "
                "WHERE os_name IS NOT NULL AND os_name != '' "
                "ORDER BY os_name"
            )
            if rows:
                for r in rows:
                    name = r['os_name'] if isinstance(r, dict) else r[0]
                    self.group_selector.addItem(name)
        except Exception as e:
            print(f"Error loading groups: {e}")
        self.group_selector.blockSignals(False)
        self.populate_table()

    # ── Data loading ──────────────────────────────────────────────────────

    def populate_table(self):
        group = self.group_selector.currentText()
        if not group:
            return

        date_start, date_end = self.date_range_widget.get_date_range()
        is_range = self.date_range_widget.is_range_mode()
        reg_filter = self.reg_filter_selector.currentData() if hasattr(self, 'reg_filter_selector') else "all"

        # Build registration clause
        reg_clause = ""
        if reg_filter == "registered":
            reg_clause = "AND b.is_registered = 1"
        elif reg_filter == "not_registered":
            reg_clause = "AND (b.is_registered = 0 OR b.is_registered IS NULL)"

        if is_range and date_start != date_end:
            data_query = f"""
                SELECT b.name as branch,
                       SUM(COALESCE(dr.fund_transfer_from_branch, 0)) AS ft_from_branch,
                       SUM(COALESCE(dr.fund_transfer_to_head_office, 0)) AS ft_to_ho,
                       SUM(COALESCE(dr.fund_transfer_to_branch, 0)) AS ft_to_branch
                FROM branches b
                LEFT JOIN {self.DAILY_TABLE} dr
                    ON b.name COLLATE utf8mb4_general_ci = dr.branch COLLATE utf8mb4_general_ci
                    AND dr.date >= %s AND dr.date <= %s
                WHERE b.os_name = %s
                  {reg_clause}
                GROUP BY b.name
                ORDER BY b.name
            """
            data_params = (date_start, date_end, group)
        else:
            data_query = f"""
                SELECT b.name as branch,
                       COALESCE(dr.fund_transfer_from_branch, 0) AS ft_from_branch,
                       COALESCE(dr.fund_transfer_to_head_office, 0) AS ft_to_ho,
                       COALESCE(dr.fund_transfer_to_branch, 0) AS ft_to_branch
                FROM branches b
                LEFT JOIN {self.DAILY_TABLE} dr
                    ON b.name COLLATE utf8mb4_general_ci = dr.branch COLLATE utf8mb4_general_ci
                    AND dr.date = %s
                WHERE b.os_name = %s
                  {reg_clause}
                ORDER BY b.name
            """
            data_params = (date_start, group)

        # Remarks query — load all remarks for the date(s) so we can map them
        if is_range and date_start != date_end:
            remark_query = """
                SELECT branch, GROUP_CONCAT(remark SEPARATOR '; ') AS remark
                FROM ft_ho_remarks
                WHERE report_date >= %s AND report_date <= %s
                GROUP BY branch
            """
            remark_params = (date_start, date_end)
        else:
            remark_query = """
                SELECT branch, remark FROM ft_ho_remarks WHERE report_date = %s
            """
            remark_params = (date_start,)

        _dq, _dp, _rq, _rp = data_query, data_params, remark_query, remark_params

        def _fetch():
            rows = db_manager.execute_query(_dq, _dp) or []
            remarks = db_manager.execute_query(_rq, _rp) or []
            return {"rows": rows, "remarks": remarks}

        run_func_async(
            parent=self,
            func=_fetch,
            on_result=self._on_data_loaded,
            on_error=self._on_data_error,
            loading_message="⏳  Loading FT HO data…",
        )

    def _on_data_error(self, err):
        print(f"FT HO load error: {err}")

    def _on_data_loaded(self, data):
        rows = data["rows"]
        remark_map = {}
        for r in data["remarks"]:
            remark_map[r["branch"]] = r.get("remark", "") or ""

        self._building = True
        self.table.setRowCount(0)

        total_from = total_ho = total_to = 0.0

        for row in rows:
            idx = self.table.rowCount()
            self.table.insertRow(idx)

            branch = row["branch"]
            ft_from = float(row.get("ft_from_branch", 0) or 0)
            ft_ho = float(row.get("ft_to_ho", 0) or 0)
            ft_to = float(row.get("ft_to_branch", 0) or 0)

            total_from += ft_from
            total_ho += ft_ho
            total_to += ft_to

            # Branch (read-only)
            branch_item = QTableWidgetItem(branch)
            branch_item.setFlags(branch_item.flags() & ~Qt.ItemIsEditable)
            self.table.setItem(idx, 0, branch_item)

            # FT From Branch
            from_item = QTableWidgetItem(f"{ft_from:,.2f}")
            from_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            from_item.setFlags(from_item.flags() & ~Qt.ItemIsEditable)
            self.table.setItem(idx, 1, from_item)

            # FT To Head Office
            ho_item = QTableWidgetItem(f"{ft_ho:,.2f}")
            ho_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            ho_item.setFlags(ho_item.flags() & ~Qt.ItemIsEditable)
            self.table.setItem(idx, 2, ho_item)

            # FT To Branch
            to_item = QTableWidgetItem(f"{ft_to:,.2f}")
            to_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            to_item.setFlags(to_item.flags() & ~Qt.ItemIsEditable)
            self.table.setItem(idx, 3, to_item)

            # Remarks (editable)
            remark_item = QTableWidgetItem(remark_map.get(branch, ""))
            remark_item.setData(Qt.UserRole, "remark")
            self.table.setItem(idx, 4, remark_item)

        # Totals row
        total_idx = self.table.rowCount()
        self.table.insertRow(total_idx)
        bold = QFont()
        bold.setBold(True)
        bg = QBrush(QColor("#ecf0f1"))

        total_label = QTableWidgetItem("TOTAL")
        total_label.setFont(bold)
        total_label.setBackground(bg)
        total_label.setFlags(total_label.flags() & ~Qt.ItemIsEditable)
        total_label.setData(Qt.UserRole, "total")
        self.table.setItem(total_idx, 0, total_label)

        for col, val in [(1, total_from), (2, total_ho), (3, total_to)]:
            item = QTableWidgetItem(f"{val:,.2f}")
            item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            item.setFont(bold)
            item.setBackground(bg)
            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
            item.setData(Qt.UserRole, "total")
            self.table.setItem(total_idx, col, item)

        empty = QTableWidgetItem("")
        empty.setBackground(bg)
        empty.setFlags(empty.flags() & ~Qt.ItemIsEditable)
        empty.setData(Qt.UserRole, "total")
        self.table.setItem(total_idx, 4, empty)

        self._building = False

    # ── Remarks persistence ───────────────────────────────────────────────

    def _on_item_changed(self, item):
        if self._building:
            return
        if item.data(Qt.UserRole) != "remark":
            return

        row = item.row()
        branch_item = self.table.item(row, 0)
        if not branch_item:
            return

        branch = branch_item.text()
        remark = item.text().strip()
        date_start, _ = self.date_range_widget.get_date_range()

        try:
            if remark:
                db_manager.execute_query(
                    "INSERT INTO ft_ho_remarks (branch, report_date, remark) "
                    "VALUES (%s, %s, %s) "
                    "ON DUPLICATE KEY UPDATE remark = VALUES(remark)",
                    (branch, date_start, remark),
                )
            else:
                db_manager.execute_query(
                    "DELETE FROM ft_ho_remarks WHERE branch = %s AND report_date = %s",
                    (branch, date_start),
                )
        except Exception as e:
            print(f"Error saving remark: {e}")

    # ── Export ────────────────────────────────────────────────────────────

    def export_to_excel(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        except ImportError:
            QMessageBox.critical(self, "Error", "openpyxl is required for Excel export.")
            return

        if self.table.rowCount() == 0:
            QMessageBox.warning(self, "No Data", "No data to export.")
            return

        date_start, date_end = self.date_range_widget.get_date_range()
        is_range = self.date_range_widget.is_range_mode()
        group = self.group_selector.currentText()

        default_name = f"FT_HO_{group}_{date_start}"
        if is_range and date_start != date_end:
            default_name = f"FT_HO_{group}_{date_start}_to_{date_end}"

        path, _ = QFileDialog.getSaveFileName(
            self, "Save Excel File", default_name + ".xlsx",
            "Excel Files (*.xlsx)"
        )
        if not path:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "FT HO"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        total_fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")

        # Title
        title_text = f"FT Head Office — {group}"
        if is_range and date_start != date_end:
            title_text += f"  ({date_start} to {date_end})"
        else:
            title_text += f"  ({date_start})"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
        cell = ws.cell(row=1, column=1, value=title_text)
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal="center")

        # Headers
        headers = ["Branches", "FT From Branch", "FT To Head Office", "FT To Branch", "Remarks"]
        for col_idx, h in enumerate(headers, 1):
            c = ws.cell(row=3, column=col_idx, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center")
            c.border = thin_border

        # Data
        for row_idx in range(self.table.rowCount()):
            for col_idx in range(self.table.columnCount()):
                item = self.table.item(row_idx, col_idx)
                value = item.text() if item else ""
                c = ws.cell(row=row_idx + 4, column=col_idx + 1, value=value)
                c.border = thin_border
                if col_idx in (1, 2, 3):
                    c.alignment = Alignment(horizontal="right")
                is_total = (item and item.data(Qt.UserRole) == "total") if item else False
                if is_total:
                    c.font = Font(bold=True)
                    c.fill = total_fill

        # Column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 35

        try:
            wb.save(path)
            QMessageBox.information(self, "Exported", f"File saved to:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save file:\n{e}")
