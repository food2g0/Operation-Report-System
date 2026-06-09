from PyQt5.QtWidgets import (
    QWidget, QLabel, QVBoxLayout, QHBoxLayout, QGroupBox, QFormLayout,
    QLineEdit, QComboBox, QPushButton, QMessageBox, QDateEdit, QStackedWidget,
    QScrollArea, QFrame, QFileDialog, QDialog, QTableWidget, QTableWidgetItem, QHeaderView,
    QApplication, QSizePolicy, QCheckBox, QGridLayout
)
from PyQt5.QtGui import QFont
from PyQt5.QtCore import Qt, QDate, pyqtSignal, QTimer
from api_db_manager import db_manager
from security import SessionManager
import json
import logging
import os
import re

logger = logging.getLogger(__name__)
import sys

from constants import COLORS, FONT_SIZES, CONFIG, MESSAGES
from widgets_util import MoneyInput, LotesInput, DisplayField

from palawan_page import PalawanPage
from mc_page import MCPage
from fund_transfer import FundTransferPage
from payable_page import PayablesPage
from global_payable_page import GlobalPayablePage
from report_page import ReportPage
from admin_manage import create_corporation, create_branch, create_client, get_all_supervisors
from variance_review_page import VarianceReviewPage
from user_management_page import UserManagementPage
from daily_transaction_page import DailyTransactionPage
from new_sanla_page import NewSanlaPage
from new_renew_page import NewRenewPage
from global_other_services_page import GlobalOtherServicesPage
from ft_ho_page import FTHOPage
from depo_br_page import DepoBRPage
from review_summary_page import ReviewSummaryPage
from connection_watcher import ConnectionWatcher, ConnectionBanner


try:
    from auto_updater import check_for_updates, check_update_success
    from version import __version__
    AUTO_UPDATE_ENABLED = True
except ImportError:
    AUTO_UPDATE_ENABLED = False
    __version__ = "1.0.0"
    check_update_success = None


class AdminDashboard(QWidget):
    logout_requested = pyqtSignal()
    
    def __init__(self, account_type=2, os_group=""):
        super().__init__()

        self.account_type = account_type
        self.os_group = os_group or ""   # assigned group; "" = no restriction
        brand_label = "Brand A" if account_type == 1 else "Brand B"
        group_label = f" — {self.os_group}" if self.os_group else ""
        self.setWindowTitle(f"Admin Dashboard ({brand_label}){group_label} - Cash Management System")
        self.db = db_manager
        self._update_checker_threads = []  
        
        # Zoom functionality
        self.zoom_level = 100
        self.setFocusPolicy(Qt.StrongFocus)

        self.session = SessionManager(inactivity_timeout=1800)
        self._session_timer = QTimer(self)
        self._session_timer.timeout.connect(self._check_session_timeout)
        self._session_timer.start(60000)

        # Network connectivity monitor
        self._is_connected = True
        self._conn_watcher = ConnectionWatcher(self.db)
        self._conn_watcher.connection_lost.connect(self._on_connection_lost)
        self._conn_watcher.connection_restored.connect(self._on_connection_restored)
        self._conn_watcher.start()

        self.debit_inputs = {}
        self.credit_inputs = {}
        self.debit_lotes_inputs = {}
        self.credit_lotes_inputs = {}
        self.selected_bank_account = None  
        self.bank_account_btn = None 
        self.selected_branch_dest = None
        self.branch_dest_btn = None
        self.selected_from_branch_dest = None 
        self.from_branch_dest_btn = None

      
        brand_key = "Brand A" if account_type == 1 else "Brand B"
        field_config = self._load_field_config()
        
        if field_config and brand_key in field_config:
            brand_config = field_config[brand_key]

            self.debit_fields = {item[0]: item[2] for item in brand_config.get("debit", [])}
            self.credit_fields = {item[0]: item[2] for item in brand_config.get("credit", [])}
        else:
  
            self.debit_fields = {
                "Rescate Jewelry": "rescate_jewelry",
                "Interest": "interest",
                "Penalty": "penalty",
                "Stamp": "stamp",
                "Resguardo/Affidavit": "resguardo_affidavit",
                "HABOL Renew/Tubos": "habol_renew_tubos",
                "Habol R/T Interest&Stamp": "habol_rt_interest_stamp",
                "Jew. A.I": "jew_ai",
                "S.C": "sc",
                "Fund Transfer": "fund_transfer_from_branch",
                "Sendah Load + SC": "sendah_load_sc",
                "PPAY CO SC": "ppay_co_sc",
                "Palawan Send Out": "palawan_send_out",
                "Palawan S.C": "palawan_sc",
                "Palawan Suki Card": "palawan_suki_card",
                "Palawan Pay Cash-In + SC": "palawan_pay_cash_in_sc",
                "Palawan Pay Bills + SC": "palawan_pay_bills_sc",
                "Palawan Load": "palawan_load",
                "Palawan Change Receiver": "palawan_change_receiver",
                "MC In": "mc_in",
                "Handling fee": "handling_fee",
                "Other Penalty": "other_penalty",
                "Cash Overage": "cash_overage"
            }
            self.credit_fields = {
                "Empeno JEW. (NEW)": "empeno_jew_new",
                "Empeno JEW (RENEW)": "empeno_jew_renew",
                "Empeno Motor/Car": "empeno_motor_car",
                "Fund Transfer to HEAD OFFICE": "fund_transfer_to_head_office",
                "Fund Transfer to BRANCH": "fund_transfer_to_branch",
                "Palawan Pay Out": "palawan_pay_out",
                "Palawan Pay Out (incentives)": "palawan_pay_out_incentives",
                "Palawan Pay Cash Out": "palawan_pay_cash_out",
                "MC Out": "mc_out",
                "PC-Salary": "pc_salary",
                "PC-Rental": "pc_rental",
                "PC-Electric": "pc_electric",
                "PC-Water": "pc_water",
                "PC-Internet": "pc_internet",
                "PC-Lbc/Jrs/Jnt": "pc_lbc_jrs_jnt",
                "PC-Permits/BIR Payments": "pc_permits_bir_payments",
                "PC-Supplies/Xerox/Maintenance": "pc_supplies_xerox_maintenance",
                "PC-Transpo": "pc_transpo",
                "Palawan Cancel": "palawan_cancel",
                "Palawan Suki Discounts": "palawan_suki_discounts",
                "Palawan Suki Rebates": "palawan_suki_rebates",
                "OTHERS": "others",
                "Cash Shortage": "cash_shortage"
            }


        self.daily_table = "daily_reports_brand_a" if account_type == 1 else "daily_reports"

        self.setup_styles()
        self.build_ui()


        self._capture_base_fonts()

        QTimer.singleShot(0,   self.load_corporations)     # populates corp/group selectors
        QTimer.singleShot(200, self._ensure_review_table)  # DDL check; not time-critical
        
        # Install event filter on application for zoom
        QApplication.instance().installEventFilter(self)

        if AUTO_UPDATE_ENABLED and check_update_success:
            check_update_success(parent=self)
    
    def _capture_base_fonts(self):

        for w in self._get_zoom_target_widgets():
            font = w.font()
            point_size = font.pointSize()
            pixel_size = font.pixelSize()

            base_stylesheet = w.styleSheet() or ""
            w.setProperty('_base_stylesheet', base_stylesheet)
            w.setProperty('_base_zoom_height', max(w.minimumHeight(), w.sizeHint().height()))
            

            if point_size < 0 and pixel_size < 0:
                # Get from application default
                app_font = QApplication.font()
                point_size = app_font.pointSize()
                if point_size < 0:
                    point_size = 10  # Default fallback
            
            # Store whichever is valid
            if point_size > 0:
                w.setProperty('_base_point_size', point_size)
            elif pixel_size > 0:
                w.setProperty('_base_pixel_size', pixel_size)

    def _get_zoom_target_widgets(self):
        """Return only debit/credit amount inputs that should respond to zoom."""
        targets = []
        seen = set()
        for inp in list(self.debit_inputs.values()) + list(self.credit_inputs.values()):
            if inp is None:
                continue
            wid = id(inp)
            if wid in seen:
                continue
            seen.add(wid)
            targets.append(inp)
        return targets

    def _scale_stylesheet_font_sizes(self, stylesheet, zoom_factor):
        """Scale font-size declarations in a stylesheet by zoom factor."""
        if not stylesheet or "font-size" not in stylesheet:
            return stylesheet

        def _replace(match):
            base_size = float(match.group(1))
            scaled = max(1, min(500, int(round(base_size * zoom_factor))))
            return f"font-size: {scaled}px"

        return re.sub(
            r"font-size\s*:\s*([0-9]*\.?[0-9]+)\s*px",
            _replace,
            stylesheet,
            flags=re.IGNORECASE,
        )

    def _apply_zoom_to_all(self):
        """Apply zoom only to debit/credit amount input widgets."""
        zoom_factor = self.zoom_level / 100.0
        for w in self._get_zoom_target_widgets():
            font = w.font()
            new_size = None
            
            # Try point size first
            base_point = w.property('_base_point_size')
            if base_point is not None:
                try:
                    base_point = int(base_point)
                    if base_point > 0:
                        new_size = max(1, min(500, int(base_point * zoom_factor)))
                        font.setPointSize(new_size)
                        w.setFont(font)
                except (ValueError, TypeError, OverflowError):
                    pass
            
            # Try pixel size
            base_pixel = w.property('_base_pixel_size')
            if base_pixel is not None:
                try:
                    base_pixel = int(base_pixel)
                    if base_pixel > 0:
                        new_size = max(1, min(500, int(base_pixel * zoom_factor)))
                        font.setPixelSize(new_size)
                        w.setFont(font)
                except (ValueError, TypeError, OverflowError):
                    pass

            base_stylesheet = w.property('_base_stylesheet')
            if isinstance(base_stylesheet, str):
                scaled_stylesheet = self._scale_stylesheet_font_sizes(base_stylesheet, zoom_factor)
                if new_size is not None and "font-size" not in scaled_stylesheet.lower():
                    scaled_stylesheet = (scaled_stylesheet + f"\nfont-size: {new_size}px;").strip()
                if scaled_stylesheet != w.styleSheet():
                    w.setStyleSheet(scaled_stylesheet)

            base_height = w.property('_base_zoom_height')
            if base_height is not None:
                try:
                    scaled_height = max(20, min(500, int(int(base_height) * zoom_factor)))
                    w.setMinimumHeight(scaled_height)
                    w.updateGeometry()
                except (ValueError, TypeError, OverflowError):
                    pass

        self.update()

    def _ensure_review_table(self):

        try:
            self.db.execute_query("""
                CREATE TABLE IF NOT EXISTS admin_review_marks (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    brand VARCHAR(10) NOT NULL,
                    branch VARCHAR(255) NOT NULL,
                    report_date DATE NOT NULL,
                    reviewed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE KEY uq_review (brand, branch, report_date)
                ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
            """)
        except Exception as e:
            logger.error("[AdminDashboard] review table create: %s", e)

        # Drop legacy payable_brand_a table — replaced by payable_tbl_brand_a
        try:
            self.db.execute_query("DROP TABLE IF EXISTS payable_brand_a")
        except Exception as e:
            logger.error("[AdminDashboard] payable_brand_a table drop: %s", e)

        # Add any columns to daily_reports_brand_a that may be missing on older DB installs
        _migrations = [
            ("daily_reports_brand_a", "pc_inc_insurance",              "DECIMAL(15,2) DEFAULT 0.00"),
            ("daily_reports_brand_a", "pc_inc_insurance_lotes",        "INT DEFAULT 0"),
            ("daily_reports_brand_a", "habol_renew_tubos",             "DECIMAL(15,2) DEFAULT 0.00"),
            ("daily_reports_brand_a", "habol_renew_tubos_lotes",       "INT DEFAULT 0"),
            ("daily_reports_brand_a", "habol_rt_interest_stamp",       "DECIMAL(15,2) DEFAULT 0.00"),
            ("daily_reports_brand_a", "habol_rt_interest_stamp_lotes", "INT DEFAULT 0"),
            ("daily_reports_brand_a", "transfast",                     "DECIMAL(15,2) DEFAULT 0.00"),
            ("daily_reports_brand_a", "transfast_lotes",               "INT DEFAULT 0"),
            # lotes columns for payable_tbl_brand_a (added for PalawanPayableContainer)
            ("payable_tbl_brand_a",   "sendout_lotes",                 "INT DEFAULT 0"),
            ("payable_tbl_brand_a",   "payout_lotes",                  "INT DEFAULT 0"),
            ("payable_tbl_brand_a",   "international_lotes",           "INT DEFAULT 0"),
        ]
        for table, col, typedef in _migrations:
            try:
                exists = self.db.execute_query(
                    "SELECT COUNT(*) as cnt FROM information_schema.COLUMNS "
                    "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = %s AND COLUMN_NAME = %s",
                    (table, col)
                )
                if not exists or exists[0].get('cnt', 0) == 0:
                    self.db.execute_query(
                        f"ALTER TABLE {table} ADD COLUMN {col} {typedef}"
                    )
            except Exception:
                pass  # Safe to ignore

    def _load_field_config(self):

        try:
            result = self.db.execute_query(
                "SELECT config_value FROM field_config WHERE config_key = 'field_definitions'"
            )
            if result and result[0].get('config_value'):
                cfg = json.loads(result[0]['config_value'])
                for brand in ("Brand A", "Brand B"):
                    cfg.setdefault(brand, {})
                    cfg[brand].setdefault("debit", [])
                    cfg[brand].setdefault("credit", [])
                return cfg
        except Exception as e:
            logger.error("[AdminDashboard] Failed to load config from DB: %s", e)

        try:
            if getattr(sys, 'frozen', False):
                config_dir = os.path.dirname(sys.executable)
            else:
                config_dir = os.path.dirname(os.path.abspath(__file__))
            
            config_path = os.path.join(config_dir, 'field_config.json')
            
            if os.path.exists(config_path):
                with open(config_path, 'r', encoding='utf-8') as f:
                    return json.load(f)
            else:
                logger.warning("field_config.json not found at %s", config_path)
                return None
        except Exception as e:
            logger.error("Error loading field_config.json: %s", e)
            return None

    def setup_styles(self):

        self.setStyleSheet("""

            QWidget {
                background-color: #f5f6fa;
                font-family: 'Segoe UI', Arial, sans-serif;
                font-size: 12px;
            }

            /* Navigation Buttons */
            QPushButton {
                background-color: #3498db;
                color: black;
                border: none;
                padding: 8px 16px;
                border-radius: 6px;
                font-weight: bold;
                font-size: 12px;
                min-height: 30px;
            }

            QPushButton:hover {
                background-color: #2980b9;
            }

            QPushButton:pressed {
                background-color: #21618c;
            }

            QPushButton#saveButton {
                background-color: #27ae60;
                font-size: 13px;
                padding: 10px 20px;
                min-width: 120px;
            }

            QPushButton#saveButton:hover {
                background-color: #219a52;
            }

            QPushButton#loadButton {
                background-color: #f39c12;
                font-size: 11px;
                padding: 6px 12px;
            }

            QPushButton#loadButton:hover {
                background-color: #e67e22;
            }

            /* Group Boxes */
            QGroupBox {
                font-weight: bold;
                font-size: 13px;
                color: #2c3e50;
                border: 2px solid #bdc3c7;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 15px;
                background-color: white;
            }

            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 8px 0 8px;
                color: #2c3e50;
                font-weight: bold;
                font-size: 14px;
                background-color: white;
            }

            /* Labels */
            QLabel {
                color: #2c3e50;
                font-size: 11px;
                font-weight: bold;
            }

            QLabel[class="header"] {
                font-weight: bold;
                font-size: 13px;
                color: #2c3e50;
            }

            QLabel[class="section"] {
                font-weight: bold;
                font-size: 14px;
                color: #34495e;
                background-color: #ecf0f1;
                padding: 5px;
                border-radius: 4px;
            }

            /* Input Fields */
            QLineEdit {
                border: 1px solid #bdc3c7;
                border-radius: 4px;
                padding: 5px 8px;
                font-size: 11px;
                background-color: white;
                min-height: 20px;
            }

            QLineEdit:focus {
                border: 2px solid #3498db;
            }

            QLineEdit[readOnly="true"] {
                background-color: #f8f9fa;
                border: 1px solid #dee2e6;
                font-weight: bold;
                color: #495057;
            }

            QLineEdit[class="money"] {
                font-weight: bold;
                text-align: right;
                color: #27ae60;
                font-size: 12px;
            }

            QLineEdit[class="result"] {
                background-color: #fff3cd;
                border: 2px solid #ffeaa7;
                font-weight: bold;
                font-size: 12px;
                color: #856404;
            }

            /* Dropdowns */
            QComboBox {
                border: 1px solid #bdc3c7;
                border-radius: 4px;
                padding: 5px 8px;
                background-color: white;
                font-size: 11px;
                min-width: 120px;
                min-height: 25px;
            }

            QComboBox:focus {
                border: 2px solid #3498db;
            }

            QComboBox::drop-down {
                border: none;
                width: 20px;
            }

            QComboBox::down-arrow {
                width: 12px;
                height: 12px;
            }

            /* Date Edit */
            QDateEdit {
                border: 1px solid #bdc3c7;
                border-radius: 4px;
                padding: 5px 28px 5px 8px;
                background-color: white;
                font-size: 11px;
                min-height: 25px;
                min-width: 130px;
            }

            QDateEdit:focus {
                border: 2px solid #3498db;
            }

            QDateEdit::drop-down {
                subcontrol-origin: border;
                subcontrol-position: center right;
                width: 28px;
                border-left: 1px solid #bdc3c7;
                background-color: #ecf0f1;
                border-top-right-radius: 4px;
                border-bottom-right-radius: 4px;
            }

            QDateEdit::drop-down:hover {
                background-color: #d5dbdb;
            }

            QDateEdit::down-arrow {
                width: 10px;
                height: 10px;
            }

            /* Calendar Popup — Standardized */
            QCalendarWidget {
                min-width: 340px;
                min-height: 280px;
                background: white;
                border: 1px solid #dee2e6;
                border-radius: 6px;
            }

            QCalendarWidget QWidget#qt_calendar_navigationbar {
                background-color: #343a40;
                min-height: 42px;
                padding: 4px 6px;
                border-radius: 4px 4px 0 0;
            }

            QCalendarWidget QToolButton {
                color: #ecf0f1;
                font-size: 14px;
                font-weight: bold;
                background-color: transparent;
                padding: 6px 10px;
                border-radius: 4px;
                margin: 2px;
            }

            QCalendarWidget QToolButton:hover {
                background-color: #007bff;
                color: white;
            }

            QCalendarWidget QToolButton:pressed {
                background-color: #0056b3;
                color: white;
            }

            QCalendarWidget QSpinBox {
                color: #2c3e50;
                background-color: #ecf0f1;
                font-size: 13px;
                font-weight: bold;
                border: 1px solid #bdc3c7;
                border-radius: 4px;
                padding: 4px 8px;
                selection-background-color: #007bff;
                selection-color: white;
            }

            QCalendarWidget QAbstractItemView {
                background: white;
                selection-background-color: #007bff;
                selection-color: white;
                font-size: 12px;
                alternate-background-color: #f8f9fa;
            }

            QCalendarWidget QAbstractItemView::item {
                padding: 6px;
                border-radius: 4px;
            }

            QCalendarWidget QAbstractItemView::item:alternate {
                background-color: #f8f9fa;
            }

            QCalendarWidget QAbstractItemView::item:selected {
                background-color: #007bff;
                color: white;
                font-weight: bold;
            }

            QCalendarWidget QAbstractItemView:disabled {
                color: #bdc3c7;
            }

            /* Scroll Area */
            QScrollArea {
                border: none;
                background-color: transparent;
            }

            /* Navigation Bar */
            QFrame#navBar {
                background-color: #2c3e50;
                border-radius: 0px;
                margin-bottom: 0px;
                padding: 0px;
            }

            QFrame#navBar QPushButton {
                background-color: transparent;
                border: none;
                border-bottom: 3px solid transparent;
                color: #bdc3c7;
                font-size: 11px;
                font-weight: bold;
                padding: 10px 6px;
                border-radius: 0px;
            }

            QFrame#navBar QPushButton:hover {
                background-color: rgba(52, 152, 219, 0.15);
                color: #ecf0f1;
                border-bottom: 3px solid #3498db;
            }

            QFrame#navBar QPushButton:checked {
                background-color: rgba(52, 152, 219, 0.25);
                color: #ffffff;
                border-bottom: 3px solid #3498db;
            }
        """)
        # Inject calendar PNG icon into the QDateEdit dropdown button.
        # Appending overrides the earlier ::down-arrow rule via CSS cascade.
        _cal = os.path.join(
            os.path.dirname(os.path.abspath(__file__)), 'assets', 'calendar.png'
        ).replace('\\', '/')
        self.setStyleSheet(self.styleSheet() + f"""
            QDateEdit::down-arrow {{
                image: url({_cal});
                width: 14px;
                height: 14px;
            }}
        """)

    def build_ui(self):
        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(10, 8, 10, 8)
        main_layout.setSpacing(6)

        # Connectivity banner (hidden until connection drops)
        self._conn_banner = ConnectionBanner()
        main_layout.addWidget(self._conn_banner)

        header_frame = QFrame()
        header_frame.setStyleSheet("""
            QFrame {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 #3498db, stop:1 #2980b9);
                border-radius: 8px;
                padding: 8px;
                margin-bottom: 4px;
            }
        """)
        header_layout = QHBoxLayout(header_frame)
        header_layout.setContentsMargins(10, 4, 10, 4)

        title_label = QLabel("Admin Dashboard")
        title_label.setStyleSheet("font-size: 20px; font-weight: bold; color: white;")
        
        if AUTO_UPDATE_ENABLED:
            update_button = QPushButton(f"ℹ️ v{__version__}")
            update_button.setStyleSheet("""
                QPushButton {
                    background-color: #16a085;
                    color: white;
                    border: none;
                    padding: 10px 20px;
                    border-radius: 6px;
                    font-weight: bold;
                    font-size: 11px;
                    min-width: 80px;
                    max-height: 40px;
                }
                QPushButton:hover {
                    background-color: #1abc9c;
                }
                QPushButton:pressed {
                    background-color: #138d75;
                }
            """)
            update_button.clicked.connect(self.check_for_updates)
            update_button.setToolTip("Check for updates")
        
        logout_button = QPushButton("Logout")
        logout_button.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                border: none;
                padding: 10px 20px;
                border-radius: 6px;
                font-weight: bold;
                font-size: 12px;
                min-width: 100px;
                max-height: 40px;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
            QPushButton:pressed {
                background-color: #a93226;
            }
        """)
        logout_button.clicked.connect(self.handle_logout)
        
        ver_lbl = QLabel(f"v{__version__}")
        ver_lbl.setStyleSheet("""
            QLabel {
                color: rgba(255,255,255,0.6);
                font-size: 11px;
                font-weight: 600;
                padding: 0 8px;
                background: transparent;
            }
        """)
        ver_lbl.setToolTip(f"Operation Report System v{__version__}")

        header_layout.addWidget(title_label)
        header_layout.addStretch()
        header_layout.addWidget(ver_lbl)
        if AUTO_UPDATE_ENABLED:
            header_layout.addWidget(update_button)
        header_layout.addWidget(logout_button)
        
        main_layout.addWidget(header_frame)

        nav_frame = QFrame()
        nav_frame.setObjectName("navBar")
        nav_layout = QHBoxLayout(nav_frame)
        nav_layout.setContentsMargins(0, 0, 0, 0)
        nav_layout.setSpacing(0)

        self.daily_btn = QPushButton("Daily Cash Count")
        self.variance_btn = QPushButton("Variance Review")
        self.palawan_btn = QPushButton("Palawan")
        self.mc_btn = QPushButton("MC")
        self.fund_btn = QPushButton("Fund Transfer")
        self.payable_btn = QPushButton("Payable")
        self.global_payable_btn = QPushButton("Global Payable")
        self.report_btn = QPushButton("Payable Reports")
        self.daily_txn_btn = QPushButton("Daily Transaction")
        self.new_sanla_btn = QPushButton("New Sanla")
        self.new_renew_btn = QPushButton("New & Renew")
        self.global_os_btn = QPushButton("Global Other Services")
        self.ft_ho_btn = QPushButton("FT HO")
        self.depo_br_btn = QPushButton("DEPO BR")
        self.admin_btn = QPushButton("User Management")
        self.review_summary_btn = QPushButton("Review Summary")


        for btn in [self.daily_btn, self.variance_btn, self.palawan_btn, self.mc_btn,
                    self.fund_btn, self.payable_btn, self.global_payable_btn, self.report_btn, self.daily_txn_btn,
                    self.new_sanla_btn, self.new_renew_btn, self.global_os_btn, self.ft_ho_btn, self.depo_br_btn,
                    self.admin_btn, self.review_summary_btn]:
            btn.setCheckable(True)
        self.daily_btn.setChecked(True) 

        if self.account_type == 1:
  
            self.nav_buttons = [
                self.daily_btn, self.variance_btn, self.palawan_btn, self.mc_btn,
                self.fund_btn, self.payable_btn, self.daily_txn_btn, self.new_sanla_btn,
                self.new_renew_btn, self.global_os_btn, self.ft_ho_btn, self.depo_br_btn,
                self.review_summary_btn, self.admin_btn
            ]
        else:
     
            self.nav_buttons = [
                self.daily_btn, self.variance_btn, self.palawan_btn, self.mc_btn,
                self.fund_btn, self.payable_btn, self.global_payable_btn, self.report_btn, self.review_summary_btn, self.admin_btn
            ]

        for btn in self.nav_buttons:
            btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
            nav_layout.addWidget(btn)

        main_layout.addWidget(nav_frame)


        self.stack = QStackedWidget()

        self._lazy_factories = {} 
        self.daily_cash_widget = self.build_daily_cash_widget()

        if self.account_type == 1:

            self.stack.addWidget(self.daily_cash_widget)    
            self._add_lazy(1, lambda: VarianceReviewPage(account_type=self.account_type), 'variance_widget')
            self._add_lazy(2, lambda: PalawanPage(account_type=self.account_type), 'palawan_widget')
            self._add_lazy(3, lambda: MCPage(account_type=self.account_type), 'mc_widget')
            self._add_lazy(4, lambda: FundTransferPage(account_type=self.account_type), 'fund_widget')
            self._add_lazy(5, lambda: PayablesPage(account_type=self.account_type), 'payable_widget')
            self._add_lazy(6, lambda: DailyTransactionPage(), 'daily_txn_widget')
            self._add_lazy(7, lambda: NewSanlaPage(), 'new_sanla_widget')
            self._add_lazy(8, lambda: NewRenewPage(), 'new_renew_widget')
            self._add_lazy(9, lambda: GlobalOtherServicesPage(), 'global_os_widget')
            self._add_lazy(10, lambda: FTHOPage(account_type=self.account_type), 'ft_ho_widget')
            self._add_lazy(11, lambda: DepoBRPage(account_type=self.account_type), 'depo_br_widget')
            self._add_lazy(12, lambda: ReviewSummaryPage(account_type=self.account_type), 'review_summary_widget')
            self._add_lazy(13, lambda: UserManagementPage(), 'admin_widget')
        else:
      
            self.stack.addWidget(self.daily_cash_widget)         
            self._add_lazy(1, lambda: VarianceReviewPage(account_type=self.account_type), 'variance_widget')
            self._add_lazy(2, lambda: PalawanPage(account_type=self.account_type), 'palawan_widget')
            self._add_lazy(3, lambda: MCPage(account_type=self.account_type), 'mc_widget')
            self._add_lazy(4, lambda: FundTransferPage(account_type=self.account_type), 'fund_widget')
            self._add_lazy(5, lambda: PayablesPage(account_type=self.account_type), 'payable_widget')
            self._add_lazy(6, lambda: GlobalPayablePage(account_type=self.account_type), 'global_payable_widget')
            self._add_lazy(7, lambda: ReportPage(), 'report_widget')
            self._add_lazy(8, lambda: ReviewSummaryPage(account_type=self.account_type), 'review_summary_widget')
            self._add_lazy(9, lambda: UserManagementPage(), 'admin_widget')
        
        main_layout.addWidget(self.stack)


        if self.account_type == 1:
         
            self.daily_btn.clicked.connect(lambda: self.switch_view(0, self.daily_btn))
            self.variance_btn.clicked.connect(lambda: self.switch_view(1, self.variance_btn))
            self.palawan_btn.clicked.connect(lambda: self.switch_view(2, self.palawan_btn))
            self.mc_btn.clicked.connect(lambda: self.switch_view(3, self.mc_btn))
            self.fund_btn.clicked.connect(lambda: self.switch_view(4, self.fund_btn))
            self.payable_btn.clicked.connect(lambda: self.switch_view(5, self.payable_btn))
            self.daily_txn_btn.clicked.connect(lambda: self.switch_view(6, self.daily_txn_btn))
            self.new_sanla_btn.clicked.connect(lambda: self.switch_view(7, self.new_sanla_btn))
            self.new_renew_btn.clicked.connect(lambda: self.switch_view(8, self.new_renew_btn))
            self.global_os_btn.clicked.connect(lambda: self.switch_view(9, self.global_os_btn))
            self.ft_ho_btn.clicked.connect(lambda: self.switch_view(10, self.ft_ho_btn))
            self.depo_br_btn.clicked.connect(lambda: self.switch_view(11, self.depo_br_btn))
            self.review_summary_btn.clicked.connect(lambda: self.switch_view(12, self.review_summary_btn))
            self.admin_btn.clicked.connect(lambda: self.switch_view(13, self.admin_btn))
        else:

            self.daily_btn.clicked.connect(lambda: self.switch_view(0, self.daily_btn))
            self.variance_btn.clicked.connect(lambda: self.switch_view(1, self.variance_btn))
            self.palawan_btn.clicked.connect(lambda: self.switch_view(2, self.palawan_btn))
            self.mc_btn.clicked.connect(lambda: self.switch_view(3, self.mc_btn))
            self.fund_btn.clicked.connect(lambda: self.switch_view(4, self.fund_btn))
            self.payable_btn.clicked.connect(lambda: self.switch_view(5, self.payable_btn))
            self.global_payable_btn.clicked.connect(lambda: self.switch_view(6, self.global_payable_btn))
            self.report_btn.clicked.connect(lambda: self.switch_view(7, self.report_btn))
            self.review_summary_btn.clicked.connect(lambda: self.switch_view(8, self.review_summary_btn))
            self.admin_btn.clicked.connect(lambda: self.switch_view(9, self.admin_btn))

        self.setLayout(main_layout)

    def handle_logout(self):
    
        reply = QMessageBox.question(
            self,
            "Confirm Logout",
            "Are you sure you want to logout?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            self.session.logout()
            self.logout_requested.emit()
            self.close()
    
    def _on_connection_lost(self):
        self._is_connected = False
        if hasattr(self, '_conn_banner'):
            self._conn_banner.show_banner()
        self._session_timer.stop()

    def _on_connection_restored(self):
        self._is_connected = True
        if hasattr(self, '_conn_banner'):
            self._conn_banner.hide_banner()
        self._session_timer.start(60000)

    def _check_session_timeout(self):

        if self.session.check_timeout():
            self._session_timer.stop()
            QMessageBox.warning(
                self,
                "Session Expired",
                "Your session has expired due to inactivity.\nPlease log in again.",
                QMessageBox.Ok
            )
            self.logout_requested.emit()
            self.close()
    
    def mousePressEvent(self, event):
  
        self.session.update_activity()
        super().mousePressEvent(event)
    
    def eventFilter(self, obj, event):
        """Handle application-level events for zoom"""
        from PyQt5.QtCore import QEvent
        # Capture wheel events at app level
        if event.type() == QEvent.Wheel:
            # Check if Ctrl is pressed
            if event.modifiers() & Qt.ControlModifier:
                delta = event.angleDelta().y()
                if delta > 0:
                    self.zoom_in()
                else:
                    self.zoom_out()
                return True  # Consume the event
        # Check for Ctrl+0 key press
        elif event.type() == QEvent.KeyPress:
            if event.modifiers() & Qt.ControlModifier and event.key() == Qt.Key_0:
                self.reset_zoom()
                return True  # Consume the event
        return super().eventFilter(obj, event)

    def wheelEvent(self, event):
        """Handle Ctrl + mouse wheel for zoom"""
        if event.modifiers() & Qt.ControlModifier:
            # Zoom in (scroll up) or zoom out (scroll down)
            delta = event.angleDelta().y()
            if delta > 0:
                self.zoom_in()
            else:
                self.zoom_out()
            event.accept()
        else:
            super().wheelEvent(event)
    
    def keyPressEvent(self, event):
        # Handle Ctrl+0 to reset zoom
        if event.modifiers() & Qt.ControlModifier and event.key() == Qt.Key_0:
            self.reset_zoom()
            event.accept()
        else:
            self.session.update_activity()
            super().keyPressEvent(event)
    
    def zoom_in(self):
        """Increase zoom level by 20%"""
        self.set_zoom_level(self.zoom_level + 20)
    
    def zoom_out(self):
        """Decrease zoom level by 20%"""
        self.set_zoom_level(self.zoom_level - 20)
    
    def reset_zoom(self):
        """Reset zoom to 100%"""
        self.set_zoom_level(100)
    
    def set_zoom_level(self, level):
        """Set zoom level and apply to all widgets"""
        # Clamp zoom level between 50% and 200%
        level = max(50, min(200, level))
        if level == self.zoom_level:
            return  # No change
        self.zoom_level = level
        self._apply_zoom_to_all()
    
    def check_for_updates(self):
       
        if AUTO_UPDATE_ENABLED:
            check_for_updates(parent=self, silent=False)
        else:
            QMessageBox.information(
                self,
                "Auto-Updater",
                "Auto-updater is not enabled.\n\n"
                "To enable it, install required dependencies:\n"
                "pip install requests packaging"
            )
    
    def closeEvent(self, event):

        if hasattr(self, '_update_checker_threads'):
            for thread in self._update_checker_threads[:]:
                if thread.isRunning():
                    thread.quit()
                    thread.wait(2000) 
        event.accept()

    def _add_lazy(self, index, factory, attr_name):

        self._lazy_factories[index] = (factory, attr_name)
        self.stack.addWidget(QWidget())

    def switch_view(self, index, active_button):

        if index in self._lazy_factories:
            factory, attr_name = self._lazy_factories.pop(index)
            QApplication.setOverrideCursor(Qt.WaitCursor)
            try:
                widget = factory()
                setattr(self, attr_name, widget)
                old = self.stack.widget(index)
                self.stack.removeWidget(old)
                old.deleteLater()
                self.stack.insertWidget(index, widget)
            finally:
                QApplication.restoreOverrideCursor()

        self.stack.setCurrentIndex(index)

        for btn in self.nav_buttons:
            if btn:
                btn.setChecked(False)

        if active_button:
            active_button.setChecked(True)

    def build_daily_cash_widget(self):

        main_widget = QWidget()
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setWidget(main_widget)

        layout = QVBoxLayout(main_widget)
        layout.setSpacing(15)


        header_frame = QFrame()
        header_frame.setStyleSheet("background-color: white; border-radius: 8px; padding: 10px;")
        header_layout = QVBoxLayout(header_frame)


        filter_type_layout = QHBoxLayout()
        filter_type_label = QLabel("Filter By:")
        filter_type_label.setProperty("class", "header")
        self.filter_type_selector = QComboBox()
        self.filter_type_selector.addItem("Corporation", "corporation")
        self.filter_type_selector.addItem("Group", "group")
        self.filter_type_selector.currentIndexChanged.connect(self.on_filter_type_changed)
        filter_type_layout.addWidget(filter_type_label)
        filter_type_layout.addWidget(self.filter_type_selector)
        filter_type_layout.addStretch()
        header_layout.addLayout(filter_type_layout)

        selection_layout = QHBoxLayout()
        selection_layout.setSpacing(15)

        self.corp_label = QLabel("Corporation:")
        self.corp_label.setProperty("class", "header")
        self.corp_selector = QComboBox()
        self.corp_selector.currentTextChanged.connect(self.load_branches)

        self.os_label = QLabel("Group:")
        self.os_label.setProperty("class", "header")
        self.os_selector = QComboBox()
        self.os_selector.currentTextChanged.connect(self.load_branches_by_os)
        self.os_label.setVisible(False)
        self.os_selector.setVisible(False)

        branch_label = QLabel("Branch:")
        branch_label.setProperty("class", "header")
        self.branch_selector = QComboBox()

        date_label = QLabel("Date:")
        date_label.setProperty("class", "header")
        self.date_picker = QDateEdit()
        self.date_picker.setDisplayFormat("dd MMM yyyy")
        self.date_picker.setCalendarPopup(True)
        self.date_picker.setDate(QDate.currentDate())
        # Apply standardized calendar styling to match other tabs
        _cal_style = (
            "QDateEdit{border:1px solid #bdc3c7;border-radius:4px;padding:5px 28px 5px 8px;"
            "background-color:white;font-size:11px;min-height:25px;min-width:130px;}"
            "QDateEdit:focus{border:2px solid #3498db;}"
            "QDateEdit::drop-down{subcontrol-origin:border;subcontrol-position:center right;"
            "width:28px;border-left:1px solid #bdc3c7;background-color:#ecf0f1;border-top-right-radius:4px;border-bottom-right-radius:4px;}"
            "QDateEdit::drop-down:hover{background-color:#d5dbdb;}"
            "QDateEdit::down-arrow{width:10px;height:10px;}"
            "QCalendarWidget{min-width:340px;min-height:280px;background:white;border:1px solid #dee2e6;border-radius:6px;}"
            "QCalendarWidget QWidget#qt_calendar_navigationbar{background-color:#343a40;min-height:42px;padding:4px 6px;border-radius:4px 4px 0 0;}"
            "QCalendarWidget QToolButton{color:#ecf0f1;font-size:14px;font-weight:bold;background-color:transparent;padding:6px 10px;border-radius:4px;margin:2px;}"
            "QCalendarWidget QToolButton:hover{background-color:#007bff;color:white;}"
            "QCalendarWidget QToolButton:pressed{background-color:#0056b3;color:white;}"
            "QCalendarWidget QSpinBox{color:#2c3e50;background-color:#ecf0f1;font-size:13px;font-weight:bold;border:1px solid #bdc3c7;border-radius:4px;padding:4px 8px;selection-background-color:#007bff;selection-color:white;}"
            "QCalendarWidget QAbstractItemView{background:white;selection-background-color:#007bff;selection-color:white;font-size:12px;alternate-background-color:#f8f9fa;}"
            "QCalendarWidget QAbstractItemView::item{padding:6px;border-radius:4px;}"
            "QCalendarWidget QAbstractItemView::item:alternate{background-color:#f8f9fa;}"
            "QCalendarWidget QAbstractItemView::item:selected{background-color:#007bff;color:white;font-weight:bold;}"
        )
        self.date_picker.setStyleSheet(_cal_style)

        self.load_button = QPushButton("Load Entry")
        self.load_button.setObjectName("loadButton")
        self.load_button.clicked.connect(self.load_entry_by_date)

        selection_layout.addWidget(self.corp_label)
        selection_layout.addWidget(self.corp_selector, 1)
        selection_layout.addWidget(self.os_label)
        selection_layout.addWidget(self.os_selector, 1)
        selection_layout.addWidget(branch_label)
        selection_layout.addWidget(self.branch_selector, 1)
        selection_layout.addWidget(date_label)
        selection_layout.addWidget(self.date_picker)
        selection_layout.addWidget(self.load_button)


        self.reviewed_checkbox = QCheckBox("Pending review")
        self.reviewed_checkbox.setStyleSheet("""
            QCheckBox {
                font-weight: bold; font-size: 12px; padding: 5px 10px;
                color: #c0392b;
            }
            QCheckBox::indicator { width: 18px; height: 18px; }
            QCheckBox::indicator:checked {
                background-color: #27ae60; border: 2px solid #1e8449; border-radius: 3px;
            }
            QCheckBox::indicator:unchecked {
                background-color: white; border: 2px solid #bdc3c7; border-radius: 3px;
            }
        """)
        self.reviewed_checkbox.setEnabled(False)
        self.reviewed_checkbox.toggled.connect(self._on_review_toggled)
        selection_layout.addWidget(self.reviewed_checkbox)

        header_layout.addLayout(selection_layout)


        balance_layout = QHBoxLayout()
        balance_label = QLabel("Beginning Balance:")
        balance_label.setProperty("class", "important")
        self.beginning_balance_input = self.create_money_input()
        self.beginning_balance_input.setReadOnly(True)
        balance_layout.addWidget(balance_label)
        balance_layout.addWidget(self.beginning_balance_input)
        balance_layout.addStretch()

        header_layout.addLayout(balance_layout)
        layout.addWidget(header_frame)


        columns_layout = QHBoxLayout()
        columns_layout.setSpacing(15)

    
        debit_box = QGroupBox("CREDIT")
        debit_form = QFormLayout()
        debit_form.setSpacing(8)
        debit_form.setLabelAlignment(Qt.AlignLeft)

        for label in self.debit_fields.keys():
            field_input = self.create_money_input()
            field_input.setReadOnly(False)
            self.debit_inputs[label] = field_input

            lotes_input = self.create_lotes_input(read_only=False) 
            self.debit_lotes_inputs[label] = lotes_input

            row = QHBoxLayout()
            row.setContentsMargins(0, 0, 0, 0)
            row.setSpacing(6)
            row.addWidget(field_input, 2)
            lotes_label = QLabel("Lotes:")
            row.addWidget(lotes_label)
            row.addWidget(lotes_input)

            field_label = QLabel(label)
            if any(keyword in label.lower() for keyword in ['interest', 'penalty', 'rescate']):
                field_label.setProperty("class", "important")
            
            if label == "MC In":
                mc_in_btn = QPushButton("View")
                mc_in_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #3B82F6; color: white;
                        border: none; border-radius: 5px;
                        font-size: 11px; font-weight: 700;
                        padding: 4px 8px;
                    }
                    QPushButton:hover { background-color: #2563EB; }
                """)
                mc_in_btn.setToolTip("Show MC In currency breakdown")
                mc_in_btn.clicked.connect(lambda checked, ft="MC In": self.show_mc_breakdown(ft))
                row.addWidget(mc_in_btn)
     
            elif label in ("Fund Transfer", "Fund Transfer from BRANCH"):
                from_branch_btn = QPushButton("View")
                from_branch_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #059669; color: white;
                        border: none; border-radius: 5px;
                        font-size: 11px; font-weight: 700;
                        padding: 4px 10px;
                    }
                    QPushButton:hover { background-color: #047857; }
                """)
                from_branch_btn.setToolTip("View source branch for this fund transfer")
                from_branch_btn.clicked.connect(self.show_from_branch_dest_info)
                self.from_branch_dest_btn = from_branch_btn
                row.addWidget(from_branch_btn)
            
            debit_form.addRow(field_label, row)

        debit_box.setLayout(debit_form)


        credit_box = QGroupBox("DEBIT")
        credit_form = QFormLayout()
        credit_form.setSpacing(8)
        credit_form.setLabelAlignment(Qt.AlignLeft)

        from Client.salary_detail_dialog import SalaryDetailDialog
        for label in self.credit_fields.keys():
            field_input = self.create_money_input()
            field_input.setReadOnly(False) 
            self.credit_inputs[label] = field_input

            lotes_input = self.create_lotes_input(read_only=False) 
            self.credit_lotes_inputs[label] = lotes_input

            row = QHBoxLayout()
            row.setContentsMargins(0, 0, 0, 0)
            row.setSpacing(6)
            row.addWidget(field_input, 2)

     
            if label == "Fund Transfer to HEAD OFFICE":
                bank_btn = QPushButton("View")
                bank_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #8B5CF6; color: white;
                        border: none; border-radius: 5px;
                        font-size: 11px; font-weight: 700;
                        padding: 4px 10px;
                    }
                    QPushButton:hover { background-color: #7C3AED; }
                """)
                bank_btn.setToolTip("View fund transfer breakdown")
                def show_ft_ho_breakdown():
                    from Client.client_dashboard import FundTransferHODialog
                    entry = self.get_current_entry_data()
                    if not entry:
                        QMessageBox.information(self, "No Entry Loaded",
                            "Please load an entry first by selecting a branch and date, then clicking Load.")
                        return
                    breakdown = []
                    raw = entry.get('ft_ho_breakdown')
                    if raw:
                        try:
                            breakdown = json.loads(raw)
                        except Exception:
                            breakdown = []
                    if not breakdown:
                  
                        if self.selected_bank_account:
                            self.show_bank_account_info()
                        else:
                            QMessageBox.information(self, "No Breakdown",
                                "No Fund Transfer to HEAD OFFICE breakdown data found.\n\n"
                                "The client needs to submit using the new breakdown format.")
                        return
                    dlg = FundTransferHODialog("Fund Transfer to HEAD OFFICE", parent=self)
                    dlg.setWindowTitle("Fund Transfer to HEAD OFFICE Breakdown (View Only)")
                    dlg.setMinimumSize(750, 460)
            
                    while dlg.table.rowCount() > 0:
                        dlg.table.removeRow(0)
                    dlg._rows_data = []
                    for bank_display, bank_id, amt in breakdown:
                        row_idx = dlg.table.rowCount()
                        dlg.table.insertRow(row_idx)
                  
                        bank_label = QLabel(f"  {bank_display}")
                        bank_label.setStyleSheet(
                            "font-size: 12px; font-weight: 600; color: #1E293B; padding: 4px 6px;"
                        )
                        bank_label.setToolTip(bank_display)
                        dlg.table.setCellWidget(row_idx, 0, bank_label)
                        amt_label = QLabel(f"  {amt:,.2f}")
                        amt_label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
                        amt_label.setStyleSheet(
                            "font-size: 13px; font-weight: 700; color: #0F766E; padding: 4px 8px;"
                        )
                        dlg.table.setCellWidget(row_idx, 1, amt_label)
             
                        empty = QLabel("")
                        dlg.table.setCellWidget(row_idx, 2, empty)
                    dlg._recalc()
            
                    for child in dlg.findChildren(QPushButton):
                        if "Add" in child.text():
                            child.setVisible(False)
                        if child.text() == "Post":
                            child.setText("Close")
                    dlg.exec_()
                bank_btn.clicked.connect(show_ft_ho_breakdown)
                self.bank_account_btn = bank_btn
                row.addWidget(bank_btn)
   
            elif label == "Fund Transfer to BRANCH":
                branch_btn = QPushButton("View")
                branch_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #059669; color: white;
                        border: none; border-radius: 5px;
                        font-size: 11px; font-weight: 700;
                        padding: 4px 10px;
                    }
                    QPushButton:hover { background-color: #047857; }
                """)
                branch_btn.setToolTip("View destination branch for this fund transfer")
                branch_btn.clicked.connect(self.show_branch_dest_info)
                self.branch_dest_btn = branch_btn
                row.addWidget(branch_btn)
            else:
                lotes_label = QLabel("Lotes:")
                row.addWidget(lotes_label)
                row.addWidget(lotes_input)

            field_label = QLabel(label)
            if any(keyword in label.lower() for keyword in ['empeno', 'fund transfer', 'salary']):
                field_label.setProperty("class", "important")

     
            if label == "PC-Salary" and self.account_type == 1:
                breakdown_btn = QPushButton("View")
                breakdown_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #2563EB; color: white;
                        border: none; border-radius: 5px;
                        font-size: 12px; font-weight: 700;
                        padding: 4px 12px;
                    }
                    QPushButton:hover { background-color: #1D4ED8; }
                """)
                breakdown_btn.setToolTip("Show salary breakdown for P.C. Salary")
                def show_salary_breakdown():
                    from Client.salary_detail_dialog import SalaryDetailDialog
         
                    entry = self.get_current_entry_data()
                    breakdown = []
                    if entry and entry.get('pc_salary_breakdown'):
                        try:
                            breakdown = json.loads(entry['pc_salary_breakdown'])
                        except Exception:
                            breakdown = []
                    if not breakdown:
                        QMessageBox.information(self, "No Breakdown", 
                            "No salary breakdown available for this entry.\n\n"
                            "Please load an entry with salary breakdown data first.")
                        return
                    dlg = SalaryDetailDialog(parent=self)
                    dlg.setWindowTitle("P.C. Salary Breakdown (View Only)")
                
                    while dlg.table.rowCount() > 0:
                        dlg.table.removeRow(0)
                    dlg._rows_data = []
                
                    for name, salary in breakdown:
                        dlg._add_row()
                        row_idx = dlg.table.rowCount() - 1
                        dlg.table.cellWidget(row_idx, 0).setText(str(name))
                        dlg.table.cellWidget(row_idx, 1).setText(str(salary))
               
                        dlg.table.cellWidget(row_idx, 0).setReadOnly(True)
                        dlg.table.cellWidget(row_idx, 1).setReadOnly(True)
                    dlg._recalc()
                    dlg.exec_()
                breakdown_btn.clicked.connect(show_salary_breakdown)
                row.addWidget(breakdown_btn)

        
            if label == "Empeno Motor/Car":
                motor_btn = QPushButton("View")
                motor_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #D97706; color: white;
                        border: none; border-radius: 5px;
                        font-size: 11px; font-weight: 700;
                        padding: 4px 8px;
                    }
                    QPushButton:hover { background-color: #B45309; }
                """)
                motor_btn.setToolTip("Show Motor/Car breakdown")
                def show_motor_breakdown():
                    from Client.client_dashboard import MotorCarDetailDialog
                    entry = self.get_current_entry_data()
                    breakdown = []
                    if entry and entry.get('empeno_motor_car_breakdown'):
                        try:
                            breakdown = json.loads(entry['empeno_motor_car_breakdown'])
                        except Exception:
                            breakdown = []
                    if not breakdown:
                        QMessageBox.information(self, "No Breakdown",
                            "No Motor/Car breakdown available for this entry.\n\n"
                            "Please load an entry with Motor/Car breakdown data first.")
                        return
                    dlg = MotorCarDetailDialog("Empeno Motor/Car", parent=self)
                    dlg.setWindowTitle("Empeno Motor/Car Breakdown (View Only)")
     
                    while dlg.table.rowCount() > 0:
                        dlg.table.removeRow(0)
                    dlg._rows_data = []
                    for pct_str, amt in breakdown:
                        dlg._add_row()
                        row_idx = dlg.table.rowCount() - 1
                        combo = dlg.table.cellWidget(row_idx, 0)
                        idx = combo.findText(pct_str)
                        if idx >= 0:
                            combo.setCurrentIndex(idx)
                        combo.setEnabled(False)
                        amt_edit = dlg.table.cellWidget(row_idx, 1)
                        amt_edit.setText(f"{amt:.2f}")
                        amt_edit.setReadOnly(True)
                        rem_btn = dlg.table.cellWidget(row_idx, 3)
                        if rem_btn:
                            rem_btn.setVisible(False)
                    dlg._recalc()
                    dlg.exec_()
                motor_btn.clicked.connect(show_motor_breakdown)
                row.addWidget(motor_btn)

    
            if label == "MC Out":
                mc_out_btn = QPushButton("View")
                mc_out_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #DC2626; color: white;
                        border: none; border-radius: 5px;
                        font-size: 11px; font-weight: 700;
                        padding: 4px 8px;
                    }
                    QPushButton:hover { background-color: #B91C1C; }
                """)
                mc_out_btn.setToolTip("Show MC Out currency breakdown")
                mc_out_btn.clicked.connect(lambda checked, ft="MC Out": self.show_mc_breakdown(ft))
                row.addWidget(mc_out_btn)

            credit_form.addRow(field_label, row)

        credit_box.setLayout(credit_form)

        columns_layout.addWidget(debit_box)
        columns_layout.addWidget(credit_box)
        layout.addLayout(columns_layout)


        totals_frame = QFrame()
        totals_frame.setStyleSheet(
            "background-color: #e8f5e9; border: 2px solid #81c784; border-radius: 8px; padding: 15px;")
        totals_layout = QHBoxLayout(totals_frame)

 
        debit_total_label = QLabel("Total Cash Receipt:")
        debit_total_label.setProperty("class", "important")
        debit_total_label.setStyleSheet("font-weight: bold; font-size: 12px; color: #2e7d32;")
        self.debit_total_display = self.create_display_field()
        self.debit_total_display.setStyleSheet(
            "background-color: #c8e6c9; border: 2px solid #66bb6a; font-weight: bold; font-size: 12px; color: #1b5e20;")


        credit_total_label = QLabel("Total Cash out:")
        credit_total_label.setProperty("class", "important")
        credit_total_label.setStyleSheet("font-weight: bold; font-size: 12px; color: #c62828;")
        self.credit_total_display = self.create_display_field()
        self.credit_total_display.setStyleSheet(
            "background-color: #ffcdd2; border: 2px solid #e57373; font-weight: bold; font-size: 12px; color: #b71c1c;")

        totals_layout.addWidget(debit_total_label)
        totals_layout.addWidget(self.debit_total_display, 1)
        totals_layout.addWidget(credit_total_label)
        totals_layout.addWidget(self.credit_total_display, 1)

        layout.addWidget(totals_frame)


        results_frame = QFrame()
        results_frame.setStyleSheet(
            "background-color: #fff3cd; border: 2px solid #ffeaa7; border-radius: 8px; padding: 15px;")
        results_layout = QHBoxLayout(results_frame)


        ending_label = QLabel("Ending Balance:")
        ending_label.setProperty("class", "important")
        self.ending_balance_display = self.create_display_field()
        self.ending_balance_display.setProperty("class", "result")


        cash_label = QLabel("Cash Count:")
        cash_label.setProperty("class", "important")
        self.cash_count_input = self.create_money_input()
        self.cash_count_input.setReadOnly(False)


        result_label = QLabel("⚖️ Short/Over:")
        result_label.setProperty("class", "important")
        self.cash_result_display = self.create_display_field()
        self.cash_result_display.setProperty("class", "result")

        results_layout.addWidget(ending_label)
        results_layout.addWidget(self.ending_balance_display, 1)
        results_layout.addWidget(cash_label)
        results_layout.addWidget(self.cash_count_input, 1)
        results_layout.addWidget(result_label)
        results_layout.addWidget(self.cash_result_display, 1)
        

        status_label = QLabel("Status:")
        status_label.setProperty("class", "important")
        self.variance_status_display = QLabel("—")
        self.variance_status_display.setStyleSheet(
            "font-weight: bold; font-size: 12px; padding: 5px 10px; border-radius: 4px;"
        )
        results_layout.addWidget(status_label)
        results_layout.addWidget(self.variance_status_display)

        layout.addWidget(results_frame)
        

        action_layout = QHBoxLayout()
        action_layout.addStretch()
        

        save_button = QPushButton("Save Changes")
        save_button.setObjectName("saveButton")
        save_button.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
            }
            QPushButton:hover {
                background-color: #219a52;
            }
        """)
        save_button.clicked.connect(self.save_entry)
        
        reset_button = QPushButton("Reset Entry")
        reset_button.setObjectName("resetButton")
        reset_button.setStyleSheet("""
            QPushButton {
                background-color: #E67E22;
                color: white;
            }
            QPushButton:hover {
                background-color: #D35400;
            }
        """)
        reset_button.clicked.connect(self.reset_entry)
        
        export_button = QPushButton("Export to Excel")
        export_button.setObjectName("exportButton")
        export_button.setStyleSheet("""
            QPushButton {
                background-color: #217346;
                color: white;
            }
            QPushButton:hover {
                background-color: #1a5c38;
            }
        """)
        export_button.clicked.connect(self.export_daily_cash_to_excel)
        

        full_brand_btn = QPushButton("Generate Report")
        full_brand_btn.setObjectName("fullBrandReportButton")
        full_brand_btn.setStyleSheet("""
            QPushButton {
                background-color: #27AE60;
                color: white;
            }
            QPushButton:hover {
                background-color: #1E8449;
            }
        """)
        full_brand_btn.clicked.connect(self.show_full_brand_report_dialog)

        action_layout.addWidget(save_button)
        action_layout.addWidget(reset_button)
        action_layout.addWidget(export_button)
        action_layout.addWidget(full_brand_btn)
        layout.addLayout(action_layout)

        # ── Palawan Details collapsible ─────────────────────────────────────
        self.palawan_inputs = {}
        self.palawan_total_displays = {}
        palawan_collapsible = self._build_palawan_collapsible()
        layout.addWidget(palawan_collapsible)


        # ── Live recalculation: connect every editable input ──────────────────
        for inp in self.debit_inputs.values():
            inp.textChanged.connect(self._recalc_totals)
        for inp in self.credit_inputs.values():
            inp.textChanged.connect(self._recalc_totals)
        self.beginning_balance_input.textChanged.connect(self._recalc_totals)
        self.cash_count_input.textChanged.connect(self._recalc_totals)

        container = QWidget()
        container_layout = QVBoxLayout(container)
        container_layout.addWidget(scroll_area)
        return container

    def create_money_input(self):
        """Create a currency input using reusable MoneyInput widget."""
        field = MoneyInput(placeholder="0.00")
        field.setProperty("class", "money")
        return field

    def create_display_field(self):
        """Create a read-only display field using reusable DisplayField widget."""
        field = DisplayField()
        field.setProperty("class", "result")
        return field

    def create_lotes_input(self, read_only=False):
        """Create an integer lotes input using reusable LotesInput widget."""
        field = LotesInput(read_only=read_only)
        field.setMaximumWidth(70)
        return field

    # ── Palawan Details collapsible ───────────────────────────────────────────
    def _build_palawan_collapsible(self):
        """Build a collapsible Palawan Details section for the Daily Cash Count tab."""
        brand_label = "" if self.account_type == 1 else ""
        self.palawan_inputs = {}   # {db_col: QLineEdit}
        self.palawan_total_displays = {}  # {section: QLineEdit (read-only)}

        # Outer wrapper
        wrapper = QFrame()
        wrapper.setStyleSheet(
            "QFrame { border: 1px solid #BAE6FD; border-radius: 8px; "
            "background-color: #F0F9FF; margin-top: 4px; }"
        )
        wrapper_layout = QVBoxLayout(wrapper)
        wrapper_layout.setContentsMargins(0, 0, 0, 0)
        wrapper_layout.setSpacing(0)

        # Toggle button
        toggle_btn = QPushButton(f"▶  Palawan Details {brand_label}")
        toggle_btn.setCheckable(True)
        toggle_btn.setChecked(False)
        toggle_btn.setStyleSheet("""
            QPushButton {
                text-align: left; padding: 10px 14px;
                font-weight: 700; font-size: 13px;
                color: #0369A1; background-color: #E0F2FE;
                border: none; border-radius: 8px;
            }
            QPushButton:checked {
                background-color: #BAE6FD; border-bottom-left-radius: 0; border-bottom-right-radius: 0;
            }
            QPushButton:hover { background-color: #BAE6FD; }
        """)
        wrapper_layout.addWidget(toggle_btn)

        # Content area (hidden by default)
        content = QFrame()
        content.setVisible(False)
        content.setStyleSheet(
            "QFrame { background-color: #FFFFFF; border: none; "
            "border-top: 1px solid #BAE6FD; border-radius: 0; }"
        )
        content_layout = QVBoxLayout(content)
        content_layout.setContentsMargins(12, 12, 12, 12)
        content_layout.setSpacing(10)

        # ── Section builder ──
        def _make_section(title, section_key, color):
            box = QGroupBox(title)
            box.setStyleSheet(f"""
                QGroupBox {{
                    border: 1px solid #E2E8F0; border-radius: 6px;
                    margin-top: 20px; padding: 16px 14px 14px 14px;
                    background-color: #FFFFFF;
                }}
                QGroupBox::title {{
                    color: {color}; font-weight: 800; font-size: 12px;
                    padding: 1px 8px; background-color: #FFFFFF;
                }}
            """)
            form = QFormLayout()
            form.setSpacing(8)
            form.setContentsMargins(12, 20, 12, 12)

            for sub_label, db_col in [
                ("Principal", f"palawan_{section_key}_principal"),
                ("SC",        f"palawan_{section_key}_sc"),
                ("Commission",f"palawan_{section_key}_commission"),
            ]:
                inp = MoneyInput(placeholder="0.00")
                self.palawan_inputs[db_col] = inp
                lbl = QLabel(sub_label + ":")
                lbl.setStyleSheet("font-size: 13px; font-weight: 600; color: #334155;")
                form.addRow(lbl, inp)

            # Lotes
            lotes_col = f"palawan_{section_key}_lotes_total"
            lotes_inp = LotesInput(read_only=False)
            self.palawan_inputs[lotes_col] = lotes_inp
            lotes_lbl = QLabel("Lotes:")
            lotes_lbl.setStyleSheet("font-size: 13px; font-weight: 600; color: #334155;")
            form.addRow(lotes_lbl, lotes_inp)

            # Total (auto-calc)
            total_col = f"palawan_{section_key}_regular_total"
            total_disp = DisplayField()
            total_disp.setStyleSheet(
                f"font-weight: 800; font-size: 13px; color: {color}; "
                "background-color: #F0F9FF; border: 1px solid #BAE6FD; "
                "border-radius: 5px; padding: 5px 10px;"
            )
            self.palawan_inputs[total_col] = total_disp
            self.palawan_total_displays[section_key] = total_disp
            total_lbl = QLabel("TOTAL:")
            total_lbl.setStyleSheet(f"font-weight: 700; color: {color}; font-size: 13px;")
            form.addRow(total_lbl, total_disp)

            # Wire principal/sc/commission to auto-calc total
            def _recalc(_, sk=section_key):
                p  = float(self.palawan_inputs.get(f"palawan_{sk}_principal",  QLineEdit()).text() or 0)
                sc = float(self.palawan_inputs.get(f"palawan_{sk}_sc",         QLineEdit()).text() or 0)
                cm = float(self.palawan_inputs.get(f"palawan_{sk}_commission",  QLineEdit()).text() or 0)
                self.palawan_total_displays[sk].setText(f"{p + sc + cm:.2f}")
                self.palawan_inputs[f"palawan_{sk}_regular_total"].setText(f"{p + sc + cm:.2f}")

            for sub in ("principal", "sc", "commission"):
                self.palawan_inputs[f"palawan_{section_key}_{sub}"].textChanged.connect(_recalc)

            box.setLayout(form)
            return box

        # Three section groups in a grid
        grid_frame = QFrame()
        grid_layout = QGridLayout(grid_frame)
        grid_layout.setSpacing(10)
        grid_layout.addWidget(_make_section("PALAWAN SEND-OUT",      "sendout",       "#1b75bc"), 0, 0)
        grid_layout.addWidget(_make_section("PALAWAN PAY-OUT",       "payout",        "#1b75bc"), 0, 1)
        grid_layout.addWidget(_make_section("PALAWAN INTERNATIONAL", "international", "#1b75bc"), 1, 0, 1, 2)
        content_layout.addWidget(grid_frame)

        # Auto-sync: sendout principal/sc → daily cash count debit fields
        def _sync_sendout_principal(text):
            if "Palawan Send Out" in self.debit_inputs:
                self.debit_inputs["Palawan Send Out"].setText(text)

        def _sync_sendout_sc(text):
            if "Palawan S.C" in self.debit_inputs:
                self.debit_inputs["Palawan S.C"].setText(text)

        def _sync_payout_principal(text):
            if "Palawan Pay Out" in self.credit_inputs:
                self.credit_inputs["Palawan Pay Out"].setText(text)

        self.palawan_inputs["palawan_sendout_principal"].textChanged.connect(_sync_sendout_principal)
        self.palawan_inputs["palawan_sendout_sc"].textChanged.connect(_sync_sendout_sc)
        self.palawan_inputs["palawan_payout_principal"].textChanged.connect(_sync_payout_principal)

        # Adjustments section
        adj_box = QGroupBox("PALAWAN ADJUSTMENTS")
        adj_box.setStyleSheet("""
            QGroupBox {
                border: 2px solid #F59E0B; border-radius: 6px;
                margin-top: 20px; padding: 16px 14px 14px 14px;
                background-color: #FFFBEB;
            }
            QGroupBox::title {
                color: #D97706; font-weight: 800; font-size: 12px;
                padding: 1px 8px; background-color: #FFFBEB;
            }
        """)
        adj_form = QFormLayout()
        adj_form.setSpacing(8)
        adj_form.setContentsMargins(12, 20, 12, 12)

        # Mapping: adj_col (db key) → credit_fields UI label
        _adj_credit_label = {
            "palawan_pay_out_incentives": "Palawan Pay Out (incentives)",
            "palawan_suki_discounts":     "Palawan Suki Discounts",
            "palawan_suki_rebates":       "Palawan Suki Rebates",
            "palawan_cancel":             "Palawan Cancel",
        }

        def _make_adj_sync(db_col):
            cf_label = _adj_credit_label.get(db_col)
            def _sync(value):
                # value is float from MoneyInput.valueChanged signal
                if cf_label and cf_label in self.credit_inputs:
                    self.credit_inputs[cf_label].setText(f"{value:.2f}")
            return _sync

        for adj_label, adj_col in [
            ("Pay Out Incentives", "palawan_pay_out_incentives"),
            ("Suki Discounts",     "palawan_suki_discounts"),
            ("Suki Rebates",       "palawan_suki_rebates"),
            ("Cancel",             "palawan_cancel"),
        ]:
            inp = MoneyInput(placeholder="0.00")
            self.palawan_inputs[adj_col] = inp
            inp.valueChanged.connect(_make_adj_sync(adj_col))
            lbl = QLabel(adj_label + ":")
            lbl.setStyleSheet("font-size: 13px; font-weight: 600; color: #92400E;")
            adj_form.addRow(lbl, inp)
        adj_box.setLayout(adj_form)
        content_layout.addWidget(adj_box)

        wrapper_layout.addWidget(content)

        def _on_toggle(checked):
            content.setVisible(checked)
            toggle_btn.setText(
                f"▼  Palawan Details {brand_label}" if checked
                else f"▶  Palawan Details {brand_label}"
            )

        toggle_btn.toggled.connect(_on_toggle)
        return wrapper

    def _load_palawan_details(self, data: dict):
        """Populate palawan detail inputs from a loaded DB row."""
        if self.account_type == 1:
            # Brand A: sendout/payout/international data lives in payable_tbl_brand_a
            # (daily_reports_brand_a does NOT have palawan_sendout_principal etc.)
            corporation  = (data or {}).get('corporation') or self.corp_selector.currentText().strip()
            branch_name   = self.branch_selector.currentText()
            selected_date = self.date_picker.date().toString("yyyy-MM-dd")
            try:
                result = []
                if corporation:
                    result = self.db.execute_query(
                        "SELECT * FROM payable_tbl_brand_a WHERE corporation=%s AND branch=%s AND date=%s LIMIT 1",
                        (corporation, branch_name, selected_date)
                    )
                if not result:
                    # Compatibility fallback for legacy rows without corporation.
                    result = self.db.execute_query(
                        "SELECT * FROM payable_tbl_brand_a WHERE branch=%s AND date=%s LIMIT 1",
                        (branch_name, selected_date)
                    )
                if not result:
                    # New-structure fallback: branch posted to daily_reports_brand_a only.
                    # Map daily_reports fields to the payable_tbl schema best-effort.
                    try:
                        dr = self.db.execute_query(
                            "SELECT * FROM daily_reports_brand_a WHERE branch=%s AND date=%s LIMIT 1",
                            (branch_name, selected_date)
                        )
                        if dr:
                            d = dr[0]
                            so     = float(d.get('palawan_send_out', 0) or 0)
                            so_sc  = float(d.get('palawan_sc', 0) or 0)
                            po     = float(d.get('palawan_pay_out', 0) or 0)
                            po_inc = float(d.get('palawan_pay_out_incentives', 0) or 0)
                            result = [{
                                'sendout_capital':             so,
                                'sendout_sc':                  so_sc,
                                'sendout_commission':          0,
                                'sendout_lotes':               int(d.get('palawan_send_out_lotes', 0) or 0),
                                'sendout_total':               so + so_sc,
                                'payout_capital':              po,
                                'payout_sc':                   0,
                                'payout_commission':           0,
                                'payout_lotes':                int(d.get('palawan_pay_out_lotes', 0) or 0),
                                'payout_total':                po + po_inc,
                                'international_capital':       0,
                                'international_sc':            0,
                                'international_commission':    0,
                                'international_lotes':         0,
                                'international_total':         0,
                                'skid':        float(d.get('palawan_suki_discounts', 0) or 0),
                                'skir':        float(d.get('palawan_suki_rebates', 0) or 0),
                                'cancellation':float(d.get('palawan_cancel', 0) or 0),
                                'inc':         po_inc,
                            }]
                    except Exception as _fe:
                        logger.error("_load_palawan_details new-structure fallback: %s", _fe)
                if result:
                    r = result[0]
                    payable_map = {
                        'palawan_sendout_principal':            r.get('sendout_capital', 0) or 0,
                        'palawan_sendout_sc':                   r.get('sendout_sc', 0) or 0,
                        'palawan_sendout_commission':           r.get('sendout_commission', 0) or 0,
                        'palawan_sendout_lotes_total':          r.get('sendout_lotes', 0) or 0,
                        'palawan_sendout_regular_total':        r.get('sendout_total', 0) or 0,
                        'palawan_payout_principal':             r.get('payout_capital', 0) or 0,
                        'palawan_payout_sc':                    r.get('payout_sc', 0) or 0,
                        'palawan_payout_commission':            r.get('payout_commission', 0) or 0,
                        'palawan_payout_lotes_total':           r.get('payout_lotes', 0) or 0,
                        'palawan_payout_regular_total':         r.get('payout_total', 0) or 0,
                        'palawan_international_principal':      r.get('international_capital', 0) or 0,
                        'palawan_international_sc':             r.get('international_sc', 0) or 0,
                        'palawan_international_commission':     r.get('international_commission', 0) or 0,
                        'palawan_international_lotes_total':    r.get('international_lotes', 0) or 0,
                        'palawan_international_regular_total':  r.get('international_total', 0) or 0,
                    }
                    data = {**data, **payable_map}
            except Exception as e:
                logger.error("_load_palawan_details Brand A payable query: %s", e)

        for db_col, widget in getattr(self, 'palawan_inputs', {}).items():
            val = data.get(db_col, 0) or 0
            widget.blockSignals(True)
            try:
                if widget.isReadOnly():
                    widget.setText(f"{float(val):.2f}" if float(val) else "")
                elif widget.validator() and hasattr(widget.validator(), 'decimals'):
                    widget.setText(f"{float(val):.2f}" if float(val) else "")
                else:
                    widget.setText(str(int(val)) if int(float(val)) else "")
            except (TypeError, ValueError):
                widget.setText("")
            widget.blockSignals(False)
        # Recalc totals
        for section in ("sendout", "payout", "international"):
            p  = float(getattr(self, 'palawan_inputs', {}).get(f"palawan_{section}_principal",  QLineEdit()).text() or 0)
            sc = float(getattr(self, 'palawan_inputs', {}).get(f"palawan_{section}_sc",         QLineEdit()).text() or 0)
            cm = float(getattr(self, 'palawan_inputs', {}).get(f"palawan_{section}_commission",  QLineEdit()).text() or 0)
            total_disp = getattr(self, 'palawan_total_displays', {}).get(section)
            if total_disp:
                total_disp.setText(f"{p + sc + cm:.2f}")

    BANK_ACCOUNTS = [
                {"id": 1, "bank_name": "CIB-BDO", "account_name": "Global Reliance", "account_number": "0077-9002-3923"},
        {"id": 2, "bank_name": "CIB-BPI", "account_name": "Kristal Clear Diamond and Gold Pawnshop", "account_number": "0091-0692-29"},
        {"id": 3, "bank_name": "CIB-BDO", "account_name": "Kristal Clear", "account_number": "0077-9001-8784"},
        {"id": 4, "bank_name": "CIB-Union Bank", "account_name": "Golbal Reliance Mgmt and Holdings Corp", "account_number": "0015-6000-5790"},
        {"id": 5, "bank_name": "CIB-BDO", "account_name": "Europacific Management & Holdings Corp", "account_number": "0038-1801-5838"},
        {"id": 6, "bank_name": "CIB-BPI", "account_name": "Europacific Management & Holdings Corp", "account_number": "3541-0035-67"},
        {"id": 7, "bank_name": "CIB-UB", "account_name": "Europacific Management & Holdings Corp", "account_number": "0021-7001-7921"},
    ]
    def show_bank_account_info(self):

        dialog = QDialog(self)
        dialog.setWindowTitle("Bank Account")
        dialog.setMinimumWidth(400)
        
        layout = QVBoxLayout(dialog)
        

        if self.selected_bank_account:

            selected_bank = None
            for bank in self.BANK_ACCOUNTS:
                if bank['id'] == self.selected_bank_account:
                    selected_bank = bank
                    break
            
            if selected_bank:

                header = QLabel("Client Selected Bank Account")
                header.setStyleSheet("font-size: 14px; font-weight: bold; color: #10B981; padding: 10px;")
                layout.addWidget(header)
                

                info_frame = QFrame()
                info_frame.setStyleSheet("""
                    QFrame {
                        background-color: #ECFDF5;
                        border: 2px solid #10B981;
                        border-radius: 8px;
                        padding: 15px;
                    }
                """)
                info_layout = QVBoxLayout(info_frame)
                info_layout.setSpacing(8)
                
                bank_label = QLabel(f"<b>Bank:</b> {selected_bank['bank_name']}")
                bank_label.setStyleSheet("font-size: 14px; color: #065F46;")
                info_layout.addWidget(bank_label)
                
                account_label = QLabel(f"<b>Account Name:</b> {selected_bank['account_name']}")
                account_label.setStyleSheet("font-size: 13px; color: #065F46;")
                info_layout.addWidget(account_label)
                
                if selected_bank.get('account_number'):
                    number_label = QLabel(f"<b>Account #:</b> {selected_bank['account_number']}")
                    number_label.setStyleSheet("font-size: 13px; color: #065F46;")
                    info_layout.addWidget(number_label)
                
                layout.addWidget(info_frame)
            else:
         
                header = QLabel("Bank Account Not Found")
                header.setStyleSheet("font-size: 14px; font-weight: bold; color: #F59E0B; padding: 10px;")
                layout.addWidget(header)
                note = QLabel(f"Bank account ID {self.selected_bank_account} not found in system.")
                note.setStyleSheet("font-size: 11px; color: #64748B; padding: 5px 10px;")
                layout.addWidget(note)
        else:
     
            header = QLabel("No Bank Account Selected")
            header.setStyleSheet("font-size: 14px; font-weight: bold; color: #64748B; padding: 10px;")
            layout.addWidget(header)
            
            note = QLabel("The client has not selected a bank account for this fund transfer.")
            note.setStyleSheet("font-size: 11px; color: #64748B; padding: 5px 10px;")
            layout.addWidget(note)
        
        layout.addStretch()
        
        close_btn = QPushButton("Close")
        close_btn.setStyleSheet("""
            QPushButton {
                background-color: #8B5CF6; color: white;
                border: none; border-radius: 5px;
                font-size: 12px; font-weight: 700;
                padding: 8px 20px;
            }
            QPushButton:hover { background-color: #7C3AED; }
        """)
        close_btn.clicked.connect(dialog.accept)
        layout.addWidget(close_btn)
        
        dialog.exec_()

    def show_branch_dest_info(self):
        """Show destination branch info for Fund Transfer to BRANCH (view only for admin)"""

        dialog = QDialog(self)
        dialog.setWindowTitle("Destination Branch")
        dialog.setMinimumWidth(400)
        
        layout = QVBoxLayout(dialog)
        

        if self.selected_branch_dest:

            header = QLabel("Fund Transfer Destination Branch")
            header.setStyleSheet("font-size: 14px; font-weight: bold; color: #059669; padding: 10px;")
            layout.addWidget(header)
            

            info_frame = QFrame()
            info_frame.setStyleSheet("""
                QFrame {
                    background-color: #ECFDF5;
                    border: 2px solid #059669;
                    border-radius: 8px;
                    padding: 15px;
                }
            """)
            info_layout = QVBoxLayout(info_frame)
            info_layout.setSpacing(8)
            
            branch_label = QLabel(f"<b>Destination Branch:</b> {self.selected_branch_dest}")
            branch_label.setStyleSheet("font-size: 14px; color: #065F46;")
            info_layout.addWidget(branch_label)
            
            layout.addWidget(info_frame)
        else:

            header = QLabel("No Destination Branch")
            header.setStyleSheet("font-size: 14px; font-weight: bold; color: #64748B; padding: 10px;")
            layout.addWidget(header)
            
            note = QLabel("The client has not specified a destination branch for this fund transfer.")
            note.setStyleSheet("font-size: 11px; color: #64748B; padding: 5px 10px;")
            layout.addWidget(note)
        
        layout.addStretch()
        

        close_btn = QPushButton("Close")
        close_btn.setStyleSheet("""
            QPushButton {
                background-color: #059669; color: white;
                border: none; border-radius: 5px;
                font-size: 12px; font-weight: 700;
                padding: 8px 20px;
            }
            QPushButton:hover { background-color: #047857; }
        """)
        close_btn.clicked.connect(dialog.accept)
        layout.addWidget(close_btn)
        
        dialog.exec_()

    def show_from_branch_dest_info(self):

        dialog = QDialog(self)
        dialog.setWindowTitle("Source Branch")
        dialog.setMinimumWidth(400)
        
        layout = QVBoxLayout(dialog)
        
        if self.selected_from_branch_dest:
            header = QLabel("Fund Transfer Source Branch")
            header.setStyleSheet("font-size: 14px; font-weight: bold; color: #059669; padding: 10px;")
            layout.addWidget(header)
            
            info_frame = QFrame()
            info_frame.setStyleSheet("""
                QFrame {
                    background-color: #ECFDF5;
                    border: 2px solid #059669;
                    border-radius: 8px;
                    padding: 15px;
                }
            """)
            info_layout = QVBoxLayout(info_frame)
            info_layout.setSpacing(8)
            
            branch_label = QLabel(f"<b>Source Branch:</b> {self.selected_from_branch_dest}")
            branch_label.setStyleSheet("font-size: 14px; color: #065F46;")
            info_layout.addWidget(branch_label)
            
            layout.addWidget(info_frame)
        else:
            header = QLabel("No Source Branch")
            header.setStyleSheet("font-size: 14px; font-weight: bold; color: #64748B; padding: 10px;")
            layout.addWidget(header)
            
            note = QLabel("The client has not specified a source branch for this fund transfer.")
            note.setStyleSheet("font-size: 11px; color: #64748B; padding: 5px 10px;")
            layout.addWidget(note)
        
        layout.addStretch()
        
        close_btn = QPushButton("Close")
        close_btn.setStyleSheet("""
            QPushButton {
                background-color: #059669; color: white;
                border: none; border-radius: 5px;
                font-size: 12px; font-weight: 700;
                padding: 8px 20px;
            }
            QPushButton:hover { background-color: #047857; }
        """)
        close_btn.clicked.connect(dialog.accept)
        layout.addWidget(close_btn)
        
        dialog.exec_()

    def build_admin_widget(self):
        """Build Admin Manage UI for corporations, branches, and clients"""
        widget = QWidget()
        
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_content = QWidget()
        layout = QVBoxLayout(scroll_content)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)

        corp_box = QGroupBox("Manage Corporations")
        corp_box.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                padding-top: 20px;
            }
        """)
        corp_layout = QVBoxLayout()
        
        corp_form = QFormLayout()
        corp_form.setSpacing(10)
        self.corp_name_input = QLineEdit()
        self.corp_name_input.setPlaceholderText("Enter corporation name")
        
        corp_add_btn = QPushButton("Add Corporation")
        corp_add_btn.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                padding: 8px 16px;
            }
            QPushButton:hover {
                background-color: #229954;
            }
        """)
        corp_add_btn.clicked.connect(self._on_add_corporation)
        
        corp_form.addRow(QLabel("Corporation Name:"), self.corp_name_input)
        corp_form.addRow(corp_add_btn)
        

        self.corp_list_display = QLabel("No corporations loaded")
        self.corp_list_display.setStyleSheet("""
            QLabel {
                background-color: #f8f9fa;
                border: 1px solid #dee2e6;
                border-radius: 4px;
                padding: 10px;
                font-family: 'Courier New', monospace;
                font-size: 10px;
                min-height: 100px;
            }
        """)
        self.corp_list_display.setWordWrap(True)
        self.corp_list_display.setTextInteractionFlags(Qt.TextSelectableByMouse)
        
        corp_refresh_btn = QPushButton("Refresh List")
        corp_refresh_btn.clicked.connect(self._refresh_corporation_display)
        
        corp_layout.addLayout(corp_form)
        corp_layout.addWidget(QLabel("Existing Corporations:"))
        corp_layout.addWidget(self.corp_list_display)
        corp_layout.addWidget(corp_refresh_btn)
        corp_box.setLayout(corp_layout)


        branch_box = QGroupBox("Manage Branches")
        branch_box.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                padding-top: 20px;
            }
        """)
        branch_layout = QVBoxLayout()

        branch_form = QFormLayout()
        branch_form.setSpacing(10)
        self.branch_corp_selector = QComboBox()
        self.branch_corp_selector.setMinimumWidth(200)
        self.branch_name_input = QLineEdit()
        self.branch_name_input.setPlaceholderText("Enter branch name")

        self.branch_os_selector = QComboBox()
        self.branch_os_selector.setMinimumWidth(200)
        self.branch_os_selector.addItem("-- Select OS (optional) --", None)
        self._load_os_options_for_branch()
        
        branch_add_btn = QPushButton("Add Branch")
        branch_add_btn.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                padding: 8px 16px;
            }
            QPushButton:hover {
                background-color: #229954;
            }
        """)
        branch_add_btn.clicked.connect(self._on_add_branch)
        
        branch_form.addRow(QLabel("Select Corporation:"), self.branch_corp_selector)
        branch_form.addRow(QLabel("Branch Name:"), self.branch_name_input)
        branch_form.addRow(QLabel("Operation Supervisor:"), self.branch_os_selector)
        branch_form.addRow(branch_add_btn)
        
        self.branch_list_display = QLabel("No branches loaded")
        self.branch_list_display.setStyleSheet("""
            QLabel {
                background-color: #f8f9fa;
                border: 1px solid #dee2e6;
                border-radius: 4px;
                padding: 10px;
                font-family: 'Courier New', monospace;
                font-size: 10px;
                min-height: 100px;
            }
        """)
        self.branch_list_display.setWordWrap(True)
        self.branch_list_display.setTextInteractionFlags(Qt.TextSelectableByMouse)
        
        branch_refresh_btn = QPushButton("Refresh List")
        branch_refresh_btn.clicked.connect(self._refresh_branch_display)
        
        branch_layout.addLayout(branch_form)
        branch_layout.addWidget(QLabel("Existing Branches:"))
        branch_layout.addWidget(self.branch_list_display)
        branch_layout.addWidget(branch_refresh_btn)
        branch_box.setLayout(branch_layout)


        client_box = QGroupBox("Manage Clients")
        client_box.setStyleSheet("""
            QGroupBox {
                font-size: 14px;
                font-weight: bold;
                padding-top: 20px;
            }
        """)
        client_layout = QVBoxLayout()
        

        client_form = QFormLayout()
        client_form.setSpacing(10)
        

        self.client_username_display = QLineEdit()
        self.client_username_display.setReadOnly(True)
        self.client_username_display.setPlaceholderText("Auto-generated (e.g., CL-0001)")
        self.client_username_display.setStyleSheet("""
            QLineEdit {
                background-color: #e9ecef;
                font-weight: bold;
                color: #495057;
                border: 2px solid #ced4da;
            }
        """)
        
        self.client_first_input = QLineEdit()
        self.client_first_input.setPlaceholderText("Enter first name")
        self.client_last_input = QLineEdit()
        self.client_last_input.setPlaceholderText("Enter last name")
        self.client_corp_selector = QComboBox()
        self.client_corp_selector.setMinimumWidth(200)
        self.client_branch_selector = QComboBox()
        self.client_branch_selector.setMinimumWidth(200)
        self.client_password_input = QLineEdit()
        self.client_password_input.setPlaceholderText("Enter password")
        self.client_password_input.setEchoMode(QLineEdit.Password)
        
        preview_btn = QPushButton("Preview Username")
        preview_btn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                padding: 6px 12px;
                font-size: 10px;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        preview_btn.clicked.connect(self._preview_username)
        
        client_add_btn = QPushButton("Add Client")
        client_add_btn.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                padding: 8px 16px;
            }
            QPushButton:hover {
                background-color: #229954;
            }
        """)
        client_add_btn.clicked.connect(self._on_add_client)
        
        client_form.addRow(QLabel("Username (Auto):"), self.client_username_display)
        client_form.addRow(preview_btn)
        client_form.addRow(QLabel("First Name:"), self.client_first_input)
        client_form.addRow(QLabel("Last Name:"), self.client_last_input)
        client_form.addRow(QLabel("Corporation:"), self.client_corp_selector)
        client_form.addRow(QLabel("Branch:"), self.client_branch_selector)
        client_form.addRow(QLabel("Password:"), self.client_password_input)
        client_form.addRow(client_add_btn)
        

        self.client_list_display = QLabel("No clients loaded")
        self.client_list_display.setStyleSheet("""
            QLabel {
                background-color: #f8f9fa;
                border: 1px solid #dee2e6;
                border-radius: 4px;
                padding: 10px;
                font-family: 'Courier New', monospace;
                font-size: 10px;
                min-height: 150px;
            }
        """)
        self.client_list_display.setWordWrap(True)
        self.client_list_display.setTextInteractionFlags(Qt.TextSelectableByMouse)
        
        client_refresh_btn = QPushButton("Refresh List")
        client_refresh_btn.clicked.connect(self._refresh_client_display)
        
        client_layout.addLayout(client_form)
        client_layout.addWidget(QLabel("Existing Clients:"))
        client_layout.addWidget(self.client_list_display)
        client_layout.addWidget(client_refresh_btn)
        client_box.setLayout(client_layout)

        layout.addWidget(corp_box)
        layout.addWidget(branch_box)
        layout.addWidget(client_box)
        layout.addStretch()


        scroll_area.setWidget(scroll_content)
        
        main_layout = QVBoxLayout(widget)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.addWidget(scroll_area)

        self._refresh_admin_corporations()
        self._refresh_corporation_display()
        self._refresh_branch_display()
        self._refresh_client_display()

        self.branch_corp_selector.currentIndexChanged.connect(
            lambda: self._refresh_admin_branches(self.branch_corp_selector.currentData())
        )
        self.client_corp_selector.currentIndexChanged.connect(
            lambda: self._refresh_admin_branches(self.client_corp_selector.currentData(), target='client')
        )

        return widget

    def _preview_username(self):
       
        try:
            row = self.db.execute_query(
                "SELECT MAX(CAST(SUBSTRING(username,4) AS UNSIGNED)) AS maxnum FROM users WHERE username LIKE 'CL-%'"
            )
            maxnum = 0
            if row and row[0] and row[0].get('maxnum') is not None:
                try:
                    maxnum = int(row[0]['maxnum'])
                except Exception:
                    maxnum = 0
            
            next_num = maxnum + 1
            username = f"CL-{str(next_num).zfill(4)}"
            self.client_username_display.setText(username)
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to preview username: {e}")

    def _refresh_corporation_display(self):
      
        try:
            rows = self.db.execute_query("SELECT id, name, created_at FROM corporations ORDER BY name")
            if not rows:
                self.corp_list_display.setText("No corporations found")
                return
            
   
            display_text = "┌─────┬────────────────────────────────┬─────────────────────┐\n"
            display_text += "│ ID  │ Corporation Name               │ Created At          │\n"
            display_text += "├─────┼────────────────────────────────┼─────────────────────┤\n"
            
            for r in rows:
                corp_id = str(r['id']).ljust(3)
                name = str(r['name'])[:30].ljust(30)
                created = str(r.get('created_at', 'N/A'))[:19].ljust(19)
                display_text += f"│ {corp_id} │ {name} │ {created} │\n"
            
            display_text += "└─────┴────────────────────────────────┴─────────────────────┘"
            self.corp_list_display.setText(display_text)
            
        except Exception as e:
            self.corp_list_display.setText(f"Error loading corporations: {e}")

    def _refresh_branch_display(self):
 
        try:
            query = """
                SELECT b.id, b.name, b.corporation_id, c.name as corp_name, b.created_at,
                       b.sub_corporation_id, sc.name as sub_corp_name
                FROM branches b
                LEFT JOIN corporations c ON b.corporation_id = c.id
                LEFT JOIN corporations sc ON b.sub_corporation_id = sc.id
                ORDER BY c.name, b.name
            """
            rows = self.db.execute_query(query)
            if not rows:
                self.branch_list_display.setText("No branches found")
                return
            
            display_text = "┌─────┬──────────────────────┬──────────────────────────────────────┬─────────────────────┐\n"
            display_text += "│ ID  │ Branch Name          │ Corporation                          │ Created At          │\n"
            display_text += "├─────┼──────────────────────┼──────────────────────────────────────┼─────────────────────┤\n"
            
            for r in rows:
                branch_id = str(r['id']).ljust(3)
                name = str(r['name'])[:20].ljust(20)
                corp_display = str(r.get('corp_name', 'N/A'))
                if r.get('sub_corp_name'):
                    corp_display += f" + {r.get('sub_corp_name')}"
                corp_display = corp_display[:36].ljust(36)
                created = str(r.get('created_at', 'N/A'))[:19].ljust(19)
                display_text += f"│ {branch_id} │ {name} │ {corp_display} │ {created} │\n"
            
            display_text += "└─────┴──────────────────────┴──────────────────────────────────────┴─────────────────────┘"
            self.branch_list_display.setText(display_text)
            
        except Exception as e:
            self.branch_list_display.setText(f"Error loading branches: {e}")

    def _refresh_client_display(self):
        """Refresh the client list display"""
        try:
            query = """
                SELECT u.id, u.username, u.first_name, u.last_name, 
                       u.corporation as corp_name, u.branch as branch_name, u.created_at
                FROM users u
                WHERE u.role = 'user'
                ORDER BY u.id DESC
                LIMIT 50
            """
            rows = self.db.execute_query(query)
            if not rows:
                self.client_list_display.setText("No clients found")
                return
            

            display_text = "┌─────┬──────────┬──────────────────────┬──────────────────┬──────────────────┬─────────────────────┐\n"
            display_text += "│ ID  │ Username │ Name                 │ Corporation      │ Branch           │ Created At          │\n"
            display_text += "├─────┼──────────┼──────────────────────┼──────────────────┼──────────────────┼─────────────────────┤\n"
            
            for r in rows:
                client_id = str(r['id']).ljust(3)
                username = str(r['username']).ljust(8)
                full_name = f"{r.get('first_name', '')} {r.get('last_name', '')}"[:20].ljust(20)
                corp = str(r.get('corp_name', 'N/A'))[:16].ljust(16)
                branch = str(r.get('branch_name', 'N/A'))[:16].ljust(16)
                created = str(r.get('created_at', 'N/A'))[:19].ljust(19)
                display_text += f"│ {client_id} │ {username} │ {full_name} │ {corp} │ {branch} │ {created} │\n"
            
            display_text += "└─────┴──────────┴──────────────────────┴──────────────────┴──────────────────┴─────────────────────┘"
            display_text += f"\n\nShowing last 50 clients (Total in database may be more)"
            self.client_list_display.setText(display_text)
            
        except Exception as e:
            self.client_list_display.setText(f"Error loading clients: {e}")

    def load_corporations(self):
        try:
            self.corp_selector.clear()
            # Query all corporations from corporations table
            result = self.db.execute_query("SELECT name as corporation FROM corporations ORDER BY name")
            if result:
                for row in result:
                    if row['corporation']:
                        self.corp_selector.addItem(row['corporation'])
            # Also load OS options
            self.load_os_options()

        except Exception as e:
            logger.error("Error loading corporations: %s", e)
            QMessageBox.critical(self, "Database Error", f"Failed to load corporations: {e}")

    def load_os_options(self):
        try:
            self.os_selector.clear()
            if self.os_group:
                # Restricted admin: only show their assigned group
                self.os_selector.addItem(self.os_group)
                self.os_selector.setCurrentIndex(0)
                self.os_selector.setEnabled(False)
            else:
                self.os_selector.setEnabled(True)
                result = self.db.execute_query("""
                    SELECT DISTINCT os_name FROM branches
                    WHERE os_name IS NOT NULL AND os_name != ''
                    ORDER BY os_name
                """)
                if result:
                    for row in result:
                        os_name = row['os_name'] if isinstance(row, dict) else row[0]
                        if os_name:
                            self.os_selector.addItem(os_name)
        except Exception as e:
            logger.error("Error loading OS options: %s", e)

    def on_filter_type_changed(self):
        filter_type = self.filter_type_selector.currentData()
        if filter_type == "corporation":
            self.corp_label.setVisible(True)
            self.corp_selector.setVisible(True)
            self.os_label.setVisible(False)
            self.os_selector.setVisible(False)
            self.load_branches()
        else:
            self.corp_label.setVisible(False)
            self.corp_selector.setVisible(False)
            self.os_label.setVisible(True)
            self.os_selector.setVisible(True)
            self.load_branches_by_os()

    def load_branches(self):
    
        try:
            self.branch_selector.clear()
            corp_name = self.corp_selector.currentText()
            if corp_name:

                query = """
                    SELECT b.name as branch
                    FROM branches b
                    LEFT JOIN corporations c ON b.corporation_id = c.id
                    LEFT JOIN corporations sc ON b.sub_corporation_id = sc.id
                    WHERE c.name = %s OR sc.name = %s
                    ORDER BY b.name
                """
                result = self.db.execute_query(query, [corp_name, corp_name])
                if result:
                    for row in result:
                        if row['branch']:
                            self.branch_selector.addItem(row['branch'])
        except Exception as e:
            logger.error("Error loading branches: %s", e)
            QMessageBox.critical(self, "Database Error", f"Failed to load branches: {e}")

    def load_branches_by_os(self):
        """Load branches filtered by OS name"""
        try:
            self.branch_selector.clear()
            os_name = self.os_selector.currentText()
            if os_name:

                query = """
                    SELECT name as branch
                    FROM branches
                    WHERE os_name = %s
                    ORDER BY name
                """
                result = self.db.execute_query(query, [os_name])
                if result:
                    for row in result:
                        if row['branch']:
                            self.branch_selector.addItem(row['branch'])
        except Exception as e:
            logger.error("Error loading branches by OS: %s", e)
            QMessageBox.critical(self, "Database Error", f"Failed to load branches: {e}")

    def _refresh_admin_corporations(self):
        try:
            rows = self.db.execute_query("SELECT id, name FROM corporations ORDER BY name")
            self.branch_corp_selector.clear()
            self.client_corp_selector.clear()
            if rows:
                for r in rows:
                    self.branch_corp_selector.addItem(r['name'], r['id'])
                    self.client_corp_selector.addItem(r['name'], r['id'])
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load corporations: {e}")

    def _refresh_admin_branches(self, corp_id=None, target='both'):
        try:
            if corp_id is None:
                corp_id = self.branch_corp_selector.currentData()
            if not corp_id:
                self.client_branch_selector.clear()
                return


            rows = self.db.execute_query("SELECT id, name FROM branches WHERE corporation_id=%s OR sub_corporation_id=%s ORDER BY name", (corp_id, corp_id))
            self.client_branch_selector.clear()
            if rows:
                for r in rows:
                    self.client_branch_selector.addItem(r['name'], r['id'])
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load branches: {e}")

    def _on_add_corporation(self):
        name = self.corp_name_input.text().strip()
        if not name:
            QMessageBox.warning(self, "Input Required", "Please enter a corporation name.")
            return
        try:
            cid = create_corporation(name)
            if cid:
                QMessageBox.information(self, "Created", f"Corporation '{name}' created successfully (ID: {cid}).")
                self.corp_name_input.clear()
                self._refresh_admin_corporations()
                self._refresh_corporation_display()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to create corporation: {e}")

    def _load_os_options_for_branch(self):

        try:
            supervisors = get_all_supervisors()
            for sup in supervisors:
                self.branch_os_selector.addItem(sup['name'], sup['name'])
        except Exception as e:
            logger.error("Error loading OS options: %s", e)

    def _on_add_branch(self):
        name = self.branch_name_input.text().strip()
        corp_id = self.branch_corp_selector.currentData()
        os_name = self.branch_os_selector.currentData()
        if not corp_id:
            QMessageBox.warning(self, "Selection Required", "Please select a corporation for this branch.")
            return
        if not name:
            QMessageBox.warning(self, "Input Required", "Please enter a branch name.")
            return
        try:
            bid = create_branch(name, corp_id, os_name)
            if bid:
                QMessageBox.information(self, "Created", f"Branch '{name}' created successfully (ID: {bid}).")
                self.branch_name_input.clear()
                self.branch_os_selector.setCurrentIndex(0)
                self._refresh_admin_branches(corp_id)
                self._refresh_branch_display()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to create branch: {e}")

    def _on_add_client(self):
        first = self.client_first_input.text().strip()
        last = self.client_last_input.text().strip()
        corp_id = self.client_corp_selector.currentData()
        branch_id = self.client_branch_selector.currentData()
        password = self.client_password_input.text() or None

        if not (first and last):
            QMessageBox.warning(self, "Input Required", "Please enter client's first and last names.")
            return
        if not corp_id or not branch_id:
            QMessageBox.warning(self, "Selection Required", "Please select corporation and branch for the client.")
            return

        try:
            row = create_client(first, last, corp_id, branch_id, password)
            if row:
                QMessageBox.information(
                    self, 
                    "✅ Created", 
                    f"Client created successfully!\n\n"
                    f"Username: {row['username']}\n"
                    f"ID: {row['id']}\n"
                    f"Name: {first} {last}"
                )
                # Clear inputs
                self.client_first_input.clear()
                self.client_last_input.clear()
                self.client_password_input.clear()
                self.client_username_display.clear()
                self._refresh_client_display()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to create client: {e}")

    def reset_entry(self):
  
        try:
            branch_name = self.branch_selector.currentText()
            selected_date = self.date_picker.date().toString("yyyy-MM-dd")

            if not branch_name:
                QMessageBox.warning(self, "Selection Required", "Please select a branch.")
                return

            reply = QMessageBox.question(
                self,
                "Confirm Reset",
                f"Are you sure you want to reset the entry for:\n\n"
                f"Branch: {branch_name}\n"
                f"Date: {selected_date}\n\n"
                f"This will reset BOTH Brand A and Brand B,\n"
                f"allowing the branch to edit and resubmit their report.",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )

            if reply == QMessageBox.No:
                return

   
            main_tables = ["daily_reports_brand_a", "daily_reports"]
            supp_tables = ["payable_tbl_brand_a", "cash_float_tbl"]
            found = False

            for table in main_tables:
                try:
                    col_check = self.db.execute_query(
                        "SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS "
                        "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = %s AND COLUMN_NAME = 'is_locked'",
                        [table]
                    )
                    if not col_check:
                        self.db.execute_query(
                            f"ALTER TABLE {table} ADD COLUMN is_locked TINYINT(1) NOT NULL DEFAULT 1",
                            []
                        )
                except Exception:
                    pass

                check = self.db.execute_query(
                    f"SELECT COUNT(*) AS cnt FROM {table} WHERE branch = %s AND date = %s",
                    [branch_name, selected_date]
                )
                if check and check[0].get('cnt', 0) > 0:
                    found = True
                    self.db.execute_query(
                        f"UPDATE {table} SET is_locked = 0 WHERE branch = %s AND date = %s",
                        [branch_name, selected_date]
                    )

            # Clear stale old-structure data so re-submissions start fresh
            for table in supp_tables:
                try:
                    self.db.execute_query(
                        f"DELETE FROM {table} WHERE branch = %s AND date = %s",
                        [branch_name, selected_date]
                    )
                except Exception:
                    pass

            if found:
                QMessageBox.information(
                    self,
                    "Entry Reset",
                    f"Entry for {branch_name} on {selected_date} has been reset.\n\n"
                    f"Both Brand A and Brand B are now unlocked.\n"
                    f"The branch can edit and resubmit their report."
                )
            else:
                QMessageBox.information(
                    self,
                    "No Entry Found",
                    f"No entry found for {branch_name} on {selected_date}."
                )

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to reset entry: {e}")

    def export_daily_cash_to_excel(self):
   
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            QMessageBox.critical(
                self,
                "Missing Dependency",
                "The openpyxl package is required to export to Excel.\nInstall with: pip install openpyxl"
            )
            return
        
        filter_type = self.filter_type_selector.currentData()
        if filter_type == "corporation":
            filter_label = "Corporation"
            filter_value = self.corp_selector.currentText()
        else:
            filter_label = "Group"
            filter_value = self.os_selector.currentText()
        
        branch_name = self.branch_selector.currentText()
        selected_date = self.date_picker.date().toString("yyyy-MM-dd")
        
        if not branch_name:
            QMessageBox.warning(self, "Selection Required", "Please select a branch.")
            return
        
        # File dialog for save location
        default_filename = f"DailyCashCount_{filter_value}_{branch_name}_{selected_date}.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Excel File",
            default_filename,
            "Excel Files (*.xlsx);;All Files (*)"
        )
        
        if not file_path:
            return  # User cancelled
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Daily Cash Count"
            
            # Styles
            title_font = Font(bold=True, size=16)
            header_font = Font(bold=True, size=11)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font_white = Font(bold=True, size=11, color="FFFFFF")
            money_font = Font(size=10)
            total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Title
            ws.merge_cells('A1:D1')
            ws['A1'] = "Daily Cash Count Report"
            ws['A1'].font = title_font
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Info section
            ws['A3'] = f"{filter_label}:"
            ws['B3'] = filter_value
            ws['A4'] = "Branch:"
            ws['B4'] = branch_name
            ws['A5'] = "Date:"
            ws['B5'] = selected_date
            ws['A6'] = "Beginning Balance:"
            ws['B6'] = float(self.beginning_balance_input.text() or 0)
            ws['B6'].number_format = '#,##0.00'
            
            for row in range(3, 7):
                ws[f'A{row}'].font = header_font
            
            # Debit Section
            row = 8
            ws.merge_cells(f'A{row}:D{row}')
            ws[f'A{row}'] = "💸 DEBIT TRANSACTIONS"
            ws[f'A{row}'].font = header_font_white
            ws[f'A{row}'].fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
            ws[f'A{row}'].alignment = Alignment(horizontal='center')
            
            row += 1
            ws[f'A{row}'] = "Description"
            ws[f'B{row}'] = "Amount"
            ws[f'C{row}'] = "Lotes"
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].font = header_font
                ws[f'{col}{row}'].border = border
            
            row += 1
            for label, db_col in self.debit_fields.items():
                ws[f'A{row}'] = label
                amount = float(self.debit_inputs[label].text() or 0) if label in self.debit_inputs else 0
                lotes = int(self.debit_lotes_inputs[label].text() or 0) if label in self.debit_lotes_inputs else 0
                ws[f'B{row}'] = amount
                ws[f'B{row}'].number_format = '#,##0.00'
                ws[f'C{row}'] = lotes
                for col in ['A', 'B', 'C']:
                    ws[f'{col}{row}'].border = border
                row += 1
            
            # Debit Total
            ws[f'A{row}'] = "Total Debit"
            ws[f'A{row}'].font = header_font
            ws[f'B{row}'] = float(self.debit_total_display.text() or 0)
            ws[f'B{row}'].number_format = '#,##0.00'
            ws[f'B{row}'].font = header_font
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].fill = total_fill
                ws[f'{col}{row}'].border = border
            
            row += 2
            
            # Credit Section
            ws.merge_cells(f'A{row}:D{row}')
            ws[f'A{row}'] = "💳 CREDIT TRANSACTIONS"
            ws[f'A{row}'].font = header_font_white
            ws[f'A{row}'].fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
            ws[f'A{row}'].alignment = Alignment(horizontal='center')
            
            row += 1
            ws[f'A{row}'] = "Description"
            ws[f'B{row}'] = "Amount"
            ws[f'C{row}'] = "Lotes"
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].font = header_font
                ws[f'{col}{row}'].border = border
            
            row += 1
            for label, db_col in self.credit_fields.items():
                ws[f'A{row}'] = label
                amount = float(self.credit_inputs[label].text() or 0) if label in self.credit_inputs else 0
                lotes = int(self.credit_lotes_inputs[label].text() or 0) if label in self.credit_lotes_inputs else 0
                ws[f'B{row}'] = amount
                ws[f'B{row}'].number_format = '#,##0.00'
                ws[f'C{row}'] = lotes
                for col in ['A', 'B', 'C']:
                    ws[f'{col}{row}'].border = border
                row += 1
            
            # Credit Total
            ws[f'A{row}'] = "Total Credit"
            ws[f'A{row}'].font = header_font
            ws[f'B{row}'] = float(self.credit_total_display.text() or 0)
            ws[f'B{row}'].number_format = '#,##0.00'
            ws[f'B{row}'].font = header_font
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].fill = total_fill
                ws[f'{col}{row}'].border = border
            
            row += 2
            
            # Results Section
            ws.merge_cells(f'A{row}:D{row}')
            ws[f'A{row}'] = "📊 SUMMARY"
            ws[f'A{row}'].font = header_font_white
            ws[f'A{row}'].fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
            ws[f'A{row}'].alignment = Alignment(horizontal='center')
            
            row += 1
            ws[f'A{row}'] = "Ending Balance:"
            ws[f'B{row}'] = float(self.ending_balance_display.text() or 0)
            ws[f'B{row}'].number_format = '#,##0.00'
            ws[f'A{row}'].font = header_font
            
            row += 1
            ws[f'A{row}'] = "Cash Count:"
            ws[f'B{row}'] = float(self.cash_count_input.text() or 0)
            ws[f'B{row}'].number_format = '#,##0.00'
            ws[f'A{row}'].font = header_font
            
            row += 1
            ws[f'A{row}'] = "Short/Over:"
            ws[f'B{row}'] = float(self.cash_result_display.text() or 0)
            ws[f'B{row}'].number_format = '#,##0.00'
            ws[f'A{row}'].font = header_font
            
            # Auto-adjust column widths
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 10
            ws.column_dimensions['D'].width = 10
            
            wb.save(file_path)
            QMessageBox.information(self, "Export Successful", f"Daily Cash Count exported to:\n{file_path}")
            
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Error exporting to Excel: {str(e)}")

    def show_full_brand_report_dialog(self):
        """Dialog to generate a multi-sheet Full Brand Report for all modules."""
        brand_label = "Brand A" if self.account_type == 1 else "Brand B"
        dialog = QDialog(self)
        dialog.setWindowTitle(f"Full Brand Report – {brand_label}")
        dialog.setMinimumWidth(460)

        layout = QVBoxLayout(dialog)
        layout.setSpacing(12)

        if self.account_type == 1:
            sheets_info = (
                "Daily Cash Count &nbsp;·&nbsp; Palawan &nbsp;·&nbsp; "
                "MC &nbsp;·&nbsp; Fund Transfer &nbsp;·&nbsp; Payable &nbsp;·&nbsp; "
                "Daily Transaction &nbsp;·&nbsp; Other Services &nbsp;·&nbsp; "
                "P&amp;L &nbsp;·&nbsp; New Sanla &nbsp;·&nbsp; New Renew &nbsp;·&nbsp; "
                "Global Other Services &nbsp;·&nbsp; FT HO"
            )
        else:
            sheets_info = (
                "Daily Cash Count &nbsp;·&nbsp; Palawan &nbsp;·&nbsp; "
                "MC &nbsp;·&nbsp; Fund Transfer &nbsp;·&nbsp; Payable &nbsp;·&nbsp; "
                "Global Payable &nbsp;·&nbsp; Payable Reports"
            )
        info_lbl = QLabel(
            f"Generates a comprehensive Excel workbook for <b>{brand_label}</b>.<br><br>"
            f"<b>Sheets:</b> {sheets_info}"
        )
        info_lbl.setWordWrap(True)
        info_lbl.setTextFormat(Qt.RichText)
        info_lbl.setStyleSheet("color:#2c3e50;font-size:11px;padding:6px;")
        layout.addWidget(info_lbl)

        # ── Filter type (Corporation / Group) ──────────────────────────────
        filter_grp = QGroupBox("Filter")
        filter_lay = QVBoxLayout(filter_grp)

        type_row = QHBoxLayout()
        from PyQt5.QtWidgets import QRadioButton
        self._fbr_corp_radio = QRadioButton("By Corporation")
        self._fbr_os_radio   = QRadioButton("By Group")
        self._fbr_corp_radio.setChecked(True)
        type_row.addWidget(self._fbr_corp_radio)
        type_row.addWidget(self._fbr_os_radio)
        type_row.addStretch()
        filter_lay.addLayout(type_row)

        sel_row = QHBoxLayout()
        self._fbr_corp_lbl  = QLabel("Corporation:")
        self._fbr_corp_lbl.setMinimumWidth(90)
        self._fbr_corp_combo = QComboBox()
        self._fbr_corp_combo.setMinimumWidth(260)
        try:
            rows = db_manager.execute_query(
                "SELECT name FROM corporations ORDER BY name"
            ) or []
            for r in rows:
                self._fbr_corp_combo.addItem(r['name'] if isinstance(r, dict) else r[0])
        except Exception:
            pass

        self._fbr_os_lbl  = QLabel("Group:")
        self._fbr_os_lbl.setMinimumWidth(90)
        self._fbr_os_combo = QComboBox()
        self._fbr_os_combo.setMinimumWidth(260)
        try:
            rows = db_manager.execute_query(
                "SELECT DISTINCT os_name FROM branches "
                "WHERE os_name IS NOT NULL AND os_name != '' ORDER BY os_name"
            ) or []
            for r in rows:
                self._fbr_os_combo.addItem(r['os_name'] if isinstance(r, dict) else r[0])
        except Exception:
            pass

        self._fbr_os_lbl.setVisible(False)
        self._fbr_os_combo.setVisible(False)

        sel_row.addWidget(self._fbr_corp_lbl)
        sel_row.addWidget(self._fbr_corp_combo)
        sel_row.addWidget(self._fbr_os_lbl)
        sel_row.addWidget(self._fbr_os_combo)
        sel_row.addStretch()
        filter_lay.addLayout(sel_row)
        layout.addWidget(filter_grp)

        def _toggle_fbr_filter():
            by_corp = self._fbr_corp_radio.isChecked()
            self._fbr_corp_lbl.setVisible(by_corp)
            self._fbr_corp_combo.setVisible(by_corp)
            self._fbr_os_lbl.setVisible(not by_corp)
            self._fbr_os_combo.setVisible(not by_corp)

        self._fbr_corp_radio.toggled.connect(_toggle_fbr_filter)

        # ── Branch Status filter ───────────────────────────────────────────
        reg_grp = QGroupBox("Branch Status")
        reg_lay = QHBoxLayout(reg_grp)
        reg_lay.addWidget(QLabel("Show:"))
        self._fbr_reg_filter = QComboBox()
        self._fbr_reg_filter.addItem("Registered Only", "registered")
        self._fbr_reg_filter.addItem("Not Registered",  "not_registered")
        self._fbr_reg_filter.addItem("All Branches",    "all")
        reg_lay.addWidget(self._fbr_reg_filter)
        reg_lay.addStretch()
        layout.addWidget(reg_grp)

        # ── Single date picker ─────────────────────────────────────────────
        date_grp = QGroupBox("Date")
        date_lay = QHBoxLayout(date_grp)
        date_lay.setSpacing(8)
        date_lay.addWidget(QLabel("Date:"))
        self._fbr_date = QDateEdit()
        self._fbr_date.setDisplayFormat("dd MMM yyyy")
        self._fbr_date.setCalendarPopup(True)
        self._fbr_date.setDate(QDate.currentDate())
        self._fbr_date.setMinimumWidth(150)
        date_lay.addWidget(self._fbr_date)
        date_lay.addStretch()
        layout.addWidget(date_grp)

        btn_lay = QHBoxLayout()
        gen_btn = QPushButton("Generate Report")
        gen_btn.setStyleSheet(
            "QPushButton{background-color:#27AE60;color:white;padding:8px 20px;"
            "font-weight:bold;border-radius:4px;}"
            "QPushButton:hover{background-color:#1E8449;}"
        )
        cancel_btn = QPushButton("Cancel")
        gen_btn.clicked.connect(lambda: self._generate_full_brand_report(dialog))
        cancel_btn.clicked.connect(dialog.reject)
        btn_lay.addStretch()
        btn_lay.addWidget(cancel_btn)
        btn_lay.addWidget(gen_btn)
        layout.addLayout(btn_lay)

        dialog.exec_()

    def _generate_full_brand_report(self, dialog):
        """Generate a multi-sheet Excel workbook with all module reports for the brand."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            QMessageBox.critical(self, "Missing Dependency",
                "The openpyxl package is required.\nInstall with: pip install openpyxl")
            return

        selected_date = self._fbr_date.date().toString("yyyy-MM-dd")

        # ── Filter type & value ───────────────────────────────────────────
        by_corp = self._fbr_corp_radio.isChecked()
        if by_corp:
            filter_type  = "corporation"
            filter_value = self._fbr_corp_combo.currentText().strip()
            filter_label = "Corporation"
        else:
            filter_type  = "os"
            filter_value = self._fbr_os_combo.currentText().strip()
            filter_label = "Group"

        if not filter_value:
            QMessageBox.warning(self, "Selection Required",
                f"Please select a {filter_label}.")
            return

        reg_filter = self._fbr_reg_filter.currentData()   # "registered" | "not_registered" | "all"

        brand_label = "Brand A" if self.account_type == 1 else "Brand B"
        safe_brand  = brand_label.replace(" ", "_")
        safe_filter = filter_value.replace(" ", "_").replace("/", "_")[:40]
        default_fn  = f"FullBrandReport_{safe_brand}_{safe_filter}_{selected_date}.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save Excel File", default_fn, "Excel Files (*.xlsx);;All Files (*)"
        )
        if not file_path:
            return

        try:
            from daily_transaction_page import (
                COLUMN_GROUPS          as DT_COLUMN_GROUPS,
                OTHER_SERVICES_COLUMN_GROUPS,
                PL_COLUMN_GROUPS,
            )
        except Exception:
            DT_COLUMN_GROUPS = []
            OTHER_SERVICES_COLUMN_GROUPS = []
            PL_COLUMN_GROUPS = []

        try:
            import datetime as _dt

            # ── Style helpers ────────────────────────────────────────────────
            thin   = Side(style='thin')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            def _fill(hex_color):
                return PatternFill("solid", fgColor=hex_color.upper())

            TITLE_FILL   = _fill("1F3864")
            INFO_FILL    = _fill("E8F4F8")
            HDR_FILL     = _fill("4472C4")
            TOTAL_FILL   = _fill("E2EFDA")
            GRAND_FILL   = _fill("D5A6BD")
            LOTES_FILL   = _fill("EBF5FB")
            DEBIT_FILL   = _fill("27AE60")
            CREDIT_FILL  = _fill("E74C3C")
            SUMMARY_FILL = _fill("F39C12")

            HDR_FONT   = Font(bold=True, size=9,  color="FFFFFF")
            TOTAL_FONT = Font(bold=True, size=10)
            BOLD_FONT  = Font(bold=True, size=10)
            GLOBAL_FILL = _fill("FFFF99")  # Yellow for global-tagged branches

            # Group → header fill colour (hex, no #)
            GROUP_FILL_MAP = {
                "JEWELRY":           "DC3545", "STORAGE":            "E67E22",
                "MOTOR/CAR":         "8E44AD", "MC":                 "2980B9",
                "SILVER":            "7F8C8D", "PALAWAN":            "16A085",
                "PALAWAN SEND OUT":  "16A085", "PALAWAN PAY OUT":    "1ABC9C",
                "INSURANCE":         "C0392B", "O.S.F":              "27AE60",
                "RESCATE JEW.":      "E74C3C", "RESCATE STO.":       "F39C12",
                "GCASH IN":          "1ABC9C", "GCASH OUT":          "1ABC9C",
                "MONEYGRAM":         "2C3E50", "TRANSFAST":          "34495E",
                "RIA":               "E74C3C", "I2I REM. IN":        "2980B9",
                "I2I BILLS":         "2980B9", "I2I INSTAPAY":       "2980B9",
                "SENDAH LOAD":       "8E44AD", "SENDAH BILLS":       "8E44AD",
                "PAYMAYA":           "27AE60", "SMART $ IN":         "117864",
                "SMART $ OUT":       "117864", "GCASH PADALA":       "148F77",
                "PAL PAY IN":        "16A085", "PAL PAY OUT":        "16A085",
                "REMITLY":           "7D3C98", "SEND OUT":           "27AE60",
                "PAY OUT":           "E74C3C", "INTERNATIONAL":      "8E44AD",
                "OTHER":             "F39C12", "FUND TRANSFER":      "2980B9",
                "FUND TRANSFER HO":  "2980B9", "SMART MONEY OUT":    "117864",
                "ABRA OUT":          "1A5276", "PAL PAY CASH OUT":   "196F3D",
                "MC OUT":            "2980B9", "MC IN (SELLING)":    "27AE60",
                "MC OUT (BUYING)":   "E74C3C", "DEBIT":              "27AE60",
                "CREDIT":            "E74C3C", "SUMMARY":            "F39C12",
                "JEWELRY EMPENO":    "DC3545", "STORAGE EMPENO":     "E67E22",
                "JEWELRY":           "DC3545",
            }

            def _group_fill(group_name):
                hex_c = GROUP_FILL_MAP.get(group_name, "4472C4")
                return _fill(hex_c)

            # ── Write info header rows (rows 1-5) ────────────────────────────
            def _write_info(ws):
                ws['A1'] = f"{brand_label} – {ws.title}"
                ws['A1'].font  = Font(bold=True, size=14, color="FFFFFF")
                ws['A1'].fill  = TITLE_FILL
                ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
                ws.row_dimensions[1].height = 22

                ws['A3'] = "Brand:";            ws['B3'] = brand_label
                ws['A4'] = f"{filter_label}:"; ws['B4'] = filter_value
                ws['A5'] = "Date:";             ws['B5'] = selected_date
                ws['A6'] = "Generated:";        ws['B6'] = _dt.datetime.now().strftime("%Y-%m-%d %H:%M")
                for r in range(3, 7):
                    ws.cell(row=r, column=1).font = BOLD_FONT
                    ws.cell(row=r, column=2).font = Font(size=10)
                    ws.cell(row=r, column=1).fill = INFO_FILL
                    ws.cell(row=r, column=2).fill = INFO_FILL

            _table_cols_cache = {}

            def _get_table_cols(tbl):
                if tbl not in _table_cols_cache:
                    try:
                        rows = db_manager.execute_query(
                            "SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS "
                            "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = %s",
                            (tbl,)
                        )
                        _table_cols_cache[tbl] = {r['COLUMN_NAME'] for r in rows} if rows else set()
                    except Exception:
                        _table_cols_cache[tbl] = set()
                return _table_cols_cache[tbl]

            def _write_grouped_sheet(ws, col_groups, table, use_branch_join=False, category_filter=None, show_lotes_total=False, show_amt_total=True):
                _write_info(ws)

                table_cols = _get_table_cols(table)

                flat = []   
                needed = set()
                for grp, subs in col_groups:
                    for sub_label, db_cols, is_lotes in subs:
                        valid = [c for c in db_cols if c and (not table_cols or c in table_cols)]
                        if not valid:
                            continue
                        needed.update(valid)
                        flat.append((sub_label, valid[0], is_lotes, grp, valid))

                if not flat:
                    ws['A7'] = "No columns defined."
                    return

                # Build SELECT: for multi-col entries, wrap in SUM(COALESCE(...)+COALESCE(...))
                seen_keys    = set()
                select_parts = []
                for _, primary_key, _, _, all_cols in flat:
                    if primary_key in seen_keys:
                        continue
                    seen_keys.add(primary_key)
                    if len(all_cols) == 1:
                        if table == "payable_tbl_brand_a" and all_cols[0] == "inc":
                            # INC should be a single branch value in report sheets; using MAX
                            # avoids accidental doubling when duplicate payable rows exist.
                            select_parts.append(
                                f"MAX(COALESCE(dr.`{all_cols[0]}`, 0)) AS `{all_cols[0]}`"
                            )
                        else:
                            select_parts.append(
                                f"SUM(COALESCE(dr.`{all_cols[0]}`, 0)) AS `{all_cols[0]}`"
                            )
                    else:
                        # Sum multiple columns into the first key
                        inner = " + ".join(f"COALESCE(dr.`{c}`, 0)" for c in all_cols)
                        select_parts.append(f"SUM({inner}) AS `{all_cols[0]}`")

                # ── Build WHERE clause with corp/os filter ────────────────
                cat_join = ""  # extra JOIN for 30%/60% category filter

                if filter_type == "corporation" and not use_branch_join:
                    if category_filter == "30":
                        cat_join = ("INNER JOIN branches b ON b.name COLLATE utf8mb4_general_ci "
                                    "= dr.branch COLLATE utf8mb4_general_ci")
                        where_clause = "AND dr.corporation = %s AND b.global_tag = 'GLOBAL'"
                        sql_params   = (selected_date, filter_value)
                    elif category_filter == "60":
                        # 60% = non-global-tagged branches only
                        cat_join = (
                            "INNER JOIN branches b ON b.name COLLATE utf8mb4_general_ci "
                            "= dr.branch COLLATE utf8mb4_general_ci"
                        )
                        where_clause = "AND dr.corporation = %s AND (b.global_tag IS NULL OR UPPER(TRIM(b.global_tag)) <> 'GLOBAL')"
                        sql_params   = (selected_date, filter_value)
                    else:
                        where_clause = "AND dr.corporation = %s"
                        sql_params   = (selected_date, filter_value)
                elif filter_type == "corporation" and use_branch_join:
                    where_clause = "AND b.global_tag = 'GLOBAL'"
                    sql_params   = (selected_date,)
                else:  # os/group – always need branches join
                    if category_filter == "30":
                        where_clause = "AND b.global_tag = 'GLOBAL'"
                    elif category_filter == "60":
                        # 60% = non-global-tagged branches only
                        where_clause = "AND (b.global_tag IS NULL OR UPPER(TRIM(b.global_tag)) <> 'GLOBAL')"
                    else:
                        # For global-tagged tables (e.g. global_other_services_tbl),
                        # restrict to branches with a global tag
                        where_clause = "AND b.global_tag = 'GLOBAL'" if use_branch_join else ""
                    sql_params = (selected_date,)

                try:
                    if category_filter:
                        # ── Branches-first approach for payable 60%/30% sheets ──────────
                        # Querying FROM the data table excludes branches with no data for
                        # the selected date.  Start from `branches` instead, and move the
                        # date condition into the LEFT JOIN ON clause so every branch
                        # (even those with no submission) always appears.
                        global_where = ""
                        if category_filter == "30":
                            global_where = "AND b.global_tag = 'GLOBAL'"
                        # 60%: show all branches (including global-tagged), highlight in yellow

                        if filter_type == "os":
                            if table == "payable_tbl_brand_a":
                                sql = (
                                    f"SELECT b.name AS branch, MAX(b.global_tag) AS global_tag, {', '.join(select_parts)} "
                                    f"FROM branches b "
                                    f"LEFT JOIN `{table}` dr "
                                    f"  ON b.name COLLATE utf8mb4_general_ci = dr.branch COLLATE utf8mb4_general_ci "
                                    f"  AND dr.date = %s "
                                    f"WHERE b.os_name = %s {global_where} "
                                    f"GROUP BY b.name ORDER BY b.name"
                                )
                            else:
                                sql = (
                                    f"SELECT b.name AS branch, MAX(b.global_tag) AS global_tag, {', '.join(select_parts)} "
                                    f"FROM branches b "
                                    f"INNER JOIN corporations c "
                                    f"  ON c.id = COALESCE(b.sub_corporation_id, b.corporation_id) "
                                    f"LEFT JOIN `{table}` dr "
                                    f"  ON b.name COLLATE utf8mb4_general_ci = dr.branch COLLATE utf8mb4_general_ci "
                                    f"  AND dr.corporation COLLATE utf8mb4_general_ci = c.name COLLATE utf8mb4_general_ci "
                                    f"  AND dr.date = %s "
                                    f"WHERE b.os_name = %s {global_where} "
                                    f"GROUP BY b.name ORDER BY b.name"
                                )
                            sql_params = (selected_date, filter_value)
                        else:
                            # corporation filter – join corps table to filter by name
                            # Also apply registration filter if specified
                            if reg_filter == "registered":
                                reg_clause_corp = "AND b.is_registered = 1"
                            elif reg_filter == "not_registered":
                                reg_clause_corp = "AND (b.is_registered = 0 OR b.is_registered IS NULL)"
                            else:
                                reg_clause_corp = ""

                            if table == "payable_tbl_brand_a":
                                sql = (
                                    f"SELECT b.name AS branch, MAX(b.global_tag) AS global_tag, {', '.join(select_parts)} "
                                    f"FROM branches b "
                                    f"INNER JOIN corporations c "
                                    f"  ON c.id = COALESCE(b.sub_corporation_id, b.corporation_id) "
                                    f"  AND c.name COLLATE utf8mb4_general_ci = %s "
                                    f"LEFT JOIN `{table}` dr "
                                    f"  ON b.name COLLATE utf8mb4_general_ci = dr.branch COLLATE utf8mb4_general_ci "
                                    f"  AND dr.date = %s "
                                    f"WHERE 1=1 {global_where} {reg_clause_corp} "
                                    f"GROUP BY b.name ORDER BY b.name"
                                )
                            else:
                                sql = (
                                    f"SELECT b.name AS branch, MAX(b.global_tag) AS global_tag, {', '.join(select_parts)} "
                                    f"FROM branches b "
                                    f"INNER JOIN corporations c "
                                    f"  ON c.id = COALESCE(b.sub_corporation_id, b.corporation_id) "
                                    f"  AND c.name COLLATE utf8mb4_general_ci = %s "
                                    f"LEFT JOIN `{table}` dr "
                                    f"  ON b.name COLLATE utf8mb4_general_ci = dr.branch COLLATE utf8mb4_general_ci "
                                    f"  AND dr.corporation COLLATE utf8mb4_general_ci = c.name COLLATE utf8mb4_general_ci "
                                    f"  AND dr.date = %s "
                                    f"WHERE 1=1 {global_where} {reg_clause_corp} "
                                    f"GROUP BY b.name ORDER BY b.name"
                                )
                            sql_params = (filter_value, selected_date)

                    elif filter_type == "os":
                        if reg_filter == "registered":
                            reg_clause_gs = "AND b.is_registered = 1"
                        elif reg_filter == "not_registered":
                            reg_clause_gs = "AND (b.is_registered = 0 OR b.is_registered IS NULL)"
                        else:
                            reg_clause_gs = ""
                        sql_params = (selected_date, filter_value)
                        sql = (
                            f"SELECT b.name AS branch, {', '.join(select_parts)} "
                            f"FROM branches b "
                            f"LEFT JOIN `{table}` dr "
                            f"  ON b.name COLLATE utf8mb4_general_ci = dr.branch COLLATE utf8mb4_general_ci "
                            f"  AND dr.date = %s "
                            f"WHERE b.os_name = %s {reg_clause_gs} "
                            f"GROUP BY b.name ORDER BY b.name"
                        )
                    elif use_branch_join:
                        # ── Branches-first approach to include branches without entries ──
                        # Start from branches, move date condition into LEFT JOIN ON so
                        # every branch appears (even those with no submission for the date).
                        sql = (
                            f"SELECT b.name AS branch, {', '.join(select_parts)} "
                            f"FROM branches b "
                            f"LEFT JOIN `{table}` dr "
                            f"  ON b.name COLLATE utf8mb4_general_ci = dr.branch COLLATE utf8mb4_general_ci "
                            f"  AND dr.date = %s "
                            f"WHERE 1=1 {where_clause} "
                            f"GROUP BY b.name ORDER BY b.name"
                        )
                        # sql_params already set correctly above
                    else:
                        # ── Branches-first approach for corporation filter without category ──
                        # Include branches even if they have no entries for the selected date.
                        if filter_type == "corporation":
                            # For corporation filter, join corporations to filter by name
                            # Also apply registration filter if specified
                            if reg_filter == "registered":
                                reg_clause_corp = "AND b.is_registered = 1"
                            elif reg_filter == "not_registered":
                                reg_clause_corp = "AND (b.is_registered = 0 OR b.is_registered IS NULL)"
                            else:
                                reg_clause_corp = ""

                            if table == "payable_tbl_brand_a":
                                sql = (
                                    f"SELECT b.name AS branch, {', '.join(select_parts)} "
                                    f"FROM branches b "
                                    f"INNER JOIN corporations c "
                                    f"  ON c.id = COALESCE(b.sub_corporation_id, b.corporation_id) "
                                    f"  AND c.name COLLATE utf8mb4_general_ci = %s "
                                    f"LEFT JOIN `{table}` dr "
                                    f"  ON b.name COLLATE utf8mb4_general_ci = dr.branch COLLATE utf8mb4_general_ci "
                                    f"  AND dr.date = %s "
                                    f"WHERE 1=1 {reg_clause_corp} "
                                    f"GROUP BY b.name ORDER BY b.name"
                                )
                            else:
                                sql = (
                                    f"SELECT b.name AS branch, {', '.join(select_parts)} "
                                    f"FROM branches b "
                                    f"INNER JOIN corporations c "
                                    f"  ON c.id = COALESCE(b.sub_corporation_id, b.corporation_id) "
                                    f"  AND c.name COLLATE utf8mb4_general_ci = %s "
                                    f"LEFT JOIN `{table}` dr "
                                    f"  ON b.name COLLATE utf8mb4_general_ci = dr.branch COLLATE utf8mb4_general_ci "
                                    f"  AND dr.corporation COLLATE utf8mb4_general_ci = c.name COLLATE utf8mb4_general_ci "
                                    f"  AND dr.date = %s "
                                    f"WHERE 1=1 {reg_clause_corp} "
                                    f"GROUP BY b.name ORDER BY b.name"
                                )
                            sql_params = (filter_value, selected_date)
                        else:
                            # Fallback for non-corporation filters (shouldn't normally reach here)
                            sql = (
                                f"SELECT dr.branch, {', '.join(select_parts)} "
                                f"FROM `{table}` dr {cat_join} "
                                f"WHERE dr.date = %s {where_clause} "
                                f"GROUP BY dr.branch ORDER BY dr.branch"
                            )
                            # sql_params already set from earlier (line 3367 or 3370)
                    logger.debug(f"Report query: {sql[:200]}... | Params: {sql_params}")
                    results = db_manager.execute_query(sql, sql_params) or []
                except Exception as ex:
                    logger.error(f"Report query failed: {ex} | SQL: {sql} | Params: {sql_params}")
                    ws['A7'] = f"Error loading data: {ex}"
                    return

                GRP_HDR_ROW = 7   # rows 1-6 are title/info
                SUB_HDR_ROW = 8   # sub-label row
                HDR_ROW     = SUB_HDR_ROW  # data starts after this

 
                ws.merge_cells(start_row=GRP_HDR_ROW, start_column=1,
                               end_row=SUB_HDR_ROW,   end_column=1)
                c = ws.cell(row=GRP_HDR_ROW, column=1, value="Branch")
                c.font = HDR_FONT; c.fill = HDR_FILL
                c.border = border
                c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                # Build group → list of column indices
                col_idx = 2
                grp_spans = []   # (group_name, start_col, end_col)
                for label, _, _, grp, _ in flat:
                    grp_spans.append((grp, col_idx))
                    col_idx += 1

                # Merge consecutive columns that share the same group name
                merged_groups = []
                i = 0
                while i < len(grp_spans):
                    grp_name, start = grp_spans[i]
                    end = start
                    while i + 1 < len(grp_spans) and grp_spans[i + 1][0] == grp_name:
                        i += 1
                        end = grp_spans[i][1]
                    merged_groups.append((grp_name, start, end))
                    i += 1

                for grp_name, start, end in merged_groups:
                    if start != end:
                        ws.merge_cells(start_row=GRP_HDR_ROW, start_column=start,
                                       end_row=GRP_HDR_ROW,   end_column=end)
                    c = ws.cell(row=GRP_HDR_ROW, column=start, value=grp_name.upper())
                    c.font  = HDR_FONT
                    c.fill  = _group_fill(grp_name)
                    c.border = border
                    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    # Apply border to every cell in the merged range
                    for ci in range(start, end + 1):
                        ws.cell(row=GRP_HDR_ROW, column=ci).border = border

                # "TOTAL LOTES" column (optional)
                lotes_total_col = None
                if show_lotes_total:
                    lotes_total_col = col_idx
                    ws.merge_cells(start_row=GRP_HDR_ROW, start_column=lotes_total_col,
                                   end_row=SUB_HDR_ROW,   end_column=lotes_total_col)
                    c = ws.cell(row=GRP_HDR_ROW, column=lotes_total_col, value="TOTAL LOTES")
                    c.font = HDR_FONT; c.fill = LOTES_FILL
                    c.border = border
                    c.alignment = Alignment(horizontal='center', vertical='center')
                    col_idx += 1

                # "AMT TOTAL" column (optional)
                total_col = None
                if show_amt_total:
                    total_col = col_idx
                    ws.merge_cells(start_row=GRP_HDR_ROW, start_column=total_col,
                                   end_row=SUB_HDR_ROW,   end_column=total_col)
                    c = ws.cell(row=GRP_HDR_ROW, column=total_col, value="AMT TOTAL")
                    c.font = HDR_FONT; c.fill = GRAND_FILL
                    c.border = border
                    c.alignment = Alignment(horizontal='center', vertical='center')

                # ── Row 8: Sub-label headers ──────────────────────────────────
                ws.cell(row=SUB_HDR_ROW, column=1).border = border  # already merged

                col_idx = 2
                for label, _, _, grp, _ in flat:
                    c = ws.cell(row=SUB_HDR_ROW, column=col_idx, value=label.upper())
                    c.font  = HDR_FONT
                    c.fill  = _group_fill(grp)
                    c.border = border
                    c.alignment = Alignment(horizontal='center', wrap_text=True)
                    col_idx += 1

                if total_col is not None:
                    ws.cell(row=SUB_HDR_ROW, column=total_col).border = border  # already merged

                # Data rows
                grand_totals = {primary_key: 0.0 for _, primary_key, _, _, _ in flat}
                grand_lotes_total = 0
                for ri, row_data in enumerate(results, HDR_ROW + 1):
                    branch_cell = ws.cell(row=ri, column=1, value=row_data.get('branch', ''))
                    branch_cell.border = border
                    # Yellow fill for globally-tagged branches in 60% sheet
                    if category_filter == "60" and (row_data.get('global_tag') or '').upper() == 'GLOBAL':
                        branch_cell.fill = GLOBAL_FILL
                    row_amt = 0.0
                    row_lotes = 0
                    for ci, (_, primary_key, is_lotes, _, all_cols) in enumerate(flat, 2):
                        raw = row_data.get(primary_key, 0) or 0
                        val = int(float(raw)) if is_lotes else float(raw)
                        cell = ws.cell(row=ri, column=ci, value=val)
                        cell.border = border
                        if is_lotes:
                            cell.alignment = Alignment(horizontal='center')
                            cell.fill = LOTES_FILL
                            row_lotes += int(float(raw))
                            grand_totals[primary_key] = grand_totals.get(primary_key, 0.0) + int(float(raw))
                        else:
                            cell.number_format = '#,##0.00'
                            cell.alignment = Alignment(horizontal='right')
                            row_amt += float(val)
                            grand_totals[primary_key] = grand_totals.get(primary_key, 0.0) + float(val)

                    if lotes_total_col is not None:
                        c = ws.cell(row=ri, column=lotes_total_col, value=row_lotes)
                        c.alignment = Alignment(horizontal='center')
                        c.fill = LOTES_FILL; c.border = border
                        grand_lotes_total += row_lotes

                    if total_col is not None:
                        c = ws.cell(row=ri, column=total_col, value=row_amt)
                        c.number_format = '#,##0.00'; c.font = TOTAL_FONT
                        c.border = border; c.fill = TOTAL_FILL
                        c.alignment = Alignment(horizontal='right')

                # Total row
                tr = HDR_ROW + len(results) + 1
                c = ws.cell(row=tr, column=1, value="TOTAL")
                c.font = TOTAL_FONT; c.fill = TOTAL_FILL; c.border = border

                grand_sum = 0.0
                for ci, (_, primary_key, is_lotes, _, _) in enumerate(flat, 2):
                    val = grand_totals.get(primary_key, 0.0)
                    cell = ws.cell(row=tr, column=ci, value=int(val) if is_lotes else val)
                    cell.font = TOTAL_FONT; cell.fill = TOTAL_FILL; cell.border = border
                    if is_lotes:
                        cell.alignment = Alignment(horizontal='center')
                    else:
                        cell.number_format = '#,##0.00'
                        cell.alignment = Alignment(horizontal='right')
                        grand_sum += val

                if lotes_total_col is not None:
                    lotes_grand = sum(int(grand_totals.get(pk, 0)) for _, pk, is_l, _, _ in flat if is_l)
                    c = ws.cell(row=tr, column=lotes_total_col, value=lotes_grand)
                    c.alignment = Alignment(horizontal='center')
                    c.font = TOTAL_FONT; c.fill = GRAND_FILL; c.border = border

                if total_col is not None:
                    c = ws.cell(row=tr, column=total_col, value=grand_sum)
                    c.number_format = '#,##0.00'; c.font = TOTAL_FONT
                    c.fill = GRAND_FILL; c.border = border
                    c.alignment = Alignment(horizontal='right')

                # Column widths
                ws.column_dimensions['A'].width = 25
                for ci, (_, _, is_lotes, _, _) in enumerate(flat, 2):
                    ws.column_dimensions[get_column_letter(ci)].width = 8 if is_lotes else 13
                if lotes_total_col is not None:
                    ws.column_dimensions[get_column_letter(lotes_total_col)].width = 12
                if total_col is not None:
                    ws.column_dimensions[get_column_letter(total_col)].width = 14
                ws.freeze_panes = f'B{HDR_ROW + 1}'

            # ── Check existing columns in daily_table ────────────────────────
            try:
                col_rows = db_manager.execute_query(
                    "SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS "
                    "WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = %s",
                    (self.daily_table,)
                )
                existing_cols = {r['COLUMN_NAME'] for r in col_rows} if col_rows else set()
            except Exception:
                existing_cols = set()

            # ── Workbook ─────────────────────────────────────────────────────
            wb = Workbook()

            # ════════════════════════════════════════════════════════════════
            # Sheet 1 – Daily Cash Count  (matches date-range DCC format)
            # ════════════════════════════════════════════════════════════════
            def _write_dcc_sheet(ws):
                """Write the Daily Cash Count sheet in the same vertical layout as the
                date-range Daily Cash Count report (fields in rows, branches in columns)."""

                # ── Styles matching date range report ────────────────────────
                title_font_dcc   = Font(bold=True, size=14)
                header_font_dcc  = Font(bold=True, size=9, color="FFFFFF")
                date_fill_dcc    = PatternFill("solid", fgColor="9B59B6")
                debit_fill_c     = PatternFill("solid", fgColor="27AE60")
                credit_fill_c    = PatternFill("solid", fgColor="E74C3C")
                summary_fill_c   = PatternFill("solid", fgColor="F39C12")
                total_fill_c     = PatternFill("solid", fgColor="E2EFDA")
                grand_total_fill = PatternFill("solid", fgColor="D5A6BD")
                info_fill_dcc    = PatternFill("solid", fgColor="E8F4F8")
                thin_s           = Side(style='thin')
                bdr              = Border(left=thin_s, right=thin_s, top=thin_s, bottom=thin_s)

                # PC field lists (same as date range report)
                salary_fields     = ['pc_salary']
                pc_fields_no_sal  = [
                    'pc_inc_emp', 'pc_inc_motor', 'pc_inc_suki_card', 'pc_inc_insurance',
                    'pc_inc_mc', 'pc_rental', 'pc_electric', 'pc_water', 'pc_internet',
                    'pc_lbc_jrs_jnt', 'pc_permits_bir_payments',
                    'pc_supplies_xerox_maintenance', 'pc_transpo'
                ]

                # ── Title & info ─────────────────────────────────────────────
                ws['A1'] = f"{brand_label} - Daily Cash Count Report"
                ws['A1'].font = title_font_dcc

                ws['A3'] = f"{filter_label}:";  ws['B3'] = filter_value
                ws['A4'] = "Date:";             ws['B4'] = selected_date
                ws['A3'].font = Font(bold=True)
                ws['A4'].font = Font(bold=True)

                # ── Determine available columns ──────────────────────────────
                tbl_cols = _get_table_cols(self.daily_table)
                d_cols = [c for c in self.debit_fields.values()
                          if not tbl_cols or c in tbl_cols]
                c_cols = [c for c in self.credit_fields.values()
                          if not tbl_cols or c in tbl_cols]

                # ── Build SELECT with GROUP BY aggregation ───────────────────
                # For Brand A: daily_reports_brand_a has N rows per branch/date (one per user).
                # Palawan adjustment fields are BRANCH-LEVEL singletons (same value in every
                # user row) → use MAX to avoid fan-out multiplication.
                # Regular transaction fields are per-user → use SUM.
                # For Brand B: 1 row per branch, so SUM/MAX behave identically.
                BRAND_A_SINGLETON_FIELDS = {
                    'palawan_cancel', 'palawan_suki_discounts', 'palawan_suki_rebates',
                    'palawan_pay_out_incentives', 'palawan_suki_card',
                }
                def _agg_fn(col):
                    """Return MAX for branch-level singletons, SUM for transaction fields."""
                    if self.account_type == 1 and col in BRAND_A_SINGLETON_FIELDS:
                        return "MAX"
                    return "SUM"

                sel_parts = [
                    "dr.branch",
                    "MAX(COALESCE(dr.`beginning_balance`, 0)) AS `beginning_balance`",
                ]
                for col in d_cols:
                    fn = _agg_fn(col)
                    sel_parts.append(f"{fn}(COALESCE(dr.`{col}`, 0)) AS `{col}`")
                    lc = f"{col}_lotes"
                    if not tbl_cols or lc in tbl_cols:
                        sel_parts.append(f"{fn}(COALESCE(dr.`{lc}`, 0)) AS `{lc}`")
                        # Capture text notes clients type in the lotes field
                        sel_parts.append(
                            f"GROUP_CONCAT(DISTINCT IF(dr.`{lc}` REGEXP '[a-zA-Z]',"
                            f" dr.`{lc}`, NULL) SEPARATOR ' | ') AS `{lc}_note`"
                        )
                    else:
                        sel_parts.append(f"0 AS `{lc}`")
                sel_parts.append("MAX(COALESCE(dr.`debit_total`, 0)) AS `debit_total`")
                for col in c_cols:
                    fn = _agg_fn(col)
                    sel_parts.append(f"{fn}(COALESCE(dr.`{col}`, 0)) AS `{col}`")
                    lc = f"{col}_lotes"
                    if not tbl_cols or lc in tbl_cols:
                        sel_parts.append(f"{fn}(COALESCE(dr.`{lc}`, 0)) AS `{lc}`")
                        # Capture text notes clients type in the lotes field
                        sel_parts.append(
                            f"GROUP_CONCAT(DISTINCT IF(dr.`{lc}` REGEXP '[a-zA-Z]',"
                            f" dr.`{lc}`, NULL) SEPARATOR ' | ') AS `{lc}_note`"
                        )
                    else:
                        sel_parts.append(f"0 AS `{lc}`")
                sel_parts.append("MAX(COALESCE(dr.`credit_total`, 0)) AS `credit_total`")
                sel_parts += [
                    "MAX(COALESCE(dr.`ending_balance`, 0)) AS `ending_balance`",
                    "MAX(COALESCE(dr.`cash_count`, 0)) AS `cash_count`",
                    "MAX(COALESCE(dr.`cash_result`, 0)) AS `cash_result`",
                ]
                for sf in salary_fields + pc_fields_no_sal:
                    sel_parts.append(
                        f"SUM(COALESCE(dr.`{sf}`, 0)) AS `{sf}`"
                        if (not tbl_cols or sf in tbl_cols) else f"0 AS `{sf}`"
                    )
                # Extra note columns for Fund Transfer and PC-Salary
                for _enc in ('fund_transfer_bank_account', 'fund_transfer_to_branch_dest',
                             'fund_transfer_from_branch_dest', 'pc_salary_breakdown',
                             'ft_ho_breakdown'):
                    if not tbl_cols or _enc in tbl_cols:
                        sel_parts.append(f"MAX(dr.`{_enc}`) AS `{_enc}`")
                sel_clause = ", ".join(sel_parts)

                try:
                    if filter_type == "os":
                        if reg_filter == "registered":
                            reg_clause_dcc = "AND b.is_registered = 1"
                        elif reg_filter == "not_registered":
                            reg_clause_dcc = "AND (b.is_registered = 0 OR b.is_registered IS NULL)"
                        else:
                            reg_clause_dcc = ""
                        dcc_sel_clause = sel_clause.replace("dr.branch", "b.name AS branch", 1)
                        sql = (
                            f"SELECT {dcc_sel_clause} FROM branches b "
                            f"LEFT JOIN `{self.daily_table}` dr "
                            "ON b.name COLLATE utf8mb4_general_ci = dr.branch COLLATE utf8mb4_general_ci "
                            "AND dr.date = %s "
                            f"WHERE b.os_name = %s {reg_clause_dcc} "
                            "GROUP BY b.name ORDER BY b.name"
                        )
                        rows = db_manager.execute_query(sql, (selected_date, filter_value)) or []
                    else:
                        sql = (
                            f"SELECT {sel_clause} FROM `{self.daily_table}` dr "
                            "WHERE dr.corporation = %s AND dr.date = %s "
                            "GROUP BY dr.branch ORDER BY dr.branch"
                        )
                        rows = db_manager.execute_query(sql, (filter_value, selected_date)) or []
                except Exception as ex:
                    ws.cell(row=8, column=1, value=f"Error loading data: {ex}")
                    return

                # ── Aggregate by branch ──────────────────────────────────────
                branch_totals = {}
                branch_notes  = {}  # text notes from lotes fields
                branches_list = []
                for row_data in rows:
                    bn = row_data.get('branch', 'Unknown')
                    if bn not in branch_totals:
                        branch_totals[bn] = {}
                        branch_notes[bn]  = {}
                        branches_list.append(bn)
                    for col in (['beginning_balance'] + d_cols +
                                ['debit_total'] + c_cols +
                                ['credit_total', 'ending_balance', 'cash_count', 'cash_result']):
                        branch_totals[bn][col] = (
                            branch_totals[bn].get(col, 0.0) + float(row_data.get(col, 0) or 0)
                        )
                    for col in d_cols + c_cols:
                        lc = f"{col}_lotes"
                        branch_totals[bn][lc] = (
                            branch_totals[bn].get(lc, 0) + int(float(row_data.get(lc, 0) or 0))
                        )
                        # Collect text note if client typed in lotes field
                        note_val = row_data.get(f"{lc}_note", '') or ''
                        if note_val:
                            branch_notes[bn][lc] = note_val
                    for sf in salary_fields:
                        branch_totals[bn]['salary'] = (
                            branch_totals[bn].get('salary', 0.0) + float(row_data.get(sf, 0) or 0)
                        )
                    for sf in pc_fields_no_sal:
                        branch_totals[bn]['total_pc'] = (
                            branch_totals[bn].get('total_pc', 0.0) + float(row_data.get(sf, 0) or 0)
                        )
                    # FT and PC-Salary notes
                    if True:
                        # Build a full bank-account detail map once per row
                        _BANK_DETAIL = {b['id']: b for b in getattr(self, 'BANK_ACCOUNTS', [])}
                        # ft_ho_breakdown takes priority (multiple FT entries)
                        _ft_ho_bd = (row_data.get('ft_ho_breakdown') or '').strip()
                        if _ft_ho_bd:
                            try:
                                _bd_items = json.loads(_ft_ho_bd)
                                _ft_parts = []
                                for _item in _bd_items:
                                    if isinstance(_item, dict):
                                        # dict format: {bank_account_id, amount, ...}
                                        _bid = _item.get('bank_account_id') or _item.get('id')
                                        _bdt = _BANK_DETAIL.get(int(_bid)) if _bid else None
                                        if _bdt:
                                            _ft_parts.append(
                                                f"{_bdt['bank_name']} - {_bdt['account_name']}"
                                                f" ({_bdt.get('account_number', '')})"
                                                f": {_item.get('amount', '')}"
                                            )
                                        else:
                                            _ft_parts.append(str(_item))
                                    elif isinstance(_item, (list, tuple)) and len(_item) >= 3:
                                        # list format: [bank_display, bank_id, amount]
                                        _ft_parts.append(f"{_item[0]}: {_item[2]}")
                                    elif isinstance(_item, (list, tuple)) and len(_item) >= 2:
                                        _ft_parts.append(f"{_item[0]}: {_item[1]}")
                                if _ft_parts:
                                    branch_notes[bn]['fund_transfer_to_head_office_lotes'] = (
                                        'Fund Transfer to HO:\n' + '\n'.join(_ft_parts)
                                    )
                            except Exception:
                                pass
                        else:
                            _bank_id = row_data.get('fund_transfer_bank_account')
                            if _bank_id:
                                _bdt = _BANK_DETAIL.get(int(_bank_id))
                                if _bdt:
                                    branch_notes[bn]['fund_transfer_to_head_office_lotes'] = (
                                        f"Bank: {_bdt['bank_name']}\n"
                                        f"Account: {_bdt['account_name']}\n"
                                        f"No: {_bdt.get('account_number', 'N/A')}"
                                    )
                                else:
                                    branch_notes[bn]['fund_transfer_to_head_office_lotes'] = f"Bank ID: {_bank_id}"
                        _ft_to_dest = (row_data.get('fund_transfer_to_branch_dest') or '').strip()
                        if _ft_to_dest:
                            branch_notes[bn]['fund_transfer_to_branch_lotes'] = f"To Branch: {_ft_to_dest}"
                        _ft_from_src = (row_data.get('fund_transfer_from_branch_dest') or '').strip()
                        if _ft_from_src:
                            branch_notes[bn]['fund_transfer_from_branch_lotes'] = f"From Branch: {_ft_from_src}"
                        _pc_sal_bd = (row_data.get('pc_salary_breakdown') or '').strip()
                        if _pc_sal_bd:
                            try:
                                _bd_items = json.loads(_pc_sal_bd)
                                _parts = [f"{_it[0]}: {_it[1]}" for _it in _bd_items
                                          if isinstance(_it, (list, tuple)) and len(_it) >= 2]
                                if _parts:
                                    branch_notes[bn]['pc_salary_lotes'] = '\n'.join(_parts)
                            except Exception:
                                pass
                branches_sorted = sorted(branches_list)

                # Recalculate derived totals from actual field values (fixes stale DB values)
                for bn in branches_list:
                    bt = branch_totals[bn]
                    _beginning  = bt.get('beginning_balance', 0.0)
                    _debit_sum  = sum(bt.get(col, 0.0) for col in d_cols)
                    _credit_sum = sum(bt.get(col, 0.0) for col in c_cols)
                    bt['debit_total']    = _beginning + _debit_sum
                    bt['credit_total']   = _credit_sum
                    bt['ending_balance'] = bt['debit_total'] - bt['credit_total']
                    bt['cash_result']    = bt.get('cash_count', 0.0) - bt['ending_balance']

                # ── Header row 6: "Field" | Branch names (merged 2 cols) | TOTAL ──
                HDR = 6
                cell = ws.cell(row=HDR, column=1, value="Field")
                cell.font = header_font_dcc; cell.fill = date_fill_dcc
                cell.border = bdr; cell.alignment = Alignment(horizontal='center', wrap_text=True)

                col_idx = 2
                for bn in branches_sorted:
                    ws.merge_cells(start_row=HDR, start_column=col_idx,
                                   end_row=HDR, end_column=col_idx + 1)
                    cell = ws.cell(row=HDR, column=col_idx, value=bn)
                    cell.font = header_font_dcc
                    cell.fill = PatternFill("solid", fgColor="4472C4")
                    cell.border = bdr
                    cell.alignment = Alignment(horizontal='center', wrap_text=True)
                    col_idx += 2

                ws.merge_cells(start_row=HDR, start_column=col_idx,
                               end_row=HDR, end_column=col_idx + 1)
                cell = ws.cell(row=HDR, column=col_idx, value="TOTAL")
                cell.font = header_font_dcc; cell.fill = grand_total_fill
                cell.border = bdr; cell.alignment = Alignment(horizontal='center', wrap_text=True)
                total_col_start = col_idx

                # ── Sub-header row 7: Lotes / Amount per branch ──────────────
                SHDR = HDR + 1
                cell = ws.cell(row=SHDR, column=1, value="")
                cell.font = header_font_dcc; cell.fill = date_fill_dcc; cell.border = bdr

                col_idx = 2
                for bn in branches_sorted:
                    for sub_lbl in ("Lotes", "Amount"):
                        cell = ws.cell(row=SHDR, column=col_idx, value=sub_lbl)
                        cell.font = Font(bold=True, size=8, color="FFFFFF")
                        cell.fill = PatternFill("solid", fgColor="4472C4")
                        cell.border = bdr
                        cell.alignment = Alignment(horizontal='center', wrap_text=True)
                        col_idx += 1

                for sub_lbl in ("Lotes", "Amount"):
                    cell = ws.cell(row=SHDR, column=col_idx, value=sub_lbl)
                    cell.font = Font(bold=True, size=8, color="FFFFFF")
                    cell.fill = grand_total_fill
                    cell.border = bdr
                    cell.alignment = Alignment(horizontal='center', wrap_text=True)
                    col_idx += 1

                # ── Helper: write one data row ───────────────────────────────
                current_row = SHDR + 1

                def _dcc_row(label, db_col, show_lotes=False,
                             row_font=None, row_fill=None, row_val_font=None):
                    """Write a single data row; returns (total_lotes, total_amount)."""
                    nonlocal current_row
                    c = ws.cell(row=current_row, column=1, value=label)
                    c.border = bdr
                    if row_font:
                        c.font = row_font
                    if row_fill:
                        c.fill = row_fill

                    tot_lotes, tot_amt = 0, 0.0
                    ci = 2
                    for bn in branches_sorted:
                        bt = branch_totals.get(bn, {})
                        lotes_val = int(bt.get(f"{db_col}_lotes", 0)) if show_lotes else ""
                        amt_val   = float(bt.get(db_col, 0))

                        lc = ws.cell(row=current_row, column=ci, value=lotes_val)
                        lc.border = bdr; lc.alignment = Alignment(horizontal='center')
                        if show_lotes:
                            _note = branch_notes.get(bn, {}).get(f"{db_col}_lotes", '')
                            if _note:
                                try:
                                    from openpyxl.comments import Comment as _OXLComment
                                    lc.comment = _OXLComment(_note, "Client Note")
                                    lc.comment.width = 300
                                    lc.comment.height = 80
                                except Exception:
                                    pass
                        ci += 1

                        ac = ws.cell(row=current_row, column=ci, value=amt_val)
                        ac.number_format = '#,##0.00'
                        ac.border = bdr; ac.alignment = Alignment(horizontal='right')
                        if row_val_font:
                            ac.font = row_val_font
                        ci += 1

                        if show_lotes:
                            tot_lotes += lotes_val
                        tot_amt += amt_val

                    # Total Lotes
                    tlc = ws.cell(row=current_row, column=total_col_start,
                                  value=tot_lotes if show_lotes else "")
                    tlc.border = bdr; tlc.alignment = Alignment(horizontal='center')
                    tlc.fill = total_fill_c

                    # Total Amount
                    tac = ws.cell(row=current_row, column=total_col_start + 1, value=tot_amt)
                    tac.number_format = '#,##0.00'
                    tac.border = bdr; tac.alignment = Alignment(horizontal='right')
                    tac.fill = total_fill_c
                    if row_val_font:
                        tac.font = row_val_font
                    elif row_font:
                        tac.font = row_font

                    current_row += 1
                    return tot_lotes, tot_amt

                def _dcc_section_header(label, fill):
                    nonlocal current_row
                    c = ws.cell(row=current_row, column=1, value=label)
                    c.font = Font(bold=True, color="FFFFFF")
                    c.fill = fill; c.border = bdr
                    current_row += 1

                # ── Beginning Balance ────────────────────────────────────────
                _dcc_row("Beginning Balance", "beginning_balance",
                         row_font=Font(bold=True), row_fill=info_fill_dcc)

                # ── CASH RECEIPT (DEBIT) ─────────────────────────────────────
                _dcc_section_header("CASH RECEIPT (DEBIT)", debit_fill_c)
                for lbl, db_col in self.debit_fields.items():
                    if db_col in d_cols:
                        _dcc_row(lbl, db_col, show_lotes=True)
                _dcc_row("Total Cash Receipt", "debit_total",
                         row_font=Font(bold=True), row_fill=info_fill_dcc)
                current_row += 1   # blank separator row

                # ── CASH OUT (CREDIT) ────────────────────────────────────────
                _dcc_section_header("CASH OUT (CREDIT)", credit_fill_c)
                for lbl, db_col in self.credit_fields.items():
                    _dcc_row(lbl, db_col, show_lotes=True)
                _dcc_row("Total Cash Out", "credit_total",
                         row_font=Font(bold=True), row_fill=info_fill_dcc)
                current_row += 1   # blank separator row

                # ── SUMMARY ──────────────────────────────────────────────────
                _dcc_section_header("SUMMARY", summary_fill_c)
                for lbl, db_col in [("Ending Balance", "ending_balance"),
                                     ("Cash Count",     "cash_count"),
                                     ("Variance",       "cash_result")]:
                    _dcc_row(lbl, db_col)

                # ── SALARY & Total PC ────────────────────────────────────────
                _dcc_row("SALARY",   "salary",   row_font=Font(bold=True), row_fill=info_fill_dcc,
                         row_val_font=Font(bold=True, color="9B59B6"))
                _dcc_row("Total PC", "total_pc", row_font=Font(bold=True), row_fill=info_fill_dcc,
                         row_val_font=Font(bold=True, color="9B59B6"))

                # ── Column widths ────────────────────────────────────────────
                ws.column_dimensions['A'].width = 25
                ci = 2
                for _ in branches_sorted:
                    ws.column_dimensions[get_column_letter(ci)].width = 10      # Lotes
                    ws.column_dimensions[get_column_letter(ci + 1)].width = 12  # Amount
                    ci += 2
                ws.column_dimensions[get_column_letter(total_col_start)].width = 10
                ws.column_dimensions[get_column_letter(total_col_start + 1)].width = 12
                ws.freeze_panes = 'B8'

            ws1 = wb.active
            ws1.title = "Daily Cash Count"
            _write_dcc_sheet(ws1)

            # ════════════════════════════════════════════════════════════════
            # Sheet 2 – Palawan  (mirrors palawan_page.py exactly)
            # Columns: Branch | Palawan In | Lotes In | Palawan Out | Lotes Out | Total
            # ════════════════════════════════════════════════════════════════
            def _write_palawan_sheet(ws):
                _write_info(ws)

                # ── Build query identical to palawan_page.py ──────────────
                select_cols = """
                    SELECT b.name AS branch,
                           COALESCE(dr.palawan_send_out, 0)                    AS palawan_send_out,
                           COALESCE(dr.palawan_sc, 0)                          AS palawan_sc,
                           COALESCE(dr.palawan_pay_out, 0)                     AS palawan_pay_out,
                           COALESCE(dr.palawan_pay_out_incentives, 0)          AS palawan_pay_out_incentives,
                           COALESCE(dr.palawan_send_out_lotes, 0)              AS palawan_send_out_lotes,
                           COALESCE(dr.palawan_sc_lotes, 0)                    AS palawan_sc_lotes,
                           COALESCE(dr.palawan_pay_out_lotes, 0)               AS palawan_pay_out_lotes,
                           COALESCE(dr.palawan_pay_out_incentives_lotes, 0)    AS palawan_pay_out_incentives_lotes
                """

                if reg_filter == "registered":
                    reg_clause = "AND b.is_registered = 1"
                elif reg_filter == "not_registered":
                    reg_clause = "AND b.is_registered = 0"
                else:
                    reg_clause = ""

                if filter_type == "corporation":
                    query = f"""
                        {select_cols}
                        FROM branches b
                        LEFT JOIN corporations c
                            ON (b.corporation_id = c.id OR b.sub_corporation_id = c.id)
                        LEFT JOIN `{self.daily_table}` dr
                            ON b.name COLLATE utf8mb4_general_ci = dr.branch COLLATE utf8mb4_general_ci
                           AND dr.corporation = %s
                           AND dr.date = %s
                        WHERE (b.corporation_id  = (SELECT id FROM corporations WHERE name = %s)
                            OR b.sub_corporation_id = (SELECT id FROM corporations WHERE name = %s))
                        {reg_clause}
                        ORDER BY b.name
                    """
                    params = (filter_value, selected_date, filter_value, filter_value)
                else:  # os / group
                    query = f"""
                        {select_cols}
                        FROM branches b
                        LEFT JOIN `{self.daily_table}` dr
                            ON b.name COLLATE utf8mb4_general_ci = dr.branch COLLATE utf8mb4_general_ci
                           AND dr.date = %s
                        WHERE b.os_name = %s
                        {reg_clause}
                        ORDER BY b.name
                    """
                    params = (selected_date, filter_value)

                try:
                    results = db_manager.execute_query(query, params) or []
                except Exception as ex:
                    ws['A7'] = f"Error loading data: {ex}"
                    return

                # ── Column headers (row 7) ────────────────────────────────
                headers = ["Branch", "Palawan In", "Lotes In", "Palawan Out", "Lotes Out"]
                hdr_fills = [HDR_FILL, _fill("16A085"), LOTES_FILL, _fill("E74C3C"), LOTES_FILL]
                for col_idx, (hdr, fill) in enumerate(zip(headers, hdr_fills), start=1):
                    c = ws.cell(row=7, column=col_idx, value=hdr)
                    c.font      = HDR_FONT
                    c.fill      = fill
                    c.border    = border
                    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.row_dimensions[7].height = 20

                # ── Data rows (row 8 onwards) ─────────────────────────────
                total_in      = 0.0
                total_out     = 0.0
                total_lotes_in  = 0
                total_lotes_out = 0
                data_row = 8

                for row_data in results:
                    branch_name  = row_data['branch']
                    send_out     = float(row_data['palawan_send_out']              or 0)
                    sc           = float(row_data['palawan_sc']                    or 0)
                    pay_out      = float(row_data['palawan_pay_out']               or 0)
                    incentives   = float(row_data['palawan_pay_out_incentives']    or 0)
                    so_lotes     = int(row_data['palawan_send_out_lotes']          or 0)
                    sc_lotes     = int(row_data['palawan_sc_lotes']                or 0)
                    po_lotes     = int(row_data['palawan_pay_out_lotes']           or 0)
                    inc_lotes    = int(row_data['palawan_pay_out_incentives_lotes']or 0)

                    palawan_in  = send_out + sc
                    palawan_out = pay_out + incentives
                    lotes_in    = so_lotes + sc_lotes
                    lotes_out   = po_lotes + inc_lotes

                    values = [branch_name, palawan_in, lotes_in, palawan_out, lotes_out]
                    for col_idx, val in enumerate(values, start=1):
                        c = ws.cell(row=data_row, column=col_idx, value=val)
                        c.border = border
                        if col_idx > 1:
                            if col_idx in (3, 5):  # lotes columns → integer
                                c.number_format = '0'
                            else:
                                c.number_format = '#,##0.00'
                            c.alignment = Alignment(horizontal='right')
                        else:
                            c.alignment = Alignment(horizontal='left')

                    total_in      += palawan_in
                    total_out     += palawan_out
                    total_lotes_in  += lotes_in
                    total_lotes_out += lotes_out
                    data_row += 1

                # ── Totals row ────────────────────────────────────────────
                totals = ["TOTAL", total_in, total_lotes_in, total_out, total_lotes_out]
                for col_idx, val in enumerate(totals, start=1):
                    c = ws.cell(row=data_row, column=col_idx, value=val)
                    c.font   = TOTAL_FONT
                    c.fill   = TOTAL_FILL
                    c.border = border
                    if col_idx > 1:
                        if col_idx in (3, 5):
                            c.number_format = '0'
                        else:
                            c.number_format = '#,##0.00'
                        c.alignment = Alignment(horizontal='right')
                    else:
                        c.alignment = Alignment(horizontal='left')

                # ── Column widths ─────────────────────────────────────────
                ws.column_dimensions['A'].width = 40
                for col_letter in ['B', 'C', 'D', 'E']:
                    ws.column_dimensions[col_letter].width = 16

            ws2 = wb.create_sheet(title="Palawan")
            _write_palawan_sheet(ws2)

            ws3 = wb.create_sheet(title="MC")
            mc_groups = [
                ("MC IN (SELLING)", [
                    ("Lotes",  ["mc_in_lotes"], True),
                    ("Amount", ["mc_in"],       False),
                ]),
                ("MC OUT (BUYING)", [
                    ("Lotes",  ["mc_out_lotes"], True),
                    ("Amount", ["mc_out"],       False),
                ]),
            ]
            _write_grouped_sheet(ws3, mc_groups, self.daily_table, show_amt_total=False)

            def _write_ft_sheet(ws):
  
                title_font_ft    = Font(bold=True, size=16)
                subtitle_font_ft = Font(bold=True, size=12)
                date_font_ft     = Font(size=11)
                header_font_ft   = Font(bold=True, size=11, color="FFFFFF")
                header_fill_ft   = PatternFill("solid", fgColor="4472C4")
                area_hdr_fill    = PatternFill("solid", fgColor="D4EDDA")
                area_hdr_font    = Font(bold=True, size=11)
                total_fill_ft    = PatternFill("solid", fgColor="E9ECEF")
                total_font_ft    = Font(bold=True)
                es_fill_ft       = PatternFill("solid", fgColor="FFF8E1")
                es_font_ft       = Font(bold=True)
                gt_fill_ft       = PatternFill("solid", fgColor="4472C4")
                gt_font_ft       = Font(bold=True, size=12, color="FFFFFF")
                thin_ft          = Side(style='thin')
                bdr_ft           = Border(left=thin_ft, right=thin_ft,
                                          top=thin_ft,  bottom=thin_ft)

                COL_LETTERS = ['A','B','C','D','E','F','G','H','I','J','K','L']

                # ── Title / info rows ────────────────────────────────────────
                ws.merge_cells('A1:L1')
                ws['A1'] = "FUND TRANSFER"
                ws['A1'].font = title_font_ft
                ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
                ws.row_dimensions[1].height = 20

                ws.merge_cells('A2:L2')
                ws['A2'] = f"{filter_label.upper()} {filter_value.upper()}"
                ws['A2'].font = subtitle_font_ft
                ws['A2'].alignment = Alignment(horizontal='left', vertical='center')

                ws.merge_cells('A3:L3')
                try:
                    import datetime as _dtft2
                    _dobj = _dtft2.datetime.strptime(str(selected_date), '%Y-%m-%d')
                    _fdate = _dobj.strftime('%A, %B %d, %Y')
                except Exception:
                    _fdate = str(selected_date)
                ws['A3'] = _fdate
                ws['A3'].font = date_font_ft
                ws['A3'].alignment = Alignment(horizontal='left', vertical='center')

                # Row 5: Column headers
                hdr_vals = ['AREA', '#', 'CORPORATION', 'LOB', 'GLOBAL', 'SUNDAY',
                            'Branch Name', 'Invty', 'CASH FLOAT', 'CASH COUNT',
                            'BR to HO', 'BR to BR']
                for cl, hval in zip(COL_LETTERS, hdr_vals):
                    c = ws[f'{cl}5']
                    c.value = hval
                    c.font = header_font_ft; c.fill = header_fill_ft
                    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    c.border = bdr_ft

                # Row 6: Sub-headers (blue; only K/L have text)
                for cl in COL_LETTERS:
                    c = ws[f'{cl}6']
                    c.value = ('BR to HO' if cl == 'K' else
                               'BR to BR' if cl == 'L' else '')
                    c.font = header_font_ft; c.fill = header_fill_ft
                    c.alignment = Alignment(horizontal='center', vertical='center')
                    c.border = bdr_ft

                # ── Column availability ──────────────────────────────────────
                tbl_cols = _get_table_cols(self.daily_table)
                def _col(name):
                    return name if (not tbl_cols or name in tbl_cols) else None

                cc_col  = _col('cash_count')

                def _expr(col):
                    return (f"COALESCE(SUM(dr.`{col}`), 0)" if col else "0")

                # Brand B has no cash float — skip the cash_float_tbl JOIN entirely
                _is_brand_b = (self.account_type != 1)
                _cf_select  = "0 AS cash_float" if _is_brand_b else "COALESCE(SUM(cf.cash_float), 0) AS cash_float"
                sel_ft = (
                    "dr.branch AS branch, "
                    "COALESCE(b.area, 'UNASSIGNED') AS area, "
                    "COALESCE(MAX(c.name), '') AS corporation_name, "
                    "COALESCE(b.line_of_business, '') AS line_of_business, "
                    "COALESCE(b.global_tag, '') AS global_tag, "
                    "COALESCE(b.sunday, '') AS sunday, "
                    f"{_expr(cc_col)}  AS cash_count, "
                    f"{_cf_select}"
                )
                join_c  = ("LEFT JOIN corporations c "
                           "ON c.id = COALESCE(b.sub_corporation_id, b.corporation_id)")
                join_cf = ("" if _is_brand_b else
                           "LEFT JOIN cash_float_tbl cf "
                           "ON cf.branch COLLATE utf8mb4_general_ci = dr.branch COLLATE utf8mb4_general_ci "
                           "AND cf.date = dr.date")
                grp     = ("GROUP BY dr.branch, b.area, b.line_of_business, b.global_tag, b.sunday "
                           "ORDER BY COALESCE(b.area, 'ZZZZZ'), dr.branch")

                try:
                    if filter_type == "os":
                        if reg_filter == "registered":
                            reg_clause_ft = "AND b.is_registered = 1"
                        elif reg_filter == "not_registered":
                            reg_clause_ft = "AND (b.is_registered = 0 OR b.is_registered IS NULL)"
                        else:
                            reg_clause_ft = ""
                        sel_ft_os = sel_ft.replace("dr.branch AS branch", "b.name AS branch", 1)
                        grp_os    = grp.replace("dr.branch", "b.name")
                        sql_ft = (
                            f"SELECT {sel_ft_os} FROM branches b "
                            f"LEFT JOIN `{self.daily_table}` dr "
                            "ON b.name COLLATE utf8mb4_general_ci = dr.branch COLLATE utf8mb4_general_ci "
                            "AND dr.date = %s "
                            f"{join_c} {join_cf} "
                            f"WHERE b.os_name = %s {reg_clause_ft} {grp_os}"
                        )
                        ft_rows = db_manager.execute_query(
                            sql_ft, (selected_date, filter_value)) or []
                    else:
                        sql_ft = (
                            f"SELECT {sel_ft} FROM `{self.daily_table}` dr "
                            "LEFT JOIN branches b "
                            "ON b.name COLLATE utf8mb4_general_ci = dr.branch COLLATE utf8mb4_general_ci "
                            f"{join_c} {join_cf} "
                            f"WHERE dr.corporation = %s AND dr.date = %s {grp}"
                        )
                        ft_rows = db_manager.execute_query(
                            sql_ft, (filter_value, selected_date)) or []
                except Exception as ex:
                    ws['A7'] = f"Error loading fund transfer data: {ex}"
                    return

                # Extra space value
                try:
                    _es = db_manager.execute_query(
                        "SELECT amount FROM extra_space_fund_transfer WHERE report_date = %s",
                        (selected_date,))
                    extra_space_val = float(_es[0].get('amount', 0) or 0) if _es else 0.0
                except Exception:
                    extra_space_val = 0.0

                # ── Group by area ────────────────────────────────────────────
                from collections import OrderedDict as _OD2
                area_groups = _OD2()
                for rd in ft_rows:
                    area_groups.setdefault(rd.get('area') or 'UNASSIGNED', []).append(rd)

                # ── Write data rows ──────────────────────────────────────────
                excel_row   = 7
                grand_total = 0.0
                branch_num  = 1

                for area_name, area_branches in area_groups.items():
                    # Area header (merged, green)
                    ws.merge_cells(f'A{excel_row}:L{excel_row}')
                    c = ws[f'A{excel_row}']
                    c.value = f"{area_name} AREA"
                    c.fill = area_hdr_fill; c.font = area_hdr_font
                    c.alignment = Alignment(horizontal='center', vertical='center')
                    c.border = bdr_ft
                    for cl in ['B','C','D','E','F','G','H','I','J','K','L']:
                        ws[f'{cl}{excel_row}'].fill  = area_hdr_fill
                        ws[f'{cl}{excel_row}'].border = bdr_ft
                    excel_row += 1

                    area_cc = 0.0
                    for rd in area_branches:
                        sunday = rd.get('sunday', '') or ''
                        if sunday == 'NO':
                            sunday = 'NO SUNDAY'
                        cash_count = float(rd.get('cash_count', 0) or 0)
                        cash_float = 0.0 if _is_brand_b else float(rd.get('cash_float', 0) or 0)

                        row_vals = [
                            (rd.get('area', '') or '',        'center', None),
                            (branch_num,                       'center', None),
                            (rd.get('corporation_name','')or '','center', None),
                            (rd.get('line_of_business','')or '','center', None),
                            (rd.get('global_tag', '') or '',   'center', None),
                            (sunday,                           'center', None),
                            (rd.get('branch', ''),             'left',   None),
                            ('',                               'center', None),
                            ('' if _is_brand_b else (cash_float if cash_float else ''), 'center', '#,##0.00'),
                            (cash_count,                       'right',  '#,##0.00'),
                            ('',                               'right',  None),
                            ('',                               'right',  None),
                        ]
                        for ci, (val, align, nfmt) in enumerate(row_vals, 1):
                            c = ws.cell(row=excel_row, column=ci, value=val)
                            c.alignment = Alignment(horizontal=align)
                            c.border = bdr_ft
                            if nfmt and isinstance(val, float):
                                c.number_format = nfmt

                        area_cc     += cash_count
                        grand_total += cash_count
                        branch_num  += 1
                        excel_row   += 1

                    # Area total row (grey)
                    for ci in range(1, 13):
                        c = ws.cell(row=excel_row, column=ci)
                        c.fill = total_fill_ft; c.font = total_font_ft; c.border = bdr_ft
                        if ci == 7:
                            c.value = f"TOTAL {area_name}"
                            c.alignment = Alignment(horizontal='right')
                        elif ci == 10:
                            c.value = area_cc
                            c.number_format = '#,##0.00'
                            c.alignment = Alignment(horizontal='right')
                        else:
                            c.value = ''; c.alignment = Alignment(horizontal='center')
                    excel_row += 1

                # Extra space row (yellow)
                for ci in range(1, 13):
                    c = ws.cell(row=excel_row, column=ci)
                    c.fill = es_fill_ft; c.font = es_font_ft; c.border = bdr_ft
                    if ci == 7:
                        c.value = "EXTRA SPACE"
                        c.alignment = Alignment(horizontal='center')
                    elif ci == 10:
                        if extra_space_val:
                            c.value = extra_space_val
                            c.number_format = '#,##0.00'
                        c.alignment = Alignment(horizontal='right')
                    else:
                        c.value = ''; c.alignment = Alignment(horizontal='center')
                excel_row += 1

                # Spacer row
                for ci in range(1, 13):
                    ws.cell(row=excel_row, column=ci).border = bdr_ft
                excel_row += 1

                # Grand total row (blue, white font)
                for ci in range(1, 13):
                    c = ws.cell(row=excel_row, column=ci)
                    c.fill = gt_fill_ft; c.font = gt_font_ft; c.border = bdr_ft
                    if ci == 7:
                        c.value = "GRAND TOTAL"
                        c.alignment = Alignment(horizontal='right')
                    elif ci == 10:
                        c.value = grand_total
                        c.number_format = '#,##0.00'
                        c.alignment = Alignment(horizontal='right')
                    else:
                        c.value = ''; c.alignment = Alignment(horizontal='center')

                # ── Column widths ────────────────────────────────────────────
                ws.column_dimensions['A'].width = 18
                ws.column_dimensions['B'].width = 5
                ws.column_dimensions['C'].width = 18
                ws.column_dimensions['D'].width = 12
                ws.column_dimensions['E'].width = 12
                ws.column_dimensions['F'].width = 12
                ws.column_dimensions['G'].width = 20
                ws.column_dimensions['H'].width = 12
                ws.column_dimensions['I'].width = 15
                ws.column_dimensions['J'].width = 15
                ws.column_dimensions['K'].width = 12
                ws.column_dimensions['L'].width = 12
                ws.freeze_panes = 'A7'

            ws4 = wb.create_sheet(title="Fund Transfer")
            _write_ft_sheet(ws4)

            # ════════════════════════════════════════════════════════════════
            # Sheet 5 – Payable (Palawan Reconciliation)
            # ════════════════════════════════════════════════════════════════
            _payable_groups_a = [
                ("SEND OUT", [
                    ("S.O. Lotes",   ["sendout_lotes"],        True),
                    ("S.O. Capital", ["sendout_capital"],      False),
                    ("S.O. SC",      ["sendout_sc"],           False),
                    ("S.O. Comm.",   ["sendout_commission"],   False),
                    ("S.O. Total",   ["sendout_total"],        False),
                ]),
                ("PAY OUT", [
                    ("P.O. Lotes",   ["payout_lotes"],         True),
                    ("P.O. Capital", ["payout_capital"],       False),
                    ("P.O. SC",      ["payout_sc"],            False),
                    ("P.O. Comm.",   ["payout_commission"],    False),
                    ("P.O. Total",   ["payout_total"],         False),
                ]),
                ("INTERNATIONAL", [
                    ("Int. Lotes",   ["international_lotes"],        True),
                    ("Int. Capital", ["international_capital"],      False),
                    ("Int. SC",      ["international_sc"],           False),
                    ("Int. Comm.",   ["international_commission"],   False),
                    ("Int. Total",   ["international_total"],        False),
                ]),
                ("OTHER", [
                    ("SKID",      ["skid"],         False),
                    ("SKIR",      ["skir"],         False),
                    ("Cancel",    ["cancellation"], False),
                    ("P.O. Inc.", ["inc"],          False),
                ]),
            ]
            if self.account_type == 1:
                # Brand A: split payable into 60% (specific corps) and 30% (Global) sheets
                ws5a = wb.create_sheet(title="Palawan 60%")
                _write_grouped_sheet(ws5a, _payable_groups_a, "payable_tbl_brand_a", category_filter="60", show_amt_total=False)
                ws5b = wb.create_sheet(title="Palawan 30%")
                _write_grouped_sheet(ws5b, _payable_groups_a, "payable_tbl_brand_a", category_filter="30", show_amt_total=False)
            else:
                ws5 = wb.create_sheet(title="Payable")
                _write_grouped_sheet(ws5, _payable_groups_a, "payable_tbl_brand_a", show_amt_total=False)

            # ════════════════════════════════════════════════════════════════
            # Sheet 6+ – Brand-specific sheets
            # ════════════════════════════════════════════════════════════════
            if self.account_type == 1:
                # ── Brand A: Daily Transaction, Other Services, P&L, New Sanla, New Renew, GOS, FT HO ──
                ws6 = wb.create_sheet(title="Daily Transaction")
                _write_grouped_sheet(ws6, DT_COLUMN_GROUPS, "daily_reports_brand_a", show_amt_total=False)

                ws7 = wb.create_sheet(title="Other Services")
                _write_grouped_sheet(ws7, OTHER_SERVICES_COLUMN_GROUPS, "daily_reports_brand_a", show_amt_total=False)

                ws8 = wb.create_sheet(title="P&L")
                _write_grouped_sheet(ws8, PL_COLUMN_GROUPS, "daily_reports_brand_a", show_amt_total=False)

                ws9 = wb.create_sheet(title="New Sanla")
                sanla_groups = [
                    ("JEWELRY EMPENO", [
                        ("Lotes",   ["empeno_jew_new_lotes"], True),
                        ("Capital", ["empeno_jew_new"],       False),
                    ]),
                    ("STORAGE EMPENO", [
                        ("Lotes",   ["empeno_sto_new_lotes"], True),
                        ("Capital", ["empeno_sto_new"],       False),
                    ]),
                ]
                _write_grouped_sheet(ws9, sanla_groups, self.daily_table, show_amt_total=False)

                ws10 = wb.create_sheet(title="New Renew")
                renew_groups = [
                    ("JEWELRY", [
                        ("JEW NEW Lotes",   ["empeno_jew_new_lotes"],           True),
                        ("JEW NEW Capital", ["empeno_jew_new"],                 False),
                        ("JEW RENEW Lotes", ["empeno_jew_renew_lotes"],         True),
                        ("JEW RENEW Cap.",  ["empeno_jew_renew"],               False),
                    ]),
                    ("STORAGE", [
                        ("STO NEW Lotes",   ["empeno_sto_new_lotes"],           True),
                        ("STO NEW Capital", ["empeno_sto_new"],                 False),
                        ("STO RENEW Lotes", ["fund_empeno_sto_renew_lotes"],    True),
                        ("STO RENEW Cap.",  ["fund_empeno_sto_renew"],          False),
                    ]),
                ]
                _write_grouped_sheet(ws10, renew_groups, self.daily_table, show_amt_total=False)

                ws11 = wb.create_sheet(title="Global Other Services")
                gos_groups = [
                    ("GCASH OUT",        [("Lotes", ["gcash_out_lotes"],        True),  ("Capital", ["gcash_out"],        False)]),
                    ("MONEYGRAM",        [("Lotes", ["moneygram_lotes"],        True),  ("Capital", ["moneygram"],        False)]),
                    ("TRANSFAST",        [("Lotes", ["transfast_lotes"],        True),  ("Capital", ["transfast"],        False)]),
                    ("RIA",              [("Lotes", ["ria_lotes"],              True),  ("Capital", ["ria"],              False)]),
                    ("SMART MONEY OUT",  [("Lotes", ["smart_money_out_lotes"],  True),  ("Capital", ["smart_money_out"],  False)]),
                    ("GCASH PADALA",     [("Lotes", ["gcash_padala_lotes"],     True),  ("Capital", ["gcash_padala"],     False)]),
                    ("ABRA OUT",         [("Lotes", ["abra_out_lotes"],         True),  ("Capital", ["abra_out"],         False)]),
                    ("REMITLY",          [("Lotes", ["remitly_lotes"],          True),  ("Capital", ["remitly"],          False)]),
                    ("PAL PAY CASH OUT", [("Lotes", ["pal_pay_cash_out_lotes"], True),  ("Capital", ["pal_pay_cash_out"], False)]),
                    ("MC OUT",           [("Lotes", ["mc_out_lotes"],           True),  ("Capital", ["mc_out"],           False)]),
                    ("EC PAY OUT",       [("Capital", ["ec_pay_out"],                       False)]),
                ]
                _write_grouped_sheet(ws11, gos_groups, "daily_reports_brand_a", use_branch_join=True, show_amt_total=False)

                ws12 = wb.create_sheet(title="FT HO")
                ft_ho_groups = [
                    ("FUND TRANSFER HO", [
                        ("FT From Branch", ["fund_transfer_from_branch"],    False),
                        ("FT To HO",       ["fund_transfer_to_head_office"], False),
                        ("FT To Branch",   ["fund_transfer_to_branch"],      False),
                    ]),
                ]
                _write_grouped_sheet(ws12, ft_ho_groups, self.daily_table, show_amt_total=False)

                # ── DEPO BR sheet ────────────────────────────────────────────
                ws13 = wb.create_sheet(title="DEPO BR")

                def _write_depo_br_sheet(ws):
                    _write_info(ws)

                    depo_headers = [
                        "BRANCHES",
                        "FT FROM BRANCH",
                        "CR BRANCH NAME",
                        "FT TO HEAD OFFICE",
                        "REMARKS DEPO",
                        "FT TO BRANCH",
                        "CT BRANCH NAME",
                    ]

                    HDR_ROW  = 7
                    DATA_START = 8

                    for c_idx, h in enumerate(depo_headers, 1):
                        c = ws.cell(row=HDR_ROW, column=c_idx, value=h)
                        c.font = HDR_FONT; c.fill = HDR_FILL; c.border = border
                        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                    # Build query — branches-first so every branch appears
                    reg_where = ""
                    if reg_filter == "registered":
                        reg_where = "AND b.is_registered = 1"
                    elif reg_filter == "not_registered":
                        reg_where = "AND (b.is_registered = 0 OR b.is_registered IS NULL)"

                    try:
                        if filter_type == "os":
                            sql = f"""
                                SELECT b.name AS branch,
                                       COALESCE(dr.fund_transfer_from_branch, 0)        AS ft_from_branch,
                                       COALESCE(dr.fund_transfer_from_branch_dest, '')   AS cr_branch_name,
                                       COALESCE(dr.fund_transfer_to_head_office, 0)      AS ft_to_ho,
                                       dr.fund_transfer_bank_account                     AS raw_bank_id,
                                       dr.ft_ho_breakdown                                AS raw_ft_breakdown,
                                       COALESCE(dr.fund_transfer_to_branch, 0)           AS ft_to_branch,
                                       COALESCE(dr.fund_transfer_to_branch_dest, '')     AS ct_branch_name
                                FROM branches b
                                LEFT JOIN `{self.daily_table}` dr
                                    ON b.name COLLATE utf8mb4_general_ci
                                     = dr.branch COLLATE utf8mb4_general_ci
                                    AND dr.date = %s
                                WHERE b.os_name = %s {reg_where}
                                ORDER BY b.name
                            """
                            sql_params = (selected_date, filter_value)
                        else:
                            sql = f"""
                                SELECT b.name AS branch,
                                       COALESCE(dr.fund_transfer_from_branch, 0)        AS ft_from_branch,
                                       COALESCE(dr.fund_transfer_from_branch_dest, '')   AS cr_branch_name,
                                       COALESCE(dr.fund_transfer_to_head_office, 0)      AS ft_to_ho,
                                       dr.fund_transfer_bank_account                     AS raw_bank_id,
                                       dr.ft_ho_breakdown                                AS raw_ft_breakdown,
                                       COALESCE(dr.fund_transfer_to_branch, 0)           AS ft_to_branch,
                                       COALESCE(dr.fund_transfer_to_branch_dest, '')     AS ct_branch_name
                                FROM branches b
                                INNER JOIN corporations c
                                    ON c.id = COALESCE(b.sub_corporation_id, b.corporation_id)
                                    AND c.name = %s
                                LEFT JOIN `{self.daily_table}` dr
                                    ON b.name COLLATE utf8mb4_general_ci
                                     = dr.branch COLLATE utf8mb4_general_ci
                                    AND dr.date = %s
                                WHERE 1=1 {reg_where}
                                ORDER BY b.name
                            """
                            sql_params = (filter_value, selected_date)

                        results = db_manager.execute_query(sql, sql_params) or []
                    except Exception as ex:
                        ws.cell(row=DATA_START, column=1, value=f"Error loading data: {ex}")
                        return

                    # Bank account ID lookup (same list as BANK_ACCOUNTS on AdminDashboard)
                    _bank_detail = {b['id']: b for b in getattr(self, 'BANK_ACCOUNTS', [])}

                    def _resolve_depo(raw_bank_id, raw_ft_breakdown):
                        _bd = (raw_ft_breakdown or '').strip()
                        if _bd:
                            try:
                                items = json.loads(_bd)
                                parts = []
                                for item in items:
                                    if isinstance(item, dict):
                                        bid = item.get('bank_account_id') or item.get('id')
                                        bdt = _bank_detail.get(int(bid)) if bid else None
                                        if bdt:
                                            parts.append(
                                                f"{bdt['bank_name']} - {bdt['account_name']}"
                                                f" ({bdt.get('account_number', '')})"
                                                f": {item.get('amount', '')}"
                                            )
                                        else:
                                            parts.append(str(item))
                                    elif isinstance(item, (list, tuple)) and len(item) >= 2:
                                        # Format: [display_name, bank_id, amount]
                                        display_name = str(item[0])
                                        amount = item[2] if len(item) >= 3 else ''
                                        parts.append(f"{display_name}: {amount}" if amount != '' else display_name)
                                if parts:
                                    return '; '.join(parts)
                            except Exception:
                                pass
                        if raw_bank_id:
                            try:
                                bdt = _bank_detail.get(int(raw_bank_id))
                            except (ValueError, TypeError):
                                bdt = None
                            if bdt:
                                return f"{bdt['bank_name']} - {bdt['account_name']} ({bdt.get('account_number', '')})"
                            return str(raw_bank_id)
                        return ''

                    total_from = total_ho = total_to = 0.0

                    for r_idx, row in enumerate(results, DATA_START):
                        ft_from  = float(row.get('ft_from_branch', 0) or 0)
                        cr_name  = str(row.get('cr_branch_name', '') or '')
                        ft_ho    = float(row.get('ft_to_ho', 0) or 0)
                        rem_depo = _resolve_depo(row.get('raw_bank_id'), row.get('raw_ft_breakdown'))
                        ft_to    = float(row.get('ft_to_branch', 0) or 0)
                        ct_name  = str(row.get('ct_branch_name', '') or '')

                        total_from += ft_from
                        total_ho   += ft_ho
                        total_to   += ft_to

                        values = [
                            row.get('branch', ''),
                            ft_from,
                            cr_name,
                            ft_ho,
                            rem_depo,
                            ft_to,
                            ct_name,
                        ]
                        for c_idx, val in enumerate(values, 1):
                            cell = ws.cell(row=r_idx, column=c_idx, value=val)
                            cell.border = border
                            if c_idx in (2, 4, 6):
                                cell.number_format = '#,##0.00'
                                cell.alignment = Alignment(horizontal='right')
                            else:
                                cell.alignment = Alignment(horizontal='left')

                    # Totals row
                    tot_row = DATA_START + len(results)
                    tot_vals = ["TOTAL", total_from, "", total_ho, "", total_to, ""]
                    for c_idx, val in enumerate(tot_vals, 1):
                        cell = ws.cell(row=tot_row, column=c_idx, value=val)
                        cell.font = TOTAL_FONT; cell.fill = TOTAL_FILL; cell.border = border
                        if c_idx in (2, 4, 6):
                            cell.number_format = '#,##0.00'
                            cell.alignment = Alignment(horizontal='right')

                    # Column widths
                    ws.column_dimensions['A'].width = 28
                    ws.column_dimensions['B'].width = 18
                    ws.column_dimensions['C'].width = 22
                    ws.column_dimensions['D'].width = 18
                    ws.column_dimensions['E'].width = 22
                    ws.column_dimensions['F'].width = 18
                    ws.column_dimensions['G'].width = 22
                    ws.freeze_panes = 'B8'

                _write_depo_br_sheet(ws13)

            else:
                # ── Brand B: Global Payable and Payable Reports ────────────
                ws6 = wb.create_sheet(title="Global Payable")
                global_payable_groups = [
                    ("SEND OUT", [
                        ("S.O. Lotes",   ["sendout_lotes"],   True),
                        ("S.O. Capital", ["sendout_capital"],  False),
                        ("S.O. SC",      ["sendout_sc"],       False),
                        ("S.O. Comm.",   ["sendout_commission"],False),
                        ("S.O. Total",   ["sendout_total"],    False),
                    ]),
                    ("PAY OUT", [
                        ("P.O. Lotes",   ["payout_lotes"],   True),
                        ("P.O. Capital", ["payout_capital"],  False),
                        ("P.O. SC",      ["payout_sc"],       False),
                        ("P.O. Comm.",   ["payout_commission"],False),
                        ("P.O. Total",   ["payout_total"],    False),
                    ]),
                    ("INTERNATIONAL", [
                        ("Int. Lotes",   ["international_lotes"],   True),
                        ("Int. Capital", ["international_capital"],  False),
                        ("Int. SC",      ["international_sc"],       False),
                        ("Int. Comm.",   ["international_commission"],False),
                        ("Int. Total",   ["international_total"],    False),
                    ]),
                    ("OTHER", [
                        ("SKID",       ["skid"],        False),
                        ("SKIR",       ["skir"],        False),
                        ("Cancel",     ["cancellation"],False),
                        ("P.O. Inc.",  ["inc"],         False),
                    ]),
                ]
                _write_grouped_sheet(ws6, global_payable_groups, "payable_tbl_brand_a", use_branch_join=True, show_amt_total=False)

                ws7 = wb.create_sheet(title="Payable Reports")

                def _write_pepp_report_sheet(ws):
                    from decimal import Decimal, ROUND_HALF_UP

                    def _r2(v):
                        return float(Decimal(str(v)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))

                    def _m2(a, b):
                        return float((Decimal(str(a)) * Decimal(str(b))).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP))

                    _corp_abbrev_map = {
                        'SILVERSTAR JEWELRY PAWNSHOP INC':                'SJPI',
                        'ALEXITE JEWELRY PAWNSHOP INC':                   'AJPI',
                        'SAN RAMON PLATINUM PAWNSHOP INC':                'SRPPI',
                        'HOMENEEDS PAWNSHOP INC':                         'HPI',
                        'KRISTAL CLEAR DIAMOND AND GOLD PAWNSHOP INC':    'KCDGPI',
                        'SAFELOCK PAWNSHOP INC':                          'SPI',
                        'MEGAWORLD DOMESTIC PAWNSHOP INC':                'MDPI',
                        'GLOBAL RELIANCE MANAGEMENT & HOLDINGS CORP.':    'GRMHC',
                    }
                    _registry_map = {
                        'SILVERSTAR JEWELRY PAWNSHOP INC':                'P250682A',
                        'ALEXITE JEWELRY PAWNSHOP INC':                   'P250683A',
                        'SAN RAMON PLATINUM PAWNSHOP INC':                'P250681A',
                        'HOMENEEDS PAWNSHOP INC':                         'P250677A',
                        'KRISTAL CLEAR DIAMOND AND GOLD PAWNSHOP INC':    'P250678A',
                        'SAFELOCK PAWNSHOP INC':                          'P250680A',
                        'MEGAWORLD DOMESTIC PAWNSHOP INC':                'P250679A',
                        'GLOBAL RELIANCE MANAGEMENT & HOLDINGS CORP.':    'P210021A',
                    }

                    corp_name   = filter_value
                    corp_upper  = corp_name.upper().strip().rstrip('.')
                    corp_abbrev = next((v for k, v in _corp_abbrev_map.items() if k.upper().strip().rstrip('.') == corp_upper), corp_name[:4].upper())
                    registry    = next((v for k, v in _registry_map.items()    if k.upper().strip().rstrip('.') == corp_upper), '')
                    is_global   = 'GLOBAL RELIANCE' in corp_upper
                    payable_tbl = "payable_tbl_brand_a"

                    if reg_filter == "registered":
                        reg_clause_payable = "AND b.is_registered = 1"
                    elif reg_filter == "not_registered":
                        reg_clause_payable = "AND (b.is_registered = 0 OR b.is_registered IS NULL)"
                    else:
                        reg_clause_payable = ""

                    try:
                        if filter_type == "os":
                            result = db_manager.execute_query(f"""
                                SELECT SUM(COALESCE(p.sendout_capital, 0))        AS total_sendout_capital,
                                       SUM(COALESCE(p.sendout_commission, 0))     AS total_sendout_commission,
                                       SUM(COALESCE(p.sendout_sc, 0))             AS total_sendout_sc,
                                       SUM(COALESCE(p.payout_capital, 0))         AS total_payout_capital,
                                       SUM(COALESCE(p.payout_commission, 0))      AS total_payout_commission,
                                       SUM(COALESCE(p.payout_sc, 0))              AS total_payout_sc,
                                       SUM(COALESCE(p.international_commission, 0)) AS total_international_commission,
                                       SUM(COALESCE(p.skid, 0))                   AS total_skid,
                                       SUM(COALESCE(p.skir, 0))                   AS total_skir,
                                       SUM(COALESCE(p.cancellation, 0))           AS total_cancellation,
                                       SUM(COALESCE(p.inc, 0))                    AS total_inc
                                FROM {payable_tbl} p
                                INNER JOIN branches b ON p.branch COLLATE utf8mb4_general_ci = b.name COLLATE utf8mb4_general_ci
                                INNER JOIN corporations c ON c.id = COALESCE(b.sub_corporation_id, b.corporation_id)
                                                          AND p.corporation COLLATE utf8mb4_general_ci = c.name COLLATE utf8mb4_general_ci
                                WHERE b.os_name = %s AND p.date = %s {reg_clause_payable}
                            """, (filter_value, selected_date))
                        elif is_global:
                            result = db_manager.execute_query(f"""
                                SELECT SUM(COALESCE(p.sendout_capital, 0))        AS total_sendout_capital,
                                       SUM(COALESCE(p.sendout_commission, 0))     AS total_sendout_commission,
                                       SUM(COALESCE(p.sendout_sc, 0))             AS total_sendout_sc,
                                       SUM(COALESCE(p.payout_capital, 0))         AS total_payout_capital,
                                       SUM(COALESCE(p.payout_commission, 0))      AS total_payout_commission,
                                       SUM(COALESCE(p.payout_sc, 0))              AS total_payout_sc,
                                       SUM(COALESCE(p.international_commission, 0)) AS total_international_commission,
                                       SUM(COALESCE(p.skid, 0))                   AS total_skid,
                                       SUM(COALESCE(p.skir, 0))                   AS total_skir,
                                       SUM(COALESCE(p.cancellation, 0))           AS total_cancellation,
                                       SUM(COALESCE(p.inc, 0))                    AS total_inc
                                FROM {payable_tbl} p
                                INNER JOIN branches b ON p.branch COLLATE utf8mb4_general_ci = b.name COLLATE utf8mb4_general_ci
                                INNER JOIN corporations c ON c.id = COALESCE(b.sub_corporation_id, b.corporation_id)
                                                          AND p.corporation COLLATE utf8mb4_general_ci = c.name COLLATE utf8mb4_general_ci
                                WHERE b.global_tag = 'GLOBAL' AND p.date = %s {reg_clause_payable}
                            """, (selected_date,))
                        else:
                            result = db_manager.execute_query(f"""
                                SELECT SUM(COALESCE(p.sendout_capital, 0))        AS total_sendout_capital,
                                       SUM(COALESCE(p.sendout_commission, 0))     AS total_sendout_commission,
                                       SUM(COALESCE(p.sendout_sc, 0))             AS total_sendout_sc,
                                       SUM(COALESCE(p.payout_capital, 0))         AS total_payout_capital,
                                       SUM(COALESCE(p.payout_commission, 0))      AS total_payout_commission,
                                       SUM(COALESCE(p.payout_sc, 0))              AS total_payout_sc,
                                       SUM(COALESCE(p.international_commission, 0)) AS total_international_commission,
                                       SUM(COALESCE(p.skid, 0))                   AS total_skid,
                                       SUM(COALESCE(p.skir, 0))                   AS total_skir,
                                       SUM(COALESCE(p.cancellation, 0))           AS total_cancellation,
                                       SUM(COALESCE(p.inc, 0))                    AS total_inc
                                FROM {payable_tbl} p
                                INNER JOIN branches b ON p.branch COLLATE utf8mb4_general_ci = b.name COLLATE utf8mb4_general_ci
                                INNER JOIN corporations c ON c.id = COALESCE(b.sub_corporation_id, b.corporation_id)
                                                          AND p.corporation COLLATE utf8mb4_general_ci = c.name COLLATE utf8mb4_general_ci
                                WHERE p.corporation = %s AND p.date = %s {reg_clause_payable}
                            """, (filter_value, selected_date))
                    except Exception as ex:
                        ws['A1'] = f"Error loading data: {ex}"
                        return

                    if not result or not result[0]:
                        ws['A1'] = "No data found."
                        return

                    row_data = result[0]
                    if not isinstance(row_data, dict):
                        _keys = [
                            'total_sendout_capital', 'total_sendout_commission', 'total_sendout_sc',
                            'total_payout_capital',  'total_payout_commission',  'total_payout_sc',
                            'total_international_commission',
                            'total_skid', 'total_skir', 'total_cancellation', 'total_inc',
                        ]
                        row_data = dict(zip(_keys, row_data))

                    if all(v is None for v in row_data.values()):
                        ws['A1'] = "No data found."
                        return

                    sendout_capital          = float(row_data.get('total_sendout_capital')          or 0)
                    sendout_commission       = float(row_data.get('total_sendout_commission')       or 0)
                    sendout_sc               = float(row_data.get('total_sendout_sc')               or 0)
                    payout_capital           = float(row_data.get('total_payout_capital')           or 0)
                    payout_commission        = float(row_data.get('total_payout_commission')        or 0)
                    payout_sc                = float(row_data.get('total_payout_sc')                or 0)
                    international_commission = float(row_data.get('total_international_commission') or 0)
                    total_skid               = float(row_data.get('total_skid')                     or 0)
                    total_skir               = float(row_data.get('total_skir')                     or 0)
                    total_cancellation       = float(row_data.get('total_cancellation')             or 0)
                    total_inc                = float(row_data.get('total_inc')                      or 0)

                    pepp_commission_61       = _m2(sendout_commission, 0.61)
                    skid_61                  = _m2(total_skid, 0.61)
                    corp_commission_43       = _m2(payout_commission, 0.43)
                    corp_international_80    = _m2(international_commission, 0.80)
                    skir_57                  = _m2(total_skir, 0.57)

                    send_subtotal                = _r2(sendout_capital + pepp_commission_61 + sendout_sc)
                    send_subtotal_after_discount = _r2(send_subtotal - skid_61)
                    total_net_send               = _r2(send_subtotal_after_discount - total_cancellation)
                    release_subtotal             = _r2(payout_capital + corp_commission_43 + corp_international_80)
                    release_subtotal_with_inc    = _r2(release_subtotal + total_inc)
                    total_net_released           = _r2(release_subtotal_with_inc + skir_57)
                    net_receivable_payable       = _r2(total_net_send - total_net_released)

                    hdr_font_pr = Font(bold=True, size=12)
                    sec_font_pr = Font(bold=True, size=11)
                    reg_font_pr = Font(size=10)
                    tot_font_pr = Font(bold=True, size=11)
                    hdr_fill_pr = PatternFill("solid", fgColor="E6F3FF")
                    sub_fill_pr = PatternFill("solid", fgColor="F0F0F0")
                    tot_fill_pr = PatternFill("solid", fgColor="D9D9D9")
                    right_al    = Alignment(horizontal='right')
                    center_al   = Alignment(horizontal='center')

                    r = 1
                    ws.merge_cells(f'A{r}:D{r}')
                    ws[f'A{r}'] = f"Palawan Express Pera Padala - {corp_name}"
                    ws[f'A{r}'].font      = hdr_font_pr
                    ws[f'A{r}'].alignment = center_al
                    ws[f'A{r}'].fill      = hdr_fill_pr
                    r += 1

                    ws[f'A{r}'] = f"PEPP Reconciliation for {selected_date}"
                    ws[f'C{r}'] = "Partner Registry No."
                    ws[f'D{r}'] = registry
                    for cell in [ws[f'A{r}'], ws[f'C{r}'], ws[f'D{r}']]:
                        cell.font = hdr_font_pr
                    ws[f'C{r}'].alignment = right_al
                    ws[f'D{r}'].alignment = right_al
                    r += 2

                    rows_pr = [
                        ("Send Transaction",                                                   "", "",                               "",                                    "section"),
                        (f"    PEPP Remittance from {corp_name}",                              "", "P",                              f"{sendout_capital:,.2f}",              "indent"),
                        ("    PEPP share: 61% of commission",                                  "P", f"{sendout_commission:,.2f}",    f"{pepp_commission_61:,.2f}",           "indent"),
                        ("    PEPP share: Service Charge",                                     "", f"{sendout_sc:,.2f}",             f"{sendout_sc:,.2f}",                   "indent"),
                        ("        Subtotal",                                                   "", "",                               f"{send_subtotal:,.2f}",                "subtotal"),
                        ("    Less: Discount (Suki Card)",                                     "", f"({total_skid:,.2f})",           f"({skid_61:,.2f})",                    "indent"),
                        ("        Subtotal",                                                   "", "",                               f"{send_subtotal_after_discount:,.2f}", "subtotal"),
                        ("    Less: Cancellation",                                             "", f"({total_cancellation:,.2f})",   f"({total_cancellation:,.2f})",         "indent"),
                        ("    Total Net Send",                                                 "", "",                               f"{total_net_send:,.2f}",               "total"),
                        ("", "", "", "", "blank"),
                        (f"    RELEASE Transaction (Payable to {corp_abbrev})",                "", "",                               "",                                    "section"),
                        (f"    PEPP Remittances released at {corp_abbrev}",                    "", "P",                              f"{payout_capital:,.2f}",               "indent"),
                        (f"    {corp_abbrev} share: 43% of commission",                        "P", f"{payout_commission:,.2f}",     f"{corp_commission_43:,.2f}",           "indent"),
                        (f"    {corp_abbrev} share: 50% of commission (LBC Domestic Payout)",  "", "",                               "",                                    "indent"),
                        (f"    {corp_abbrev} share: 80% of commission (International Payout)", "", f"{international_commission:,.2f}", f"{corp_international_80:,.2f}",    "indent"),
                        ("    Service Charge",                                                 "", f"{payout_sc:,.2f}",              "-",                                   "indent"),
                        ("        Subtotal",                                                   "", "",                               f"{release_subtotal:,.2f}",             "subtotal"),
                        (f"    Add: {corp_abbrev} Branch Incentives released",                 "", "",                               f"{total_inc:,.2f}",                   "indent"),
                        ("        Subtotal",                                                   "", "",                               f"{release_subtotal_with_inc:,.2f}",    "subtotal"),
                        ("    Add: Rebates (Suki Card)",                                       "", f"{total_skir:,.2f}",             f"{skir_57:,.2f}",                     "indent"),
                        ("    Total Net Released",                                             "", "",                               f"{total_net_released:,.2f}",           "total"),
                        ("", "", "", "", "blank"),
                        ("    Net Send",                   "", "", f"{total_net_send:,.2f}",         "regular"),
                        ("    Less : Net Released",        "", "", f"{total_net_released:,.2f}",     "regular"),
                        ("    Net Receivable / (Payable)", "", "", f"{net_receivable_payable:,.2f}", "total"),
                    ]

                    for col1, col2, col3, col4, row_type in rows_pr:
                        ws[f'A{r}'] = col1
                        ws[f'B{r}'] = col2
                        ws[f'C{r}'] = col3
                        ws[f'D{r}'] = col4

                        if row_type == "section":
                            for c in ['A', 'B', 'C', 'D']:
                                ws[f'{c}{r}'].font = sec_font_pr
                        elif row_type == "total":
                            for c in ['A', 'B', 'C', 'D']:
                                ws[f'{c}{r}'].font = tot_font_pr
                                ws[f'{c}{r}'].fill = tot_fill_pr
                        elif row_type == "subtotal":
                            for c in ['A', 'B', 'C', 'D']:
                                ws[f'{c}{r}'].font = tot_font_pr
                                ws[f'{c}{r}'].fill = sub_fill_pr
                        else:
                            for c in ['A', 'B', 'C', 'D']:
                                ws[f'{c}{r}'].font = reg_font_pr

                        ws[f'C{r}'].alignment = right_al
                        ws[f'D{r}'].alignment = right_al
                        r += 1

                    r += 2
                    ws[f'A{r}'] = "Prepared by:"
                    ws[f'C{r}'] = "Noted by:"
                    r += 2
                    ws[f'A{r}'] = "Rochelle G. Serrano"
                    ws[f'C{r}'] = "Aimee M. Martinez"

                    ws.column_dimensions['A'].width = 55
                    ws.column_dimensions['B'].width = 8
                    ws.column_dimensions['C'].width = 18
                    ws.column_dimensions['D'].width = 18

                _write_pepp_report_sheet(ws7)

            # ── Save ─────────────────────────────────────────────────────────
            wb.save(file_path)
            dialog.accept()
            QMessageBox.information(
                self, "Export Successful",
                f"Full Brand Report exported to:\n{file_path}\n\n"
                f"{filter_label}: {filter_value}\n"
                f"Date: {selected_date}\n"
                f"Sheets generated: {len(wb.sheetnames)}"
            )

        except Exception as e:
            import traceback
            QMessageBox.critical(
                self, "Export Error",
                f"Error exporting: {str(e)}\n\n{traceback.format_exc()[:600]}"
            )

    def load_entry_by_date(self):
        filter_type = self.filter_type_selector.currentData()
        branch_name = self.branch_selector.currentText()
        selected_date = self.date_picker.date().toString("yyyy-MM-dd")

        if not branch_name:
            QMessageBox.warning(self, "Missing Selection", "Please select a branch.")
            return

        try:
            where_clauses = ["branch = %s", "date = %s"]
            query_params = [branch_name, selected_date]

            if filter_type == "corporation":
                selected_corporation = self.corp_selector.currentText().strip()
                if selected_corporation:
                    where_clauses.append("corporation = %s")
                    query_params.append(selected_corporation)

            query = f"""
                SELECT *
                FROM {self.daily_table}
                WHERE {' AND '.join(where_clauses)}
                ORDER BY id DESC
                LIMIT 1
            """
            result = self.db.execute_query(query, query_params)

            if not result:
                QMessageBox.information(self, "No Entry", f"No entry found for {selected_date}.")
                self.clear_all_fields()
                return

            data = result[0]
            self._current_entry_data = data  

            beginning = float(data.get('beginning_balance') or 0)
            self.beginning_balance_input.setText(f"{beginning:.2f}")
            self.cash_count_input.setText(f"{float(data.get('cash_count') or 0):.2f}")

            for ui_label, db_column in self.debit_fields.items():
                if db_column in data and data[db_column] is not None:
                    self.debit_inputs[ui_label].setText(str(data[db_column]))
                else:
                    self.debit_inputs[ui_label].setText("0.00")

                lotes_col = db_column + "_lotes"
                if lotes_col in data and data[lotes_col] is not None:
                    self.debit_lotes_inputs[ui_label].setText(str(data[lotes_col]))
                else:
                    self.debit_lotes_inputs[ui_label].setText("0")

            for ui_label, db_column in self.credit_fields.items():
                if db_column in data and data[db_column] is not None:
                    self.credit_inputs[ui_label].setText(str(data[db_column]))
                else:
                    self.credit_inputs[ui_label].setText("0.00")

                lotes_col = db_column + "_lotes"
                if lotes_col in data and data[lotes_col] is not None:
                    self.credit_lotes_inputs[ui_label].setText(str(data[lotes_col]))
                else:
                    self.credit_lotes_inputs[ui_label].setText("0")

            # Recalculate totals from actual field values (fixes stale stored values)
            debit_sum = sum(
                float(self.debit_inputs[lbl].text().strip() or 0)
                for lbl in self.debit_fields
            )
            credit_sum = sum(
                float(self.credit_inputs[lbl].text().strip() or 0)
                for lbl in self.credit_fields
            )
            debit_total = beginning + debit_sum
            credit_total = credit_sum
            ending_balance = debit_total - credit_total
            cash_count = float(data.get('cash_count') or 0)
            cash_result = cash_count - ending_balance

            self.debit_total_display.setText(f"{debit_total:.2f}")
            self.credit_total_display.setText(f"{credit_total:.2f}")
            self.ending_balance_display.setText(f"{ending_balance:.2f}")
            self.cash_result_display.setText(f"{cash_result:.2f}")

            variance_status = data.get('variance_status', '')
            if abs(cash_result) < 0.01:
                variance_status = 'balanced'
            elif cash_result > 0:
                variance_status = 'over'
            else:
                variance_status = 'short'
            if variance_status == 'short':
                self.variance_status_display.setText("SHORT")
                self.variance_status_display.setStyleSheet(
                    "font-weight: bold; font-size: 12px; padding: 5px 10px; border-radius: 4px; background-color: #ffcdd2; color: #c62828;"
                )
            elif variance_status == 'over':
                self.variance_status_display.setText("OVER")
                self.variance_status_display.setStyleSheet(
                    "font-weight: bold; font-size: 12px; padding: 5px 10px; border-radius: 4px; background-color: #fff3cd; color: #856404;"
                )
            else:
                self.variance_status_display.setText("✓ Balanced")
                self.variance_status_display.setStyleSheet(
                    "font-weight: bold; font-size: 12px; padding: 5px 10px; border-radius: 4px; background-color: #c8e6c9; color: #2e7d32;"
                )

            self.selected_bank_account = data.get('fund_transfer_bank_account')
            if self.bank_account_btn:

                ft_ho_bd = data.get('ft_ho_breakdown')
                if ft_ho_bd:
                    try:
                        bd_list = json.loads(ft_ho_bd)
                        self.bank_account_btn.setText(f"🏦 View ({len(bd_list)})")
                        self.bank_account_btn.setToolTip(f"View {len(bd_list)} fund transfer(s) breakdown")
                    except Exception:
                        self.bank_account_btn.setText("View")
                        self.bank_account_btn.setToolTip("View fund transfer breakdown")
                elif self.selected_bank_account:

                    bank_name = "View"
                    for bank in self.BANK_ACCOUNTS:
                        if bank['id'] == self.selected_bank_account:
                            bank_name = bank['bank_name']
                            break
                    self.bank_account_btn.setText(f"🏦 {bank_name}")
                    self.bank_account_btn.setToolTip("View selected bank account")
                else:
                    self.bank_account_btn.setText("View")
                    self.bank_account_btn.setToolTip("No bank account selected")


            self.selected_branch_dest = data.get('fund_transfer_to_branch_dest')
            if self.branch_dest_btn:
                self.branch_dest_btn.setText("View")
                if self.selected_branch_dest:
                    self.branch_dest_btn.setToolTip(f"Destination: {self.selected_branch_dest}")
                else:
                    self.branch_dest_btn.setToolTip("No destination branch specified")

       
            self.selected_from_branch_dest = data.get('fund_transfer_from_branch_dest')
            if self.from_branch_dest_btn:
                self.from_branch_dest_btn.setText("View")
                if self.selected_from_branch_dest:
                    self.from_branch_dest_btn.setToolTip(f"Source: {self.selected_from_branch_dest}")
                else:
                    self.from_branch_dest_btn.setToolTip("No source branch specified")

            self.current_record_id = data.get('id')

            # Check review status
            brand_key = "A" if self.account_type == 1 else "B"
            try:
                review_row = self.db.execute_query(
                    "SELECT id FROM admin_review_marks WHERE brand = %s AND branch = %s AND report_date = %s",
                    [brand_key, branch_name, selected_date]
                )
                is_reviewed = bool(review_row)
                self.reviewed_checkbox.blockSignals(True)
                self.reviewed_checkbox.setChecked(is_reviewed)
                self.reviewed_checkbox.setText("✅ Reviewed" if is_reviewed else "Pending review")
                self.reviewed_checkbox.setStyleSheet(self.reviewed_checkbox.styleSheet().replace(
                    "color: #c0392b;" if is_reviewed else "color: #2e7d32;",
                    "color: #2e7d32;" if is_reviewed else "color: #c0392b;"
                ))
                self.reviewed_checkbox.blockSignals(False)
                self.reviewed_checkbox.setEnabled(True)
            except Exception:
                self.reviewed_checkbox.setEnabled(False)

            # Load palawan details into collapsible
            self._load_palawan_details(data)

            QMessageBox.information(self, "✅ Loaded", f"Entry for {selected_date} loaded successfully!")

        except Exception as e:
            logger.error("Error loading entry: %s", e)
            QMessageBox.critical(self, "Error", f"Failed to load entry: {e}")

    def _on_review_toggled(self, checked):
        """Save or remove the review mark for the current entry."""
        branch_name = self.branch_selector.currentText()
        selected_date = self.date_picker.date().toString("yyyy-MM-dd")
        brand_key = "A" if self.account_type == 1 else "B"
        if not branch_name or not selected_date:
            return
        try:
            if checked:
                self.db.execute_query(
                    "INSERT IGNORE INTO admin_review_marks (brand, branch, report_date) VALUES (%s, %s, %s)",
                    [brand_key, branch_name, selected_date]
                )
                self.reviewed_checkbox.setText("Reviewed")
                self.reviewed_checkbox.setStyleSheet(self.reviewed_checkbox.styleSheet().replace(
                    "color: #c0392b;", "color: #2e7d32;"
                ))
            else:
                self.db.execute_query(
                    "DELETE FROM admin_review_marks WHERE brand = %s AND branch = %s AND report_date = %s",
                    [brand_key, branch_name, selected_date]
                )
                self.reviewed_checkbox.setText("Pending review")
                self.reviewed_checkbox.setStyleSheet(self.reviewed_checkbox.styleSheet().replace(
                    "color: #2e7d32;", "color: #c0392b;"
                ))
        except Exception as e:
            logger.error("Error toggling review mark: %s", e)

    def get_current_entry_data(self):
        """Return the current loaded entry data for breakdown views"""
        return getattr(self, '_current_entry_data', None)

    def show_mc_breakdown(self, field_type):
        """Show MC currency breakdown in a dialog"""
        entry = self.get_current_entry_data()
        if not entry:
            QMessageBox.information(self, "No Entry Loaded", 
                "Please load an entry first to view MC breakdown.")
            return
        
        # Determine which details field to use
        details_key = "mc_in_details" if field_type == "MC In" else "mc_out_details"
        breakdown = []
        
        if entry.get(details_key):
            try:
                breakdown = json.loads(entry[details_key])
            except Exception:
                breakdown = []
        
        if not breakdown:
            QMessageBox.information(self, "No Breakdown", 
                f"No {field_type} currency breakdown available for this entry.")
            return
        
        # Create breakdown dialog
        dialog = QDialog(self)
        dialog.setWindowTitle(f"{field_type} Currency Breakdown (View Only)")
        dialog.setMinimumWidth(600)
        dialog.setMinimumHeight(300)
        
        layout = QVBoxLayout(dialog)
        
        # Header
        header_text = "SELLING Currency (Money In)" if field_type == "MC In" else "BUYING Currency (Money Out)"
        header_color = "#22C55E" if field_type == "MC In" else "#DC2626"
        
        header = QLabel(f"💱 {header_text}")
        header.setStyleSheet(f"font-size: 14px; font-weight: 800; color: {header_color}; padding: 10px;")
        layout.addWidget(header)
        
        # Table for breakdown
        table = QTableWidget()
        table.setColumnCount(5)
        table.setHorizontalHeaderLabels(["Currency", "Pcs", "Denomination", "Rate", "Total"])
        table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        table.setEditTriggers(QTableWidget.NoEditTriggers)
        
        total = 0.0
        for entry_data in breakdown:
            row = table.rowCount()
            table.insertRow(row)
            
            currency = entry_data.get('currency', 'Unknown')
            qty = entry_data.get('quantity', 0)
            denom = entry_data.get('denomination', 0.0)
            rate = entry_data.get('rate', 0.0)
            total_php = entry_data.get('total_php', 0)
            if total_php == 0 and qty > 0 and rate > 0:
                if denom > 0:
                    total_php = qty * denom * rate
                else:
                    total_php = qty * rate
            total += total_php
            
            table.setItem(row, 0, QTableWidgetItem(str(currency)))
            table.setItem(row, 1, QTableWidgetItem(str(qty)))
            table.setItem(row, 2, QTableWidgetItem(f"{denom:,.2f}" if denom else "-"))
            table.setItem(row, 3, QTableWidgetItem(f"{rate:,.2f}"))
            table.setItem(row, 4, QTableWidgetItem(f"₱{total_php:,.2f}"))
        
        layout.addWidget(table)
        

        total_label = QLabel(f"TOTAL: ₱{total:,.2f}")
        total_label.setStyleSheet(f"font-size: 16px; font-weight: 800; color: {header_color}; padding: 10px;")
        total_label.setAlignment(Qt.AlignRight)
        layout.addWidget(total_label)
        

        close_btn = QPushButton("Close")
        close_btn.setStyleSheet("""
            QPushButton {
                background-color: #6B7280; color: white;
                padding: 10px 24px; border-radius: 6px;
                font-weight: 600;
            }
            QPushButton:hover { background-color: #4B5563; }
        """)
        close_btn.clicked.connect(dialog.accept)
        layout.addWidget(close_btn)
        
        dialog.exec_()

    def clear_all_fields(self):
        self._current_entry_data = None
        self.beginning_balance_input.setText("0.00")
        self.cash_count_input.setText("0.00")
        self.ending_balance_display.setText("0.00")
        self.cash_result_display.setText("0.00")
        self.debit_total_display.setText("0.00")
        self.credit_total_display.setText("0.00")

        for input_field in self.debit_inputs.values():
            input_field.setText("0.00")
        for input_field in self.credit_inputs.values():
            input_field.setText("0.00")
        for input_field in self.debit_lotes_inputs.values():
            input_field.setText("0")
        for input_field in self.credit_lotes_inputs.values():
            input_field.setText("0")
        
        # Reset bank account selection
        self.selected_bank_account = None
        if self.bank_account_btn:
            self.bank_account_btn.setText("View")
            self.bank_account_btn.setToolTip("No bank account selected")
        
        # Reset branch destination selection
        self.selected_branch_dest = None
        if self.branch_dest_btn:
            self.branch_dest_btn.setText("View")
            self.branch_dest_btn.setToolTip("No destination branch specified")
        
        # Reset from branch source selection
        self.selected_from_branch_dest = None
        if self.from_branch_dest_btn:
            self.from_branch_dest_btn.setText("View")
            self.from_branch_dest_btn.setToolTip("No source branch specified")
        
        # Reset variance status display
        self.variance_status_display.setText("—")
        self.variance_status_display.setStyleSheet(
            "font-weight: bold; font-size: 12px; padding: 5px 10px; border-radius: 4px;"
        )

        # Reset reviewed checkbox
        self.reviewed_checkbox.blockSignals(True)
        self.reviewed_checkbox.setChecked(False)
        self.reviewed_checkbox.setText("Pending review")
        self.reviewed_checkbox.setStyleSheet(self.reviewed_checkbox.styleSheet().replace(
            "color: #2e7d32;", "color: #c0392b;"
        ))
        self.reviewed_checkbox.blockSignals(False)
        self.reviewed_checkbox.setEnabled(False)

    def _recalc_totals(self):
        """Recalculate and display Total Cash Receipt, Total Cash Out,
        Ending Balance, Cash Result and Variance live as the admin edits fields."""
        try:
            beginning = float(self.beginning_balance_input.text().strip().replace(',', '') or 0)
        except ValueError:
            beginning = 0.0
        try:
            cash_count = float(self.cash_count_input.text().strip().replace(',', '') or 0)
        except ValueError:
            cash_count = 0.0

        debit_sum = 0.0
        for inp in self.debit_inputs.values():
            try:
                debit_sum += float(inp.text().strip().replace(',', '') or 0)
            except ValueError:
                pass

        credit_sum = 0.0
        for inp in self.credit_inputs.values():
            try:
                credit_sum += float(inp.text().strip().replace(',', '') or 0)
            except ValueError:
                pass

        debit_total    = beginning + debit_sum
        credit_total   = credit_sum
        ending_balance = debit_total - credit_total
        cash_result    = cash_count - ending_balance

        self.debit_total_display.setText(f"{debit_total:.2f}")
        self.credit_total_display.setText(f"{credit_total:.2f}")
        self.ending_balance_display.setText(f"{ending_balance:.2f}")
        self.cash_result_display.setText(f"{cash_result:.2f}")

        if abs(cash_result) < 0.01:
            self.variance_status_display.setText("✓ Balanced")
            self.variance_status_display.setStyleSheet(
                "font-weight: bold; font-size: 12px; padding: 5px 10px; border-radius: 4px;"
                " background-color: #c8e6c9; color: #2e7d32;"
            )
        elif cash_result > 0:
            self.variance_status_display.setText("OVER")
            self.variance_status_display.setStyleSheet(
                "font-weight: bold; font-size: 12px; padding: 5px 10px; border-radius: 4px;"
                " background-color: #fff3cd; color: #856404;"
            )
        else:
            self.variance_status_display.setText("SHORT")
            self.variance_status_display.setStyleSheet(
                "font-weight: bold; font-size: 12px; padding: 5px 10px; border-radius: 4px;"
                " background-color: #ffcdd2; color: #c62828;"
            )

    def save_entry(self):
        """Save edited entry to the database"""
        if not hasattr(self, 'current_record_id') or not self.current_record_id:
            QMessageBox.warning(self, "No Entry Loaded", "Please load an entry first before saving.")
            return

        branch_name = self.branch_selector.currentText()
        selected_date = self.date_picker.date().toString("yyyy-MM-dd")

        if not branch_name:
            QMessageBox.warning(self, "Selection Required", "Please select a branch.")
            return

        reply = QMessageBox.question(
            self,
            "Confirm Save",
            f"Are you sure you want to save changes for:\n\n"
            f"Branch: {branch_name}\n"
            f"Date: {selected_date}\n\n"
            f"This will update the existing entry.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )

        if reply == QMessageBox.No:
            return

        try:

            update_data = {}
            debit_sum = 0
            for ui_label, db_column in self.debit_fields.items():
                try:
                    val = float(self.debit_inputs[ui_label].text().strip() or 0)
                except ValueError:
                    val = 0
                update_data[db_column] = val
                debit_sum += val
                

                lotes_col = db_column + "_lotes"
                try:
                    lotes_val = int(self.debit_lotes_inputs[ui_label].text().strip() or 0)
                except ValueError:
                    lotes_val = 0
                update_data[lotes_col] = lotes_val

            # Collect all credit values
            credit_sum = 0
            for ui_label, db_column in self.credit_fields.items():
                try:
                    val = float(self.credit_inputs[ui_label].text().strip() or 0)
                except ValueError:
                    val = 0
                update_data[db_column] = val
                credit_sum += val
                
                # Lotes
                lotes_col = db_column + "_lotes"
                try:
                    lotes_val = int(self.credit_lotes_inputs[ui_label].text().strip() or 0)
                except ValueError:
                    lotes_val = 0
                update_data[lotes_col] = lotes_val

            # Calculate totals
            try:
                beginning = float(self.beginning_balance_input.text().strip() or 0)
            except ValueError:
                beginning = 0
            
            try:
                cash_count = float(self.cash_count_input.text().strip() or 0)
            except ValueError:
                cash_count = 0

            debit_total = beginning + debit_sum
            credit_total = credit_sum
            ending_balance = debit_total - credit_total
            cash_result = cash_count - ending_balance


            if abs(cash_result) < 0.01:
                variance_status = "balanced"
            elif cash_result > 0:
                variance_status = "over"
            else:
                variance_status = "short"

            update_data['beginning_balance'] = beginning
            update_data['cash_count'] = cash_count
            update_data['debit_total'] = debit_total
            update_data['credit_total'] = credit_total
            update_data['ending_balance'] = ending_balance
            update_data['cash_result'] = cash_result
            update_data['variance_status'] = variance_status

            # Merge palawan detail values
            # Brand A: sendout/payout/international go to payable_tbl_brand_a
            # (those columns don't exist in daily_reports_brand_a)
            _PAYABLE_SECTION_COLS = {
                'palawan_sendout_principal', 'palawan_sendout_sc', 'palawan_sendout_commission',
                'palawan_sendout_lotes_total', 'palawan_sendout_regular_total',
                'palawan_payout_principal', 'palawan_payout_sc', 'palawan_payout_commission',
                'palawan_payout_lotes_total', 'palawan_payout_regular_total',
                'palawan_international_principal', 'palawan_international_sc',
                'palawan_international_commission', 'palawan_international_lotes_total',
                'palawan_international_regular_total',
                # adjustments — stored in payable_tbl_brand_a as skid/skir/cancellation/inc
                'palawan_suki_discounts', 'palawan_suki_rebates',
                'palawan_cancel', 'palawan_pay_out_incentives',
            }
            payable_a_vals = {}
            for db_col, widget in getattr(self, 'palawan_inputs', {}).items():
                if widget.isReadOnly():
                    continue  
                raw_text = widget.text().strip()
                if self.account_type == 1 and db_col in _PAYABLE_SECTION_COLS and raw_text == "":
                    # Keep existing payable values when an admin leaves a field blank.
                    continue
                try:
                    val = float(raw_text or 0)
                except ValueError:
                    val = 0
                if self.account_type == 1 and db_col in _PAYABLE_SECTION_COLS:
                    payable_a_vals[db_col] = val  # goes to payable_tbl_brand_a
                else:
                    update_data[db_col] = val
            # Also collect auto-calc totals
            for section in ("sendout", "payout", "international"):
                disp = getattr(self, 'palawan_total_displays', {}).get(section)
                if disp:
                    section_input_keys = (
                        f"palawan_{section}_principal",
                        f"palawan_{section}_sc",
                        f"palawan_{section}_commission",
                        f"palawan_{section}_lotes_total",
                    )
                    has_section_input = any(
                        getattr(self, 'palawan_inputs', {}).get(k, QLineEdit()).text().strip() != ""
                        for k in section_input_keys
                    )
                    try:
                        total_val = float(disp.text() or 0)
                    except ValueError:
                        total_val = 0
                    col = f"palawan_{section}_regular_total"
                    if self.account_type == 1 and has_section_input:
                        payable_a_vals[col] = total_val
                    elif self.account_type != 1:
                        update_data[col] = total_val

            set_clauses = []
            values = []
            for col, val in update_data.items():
                set_clauses.append(f"`{col}` = %s")
                values.append(val)

            # Use the primary key (id) so we update exactly the one record loaded,
            # regardless of how many corporation rows share the same branch+date.
            values.append(self.current_record_id)

            update_query = f"""
                UPDATE {self.daily_table}
                SET {', '.join(set_clauses)}
                WHERE id = %s
            """
            
            rows_affected = self.db.execute_query(update_query, values)

          
            if self.account_type == 1 and payable_a_vals:
                _col_map = {
                    'palawan_sendout_principal':           'sendout_capital',
                    'palawan_sendout_sc':                  'sendout_sc',
                    'palawan_sendout_commission':          'sendout_commission',
                    'palawan_sendout_lotes_total':         'sendout_lotes',
                    'palawan_sendout_regular_total':       'sendout_total',
                    'palawan_payout_principal':            'payout_capital',
                    'palawan_payout_sc':                   'payout_sc',
                    'palawan_payout_commission':           'payout_commission',
                    'palawan_payout_lotes_total':          'payout_lotes',
                    'palawan_payout_regular_total':        'payout_total',
                    'palawan_international_principal':     'international_capital',
                    'palawan_international_sc':            'international_sc',
                    'palawan_international_commission':    'international_commission',
                    'palawan_international_lotes_total':   'international_lotes',
                    'palawan_international_regular_total': 'international_total',
                    'palawan_suki_discounts':              'skid',
                    'palawan_suki_rebates':                'skir',
                    'palawan_cancel':                      'cancellation',
                    'palawan_pay_out_incentives':          'inc',
                }
                entry_data   = getattr(self, '_current_entry_data', {}) or {}
                corporation  = (entry_data.get('corporation') or "").strip()
                if not corporation and getattr(self, 'current_record_id', None):
                    try:
                        rec = self.db.execute_query(
                            f"SELECT corporation FROM {self.daily_table} WHERE id=%s LIMIT 1",
                            [self.current_record_id]
                        )
                        if rec and rec[0].get('corporation'):
                            corporation = str(rec[0].get('corporation')).strip()
                    except Exception as ce:
                        logger.error("_save resolve corporation by id: %s", ce)
                if not corporation:
                    corporation = self.corp_selector.currentText().strip()
                branch_name  = self.branch_selector.currentText()
                p_cols        = {_col_map[k]: v for k, v in payable_a_vals.items() if k in _col_map}
                if p_cols:
                    col_names    = ', '.join(p_cols.keys())
                    set_parts    = ', '.join(f"{c}=VALUES({c})" for c in p_cols.keys())
                    placeholders = ', '.join(['%s'] * len(p_cols))
                    upsert_sql   = (
                        "INSERT INTO payable_tbl_brand_a "
                        "(corporation, branch, date, " + col_names + ") "
                        "VALUES (%s, %s, %s, " + placeholders + ") "
                        "ON DUPLICATE KEY UPDATE " + set_parts + ", updated_at=CURRENT_TIMESTAMP"
                    )
                    try:
                        self.db.execute_query(
                            upsert_sql,
                            [corporation, branch_name, selected_date] + list(p_cols.values())
                        )
                    except Exception as pe:
                        logger.error("_save payable_tbl_brand_a upsert: %s", pe)

            if rows_affected is not None and rows_affected > 0:

                self.debit_total_display.setText(f"{debit_total:.2f}")
                self.credit_total_display.setText(f"{credit_total:.2f}")
                self.ending_balance_display.setText(f"{ending_balance:.2f}")
                self.cash_result_display.setText(f"{cash_result:.2f}")
                
                if variance_status == 'short':
                    self.variance_status_display.setText("SHORT")
                    self.variance_status_display.setStyleSheet(
                        "font-weight: bold; font-size: 12px; padding: 5px 10px; border-radius: 4px; background-color: #ffcdd2; color: #c62828;"
                    )
                elif variance_status == 'over':
                    self.variance_status_display.setText("OVER")
                    self.variance_status_display.setStyleSheet(
                        "font-weight: bold; font-size: 12px; padding: 5px 10px; border-radius: 4px; background-color: #fff3cd; color: #856404;"
                    )
                else:
                    self.variance_status_display.setText("✓ Balanced")
                    self.variance_status_display.setStyleSheet(
                        "font-weight: bold; font-size: 12px; padding: 5px 10px; border-radius: 4px; background-color: #c8e6c9; color: #2e7d32;"
                    )
                
           

                QMessageBox.information(
                    self,
                    "✅ Saved",
                    f"Entry for {selected_date} has been updated successfully!"
                )
            else:
                QMessageBox.warning(
                    self,
                    "No Changes",
                    f"No entry was updated. The entry may not exist."
                )

        except Exception as e:
            logger.error("Error saving entry: %s", e)
            QMessageBox.critical(self, "Error", f"Failed to save entry: {e}")

    def _update_supplementary_tables(self, selected_date, all_vals: dict):
     
        logger.info("_update_supplementary_tables() called but skipped - data is in canonical tables only")
        pass

    def closeEvent(self, event):
       
        event.accept()


# Optional: to run the widget directly for testing
if __name__ == "__main__":
    import sys
    from PyQt5.QtWidgets import QApplication

    app = QApplication(sys.argv)
    window = AdminDashboard()
    window.show()
    sys.exit(app.exec_())