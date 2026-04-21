from PyQt5.QtWidgets import (
    QWidget, QLabel, QVBoxLayout, QComboBox, QTableWidget, QTableWidgetItem,
    QDateEdit, QMessageBox, QHeaderView, QSizePolicy, QPushButton, QHBoxLayout,
    QApplication, QDesktopWidget, QScrollArea, QFrame, QFileDialog
)
from PyQt5.QtCore import Qt, QDate, QTimer
from PyQt5.QtGui import QFont, QColor, QBrush, QTextDocument, QDoubleValidator, QPainter
from PyQt5.QtPrintSupport import QPrinter, QPrintDialog
from db_connect_pooled import db_manager
from date_range_widget import DateRangeWidget
import datetime


class ColoredHeaderView(QHeaderView):
    """Custom header view with colored sections for each column group"""
    
    def __init__(self, orientation, parent=None):
        super().__init__(orientation, parent)
        self.colors = {}
        self.setFont(QFont("", 9, QFont.Bold))
    
    def paintSection(self, painter, rect, logicalIndex):
        painter.save()
        
        # Get color for this section
        if logicalIndex in self.colors:
            painter.fillRect(rect, self.colors[logicalIndex])
        else:
            painter.fillRect(rect, QColor("#f5f5f5"))
        
        # Draw border
        pen = painter.pen()
        pen.setColor(QColor("#444"))
        painter.setPen(pen)
        painter.drawRect(rect.adjusted(0, 0, -1, -1))
        
        # Draw text
        if logicalIndex in self.colors:
            painter.setPen(QColor("white"))
        else:
            painter.setPen(QColor("#333"))
        
        font = QFont("", 9, QFont.Bold)
        painter.setFont(font)
        
        text = self.model().headerData(logicalIndex, self.orientation(), Qt.DisplayRole)
        painter.drawText(rect, Qt.AlignCenter, str(text) if text else "")
        painter.restore()


class PayablesPage(QWidget):
    def __init__(self, account_type=2):
        super().__init__()
        self.account_type = account_type
        # Set correct table based on brand: Brand A -> daily_reports_brand_a, Brand B -> daily_reports
        self.daily_table = "daily_reports_brand_a" if account_type == 1 else "daily_reports"
        # Set correct payable table based on brand
        self.payable_table = "payable_tbl_brand_a" if account_type == 1 else "payable_tbl"
        self.setWindowTitle("Palawan Transactions - Detailed View")

        self._is_loading = False  # ← FLAG: prevents auto-save during table population

        # Main layout directly on the widget (no intermediate scroll area)
        self.main_layout = QVBoxLayout(self)
        self.main_layout.setContentsMargins(8, 6, 8, 6)
        self.main_layout.setSpacing(4)

        self.create_controls()
        self.create_table()
        self.create_buttons()

        self.load_corporations()
        self.verify_table_structure()

        self.installEventFilter(self)

    # ─────────────────────────────────────────────────────────────────────────
    def verify_table_structure(self):
        """Ensures payable table has the UNIQUE constraint for ON DUPLICATE KEY UPDATE."""
        constraint_name = "uq_corp_branch_date_a" if self.account_type == 1 else "uq_corp_branch_date"
        try:
            result = db_manager.execute_query("""
                SELECT COUNT(*) as cnt
                FROM information_schema.TABLE_CONSTRAINTS
                WHERE TABLE_SCHEMA    = DATABASE()
                  AND TABLE_NAME      = %s
                  AND CONSTRAINT_TYPE = 'UNIQUE'
                  AND CONSTRAINT_NAME = %s
            """, (self.payable_table, constraint_name))
            exists = result[0]['cnt'] > 0 if result else False
            if not exists:
                db_manager.execute_query(f"""
                    ALTER TABLE {self.payable_table}
                    ADD CONSTRAINT {constraint_name}
                    UNIQUE (corporation, branch, date)
                """)
                print(f"✅ Added unique constraint to {self.payable_table}")
            else:
                print(f"✅ Unique constraint already present on {self.payable_table}")
        except Exception as e:
            print(f"⚠️  verify_table_structure: {e}")

    # ─────────────────────────────────────────────────────────────────────────
    def populate_table(self):
        """Load data into table. _is_loading flag blocks auto-save during fill."""
        corp          = self.corp_selector.currentText()
        date_start, date_end = self.date_range_widget.get_date_range()
        is_range      = self.date_range_widget.is_range_mode()
        reg_filter    = self.reg_filter_selector.currentData()
        os_filter     = self.os_filter_selector.currentData()  # None or OS name
        category_filter = self.category_filter_selector.currentData() if hasattr(self, 'category_filter_selector') else "all"

        # 60% corporations list
        SIXTY_PERCENT_CORPS = [
            "Allexite Jewelry Pawnshop Inc.",
            "Homeneeds Pawnshop Inc.",
            "Kristal Clear Diamond and Gold Pawnshop Inc.",
            "Safelock Pawnshop Inc.",
            "Megaworld Domestic Pawnshop Inc.",
            "San Ramon Platinum Pawnshop Inc."
        ]

        # For Brand A admin, don't require corporation - only need OS or category filter
        if self.account_type == 1:
            # Brand A admin: don't filter by corporation, only by Group/Category
            if not os_filter and category_filter == "all":
                return  # Need at least Group or Category filter
        else:
            # Non-Brand A: Need corporation OR OS selected
            if not corp and not os_filter:
                return

        # ── SET FLAG: stops on_item_changed from saving while we fill the table ──
        self._is_loading = True

        # ── Disable UI updates during population for 600+ row performance ──
        self.table.setUpdatesEnabled(False)
        self.table.blockSignals(True)

        try:
            if is_range:
                select_cols = f"""
                    SELECT b.name as branch,
                           SUM(COALESCE(dr.palawan_sendout_lotes_total, 0))         AS so_lotes,
                           SUM(COALESCE(dr.palawan_sendout_principal, 0))           AS so_capital,
                           SUM(COALESCE(dr.palawan_sendout_sc, 0))                  AS so_sc,
                           SUM(COALESCE(dr.palawan_sendout_commission, 0))          AS so_commission,
                           SUM(COALESCE(dr.palawan_sendout_regular_total, 0))       AS so_total,
                           SUM(COALESCE(dr.palawan_payout_lotes_total, 0))          AS po_lotes,
                           SUM(COALESCE(dr.palawan_payout_principal, 0))            AS po_capital,
                           SUM(COALESCE(dr.palawan_payout_sc, 0))                   AS po_sc,
                           SUM(COALESCE(dr.palawan_payout_commission, 0))           AS po_commission,
                           SUM(COALESCE(dr.palawan_payout_regular_total, 0))        AS po_total,
                           SUM(COALESCE(dr.palawan_international_lotes_total, 0))   AS int_lotes,
                           SUM(COALESCE(dr.palawan_international_principal, 0))     AS int_capital,
                           SUM(COALESCE(dr.palawan_international_sc, 0))            AS int_sc,
                           SUM(COALESCE(dr.palawan_international_commission, 0))    AS int_commission,
                           SUM(COALESCE(dr.palawan_international_regular_total, 0)) AS int_total,
                           SUM(COALESCE(dr.palawan_suki_discounts, 0))              AS skid,
                           SUM(COALESCE(dr.palawan_suki_rebates, 0))                AS skir,
                           SUM(COALESCE(dr.palawan_cancel, 0))                      AS cancellation,
                           SUM(COALESCE(dr.palawan_pay_out_incentives, 0))          AS po_incentives,
                           c.name                                                   AS corp_name
                    FROM branches b
                    LEFT JOIN corporations c ON (b.corporation_id = c.id OR b.sub_corporation_id = c.id)
                    LEFT JOIN {self.daily_table} dr ON b.name COLLATE utf8mb4_general_ci = dr.branch COLLATE utf8mb4_general_ci
                """
            else:
                select_cols = f"""
                    SELECT b.name as branch,
                           COALESCE(dr.palawan_sendout_lotes_total, 0)         AS so_lotes,
                           COALESCE(dr.palawan_sendout_principal, 0)           AS so_capital,
                           COALESCE(dr.palawan_sendout_sc, 0)                  AS so_sc,
                           COALESCE(dr.palawan_sendout_commission, 0)          AS so_commission,
                           COALESCE(dr.palawan_sendout_regular_total, 0)       AS so_total,
                           COALESCE(dr.palawan_payout_lotes_total, 0)          AS po_lotes,
                           COALESCE(dr.palawan_payout_principal, 0)            AS po_capital,
                           COALESCE(dr.palawan_payout_sc, 0)                   AS po_sc,
                           COALESCE(dr.palawan_payout_commission, 0)           AS po_commission,
                           COALESCE(dr.palawan_payout_regular_total, 0)        AS po_total,
                           COALESCE(dr.palawan_international_lotes_total, 0)   AS int_lotes,
                           COALESCE(dr.palawan_international_principal, 0)     AS int_capital,
                           COALESCE(dr.palawan_international_sc, 0)            AS int_sc,
                           COALESCE(dr.palawan_international_commission, 0)    AS int_commission,
                           COALESCE(dr.palawan_international_regular_total, 0) AS int_total,
                           COALESCE(dr.palawan_suki_discounts, 0)              AS skid,
                           COALESCE(dr.palawan_suki_rebates, 0)                AS skir,
                           COALESCE(dr.palawan_cancel, 0)                      AS cancellation,
                           COALESCE(dr.palawan_pay_out_incentives, 0)          AS po_incentives,
                           c.name                                              AS corp_name
                    FROM branches b
                    LEFT JOIN corporations c ON (b.corporation_id = c.id OR b.sub_corporation_id = c.id)
                    LEFT JOIN {self.daily_table} dr ON b.name COLLATE utf8mb4_general_ci = dr.branch COLLATE utf8mb4_general_ci
                """

            # ── Build WHERE clauses dynamically ───────────────────────────────
            if is_range:
                where_parts = ["dr.date >= %s", "dr.date <= %s"]
                params = [date_start, date_end]
            else:
                where_parts = ["dr.date = %s"]
                params = [date_start]

            if corp:
                where_parts.append("(b.corporation_id = (SELECT id FROM corporations WHERE name = %s) OR b.sub_corporation_id = (SELECT id FROM corporations WHERE name = %s))")
                params.extend([corp, corp])

            # Registration filter: "all" shows everything, "registered" and "not_registered" filter
            if reg_filter == "registered":
                where_parts.append("b.is_registered = 1")
            elif reg_filter == "not_registered":
                where_parts.append("(b.is_registered = 0 OR b.is_registered IS NULL)")
            # "all" - no filter needed

            if os_filter:
                where_parts.append("b.os_name = %s")
                params.append(os_filter)

            # Category filter: 60% = specific corporations, 30% = Global tagging
            if category_filter == "60":
                corps_to_filter = SIXTY_PERCENT_CORPS
                placeholders = ", ".join(["%s"] * len(corps_to_filter))
                where_parts.append(f"c.name IN ({placeholders})")
                params.extend(corps_to_filter)
            elif category_filter == "30":
                where_parts.append("b.global_tag = 'GLOBAL'")

            group_by = " GROUP BY b.name, c.name" if is_range else ""
            daily_query = select_cols + " WHERE " + " AND ".join(where_parts) + group_by + " ORDER BY b.name"

            results = db_manager.execute_query(daily_query, tuple(params))

            self.table.setRowCount(0)

            if not results:
                self._is_loading = False
                self.table.blockSignals(False)
                self.table.setUpdatesEnabled(True)
                QMessageBox.information(self, "No Data", f"No data found for the selected date(s).")
                return

            column_totals = [0.0] * (self.table.columnCount() - 1)
            
            # ── Pre-allocate rows for better performance with 600+ branches ──
            self.table.setRowCount(len(results))

            for row_count, row_data in enumerate(results):

                branch_name = row_data['branch'] if isinstance(row_data, dict) else row_data[0]
                # Store the corporation this branch belongs to (for cross-corp OS filtering)
                row_corp = row_data['corp_name'] if isinstance(row_data, dict) else ''
                self.table.setItem(row_count, 0, QTableWidgetItem(branch_name))
                # Stash corp in the branch item's data role so _sync can retrieve it
                self.table.item(row_count, 0).setData(Qt.UserRole, row_corp)

                # Columns 1-15: daily report values (read-only)
                if isinstance(row_data, dict):
                    values = [
                        float(row_data['so_lotes']),   float(row_data['so_capital']),
                        float(row_data['so_sc']),      float(row_data['so_commission']),
                        float(row_data['so_total']),
                        float(row_data['po_lotes']),   float(row_data['po_capital']),
                        float(row_data['po_sc']),      float(row_data['po_commission']),
                        float(row_data['po_total']),
                        float(row_data['int_lotes']),  float(row_data['int_capital']),
                        float(row_data['int_sc']),     float(row_data['int_commission']),
                        float(row_data['int_total']),
                    ]
                    # SKID, SKIR, CANCELLATION, INC from daily_reports
                    skid_val = float(row_data['skid'])
                    skir_val = float(row_data['skir'])
                    cancel_val = float(row_data['cancellation'])
                    inc_val = float(row_data['po_incentives'])
                else:
                    values = [float(x) for x in row_data[1:16]]
                    # SKID, SKIR, CANCELLATION, INC from tuple indices 16, 17, 18, 19
                    skid_val = float(row_data[16]) if len(row_data) > 16 else 0.0
                    skir_val = float(row_data[17]) if len(row_data) > 17 else 0.0
                    cancel_val = float(row_data[18]) if len(row_data) > 18 else 0.0
                    inc_val = float(row_data[19]) if len(row_data) > 19 else 0.0

                for i, value in enumerate(values, 1):
                    item = QTableWidgetItem(f"{value:.2f}")
                    item.setTextAlignment(Qt.AlignCenter)
                    self.table.setItem(row_count, i, item)
                    column_totals[i - 1] += value

                # Columns 16-18: SKID, SKIR, CANCELLATION from daily_reports (read-only)
                for col_index, val in [(16, skid_val), (17, skir_val), (18, cancel_val)]:
                    item = QTableWidgetItem(f"{val:.2f}")
                    item.setTextAlignment(Qt.AlignCenter)
                    item.setFlags(item.flags() & ~Qt.ItemIsEditable)  # Read-only
                    self.table.setItem(row_count, col_index, item)
                    column_totals[col_index - 1] += val

                # Column 19: INC from daily_reports.palawan_pay_out_incentives (read-only)
                item = QTableWidgetItem(f"{inc_val:.2f}")
                item.setTextAlignment(Qt.AlignCenter)
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)  # Read-only – carried from cash flow
                self.table.setItem(row_count, 19, item)
                column_totals[18] += inc_val

            self.add_totals_row(column_totals)
            self.add_group_headers_visual()
            QTimer.singleShot(200, self.adjust_responsive_widths)

            # ── Auto-sync INC values to payable_tbl so report_page can read them ──
            if not is_range:
                self._sync_inc_to_payable(date_start)

        except Exception as e:
            QMessageBox.critical(self, "Database Error", f"Error loading data: {str(e)}")
            import traceback; traceback.print_exc()

        finally:
            # ── RE-ENABLE UI updates and signals ──────────────────────────────
            self.table.blockSignals(False)
            self.table.setUpdatesEnabled(True)
            # ── CLEAR FLAG: user edits can now trigger auto-save again ──────
            self._is_loading = False

    # ─────────────────────────────────────────────────────────────────────────
    def _sync_inc_to_payable(self, selected_date):
        """Sync palawan_pay_out_incentives (INC) from daily_reports into payable table
        so that report_page can read them. Uses batch execution for performance."""
        try:
            totals_row = self.table.rowCount() - 1
            
            # Collect all rows data first
            batch_params = []
            for row in range(totals_row):
                branch_item = self.table.item(row, 0)
                if not branch_item:
                    continue
                branch = branch_item.text()
                # Get the corporation stored per-row (handles cross-corp OS filter)
                row_corp = branch_item.data(Qt.UserRole) or self.corp_selector.currentText()
                if not row_corp:
                    continue
                inc_item = self.table.item(row, 19)
                inc_val = float(inc_item.text()) if inc_item and inc_item.text().strip() else 0.0

                def get(c):
                    it = self.table.item(row, c)
                    try:
                        return float(it.text()) if it and it.text().strip() else 0.0
                    except ValueError:
                        return 0.0

                batch_params.append((
                    row_corp, branch, selected_date,
                    get(2),  get(3),  get(4),  get(5),
                    get(7),  get(8),  get(9),  get(10),
                    get(12), get(13), get(14), get(15),
                    get(16), get(17), get(18), inc_val,
                ))

            # Execute batch upsert if we have data
            if batch_params:
                db_manager.execute_many(f"""
                    INSERT INTO {self.payable_table} (
                        corporation, branch, date,
                        sendout_capital, sendout_sc, sendout_commission, sendout_total,
                        payout_capital,  payout_sc,  payout_commission,  payout_total,
                        international_capital, international_sc,
                        international_commission, international_total,
                        skid, skir, cancellation, inc
                    ) VALUES (
                        %s, %s, %s,
                        %s, %s, %s, %s,
                        %s, %s, %s, %s,
                        %s, %s, %s, %s,
                        %s, %s, %s, %s
                    )
                    ON DUPLICATE KEY UPDATE
                        inc        = VALUES(inc),
                        updated_at = CURRENT_TIMESTAMP
                """, batch_params)
        except Exception as e:
            print(f"INC auto-sync error: {e}")

    # ─────────────────────────────────────────────────────────────────────────
    def on_item_changed(self, item):
        """
        Fires whenever any cell changes.
        Guard with _is_loading so it does NOT run during populate_table.
        Guard with blockSignals so setText() doesn't cause infinite recursion.
        """
        if self._is_loading:
            return  # ← table is being filled — do nothing

        row = item.row()
        col = item.column()

        # Only INC column (19) is editable now
        if col == 19:
            # Block signals while we reformat the text to avoid recursion
            self.table.blockSignals(True)
            try:
                value = float(item.text()) if item.text().strip() else 0.0
                item.setText(f"{value:.2f}")
            except ValueError:
                item.setText("0.00")
                QMessageBox.warning(self, "Invalid Input", "Please enter a valid number.")
                return
            finally:
                self.table.blockSignals(False)

            self.calculate_adjustment_totals()
            self.save_single_row_adjustments(row)

    # ─────────────────────────────────────────────────────────────────────────
    def save_single_row_adjustments(self, row):
        """Auto-save INC value for one row immediately after edit using upsert."""
        if self._is_loading:
            return
        if row >= self.table.rowCount() - 1:  # skip totals row
            return

        selected_date = self.date_range_widget.get_date_range()[0]  # use start date
        branch_item   = self.table.item(row, 0)

        if not branch_item:
            return

        # Get per-row corporation stored during populate_table
        corp = branch_item.data(Qt.UserRole) or self.corp_selector.currentText()
        if not corp:
            return

        branch = branch_item.text()

        def get(col):
            item = self.table.item(row, col)
            try:
                return float(item.text()) if item and item.text().strip() else 0.0
            except ValueError:
                return 0.0

        try:
            # Get all values from the table
            skid         = get(16)  # From daily_reports (read-only in UI)
            skir         = get(17)  # From daily_reports (read-only in UI)
            cancellation = get(18)  # From daily_reports (read-only in UI)
            inc          = get(19)  # Editable by user

            # Use upsert - no need for a separate SELECT check
            db_manager.execute_query(f"""
                INSERT INTO {self.payable_table} (
                    corporation, branch, date,
                    sendout_capital, sendout_sc, sendout_commission, sendout_total,
                    payout_capital,  payout_sc,  payout_commission,  payout_total,
                    international_capital, international_sc,
                    international_commission, international_total,
                    skid, skir, cancellation, inc
                ) VALUES (
                    %s, %s, %s,
                    %s, %s, %s, %s,
                    %s, %s, %s, %s,
                    %s, %s, %s, %s,
                    %s, %s, %s, %s
                )
                ON DUPLICATE KEY UPDATE
                    skid         = VALUES(skid),
                    skir         = VALUES(skir),
                    cancellation = VALUES(cancellation),
                    inc          = VALUES(inc),
                    updated_at   = CURRENT_TIMESTAMP
            """, (
                corp, branch, selected_date,
                get(2),  get(3),  get(4),  get(5),
                get(7),  get(8),  get(9),  get(10),
                get(12), get(13), get(14), get(15),
                skid, skir, cancellation, inc,
            ))

        except Exception as e:
            print(f"Auto-save error for branch '{branch}': {e}")

    # ─────────────────────────────────────────────────────────────────────────
    def save_to_database(self):
        """Manual save — writes every row to payable table using batch execution."""
        selected_date = self.date_range_widget.get_date_range()[0]  # use start date

        if self.table.rowCount() == 0:
            QMessageBox.warning(self, "No Data", "No data to save.")
            return

        try:
            totals_row    = self.table.rowCount() - 1

            upsert = f"""
                INSERT INTO {self.payable_table} (
                    corporation, branch, date,
                    sendout_capital, sendout_sc, sendout_commission, sendout_total,
                    payout_capital,  payout_sc,  payout_commission,  payout_total,
                    international_capital, international_sc,
                    international_commission, international_total,
                    skid, skir, cancellation, inc
                ) VALUES (
                    %s, %s, %s,
                    %s, %s, %s, %s,
                    %s, %s, %s, %s,
                    %s, %s, %s, %s,
                    %s, %s, %s, %s
                )
                ON DUPLICATE KEY UPDATE
                    sendout_capital          = VALUES(sendout_capital),
                    sendout_sc               = VALUES(sendout_sc),
                    sendout_commission       = VALUES(sendout_commission),
                    sendout_total            = VALUES(sendout_total),
                    payout_capital           = VALUES(payout_capital),
                    payout_sc                = VALUES(payout_sc),
                    payout_commission        = VALUES(payout_commission),
                    payout_total             = VALUES(payout_total),
                    international_capital    = VALUES(international_capital),
                    international_sc         = VALUES(international_sc),
                    international_commission = VALUES(international_commission),
                    international_total      = VALUES(international_total),
                    skid                     = VALUES(skid),
                    skir                     = VALUES(skir),
                    cancellation             = VALUES(cancellation),
                    inc                      = VALUES(inc),
                    updated_at               = CURRENT_TIMESTAMP
            """

            def get(r, c):
                item = self.table.item(r, c)
                try:
                    return float(item.text()) if item and item.text().strip() else 0.0
                except ValueError:
                    return 0.0

            # Collect all row data for batch execution
            batch_params = []
            for row in range(totals_row):
                branch_item = self.table.item(row, 0)
                if not branch_item:
                    continue
                branch = branch_item.text()
                # Get per-row corporation stored during populate_table
                row_corp = branch_item.data(Qt.UserRole) or self.corp_selector.currentText()
                if not row_corp:
                    continue

                batch_params.append((
                    row_corp, branch, selected_date,
                    get(row,2),  get(row,3),  get(row,4),  get(row,5),
                    get(row,7),  get(row,8),  get(row,9),  get(row,10),
                    get(row,12), get(row,13), get(row,14), get(row,15),
                    get(row,16), get(row,17), get(row,18), get(row,19),
                ))

            # Execute batch upsert
            if batch_params:
                result = db_manager.execute_many(upsert, batch_params)
                saved_count = len(batch_params) if result is not None else 0
                error_count = 0 if result is not None else len(batch_params)
            else:
                saved_count = 0
                error_count = 0

            msg = f"Save Complete!\n\nRecords saved: {saved_count}\n"
            if error_count:
                msg += f"Errors:        {error_count}\n"
            corp_display = self.corp_selector.currentText() or "All Corporations"
            msg += f"\nCorporation: {corp_display}\nDate: {selected_date}"
            QMessageBox.information(self, "Save Successful", msg)

        except Exception as e:
            QMessageBox.critical(self, "Database Error", f"Error saving: {str(e)}")

    # ─────────────────────────────────────────────────────────────────────────
    # All remaining methods — unchanged
    # ─────────────────────────────────────────────────────────────────────────

    def create_controls(self):
        controls_frame = QFrame()
        controls_frame.setFrameStyle(QFrame.Box)
        controls_frame.setLineWidth(1)
        controls_frame.setStyleSheet("""
            QFrame {
                background-color: #f8f9fa; border: 1px solid #dee2e6;
                border-radius: 5px; padding: 4px;
            }
        """)
        controls_layout = QVBoxLayout(controls_frame)
        controls_layout.setContentsMargins(6, 4, 6, 4)
        controls_layout.setSpacing(4)

        _combo_ss = """
            QComboBox { padding:4px 6px; border:2px solid #dee2e6; border-radius:4px;
                        background-color:white; font-size:12px; }
            QComboBox:focus { border-color:#007bff; }
            QComboBox::drop-down { width: 24px; }
        """

        # ── Row 1: Corporation + Date ────────────────────────────────────
        row1 = QHBoxLayout()
        row1.setSpacing(8)

        # Corporation filter (hidden for Brand A admin)
        corp_label = QLabel("Corporation:")
        corp_label.setFont(QFont("Arial", 10, QFont.Bold))
        self.corp_selector = QComboBox()
        self.corp_selector.setMinimumHeight(30)
        self.corp_selector.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.corp_selector.currentTextChanged.connect(self.populate_table)
        self.corp_selector.setStyleSheet(_combo_ss)

        # Hide corporation filter for Brand A admin (account_type == 1)
        self.corp_label = corp_label
        if self.account_type == 1:
            corp_label.hide()
            self.corp_selector.hide()

        date_label = QLabel("Date:")
        date_label.setFont(QFont("Arial", 10, QFont.Bold))
        self.date_range_widget = DateRangeWidget()
        self.date_range_widget.dateRangeChanged.connect(self.populate_table)
        self.date_selector = self.date_range_widget  # backward-compat

        if self.account_type != 1:
            row1.addWidget(corp_label)
            row1.addWidget(self.corp_selector, 1)

        row1.addWidget(date_label)
        row1.addWidget(self.date_range_widget, 2)
        controls_layout.addLayout(row1)

        # ── Row 2: Branch Status + Group Filter + Category ───────────────
        row2 = QHBoxLayout()
        row2.setSpacing(8)

        # Registration status filter
        reg_filter_label = QLabel("Branch Status:")
        reg_filter_label.setFont(QFont("Arial", 10, QFont.Bold))
        self.reg_filter_selector = QComboBox()
        self.reg_filter_selector.setMinimumHeight(30)
        self.reg_filter_selector.addItem("All", "all")
        self.reg_filter_selector.addItem("Registered Only", "registered")
        self.reg_filter_selector.addItem("Unregistered", "not_registered")
        self.reg_filter_selector.setCurrentIndex(1)  # Default to "Registered Only"
        self.reg_filter_selector.currentIndexChanged.connect(self.populate_table)
        self.reg_filter_selector.setStyleSheet(_combo_ss)

        # OS (Operation Supervisor) filter
        os_filter_label = QLabel("Group Filter:" if self.account_type != 1 else "Group:")
        os_filter_label.setFont(QFont("Arial", 10, QFont.Bold))
        self.os_filter_selector = QComboBox()
        self.os_filter_selector.setMinimumHeight(30)
        if self.account_type != 1:
            self.os_filter_selector.addItem("All (by Corporation)", None)
        self.os_filter_selector.currentIndexChanged.connect(self.populate_table)
        self.os_filter_selector.setStyleSheet(_combo_ss)

        # Category filter (60% / 30%)
        category_filter_label = QLabel("Category:")
        category_filter_label.setFont(QFont("Arial", 10, QFont.Bold))
        self.category_filter_selector = QComboBox()
        self.category_filter_selector.setMinimumHeight(30)
        self.category_filter_selector.addItem("All", "all")
        self.category_filter_selector.addItem("60%", "60")
        self.category_filter_selector.addItem("30%", "30")
        self.category_filter_selector.currentIndexChanged.connect(self.populate_table)
        self.category_filter_selector.setStyleSheet(_combo_ss)

        row2.addWidget(reg_filter_label)
        row2.addWidget(self.reg_filter_selector, 1)
        row2.addWidget(os_filter_label)
        row2.addWidget(self.os_filter_selector, 1)
        row2.addWidget(category_filter_label)
        row2.addWidget(self.category_filter_selector, 1)
        row2.addStretch()
        controls_layout.addLayout(row2)

        selectors_layout = QHBoxLayout()  # kept for backward-compat (create_group_headers_layout)

        controls_layout.addLayout(selectors_layout)
        self.main_layout.addWidget(controls_frame, 0)  # stretch=0, fixed size
        self.create_group_headers_layout(controls_layout)

    def create_group_headers_layout(self, parent_layout):
        # Group headers are no longer needed since column headers have prefixes (SO:, PO:, IN:)
        # This provides better visibility on small screens
        self.header_widgets = []

    def adjust_header_widths(self):
        if not hasattr(self, 'table') or not hasattr(self, 'header_widgets'):
            return
        current_col = 0
        for header in self.header_widgets:
            total_width = 0
            for i in range(header.column_count):
                if current_col < self.table.columnCount():
                    total_width += self.table.columnWidth(current_col)
                    current_col += 1
            if total_width > 0:
                header.setFixedWidth(max(total_width, 100))

    def create_table(self):
        # ══════════════════════════════════════════════════════════════════════
        # Create scroll area for the table
        # ══════════════════════════════════════════════════════════════════════
        table_scroll = QScrollArea()
        table_scroll.setWidgetResizable(True)
        table_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        table_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        table_scroll.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        
        self.table = QTableWidget()
        self.table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.table.setColumnCount(20)
        
        # Include section prefix in each column header so it's always visible
        self.table.setHorizontalHeaderLabels([
            "Branch",
            "SO:Lotes", "SO:Cap", "SO:SC", "SO:Com", "SO:TOTAL",
            "PO:Lotes", "PO:Cap", "PO:SC", "PO:Com", "PO:TOTAL",
            "IN:Lotes", "IN:Cap", "IN:SC", "IN:Com", "IN:TOTAL",
            "SKID", "SKIR", "CANCEL", "INC",
        ])
        # Set minimum width for horizontal scrolling - reduced for small screens
        self.table.setMinimumWidth(1000)
        self.table.setMinimumHeight(300)
        self.table.verticalHeader().setDefaultSectionSize(35)  # Taller rows
        self.setup_responsive_columns()
        self.style_table_with_grouped_headers()
        self.table.itemChanged.connect(self.on_item_changed)

        table_scroll.setWidget(self.table)
        self.main_layout.addWidget(table_scroll, 1)  # stretch factor 1 to expand

    def setup_responsive_columns(self):
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.table.setColumnWidth(0, 110)  # Branch column - wider
        
        # Set column widths with wider separators at section starts
        for i in range(1, self.table.columnCount()):
            # Total columns (5, 10, 15) - slightly wider
            if i in [5, 10, 15]:
                self.table.setColumnWidth(i, 85)
            # First column of each section (1, 6, 11, 16) - extra left padding
            elif i in [1, 6, 11, 16]:
                self.table.setColumnWidth(i, 80)
            else:
                self.table.setColumnWidth(i, 70)
            header.setSectionResizeMode(i, QHeaderView.Interactive)
        header.setStretchLastSection(False)
        QTimer.singleShot(100, self.adjust_responsive_widths)

    def adjust_responsive_widths(self):
        if not self.table.isVisible():
            return
        available_width     = self.table.viewport().width()
        current_total_width = sum(self.table.columnWidth(i) for i in range(self.table.columnCount()))
        if available_width > current_total_width:
            extra      = available_width - current_total_width
            extra_each = extra // (self.table.columnCount() - 1)
            for i in range(1, self.table.columnCount()):
                self.table.setColumnWidth(i, min(self.table.columnWidth(i) + extra_each, 120))
        self.adjust_header_widths()

    def resizeEvent(self, event):
        super().resizeEvent(event)
        if hasattr(self, 'resize_timer'):
            self.resize_timer.stop()
        else:
            self.resize_timer = QTimer()
            self.resize_timer.setSingleShot(True)
            self.resize_timer.timeout.connect(self.adjust_responsive_widths)
        self.resize_timer.start(150)

    def style_table_with_grouped_headers(self):
        # Use stylesheet with specific section coloring in header
        self.table.setStyleSheet("""
            QTableWidget {
                gridline-color:#d0d0d0; border:1px solid #c0c0c0;
                background-color:white; alternate-background-color:#f8f9fa;
                font-size:12px; selection-background-color:#e3f2fd;
            }
            QTableWidget::item { border:1px solid #e0e0e0; padding:10px 6px; }
            QTableWidget::item:selected { background-color:#e3f2fd; color:black; }
            QTableWidget::item:focus { background-color:#bbdefb; border:2px solid #2196f3; }
            QHeaderView::section {
                padding:10px 4px; font-weight:bold; font-size:11px;
                border:1px solid #666; color: white;
            }
            QHeaderView::section:hover { opacity: 0.9; }
        """)
        self.table.setAlternatingRowColors(True)
        
        # Set custom header with colors using a model
        self._apply_header_colors()
    
    def _apply_header_colors(self):
        """Apply colors to header using the ColoredHeaderView class"""
        # Create and set the custom header
        colored_header = ColoredHeaderView(Qt.Horizontal, self.table)
        colored_header.colors = {
            0: QColor("#495057"),  # Branch - gray
            1: QColor("#dc3545"), 2: QColor("#dc3545"), 3: QColor("#dc3545"), 4: QColor("#dc3545"), 5: QColor("#b02a37"),  # SO - red (darker for total)
            6: QColor("#28a745"), 7: QColor("#28a745"), 8: QColor("#28a745"), 9: QColor("#28a745"), 10: QColor("#1e7e34"),  # PO - green (darker for total)
            11: QColor("#007bff"), 12: QColor("#007bff"), 13: QColor("#007bff"), 14: QColor("#007bff"), 15: QColor("#0056b3"),  # INT - blue (darker for total)
            16: QColor("#6f42c1"), 17: QColor("#6f42c1"), 18: QColor("#6f42c1"), 19: QColor("#6f42c1"),  # ADJ - purple
        }
        colored_header.setDefaultAlignment(Qt.AlignCenter)
        colored_header.setMinimumSectionSize(50)
        colored_header.setDefaultSectionSize(70)
        self.table.setHorizontalHeader(colored_header)

    def create_buttons(self):
        button_frame = QFrame()
        button_frame.setFixedHeight(80)  # Fixed height ensures buttons are always visible
        button_layout = QHBoxLayout(button_frame)
        button_layout.setContentsMargins(0, 10, 0, 10)

        def styled(c1, c2, c3):
            return f"""
                QPushButton {{
                    padding:14px 28px; border:2px solid {c1}; border-radius:8px;
                    background-color:{c1}; color:white; font-weight:bold;
                    font-size:13px; min-width:140px; min-height:45px;
                }}
                QPushButton:hover   {{ background-color:{c2}; border-color:{c2}; }}
                QPushButton:pressed {{ background-color:{c3}; }}
            """

        self.export_button = QPushButton("📊 Export to Excel")
        self.export_button.clicked.connect(self.export_to_excel)
        self.export_button.setStyleSheet(styled("#217346","#1a5c38","#155724"))

        self.print_button = QPushButton("🖨️ Print Report")
        self.print_button.clicked.connect(self.print_table)
        self.print_button.setStyleSheet(styled("#6f42c1","#5a2d91","#4c1f75"))

        button_layout.addStretch()
        button_layout.addWidget(self.export_button)
        button_layout.addSpacing(15)
        button_layout.addWidget(self.print_button)
        button_layout.addStretch()

        self.main_layout.addWidget(button_frame, 0)  # stretch=0, fixed at bottom

    def load_corporations(self):
        self.corp_selector.blockSignals(True)
        self.corp_selector.clear()
        self.corp_selector.addItem("")  # empty = All Corporations (when using OS filter)
        try:
            results = db_manager.execute_query(
                f"SELECT DISTINCT corporation FROM {self.daily_table} ORDER BY corporation"
            )
            for corp in results:
                self.corp_selector.addItem(corp['corporation'])
        except Exception as e:
            QMessageBox.critical(self, "Database Error", f"Error loading corporations: {str(e)}")
        finally:
            self.corp_selector.blockSignals(False)

        # Also refresh OS filter dropdown
        self._load_os_options()

    def _load_os_options(self):
        """Populate the OS filter dropdown with distinct OS names from branches."""
        self.os_filter_selector.blockSignals(True)
        current = self.os_filter_selector.currentData()
        self.os_filter_selector.clear()
        if self.account_type != 1:
            self.os_filter_selector.addItem("All (by Corporation)", None)
        try:
            rows = db_manager.execute_query(
                "SELECT DISTINCT os_name FROM branches WHERE os_name IS NOT NULL AND os_name != '' ORDER BY os_name"
            )
            if rows:
                for r in rows:
                    os_name = r['os_name'] if isinstance(r, dict) else r[0]
                    self.os_filter_selector.addItem(os_name, os_name)
            # Restore previous selection if still valid
            if current:
                idx = self.os_filter_selector.findData(current)
                if idx >= 0:
                    self.os_filter_selector.setCurrentIndex(idx)
        except Exception as e:
            print(f"Error loading OS options: {e}")
        finally:
            self.os_filter_selector.blockSignals(False)

    def add_totals_row(self, column_totals):
        row_count   = self.table.rowCount()
        self.table.insertRow(row_count)
        bold_font   = QFont(); bold_font.setBold(True)
        gray_brush  = QBrush(QColor("#e9ecef"))

        item = QTableWidgetItem("TOTAL")
        item.setFont(bold_font); item.setBackground(gray_brush)
        item.setTextAlignment(Qt.AlignCenter)
        self.table.setItem(row_count, 0, item)

        for i, total in enumerate(column_totals, 1):
            item = QTableWidgetItem(f"{total:.2f}")
            item.setFont(bold_font); item.setBackground(gray_brush)
            item.setTextAlignment(Qt.AlignCenter)
            item.setFlags(item.flags() & ~Qt.ItemIsEditable)
            self.table.setItem(row_count, i, item)

    def add_group_headers_visual(self):
        """Apply color coding to table cells matching the section headers"""
        if self.table.rowCount() == 0:
            return
        
        # Colors matching the header sections - more saturated for visibility
        colors = {
            'so':       QColor("#ffcdd2"),   # Light red for Send Out
            'so_total': QColor("#ef9a9a"),   # Darker red for SO Total
            'po':       QColor("#c8e6c9"),   # Light green for Payout
            'po_total': QColor("#a5d6a7"),   # Darker green for PO Total
            'int':      QColor("#bbdefb"),   # Light blue for International
            'int_total':QColor("#90caf9"),   # Darker blue for INT Total
            'adj':      QColor("#e1bee7"),   # Light purple for Adjustments
            'tot':      QColor("#dee2e6"),   # Gray for totals row
            'sep':      QColor("#6c757d"),   # Dark separator
        }
        
        last = self.table.rowCount() - 1
        for row in range(self.table.rowCount()):
            is_last = (row == last)
            
            # Apply background colors to Send Out columns (1-5)
            for col in range(1, 6):
                item = self.table.item(row, col)
                if item:
                    if is_last:
                        item.setBackground(QBrush(colors['tot']))
                        item.setFont(QFont("", 0, QFont.Bold))
                    elif col == 5:
                        item.setBackground(QBrush(colors['so_total']))
                        item.setFont(QFont("", 0, QFont.Bold))
                    else:
                        item.setBackground(QBrush(colors['so']))
            
            # Apply background colors to Payout columns (6-10)
            for col in range(6, 11):
                item = self.table.item(row, col)
                if item:
                    if is_last:
                        item.setBackground(QBrush(colors['tot']))
                        item.setFont(QFont("", 0, QFont.Bold))
                    elif col == 10:
                        item.setBackground(QBrush(colors['po_total']))
                        item.setFont(QFont("", 0, QFont.Bold))
                    else:
                        item.setBackground(QBrush(colors['po']))
            
            # Apply background colors to International columns (11-15)
            for col in range(11, 16):
                item = self.table.item(row, col)
                if item:
                    if is_last:
                        item.setBackground(QBrush(colors['tot']))
                        item.setFont(QFont("", 0, QFont.Bold))
                    elif col == 15:
                        item.setBackground(QBrush(colors['int_total']))
                        item.setFont(QFont("", 0, QFont.Bold))
                    else:
                        item.setBackground(QBrush(colors['int']))
            
            # Apply background colors to Adjustment columns (16-19)
            for col in range(16, 20):
                item = self.table.item(row, col)
                if item:
                    item.setBackground(QBrush(colors['tot'] if is_last else colors['adj']))
                    if is_last: 
                        item.setFont(QFont("", 0, QFont.Bold))
            
            # Branch column
            if is_last:
                item = self.table.item(row, 0)
                if item:
                    item.setBackground(QBrush(colors['tot']))
                    item.setFont(QFont("", 0, QFont.Bold))

    def calculate_adjustment_totals(self):
        if self.table.rowCount() == 0:
            return
        totals_row = self.table.rowCount() - 1
        for col in range(16, 20):
            total = 0.0
            for row in range(totals_row):
                item = self.table.item(row, col)
                if item and item.text().strip():
                    try:
                        total += float(item.text())
                    except ValueError:
                        pass
            total_item = self.table.item(totals_row, col)
            if total_item:
                total_item.setText(f"{total:.2f}")
                total_item.setBackground(QBrush(QColor(
                    "#fff3cd" if abs(total) > 0.01 else "#e9ecef"
                )))

    def export_to_excel(self):
        """Export table data to Excel with file dialog for save location"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            QMessageBox.critical(
                self,
                "Missing Dependency",
                "The openpyxl package is required to export to Excel.\nInstall with: pip install openpyxl"
            )
            return
        
        corp = self.corp_selector.currentText()
        date_start, date_end = self.date_range_widget.get_date_range()
        date = date_start if date_start == date_end else f"{date_start}_to_{date_end}"
        
        rows = self.table.rowCount()
        cols = self.table.columnCount()
        if rows == 0:
            QMessageBox.warning(self, "No Data", "No data to export.")
            return
        
        # File dialog for save location
        default_filename = f"Palawan_Report_{corp}_{date}.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Excel File",
            default_filename,
            "Excel Files (*.xlsx);;All Files (*)"
        )
        
        if not file_path:
            return  # User cancelled
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Palawan Report"
            
            # Styles
            title_font = Font(bold=True, size=16)
            header_font = Font(bold=True, size=10, color="FFFFFF")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Color fills for sections
            so_fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")  # Red
            po_fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")  # Green
            int_fill = PatternFill(start_color="007BFF", end_color="007BFF", fill_type="solid")  # Blue
            adj_fill = PatternFill(start_color="6F42C1", end_color="6F42C1", fill_type="solid")  # Purple
            gray_fill = PatternFill(start_color="495057", end_color="495057", fill_type="solid")
            total_fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
            
            # Title
            ws.merge_cells('A1:T1')
            ws['A1'] = "Detailed Palawan Transaction Report"
            ws['A1'].font = title_font
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Info section
            ws['A3'] = "Corporation:"
            ws['B3'] = corp
            ws['A4'] = "Date:"
            ws['B4'] = date
            ws['A5'] = "Generated:"
            ws['B5'] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            for row_num in range(3, 6):
                ws[f'A{row_num}'].font = Font(bold=True)
            
            # Headers (row 7)
            header_row = 7
            for col in range(cols):
                cell = ws.cell(row=header_row, column=col+1)
                header_item = self.table.horizontalHeaderItem(col)
                cell.value = header_item.text() if header_item else ""
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', wrap_text=True)
                
                # Apply section colors to headers
                if col == 0:
                    cell.fill = gray_fill
                elif 1 <= col <= 5:
                    cell.fill = so_fill
                elif 6 <= col <= 10:
                    cell.fill = po_fill
                elif 11 <= col <= 15:
                    cell.fill = int_fill
                else:
                    cell.fill = adj_fill
            
            # Data rows
            for row in range(rows):
                excel_row = header_row + 1 + row
                is_total_row = (row == rows - 1)
                
                for col in range(cols):
                    cell = ws.cell(row=excel_row, column=col+1)
                    item = self.table.item(row, col)
                    
                    if item:
                        text = item.text()
                        # Try to convert numeric values
                        if col > 0:  # Not branch column
                            try:
                                cell.value = float(text) if text else 0
                                cell.number_format = '#,##0.00'
                            except ValueError:
                                cell.value = text
                        else:
                            cell.value = text
                    
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')
                    
                    if is_total_row:
                        cell.fill = total_fill
                        cell.font = Font(bold=True)
            
            # Auto-adjust column widths
            from openpyxl.utils import get_column_letter
            ws.column_dimensions['A'].width = 15  # Branch
            for col in range(2, cols + 1):
                ws.column_dimensions[get_column_letter(col)].width = 12
            
            wb.save(file_path)
            QMessageBox.information(self, "Export Successful", f"Report exported to:\n{file_path}")
            
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Error exporting to Excel: {str(e)}")

    def print_table(self):
        try:
            if self.table.rowCount() == 0:
                QMessageBox.warning(self, "No Data", "No data to print."); return
            doc  = QTextDocument()
            html = """<html><head><style>
                body{font-family:Arial,sans-serif;margin:20px}
                h2{color:#2c3e50;text-align:center;margin-bottom:10px}
                .info{margin-bottom:20px;font-size:12px}
                table{border-collapse:collapse;width:100%;font-size:9px}
                th,td{border:1px solid #ddd;padding:4px 6px;text-align:center}
                th{background-color:#f8f9fa;font-weight:bold;color:#495057}
                .totals-row{background-color:#e9ecef;font-weight:bold}
                .so{background-color:#ffebee}.po{background-color:#e8f5e8}
                .intl{background-color:#e3f2fd}.adj{background-color:#f3e5f5}
            </style></head><body>"""
            html += "<h2>Detailed Palawan Transaction Report</h2>"
            date_start, date_end = self.date_range_widget.get_date_range()
            date_display = date_start if date_start == date_end else f"{date_start} to {date_end}"
            html += (f'<div class="info">'
                     f'<strong>Corporation:</strong> {self.corp_selector.currentText()}<br>'
                     f'<strong>Date:</strong> {date_display}<br>'
                     f'<strong>Generated:</strong> '
                     f'{datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</div>')
            html += "<table><tr>"
            for c in range(self.table.columnCount()):
                h = self.table.horizontalHeaderItem(c)
                html += f"<th>{h.text() if h else ''}</th>"
            html += "</tr>"
            for r in range(self.table.rowCount()):
                is_last = (r == self.table.rowCount() - 1)
                html += f'<tr class="{"totals-row" if is_last else ""}">'
                for c in range(self.table.columnCount()):
                    item = self.table.item(r, c)
                    text = item.text() if item else ""
                    cls  = ""
                    if not is_last:
                        if c == 5:          cls = "so"
                        elif c == 10:       cls = "po"
                        elif c == 15:       cls = "intl"
                        elif 16 <= c <= 19: cls = "adj"
                    html += f"<td class='{cls}'>{text}</td>"
                html += "</tr>"
            html += "</table></body></html>"
            doc.setHtml(html)

            printer = QPrinter()
            printer.setPageSize(QPrinter.A4)
            printer.setPageMargins(10, 10, 10, 10, QPrinter.Millimeter)
            dialog = QPrintDialog(printer, self)
            if dialog.exec_() == QPrintDialog.Accepted:
                doc.print_(printer)
                QMessageBox.information(self, "Print Successful", "Printed successfully!")
        except Exception as e:
            QMessageBox.critical(self, "Print Error", f"Error: {str(e)}")

    def showEvent(self, event):
        super().showEvent(event)
        QTimer.singleShot(300, self.adjust_responsive_widths)

    def update_totals(self):
        self.calculate_adjustment_totals()


def get_optimal_font_size(widget_width):
    if widget_width < 1200: return 9
    elif widget_width < 1600: return 10
    else: return 11

def apply_responsive_styling(widget, screen_width):
    if screen_width < 1366:   return "font-size:9px; padding:6px;"
    elif screen_width < 1920: return "font-size:10px; padding:8px;"
    else:                     return "font-size:11px; padding:10px;"